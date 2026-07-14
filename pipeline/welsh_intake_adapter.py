"""Generate checksum-bound, privacy-safe Welsh 2024-25 intake candidates."""

from __future__ import annotations

import argparse
import contextlib
import csv
import hashlib
import io
import json
import os
import re
import tempfile
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline.__main__ import (
    EXPOSURE_DECLARED_GRAIN_FIELD,
    EXPOSURE_LOCATOR_FIELDS,
    LOCATOR_FIELDS,
    TEAM_ALIAS_MAP_PATH,
    TEAM_KEY_ALIAS_LOOKUP_NAMES,
    URC_FIXTURES_2024_25_CORRECTED_PATH,
    URC_FIXTURES_2024_25_CORRECTED_SHA256,
    clean_cell,
    clean_exposure,
    load_fixture_team_aliases,
    parse_flexible_date,
    sha256_file,
    stable_uid,
    received_in_team_status,
    validate_intake_profile_manifest,
)
from pipeline.welsh_profile_package import (
    ADAPTER_VERSION,
    CONFIG,
    INTAKE,
    ROOT,
    SEASON,
    SHEET,
    SOURCE_ROOT,
    canonical_team_name,
)


CROSSWALK = SOURCE_ROOT / "players - codebookN.csv"
CROSSWALK_SHA256 = "aaa12b1a69b0a2e4779cc20194b05c9a77c5b06f21e9a8eed21327f659e62180"
GENERATOR_VERSION = "welsh_privacy_safe_intake_adapter_2026-07-14_v1"
WINDOW_START = "2024-09-20"
WINDOW_END = "2025-06-14"
PLAN_NAME = "source_adapter_plan.v1.draft.json"
BASE_DECISION_APPLICATIONS = INTAKE / "wales" / "decision_applications.v1.json"
REVISED_DECISION_APPLICATIONS = INTAKE / "wales" / "decision_applications.v2.json"
COHORT_DECISION_APPLICATIONS = INTAKE / "wales" / "decision_applications.v3.json"
FINAL_DECISION_APPLICATIONS = INTAKE / "wales" / "decision_applications.v4.json"
REDACTED = "[REDACTED_PROTECTED_METADATA]"
MEDICAL_EXPECTED = {"cardiff": 0, "dragons": 46, "ospreys": 34, "scarlets": 46}
LINK_EXPECTED = {
    "cardiff": {"pairs": 33, "remapped_rows": 82},
    "dragons": {"pairs": 49, "remapped_rows": 5693},
    "ospreys": {"pairs": 45, "remapped_rows": 265},
    "scarlets": {"pairs": 37, "remapped_rows": 425},
}
OVERRIDES = {
    "occasion_category": ("Adapter Canonical Activity Context", "Adapter Canonical Activity Context Origin"),
    "contact_context": ("Adapter Canonical Contact Context", "Adapter Canonical Contact Context Origin"),
    "recurrence": ("Adapter Canonical Recurrence Status", "Adapter Canonical Recurrence Status Origin"),
    "problem_type": ("Adapter Canonical Problem Type", "Adapter Canonical Problem Type Origin"),
    "body_location": ("Adapter Canonical Body Location", "Adapter Canonical Body Location Origin"),
    "tissue_pathology": ("Adapter Canonical Tissue Pathology", "Adapter Canonical Tissue Pathology Origin"),
}


def _norm(value: Any) -> str:
    return " ".join(unicodedata.normalize("NFKC", str(value or "")).strip().casefold().split())


def _json_write(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, indent=2, sort_keys=True) + "\n"
    fd, temporary = tempfile.mkstemp(dir=path.parent, prefix=f".{path.name}.", suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(text)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)


def _csv_write(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile.mkstemp(dir=path.parent, prefix=f".{path.name}.", suffix=".tmp")
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)


def _immutable_write(path: Path, data: bytes) -> None:
    if path.exists():
        if path.read_bytes() != data:
            raise ValueError(f"immutable approval artifact drift: {path}")
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile.mkstemp(dir=path.parent, prefix=f".{path.name}.", suffix=".tmp")
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)


def _read_xlsx(path: Path, sheet_name: str = SHEET) -> tuple[list[str], list[tuple[int, dict[str, Any]]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        values = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(values)]
        rows = []
        for row_number, values_row in enumerate(values, start=2):
            row = {header: values_row[index] if index < len(values_row) else None
                   for index, header in enumerate(headers) if header}
            rows.append((row_number, row))
        return [header for header in headers if header], rows, worksheet.max_row - 1
    finally:
        workbook.close()


@dataclass(frozen=True)
class Identity:
    pseudonym: str
    status: str
    collision_hash: str


class Crosswalk:
    def __init__(self, path: Path = CROSSWALK) -> None:
        if sha256_file(path) != CROSSWALK_SHA256:
            raise ValueError("Welsh player crosswalk checksum drift")
        with path.open(newline="", encoding="cp1252") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames != ["original_id", "new_id"]:
                raise ValueError("Welsh player crosswalk schema drift")
            rows = list(reader)
        self.original_to_new: dict[str, set[str]] = defaultdict(set)
        self.new_to_original: dict[str, str] = {}
        self.new_spelling: dict[str, str] = {}
        for row in rows:
            original, pseudonym = _norm(row["original_id"]), _norm(row["new_id"])
            if not original or not pseudonym:
                raise ValueError("Welsh player crosswalk contains a blank key")
            if pseudonym in self.new_to_original:
                raise ValueError("Welsh player crosswalk new_id is not unique")
            self.original_to_new[original].add(pseudonym)
            self.new_to_original[pseudonym] = original
            self.new_spelling[pseudonym] = str(row["new_id"]).strip()
        self.direct_only = set(self.original_to_new) - set(self.new_to_original)

    def resolve(self, value: Any, *, team_key: str) -> Identity:
        key = _norm(value)
        if not key:
            raise ValueError("candidate player identifier is blank")
        if key in self.new_to_original:
            original = self.new_to_original[key]
            collision = len(self.original_to_new[original]) > 1
            return Identity(
                self.new_spelling[key],
                "collision_quarantined" if collision else "existing_pseudonym",
                stable_uid("idq", CROSSWALK_SHA256, team_key, original) if collision else "",
            )
        targets = self.original_to_new.get(key, set())
        if len(targets) != 1:
            raise ValueError("candidate identifier is absent or ambiguous in the approved crosswalk")
        pseudonym_key = next(iter(targets))
        if self.new_to_original[pseudonym_key] != key:
            raise ValueError("candidate identifier does not have an exact one-to-one crosswalk link")
        return Identity(self.new_spelling[pseudonym_key], "exact_one_to_one_mapped", "")


@dataclass(frozen=True)
class TeamLinkage:
    exposure_target: dict[str, str]
    injury_status: dict[str, str]
    exposure_status: dict[str, str]
    injury_collision_hash: dict[str, str]
    exposure_collision_hash: dict[str, str]
    counts: dict[str, int]


def _loose_original(value: str) -> str:
    return re.sub(r"[\s\-'’]+", "", _norm(value))


def _build_team_linkage(
    team_key: str,
    injury_rows: list[dict[str, Any]],
    exposure_rows: list[dict[str, Any]],
    crosswalk: Crosswalk,
) -> TeamLinkage:
    injury_spelling: dict[str, str] = {}
    exposure_spelling: dict[str, str] = {}
    for row in injury_rows:
        identity = crosswalk.resolve(row.get("PlayerID"), team_key=team_key)
        injury_spelling[_norm(identity.pseudonym)] = identity.pseudonym
    for row in exposure_rows:
        identity = crosswalk.resolve(row.get("name"), team_key=team_key)
        exposure_spelling[_norm(identity.pseudonym)] = identity.pseudonym

    injury_remaining, exposure_remaining = set(injury_spelling), set(exposure_spelling)
    target: dict[str, str] = {}
    injury_status = {key: "unmatched_distinct" for key in injury_remaining}
    exposure_status = {key: "unmatched_distinct" for key in exposure_remaining}
    counts = Counter()

    for key in sorted(injury_remaining & exposure_remaining):
        target[key] = injury_spelling[key]
        injury_status[key] = exposure_status[key] = "tier0_identical_pseudonym"
        counts["tier0_pairs"] += 1
    injury_remaining -= set(target)
    exposure_remaining -= set(target)

    def unique_pairs(keys_i: set[str], keys_e: set[str], key_function: Any) -> list[tuple[str, str]]:
        grouped_i: dict[str, list[str]] = defaultdict(list)
        grouped_e: dict[str, list[str]] = defaultdict(list)
        for key in keys_i:
            grouped_i[key_function(key)].append(key)
        for key in keys_e:
            grouped_e[key_function(key)].append(key)
        return [
            (values[0], grouped_e[group][0])
            for group, values in grouped_i.items()
            if len(values) == 1 and len(grouped_e.get(group, [])) == 1
        ]

    original = lambda pseudonym: crosswalk.new_to_original[pseudonym]
    tier1 = unique_pairs(injury_remaining, exposure_remaining, original)
    for injury_key, exposure_key in tier1:
        target[exposure_key] = injury_spelling[injury_key]
        injury_status[injury_key] = exposure_status[exposure_key] = "tier1_exact_shared_original"
    injury_remaining -= {item[0] for item in tier1}
    exposure_remaining -= {item[1] for item in tier1}
    counts["tier1_pairs"] = len(tier1)

    grouped_i: dict[str, list[str]] = defaultdict(list)
    grouped_e: dict[str, list[str]] = defaultdict(list)
    for key in injury_remaining:
        grouped_i[_loose_original(original(key))].append(key)
    for key in exposure_remaining:
        grouped_e[_loose_original(original(key))].append(key)
    injury_collision: dict[str, str] = {}
    exposure_collision: dict[str, str] = {}
    for group in sorted(set(grouped_i) | set(grouped_e)):
        members_i, members_e = grouped_i.get(group, []), grouped_e.get(group, [])
        if len(members_i) <= 1 and len(members_e) <= 1:
            continue
        collision_hash = stable_uid("idq", CROSSWALK_SHA256, team_key, group)
        for key in members_i:
            injury_status[key] = "collision_quarantined"
            injury_collision[key] = collision_hash
        for key in members_e:
            exposure_status[key] = "collision_quarantined"
            exposure_collision[key] = collision_hash
        counts["collision_groups"] += 1

    tier2 = [
        (values[0], grouped_e[group][0])
        for group, values in grouped_i.items()
        if len(values) == 1 and len(grouped_e.get(group, [])) == 1
    ]
    for injury_key, exposure_key in tier2:
        target[exposure_key] = injury_spelling[injury_key]
        injury_status[injury_key] = exposure_status[exposure_key] = "tier2_normalized_original"
    counts["tier2_pairs"] = len(tier2)
    counts["linked_pairs"] = counts["tier0_pairs"] + counts["tier1_pairs"] + counts["tier2_pairs"]
    if counts["linked_pairs"] != LINK_EXPECTED[team_key]["pairs"]:
        raise ValueError(f"{team_key}: tiered cross-file linkage pair-count drift")
    return TeamLinkage(target, injury_status, exposure_status, injury_collision, exposure_collision, dict(counts))


def _load_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text())
    if not isinstance(value, dict):
        raise ValueError(f"expected JSON object: {path.name}")
    return value


def _decision_output(team_key: str) -> tuple[dict[str, Any], Path, Path]:
    pointer = next(
        (path for path in (FINAL_DECISION_APPLICATIONS, COHORT_DECISION_APPLICATIONS,
                           REVISED_DECISION_APPLICATIONS, BASE_DECISION_APPLICATIONS)
         if path.exists()),
        BASE_DECISION_APPLICATIONS,
    )
    record = _load_json(pointer)
    if record.get("approval_granted") is not False:
        raise ValueError("Welsh decision-application record must not grant profile approval")
    matches = [item for item in record.get("outputs", []) if item.get("team_key") == team_key]
    if len(matches) != 1 or matches[0].get("validation_status") != "PASS":
        raise ValueError(f"{team_key}: immutable decision-application output is unavailable")
    output = matches[0]
    profile_path, mapping_path = ROOT / output["profile_path"], ROOT / output["mapping_path"]
    if sha256_file(profile_path) != output.get("profile_sha256") \
            or sha256_file(mapping_path) != output.get("mapping_sha256"):
        raise ValueError(f"{team_key}: immutable decision-application checksum drift")
    return output, profile_path, mapping_path


def _protected_received_key(team_key: str, source_sha: str, value: Any) -> str:
    return stable_uid("teamtok", team_key, source_sha, "Received/Injured In Team", _norm(value))


def _mapped_received_status(
    mapping: dict[str, Any], team_key: str, source_sha: str, value: Any
) -> tuple[str, str] | None:
    key = _protected_received_key(team_key, source_sha, value)
    matches = [
        entry for entry in mapping.get("mappings", [])
        if entry.get("canonical_field") == "received_in_team_status"
        and entry.get("source_evidence") == {"protected_source_token_key": key}
    ]
    if len(matches) > 1 or (matches and (
        matches[0].get("canonical_value") != "own_team"
        or matches[0].get("evidence_class") != "manual_adjudication"
        or matches[0].get("adjudication_id") != "welsh_dragons_received_cohort_20260714"
    )):
        raise ValueError(f"{team_key}: invalid protected received-team mapping")
    return (
        "own_team", "manual_adjudication:welsh_dragons_received_cohort_20260714"
    ) if matches else None


def _verify_team_bindings(team_key: str) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any], dict[str, Any], Path, Path]:
    team_dir = INTAKE / team_key
    decision_output, profile_path, mapping_path = _decision_output(team_key)
    plan_path = team_dir / PLAN_NAME
    profile, mapping, plan = map(_load_json, (profile_path, mapping_path, plan_path))
    canonical = canonical_team_name(team_key)
    if profile.get("team") != canonical or profile.get("season") != SEASON \
            or profile.get("decision") != "adapter_required" \
            or profile.get("approval_status") != "pending" \
            or profile.get("approved_by") is not None or profile.get("approved_at") is not None \
            or profile.get("approved_input_sha256s") != [] \
            or profile.get("unresolved_adjudication_ids") != []:
        raise ValueError(f"{team_key}: reviewed profile is not a pending adapter-required envelope")
    expected_mapping_path = os.path.relpath(mapping_path, team_dir)
    if profile.get("mapping_path") != expected_mapping_path \
            or profile.get("mapping_sha256") != sha256_file(mapping_path) \
            or mapping.get("mapping_version") != profile.get("mapping_version") \
            or mapping.get("team") != canonical or mapping.get("status") != "reviewed_pending_approval":
        raise ValueError(f"{team_key}: reviewed mapping/profile binding drift")
    if plan.get("adapter_plan_version") != ADAPTER_VERSION or plan.get("team_key") != team_key:
        raise ValueError(f"{team_key}: adapter-plan drift")
    source_dir = SOURCE_ROOT / CONFIG[team_key]["source_dir"]
    file_by_id = {item["id"]: source_dir / item["file"] for item in plan["source_inventory"]}
    if set(file_by_id) != set(plan["source_bindings"]):
        raise ValueError(f"{team_key}: adapter-plan source binding shape drift")
    for source_id, expected in plan["source_bindings"].items():
        if sha256_file(file_by_id[source_id]) != expected:
            raise ValueError(f"{team_key}: source checksum drift for {source_id}")
    return profile, mapping, plan, decision_output, profile_path, mapping_path


def _protected_values() -> set[str]:
    aliases = _load_json(TEAM_ALIAS_MAP_PATH).get("fixture_team_aliases")
    if not isinstance(aliases, dict) or not aliases:
        raise ValueError("protected team-alias map is unavailable")
    return {_norm(value) for value in aliases.values() if _norm(value)}


def _sanitize(row: dict[str, Any], protected: set[str], crosswalk: Crosswalk) -> tuple[dict[str, str], int]:
    output: dict[str, str] = {}
    redacted = 0
    for field, raw in row.items():
        value = clean_cell(raw)
        normalized = _norm(value)
        if normalized in protected or re.fullmatch(r"team [a-z]", normalized):
            value, redacted = REDACTED, redacted + 1
        elif normalized in crosswalk.direct_only:
            raise ValueError("direct player identifier found outside the adapter identifier boundary")
        output[field] = value
    return output, redacted


def _mapping(mapping: dict[str, Any], field: str, row: dict[str, str]) -> tuple[str, str]:
    matches = []
    for entry in mapping.get("mappings", []):
        if entry.get("canonical_field") != field or entry.get("evidence_source_id") != "injury":
            continue
        evidence = {**entry.get("source_evidence", {}), **entry.get("supporting_evidence", {})}
        if evidence and all(clean_cell(row.get(key)) == str(value) for key, value in evidence.items()):
            matches.append(entry)
    specificity = max((len(item.get("source_evidence", {})) + len(item.get("supporting_evidence", {})) for item in matches), default=0)
    matches = [item for item in matches if len(item.get("source_evidence", {})) + len(item.get("supporting_evidence", {})) == specificity]
    targets = {entry["canonical_value"] for entry in matches}
    if len(targets) > 1:
        raise ValueError(f"conflicting reviewed mappings for {field}")
    if not matches:
        return "unknown", "approved_mapping:no_supported_evidence"
    chosen = matches[0]
    adjudication = chosen.get("adjudication_id")
    origin = f"manual_adjudication:{adjudication}" if adjudication else f"approved_mapping:{chosen['evidence_class']}"
    return chosen["canonical_value"], origin


def _valid_date(value: Any) -> datetime | None:
    return parse_flexible_date(value, "day-first")


def _valid_days(value: Any) -> int | None:
    if value is None or isinstance(value, (bool, datetime)):
        return None
    text = clean_cell(value)
    if not text or ":" in text or "/" in text or "-" in text[1:]:
        return None
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    if number != number.to_integral_value() or number < 0 or number >= 10000:
        return None
    return int(number)


def _template_nonrecord(row: dict[str, Any]) -> bool:
    days = _valid_days(row.get("Days Injured"))
    evidence_fields = [
        "PlayerID", "Date Injured", "Fit For Selection Date", "Confirmed Return Date",
        "Body Part", "Orchard Code", "Problem type", "Occasion category", "Description",
    ]
    return days == 0 and not any(clean_cell(row.get(field)) for field in evidence_fields)


def _fixture_dates(team_key: str) -> set[str]:
    path = ROOT / URC_FIXTURES_2024_25_CORRECTED_PATH
    if sha256_file(path) != URC_FIXTURES_2024_25_CORRECTED_SHA256:
        raise ValueError("frozen URC fixture checksum drift")
    aliases = _load_json(TEAM_ALIAS_MAP_PATH)["fixture_team_aliases"]
    team_alias = aliases.get(TEAM_KEY_ALIAS_LOOKUP_NAMES[team_key])
    if not team_alias:
        raise ValueError(f"{team_key}: protected fixture alias unavailable")
    by_date: dict[str, set[tuple[str, str]]] = defaultdict(set)
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            pair = (row["home_team_alias"].strip(), row["away_team_alias"].strip())
            if team_alias in pair:
                by_date[row["corrected_date"].strip()].add(pair)
    return {date_value for date_value, fixtures in by_date.items() if len(fixtures) == 1}


def _source_hash(row: dict[str, Any]) -> str:
    payload = {key: clean_cell(value) for key, value in row.items()}
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode()).hexdigest()


def _raw_return_index(
    team_key: str,
    config: dict[str, Any],
    crosswalk: Crosswalk,
    injury_pseudonyms: set[str],
) -> tuple[dict[tuple[str, str, str], tuple[int, datetime]], str]:
    raw_name = config.get("raw_injury")
    if not raw_name:
        return {}, ""
    raw_path = SOURCE_ROOT / config["source_dir"] / raw_name
    _, physical_rows, _ = _read_xlsx(raw_path, "Injury Records")
    groups: dict[tuple[str, str, str], list[tuple[int, datetime]]] = defaultdict(list)
    for row_number, row in physical_rows:
        if not clean_cell(row.get("Full name and DoB")):
            continue
        try:
            identity = crosswalk.resolve(row.get("Full name and DoB"), team_key=team_key)
        except ValueError:
            targets = crosswalk.original_to_new.get(_norm(row.get("Full name and DoB")), set()) & injury_pseudonyms
            if len(targets) != 1:
                continue
            identity = Identity(crosswalk.new_spelling[next(iter(targets))], "team_scoped_exact_target", "")
        injured = _valid_date(row.get("Date of First Symptoms\n dd/mm/yy"))
        returned = _valid_date(row.get("Date of return to availability for selection\n dd/mm/yy"))
        if injured and returned and returned >= injured:
            groups[(_norm(identity.pseudonym), injured.date().isoformat(), returned.date().isoformat())].append((row_number, returned))
    return ({key: values[0] for key, values in groups.items() if len(values) == 1}, sha256_file(raw_path))


def _relative(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def _validate_no_identifiers(path: Path, crosswalk: Crosswalk, protected: set[str]) -> None:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            if clean_cell(row.get("DOB")):
                raise ValueError(f"{path.name}: DOB remains populated")
            for value in row.values():
                normalized = _norm(value)
                if normalized in crosswalk.direct_only or normalized in protected \
                        or re.fullmatch(r"team [a-z]", normalized):
                    raise ValueError(f"{path.name}: protected identifier or alias remains")


def generate_team(team_key: str, *, output_root: Path = INTAKE) -> dict[str, Any]:
    if team_key not in CONFIG:
        raise ValueError(f"unsupported Welsh team: {team_key}")
    team_dir = output_root / team_key
    existing_manifest_path = team_dir / "intake_manifest.json"
    if existing_manifest_path.exists():
        existing_manifest = _load_json(existing_manifest_path)
        existing_profile = existing_manifest.get("intake_profile", {})
        if existing_profile.get("approved_by") or existing_profile.get("approved_at") \
                or existing_manifest.get("profile_approval_application"):
            raise ValueError(
                f"{team_key}: approved intake manifest is immutable; generate into a distinct empty output root"
            )
    profile, mapping, plan, decision_output, profile_path, mapping_path = _verify_team_bindings(team_key)
    crosswalk, protected = Crosswalk(), _protected_values()
    config, canonical = CONFIG[team_key], canonical_team_name(team_key)
    source_dir = SOURCE_ROOT / config["source_dir"]
    injury_path = source_dir / config["injury"]
    exposure_path = source_dir / config["exposure"]
    raw_exposure_path = source_dir / config["raw_exposure"]
    injury_headers, physical_injury, injury_physical_count = _read_xlsx(injury_path)
    exposure_headers, physical_exposure, exposure_physical_count = _read_xlsx(exposure_path)
    with raw_exposure_path.open(newline="", encoding="utf-8-sig") as handle:
        raw_exposure = list(csv.DictReader(handle))
    substantive_injury = [row for _, row in physical_injury if any(clean_cell(v) for v in row.values()) and not _template_nonrecord(row)]
    substantive_exposure = [row for _, row in physical_exposure if any(clean_cell(v) for v in row.values())]
    if len(raw_exposure) != len(substantive_exposure):
        raise ValueError(f"{team_key}: raw/standard exposure row-count drift")
    linkage = _build_team_linkage(team_key, substantive_injury, substantive_exposure, crosswalk)

    fixture_dates = _fixture_dates(team_key)
    own_team_alias = load_fixture_team_aliases()[TEAM_KEY_ALIAS_LOOKUP_NAMES[team_key]]
    injury_pseudonyms = {
        _norm(crosswalk.resolve(row.get("PlayerID"), team_key=team_key).pseudonym)
        for row in substantive_injury
    }
    raw_return_by_key, raw_return_hash = _raw_return_index(
        team_key, config, crosswalk, injury_pseudonyms
    )
    injury_rows: list[dict[str, Any]] = []
    injury_events: list[dict[str, Any]] = []
    collision_rows: list[dict[str, Any]] = []
    injury_counts = Counter(physical_rows=injury_physical_count)
    injury_source_hash = sha256_file(injury_path)
    raw_received_tokens = {
        _norm(row.get("Received/Injured In Team"))
        for row in substantive_injury
        if clean_cell(row.get("Received/Injured In Team"))
    }
    for source_row, raw in physical_injury:
        if not any(clean_cell(value) for value in raw.values()):
            injury_counts["blank_physical_rows"] += 1
            continue
        if _template_nonrecord(raw):
            injury_counts["structural_zero_nonrecords"] += 1
            injury_events.append({"source_row_number": source_row, "action": "reconcile_nonrecord", "reason": "structural_zero_template_row"})
            continue
        identity = crosswalk.resolve(raw.get("PlayerID"), team_key=team_key)
        identity_key = _norm(identity.pseudonym)
        safe_raw = dict(raw)
        safe_raw["PlayerID"] = identity.pseudonym
        safe_raw["DOB"] = ""
        raw_received = clean_cell(raw.get("Received/Injured In Team"))
        # Team/squad labels are protected metadata. Classify on the raw value in
        # memory, then serialize only a constant marker (or a blank).
        safe_raw["Received/Injured In Team"] = REDACTED if raw_received else ""
        mapped_received = _mapped_received_status(
            mapping, team_key, injury_source_hash, raw_received
        )
        if mapped_received is None and _norm(raw_received) == _norm(own_team_alias):
            mapped_received = ("own_team", "approved_mapping:exact_current_team_alias")
        row, redacted = _sanitize(safe_raw, protected, crosswalk)
        if raw_received:
            redacted += 1
        if mapped_received:
            received_status, received_origin = mapped_received
            row["Adapter Canonical Received In Team Status"] = received_status
            row["Adapter Canonical Received In Team Status Origin"] = received_origin
        else:
            row["Adapter Canonical Received In Team Status"] = ""
            row["Adapter Canonical Received In Team Status Origin"] = ""
            received_status, _ = received_in_team_status(row, own_team_alias)
        injury_counts["protected_values_redacted"] += redacted
        injured, source_return = _valid_date(raw.get("Date Injured")), _valid_date(raw.get("Confirmed Return Date"))
        fit_return = _valid_date(raw.get("Fit For Selection Date"))
        raw_return = raw_return_by_key.get(
            (identity_key, injured.date().isoformat(), fit_return.date().isoformat())
        ) if injured and fit_return else None
        accepted_return = source_return
        return_origin = "approved_mapping:source_confirmed_return_date"
        return_reference_hash = injury_source_hash
        return_reference_row = source_row
        if accepted_return is None and raw_return and fit_return and raw_return[1].date() == fit_return.date():
            return_reference_row, accepted_return = raw_return
            return_origin = "approved_mapping:raw_return_to_availability_exact_row"
            return_reference_hash = raw_return_hash
        elif accepted_return is None and fit_return:
            accepted_return = fit_return
            return_origin = "approved_mapping:standardized_return_to_availability_exact_row"
        row["Date Injured"] = injured.strftime("%d/%m/%Y") if injured else clean_cell(raw.get("Date Injured"))
        if accepted_return and injured and accepted_return >= injured:
            row["Confirmed Return Date"] = clean_cell(raw.get("Confirmed Return Date"))
            row["Fit For Selection Date"] = accepted_return.strftime("%d/%m/%Y")
            row["Adapter Canonical Confirmed Return Date"] = accepted_return.strftime("%d/%m/%Y")
            row["Adapter Canonical Confirmed Return Date Origin"] = return_origin
            row["return_reference_file_sha256"] = return_reference_hash
            row["return_reference_sheet"] = "Injury Records" if return_reference_hash == raw_return_hash and raw_return_hash else SHEET
            row["return_reference_row_number"] = str(return_reference_row)
            injury_counts["valid_return_dates_restored"] += 1
            if return_origin.endswith("raw_return_to_availability_exact_row"):
                injury_counts["raw_return_dates_restored"] += 1
            elif return_origin.endswith("standardized_return_to_availability_exact_row"):
                injury_counts["standardized_return_dates_restored"] += 1
        else:
            row["Confirmed Return Date"] = clean_cell(raw.get("Confirmed Return Date"))
            row["Adapter Canonical Confirmed Return Date"] = ""
            row["Adapter Canonical Confirmed Return Date Origin"] = ""
            row["return_reference_file_sha256"] = ""
            row["return_reference_sheet"] = ""
            row["return_reference_row_number"] = ""
            if accepted_return:
                injury_counts["invalid_return_dates_rejected"] += 1
        days = _valid_days(raw.get("Days Injured"))
        if days is not None:
            row["Days Injured"] = str(days)
            row["Adapter Days Injured Origin"] = "source_reported_days_injured"
            injury_counts["source_days_preserved"] += 1
        elif injured and accepted_return and accepted_return >= injured:
            row["Days Injured"] = ""
            row["Adapter Days Injured Origin"] = "shared_processing_derivation_from_accepted_return_date"
            injury_counts["days_derived_from_ordered_dates"] += 1
        else:
            row["Days Injured"] = ""
            row["Adapter Days Injured Origin"] = "unknown"
            if clean_cell(raw.get("Days Injured")):
                injury_counts["invalid_source_days_rejected"] += 1

        for canonical_field, (value_field, origin_field) in OVERRIDES.items():
            value, origin = _mapping(mapping, canonical_field, row)
            row[value_field], row[origin_field] = value, origin
        if row[OVERRIDES["problem_type"][0]] == "illness":
            row[OVERRIDES["body_location"][0]] = "unknown"
            row[OVERRIDES["body_location"][1]] = "approved_mapping:not_applicable_to_illness"
            row[OVERRIDES["tissue_pathology"][0]] = "unknown"
            row[OVERRIDES["tissue_pathology"][1]] = "approved_mapping:not_applicable_to_illness"
            injury_counts["medical_illness_rows"] += 1
        if team_key == "dragons":
            row[OVERRIDES["tissue_pathology"][0]] = "unknown"
            row[OVERRIDES["tissue_pathology"][1]] = "approved_mapping:ref_error_no_tissue_evidence"
        match_value, _ = _mapping(mapping, "match_type", row)
        if match_value == "training":
            row["Match Type"] = "training"
        elif clean_cell(row.get("Match Type")).casefold() != "urc":
            date_iso = injured.date().isoformat() if injured else ""
            row["Match Type"] = "URC" if row[OVERRIDES["occasion_category"][0]] == "match" and date_iso in fixture_dates else ""
            if row["Match Type"]:
                injury_counts["unique_fixture_urc_rows"] += 1

        row.update({
            "source_archive_path": str(injury_path), "source_file_sha256": injury_source_hash,
            "source_sheet": SHEET, "source_row_number": str(source_row),
            "standardised_file_sha256": injury_source_hash, "standardised_row_number": str(source_row),
            "source_locator_status": "provisional_reference_locator",
            "source_row_sha256": _source_hash(raw),
            "player_uid": stable_uid("ply", canonical, identity.pseudonym),
            "injury_uid": stable_uid("inj", canonical, SEASON, injury_source_hash, source_row),
            "identity_link_status": linkage.injury_status[identity_key],
            "identity_collision_key_hash": linkage.injury_collision_hash.get(identity_key, ""),
        })
        if identity_key in linkage.injury_collision_hash:
            collision_rows.append({"domain": "injury", "key_hash": linkage.injury_collision_hash[identity_key],
                                   "source_file_sha256": injury_source_hash, "source_sheet": SHEET,
                                   "source_row_number": source_row})
        injury_rows.append(row)
        injury_counts[f"received_status_{received_status}"] += 1
    if len(injury_rows) != config["retained_injury_rows"] \
            or injury_counts["structural_zero_nonrecords"] != config["placeholder_rows"] \
            or injury_counts["medical_illness_rows"] != MEDICAL_EXPECTED[team_key]:
        raise ValueError(f"{team_key}: injury row/adjudication reconciliation drift")
    if injury_counts["received_status_other_team"] == len(injury_rows):
        raise ValueError(f"{team_key}: unsupported received-team mapping would exclude the whole cohort")
    if team_key == "dragons" and (
        injury_counts["received_status_own_team"] != 186
        or injury_counts["received_status_other_team"] != 18
    ):
        raise ValueError("dragons: reviewed received-team mapping count drift")
    if any(
        _norm(value) in raw_received_tokens
        for row in injury_rows for value in row.values()
        if clean_cell(value)
    ):
        raise ValueError(f"{team_key}: protected received-team token serialized")
    if any(
        row.get("Received/Injured In Team") not in {"", REDACTED}
        for row in injury_rows
    ):
        raise ValueError(f"{team_key}: received-team source field was not unconditionally redacted")

    injury_output = team_dir / f"{team_key}_injury_intake_locator_enriched_2024-25.csv"
    injury_extra = [item for pair in OVERRIDES.values() for item in pair] + [
        "Adapter Canonical Confirmed Return Date", "Adapter Canonical Confirmed Return Date Origin",
        "Adapter Canonical Received In Team Status", "Adapter Canonical Received In Team Status Origin",
        "Adapter Days Injured Origin", "return_reference_file_sha256", "return_reference_sheet",
        "return_reference_row_number",
        "source_archive_path", "source_file_sha256", "source_sheet", "source_row_number",
        "standardised_file_sha256", "standardised_row_number", "source_locator_status",
        "source_row_sha256", "player_uid", "injury_uid", "identity_link_status",
        "identity_collision_key_hash",
    ]
    _csv_write(injury_output, injury_rows, injury_headers + [field for field in injury_extra if field not in injury_headers])

    exposure_rows: list[dict[str, Any]] = []
    exposure_counts = Counter(physical_rows=exposure_physical_count)
    exposure_source_hash, raw_exposure_hash = sha256_file(exposure_path), sha256_file(raw_exposure_path)
    substantive = [(number, row) for number, row in physical_exposure if any(clean_cell(v) for v in row.values())]
    comparisons = [("date", "session date"), ("week", "Week"), ("competition", "Competition"),
                   ("session", "session type"), ("duration", "minutes total"),
                   ("distance_m", "distance total"), ("distance_high_speed", "sprint distance"),
                   ("distance_very_high_speed", "high speed running minutes")]
    for (source_row, raw), raw_reference in zip(substantive, raw_exposure, strict=True):
        def comparable(value: Any) -> str:
            text = clean_cell(value)
            try:
                return format(Decimal(text).normalize(), "f") if text else ""
            except InvalidOperation:
                return text
        if any(comparable(raw_reference[a]) != comparable(raw[b]) for a, b in comparisons):
            raise ValueError(f"{team_key}: exposure safe-field physical-row reconciliation drift")
        identity = crosswalk.resolve(raw.get("name"), team_key=team_key)
        identity_key = _norm(identity.pseudonym)
        target_pseudonym = linkage.exposure_target.get(identity_key, identity.pseudonym)
        safe_raw = dict(raw)
        safe_raw["name"] = target_pseudonym
        row, redacted = _sanitize(safe_raw, protected, crosswalk)
        exposure_counts["protected_values_redacted"] += redacted
        row["high speed running distance"] = clean_cell(raw_reference["distance_high_speed"])
        row["very high speed running distance"] = clean_cell(raw_reference["distance_very_high_speed"])
        row.update({
            "minutes_total_origin": "source_reported" if clean_cell(raw.get("minutes total")) else "missing",
            "distance_total_origin": "source_reported" if clean_cell(raw.get("distance total")) else "missing",
            "high_speed_running_distance_origin": "source_preserved_only",
            "very_high_speed_running_distance_origin": "source_preserved_only",
            "restored_high_speed_metric_comparability": "non_comparable_unknown_vendor_threshold",
            "restored_high_speed_metrics_analysis_eligible": "false",
            "source_archive_path": str(exposure_path), "source_file_sha256": exposure_source_hash,
            "source_sheet": SHEET, "source_row_number": str(source_row),
            "source_row_sha256": _source_hash(raw), "standardised_file_sha256": exposure_source_hash,
            "standardised_row_number": str(source_row), "source_locator_status": "provisional_reference_locator",
            "hsr_reference_file_sha256": raw_exposure_hash, "hsr_reference_row_number": str(source_row),
            "player_uid": stable_uid("ply", canonical, target_pseudonym),
            EXPOSURE_DECLARED_GRAIN_FIELD: "session", "identity_link_status": linkage.exposure_status[identity_key],
            "identity_collision_key_hash": linkage.exposure_collision_hash.get(identity_key, ""),
        })
        if _norm(target_pseudonym) != identity_key:
            exposure_counts["identity_remapped_rows"] += 1
        if identity_key in linkage.exposure_collision_hash:
            collision_rows.append({"domain": "exposure", "key_hash": linkage.exposure_collision_hash[identity_key],
                                   "source_file_sha256": exposure_source_hash, "source_sheet": SHEET,
                                   "source_row_number": source_row})
        exposure_rows.append(row)
    if len(exposure_rows) != plan["exposure_reconciliation"]["raw_rows"]:
        raise ValueError(f"{team_key}: exposure output row reconciliation drift")
    if exposure_counts["identity_remapped_rows"] != LINK_EXPECTED[team_key]["remapped_rows"]:
        raise ValueError(f"{team_key}: tiered exposure remap row-count drift")
    exposure_counts["rows"] = len(exposure_rows)
    exposure_counts["hsr_rows_restored"] = len(exposure_rows)
    exposure_counts["vhsr_rows_restored"] = len(exposure_rows)

    prepared_output = team_dir / f"{team_key}_exposure_intake_locator_enriched_2024-25.csv"
    exposure_extra = [
        "minutes_total_origin", "distance_total_origin", "high_speed_running_distance_origin",
        "very_high_speed_running_distance_origin", "restored_high_speed_metric_comparability",
        "restored_high_speed_metrics_analysis_eligible", *EXPOSURE_LOCATOR_FIELDS,
        "hsr_reference_file_sha256", "hsr_reference_row_number", "player_uid",
        EXPOSURE_DECLARED_GRAIN_FIELD, "identity_link_status", "identity_collision_key_hash",
    ]
    exposure_fields = exposure_headers + [field for field in exposure_extra if field not in exposure_headers]
    _csv_write(prepared_output, exposure_rows, exposure_fields)

    manifest_path = team_dir / "intake_manifest.json"
    manifest = {
        "team": canonical, "season": SEASON, "generator_version": GENERATOR_VERSION,
        "adapter_plan_path": PLAN_NAME, "adapter_plan_sha256": sha256_file(INTAKE / team_key / PLAN_NAME),
        "locator_enriched_intake_file": _relative(injury_output),
        "locator_enriched_file_sha256": sha256_file(injury_output),
        "locator_enriched_row_count": len(injury_rows), "locator_fields": LOCATOR_FIELDS,
        "uid_fields": ["player_uid", "injury_uid"],
        "injury_adapter_qc_file": _relative(team_dir / f"{team_key}_injury_adapter_qc_2024-25.json"),
        "exposure_intake": {
            "team": canonical, "season": SEASON, "source_workbook": str(exposure_path),
            "source_workbook_sha256": exposure_source_hash, "source_sheet": SHEET,
            "locator_enriched_intake_file": _relative(prepared_output),
            "locator_enriched_file_sha256": sha256_file(prepared_output),
            "locator_enriched_row_count": len(exposure_rows), "locator_fields": EXPOSURE_LOCATOR_FIELDS,
            "uid_fields": ["player_uid"], "source_locator_status": "provisional_reference_locator",
            "exposure_reporting_grain": {"current_file_reporting_grain": "session", "selection_source": "required_adapter_contract"},
            "restored_high_speed_metric_policy": plan["restored_high_speed_metric_policy"],
        },
        "intake_profile": {
            key: profile.get(key) for key in (
                "team", "season", "profile_version", "decision", "mapping_path", "mapping_sha256",
                "mapping_version", "ai_review_status", "ai_reviewed_by", "ai_reviewed_at", "approved_by",
                "approved_at", "unresolved_adjudication_ids", "approved_input_sha256s",
            )
        },
    }
    manifest["intake_profile"].update({
        "profile_path": os.path.relpath(profile_path, INTAKE / team_key),
        "profile_sha256": decision_output["profile_sha256"],
    })
    _json_write(manifest_path, manifest)

    injury_qc_path = team_dir / f"{team_key}_injury_adapter_qc_2024-25.json"
    exposure_qc_path = team_dir / f"{team_key}_exposure_adapter_qc_2024-25.json"
    _json_write(injury_qc_path, {
        "rule_version": GENERATOR_VERSION, "team": canonical, "season": SEASON,
        "source_file_sha256": injury_source_hash, "output_file": _relative(injury_output),
        "output_file_sha256": sha256_file(injury_output), "counts": dict(injury_counts),
        "row_reconciliation": {"physical_rows": injury_physical_count, "blank_rows": injury_counts["blank_physical_rows"],
                               "structural_nonrecords": injury_counts["structural_zero_nonrecords"], "candidate_rows": len(injury_rows)},
        "events": injury_events, "privacy": {"direct_identifier_values": 0, "dob_values": 0, "identifying_values_logged": False},
    })
    _json_write(exposure_qc_path, {
        "rule_version": GENERATOR_VERSION, "team": canonical, "season": SEASON,
        "source_file_sha256": exposure_source_hash, "raw_reference_file_sha256": raw_exposure_hash,
        "output_file": _relative(prepared_output), "output_file_sha256": sha256_file(prepared_output),
        "counts": dict(exposure_counts), "row_reconciliation": plan["exposure_reconciliation"],
        "privacy": {"direct_identifier_values": 0, "identifying_values_logged": False},
    })
    identity_qc_path = team_dir / f"{team_key}_identity_linkage_qc_2024-25.json"
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in collision_rows:
        grouped[item["key_hash"]].append({key: value for key, value in item.items() if key != "key_hash"})
    _json_write(identity_qc_path, {
        "rule_version": GENERATOR_VERSION, "team": canonical, "crosswalk_sha256": CROSSWALK_SHA256,
        "exact_one_to_one_only": True, "linkage_counts": linkage.counts,
        "identity_remapped_exposure_rows": exposure_counts["identity_remapped_rows"], "collision_groups": [
            {"key_hash": key, "source_locators": values} for key, values in sorted(grouped.items())
        ], "direct_identifiers_serialized": False,
    })

    cleaned_output = team_dir / f"{team_key}_exposure_cleaned_2024-25.csv"
    cleaning_qc = team_dir / f"{team_key}_exposure_cleaning_qc_2024-25.json"
    args = argparse.Namespace(file=str(prepared_output), output=str(cleaned_output), qc_output=str(cleaning_qc),
                              manifest=str(manifest_path), reporting_grain="session", team=team_key, season=SEASON,
                              date_order="day-first", window_start=WINDOW_START, window_end=WINDOW_END)
    with contextlib.redirect_stdout(io.StringIO()):
        clean_exposure(args)
    for path in (injury_output, prepared_output, cleaned_output):
        _validate_no_identifiers(path, crosswalk, protected)
    manifest = _load_json(manifest_path)
    manifest["injury_adapter_qc_sha256"] = sha256_file(injury_qc_path)
    manifest["exposure_intake"]["qc_file"] = _relative(exposure_qc_path)
    manifest["exposure_intake"]["qc_file_sha256"] = sha256_file(exposure_qc_path)
    manifest["identity_linkage_qc_file"] = _relative(identity_qc_path)
    manifest["identity_linkage_qc_sha256"] = sha256_file(identity_qc_path)
    manifest["generation_validation"] = {
        "status": "PASS", "profile_approval_granted": False, "database_action": False,
        "direct_identifiers_remaining": 0, "protected_aliases_remaining": 0,
        "injury_rows": len(injury_rows), "exposure_rows": len(exposure_rows),
        "crosswalk_sha256": CROSSWALK_SHA256,
    }
    _json_write(manifest_path, manifest)
    return {
        "team": canonical, "injury_rows": len(injury_rows), "exposure_rows": len(exposure_rows),
        "injury_sha256": sha256_file(injury_output), "prepared_exposure_sha256": sha256_file(prepared_output),
        "cleaned_exposure_sha256": sha256_file(cleaned_output), "manifest_sha256": sha256_file(manifest_path),
        "collision_groups": len(grouped), "medical_illness_rows": injury_counts["medical_illness_rows"],
    }


def generate_all(*, output_root: Path = INTAKE) -> list[dict[str, Any]]:
    return [generate_team(team_key, output_root=output_root) for team_key in CONFIG]


def create_dragons_cohort_revision(*, reviewed_at: str) -> dict[str, Any]:
    """Version the aggregate-reviewed Dragons protected-token cohort mapping."""
    reviewed_time = datetime.fromisoformat(reviewed_at.replace("Z", "+00:00"))
    if reviewed_time.tzinfo is None or reviewed_time > datetime.now(reviewed_time.tzinfo):
        raise ValueError("cohort review timestamp must be timezone-aware and not future")
    base = _load_json(BASE_DECISION_APPLICATIONS)
    matches = [item for item in base["outputs"] if item["team_key"] == "dragons"]
    if len(matches) != 1:
        raise ValueError("base Dragons decision output is unavailable")
    output = matches[0]
    old_mapping_path, old_profile_path = ROOT / output["mapping_path"], ROOT / output["profile_path"]
    if sha256_file(old_mapping_path) != output["mapping_sha256"] or sha256_file(old_profile_path) != output["profile_sha256"]:
        raise ValueError("base Dragons decision output checksum drift")
    config = CONFIG["dragons"]
    source_path = SOURCE_ROOT / config["source_dir"] / config["injury"]
    source_sha = sha256_file(source_path)
    _, rows, _ = _read_xlsx(source_path)
    frequencies = Counter(
        _norm(row.get("Received/Injured In Team"))
        for _, row in rows if clean_cell(row.get("Received/Injured In Team"))
    )
    if len(frequencies) != 7 or not frequencies or frequencies.most_common(1)[0][1] != 186:
        raise ValueError("Dragons protected-token aggregate evidence drift")
    dominant = frequencies.most_common(1)[0][0]
    protected_key = _protected_received_key("dragons", source_sha, dominant)
    mapping = _load_json(old_mapping_path)
    mapping["mapping_version"] = "source_to_canonical_mapping_v3"
    entry = {
        "canonical_field": "received_in_team_status", "canonical_value": "own_team",
        "evidence_class": "manual_adjudication",
        "source_evidence": {"protected_source_token_key": protected_key},
        "supporting_evidence": {"source_workbook_sha256": source_sha, "source_row_count": "186"},
        "evidence_source_id": "injury", "evidence_sheet": SHEET,
        "specificity_change": "equivalent", "protocol_rule_id": "team_scoped_protected_token_v1",
        "adjudication_id": "welsh_dragons_received_cohort_20260714",
        "rule": "Abdel-approved manual adjudication: map only the exact hashed Dragons team-scoped legacy token to own_team; all other tokens retain frozen default handling.",
    }
    if not any(item.get("canonical_field") == "received_in_team_status" for item in mapping["mappings"]):
        mapping["mappings"].append(entry)
    fingerprint = hashlib.sha256(
        (output["profile_sha256"] + protected_key + reviewed_at).encode()
    ).hexdigest()
    run_id = f"dragons-cohort-{reviewed_time.astimezone(UTC).strftime('%Y%m%dT%H%M%S%z')}-{fingerprint[:12]}"
    run_dir = INTAKE / "dragons" / "reviewed" / run_id
    mapping_path = run_dir / "source_to_canonical_mapping.v3.json"
    _immutable_write(mapping_path, (json.dumps(mapping, indent=2, sort_keys=True) + "\n").encode())
    profile = _load_json(old_profile_path)
    profile.update({
        "profile_version": "team_intake_profile_v3",
        "mapping_version": "source_to_canonical_mapping_v3",
        "mapping_path": os.path.relpath(mapping_path, INTAKE / "dragons"),
        "mapping_sha256": sha256_file(mapping_path),
        "ai_reviewed_by": "Codex fresh Welsh protected cohort-signal review with Abdel-approved manual adjudication",
        "ai_reviewed_at": reviewed_at,
    })
    profile.setdefault("ai_review", {}).setdefault("findings", []).append({
        "finding": "The Dragons standard field contains seven protected legacy team/squad tokens; the adjudicated token occurs in 186/204 rows, appears only in Dragons across eight comparable team workbooks, and is dominant only there.",
        "disposition": "Abdel-approved manual adjudication maps only that exact team-scoped token to own_team by salted hash; the remaining 18 rows retain frozen non-own handling.",
        "status": "resolved",
    })
    profile_path = run_dir / "team_intake_profile.v3.json"
    _immutable_write(profile_path, (json.dumps(profile, indent=2, sort_keys=True) + "\n").encode())
    revised = json.loads(json.dumps(base))
    revised["schema_version"] = "welsh_step0_decision_application_v4"
    revised["run_id"] = run_id
    revised["applied_at"] = reviewed_at
    for item in revised["outputs"]:
        if item["team_key"] == "dragons":
            item.update({
                "mapping_path": str(mapping_path.relative_to(ROOT)),
                "mapping_sha256": sha256_file(mapping_path),
                "profile_path": str(profile_path.relative_to(ROOT)),
                "profile_sha256": sha256_file(profile_path),
                "validation_status": "PASS",
            })
    _immutable_write(
        FINAL_DECISION_APPLICATIONS,
        (json.dumps(revised, indent=2, sort_keys=True) + "\n").encode(),
    )
    return {
        "run_id": run_id, "profile_sha256": sha256_file(profile_path),
        "mapping_sha256": sha256_file(mapping_path), "mapped_rows": 186,
        "default_other_team_rows": 18, "protected_values_serialized": False,
    }


def stamp_approval(
    team_key: str,
    *,
    approved_at: str,
    actor_evidence_path: Path,
    team_dir: Path | None = None,
) -> dict[str, Any]:
    """Create an immutable approved profile and atomically repoint the manifest.

    This is local-only. It never runs ingest or connects to the database.
    """
    if team_key not in CONFIG:
        raise ValueError(f"unsupported Welsh team: {team_key}")
    try:
        approved_time = datetime.fromisoformat(approved_at.replace("Z", "+00:00"))
    except ValueError as exc:
        raise ValueError("approval timestamp must be ISO-8601") from exc
    if approved_time.tzinfo is None:
        raise ValueError("approval timestamp must be timezone-aware")
    if not actor_evidence_path.is_file():
        raise ValueError("structured approval authorization JSON is required")

    profile, _, _, decision_output, reviewed_profile_path, reviewed_mapping_path = _verify_team_bindings(team_key)
    reviewed_at = datetime.fromisoformat(str(profile["ai_reviewed_at"]).replace("Z", "+00:00"))
    if approved_time < reviewed_at or approved_time > datetime.now(approved_time.tzinfo):
        raise ValueError("approval timestamp must follow AI review and not be in the future")
    directory = team_dir or INTAKE / team_key
    manifest_path = directory / "intake_manifest.json"
    manifest = _load_json(manifest_path)
    pending = manifest.get("intake_profile", {})
    if pending.get("approved_by") is not None or pending.get("approved_at") is not None:
        # Idempotence is handled after the deterministic approval id is known.
        if pending.get("approved_by") != "Abdel Babiker":
            raise ValueError("manifest already points to a different approval")

    injury_path = directory / f"{team_key}_injury_intake_locator_enriched_2024-25.csv"
    exposure_path = directory / f"{team_key}_exposure_cleaned_2024-25.csv"
    injury_sha, exposure_sha = sha256_file(injury_path), sha256_file(exposure_path)
    if manifest.get("locator_enriched_file_sha256") != injury_sha:
        raise ValueError("injury candidate checksum drift before approval")
    cleaning = manifest.get("exposure_cleaning", {})
    if cleaning.get("cleaned_file_sha256") != exposure_sha:
        raise ValueError("cleaned exposure checksum drift before approval")
    authorization = _load_json(actor_evidence_path)
    expected_authorization = {
        "schema_version": "welsh_profile_authorization_v1",
        "actor": "Abdel Babiker",
        "action": "approve_team_intake_profile",
        "decision": "approved",
        "team": canonical_team_name(team_key),
        "team_key": team_key,
        "season": SEASON,
        "approved_at": approved_at,
        "reviewed_profile_sha256": decision_output["profile_sha256"],
        "mapping_sha256": decision_output["mapping_sha256"],
        "injury_candidate_sha256": injury_sha,
        "cleaned_exposure_sha256": exposure_sha,
        "database_action": False,
    }
    for key, expected in expected_authorization.items():
        if authorization.get(key) != expected:
            raise ValueError(f"approval authorization mismatch: {key}")
    if team_key == "dragons" and authorization.get("manual_adjudication") != {
        "adjudication_id": "welsh_dragons_received_cohort_20260714",
        "own_team_rows": 186,
        "frozen_non_own_rows": 18,
        "mapping_version": "source_to_canonical_mapping_v3",
    }:
        raise ValueError("approval authorization mismatch: Dragons manual adjudication")
    actor_sha = sha256_file(actor_evidence_path)
    approval_fingerprint = hashlib.sha256(
        "\x1f".join((decision_output["profile_sha256"], injury_sha, exposure_sha, actor_sha, approved_at)).encode()
    ).hexdigest()
    run_id = f"approved-{approved_time.astimezone(UTC).strftime('%Y%m%dT%H%M%S%z')}-{approval_fingerprint[:12]}"
    run_dir = directory / "reviewed" / run_id
    mapping_path = run_dir / reviewed_mapping_path.name
    evidence_copy = run_dir / "approval_authorization.v1.json"
    profile_path = run_dir / reviewed_profile_path.name
    record_path = run_dir / "profile_approval_application.v1.json"
    _immutable_write(mapping_path, reviewed_mapping_path.read_bytes())
    _immutable_write(evidence_copy, actor_evidence_path.read_bytes())

    approved_profile = dict(profile)
    approved_profile.update({
        "mapping_path": os.path.relpath(mapping_path, directory),
        "mapping_sha256": sha256_file(mapping_path),
        "approval_status": "approved",
        "approved_by": "Abdel Babiker",
        "approved_at": approved_at,
        "approved_input_sha256s": [injury_sha, exposure_sha],
    })
    profile_bytes = (json.dumps(approved_profile, indent=2, sort_keys=True) + "\n").encode()
    _immutable_write(profile_path, profile_bytes)
    profile_sha = sha256_file(profile_path)
    record = {
        "schema_version": "welsh_profile_approval_application_v1",
        "team": canonical_team_name(team_key), "team_key": team_key, "season": SEASON,
        "approved_by": "Abdel Babiker", "approved_at": approved_at,
        "actor_evidence_path": os.path.relpath(evidence_copy, directory),
        "actor_evidence_sha256": actor_sha,
        "reviewed_profile_path": decision_output["profile_path"],
        "reviewed_profile_sha256": decision_output["profile_sha256"],
        "approved_profile_path": os.path.relpath(profile_path, directory),
        "approved_profile_sha256": profile_sha,
        "mapping_path": os.path.relpath(mapping_path, directory),
        "mapping_sha256": sha256_file(mapping_path),
        "approved_inputs": [
            {"kind": "injury", "path": str(injury_path), "sha256": injury_sha},
            {"kind": "cleaned_exposure", "path": str(exposure_path), "sha256": exposure_sha},
        ],
        "database_action": False,
    }
    _immutable_write(record_path, (json.dumps(record, indent=2, sort_keys=True) + "\n").encode())

    bound_fields = (
        "team", "season", "profile_version", "decision", "mapping_path", "mapping_sha256",
        "mapping_version", "ai_review_status", "ai_reviewed_by", "ai_reviewed_at", "approved_by",
        "approved_at", "unresolved_adjudication_ids", "approved_input_sha256s",
    )
    bound = {key: approved_profile.get(key) for key in bound_fields}
    bound.update({"profile_path": os.path.relpath(profile_path, directory), "profile_sha256": profile_sha})
    updated_manifest = dict(manifest)
    updated_manifest["intake_profile"] = bound
    updated_manifest["profile_approval_application"] = {
        "path": os.path.relpath(record_path, directory), "sha256": sha256_file(record_path)
    }
    # Validate the exact would-be persisted manifest against both approved inputs.
    validate_intake_profile_manifest(updated_manifest, manifest_path, injury_sha, canonical_team_name(team_key), SEASON)
    validate_intake_profile_manifest(updated_manifest, manifest_path, exposure_sha, canonical_team_name(team_key), SEASON)
    if pending.get("approved_by") is not None and manifest.get("intake_profile") != bound:
        raise ValueError("manifest already points to a different immutable approval")
    _json_write(manifest_path, updated_manifest)
    return {
        "team": canonical_team_name(team_key), "profile_path": bound["profile_path"],
        "profile_sha256": profile_sha, "mapping_sha256": bound["mapping_sha256"],
        "approved_input_sha256s": [injury_sha, exposure_sha],
        "approval_record_sha256": sha256_file(record_path), "database_action": False,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--team", choices=[*CONFIG, "all"], default="all")
    parser.add_argument("--stamp-approval", action="store_true")
    parser.add_argument("--approved-at", default="")
    parser.add_argument("--actor-evidence", default="")
    args = parser.parse_args()
    if args.stamp_approval:
        if args.team == "all" or not args.approved_at or not args.actor_evidence:
            raise SystemExit("approval stamping requires one --team, --approved-at, and --actor-evidence")
        results = [stamp_approval(
            args.team, approved_at=args.approved_at, actor_evidence_path=Path(args.actor_evidence)
        )]
    else:
        results = generate_all() if args.team == "all" else [generate_team(args.team)]
    print(json.dumps(results, indent=2))


if __name__ == "__main__":
    main()
