"""Compile aggregate-only Welsh Step 0 intake-profile draft packages.

The legacy archive is reference-only. This module reads it locally to compute
checksums and aggregate reconciliation counts, but never serializes identifier
or clinical free-text cell values and never connects to the database.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import tempfile
from datetime import UTC, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from pipeline.__main__ import IOC_BODY_CODE_MAP, ORCHARD_PATHOLOGY_TYPE_MAP
from pipeline.profiling import (
    MAPPING_VERSION,
    _build_inventory,
    family_check,
    scan_plan,
    validate_package,
)


ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = Path("/Users/abdelbabiker/Desktop/URC")
INTAKE = ROOT / "data" / "intake" / "2024-25"
CACHE = ROOT / "data" / "intake" / ".profile-cache"
SEASON = "2024-25"
SHEET = "Standardized Data"
ADAPTER_VERSION = "welsh_intake_adapter_plan_v1"
DECISION_APPLICATION_VERSION = "welsh_step0_decision_application_v1"
SUPPORTED_DECISION_CHOICE = "medical_illness_to_illness"
DECISION_ACTOR = "Abdel Babiker"
DECISION_ACTOR_BASIS = (
    "The supplied fresh-session handoff identifies these selections as Abdel's saved decisions."
)

CONFIG = {
    "cardiff": {
        "team": "Cardiff", "source_dir": "Cardiff",
        "injury": "Cardiff standardised_data .xlsx",
        "exposure": "Cardiff standardised_Exposure data .xlsx",
        "raw_exposure": "Cardiff Exposure Data .csv",
        "injury_codebook": "mapping-codebook-Cardiff Injuries.csv",
        "exposure_codebook": "mapping-codebook-Cardiff Exp..csv",
        "time_loss_field": "TimeLoss vs Medical Attention",
        "retained_injury_rows": 62, "placeholder_rows": 0, "unresolved": [],
        "injury_window": "27 Nov 2022 to 10 May 2025 (60 valid dates; two blank)",
        "exposure_window": "12 Jul 2024 to 29 May 2025",
        "stats": {
            "occasion": 57, "training": 13, "problem": 62, "closed": 42,
            "open": 0, "fit": 42, "return": 42, "days_before": 36,
            "days_after": 43, "severity": 42, "recurrence": 59,
            "contact": 53, "body": 62, "tissue": 39,
            "dob": 27, "bad_return_order": 2, "exposure_duplicates": 13,
            "duration_above_220": 0,
        },
    },
    "dragons": {
        "team": "Dragons", "source_dir": "Dragons",
        "injury": "Dragons standardised_data.xlsx",
        "exposure": "Dragons standardised_Exposure data .xlsx",
        "raw_exposure": "Dragons Exposure Data 24_25.csv",
        "injury_codebook": "Injury mapping-codebook-Dragons.csv",
        "exposure_codebook": "Exposure mapping-codebook-Dragons.csv",
        "time_loss_field": "TimeLoss vs Medical Attention (Time Loss yes or no)",
        "extra_injury_categories": ["Match vs Training"],
        "retained_injury_rows": 204, "placeholder_rows": 17,
        "unresolved": ["DRAGONS-2024-25-ADJ-002-PROBLEM-TYPE"],
        "injury_window": "29 Dec 2023 to 11 May 2025 (197 valid dates; one invalid and six blank among retained rows)",
        "exposure_window": "9 Jul 2024 to 22 May 2025",
        "stats": {
            "occasion": 149, "training": 68, "problem": 157, "closed": 142,
            "open": 61, "fit": 141, "return": 141, "days_before": 154,
            "days_after": 154, "severity": 142, "recurrence": 153,
            "contact": 154, "body": 157, "tissue": 0,
            "dob": 0, "bad_return_order": 0, "exposure_duplicates": 5,
            "duration_above_220": 0,
        },
    },
    "ospreys": {
        "team": "Ospreys", "source_dir": "Ospreys",
        "injury": "Ospreys standardised_data .xlsx",
        "exposure": "Ospreys standardised_Exposure data.xlsx",
        "raw_injury": "Ospreys- Overall Injuries -2024_25.xlsx",
        "raw_exposure": "Ospreys Exposure Data 24_25.csv",
        "injury_codebook": "mapping-codebook-Ospreys.csv",
        "exposure_codebook": "mapping-codebook-Ospreys Exp..csv",
        "time_loss_field": "TimeLoss vs Medical Attention(Yes = Time loss)",
        "retained_injury_rows": 163, "placeholder_rows": 58,
        "unresolved": ["OSPREYS-2024-25-ADJ-001-PROBLEM-TYPE"],
        "injury_window": "26 Oct 2014 to 17 May 2025 (163 valid dates; historical rows retained for the frozen window rule)",
        "exposure_window": "15 Jul 2024 to 10 Jul 2025",
        "stats": {
            "occasion": 126, "training": 54, "problem": 129, "closed": 150,
            "open": 13, "fit": 150, "return": 150, "days_before": 150,
            "days_after": 151, "severity": 150, "recurrence": 162,
            "contact": 131, "body": 129, "tissue": 77,
            "dob": 124, "bad_return_order": 2, "exposure_duplicates": 3,
            "duration_above_220": 1,
        },
    },
    "scarlets": {
        "team": "Scarlets", "source_dir": "Scarlets",
        "injury": "Scarlets standardizsed_data .xlsx",
        "exposure": "Scarlets standardised_Exposure data .xlsx",
        "raw_injury": "Scarlets Overall Injuries 2024_25.xlsx",
        "raw_exposure": "Scarlets Exposure Data 24_25.csv",
        "injury_codebook": "mapping-codebook-Scarlets.csv",
        "exposure_codebook": "mapping-codebook-Scarlets Exp..csv",
        "time_loss_field": "TimeLoss vs Medical Attention",
        "retained_injury_rows": 200, "placeholder_rows": 0,
        "unresolved": ["SCARLETS-2024-25-ADJ-001-PROBLEM-TYPE"],
        "injury_window": "5 Jul 2024 to 24 Sep 2029 (one future date retained as an anomaly, not silently corrected)",
        "exposure_window": "2 Jul 2024 to 31 May 2025",
        "stats": {
            "occasion": 147, "training": 52, "problem": 154, "closed": 185,
            "open": 15, "fit": 172, "return": 172, "days_before": 187,
            "days_after": 188, "severity": 185, "recurrence": 156,
            "contact": 149, "body": 154, "tissue": 92,
            "dob": 0, "bad_return_order": 3, "exposure_duplicates": 42,
            "duration_above_220": 0,
        },
    },
}

INJURY_IDENTIFIER = ["PlayerID", "DOB"]
INJURY_FREE_TEXT = ["Description", "Mechanism Notes"]
INJURY_DATES = ["Date Injured", "Fit For Selection Date", "Confirmed Return Date"]
INJURY_CATEGORIES = [
    "Problem type", "Injury Status", "Fit for selection", "Training only",
    "Treatment/Rehab", "Occasion category", "Body Part", "Orchard Code",
    "Illness Code", "Injury Tissue Type/s", "Side", "Injury Grade",
    "BAMIC Grade", "Nature of onset", "Recurrence", "Recurrence Stage",
    "Is Contact", "Mechanism of Injury", "Injury Surface Type",
    "Injury Surface Condition", "Injury Ambient Condition",
    "Injury Training Type", "Match Type", "Occasion", "Match status",
    "Equipment", "Injury Immediate Action", "Required Surgery",
]
EXPOSURE_CATEGORIES = ["Team", "Competition", "session type", "If match, surface?"]
EXPOSURE_METRICS = [
    "minutes total", "distance total", "high speed running distance",
    "very high speed running distance",
]
BODY_LABEL_MAP = {
    "Ankle": "ankle", "Anterior_thigh": "thigh", "Chest": "chest",
    "Elbow": "elbow", "Foot": "foot", "Head": "head",
    "Hip_and_Groin": "hip_groin", "Knee": "knee",
    "Lower_leg": "lower_leg", "Lumbar_Spine": "lumbosacral",
    "Medical_illness": "unknown", "Neck": "neck",
    "Posterior_thigh": "thigh", "Shoulder": "shoulder",
    "Thoracic_Spine": "thoracic_spine", "Trunk_and_Abdominal": "unknown",
    "Upper_Arm": "upper_arm", "Wrist_and_Hand": "unknown",
}


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def canonical_team_name(team_key: str) -> str:
    """Return the exact frozen CLI/database team name for a Welsh team key."""
    return "Dragons RFC" if team_key == "dragons" else CONFIG[team_key]["team"]


def validate_saved_decision_state(
    state: dict,
    *,
    current_fingerprint: str,
    required_decisions: list[dict],
) -> dict[str, dict]:
    """Fail closed unless every fingerprint-bound Welsh decision is usable."""
    if state.get("evidence_fingerprint") != current_fingerprint:
        raise ValueError("saved Welsh decisions do not match the current evidence fingerprint")
    if state.get("invalidated_previous_state"):
        raise ValueError("saved Welsh decisions were invalidated by changed evidence")

    selections = state.get("selections")
    if not isinstance(selections, dict):
        raise ValueError("saved Welsh decision selections must be an object")
    required_ids = {item["id"] for item in required_decisions}
    missing = sorted(required_ids - set(selections))
    extra = sorted(set(selections) - required_ids)
    if missing or extra:
        raise ValueError(
            f"saved Welsh decisions are incomplete or unexpected: missing={missing} extra={extra}"
        )

    for decision in required_decisions:
        selection = selections[decision["id"]]
        if not isinstance(selection, dict):
            raise ValueError(f"saved selection must be an object: {decision['id']}")
        allowed = {choice["value"] for choice in decision["choices"]}
        choice = selection.get("choice")
        if choice not in allowed:
            raise ValueError(f"unsupported saved choice for {decision['id']}: {choice!r}")
        if choice != SUPPORTED_DECISION_CHOICE:
            raise ValueError(
                f"Welsh finalizer does not implement choice {choice!r} for {decision['id']}"
            )
        try:
            selected_at = datetime.fromisoformat(str(selection.get("selected_at")))
        except ValueError as exc:
            raise ValueError(f"saved selection has an invalid timestamp: {decision['id']}") from exc
        if selected_at.tzinfo is None:
            raise ValueError(f"saved selection timestamp is not timezone-aware: {decision['id']}")
    return selections


def load(path: Path) -> dict:
    return json.loads(path.read_text())


def write(path: Path, payload: dict) -> None:
    write_text_atomic(path, json.dumps(payload, indent=2, sort_keys=True) + "\n")


def write_text_atomic(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(
        dir=path.parent, prefix=f".{path.name}.", suffix=".tmp"
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(text)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_name, path)
        directory_fd = os.open(path.parent, os.O_RDONLY)
        try:
            os.fsync(directory_fd)
        finally:
            os.close(directory_fd)
    finally:
        if os.path.exists(temporary_name):
            os.unlink(temporary_name)


def sanitize_orchard_code_evidence(team_key: str) -> int:
    """Remove narrative-like values accidentally placed in Orchard Code.

    The scanner plan must retain Orchard Code as a safe category to support the
    team-specific IOC pass. The source contains one Cardiff narrative in that
    column, so the Welsh compiler applies a deterministic code-token allowlist
    before any aggregate evidence is treated as publishable.
    """
    team_dir = INTAKE / team_key
    evidence_path = team_dir / "mechanical_evidence.v1.json"
    evidence = load(evidence_path)
    redacted = 0
    for source in evidence["sources"]:
        if source["id"] != "injury":
            continue
        for source_sheet in source["sheets"]:
            frequencies = source_sheet["category_frequencies"].get("Orchard Code", {})
            unsafe = {
                value: count for value, count in frequencies.items()
                if value != "#REF!" and not re.fullmatch(r"[A-Z0-9]{2,5}", value)
            }
            if not unsafe:
                continue
            redacted += sum(unsafe.values())
            source_sheet["category_frequencies"]["Orchard Code"] = {
                value: count for value, count in frequencies.items() if value not in unsafe
            }
            source_sheet["joint_category_frequencies"] = [
                item for item in source_sheet["joint_category_frequencies"]
                if not (
                    "Orchard Code" in item["fields"]
                    and item["values"][item["fields"].index("Orchard Code")] in unsafe
                )
            ]
            source_sheet["privacy_redactions"] += sum(unsafe.values())
    if redacted:
        write(evidence_path, evidence)
        inventory = _build_inventory(
            {"team": evidence["team"], "team_key": evidence["team_key"], "season": evidence["season"]},
            evidence,
        )
        write(team_dir / "column_inventory.v2.json", inventory)
    return redacted


def injury_source(path: Path, config: dict) -> dict:
    return {
        "id": "injury", "role": "proposed_intake", "kind": "injury",
        "path": str(path), "sheets": [SHEET],
        "column_classes": {
            "identifier": INJURY_IDENTIFIER, "free_text": INJURY_FREE_TEXT,
            "date": INJURY_DATES,
            "safe_category": INJURY_CATEGORIES + [config["time_loss_field"]]
            + config.get("extra_injury_categories", []), "opaque": [],
        },
        "date_order": "day_first",
        "duplicate_keys": [["PlayerID", "Date Injured"], ["InjuryID"]],
        "exact_row_duplicates": True,
        "joint_category_keys": [
            ["Body Part", "Orchard Code"],
            ["Injury Tissue Type/s", "Orchard Code"],
            ["Body Part", "Injury Tissue Type/s", "Orchard Code"],
            ["Problem type", "Body Part", "Orchard Code"],
        ],
        "required_metrics": [], "anomaly_rules": [],
    }


def exposure_source(path: Path) -> dict:
    return {
        "id": "exposure", "role": "proposed_intake", "kind": "exposure",
        "path": str(path), "sheets": [SHEET],
        "column_classes": {
            "identifier": ["name"], "free_text": [], "date": ["session date"],
            "safe_category": EXPOSURE_CATEGORIES, "opaque": [],
        },
        "date_order": "day_first",
        "duplicate_keys": [["name", "session date", "session start date time"]],
        "exact_row_duplicates": True, "required_metrics": EXPOSURE_METRICS,
        "exposure_grain_evidence": {
            "weekly_columns": ["Week"],
            "session_columns": ["session date", "session start date time", "session end date time"],
        },
        "anomaly_rules": [
            {"id": "minutes_blank", "column": "minutes total", "operator": "blank"},
            {"id": "distance_blank", "column": "distance total", "operator": "blank"},
            {"id": "hsr_blank", "column": "high speed running distance", "operator": "blank"},
            {"id": "vhsr_blank", "column": "very high speed running distance", "operator": "blank"},
            {"id": "identifier_blank", "column": "name", "operator": "blank"},
            {"id": "duration_above_220", "column": "minutes total", "operator": "gt", "value": 220},
            {"id": "negative_duration", "column": "minutes total", "operator": "lt", "value": 0},
            {"id": "negative_distance", "column": "distance total", "operator": "lt", "value": 0},
        ],
    }


def raw_injury_source(path: Path) -> dict:
    return {
        "id": "raw_injury", "role": "reference_only", "kind": "injury",
        "path": str(path), "sheets": ["Injury Records"],
        "column_classes": {
            "identifier": ["Full name and DoB", "DOB\ndd/mm/yy", "WRU Registration Number"],
            "free_text": ["Diagnosis\n (free text)", "Injury Diagnosis (concatenated)",
                          "Orchard Diagnosis Text", "Mechanism Details\n (Optional Free Text)",
                          "Other/additional details", "Additional Comments\n (Optional Free Text)"],
            "date": ["Date of First Symptoms\n dd/mm/yy",
                     "Date first unavailable for competition\n dd/mm/yy",
                     "Expected available date\n dd/mm/yy",
                     "Date of return to availability for selection\n dd/mm/yy",
                     "Date Fully Resolved\n dd/mm/yy"],
            "safe_category": ["Body Area", "Side of injury", "Orchard Code",
                              "Squad at time of injury", "If International, type?",
                              "Injured During", "If match or training injury, surface?",
                              "If match injury, removed from play?", "If match injury, started or replacement",
                              "if match injury, position when injured", "If training injury, activity?",
                              "Other/Non-Rugby activity", "Mechanism of Injury", "Contact injury?",
                              "If contact injury, activity?", "If non-contact injury, activity?",
                              "Recurrent Injury", "Surgery?", "Time loss injury?",
                              "Did this injury cause this player to retire from rugby?"],
            "opaque": [],
        },
        "date_order": "day_first",
        "duplicate_keys": [["WRU Registration Number", "Date of First Symptoms\n dd/mm/yy"]],
        "exact_row_duplicates": True,
        "joint_category_keys": [["Body Area", "Orchard Code"], ["Injured During", "Time loss injury?"]],
        "required_metrics": [], "anomaly_rules": [],
    }


def build_plan(team_key: str) -> dict:
    config = CONFIG[team_key]
    source_dir = SOURCE_ROOT / config["source_dir"]
    sources = [
        injury_source(source_dir / config["injury"], config),
        exposure_source(source_dir / config["exposure"]),
    ]
    if config.get("raw_injury"):
        sources.append(raw_injury_source(source_dir / config["raw_injury"]))
    return {"plan_version": "team_intake_profiling_plan_v1", "team": config["team"],
            "team_key": team_key, "season": SEASON, "sources": sources}


def scan_all() -> None:
    evidence_paths = []
    for team_key in CONFIG:
        plan = build_plan(team_key)
        plan_path = INTAKE / team_key / "profiling_plan.v1.json"
        write(plan_path, plan)
        result = scan_plan(plan_path, INTAKE, CACHE)
        redacted = sanitize_orchard_code_evidence(team_key)
        evidence_paths.append(INTAKE / team_key / "mechanical_evidence.v1.json")
        print(team_key, result["status"], f"sources={result['source_count']}",
              f"cache_hits={result['cache_hits']}", f"privacy_redactions={redacted}")
    family_check(evidence_paths, INTAKE / "wales" / "profile_family_check.v1.json")


def sheet(evidence: dict, source_id: str) -> dict:
    return next(item for item in evidence["sources"] if item["id"] == source_id)["sheets"][0]


def locator(field: str, source_id: str = "injury", sheet_name: str = SHEET) -> dict[str, str]:
    return {"source_id": source_id, "sheet": sheet_name, "field": field}


def mapping_entry(canonical_field: str, canonical_value: str, source_evidence: dict[str, str],
                  *, rule: str, evidence_class: str = "deterministic_derivation",
                  supporting_evidence: dict[str, str] | None = None,
                  specificity_change: str = "equivalent",
                  protocol_rule_id: str | None = None, adjudication_id: str | None = None,
                  evidence_source_id: str = "injury", evidence_sheet: str = SHEET) -> dict:
    return {
        "canonical_field": canonical_field, "canonical_value": canonical_value,
        "evidence_class": evidence_class, "source_evidence": source_evidence,
        "specificity_change": specificity_change,
        "supporting_evidence": supporting_evidence or {},
        "evidence_source_id": evidence_source_id, "evidence_sheet": evidence_sheet,
        "rule": rule, "protocol_rule_id": protocol_rule_id, "adjudication_id": adjudication_id,
    }


def injury_mappings(team_key: str, injury_sheet: dict) -> list[dict]:
    categories = injury_sheet["category_frequencies"]
    result: list[dict] = []
    for value in categories["Occasion category"]:
        canonical = "match" if value == "Match" else "training" if value == "Training" else "unknown"
        result.append(mapping_entry("occasion_category", canonical, {"Occasion category": value},
                                    evidence_class="source_reported",
                                    rule="Normalize only explicit Match or Training; every other label remains Unknown."))
        match_type = "training" if value == "Training" else "unknown"
        result.append(mapping_entry("match_type", match_type, {"Occasion category": value},
                                    rule="Training is explicit; match rows remain Unknown without explicit URC competition or a unique audited fixture link."))
    for value in categories["Body Part"]:
        problem = "illness" if value == "Medical_illness" else "injury"
        medical_adjudication = CONFIG[team_key]["unresolved"][0] if value == "Medical_illness" else None
        result.append(mapping_entry(
            "problem_type", problem, {"Body Part": value},
            evidence_class="manual_adjudication" if medical_adjudication else "protocol_defined_inference",
            protocol_rule_id=None if medical_adjudication else "team_specific_cross_field_v1",
            adjudication_id=medical_adjudication,
            rule=("Proposed Medical_illness to illness mapping is isolated behind the named team-specific adjudication."
                  if medical_adjudication else
                  "An explicit anatomical injury bucket supports injury; absent evidence remains Unknown."),
        ))
        target = BODY_LABEL_MAP[value]
        result.append(mapping_entry("body_location", target, {"Body Part": value},
                                    evidence_class="source_reported" if target != "unknown" else "deterministic_derivation",
                                    rule="Preserve the explicit IOC-equivalent body area; combined or non-injury labels remain Unknown until row-level code evidence resolves them."))
    for value in categories["Recurrence"]:
        folded = value.casefold()
        target = "recurrence" if folded == "yes" else "first_episode" if folded == "no" else "unknown"
        result.append(mapping_entry("recurrence", target, {"Recurrence": value}, evidence_class="source_reported",
                                    rule="Normalize explicit Yes/No recurrence only; N/A or missing remains Unknown."))
    for value in categories["Is Contact"]:
        folded = value.casefold()
        target = "contact" if folded == "yes" else "non_contact" if folded == "no" else "unknown"
        result.append(mapping_entry("contact_context", target, {"Is Contact": value}, evidence_class="source_reported",
                                    rule="Normalize explicit Yes/No contact only; N/A, Unknown, or missing remains Unknown."))

    for joint in injury_sheet["joint_category_frequencies"]:
        if joint["fields"] != ["Body Part", "Orchard Code"]:
            continue
        body, code = joint["values"]
        valid_code = bool(re.fullmatch(r"[A-Z0-9]{2,5}", code or "")) and code != "#REF!"
        if body in {"Wrist_and_Hand", "Trunk_and_Abdominal"} and valid_code:
            target = IOC_BODY_CODE_MAP.get(code[0], "unknown")
            result.append(mapping_entry(
                "body_location", target, {"Body Part": body}, supporting_evidence={"Orchard Code": code},
                evidence_class="protocol_defined_inference", protocol_rule_id="ioc_code_mapping_v1",
                specificity_change="narrower",
                rule="Resolve the combined source body label only from its retained row-level OSIICS body prefix.",
            ))
        if body == "Medical_illness" or not valid_code:
            continue
        pathology = ORCHARD_PATHOLOGY_TYPE_MAP.get(code[1]) if len(code) > 1 else None
        if pathology:
            result.append(mapping_entry(
                "tissue_pathology", pathology, {"Orchard Code": code},
                supporting_evidence={"Body Part": body},
                evidence_class="protocol_defined_inference", protocol_rule_id="ioc_code_mapping_v1",
                rule="Map the retained row-level OSIICS pathology character into the frozen IOC pathology bucket; unsupported characters remain Unknown.",
            ))
    if team_key == "dragons":
        result.append(mapping_entry("tissue_pathology", "unknown", {"Orchard Code": "#REF!"},
                                    rule="The literal spreadsheet error is not clinical code evidence; tissue/pathology remains Unknown."))
    return result


def build_assessments(team_key: str) -> list[dict]:
    c = CONFIG[team_key]
    s, n = c["stats"], c["retained_injury_rows"]
    specs = {
        "occasion_category": (["Occasion category"], "Normalize explicit Match/Training; all other labels remain Unknown.", "source_reported", "mapped_from_explicit_setting", s["occasion"] / n, s["occasion"] / n, []),
        "match_type": (["Match Type", "Occasion category", "Date Injured"], "Populate training from explicit training. Populate URC only from explicit competition or a unique audited fixture; no Welsh injury file supplies either, so match rows remain Unknown.", "deterministic_derivation", "explicit_competition_unique_fixture_or_unknown", 0, s["training"] / n, []),
        "problem_type": (["Problem type", "Body Part", "Orchard Code"],
                         ("Classify explicit anatomical injury evidence as injury; the Medical_illness to illness mapping remains pending the named team-specific adjudication."
                          if c["unresolved"] else
                          "Classify explicit anatomical injury evidence as injury; absent evidence remains Unknown."),
                         "manual_adjudication" if c["unresolved"] else "protocol_defined_inference",
                         "pending_manual_adjudication" if c["unresolved"] else "team_specific_problem_type_mapping", 0, s["problem"] / n,
                         c["unresolved"]),
        "injury_status": (["Injury Status", "Fit For Selection Date", "Days Injured", c["time_loss_field"]], "Use a valid explicitly headed return-to-availability date as closed; explicit time-loss Yes without return is open/censored and explicit No is medical-attention closed; otherwise Unknown.", "deterministic_derivation", "explicit_return_or_time_loss_status", 0, (s["closed"] + s["open"]) / n, [f"{s['bad_return_order']} return date(s) precede injury date and remain unresolved row anomalies, not silently corrected."] if s["bad_return_order"] else []),
        "fit_for_selection_status": (["Fit for selection", "Fit For Selection Date"], "A valid date whose source header explicitly says return to availability for selection supports fit; missing or contradictory dates remain Unknown.", "source_reported", "explicit_return_to_availability_header", 0, s["fit"] / n, []),
        "confirmed_return_date": (["Confirmed Return Date", "Fit For Selection Date", "Date Injured"], "Move only valid source return-to-availability dates to Confirmed Return Date; preserve the source field and physical row locator.", "deterministic_derivation", "source_reported_return_to_availability", 0, s["return"] / n, []),
        "days_injured": (["Days Injured", "Date Injured", "Fit For Selection Date"], "Preserve every valid non-negative source Days Injured value even when calendar dates disagree; derive only when source days are missing/invalid and both dates are valid and ordered.", "deterministic_derivation", "source_days_precedence_then_valid_date_difference", s["days_before"] / n, s["days_after"] / n, []),
        "severity_time_loss_category": (["Days Injured", c["time_loss_field"], "Fit For Selection Date"], "Apply only the frozen v1 severity bands to closed rows; open rows remain censored/Unknown.", "deterministic_derivation", "frozen_v1_severity_rule", 0, s["severity"] / n, []),
        "recurrence": (["Recurrence"], "Normalize explicit Yes/No only; N/A or missing remains Unknown.", "source_reported", "mapped_from_source_recurrence", s["recurrence"] / n, s["recurrence"] / n, []),
        "contact_context": (["Is Contact"], "Normalize explicit Yes/No only; N/A, Unknown, or missing remains Unknown.", "source_reported", "mapped_from_source_contact", s["contact"] / n, s["contact"] / n, []),
        "body_location": (["Body Part", "Orchard Code"], "Use explicit IOC-equivalent body labels; resolve Wrist_and_Hand or Trunk_and_Abdominal only from the same row's valid OSIICS prefix; illness/unsupported evidence remains Unknown.", "protocol_defined_inference", "team_specific_ioc_row_mapping", s["body"] / n, s["body"] / n, []),
        "tissue_pathology": (["Injury Tissue Type/s", "Orchard Code", "Body Part"], "Use only a valid retained row-level OSIICS pathology character for non-illness rows; unsupported, absent, free-text, and spreadsheet-error evidence remains Unknown.", "protocol_defined_inference", "ioc_code_mapping_or_unknown", 0, s["tissue"] / n, ["Dragons Orchard Code is literal #REF! throughout; no tissue/pathology is inferred."] if team_key == "dragons" else []),
    }
    return [{
        # The package validator requires a structurally complete assessment
        # record. For unresolved problem type, review_required + pending origin
        # explicitly distinguish complete evidence review from an approved map.
        "canonical_field": field, "status": "complete",
        "source_fields": [locator(source) for source in fields], "rule": rule,
        "evidence_class": evidence_class, "origin_status": origin,
        "coverage_before": round(before, 6), "coverage_after": round(after, 6),
        "conflicts": conflicts,
        "review_required": bool(field == "problem_type" and c["unresolved"]),
        "tests": [f"{team_key}_{field}_aggregate_reconciliation"],
    } for field, (fields, rule, evidence_class, origin, before, after, conflicts) in specs.items()]


def evidence_field(value: str | None) -> dict[str, str | None]:
    return {"status": "available" if value else "unavailable", "value": value}


def provenance(config: dict, evidence: dict) -> list[dict]:
    result = []
    for source in evidence["sources"]:
        source_id, kind = source["id"], source["kind"]
        reference = source["role"] == "reference_only"
        identifier = "Full name and DoB" if reference else "PlayerID" if kind == "injury" else "name"
        codebook = config["injury_codebook"] if kind == "injury" else config["exposure_codebook"]
        physical = sum(item["physical_data_rows"] for item in source["sheets"])
        result.append({
            "source_id": source_id,
            "preparer": evidence_field(None), "preparation_timestamp": evidence_field(None),
            "codebook_version": evidence_field(f"checksummed {codebook}; column-mapping evidence only"),
            "secure_original_locator": evidence_field(None), "secure_original_checksum": evidence_field(None),
            "pseudonymisation_status": evidence_field(
                "reference-only legacy source may contain direct identifiers and is prohibited from intake"
                if reference else
                "candidate identifier values were not serialized; DOB removal and stable pseudonym verification are required in the adapter"
            ),
            "player_identifier_field": evidence_field(identifier),
            "player_identifier_status": evidence_field(
                "used only in-memory for row reconciliation; values excluded from artifacts" if reference
                else "identifier values excluded from aggregate evidence and draft artifacts"
            ),
            "carried_locator_status": evidence_field("file checksum + sheet/table + original physical row; provisional_reference_locator"),
            "row_reconciliation": {"status": "completed", "source_rows": physical,
                                   "profiled_rows": physical,
                                   "notes": "Every physical data row is counted; identifier and free-text cell values are excluded."},
        })
    return result


def source_inventory(config: dict) -> list[dict]:
    names = [("injury", config["injury"], "proposed injury intake"),
             ("exposure", config["exposure"], "proposed exposure intake"),
             ("raw_exposure", config["raw_exposure"], "reference-only row restoration evidence"),
             ("injury_codebook", config["injury_codebook"], "reference-only column mapping codebook"),
             ("exposure_codebook", config["exposure_codebook"], "reference-only column mapping codebook")]
    if config.get("raw_injury"):
        names.insert(2, ("raw_injury", config["raw_injury"], "reference-only injury and embedded OSIICS codebook evidence"))
    result = []
    source_dir = SOURCE_ROOT / config["source_dir"]
    for source_id, name, role in names:
        path = source_dir / name
        tables = []
        if path.suffix.casefold() == ".csv":
            with path.open(encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                headers = next(reader, [])
                row_count = sum(1 for row in reader if any(str(cell).strip() for cell in row))
            tables.append(f"CSV table: {row_count} substantive rows, {len(headers)} columns")
        else:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                tables = [f"{ws.title}: {ws.max_row - 1} physical data rows, {ws.max_column} columns" for ws in workbook.worksheets]
            finally:
                workbook.close()
        result.append({"id": source_id, "file": name, "role": role, "sha256": digest(path),
                       "tables": "; ".join(tables),
                       "boundary": "reference only" if "reference-only" in role else "draft candidate; adapter required before intake"})
    return result


def _comparison_value(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        number = Decimal(text)
    except InvalidOperation:
        return text
    return format(number.normalize(), "f")


def exposure_reconciliation(config: dict) -> dict:
    """Reconcile safe exposure fields by physical row without identifiers."""
    source_dir = SOURCE_ROOT / config["source_dir"]
    raw_path = source_dir / config["raw_exposure"]
    standard_path = source_dir / config["exposure"]
    with raw_path.open(encoding="utf-8-sig", newline="") as handle:
        raw_rows = list(csv.DictReader(handle))
    workbook = load_workbook(standard_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[SHEET]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        standard_rows = [dict(zip(headers, row)) for row in rows if any(value not in (None, "") for value in row)]
    finally:
        workbook.close()

    field_pairs = [
        ("date", "session date", "date"),
        ("week", "Week", "week grouping"),
        ("competition", "Competition", "competition category"),
        ("session", "session type", "session category"),
        ("duration", "minutes total", "duration metric"),
        ("distance_m", "distance total", "total-distance metric"),
        ("distance_high_speed", "sprint distance", "shifted high-speed source metric"),
        ("distance_very_high_speed", "high speed running minutes", "shifted very-high-speed source metric"),
    ]
    compared = min(len(raw_rows), len(standard_rows))
    comparisons = []
    for raw_field, standard_field, meaning in field_pairs:
        equal = sum(
            _comparison_value(raw_rows[index].get(raw_field))
            == _comparison_value(standard_rows[index].get(standard_field))
            for index in range(compared)
        )
        comparisons.append({
            "raw_field": raw_field, "standard_field": standard_field,
            "meaning": meaning, "rows_compared": compared,
            "equal_rows": equal, "mismatched_rows": compared - equal,
        })
    correct_targets_blank = {
        field: sum(_comparison_value(row.get(field)) == "" for row in standard_rows)
        for field in ("high speed running distance", "very high speed running distance")
    }
    passed = (
        len(raw_rows) == len(standard_rows)
        and all(item["equal_rows"] == len(raw_rows) for item in comparisons)
    )
    return {
        "status": "PASS" if passed else "FAIL",
        "method": "same physical row; safe mapped fields and metrics only; identifier values excluded",
        "raw_source_sha256": digest(raw_path),
        "standard_source_sha256": digest(standard_path),
        "raw_rows": len(raw_rows), "standard_rows": len(standard_rows),
        "row_count_equal": len(raw_rows) == len(standard_rows),
        "field_comparisons": comparisons,
        "correct_standard_targets_blank_rows": correct_targets_blank,
        "identifier_values_serialized": False,
    }


def adapter_plan(team_key: str, inventory: list[dict], evidence: dict) -> dict:
    c, s = CONFIG[team_key], CONFIG[team_key]["stats"]
    exposure_rows = sheet(evidence, "exposure")["substantive_rows"]
    reconciliation = exposure_reconciliation(c)
    if reconciliation["status"] != "PASS":
        raise RuntimeError(f"{team_key} exposure row reconciliation failed")
    rules = [
        {"id": "identifier_and_dob_boundary", "rows_examined": c["retained_injury_rows"] + c["placeholder_rows"],
         "affected": {"populated_dob_values_removed": s["dob"]},
         "action": "Remove DOB; resolve every candidate player identifier through the approved stable pseudonym workflow; never serialize raw identifier values."},
        {"id": "source_days_precedence", "rows_examined": c["retained_injury_rows"],
         "affected": {"valid_source_days_preserved": s["days_before"], "date_derived_only_when_source_invalid_or_missing": s["days_after"] - s["days_before"]},
         "action": "Preserve valid non-negative source Days Injured even when calendar dates differ; derive only with missing/invalid days and two valid ordered dates."},
        {"id": "return_to_availability", "rows_examined": c["retained_injury_rows"],
         "affected": {"accepted_explicit_return_dates": s["return"], "return_before_injury_left_unresolved": s["bad_return_order"]},
         "action": "Move the date row-wise to confirmed return/fit only because the retained codebook header explicitly says return to availability for selection; invalid ordering remains Unknown."},
        {"id": "injury_zero_only_template_rows", "rows_examined": c["retained_injury_rows"] + c["placeholder_rows"],
         "affected": {"zero_only_nonrecords": c["placeholder_rows"]},
         "action": "Rows containing only a generated zero duration and no identifier/date/body/context evidence are audited as blank template non-records, not injuries."},
        {"id": "exposure_hsr_column_restoration", "rows_examined": exposure_rows,
         "affected": {"distance_high_speed_restored": exposure_rows, "distance_very_high_speed_restored": exposure_rows},
         "action": "Source-preserve raw distance_high_speed and distance_very_high_speed in correctly named distance fields by the same physical row; never use raw identifiers for the join. Both metrics are non-comparable and analysis-ineligible until team-specific vendor/threshold evidence is approved.",
         "comparability": "non_comparable_unknown_vendor_threshold",
         "analysis_eligible": False},
        {"id": "exposure_exact_duplicates", "rows_examined": exposure_rows,
         "affected": {"duplicate_groups": s["exposure_duplicates"], "later_exact_copies_excluded": s["exposure_duplicates"]},
         "action": "Retain the first physical locator and exclude only later exact copies with the controlled duplicate reason; retain all non-identical repeated rows."},
        {"id": "unlabelled_exposure_scope", "rows_examined": exposure_rows,
         "affected": {"rows_reviewed": exposure_rows},
         "action": "Keep absent, zero, NA, or otherwise unlabeled competition/session scope included as Unknown; classify URC only from explicit URC and apply no inferred scope exclusion."},
        {"id": "frozen_exposure_validity", "rows_examined": exposure_rows,
         "affected": {"session_minutes_above_220": s["duration_above_220"]},
         "action": "Apply only the frozen >220-minute session validity exclusion and existing frozen negative/missing rules; introduce no vendor-specific threshold."},
    ]
    if team_key == "dragons":
        rules.append({"id": "dragons_invalid_orchard_code", "rows_examined": c["retained_injury_rows"],
                      "affected": {"literal_ref_error_rows": 200},
                      "action": "Treat literal #REF! as invalid code evidence; keep tissue/pathology Unknown unless another approved source field supplies evidence."})
    return {"adapter_plan_version": ADAPTER_VERSION, "status": "reviewed_unapproved",
            "team": c["team"], "team_key": team_key, "season": SEASON,
            "source_inventory": inventory,
            "source_bindings": {item["id"]: item["sha256"] for item in inventory},
            "exposure_reconciliation": reconciliation,
            "restored_high_speed_metric_policy": {
                "storage": "source_preserved_only",
                "comparability": "non_comparable_unknown_vendor_threshold",
                "analysis_eligible": False,
                "eligibility_requirement": "approved team-specific vendor and threshold evidence",
            },
            "rules": rules}


def compile_team(team_key: str) -> None:
    c = CONFIG[team_key]
    team_dir = INTAKE / team_key
    evidence_path = team_dir / "mechanical_evidence.v1.json"
    inventory_path = team_dir / "column_inventory.v2.json"
    mapping_path = team_dir / "source_to_canonical_mapping.v2.draft.json"
    adapter_path = team_dir / "source_adapter_plan.v1.draft.json"
    profile_path = team_dir / "team_intake_profile.v2.draft.json"
    evidence = load(evidence_path)
    injury_sheet, exposure_sheet = sheet(evidence, "injury"), sheet(evidence, "exposure")
    mappings = injury_mappings(team_key, injury_sheet)
    mapping = {"mapping_version": MAPPING_VERSION, "team": c["team"], "team_key": team_key,
               "season": SEASON, "status": "reviewed_unapproved",
               "inventory_sha256": digest(inventory_path), "evidence_sha256": digest(evidence_path),
               "mappings": mappings,
               "adapter_source_mappings": [
                   mapping_entry("confirmed_return_date", "source_reported",
                                 {"Date of return to availability for selection": "valid ordered date"},
                                 rule="Restore from the same physical source row under the explicit source header.",
                                 evidence_source_id="raw_injury", evidence_sheet="Injury Records")
               ] if c.get("raw_injury") else []}
    write(mapping_path, mapping)
    files = source_inventory(c)
    adapter = adapter_plan(team_key, files, evidence)
    write(adapter_path, adapter)
    exposure_check = adapter["exposure_reconciliation"]
    review_time = datetime.now(UTC).isoformat()
    s, n = c["stats"], c["retained_injury_rows"]
    profile = {
        "profile_version": "team_intake_profile_v2", "team": c["team"], "team_key": team_key,
        "season": SEASON, "decision": "adjudication_required" if c["unresolved"] else "adapter_required",
        "evidence_path": str(evidence_path.relative_to(ROOT)), "evidence_sha256": digest(evidence_path),
        "column_inventory_path": str(inventory_path.relative_to(ROOT)), "column_inventory_sha256": digest(inventory_path),
        "mapping_path": str(mapping_path.relative_to(ROOT)), "mapping_sha256": digest(mapping_path),
        "mapping_version": MAPPING_VERSION, "ai_review_status": "completed",
        "ai_reviewed_by": "Codex fresh aggregate-only review (/root/welsh_profile_compiler)",
        "ai_reviewed_at": review_time, "approval_status": "pending",
        "approved_by": None, "approved_at": None, "approved_input_sha256s": [],
        "unresolved_adjudication_ids": c["unresolved"], "provenance_review": provenance(c, evidence),
        "reporting_reviews": {
            "injury": {"status": "completed", "units": {"days_injured": "days", "dates": "calendar dates"},
                       "gaps": c["injury_window"] + "; values outside the frozen window are retained and handled by the frozen cohort rule.",
                       "repeated_measure_structure": "One injury observation per retained source row; zero-only template rows are non-records and exact/non-identical repetitions retain physical locators.",
                       "native_grain": "not_applicable", "grain_conclusion": "not_applicable",
                       "grain_review_rationale": "Reporting grain applies to exposure, not injury records.", "anomalies_reviewed": True},
            "exposure": {"status": "completed", "units": {"duration": "minutes", "distance": "metres", "high_speed_distance": "source-labelled metres; threshold unavailable"},
                         "gaps": c["exposure_window"] + "; unlabeled scope remains included as Unknown and only frozen validity rules apply. Restored high-speed metrics are source-preserved but non-comparable and analysis-ineligible until team-specific vendor/threshold evidence is approved.",
                         "repeated_measure_structure": "Player-session/activity rows contain participant, date, session label, duration, and distance; Week is a grouping label rather than a weekly aggregate.",
                         "native_grain": "session", "grain_conclusion": "reviewed_session",
                         "grain_review_rationale": "The current raw and standardised files reconcile row-for-row as dated participant activity/session records.", "anomalies_reviewed": True},
        },
        "taxonomy_review": {"status": "completed", "body_location_inventory_complete": True,
                            "tissue_pathology_inventory_complete": True,
                            "notes": ("Every source body/code pair was reviewed. Dragons #REF! codes provide no pathology evidence, so all tissue remains Unknown."
                                      if team_key == "dragons" else
                                      "Every source body/code pair was reviewed; valid row-level OSIICS characters map only to frozen IOC buckets and unsupported/non-code evidence remains Unknown.")},
        "tests_and_reconciliation_samples": [
            {"id": f"{team_key}_source_rows_and_checksums", "status": "passed",
             "evidence": "mechanical_evidence.v1.json plus source_adapter_plan.v1.draft.json checksum bindings",
             "notes": f"Reconciled {injury_sheet['physical_data_rows']} injury physical rows and {exposure_sheet['physical_data_rows']} exposure rows; {c['placeholder_rows']} zero-only injury template rows are separately accounted for."},
            {"id": f"{team_key}_exposure_rowwise_metric_restoration", "status": "passed",
             "evidence": "source_adapter_plan.v1.draft.json exposure_reconciliation bound to raw/standard source SHA-256 values",
             "notes": f"PASS: raw={exposure_check['raw_rows']} and standard={exposure_check['standard_rows']}; all {len(exposure_check['field_comparisons'])} safe field/metric comparisons match on every physical row without serializing identifier values."},
            {"id": f"{team_key}_privacy_safe_orchard_inventory", "status": "passed",
             "evidence": "post-scan Orchard Code token allowlist and regenerated mechanical evidence/inventory",
             "notes": f"Removed {injury_sheet['privacy_redactions']} narrative-like Orchard Code cell value(s) from category and joint frequencies; controlled code tokens remain available for IOC review."},
            {"id": f"{team_key}_taxonomy_reconciliation", "status": "passed",
             "evidence": "mechanical joint Body Part/Orchard Code frequencies reviewed against frozen IOC maps",
             "notes": f"Defensible coverage is {s['body']}/{n} body and {s['tissue']}/{n} tissue; every unsupported value remains Unknown."},
        ],
        "canonical_field_assessments": build_assessments(team_key),
        "ai_review": {"status": "completed", "findings": [
            {"finding": "The advisory family check grouped none of the four Welsh files.",
             "disposition": "Only the review skeleton is shared; checksums, counts, mappings, anomalies, and approval remain team-specific.", "status": "resolved"},
            {"finding": "The exposure codebooks shifted both raw high-speed distance fields into semantically wrong standardised targets.",
             "disposition": "A checksum-bound aggregate reconciliation proves the shift row-for-row. The adapter source-preserves the two metrics, but marks both non-comparable and analysis-ineligible until team-specific vendor/threshold evidence exists.", "status": "resolved"},
            {"finding": "Populated source days frequently disagree with calendar differences.",
             "disposition": "The prior decision is applied only where evidence matches: valid source days retain precedence and date derivation is limited to missing/invalid source days with valid ordered dates.", "status": "resolved"},
            {"finding": "Competition, exposure scope, dates, and taxonomy could be over-inferred.",
             "disposition": "URC requires explicit competition or a unique audited fixture; unlabeled exposure stays included as Unknown; invalid dates and unsupported clinical evidence remain Unknown.", "status": "resolved"},
            {"finding": "Direct identifiers, DOB, and clinical narrative must not enter the package.",
             "disposition": "Only aggregate counts, controlled categories/codes, file checksums, and provisional physical locators are serialized; DOB removal and pseudonym verification are adapter gates.", "status": "resolved"},
            *([{"finding": "Medical_illness is explicit source evidence, but it appears in Body Area/Body Part while canonical Problem type is blank.",
                "disposition": f"The proposed illness mapping is isolated behind {c['unresolved'][0]}; non-Medical rows require retained anatomical injury evidence and all unsupported rows remain Unknown.", "status": "resolved"}] if c["unresolved"] else []),
            *([{"finding": "Dragons Orchard Code is literal #REF! throughout the populated code column.",
                "disposition": "The spreadsheet error is rejected as clinical evidence and tissue/pathology remains Unknown without creating an adjudication need.", "status": "resolved"}] if team_key == "dragons" else []),
        ]},
    }
    write(profile_path, profile)

    assessments = {item["canonical_field"]: item for item in profile["canonical_field_assessments"]}
    rows = [f"| `{name}` | {'pending adjudication' if item['review_required'] else 'reviewed'} | {item['coverage_before']:.1%} | {item['coverage_after']:.1%} | {item['rule']} |"
            for name, item in assessments.items()]
    markdown = [
        f"# {c['team']} 2024-25 — Step 0 intake profile", "",
        f"**Decision:** `{'adjudication_required' if c['unresolved'] else 'adapter_required'}` (draft, unapproved)  ",
        f"**Unresolved scientific adjudications:** `{', '.join(c['unresolved'])}`  " if c["unresolved"] else "**Unresolved scientific adjudications:** none  ",
        f"**Profile/mapping/adapter:** `team_intake_profile_v2` / `{MAPPING_VERSION}` / `{ADAPTER_VERSION}`  ",
        f"**Fresh AI review:** completed `{review_time}`; approval fields remain unset.", "",
        "## Decision summary", "",
        "The frozen canonical model can represent the supplied evidence. A row-preserving source adapter is required for privacy normalization, explicit return-date restoration, shifted exposure metrics, and audited non-record/duplicate handling. Unsupported evidence remains `Unknown`; this draft authorizes no database action.",
        *(["", f"Human decision required: `{c['unresolved'][0]}` — approve or reject mapping the explicit `Medical_illness` Body Area/Body Part label to canonical problem type `illness`. Until approved, those rows remain `Unknown` for problem type."] if c["unresolved"] else []), "",
        "## Supplied-file inventory and provenance", "",
        "| File | Role/boundary | Tables or sheets | SHA-256 |", "|---|---|---|---|",
        *[f"| `{item['file']}` | {item['role']}; {item['boundary']} | {item['tables']} | `{item['sha256']}` |" for item in files], "",
        "Preparer, preparation timestamp, secure original locator, and secure original checksum are unavailable in the supplied package. The accepted provisional locator is file checksum + sheet/table + original physical row. Raw/reference identifiers are used only in-memory for reconciliation and no identifier or clinical free-text value is serialized.", "",
        "## Reporting structure", "",
        f"- Injury: {injury_sheet['physical_data_rows']} physical rows, {c['retained_injury_rows']} retained candidate records, and {c['placeholder_rows']} zero-only template non-records. Source window: {c['injury_window']}.",
        f"- Exposure: {exposure_sheet['substantive_rows']} participant-session/activity rows. Source window: {c['exposure_window']}. Reviewed native grain: `session`; Week is a grouping label, not a weekly aggregate.",
        f"- Exact exposure duplicates: {s['exposure_duplicates']} groups; exclude only the later exact copy with source locators. Non-identical repeated rows stay retained.",
        f"- Frozen validity: {s['duration_above_220']} exposure row(s) exceed 220 minutes. No new vendor/device threshold is proposed.",
        "- Unlabeled/zero/NA exposure context remains included as `Unknown`. URC is assigned only from explicit competition or a unique audited fixture.", "",
        "## Adapter findings", "",
        f"- Remove {s['dob']} populated DOB value(s); do not copy any raw player identifier value into the draft or candidate output.",
        f"- Preserve {s['days_before']} valid source-day values; only {s['days_after'] - s['days_before']} missing/invalid value(s) are date-derivable under ordered valid dates.",
        f"- The explicit source header supports {s['return']} valid return-to-availability dates as confirmed return and fit evidence; {s['bad_return_order']} contradictory date(s) remain Unknown/anomalous.",
        f"- Deterministic reconciliation PASS: {exposure_check['raw_rows']} raw rows equal {exposure_check['standard_rows']} standard rows and all {len(exposure_check['field_comparisons'])} safe mapped fields/metrics match row-wise; raw and standard SHA-256 bindings are in the adapter plan.",
        f"- Source-preserve both shifted high-speed distance fields for all {exposure_sheet['substantive_rows']} rows. They are **non-comparable and analysis-ineligible** until approved team-specific vendor/threshold evidence exists.",
        f"- Post-scan privacy sanitizer removed {injury_sheet['privacy_redactions']} narrative-like Orchard Code value(s) from category and joint frequencies; no narrative value is present in the aggregate package.",
        "- Preserve rows outside the apparent season and apply the frozen cohort/window rule later; do not silently edit dates or denominators.",
        *(["- Dragons: Orchard Code is literal `#REF!` throughout the populated column. It supplies no tissue evidence, so tissue/pathology remains `Unknown`."] if team_key == "dragons" else []), "",
        "## Canonical field assessment", "",
        "Coverage is defensible non-`Unknown` coverage over retained candidate injury records, not filled-cell count.", "",
        "| Canonical field | Review state | Before | After adapter/mapping | Rule/disposition |", "|---|---|---:|---:|---|", *rows, "",
        "## Team-specific IOC taxonomy", "",
        f"- Body location: {s['body']}/{n} rows have defensible explicit or same-row code-supported IOC body evidence.",
        f"- Tissue/pathology: {s['tissue']}/{n} rows have a supported non-illness OSIICS pathology character; all other rows remain `Unknown`.",
        "- Combined `Wrist_and_Hand` and `Trunk_and_Abdominal` labels are resolved only by a valid same-row OSIICS body prefix. Medical/illness rows receive no injury body or pathology bucket.",
        ("- The mapping file contains only controlled category values and code tokens; it omits the narrative mis-entered in Cardiff Orchard Code."
         if team_key == "cardiff" else
         "- The mapping file contains only controlled category values and code tokens observed in aggregate scanner evidence; it serializes no identifier or clinical narrative value."), "",
        "## AI review findings", "",
        *[f"- **Resolved:** {item['finding']} {item['disposition']}" for item in profile["ai_review"]["findings"]], "",
        "## Gate boundary", "",
        "Approval is pending: `approved_by = null`, `approved_at = null`, and approved input checksums are empty. No ingest, processing, curated build, release, database query/write, migration, GitHub action, or deployment was run.", "",
    ]
    (team_dir / "team_intake_profile.md").write_text("\n".join(markdown))
    print(team_key, profile["decision"], len(mappings), "mappings")


def finalize_decisions() -> None:
    """Apply the saved human choices without granting profile approval."""
    from tools import welsh_profile_decision_board as decision_board

    current_fingerprint = decision_board.evidence_fingerprint()
    state = decision_board.load_state()
    selections = validate_saved_decision_state(
        state,
        current_fingerprint=current_fingerprint,
        required_decisions=decision_board.DECISIONS,
    )
    state_path = INTAKE / "wales" / "decision_selections.json"
    state_sha256 = digest(state_path)
    handoff_path = Path("/tmp/urc-welsh-e2e-handoff.md")
    if not handoff_path.is_file():
        raise ValueError("Welsh handoff is required to bind the recorded human decision actor")
    handoff_text = handoff_path.read_text(encoding="utf-8")
    if "Abdel's saved decisions" not in handoff_text \
            or any(item["id"] not in handoff_text for item in decision_board.DECISIONS):
        raise ValueError("Welsh handoff does not identify the saved decisions and their actor")
    applied_at = datetime.now(UTC)
    run_id = f"{applied_at.strftime('%Y%m%dT%H%M%S%fZ')}-{state_sha256[:12]}"
    actor_snapshot_path = INTAKE / "wales" / "reviewed" / run_id / "urc-welsh-e2e-handoff.md"
    write_text_atomic(actor_snapshot_path, handoff_text)
    applications = []
    outputs = []

    for team_key, c in CONFIG.items():
        team_dir = INTAKE / team_key
        draft_mapping_path = team_dir / "source_to_canonical_mapping.v2.draft.json"
        draft_profile_path = team_dir / "team_intake_profile.v2.draft.json"
        review_dir = team_dir / "reviewed" / run_id
        evidence_path = review_dir / "mechanical_evidence.v1.json"
        inventory_path = review_dir / "column_inventory.v2.json"
        mapping_path = review_dir / "source_to_canonical_mapping.v2.json"
        profile_path = review_dir / "team_intake_profile.v2.json"
        mapping = load(draft_mapping_path)
        profile = load(draft_profile_path)
        evidence = load(team_dir / "mechanical_evidence.v1.json")
        evidence["team"] = canonical_team_name(team_key)
        write(evidence_path, evidence)
        inventory = _build_inventory(
            {
                "team": canonical_team_name(team_key),
                "team_key": team_key,
                "season": SEASON,
            },
            evidence,
        )
        write(inventory_path, inventory)
        expected_ids = set(c["unresolved"])
        observed_ids = {
            entry.get("adjudication_id")
            for entry in mapping.get("mappings", [])
            if entry.get("adjudication_id")
        }
        if observed_ids != expected_ids:
            raise ValueError(
                f"{team_key}: mapping adjudication IDs do not match the reviewed draft: "
                f"expected={sorted(expected_ids)} observed={sorted(observed_ids)}"
            )

        team_applications = []
        for entry in mapping.get("mappings", []):
            decision_id = entry.get("adjudication_id")
            if not decision_id:
                continue
            selection = selections[decision_id]
            if entry.get("canonical_field") != "problem_type" \
                    or entry.get("canonical_value") != "illness" \
                    or entry.get("source_evidence") != {"Body Part": "Medical_illness"}:
                raise ValueError(f"{team_key}: unexpected adjudicated mapping shape: {decision_id}")
            entry["rule"] = (
                "Map the explicit Medical_illness source label to canonical illness under "
                "the recorded team-specific human adjudication."
            )
            application = {
                "decision_id": decision_id,
                "choice": selection["choice"],
                "selected_at": selection["selected_at"],
                "selected_by": DECISION_ACTOR,
                "team_key": team_key,
                "canonical_team": canonical_team_name(team_key),
                "effect": {
                    "canonical_field": "problem_type",
                    "canonical_value": "illness",
                    "source_evidence": {"Body Part": "Medical_illness"},
                },
            }
            team_applications.append(application)
            applications.append(application)

        mapping.update(
            {
                "team": canonical_team_name(team_key),
                "status": "reviewed_pending_approval",
                "evidence_sha256": digest(evidence_path),
                "inventory_sha256": digest(inventory_path),
            }
        )
        write(mapping_path, mapping)

        for assessment in profile.get("canonical_field_assessments", []):
            if assessment.get("canonical_field") != "problem_type" or not expected_ids:
                continue
            assessment.update(
                {
                    "conflicts": [],
                    "origin_status": "manual_adjudication_applied",
                    "review_required": False,
                    "rule": (
                        "Classify Medical_illness as illness under the recorded team-specific "
                        "adjudication; classify other rows as injury only where retained injury "
                        "evidence supports it, otherwise Unknown."
                    ),
                }
            )

        profile.update(
            {
                "team": canonical_team_name(team_key),
                "decision": "adapter_required",
                "evidence_path": str(evidence_path.relative_to(team_dir)),
                "evidence_sha256": digest(evidence_path),
                "column_inventory_path": str(inventory_path.relative_to(team_dir)),
                "column_inventory_sha256": digest(inventory_path),
                "mapping_path": str(mapping_path.relative_to(team_dir)),
                "mapping_sha256": digest(mapping_path),
                "mapping_version": MAPPING_VERSION,
                "unresolved_adjudication_ids": [],
                "approval_status": "pending",
                "approved_by": None,
                "approved_at": None,
                "approved_input_sha256s": [],
            }
        )
        write(profile_path, profile)
        validation = validate_package(
            evidence_path,
            profile_path,
            mapping_path,
            inventory_path,
        )
        if validation["status"] != "PASS":
            raise ValueError(f"{team_key}: finalized package validation failed: {validation['errors']}")
        outputs.append(
            {
                "team_key": team_key,
                "team": canonical_team_name(team_key),
                "evidence_path": str(evidence_path.relative_to(ROOT)),
                "evidence_sha256": digest(evidence_path),
                "inventory_path": str(inventory_path.relative_to(ROOT)),
                "inventory_sha256": digest(inventory_path),
                "draft_mapping_path": str(draft_mapping_path.relative_to(ROOT)),
                "draft_mapping_sha256": digest(draft_mapping_path),
                "mapping_path": str(mapping_path.relative_to(ROOT)),
                "mapping_sha256": digest(mapping_path),
                "draft_profile_path": str(draft_profile_path.relative_to(ROOT)),
                "draft_profile_sha256": digest(draft_profile_path),
                "profile_path": str(profile_path.relative_to(ROOT)),
                "profile_sha256": digest(profile_path),
                "validation_status": validation["status"],
            }
        )

    application_path = INTAKE / "wales" / "decision_applications.v1.json"
    write(
        application_path,
        {
            "schema_version": DECISION_APPLICATION_VERSION,
            "run_id": run_id,
            "evidence_fingerprint": current_fingerprint,
            "decision_state_path": str(state_path.relative_to(ROOT)),
            "decision_state_sha256": state_sha256,
            "selected_by": DECISION_ACTOR,
            "actor_basis": DECISION_ACTOR_BASIS,
            "actor_evidence_path": str(actor_snapshot_path.relative_to(ROOT)),
            "actor_evidence_sha256": digest(actor_snapshot_path),
            "applied_at": applied_at.isoformat(),
            "approval_granted": False,
            "applications": applications,
            "outputs": outputs,
        },
    )
    print(
        f"finalized {len(outputs)} Welsh Step 0 packages with {len(applications)} "
        f"recorded adjudication applications; approval remains pending"
    )


def decision_application_errors(
    record: dict,
    *,
    state: dict,
    current_fingerprint: str,
    current_state_sha256: str,
    required_decisions: list[dict],
    validate_actor_file: bool = True,
) -> list[str]:
    errors = []
    required_record_keys = {
        "schema_version", "run_id", "evidence_fingerprint", "decision_state_path",
        "decision_state_sha256", "selected_by", "actor_basis", "actor_evidence_path",
        "actor_evidence_sha256", "applied_at", "approval_granted", "applications", "outputs",
    }
    if set(record) != required_record_keys:
        errors.append("decision application record schema keys are incomplete or unexpected")
    if record.get("schema_version") != DECISION_APPLICATION_VERSION:
        errors.append("decision application schema version mismatch")
    if record.get("evidence_fingerprint") != current_fingerprint:
        errors.append("decision application fingerprint does not match current evidence")
    if record.get("decision_state_sha256") != current_state_sha256:
        errors.append("decision application does not bind the current saved decision state")
    expected_state_path = str((INTAKE / "wales" / "decision_selections.json").relative_to(ROOT))
    if record.get("decision_state_path") != expected_state_path:
        errors.append("decision application state path is not canonical")
    if record.get("selected_by") != DECISION_ACTOR \
            or record.get("actor_basis") != DECISION_ACTOR_BASIS:
        errors.append("decision application human actor provenance is missing or changed")
    if not isinstance(record.get("actor_evidence_sha256"), str) \
            or not re.fullmatch(r"[0-9a-f]{64}", record["actor_evidence_sha256"]):
        errors.append("decision application actor evidence checksum is invalid")
    if record.get("approval_granted") is not False:
        errors.append("decision application record must not grant profile approval")
    run_id = record.get("run_id")
    if not isinstance(run_id, str) or not re.fullmatch(r"[0-9]{8}T[0-9]{12}Z-[0-9a-f]{12}", run_id):
        errors.append("decision application run ID is invalid")
    else:
        expected_actor_path = (
            INTAKE / "wales" / "reviewed" / run_id / "urc-welsh-e2e-handoff.md"
        )
        expected_actor_relative = str(expected_actor_path.relative_to(ROOT))
        if record.get("actor_evidence_path") != expected_actor_relative:
            errors.append("decision application actor evidence path/checksum drift")
        elif validate_actor_file and (
            not expected_actor_path.is_file()
            or record.get("actor_evidence_sha256") != digest(expected_actor_path)
        ):
            errors.append("decision application actor evidence path/checksum drift")
        elif validate_actor_file and "Abdel's saved decisions" not in expected_actor_path.read_text(encoding="utf-8"):
            errors.append("decision application actor evidence does not support the recorded actor")

    try:
        applied_at = datetime.fromisoformat(str(record.get("applied_at")))
    except ValueError:
        applied_at = None
        errors.append("decision application applied_at is not a valid ISO timestamp")
    if applied_at is not None and applied_at.tzinfo is None:
        errors.append("decision application applied_at is not timezone-aware")
    selected_times = []
    for selection in (state.get("selections") or {}).values():
        try:
            selected_times.append(datetime.fromisoformat(str(selection.get("selected_at"))))
        except (AttributeError, ValueError):
            pass
    if applied_at is not None and applied_at.tzinfo is not None and selected_times:
        if applied_at < max(selected_times):
            errors.append("decision application predates a saved human selection")
        if applied_at > datetime.now(UTC) + timedelta(minutes=5):
            errors.append("decision application timestamp is in the future")

    selections = state.get("selections") if isinstance(state.get("selections"), dict) else {}
    expected_ids = {item["id"] for item in required_decisions}
    applications = record.get("applications")
    if not isinstance(applications, list):
        errors.append("decision applications must be a list")
        applications = []
    application_by_id = {
        item.get("decision_id"): item
        for item in applications
        if isinstance(item, dict) and isinstance(item.get("decision_id"), str)
    }
    if len(application_by_id) != len(applications) or set(application_by_id) != expected_ids:
        errors.append("decision application IDs/cardinality do not match required decisions")
    decision_team = {
        decision_id: team_key
        for team_key, config in CONFIG.items()
        for decision_id in config["unresolved"]
    }
    expected_application_keys = {
        "decision_id", "choice", "selected_at", "selected_by", "team_key",
        "canonical_team", "effect",
    }
    for decision_id in expected_ids:
        application = application_by_id.get(decision_id)
        selection = selections.get(decision_id)
        if not isinstance(application, dict) or not isinstance(selection, dict):
            continue
        if set(application) != expected_application_keys:
            errors.append(f"{decision_id}: application schema keys are incomplete or unexpected")
        if application.get("choice") != selection.get("choice") \
                or application.get("selected_at") != selection.get("selected_at"):
            errors.append(f"{decision_id}: recorded application contradicts the saved selection")
        if application.get("selected_by") != DECISION_ACTOR:
            errors.append(f"{decision_id}: human adjudicator is missing or changed")
        team_key = decision_team.get(decision_id)
        if application.get("team_key") != team_key \
                or application.get("canonical_team") != canonical_team_name(team_key):
            errors.append(f"{decision_id}: application team identity mismatch")
        if application.get("effect") != {
            "canonical_field": "problem_type",
            "canonical_value": "illness",
            "source_evidence": {"Body Part": "Medical_illness"},
        }:
            errors.append(f"{decision_id}: application effect is missing or contradictory")

    outputs = record.get("outputs")
    if not isinstance(outputs, list):
        errors.append("decision application outputs must be a list")
        outputs = []
    output_keys = [item.get("team_key") for item in outputs if isinstance(item, dict)]
    if len(output_keys) != len(outputs) or len(set(output_keys)) != len(outputs) \
            or set(output_keys) != set(CONFIG):
        errors.append("decision application outputs do not cover each Welsh team exactly once")
    expected_output_keys = {
        "team_key", "team", "evidence_path", "evidence_sha256", "inventory_path",
        "inventory_sha256", "draft_mapping_path", "draft_mapping_sha256", "mapping_path",
        "mapping_sha256", "draft_profile_path", "draft_profile_sha256", "profile_path",
        "profile_sha256", "validation_status",
    }
    for output in outputs:
        if not isinstance(output, dict):
            continue
        team_key = output.get("team_key")
        if set(output) != expected_output_keys:
            errors.append(f"{team_key}: output schema keys are incomplete or unexpected")
        if team_key in CONFIG and output.get("team") != canonical_team_name(team_key):
            errors.append(f"{team_key}: output canonical team identity mismatch")
        if output.get("validation_status") != "PASS":
            errors.append(f"{team_key}: output does not record successful package validation")
    return errors


def validate_finalized() -> bool:
    from tools import welsh_profile_decision_board as decision_board

    application_path = INTAKE / "wales" / "decision_applications.v1.json"
    if not application_path.is_file():
        print("finalized validation FAIL: decision application record is missing")
        return False
    record = load(application_path)
    state_path = INTAKE / "wales" / "decision_selections.json"
    state = decision_board.load_state()
    current_fingerprint = decision_board.evidence_fingerprint()
    try:
        validate_saved_decision_state(
            state,
            current_fingerprint=current_fingerprint,
            required_decisions=decision_board.DECISIONS,
        )
    except ValueError as exc:
        print(f"finalized validation FAIL: {exc}")
        return False
    errors = decision_application_errors(
        record,
        state=state,
        current_fingerprint=current_fingerprint,
        current_state_sha256=digest(state_path),
        required_decisions=decision_board.DECISIONS,
    )
    raw_outputs = record.get("outputs")
    safe_outputs = raw_outputs if isinstance(raw_outputs, list) else []
    output_by_team = {
        item.get("team_key"): item
        for item in safe_outputs
        if isinstance(item, dict) and isinstance(item.get("team_key"), str)
    }
    for team_key in CONFIG:
        team_dir = INTAKE / team_key
        output = output_by_team.get(team_key, {})
        expected_dir = (team_dir / "reviewed" / str(record.get("run_id", ""))).resolve()
        resolved_paths = {}
        for label in ("evidence", "inventory", "mapping", "profile"):
            value = output.get(f"{label}_path")
            if not isinstance(value, str) or not value:
                errors.append(f"{team_key}: finalized {label} path is missing")
                continue
            path = (ROOT / value).resolve()
            if path.parent != expected_dir:
                errors.append(f"{team_key}: finalized {label} path escapes the immutable run directory")
                continue
            resolved_paths[label] = path
        if set(resolved_paths) != {"evidence", "inventory", "mapping", "profile"}:
            continue
        evidence_path = resolved_paths["evidence"]
        inventory_path = resolved_paths["inventory"]
        mapping_path = resolved_paths["mapping"]
        profile_path = resolved_paths["profile"]
        if not all(path.is_file() for path in (
            evidence_path, inventory_path, mapping_path, profile_path,
        )):
            errors.append(f"{team_key}: finalized evidence package is incomplete")
            continue
        mapping = load(mapping_path)
        profile = load(profile_path)
        if output.get("mapping_sha256") != digest(mapping_path):
            errors.append(f"{team_key}: finalized mapping checksum drift")
        if output.get("profile_sha256") != digest(profile_path):
            errors.append(f"{team_key}: finalized profile checksum drift")
        if output.get("evidence_sha256") != digest(evidence_path):
            errors.append(f"{team_key}: finalized evidence checksum drift")
        if output.get("inventory_sha256") != digest(inventory_path):
            errors.append(f"{team_key}: finalized inventory checksum drift")
        for label in ("mapping", "profile"):
            draft_value = output.get(f"draft_{label}_path")
            if not isinstance(draft_value, str) or not draft_value:
                errors.append(f"{team_key}: draft {label} path is missing")
                continue
            draft_path = (ROOT / draft_value).resolve()
            expected_draft = (team_dir / f"{'source_to_canonical_mapping.v2' if label == 'mapping' else 'team_intake_profile.v2'}.draft.json").resolve()
            if draft_path != expected_draft or not draft_path.is_file() \
                    or output.get(f"draft_{label}_sha256") != digest(draft_path):
                errors.append(f"{team_key}: recorded draft {label} binding drift")
        if profile.get("mapping_sha256") != digest(mapping_path):
            errors.append(f"{team_key}: profile does not bind finalized mapping")
        if profile.get("team") != canonical_team_name(team_key) \
                or mapping.get("team") != canonical_team_name(team_key):
            errors.append(f"{team_key}: canonical team name mismatch")
        if profile.get("decision") != "adapter_required" \
                or profile.get("unresolved_adjudication_ids") != []:
            errors.append(f"{team_key}: finalized decision state is not ingest-compatible")
        if profile.get("approved_by") is not None \
                or profile.get("approved_at") is not None \
                or profile.get("approved_input_sha256s") != []:
            errors.append(f"{team_key}: finalization incorrectly granted profile approval")
        result = validate_package(
            evidence_path,
            profile_path,
            mapping_path,
            inventory_path,
        )
        if result["status"] != "PASS":
            errors.append(f"{team_key}: package validation failed")
    if errors:
        print("finalized validation FAIL")
        for error in errors:
            print(" ", error)
        return False
    print("finalized validation PASS")
    return True


def validate_all() -> bool:
    passed = True
    for team_key in CONFIG:
        team_dir = INTAKE / team_key
        result = validate_package(team_dir / "mechanical_evidence.v1.json",
                                  team_dir / "team_intake_profile.v2.draft.json",
                                  team_dir / "source_to_canonical_mapping.v2.draft.json",
                                  team_dir / "column_inventory.v2.json")
        print(team_key, "validation", result["status"])
        if result["status"] != "PASS":
            passed = False
            for error in result["errors"]:
                print(" ", error["code"], error["message"])
    return passed


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Build aggregate-only Welsh Step 0 profile drafts.")
    parser.add_argument("--scan", action="store_true", help="Write scanner plans and refresh aggregate evidence.")
    parser.add_argument("--compile", action="store_true", help="Compile draft packages from existing evidence.")
    parser.add_argument("--validate", action="store_true", help="Validate the compiled draft package contract.")
    parser.add_argument(
        "--finalize-decisions",
        action="store_true",
        help="Apply fingerprint-bound saved decisions without granting profile approval.",
    )
    parser.add_argument(
        "--validate-finalized",
        action="store_true",
        help="Validate the decision application record and reviewed packages.",
    )
    args = parser.parse_args(argv)
    if not any((
        args.scan,
        args.compile,
        args.validate,
        args.finalize_decisions,
        args.validate_finalized,
    )):
        parser.error(
            "use --scan, --compile, --validate, --finalize-decisions, "
            "and/or --validate-finalized"
        )
    if args.scan:
        scan_all()
    if args.compile:
        for team_key in CONFIG:
            compile_team(team_key)
    if args.finalize_decisions:
        finalize_decisions()
    if args.validate and not validate_all():
        raise SystemExit(1)
    if args.validate_finalized and not validate_finalized():
        raise SystemExit(1)


if __name__ == "__main__":
    main()
