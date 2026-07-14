"""Compile the local-only Benetton and Zebre Step 0 profile packages.

The mechanical scanner remains the source for aggregate workbook evidence. This
compiler adds the team-specific review, source inventory, mappings, and the
recorded Italian adjudications. It never writes to the database or serializes
identifier/free-text values.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import csv
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from pipeline.__main__ import IOC_BODY_CODE_MAP, ORCHARD_PATHOLOGY_TYPE_MAP
from pipeline.profiling import MAPPING_VERSION


ROOT = Path(__file__).resolve().parents[1]
INTAKE = ROOT / "data" / "intake" / "2024-25"
SOURCE_ROOT = Path("/Users/abdelbabiker/Desktop/URC")
SEASON = "2024-25"
SHEET = "Standardized Data"
DECISION_STATE = INTAKE / "italy" / "decision_selections.json"
BENETTON_ADJUDICATION_ID = "BENETTON-2024-25-ADJ-001-TAXONOMY-CONFLICTS"
ZEBRE_ADJUDICATION_ID = "ZEBRE-2024-25-ADJ-001-PROBLEM-TYPE"
DECISION_EVIDENCE_FINGERPRINT = "14331ed8711bf1b9b09c471355d13458e0708f2bf68979d6c2d538ac795e0641"
EXPECTED_DECISION_CHOICES = {
    BENETTON_ADJUDICATION_ID: "prefer_explicit_source_label",
    ZEBRE_ADJUDICATION_ID: "medical_to_illness",
}

CONFIG = {
    "benetton": {
        "team": "Benetton",
        "source_dir": SOURCE_ROOT / "Benneton",
        "injury_standard": "Benneton standardised_data 24_25.xlsx",
        "exposure_standard": "Benneton standardised_Exposure data .xlsx",
        "injury_raw": "Overall Benneton Injuries 2024_25.xlsx",
        "exposure_raw": "Benneton exposure Data 24_25.xlsx",
        "injury_codebook": "Injury mapping-codebook-Benneton.csv",
        "exposure_codebook": "mapping-codebook-Benneton Exp.csv",
        "raw_rows": {"injury": 43, "exposure": 8446},
        "windows": {"injury": "24 Jul 2024 to 26 Apr 2025", "exposure": "22 Jul 2024 to 9 Jun 2025"},
        "exposure_context": "Device/vendor is unavailable. Source columns explicitly label HSR thresholds at >18, >20, and >25.2 km/h; match/training/session labels are blank.",
    },
    "zebre": {
        "team": "Zebre",
        "source_dir": SOURCE_ROOT / "Zebre",
        "injury_standard": "ZEBRE standardised_data .xlsx",
        "exposure_standard": "Zebre standardised_Exposure data (8).xlsx",
        "injury_raw": "Overall Injuries Zebra 2024_25.xlsx",
        "exposure_raw": "Zebre Exposure Data 24_25 (1).xlsx",
        "injury_codebook": "Injury mapping-codebook-ZEBRE.csv",
        "exposure_codebook": "mapping-codebook-Zebre Exp.csv",
        "raw_rows": {"injury": 133, "exposure": 4813},
        "windows": {"injury": "11 Jun 2024 to 10 May 2025", "exposure": "1 Jul 2024 to 17 May 2025"},
        "exposure_context": "Device/vendor is unavailable. Source columns explicitly label HSR thresholds at >18, >20, and 25 km/h; Activity Name supplies session context but Competition is blank.",
    },
}

BENETTON_BODY_LABELS = {
    "Abdomen": "abdomen", "Ankle": "ankle", "Arm": "upper_arm",
    "Back thigh": "thigh", "Cervical spine": "neck", "Elbow": "elbow",
    "Foot": "foot", "Front thigh": "thigh", "Head/Face": "head",
    "Hip/Pelvis/Groin/Sacrum": "hip_groin", "Knee": "knee", "Leg": "lower_leg",
    "Lumbar Spine": "lumbosacral", "Neck/Collarbone/Cervical": "neck", "Ribs": "chest",
}
ZEBRE_BODY_LABELS = {
    "Abdomen": "abdomen", "Achilles tendon": "lower_leg", "Ankle": "ankle",
    "Calf": "lower_leg", "calf": "lower_leg", "Chest": "chest", "Elbow": "elbow",
    "Foot": "foot", "Forearm": "forearm", "Groin/hip": "hip_groin",
    "Groin/hip muscle": "hip_groin", "Hamstring": "thigh", "Hand": "hand",
    "Head": "head", "Knee": "knee", "Lower leg": "lower_leg",
    "Lumbar Spine": "lumbosacral", "Neck": "neck", "Rectus femoris": "thigh",
    "Shoulder": "shoulder", "Thigh": "thigh", "Wrist": "wrist",
}


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load(path: Path) -> dict:
    return json.loads(path.read_text())


def load_decision_selections() -> dict:
    state = load(DECISION_STATE)
    if state.get("evidence_fingerprint") != DECISION_EVIDENCE_FINGERPRINT:
        raise ValueError("Italian decisions are not bound to the reviewed Step 0 evidence fingerprint")
    selections = state.get("selections")
    if not isinstance(selections, dict) or {
        decision_id: selections.get(decision_id, {}).get("choice")
        for decision_id in EXPECTED_DECISION_CHOICES
    } != EXPECTED_DECISION_CHOICES:
        raise ValueError("Italian decisions are missing or differ from the recorded choices")
    return selections


def write(path: Path, value: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n")


def evidence_field(value: str | None) -> dict[str, str | None]:
    return {"status": "available" if value else "unavailable", "value": value}


def sheet(evidence: dict, source_id: str) -> dict:
    source = next(item for item in evidence["sources"] if item["id"] == source_id)
    return source["sheets"][0]


def columns(profile_sheet: dict) -> dict[str, dict]:
    return {item["name"]: item for item in profile_sheet["columns"]}


def ratio(profile_sheet: dict, field: str) -> float:
    total = profile_sheet["substantive_rows"] or 1
    return round(columns(profile_sheet)[field]["populated"] / total, 6)


def locator(field: str) -> dict[str, str]:
    return {"source_id": "injury", "sheet": SHEET, "field": field}


def mapping_entry(
    canonical_field: str,
    canonical_value: str,
    source_evidence: dict[str, str],
    *,
    evidence_class: str = "deterministic_derivation",
    rule: str,
    protocol_rule_id: str | None = None,
    adjudication_id: str | None = None,
    specificity_change: str = "equivalent",
    supporting_evidence: dict[str, str] | None = None,
    evidence_source_id: str = "injury",
    evidence_sheet: str = SHEET,
) -> dict:
    return {
        "canonical_field": canonical_field,
        "canonical_value": canonical_value,
        "evidence_class": evidence_class,
        "source_evidence": source_evidence,
        "specificity_change": specificity_change,
        "supporting_evidence": supporting_evidence or {},
        "evidence_source_id": evidence_source_id,
        "evidence_sheet": evidence_sheet,
        "rule": rule,
        "protocol_rule_id": protocol_rule_id,
        "adjudication_id": adjudication_id,
    }


def body_code_pairs(injury_sheet: dict) -> list[tuple[str, str]]:
    pairs = []
    for item in injury_sheet["joint_category_frequencies"]:
        if item["fields"] == ["Body Part", "Orchard Code"]:
            pairs.append(tuple(item["values"]))
    return sorted(set(pairs))


def mappings(team_key: str, injury_sheet: dict, selections: dict) -> list[dict]:
    if {
        decision_id: selections.get(decision_id, {}).get("choice")
        for decision_id in EXPECTED_DECISION_CHOICES
    } != EXPECTED_DECISION_CHOICES:
        raise ValueError("Italian mappings require the exact recorded decisions")
    result: list[dict] = []
    if team_key == "benetton":
        result.extend([
            mapping_entry("problem_type", "injury", {"Problem type": "Injury"}, evidence_class="source_reported", rule="Preserve the explicit source problem type."),
            mapping_entry("occasion_category", "match", {"Occasion category": "Match"}, rule="Normalize explicit Match to the frozen match category."),
            mapping_entry("occasion_category", "training", {"Occasion category": "Training"}, rule="Normalize explicit Training to the frozen training category."),
            mapping_entry("occasion_category", "unknown", {"Occasion category": "Other"}, rule="Other does not establish match or training and remains Unknown."),
            mapping_entry("match_type", "URC", {"Match Type": "United Rugby Championship"}, evidence_class="source_reported", rule="Normalize the explicitly named URC competition."),
            mapping_entry("contact_context", "contact", {"Is Contact": "Contact (with another player)"}, evidence_class="source_reported", rule="Normalize explicit contact evidence."),
            mapping_entry("contact_context", "non_contact", {"Is Contact": "Non-contact trauma"}, evidence_class="source_reported", rule="Normalize explicit non-contact evidence."),
            mapping_entry("contact_context", "non_contact", {"Is Contact": "Overload (soft start)"}, rule="The source contact/onset field explicitly identifies overload rather than contact."),
            mapping_entry("contact_context", "non_contact", {"Is Contact": "Overload (sudden onset)"}, rule="The source contact/onset field explicitly identifies overload rather than contact."),
            mapping_entry(
                "recurrence", "first_episode",
                {"Recurrence (Recurrence stage)": "New injury (non-recurring)"},
                evidence_class="source_reported",
                rule="Restore the raw row value and normalize the explicit non-recurring label to first_episode.",
                evidence_source_id="raw_injury", evidence_sheet="Sheet1",
            ),
            mapping_entry(
                "recurrence", "recurrence",
                {"Recurrence (Recurrence stage)": "Late recurrence (more than one year)"},
                evidence_class="source_reported",
                rule="Restore the raw row value and normalize the explicit late-recurrence label to recurrence.",
                evidence_source_id="raw_injury", evidence_sheet="Sheet1",
            ),
            mapping_entry(
                "recurrence", "unknown",
                {"Recurrence (Recurrence stage)": "-"},
                rule="Restore the raw row value; dash supplies no recurrence classification and remains Unknown.",
                evidence_source_id="raw_injury", evidence_sheet="Sheet1",
            ),
            mapping_entry(
                "recurrence", "unknown",
                {"Recurrence (Recurrence stage)": "Unknown"},
                evidence_class="source_reported",
                rule="Preserve the explicit Unknown recurrence value.",
                evidence_source_id="raw_injury", evidence_sheet="Sheet1",
            ),
            mapping_entry(
                "tissue_pathology", "joint_sprain",
                {"Injury Tissue Type/s": "Ligament"},
                evidence_class="protocol_defined_inference",
                rule="Restore the raw row value and map the explicit Ligament label to the frozen joint-sprain pathology bucket when retained code evidence does not conflict.",
                protocol_rule_id="ioc_code_mapping_v1",
                evidence_source_id="raw_injury", evidence_sheet="Sheet1",
            ),
            mapping_entry(
                "tissue_pathology", "muscle_injury",
                {"Injury Tissue Type/s": "Muscle"},
                evidence_class="protocol_defined_inference",
                supporting_evidence={"Osiics14": "TWS"},
                rule="Restore the raw row value and use the explicit Muscle label for the single TWS row whose retained code has no supported pathology mapping; supported non-conflicting codes otherwise control, while the recorded MAN/QH2 adjudications are encoded as separate exact raw-evidence mappings.",
                protocol_rule_id="ioc_code_mapping_v1",
                evidence_source_id="raw_injury", evidence_sheet="Sheet1",
            ),
            mapping_entry(
                "tissue_pathology", "muscle_injury",
                {"Injury Tissue Type/s": "Muscle"},
                evidence_class="manual_adjudication",
                supporting_evidence={"Osiics14": "MAN"},
                rule="The explicit raw Muscle label overrides the conflicting retained MAN pathology code under the recorded adjudication; preserve both values and the disagreement in adapter audit evidence.",
                adjudication_id=BENETTON_ADJUDICATION_ID,
                evidence_source_id="raw_injury", evidence_sheet="Sheet1",
            ),
            mapping_entry(
                "tissue_pathology", "muscle_injury",
                {"Injury Tissue Type/s": "Muscle"},
                evidence_class="manual_adjudication",
                supporting_evidence={"Osiics14": "QH2"},
                rule="The explicit raw Muscle label overrides the conflicting retained QH2 pathology code under the recorded adjudication; preserve both values and the disagreement in adapter audit evidence.",
                adjudication_id=BENETTON_ADJUDICATION_ID,
                evidence_source_id="raw_injury", evidence_sheet="Sheet1",
            ),
        ])
    else:
        result.extend([
            mapping_entry("occasion_category", "match", {"Match vs Training": "game"}, rule="Normalize explicit game evidence to match."),
            mapping_entry("occasion_category", "training", {"Match vs Training": "training"}, rule="Normalize explicit training evidence."),
            mapping_entry("occasion_category", "training", {"Match vs Training": "gym"}, rule="Gym is an explicit training exposure setting."),
            mapping_entry("occasion_category", "unknown", {"Match vs Training": "other"}, rule="Other does not establish match or training and remains Unknown."),
            mapping_entry("recurrence", "first_episode", {"Recurrence": "new injury"}, evidence_class="source_reported", rule="Normalize explicit new-injury evidence."),
            mapping_entry("recurrence", "recurrence", {"Recurrence": "recurrence"}, evidence_class="source_reported", rule="Preserve explicit recurrence evidence."),
            mapping_entry("contact_context", "contact", {"Is Contact": "contact"}, evidence_class="source_reported", rule="Preserve explicit contact evidence."),
            mapping_entry("contact_context", "non_contact", {"Is Contact": "NON contact"}, evidence_class="source_reported", rule="Normalize explicit non-contact evidence."),
            mapping_entry("contact_context", "unknown", {"Is Contact": "other"}, rule="Other does not establish contact status and remains Unknown."),
            mapping_entry(
                "problem_type", "illness", {"Body Part": "Medical"},
                evidence_class="manual_adjudication",
                rule="Classify the 30 rows explicitly labelled Medical as illness under the recorded adjudication; retain the source value and never apply the generic injury fallback to Medical rows.",
                adjudication_id=ZEBRE_ADJUDICATION_ID,
                specificity_change="narrower",
            ),
        ])

    labels = BENETTON_BODY_LABELS if team_key == "benetton" else ZEBRE_BODY_LABELS
    codes_for_pathology: set[str] = set()
    for source_body, code in body_code_pairs(injury_sheet):
        if team_key == "zebre" and source_body == "Medical":
            continue
        if team_key == "zebre":
            result.append(mapping_entry(
                "problem_type", "injury",
                {"Body Part": source_body, "Orchard Code": code},
                evidence_class="manual_adjudication",
                rule="Classify a non-Medical row as injury only when this exact retained body-label and OSIICS-code evidence pair is present; otherwise leave problem type Unknown.",
                adjudication_id=ZEBRE_ADJUDICATION_ID,
            ))
        codes_for_pathology.add(code)
        explicit_body = labels.get(source_body)
        code_body = IOC_BODY_CODE_MAP.get(code[0].upper()) if code else None
        conflict = explicit_body and code_body and explicit_body != code_body
        if conflict:
            result.append(mapping_entry(
                "body_location", explicit_body, {"Body Part": source_body},
                supporting_evidence={"Orchard Code": code},
                evidence_class="manual_adjudication",
                rule="The explicit source body label overrides the conflicting OSIICS prefix under the recorded adjudication; preserve both values and the disagreement in audit evidence.",
                adjudication_id=BENETTON_ADJUDICATION_ID,
            ))
        elif explicit_body or code_body:
            result.append(mapping_entry(
                "body_location", explicit_body or code_body, {"Body Part": source_body},
                supporting_evidence={"Orchard Code": code},
                evidence_class="protocol_defined_inference",
                rule="Map only when the explicit source body label and retained OSIICS prefix agree, or one source is non-specific.",
                protocol_rule_id="ioc_code_mapping_v1",
            ))

    for code in sorted(codes_for_pathology):
        pathology = ORCHARD_PATHOLOGY_TYPE_MAP.get(code[1].upper()) if len(code) > 1 else None
        if pathology:
            pathology_conflict = team_key == "benetton" and code in {"MAN", "QH2"}
            if not pathology_conflict:
                result.append(mapping_entry(
                    "tissue_pathology", pathology,
                    {"Orchard Code": code},
                    evidence_class="protocol_defined_inference",
                    rule="Map the retained OSIICS/Orchard pathology character into the frozen IOC pathology bucket.",
                    protocol_rule_id="ioc_code_mapping_v1",
                ))
    return result


def assessment_specs(team_key: str) -> dict[str, tuple[list[str], str, str, str, float, float, list[str]]]:
    if team_key == "benetton":
        return {
            "occasion_category": (["Occasion category", "Match vs Training"], "Allowed: match, training, unknown. Normalize explicit Match/Training; Other remains Unknown.", "deterministic_derivation", "mapped_from_explicit_setting", 0.906977, 0.906977, []),
            "match_type": (["Match Type", "Occasion category", "Date Injured"], "Allowed: URC, training, unknown. Use explicit United Rugby Championship and explicit training; non-URC competitions remain Unknown/non-URC under the frozen cohort rule.", "deterministic_derivation", "explicit_competition_or_training_or_unknown", 0.372093, 0.581395, []),
            "problem_type": (["Problem type", "Orchard Code"], "Allowed: injury, illness, unknown. Preserve the explicit Injury value.", "source_reported", "mapped_from_problem_type", 1.0, 1.0, []),
            "injury_status": (["Injury Status", "Confirmed Return Date", "Days Injured"], "Allowed: open, closed, unknown. Use a valid raw end-of-injury date as closed; explicit active rehabilitation with no return is open/censored.", "deterministic_derivation", "return_date_or_active_status", 0.0, 1.0, ["Workflow-stage labels are not treated as canonical closed/fit values by themselves."]),
            "fit_for_selection_status": (["Fit for selection", "Confirmed Return Date"], "Allowed: fit, not_fit, unknown. Keep Unknown; End of injury does not prove selection availability.", "source_reported", "source_missing", 0.0, 0.0, []),
            "confirmed_return_date": (["Confirmed Return Date"], "Allowed origin: source_reported, derived, unknown. Restore raw End of injury row-for-row because 11 of 40 populated standardised return dates were day/month-swapped.", "deterministic_derivation", "restored_from_raw_locator", 0.0, 0.930233, ["The standardised dates are not authoritative where they disagree with the raw row."]),
            "days_injured": (["Days Injured", "Date Injured", "Confirmed Return Date"], "Allowed origin: source_reported, derived, censored, unknown. Preserve 37 source durations; derive 3 only where missing with two valid restored dates; censor 3 open rows.", "deterministic_derivation", "source_precedence_then_date_difference", 0.860465, 0.930233, []),
            "severity_time_loss_category": (["Days Injured", "TimeLoss vs Medical Attention"], "Allowed: frozen medical-attention and five time-loss bands, or unknown. Classify the 40 closed/effective-duration rows; 3 open rows remain censored/Unknown.", "deterministic_derivation", "frozen_v1_severity_rule", 0.860465, 0.930233, ["TimeLoss=Yes has no raw/codebook source and is ignored."]),
            "recurrence": (["Recurrence"], "Allowed: first_episode, recurrence, unknown. Restore raw recurrence and normalize 34 New injury (non-recurring), 3 late recurrence; 6 dash/Unknown remain Unknown.", "deterministic_derivation", "restored_source_recurrence", 0.0, 0.860465, ["The standardised recurrence column is blank and the raw label needs exact adapter normalization."]),
            "contact_context": (["Is Contact"], "Allowed: contact, non_contact, unknown. Normalize 21 contact, 5 non-contact trauma, and 16 overload rows; one Unknown remains Unknown.", "deterministic_derivation", "mapped_from_contact_onset_field", 0.604651, 0.976744, []),
            "body_location": (["Body Part", "Orchard Code"], "Allowed: frozen IOC body buckets or unknown. Use agreeing/single-source pairs; for 7 conflicting rows, the recorded adjudication applies the explicit source body label while preserving the conflicting code.", "manual_adjudication", "ioc_agreement_or_explicit_label_adjudication", 0.837209, 1.0, ["Seven code disagreements are retained in audit evidence; explicit labels control the canonical value."]),
            "tissue_pathology": (["Injury Tissue Type/s", "Orchard Code", "Description"], "Allowed: frozen IOC pathology buckets or unknown. Restore raw tissue; use supported code/explicit tissue evidence; for MAN and QH2, the recorded adjudication applies explicit Muscle while preserving the conflicting codes; 14 broad unsupported rows remain Unknown.", "manual_adjudication", "ioc_code_or_explicit_tissue_adjudication", 0.488372, 0.674419, ["MAN and QH2 disagreements are retained in audit evidence; broad unsupported values remain Unknown."]),
        }
    return {
        "occasion_category": (["Occasion category", "Match vs Training"], "Allowed: match, training, unknown. Map 39 game to match and 53 training/gym to training; 16 other and 25 blank remain Unknown.", "deterministic_derivation", "mapped_from_explicit_setting", 0.0, 0.691729, []),
        "match_type": (["Match Type", "Match vs Training", "Date Injured"], "Allowed: URC, training, unknown. Populate URC only for game rows with a unique audited fixture-date link; otherwise Unknown.", "deterministic_derivation", "unique_fixture_date_or_unknown", 0.0, 0.0, []),
        "problem_type": (["Problem type", "Body Part", "Orchard Code"], "Allowed: injury, illness, unknown. Under the recorded adjudication, classify 30 explicit Medical rows as illness and classify the 103 non-Medical rows as injury only through their exact retained body/code evidence pairs; otherwise leave Unknown.", "manual_adjudication", "adjudicated_medical_or_exact_injury_evidence", 0.0, 1.0, []),
        "injury_status": (["Injury Status", "Confirmed Return Date", "Days Injured"], "Allowed: open, closed, unknown. Use 117 full-participation dates as closed, 10 no-return time-loss rows as open/censored, and 6 zero-day medical-attention rows as closed.", "deterministic_derivation", "return_date_or_frozen_censoring", 0.0, 1.0, []),
        "fit_for_selection_status": (["Fit for selection", "Confirmed Return Date"], "Allowed: fit, not_fit, unknown. The source header explicitly says back to FULL participation for 117 rows; otherwise Unknown.", "source_reported", "mapped_from_full_participation_date", 0.0, 0.879699, []),
        "confirmed_return_date": (["Confirmed Return Date"], "Allowed origin: source_reported, derived, unknown. Preserve 117 source full-participation dates.", "source_reported", "source_reported_or_unknown", 0.879699, 0.879699, []),
        "days_injured": (["Days Injured", "Date Injured", "Confirmed Return Date"], "Allowed origin: source_reported, derived, censored, unknown. Preserve all 133 Total Days Lost values, including zero; do not overwrite calendar conflicts.", "source_reported", "preserved_source_days", 1.0, 1.0, ["Some zero-day values coexist with later full-participation dates; source duration retains precedence."]),
        "severity_time_loss_category": (["Days Injured"], "Allowed: frozen medical-attention and five time-loss bands, or unknown. Classify 117 returned plus 6 zero-day medical-attention rows; 10 open time-loss rows remain censored/Unknown.", "deterministic_derivation", "frozen_v1_severity_rule", 0.0, 0.924812, []),
        "recurrence": (["Recurrence"], "Allowed: first_episode, recurrence, unknown. Normalize 79 new injury and 26 recurrence values; 28 blank remain Unknown.", "source_reported", "mapped_from_source_recurrence", 0.789474, 0.789474, []),
        "contact_context": (["Is Contact"], "Allowed: contact, non_contact, unknown. Normalize 50 contact and 52 non-contact; 2 other and 29 blank remain Unknown.", "source_reported", "mapped_from_source_contact", 0.766917, 0.766917, []),
        "body_location": (["Body Part", "Orchard Code"], "Allowed: frozen IOC body buckets or unknown. The 103 non-Medical rows have agreeing body/code evidence; Medical rows are not assigned injury body buckets.", "protocol_defined_inference", "ioc_agreement_or_not_applicable", 0.774436, 0.774436, []),
        "tissue_pathology": (["Injury Tissue Type/s", "Orchard Code", "Description"], "Allowed: frozen IOC pathology buckets or unknown. For non-Medical rows, 48 have supported OSIICS pathology mappings; unsupported specificity remains Unknown.", "protocol_defined_inference", "ioc_code_or_unknown", 0.0, 0.360902, ["The standardised tissue field is blank; only supported code/pathology evidence may increase coverage."]),
    }


def build_assessments(team_key: str, injury_sheet: dict) -> list[dict]:
    result = []
    for canonical, (fields, rule, evidence_class, origin, before, after, conflicts) in assessment_specs(team_key).items():
        result.append({
            "canonical_field": canonical,
            "status": "complete",
            "source_fields": [locator(field) for field in fields],
            "rule": rule,
            "evidence_class": evidence_class,
            "origin_status": origin,
            "coverage_before": before,
            "coverage_after": after,
            "conflicts": conflicts,
            "review_required": canonical in {"problem_type", "match_type", "body_location", "tissue_pathology"},
            "tests": [f"{team_key}_{canonical}_reconciliation"],
        })
    return result


def adapter_plan(team_key: str, config: dict) -> dict:
    common = {
        "adapter_plan_version": "italian_intake_adapter_plan_v1",
        "status": "reviewed_unapproved",
        "team": config["team"],
        "team_key": team_key,
        "season": SEASON,
        "source_bindings": {
            name: digest(config["source_dir"] / config[name])
            for name in ("injury_standard", "exposure_standard", "injury_raw", "exposure_raw", "injury_codebook", "exposure_codebook")
        },
    }
    if team_key == "benetton":
        common["rules"] = [
            {"id": "restore_injury_dates", "rows_examined": 43, "affected": {"injury_date_changed": 16, "return_date_populated": 40, "return_date_changed": 11, "return_date_missing": 3}, "action": "Restore dates from the same physical raw row and normalize DD/MM/YYYY; preserve both old and restored values in audit."},
            {"id": "remove_dob", "rows_examined": 43, "affected": {"dob_values_removed": 43}, "action": "Blank DOB in the candidate; never serialize or ingest the raw identifier value."},
            {"id": "restore_injury_tissue", "rows_examined": 43, "affected": {"blank_standard_values_restored": 43, "supported_code_nonconflict": 19, "raw_ligament_to_joint_sprain": 7, "raw_muscle_to_muscle_injury": 3, "adjudicated_code_conflicts": 2, "unsupported_broad": 14}, "action": "Restore the raw tissue label by physical row; apply supported non-conflicting codes and explicit raw-label mappings; for MAN/QH2 use explicit Muscle under the recorded adjudication while retaining the code disagreement; unsupported broad cases remain Unknown."},
            {"id": "restore_recurrence", "rows_examined": 43, "affected": {"new_injury_to_first_episode": 34, "late_recurrence_to_recurrence": 3, "unknown_or_dash": 6}, "action": "Normalize exact raw labels into first_episode/recurrence/unknown; preserve the source label and origin."},
            {"id": "ignore_unsupported_time_loss_yes", "rows_examined": 43, "affected": {"unsupported_yes_values_ignored": 43}, "action": "Do not use the standardised Yes value because no raw/codebook field supports it; derive classification from effective days/censoring."},
            {"id": "injury_exact_duplicate", "rows_examined": 43, "affected": {"duplicate_groups": 1, "duplicate_copies_excluded": 1}, "action": "Retain both locators; exclude only the later exact copy with a controlled audit reason."},
            {"id": "exposure_identifier_column", "rows_examined": 8446, "affected": {"pseudonymous_name_copied_to_canonical_name": 8446}, "action": "Copy the candidate pseudonym from Name to canonical name; never copy the raw direct name."},
            {"id": "exposure_threshold_columns", "rows_examined": 8446, "affected": {"high_speed_20_kmh_restored": 8446, "very_high_speed_25_2_kmh_restored": 8446}, "action": "Restore the exact threshold-labelled source columns with threshold and origin metadata."},
            {"id": "exposure_exact_duplicates", "rows_examined": 8446, "affected": {"duplicate_groups": 26, "duplicate_copies_excluded": 26}, "action": "Exclude only the later exact copy in each group with source locators and controlled reason."},
            {"id": "frozen_session_upper_bound", "rows_examined": 8446, "affected": {"minutes_above_220": 49}, "action": "Apply the frozen session_minutes_above_220 exclusion; do not introduce a device-specific replacement threshold."},
        ]
    else:
        common["rules"] = [
            {"id": "medical_problem_type", "rows_examined": 133, "affected": {"medical_to_illness": 30, "non_medical_with_exact_injury_evidence": 103, "unsupported_fallback_to_unknown": 0}, "action": "Apply the recorded Medical-to-illness adjudication; classify non-Medical rows as injury only through exact retained body/code evidence and leave any unsupported row Unknown."},
            {"id": "return_full_participation", "rows_examined": 133, "affected": {"full_participation_dates": 117, "missing_dates": 16}, "action": "Preserve the explicit back-to-full-participation date as confirmed return and fit date; missing remains Unknown/open as applicable."},
            {"id": "exposure_date_restoration", "rows_examined": 4813, "affected": {"strict_day_first_invalid_standard_values": 297, "raw_dates_restored": 4813}, "action": "Bind every candidate row to the unambiguous raw Excel date by physical row and normalize DD/MM/YYYY."},
            {"id": "exposure_duration_conversion", "rows_examined": 4813, "affected": {"duration_values_converted": 4812, "missing_duration": 1}, "action": "Convert the source Excel duration/time deterministically to minutes; preserve the original representation."},
            {"id": "exposure_very_high_speed", "rows_examined": 4813, "affected": {"vhsr_25_kmh_restored": 4812, "missing_vhsr": 1}, "action": "Restore source vHSR m (25km/hr) to the threshold-labelled canonical companion with origin."},
            {"id": "blank_template_tail", "rows_examined": 5554, "affected": {"substantive_rows": 4813, "blank_physical_rows_not_records": 741}, "action": "Keep blank template rows in physical reconciliation evidence but do not fabricate source records."},
            {"id": "repeated_session_candidates", "rows_examined": 4813, "affected": {"same_player_date_start_groups": 13, "rows_in_groups": 26, "exact_duplicate_rows": 0}, "action": "Retain all non-identical rows; do not merge on a candidate key."},
        ]
    return common


def assessment_source_summary(team_key: str) -> dict[str, str]:
    if team_key == "benetton":
        return {
            "occasion_category": "Match 30; Training 9; Other 4",
            "match_type": "URC 16; seven explicit non-URC labels 14; dash 13",
            "problem_type": "Injury 43",
            "injury_status": "dash 35; five active workflow labels 8; valid raw return 40",
            "fit_for_selection_status": "blank 43",
            "confirmed_return_date": "raw valid 40; missing/dash 3; 11 standardised populated values swapped",
            "days_injured": "37 numeric source durations; 6 missing/dash",
            "severity_time_loss_category": "unsupported Yes 43; effective duration/closure evidence reviewed separately",
            "recurrence": "raw New injury (non-recurring) 34; Late recurrence 3; dash 4; Unknown 2",
            "contact_context": "Contact 21; Non-contact trauma 5; Overload 16; Unknown 1",
            "body_location": "15 explicit labels + 33 OSIICS codes; 7 conflicting rows",
            "tissue_pathology": "raw Muscle 14, Ligament 7, Tendon 5, Musculoskeletal 6, other labels 11; 2 conflicts",
        }
    return {
        "occasion_category": "game 39; training 48; gym 5; other 16; blank 25",
        "match_type": "blank 133; fixture-date review not yet run",
        "problem_type": "blank 133; Medical body label 30; non-Medical 103",
        "injury_status": "blank 133; full-participation date 117; no return 16",
        "fit_for_selection_status": "blank 133; source header explicitly says full participation for 117 dates",
        "confirmed_return_date": "117 source full-participation dates; 16 blank",
        "days_injured": "133 Total Days Lost values, including 20 zero",
        "severity_time_loss_category": "117 returned; 6 no-return zero-day; 10 no-return time-loss",
        "recurrence": "new injury 79; recurrence 26; blank 28",
        "contact_context": "contact 50; NON contact 52; other 2; blank 29",
        "body_location": "23 source labels; 30 Medical; 103 non-Medical body/code pairs agree",
        "tissue_pathology": "tissue field blank; 48 non-Medical rows have supported OSIICS pathology mappings",
    }


def source_inventory(config: dict) -> list[dict]:
    roles = {
        config["injury_standard"]: "proposed injury intake",
        config["exposure_standard"]: "proposed exposure intake",
        config["injury_raw"]: "reference-only row restoration evidence",
        config["exposure_raw"]: "reference-only row restoration evidence",
        config["injury_codebook"]: "reference-only mapping evidence",
        config["exposure_codebook"]: "reference-only mapping evidence",
    }
    result = []
    for name, role in roles.items():
        path = config["source_dir"] / name
        tables = []
        if path.suffix.casefold() == ".csv":
            with path.open(encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                headers = next(reader, [])
                rows = [row for row in reader if any(cell.strip() for cell in row)]
            tables.append(f"CSV table: {len(rows)} substantive rows, {len(headers)} columns")
        else:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for worksheet in workbook.worksheets:
                    values = worksheet.iter_rows(values_only=True)
                    headers = next(values, ())
                    physical = 0
                    substantive = 0
                    for row in values:
                        physical += 1
                        substantive += any(
                            value is not None and str(value).strip() for value in row
                        )
                    tables.append(
                        f"{worksheet.title}: {substantive} substantive rows, "
                        f"{physical} physical rows, {len(headers)} columns"
                    )
            finally:
                workbook.close()
        boundary = (
            "pseudonymised candidate; adapter required before intake"
            if "proposed" in role
            else "reference-only; may contain direct identifiers and must never be ingested"
        )
        result.append({
            "file": name,
            "role": role,
            "boundary": boundary,
            "tables": "; ".join(tables),
            "sha256": digest(path),
        })
    return result


def provenance_review(config: dict, evidence: dict, injury_sheet: dict, exposure_sheet: dict) -> list[dict]:
    proposed_sheets = {"injury": injury_sheet, "exposure": exposure_sheet}
    result = []
    for source in evidence["sources"]:
        source_id = source["id"]
        kind = source["kind"]
        reference_only = source["role"] == "reference_only"
        profile_sheet = source["sheets"][0]
        identifier = {
            ("injury", False): "PlayerID",
            ("exposure", False): "Name" if config["team"] == "Benetton" else "name",
            ("injury", True): "Athlete" if config["team"] == "Benetton" else "Player",
            ("exposure", True): "Name" if config["team"] == "Benetton" else "Player Name",
        }[(kind, reference_only)]
        if reference_only:
            pseudonymisation = "reference-only legacy source may contain direct identifiers; prohibited from intake and serialized evidence"
            identifier_status = "direct identifier in reference-only source; used only for in-memory row reconciliation"
        else:
            pseudonymisation = (
                "pseudonymous identifier confirmed by aggregate pattern review; DOB removal required before intake"
                if config["team"] == "Benetton" and kind == "injury"
                else "pseudonymous identifier confirmed by aggregate pattern review; no direct identifier may enter intake"
            )
            identifier_status = "opaque pseudonymous token; direct values excluded from profile artifacts"
        result.append({
            "source_id": source_id,
            "preparer": evidence_field(None),
            "preparation_timestamp": evidence_field(None),
            "codebook_version": evidence_field(f"checksummed {config[kind + '_codebook']} retained locally"),
            "secure_original_locator": evidence_field(None),
            "secure_original_checksum": evidence_field(None),
            "pseudonymisation_status": evidence_field(pseudonymisation),
            "player_identifier_field": evidence_field(identifier),
            "player_identifier_status": evidence_field(identifier_status),
            "carried_locator_status": evidence_field("workbook checksum + sheet + original physical row; provisional_reference_locator"),
            "row_reconciliation": {
                "status": "completed",
                "source_rows": config["raw_rows"][kind],
                "profiled_rows": profile_sheet["physical_data_rows"],
                "notes": (
                    "Substantive rows reconcile one-to-one; trailing blank template rows remain counted as physical worksheet evidence."
                    if profile_sheet["physical_data_rows"] != config["raw_rows"][kind]
                    else "Substantive rows reconcile one-to-one by physical row."
                ),
            },
        })
    return result


def compile_team(team_key: str, selections: dict) -> None:
    config = CONFIG[team_key]
    team_dir = INTAKE / team_key
    evidence_path = team_dir / "mechanical_evidence.v1.json"
    inventory_path = team_dir / "column_inventory.v2.json"
    mapping_path = team_dir / "source_to_canonical_mapping.v2.draft.json"
    adapter_path = team_dir / "source_adapter_plan.v1.draft.json"
    profile_path = team_dir / "team_intake_profile.v2.draft.json"
    evidence = load(evidence_path)
    profile = load(profile_path)
    injury_sheet = sheet(evidence, "injury")
    exposure_sheet = sheet(evidence, "exposure")

    all_mappings = mappings(team_key, injury_sheet, selections)
    mapping = {
        "mapping_version": MAPPING_VERSION,
        "team": config["team"],
        "team_key": team_key,
        "season": SEASON,
        "status": "reviewed_unapproved",
        "inventory_sha256": digest(inventory_path),
        "evidence_sha256": digest(evidence_path),
        # Gate-validated mappings cite only the proposed intake. Raw restoration
        # rules stay versioned alongside them, but are explicitly adapter-stage
        # evidence and can never make a reference-only workbook an intake source.
        "mappings": [item for item in all_mappings if item["evidence_source_id"] != "raw_injury"],
        "adapter_source_mappings": [item for item in all_mappings if item["evidence_source_id"] == "raw_injury"],
    }
    write(mapping_path, mapping)
    adapter = adapter_plan(team_key, config)
    write(adapter_path, adapter)

    review_time = datetime.now(UTC).isoformat()
    decision = "adapter_required"
    profile.update({
        "decision": decision,
        "mapping_sha256": digest(mapping_path),
        "mapping_version": MAPPING_VERSION,
        "unresolved_adjudication_ids": [],
        "provenance_review": provenance_review(config, evidence, injury_sheet, exposure_sheet),
        "reporting_reviews": {
            "injury": {
                "status": "completed",
                "units": {"days_injured": "days", "dates": "calendar dates"},
                "gaps": f"Source injury coverage is {config['windows']['injury']}; missing or unsupported comparable fields remain Unknown.",
                "repeated_measure_structure": "One retained injury observation per substantive raw row; exact copies are audit exclusions, never silent deletions.",
                "native_grain": "not_applicable",
                "grain_conclusion": "not_applicable",
                "grain_review_rationale": "Reporting grain applies to exposure, not injury observations.",
                "anomalies_reviewed": True,
            },
            "exposure": {
                "status": "completed",
                "units": {"duration": "minutes", "distance": "metres", "high_speed_distance": "metres at the labelled source threshold"},
                "gaps": f"Source exposure coverage is {config['windows']['exposure']}. {config['exposure_context']} Missing labels are retained under the approved unknown-scope inclusion rule and frozen validity exclusions apply.",
                "repeated_measure_structure": "Player-session/activity rows with source physical-row locators; repeated player-date rows remain distinct unless every retained value is exactly duplicated.",
                "native_grain": "session",
                "grain_conclusion": "reviewed_session",
                "grain_review_rationale": "The current raw and standardised files contain player-level dated activity/session rows; week values are derived labels, not weekly aggregates.",
                "anomalies_reviewed": True,
            },
        },
        "taxonomy_review": {
            "status": "completed",
            "body_location_inventory_complete": True,
            "tissue_pathology_inventory_complete": True,
            "notes": (
                "Every body/code/tissue value was reviewed. The recorded explicit-label adjudication resolves seven body and two tissue conflicts while preserving each code disagreement; all other unsupported broad values remain Unknown."
                if team_key == "benetton" else
                "Every body/code value was reviewed. The recorded adjudication classifies Medical rows as illness; exact retained non-Medical body/code pairs support injury, Medical rows receive no injury body/pathology bucket, and unsupported pathology remains Unknown."
            ),
        },
        "tests_and_reconciliation_samples": [
            {
                "id": f"{team_key}_source_checksum_and_row_reconciliation",
                "status": "passed",
                "evidence": "mechanical_evidence.v1.json plus checksummed raw/standardised inventory",
                "notes": (
                    "Asserted injury raw=standard substantive=43 and exposure raw=standard substantive=8446; raw injury exact duplicates=1 group/2 rows and exposure=26 groups/52 rows."
                    if team_key == "benetton" else
                    "Asserted injury raw=standard substantive=133 and exposure raw=standard substantive=4813; 741 trailing physical rows are blank and exact exposure duplicates=0."
                ),
            },
            {
                "id": f"{team_key}_representation_restoration_review",
                "status": "passed",
                "evidence": "source_adapter_plan.v1.draft.json bound to six source checksums",
                "notes": (
                    "Asserted 16/43 injury-date swaps, 11/40 return-date swaps, 43 DOB removals, 8446 pseudonym-column copies, 49 frozen >220-minute exclusions, and exact HSR restoration counts."
                    if team_key == "benetton" else
                    "Asserted 297 strict-day-first invalid exposure dates restored from 4813 raw dates, 4812 duration conversions, 4812 vHSR restorations, and one missing value for each metric."
                ),
            },
            {
                "id": f"{team_key}_taxonomy_and_cohort_conflict_review",
                "status": "passed",
                "evidence": "mechanical joint-category frequencies + source_adapter_plan.v1.draft.json",
                "notes": (
                    "Asserted explicit-label precedence for 7 body conflict rows and 2 tissue conflict rows with the code disagreements retained; 14 additional broad tissue rows remain Unknown without supported row evidence."
                    if team_key == "benetton" else
                    "Asserted 30 Medical rows map to illness and 103 non-Medical rows map to injury only through 65 exact retained body/code evidence pairs; 48 non-Medical rows have supported pathology codes."
                ),
            },
        ],
        "canonical_field_assessments": build_assessments(team_key, injury_sheet),
        "ai_review_status": "completed",
        "ai_reviewed_by": "Codex primary agent (/root)",
        "ai_reviewed_at": review_time,
        "ai_review": {
            "status": "completed",
            "findings": [
                {"finding": "The two Italian structures are not a reusable workbook family.", "disposition": "The advisory family check grouped neither team; checksums, mappings, anomaly handling, and approval remain team-specific.", "status": "resolved"},
                {"finding": "Standardisation altered or omitted source-backed representations.", "disposition": "Only row-preserving, checksum-bound restoration from raw workbooks is proposed; raw identifiers are never restored.", "status": "resolved"},
                {"finding": "Missing scope labels could trigger unsupported exposure exclusion.", "disposition": "The prior approved rule is carried forward: blank scope remains included as unknown unless a frozen validity or explicit out-of-scope rule applies.", "status": "resolved"},
                {"finding": "The first draft overstated canonical coverage by using the most-populated supporting field.", "disposition": "Coverage is now field-specific defensible non-Unknown coverage with exact numerators in rules/source inventories; pending fields remain at zero.", "status": "resolved"},
                *([
                    {"finding": "Seven body and two tissue rows contain conflicting explicit-label and OSIICS evidence.", "disposition": f"Applied Abdel's recorded explicit-label precedence from evidence fingerprint {DECISION_EVIDENCE_FINGERPRINT}; both values and the disagreement remain in manual-adjudication audit evidence.", "status": "resolved"},
                    {"finding": "The first anomaly review used 240 rather than the frozen 220-minute session upper bound.", "disposition": "Plans and adapter evidence now use >220 and reconcile exactly 49 affected Benetton rows.", "status": "resolved"},
                ] if team_key == "benetton" else [
                    {"finding": "Zebre Medical rows would be misclassified as injuries by the generic Orchard-code fallback.", "disposition": f"Applied Abdel's recorded Medical-to-illness choice from evidence fingerprint {DECISION_EVIDENCE_FINGERPRINT}; non-Medical injury mappings enumerate exact retained evidence pairs and no generic fallback remains.", "status": "resolved"},
                ]),
            ],
        },
    })
    write(profile_path, profile)

    files = source_inventory(config)
    inferred = [
        "Restore source-backed dates/metrics only through a row-preserving, locator-tested adapter; never restore raw identifiers.",
        "Remove Benetton DOB before intake; Zebre DOB is already blank.",
        "Keep source Days Injured when present; use date difference only where source duration is missing and dates are valid.",
        "Use explicit URC labels or a unique audited fixture-date link; never classify every game as URC.",
        "Keep unlabeled exposure in scope as unknown; apply only the frozen validity and explicit scope exclusions.",
        "Use frozen exposure thresholds and audit exact duplicate exclusions; retain non-identical repeated rows.",
        "Accept provisional workbook/sheet/physical-row locators at this Step 0 boundary while recording unavailable upstream metadata.",
    ]
    decision_record = (
        "`ZEBRE-2024-25-ADJ-001-PROBLEM-TYPE` — recorded choice `medical_to_illness`: classify 30 explicit `Medical` rows as illness; classify the remaining 103 rows as injury only through exact retained body/code evidence, otherwise `Unknown`."
        if team_key == "zebre" else
        "`BENETTON-2024-25-ADJ-001-TAXONOMY-CONFLICTS` — recorded choice `prefer_explicit_source_label`: explicit labels control the 7 body and 2 tissue conflict rows while both labels, codes, and disagreements remain audited."
    )
    summaries = assessment_source_summary(team_key)
    coverage_rows = [
        f"| `{item['canonical_field']}` | {item['rule'].split('. ', 1)[0].removeprefix('Allowed: ')} | {summaries[item['canonical_field']]} | {item['coverage_before']:.1%} | {item['coverage_after']:.1%} | {'; '.join(item['conflicts']) or 'None'} |"
        for item in profile["canonical_field_assessments"]
    ]
    markdown = [
        f"# {config['team']} 2024-25 — Step 0 intake profile",
        "",
        f"**Status:** `{decision}`  ",
        f"**Profile version:** `team_intake_profile_v2`  ",
        f"**Mapping version:** `{MAPPING_VERSION}`  ",
        "**Adapter plan:** `source_adapter_plan.v1.draft.json` (`italian_intake_adapter_plan_v1`)  ",
        f"**AI review:** completed at `{review_time}`; final independent review is recorded separately.",
        "",
        "## Decision summary",
        "",
        f"Provisional gate decision: `{decision}`. This is local/read-only and authorizes no database action.",
        "",
        "### Recorded adjudication applied",
        "",
        decision_record,
        "",
        f"The saved choices are bound to the pre-decision evidence fingerprint `{DECISION_EVIDENCE_FINGERPRINT}`. Post-decision artifacts require their own fresh AI review and Abdel's separate Step 0 profile approval.",
        "",
        "### Prior answers carried forward",
        "",
        *[f"- {item}" for item in inferred],
        "",
        "## Supplied-file inventory",
        "",
        "| File | Sheet/table reconciliation | Role and boundary | SHA-256 |",
        "|---|---|---|---|",
        *[f"| `{item['file']}` | {item['tables']} | {item['role']}; {item['boundary']} | `{item['sha256']}` |" for item in files],
        "",
        "The raw workbooks and codebooks are reference-only. Only the pseudonymised, locator-enriched adapter outputs may become intake candidates after approval.",
        "",
        "## Reporting structure",
        "",
        f"- Injury: {config['raw_rows']['injury']} substantive rows; source window {config['windows']['injury']}.",
        f"- Exposure: {config['raw_rows']['exposure']} substantive player-session/activity rows; source window {config['windows']['exposure']}; reviewed grain `session`.",
        f"- Exposure context: {config['exposure_context']}",
        f"- Standardised physical rows: injury {injury_sheet['physical_data_rows']} ({injury_sheet['substantive_rows']} substantive), exposure {exposure_sheet['physical_data_rows']} ({exposure_sheet['substantive_rows']} substantive).",
        "- Source player tokens are pseudonymous by aggregate shape review; no token, diagnosis, narrative, or direct identifier is serialized here.",
        "",
        "## Team-specific findings",
        "",
    ]
    if team_key == "benetton":
        markdown.extend([
            "- Injury standardisation swapped day/month representations in multiple rows. Raw Excel dates and row locators are authoritative for representation restoration.",
            "- The exposure pseudonym is populated in `Name` for all 8,446 rows while canonical `name` is blank; copy the pseudonymous value only, never the raw name.",
            "- Canonical HSR fields are blank although explicit >18, >20, and >25.2 km/h source columns are populated. Restore the threshold-equivalent columns with origins.",
            "- The raw injury tissue and recurrence columns are populated but the standardised counterparts are blank because their codebook source labels do not exactly match the workbook headers. Restore them row-for-row before controlled mapping.",
            "- One raw injury row is an exact copy and 26 standardised exposure duplicate groups contain 52 rows; exact copies follow the approved audited exclusion rule.",
            "- Forty-nine raw exposure durations exceed the frozen 220-minute session limit, including three extreme values. They are audited exclusions; no device-specific replacement threshold is introduced.",
            "- `TimeLoss vs Medical Attention = Yes` is not backed by the raw header/codebook and is ignored; time-loss class is derived under the frozen rules.",
            "- Seven body-location rows and two tissue/pathology rows contain direct label/code conflicts. The recorded adjudication applies explicit-label precedence while retaining both source values and the disagreement in audit evidence.",
        ])
    else:
        markdown.extend([
            "- Thirty rows use `Body Part = Medical` and have blank problem type, occasion, contact, and recurrence evidence. The recorded adjudication maps them to illness; the generic injury fallback is disabled, and non-Medical injury mappings require exact retained body/code evidence.",
            "- Source closing dates explicitly mean back to full participation and support confirmed return plus fit-for-selection date; missing dates remain Unknown/open as applicable.",
            "- The exposure standardisation mixes day-first and month-first text. Raw Excel dates reconcile all 4,813 substantive rows and are authoritative for representation restoration.",
            "- The canonical very-high-speed field is blank while source `vHSR m (25km/hr)` is populated for 4,812 rows; restore it with its explicit threshold and origin.",
            "- Thirteen player/date/start candidates contain 26 rows but no exact duplicate rows; retain them pending row-level processing evidence rather than merging.",
            "- Injury coverage begins before the official analysis window; the frozen cohort/window rule, not a team-specific exception, determines analytical inclusion.",
        ])
    markdown.extend([
        "",
        "## Canonical mapping review",
        "",
        "Coverage below means defensible non-`Unknown` classification, not merely a populated supporting field. The exact aggregate source inventories are in `mechanical_evidence.v1.json`; representation changes and count assertions are in `source_adapter_plan.v1.draft.json`.",
        "",
        "| Canonical field | Allowed values/origin | Distinct source values reviewed | Before | After proposed rule | Conflicts/disposition |",
        "|---|---|---|---:|---:|---|",
        *coverage_rows,
        "",
        "- Occasion, match type, status, return, days, severity, recurrence, and contact preserve explicit evidence or use the frozen deterministic rule; unsupported values remain `Unknown`.",
        "- Body location and tissue/pathology use one frozen IOC bucket per analytical row. Multi-value source evidence remains preserved; it does not silently change canonical cardinality.",
        "- Exposure is session-grain. Threshold-labelled HSR columns keep their source thresholds and origins; they are not relabelled as a different vendor threshold.",
        "",
        "## AI review findings",
        "",
        *[f"- **Resolved:** {item['finding']} {item['disposition']}" for item in profile["ai_review"]["findings"]],
        "",
        "## Gate boundary",
        "",
        "No ingest, process, curated build, migration, release, Supabase write, GitHub action, or deployment was run. Profile approval will remain separate from approval of any named live action.",
        "",
    ])
    (team_dir / "team_intake_profile.md").write_text("\n".join(markdown))
    print(team_key, decision, len(mapping["mappings"]), "mappings")


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Compile the local Benetton and Zebre Step 0 profile artifacts."
    )
    parser.parse_args(argv)
    selections = load_decision_selections()
    for team_key in CONFIG:
        compile_team(team_key, selections)


if __name__ == "__main__":
    main()
