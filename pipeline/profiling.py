"""Local, read-only intake profiling.

This module deliberately has no dependency on the database-backed pipeline.
It emits aggregate evidence only and never emits identifier or free-text values.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import tempfile
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Sequence

from openpyxl import load_workbook


PLAN_VERSION = "team_intake_profiling_plan_v1"
SCANNER_VERSION = "local_intake_profiler_2026-07-12_v3"
EVIDENCE_VERSION = "mechanical_evidence_v1"
FAMILY_CHECK_VERSION = "profile_family_check_v1"
INVENTORY_VERSION = "column_inventory_v2"
PROFILE_VERSION = "team_intake_profile_v2"
MAPPING_VERSION = "source_to_canonical_mapping_v2"

SOURCE_ROLES = {"proposed_intake", "reference_only"}
COLUMN_CLASSES = {"safe_category", "identifier", "free_text", "opaque", "date"}
EVIDENCE_CLASSES = {
    "source_reported",
    "deterministic_derivation",
    "protocol_defined_inference",
    "manual_adjudication",
}
DECISIONS = {"compatible", "adapter_required", "adjudication_required", "protocol_incompatible"}
PROTOCOL_RULE_IDS = {
    "ioc_code_mapping_v1",
    "team_specific_cross_field_v1",
    "frozen_v1_derivation_v1",
}
REQUIRED_ASSESSMENTS = (
    "occasion_category",
    "match_type",
    "problem_type",
    "injury_status",
    "fit_for_selection_status",
    "confirmed_return_date",
    "days_injured",
    "severity_time_loss_category",
    "recurrence",
    "contact_context",
    "body_location",
    "tissue_pathology",
)
CONTROLLED_VALUES: dict[str, set[str] | None] = {
    "occasion_category": {"match", "training", "unknown"},
    "match_type": {"URC", "training", "unknown"},
    "problem_type": {"injury", "illness", "unknown"},
    "injury_status": {"open", "closed", "unknown"},
    "fit_for_selection_status": {"fit", "not_fit", "unknown"},
    "confirmed_return_date": {"source_reported", "derived", "unknown"},
    "days_injured": {"source_reported", "derived", "censored", "unknown"},
    "severity_time_loss_category": {
        "medical_attention", "one_day", "two_to_three_days", "four_to_seven_days",
        "eight_to_twenty_eight_days", "greater_than_twenty_eight_days", "unknown",
    },
    "recurrence": {"first_episode", "recurrence", "unknown"},
    "contact_context": {"contact", "non_contact", "unknown"},
    "body_location": None,
    "tissue_pathology": None,
    "injury_type": None,
    "exposure_grain": {"weekly", "session"},
    "exposure_row_status": {"include", "exclude_ghost_row"},
    "time_loss_classification": {"medical_attention", "time_loss", "unknown"},
}
PROTECTED_ALIAS = re.compile(r"^Team [A-Z]$")
SENSITIVE_HEADER_TOKENS = {
    "name", "surname", "forename", "initial", "initials", "id", "uid", "identifier", "dob",
    "email", "description", "diagnosis", "dx", "narrative", "note", "notes", "comment", "comments",
    "detail", "details", "history", "symptom", "symptoms",
}
SENSITIVE_HEADER_PHRASES = {
    ("date", "of", "birth"), ("birth", "date"), ("free", "text"), ("clinical", "text"),
    ("player", "code"), ("athlete", "code"), ("participant", "code"),
    ("player", "number"), ("athlete", "number"), ("participant", "number"),
}
SENSITIVE_STANDALONE_HEADERS = {("player",), ("athlete",)}
SENSITIVE_COMPACT_HEADERS = {
    "name", "surname", "forename", "initial", "initials", "id", "uid", "identifier", "dob",
    "email", "playerid", "playeruid", "playeridentifier", "playercode", "playernumber",
    "athleteid", "athleteuid", "athleteidentifier", "athletecode", "athletenumber",
    "participantid", "participantuid", "participantidentifier", "participantcode",
    "participantnumber", "dateofbirth", "birthdate", "freetext", "clinicaltext",
    "clinicalnote", "clinicalnotes", "medicalnote", "medicalnotes",
}


class ProfileError(ValueError):
    pass


def _json_bytes(value: Any) -> bytes:
    return (json.dumps(value, indent=2, sort_keys=True, ensure_ascii=False) + "\n").encode()


def _digest_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_json_bytes(value))


def _private_directory(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    try:
        path.chmod(0o700)
    except OSError:
        pass


def _write_private_json_atomic(path: Path, value: Any) -> None:
    data = _json_bytes(value)
    _private_directory(path.parent)
    with tempfile.NamedTemporaryFile(dir=path.parent, prefix=f".{path.name}.", delete=False) as handle:
        temporary = Path(handle.name)
        handle.write(data)
    try:
        temporary.chmod(0o600)
    except OSError:
        pass
    temporary.replace(path)


def _load_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise ProfileError(f"invalid JSON: {path}") from exc


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _header_tokens(value: str) -> tuple[str, ...]:
    separated = re.sub(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])", " ", value)
    return tuple(re.findall(r"[a-z0-9]+", separated.casefold()))


def _sensitive_safe_header(value: str) -> bool:
    tokens = _header_tokens(value)
    compact = "".join(tokens)
    role_identifier = any(
        compact == role or (
            compact.startswith(role)
            and compact[len(role):] in {"id", "uid", "identifier", "code", "number", "no", "name"}
        )
        for role in ("player", "athlete", "participant")
    )
    return bool(
        role_identifier
        or compact in SENSITIVE_COMPACT_HEADERS
        or set(tokens) & SENSITIVE_HEADER_TOKENS
        or tokens in SENSITIVE_STANDALONE_HEADERS
        or any(tokens[index:index + len(phrase)] == phrase
               for phrase in SENSITIVE_HEADER_PHRASES
               for index in range(len(tokens) - len(phrase) + 1))
    )


def _snapshot_and_hash(source: Path, destination: Path) -> str:
    digest = hashlib.sha256()
    with source.open("rb") as reader, destination.open("wb") as writer:
        for chunk in iter(lambda: reader.read(1024 * 1024), b""):
            digest.update(chunk)
            writer.write(chunk)
    return digest.hexdigest()


def _contract_hash(source: dict[str, Any]) -> str:
    policy = {
        "scanner_version": SCANNER_VERSION,
        "id": source["id"],
        "role": source["role"],
        "kind": source["kind"],
        "sheets": source["sheets"],
        "column_classes": source["column_classes"],
        "date_order": source.get("date_order"),
        "duplicate_keys": source.get("duplicate_keys", []),
        "exact_row_duplicates": source.get("exact_row_duplicates", False),
        "joint_category_keys": source.get("joint_category_keys", []),
        "required_metrics": source.get("required_metrics", []),
        "exposure_grain_evidence": source.get("exposure_grain_evidence", {}),
        "anomaly_rules": source.get("anomaly_rules", []),
    }
    return _digest_bytes(json.dumps(policy, sort_keys=True, separators=(",", ":")).encode())


def _validate_plan(plan: dict[str, Any]) -> None:
    allowed_plan_keys = {"plan_version", "team", "team_key", "season", "sources"}
    if set(plan) - allowed_plan_keys:
        raise ProfileError(f"unknown plan keys: {sorted(set(plan) - allowed_plan_keys)}")
    if plan.get("plan_version") != PLAN_VERSION:
        raise ProfileError(f"plan_version must be {PLAN_VERSION}")
    for field in ("team", "team_key", "season"):
        if not isinstance(plan.get(field), str) or not plan[field].strip():
            raise ProfileError(f"plan.{field} is required")
    sources = plan.get("sources")
    if not isinstance(sources, list) or not sources:
        raise ProfileError("plan.sources must be a non-empty list")
    ids: set[str] = set()
    for source in sources:
        if not isinstance(source, dict):
            raise ProfileError("each source must be an object")
        for field in ("id", "role", "kind", "path"):
            if not isinstance(source.get(field), str) or not source[field].strip():
                raise ProfileError(f"source.{field} is required")
        allowed_source_keys = {
            "id", "role", "kind", "path", "sheets", "column_classes", "duplicate_keys",
            "required_metrics", "exposure_grain_evidence", "anomaly_rules",
            "date_order",
            "exact_row_duplicates", "joint_category_keys",
        }
        if set(source) - allowed_source_keys:
            raise ProfileError(f"source {source['id']} has unknown keys: {sorted(set(source) - allowed_source_keys)}")
        if source["id"] in ids:
            raise ProfileError(f"duplicate source id: {source['id']}")
        ids.add(source["id"])
        if source["role"] not in SOURCE_ROLES:
            raise ProfileError(f"invalid source role: {source['role']}")
        if source["kind"] not in {"injury", "exposure", "codebook", "other"}:
            raise ProfileError(f"invalid source kind: {source['kind']}")
        if Path(source["path"]).suffix.lower() != ".xlsx":
            raise ProfileError("profiler v1 accepts .xlsx sources only")
        if not isinstance(source.get("sheets"), list) or not source["sheets"] or not all(
            isinstance(item, str) and item.strip() for item in source["sheets"]
        ):
            raise ProfileError(f"source {source['id']} requires sheets")
        if len(set(source["sheets"])) != len(source["sheets"]):
            raise ProfileError(f"source {source['id']} has duplicate sheets")
        classes = source.get("column_classes")
        if not isinstance(classes, dict) or set(classes) - COLUMN_CLASSES:
            raise ProfileError(f"source {source['id']} has invalid column_classes")
        seen: set[str] = set()
        for class_name in COLUMN_CLASSES:
            columns = classes.get(class_name, [])
            if not isinstance(columns, list) or not all(isinstance(item, str) and item for item in columns):
                raise ProfileError(f"source {source['id']} column class {class_name} must be a list")
            overlap = seen.intersection(columns)
            if overlap:
                raise ProfileError(f"columns assigned to multiple privacy classes: {sorted(overlap)}")
            seen.update(columns)
        sensitive_safe_columns = [
            column for column in classes.get("safe_category", []) if _sensitive_safe_header(column)
        ]
        if sensitive_safe_columns:
            raise ProfileError(
                f"source {source['id']} assigns {len(sensitive_safe_columns)} sensitive-looking header(s) to safe_category"
            )
        if not isinstance(source.get("exact_row_duplicates", False), bool):
            raise ProfileError(f"source {source['id']} exact_row_duplicates must be boolean")
        joint_keys = source.get("joint_category_keys", [])
        if not isinstance(joint_keys, list) or any(
            not isinstance(fields, list) or not fields or len(set(fields)) != len(fields)
            or not all(isinstance(field, str) and field in classes.get("safe_category", []) for field in fields)
            for fields in joint_keys
        ) or len({tuple(fields) for fields in joint_keys}) != len(joint_keys):
            raise ProfileError(f"source {source['id']} joint_category_keys must be unique safe-category field lists")
        date_columns = classes.get("date", [])
        if date_columns and source.get("date_order") not in {"day_first", "month_first"}:
            raise ProfileError(f"source {source['id']} requires date_order day_first or month_first")
        if not date_columns and source.get("date_order") is not None:
            raise ProfileError(f"source {source['id']} date_order requires configured date columns")
        grain = source.get("exposure_grain_evidence", {})
        if not isinstance(grain, dict) or set(grain) - {"weekly_columns", "session_columns"}:
            raise ProfileError(f"source {source['id']} has invalid exposure_grain_evidence")
        if any(
            not isinstance(grain.get(key, []), list)
            or not all(isinstance(item, str) and item for item in grain.get(key, []))
            for key in ("weekly_columns", "session_columns")
        ):
            raise ProfileError(f"source {source['id']} exposure grain columns must be lists")
        for keys in source.get("duplicate_keys", []):
            if not isinstance(keys, list) or not keys or not all(isinstance(item, str) and item for item in keys):
                raise ProfileError("duplicate_keys entries must be non-empty column lists")
        for rule in source.get("anomaly_rules", []):
            if not isinstance(rule, dict) or not all(_clean(rule.get(item)) for item in ("id", "operator")):
                raise ProfileError("anomaly rules require id and operator")
            if rule["operator"] not in {
                "blank", "nonblank", "lt", "gt", "eq", "elapsed_minutes_gt", "duration_minutes_gt",
            }:
                raise ProfileError(f"invalid anomaly operator: {rule['operator']}")
            allowed_rule_keys = {"id", "column", "start_column", "end_column", "operator", "value"}
            if set(rule) - allowed_rule_keys:
                raise ProfileError(f"anomaly rule has unknown keys: {sorted(set(rule) - allowed_rule_keys)}")
            required_rule_keys = {
                "elapsed_minutes_gt": {"start_column", "end_column", "value"},
                "duration_minutes_gt": {"column", "value"},
                "lt": {"column", "value"}, "gt": {"column", "value"}, "eq": {"column", "value"},
                "blank": {"column"}, "nonblank": {"column"},
            }[rule["operator"]]
            if set(rule) != required_rule_keys | {"id", "operator"} or not all(
                key in rule and _clean(rule[key]) for key in required_rule_keys
            ):
                raise ProfileError(f"anomaly rule {rule['id']} is incomplete")


def _value_type(value: Any) -> str:
    if value is None or _clean(value) == "":
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (datetime, date)):
        return "date"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    return "text"


def _parse_date(value: Any, date_order: str) -> tuple[date | None, str | None]:
    if isinstance(value, datetime):
        return value.date(), "excel_datetime"
    if isinstance(value, date):
        return value, "excel_date"
    text = _clean(value)
    ordered_formats = [("yyyy-mm-dd", "%Y-%m-%d")]
    if date_order == "day_first":
        ordered_formats.extend((
            ("dd/mm/yyyy", "%d/%m/%Y"), ("dd/mm/yy", "%d/%m/%y"),
            ("dd-mm-yyyy", "%d-%m-%Y"), ("dd-mm-yy", "%d-%m-%y"),
        ))
    else:
        ordered_formats.extend((
            ("mm/dd/yyyy", "%m/%d/%Y"), ("mm/dd/yy", "%m/%d/%y"),
            ("mm-dd-yyyy", "%m-%d-%Y"), ("mm-dd-yy", "%m-%d-%y"),
        ))
    for label, pattern in ordered_formats:
        try:
            return datetime.strptime(text, pattern).date(), label
        except ValueError:
            pass
    return None, None


def _safe_frequency(values: Sequence[Any]) -> tuple[dict[str, int], int]:
    counts: Counter[str] = Counter()
    redacted = 0
    for value in values:
        text = _clean(value)
        if not text:
            continue
        if PROTECTED_ALIAS.fullmatch(text):
            redacted += 1
            continue
        counts[text] += 1
    return dict(sorted(counts.items())), redacted


def _duration_minutes(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel durations are fractions of a day; ordinary numeric inputs are minutes.
        return float(value) * 1440 if 0 <= float(value) < 1 else float(value)
    text = _clean(value)
    match = re.fullmatch(r"(\d+):(\d{1,2})(?::(\d{1,2}))?", text)
    if not match:
        return None
    first, second, third = (int(item) if item is not None else None for item in match.groups())
    if third is None:
        return first * 60 + second
    return first * 60 + second + third / 60


def _datetime_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = _clean(value)
    for pattern in (
        "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%m/%d/%Y %H:%M:%S",
        "%I:%M:%S %p", "%I:%M %p", "%H:%M:%S", "%H:%M",
    ):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            pass
    return None


def _rule_matches(row: dict[str, Any], rule: dict[str, Any]) -> bool:
    operator = rule["operator"]
    if operator == "elapsed_minutes_gt":
        start = _datetime_value(row.get(rule["start_column"]))
        end = _datetime_value(row.get(rule["end_column"]))
        if not start or not end:
            return False
        if end < start:
            end += timedelta(days=1)
        return (end - start).total_seconds() / 60 > float(rule["value"])
    value = row.get(rule["column"])
    if operator == "duration_minutes_gt":
        minutes = _duration_minutes(value)
        return minutes is not None and minutes > float(rule["value"])
    text = _clean(value)
    if operator == "blank":
        return not text
    if operator == "nonblank":
        return bool(text)
    if operator == "eq":
        return value == rule.get("value") or text == _clean(rule.get("value"))
    try:
        numeric = float(value)
        target = float(rule["value"])
    except (KeyError, TypeError, ValueError):
        return False
    return numeric < target if operator == "lt" else numeric > target


def _scan_sheet(worksheet: Any, source: dict[str, Any]) -> dict[str, Any]:
    values = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(values)
    except StopIteration as exc:
        raise ProfileError(f"empty sheet: {worksheet.title}") from exc
    headers = [_clean(value) for value in raw_headers]
    if not any(headers):
        raise ProfileError(f"sheet has no headers: {worksheet.title}")
    if len(set(item for item in headers if item)) != len([item for item in headers if item]):
        raise ProfileError(f"duplicate or blank headers are not supported: {worksheet.title}")
    headers = [item or f"__blank_column_{index + 1}" for index, item in enumerate(headers)]
    physical_rows: list[dict[str, Any]] = []
    substantive_rows: list[dict[str, Any]] = []
    for raw_row in values:
        row = {header: raw_row[index] if index < len(raw_row) else None for index, header in enumerate(headers)}
        physical_rows.append(row)
        if any(_clean(value) for value in row.values()):
            substantive_rows.append(row)

    configured = source["column_classes"]
    missing_configured = sorted({item for values_ in configured.values() for item in values_} - set(headers))
    if missing_configured:
        raise ProfileError(f"configured columns absent from {worksheet.title}: {missing_configured}")
    anomaly_columns = {
        column
        for rule in source.get("anomaly_rules", [])
        for column in (rule.get("column"), rule.get("start_column"), rule.get("end_column"))
        if column
    }
    missing_anomaly_columns = sorted(anomaly_columns - set(headers))
    if missing_anomaly_columns:
        raise ProfileError(f"anomaly columns absent from {worksheet.title}: {missing_anomaly_columns}")
    column_class = {
        column: next((name for name in COLUMN_CLASSES if column in configured.get(name, [])), "opaque")
        for column in headers
    }
    columns: list[dict[str, Any]] = []
    for column in headers:
        # Column inventories reconcile to the physical sheet denominator;
        # substantive_rows remains separate for category/anomaly analysis.
        column_values = [row[column] for row in physical_rows]
        populated = sum(bool(_clean(value)) for value in column_values)
        types = Counter(_value_type(value) for value in column_values if _value_type(value) != "blank")
        columns.append({
            "name": column,
            "class": column_class[column],
            "populated": populated,
            "blank": len(physical_rows) - populated,
            "observed_types": dict(sorted(types.items())),
        })

    category_frequencies: dict[str, dict[str, int]] = {}
    privacy_redactions = 0
    for column in configured.get("safe_category", []):
        frequencies, redactions = _safe_frequency([row[column] for row in substantive_rows])
        category_frequencies[column] = frequencies
        privacy_redactions += redactions

    identifier_stats: dict[str, dict[str, int]] = {}
    for column in configured.get("identifier", []):
        identifiers = [_clean(row[column]) for row in substantive_rows if _clean(row[column])]
        identifier_counts = Counter(identifiers)
        identifier_stats[column] = {
            "populated": len(identifiers),
            "unique": len(identifier_counts),
            "duplicate_value_count": sum(count > 1 for count in identifier_counts.values()),
            "rows_in_duplicate_values": sum(count for count in identifier_counts.values() if count > 1),
        }

    joint_category_frequencies: list[dict[str, Any]] = []
    for fields in source.get("joint_category_keys", []):
        tuples: Counter[tuple[str, ...]] = Counter()
        for row in substantive_rows:
            values_ = tuple(_clean(row[field]) for field in fields)
            if not all(values_) or any(PROTECTED_ALIAS.fullmatch(value) for value in values_):
                continue
            tuples[values_] += 1
        joint_category_frequencies.extend(
            {
                "source_id": source["id"],
                "sheet": worksheet.title,
                "fields": list(fields),
                "values": list(values_),
                "count": count,
            }
            for values_, count in sorted(tuples.items())
        )

    date_columns: dict[str, Any] = {}
    all_dates: list[date] = []
    for column in configured.get("date", []):
        parsed: list[date] = []
        patterns: Counter[str] = Counter()
        invalid = 0
        for row in substantive_rows:
            if not _clean(row[column]):
                continue
            parsed_value, pattern = _parse_date(row[column], source["date_order"])
            if parsed_value is None:
                invalid += 1
            else:
                parsed.append(parsed_value)
                all_dates.append(parsed_value)
                patterns[pattern or "unknown"] += 1
        date_columns[column] = {
            "start": min(parsed).isoformat() if parsed else None,
            "end": max(parsed).isoformat() if parsed else None,
            "patterns": dict(sorted(patterns.items())),
            "invalid": invalid,
        }

    duplicate_groups: list[dict[str, Any]] = []
    for keys in source.get("duplicate_keys", []):
        absent = sorted(set(keys) - set(headers))
        if absent:
            raise ProfileError(f"duplicate key columns absent from {worksheet.title}: {absent}")
        signatures = Counter(
            tuple(_clean(row[key]) for key in keys)
            for row in substantive_rows
            if all(_clean(row[key]) for key in keys)
        )
        duplicate_groups.append({
            "columns": keys,
            "groups": sum(count > 1 for count in signatures.values()),
            "rows": sum(count for count in signatures.values() if count > 1),
        })
    exact_signatures = Counter(
        tuple(_clean(row[column]) for column in headers) for row in substantive_rows
    ) if source.get("exact_row_duplicates", False) else Counter()

    anomaly_counts = {
        rule["id"]: sum(_rule_matches(row, rule) for row in physical_rows)
        for rule in source.get("anomaly_rules", [])
    }
    required_metrics = source.get("required_metrics", [])
    missing_metrics = sorted(metric for metric in required_metrics if metric not in headers)
    blank_metrics = sorted(
        metric for metric in required_metrics
        if metric in headers and not any(_clean(row[metric]) for row in substantive_rows)
    )
    grain = source.get("exposure_grain_evidence", {})
    configured_grain_columns = set(grain.get("weekly_columns", [])) | set(grain.get("session_columns", []))
    missing_grain_columns = sorted(configured_grain_columns - set(headers))
    if missing_grain_columns:
        raise ProfileError(f"exposure grain evidence columns absent from {worksheet.title}: {missing_grain_columns}")

    def grain_stats(columns_: list[str]) -> list[dict[str, Any]]:
        result = []
        for column in columns_:
            values_ = [_clean(row[column]) for row in substantive_rows if _clean(row[column])]
            counts_ = Counter(values_)
            result.append({
                "column": column,
                "populated": len(values_),
                "distinct": len(counts_),
                "repeated_value_count": sum(count > 1 for count in counts_.values()),
                "rows_in_repeated_values": sum(count for count in counts_.values() if count > 1),
            })
        return result

    weekly_stats = grain_stats(grain.get("weekly_columns", []))
    session_stats = grain_stats(grain.get("session_columns", []))
    weekly_populated = any(item["populated"] for item in weekly_stats)
    session_populated = any(item["populated"] for item in session_stats)
    candidate = "ambiguous"
    if weekly_populated and not session_populated:
        candidate = "weekly"
    elif session_populated and not weekly_populated:
        candidate = "session"
    elif not weekly_populated and not session_populated:
        candidate = "unknown"

    return {
        "name": worksheet.title,
        "dimensions": {"max_row": worksheet.max_row, "max_column": worksheet.max_column},
        "physical_data_rows": len(physical_rows),
        "substantive_rows": len(substantive_rows),
        "columns": columns,
        "category_frequencies": category_frequencies,
        "joint_category_frequencies": joint_category_frequencies,
        "identifier_stats": identifier_stats,
        "reporting_window": {
            "start": min(all_dates).isoformat() if all_dates else None,
            "end": max(all_dates).isoformat() if all_dates else None,
            "date_columns": date_columns,
        },
        "duplicate_groups": duplicate_groups,
        "exact_duplicate_groups": sum(count > 1 for count in exact_signatures.values()),
        "exact_duplicate_rows": sum(count for count in exact_signatures.values() if count > 1),
        "exposure_grain_evidence": {
            "candidate": candidate,
            "confirmed": False,
            "weekly_columns": weekly_stats,
            "session_columns": session_stats,
        },
        "missing_required_metrics": missing_metrics,
        "blank_required_metrics": blank_metrics,
        "anomalies": dict(sorted(anomaly_counts.items())),
        "privacy_redactions": privacy_redactions,
    }


def _scan_xlsx(snapshot: Path, source: dict[str, Any], workbook_loader: Callable[..., Any]) -> dict[str, Any]:
    workbook = workbook_loader(snapshot, read_only=True, data_only=True)
    try:
        missing = sorted(set(source["sheets"]) - set(workbook.sheetnames))
        if missing:
            raise ProfileError(f"source {source['id']} missing sheets: {missing}")
        return {
            "id": source["id"],
            "role": source["role"],
            "kind": source["kind"],
            "structure_contract": {
                "file_format": "xlsx",
                "sheets": list(source["sheets"]),
                "column_classes": source["column_classes"],
                "date_order": source.get("date_order"),
                "duplicate_keys": source.get("duplicate_keys", []),
                "required_metrics": source.get("required_metrics", []),
                "exposure_grain_evidence": source.get("exposure_grain_evidence", {}),
            },
            "sheets": [_scan_sheet(workbook[name], source) for name in source["sheets"]],
        }
    finally:
        workbook.close()


def _build_inventory(plan: dict[str, Any], evidence: dict[str, Any]) -> dict[str, Any]:
    return {
        "inventory_version": INVENTORY_VERSION,
        "team": plan["team"],
        "team_key": plan["team_key"],
        "season": plan["season"],
        "evidence_sha256": _digest_bytes(_json_bytes(evidence)),
        "sources": [{
            "id": source["id"],
            "role": source["role"],
            "kind": source["kind"],
            "sha256": source["sha256"],
            "sheets": [{
                "name": sheet["name"],
                "dimensions": sheet["dimensions"],
                "physical_data_rows": sheet["physical_data_rows"],
                "substantive_rows": sheet["substantive_rows"],
                "columns": sheet["columns"],
            } for sheet in source["sheets"]],
        } for source in evidence["sources"]],
    }


def _mapping_template(plan: dict[str, Any], inventory_sha256: str, evidence_sha256: str) -> dict[str, Any]:
    return {
        "mapping_version": MAPPING_VERSION,
        "team": plan["team"],
        "team_key": plan["team_key"],
        "season": plan["season"],
        "status": "draft_unapproved",
        "inventory_sha256": inventory_sha256,
        "evidence_sha256": evidence_sha256,
        "mappings": [],
    }


def _profile_template(
    plan: dict[str, Any], evidence_sha256: str, inventory_sha256: str, mapping_sha256: str
) -> dict[str, Any]:
    pending_evidence = lambda: {"status": "pending", "value": None}
    return {
        "profile_version": PROFILE_VERSION,
        "team": plan["team"],
        "team_key": plan["team_key"],
        "season": plan["season"],
        "decision": "adjudication_required",
        "evidence_path": "mechanical_evidence.v1.json",
        "evidence_sha256": evidence_sha256,
        "column_inventory_path": "column_inventory.v2.json",
        "column_inventory_sha256": inventory_sha256,
        "mapping_path": "source_to_canonical_mapping.v2.draft.json",
        "mapping_sha256": mapping_sha256,
        "mapping_version": MAPPING_VERSION,
        "ai_review_status": "pending",
        "ai_reviewed_by": None,
        "ai_reviewed_at": None,
        "approval_status": "pending",
        "approved_by": None,
        "approved_at": None,
        "approved_input_sha256s": [],
        "unresolved_adjudication_ids": [],
        "provenance_review": [{
            "source_id": source["id"],
            "preparer": pending_evidence(),
            "preparation_timestamp": pending_evidence(),
            "codebook_version": pending_evidence(),
            "secure_original_locator": pending_evidence(),
            "secure_original_checksum": pending_evidence(),
            "pseudonymisation_status": pending_evidence(),
            "player_identifier_field": pending_evidence(),
            "player_identifier_status": pending_evidence(),
            "carried_locator_status": pending_evidence(),
            "row_reconciliation": {
                "status": "pending", "source_rows": None, "profiled_rows": None, "notes": None,
            },
        } for source in plan["sources"]],
        "reporting_reviews": {
            kind: {
                "status": "pending",
                "units": {},
                "gaps": None,
                "repeated_measure_structure": None,
                "native_grain": None,
                "grain_conclusion": "pending",
                "grain_review_rationale": None,
                "anomalies_reviewed": False,
            } for kind in ("injury", "exposure")
        },
        "taxonomy_review": {
            "status": "pending",
            "body_location_inventory_complete": False,
            "tissue_pathology_inventory_complete": False,
            "notes": None,
        },
        "tests_and_reconciliation_samples": [],
        "ai_review": {"status": "pending", "findings": []},
        "canonical_field_assessments": [{
            "canonical_field": field,
            "status": "incomplete",
            "source_fields": [],
            "rule": None,
            "evidence_class": None,
            "origin_status": None,
            "coverage_before": None,
            "coverage_after": None,
            "conflicts": [],
            "review_required": True,
            "tests": [],
        } for field in REQUIRED_ASSESSMENTS],
    }


def scan_plan(
    plan: dict[str, Any] | Path,
    output_root: Path,
    cache_root: Path,
    workbook_loader: Callable[..., Any] | None = None,
) -> dict[str, Any]:
    """Scan one team's plan and emit deterministic aggregate-only drafts."""
    if isinstance(plan, Path):
        plan_path = plan
        loaded = _load_json(plan_path)
        if not isinstance(loaded, dict):
            raise ProfileError("profiling plan must be a JSON object")
        plan = loaded
        for source in plan.get("sources", []):
            if isinstance(source, dict) and isinstance(source.get("path"), str):
                source_path = Path(source["path"])
                if not source_path.is_absolute():
                    source["path"] = str(plan_path.parent / source_path)
    _validate_plan(plan)
    workbook_loader = workbook_loader or load_workbook
    output_dir = Path(output_root) / plan["team_key"]
    cache_root = Path(cache_root)
    _private_directory(cache_root)
    cache_hits = 0
    cache_misses = 0
    evidence_sources: list[dict[str, Any]] = []

    with tempfile.TemporaryDirectory(prefix="urc-profile-") as temp_dir:
        for index, source in enumerate(plan["sources"]):
            source_path = Path(source["path"])
            if not source_path.is_file():
                raise ProfileError(f"source not found: {source_path}")
            snapshot = Path(temp_dir) / f"{index}.xlsx"
            source_sha = _snapshot_and_hash(source_path, snapshot)
            contract_hash = _contract_hash(source)
            cache_key = f"{source_sha}.{contract_hash}"
            cache_path = cache_root / f"{cache_key}.json"
            scanned = None
            if cache_path.is_file():
                try:
                    cached = _load_json(cache_path)
                    payload = cached.get("payload") if isinstance(cached, dict) else None
                    if (
                        isinstance(payload, dict)
                        and cached.get("source_sha256") == source_sha
                        and cached.get("contract_hash") == contract_hash
                        and cached.get("payload_sha256") == _digest_bytes(_json_bytes(payload))
                    ):
                        scanned = payload
                        cache_hits += 1
                except ProfileError:
                    scanned = None
            if scanned is None:
                scanned = _scan_xlsx(snapshot, source, workbook_loader)
                _write_private_json_atomic(cache_path, {
                    "source_sha256": source_sha,
                    "contract_hash": contract_hash,
                    "payload_sha256": _digest_bytes(_json_bytes(scanned)),
                    "payload": scanned,
                })
                cache_misses += 1
            scanned = dict(scanned)
            scanned["sha256"] = source_sha
            evidence_sources.append(scanned)

    evidence = {
        "evidence_version": EVIDENCE_VERSION,
        "scanner_version": SCANNER_VERSION,
        "plan_version": PLAN_VERSION,
        "team": plan["team"],
        "team_key": plan["team_key"],
        "season": plan["season"],
        "sources": evidence_sources,
    }
    evidence_bytes = _json_bytes(evidence)
    evidence_sha = _digest_bytes(evidence_bytes)
    inventory = _build_inventory(plan, evidence)
    inventory_bytes = _json_bytes(inventory)
    inventory_sha = _digest_bytes(inventory_bytes)
    mapping = _mapping_template(plan, inventory_sha, evidence_sha)
    mapping_bytes = _json_bytes(mapping)
    mapping_sha = _digest_bytes(mapping_bytes)
    profile = _profile_template(plan, evidence_sha, inventory_sha, mapping_sha)

    outputs = {
        "mechanical_evidence": output_dir / "mechanical_evidence.v1.json",
        "column_inventory": output_dir / "column_inventory.v2.json",
        "profile_draft": output_dir / "team_intake_profile.v2.draft.json",
        "mapping_draft": output_dir / "source_to_canonical_mapping.v2.draft.json",
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    outputs["mechanical_evidence"].write_bytes(evidence_bytes)
    outputs["column_inventory"].write_bytes(inventory_bytes)
    outputs["mapping_draft"].write_bytes(mapping_bytes)
    outputs["profile_draft"].write_bytes(_json_bytes(profile))
    return {
        "status": "PASS",
        "team": plan["team"],
        "team_key": plan["team_key"],
        "season": plan["season"],
        "sources": evidence_sources,
        "source_count": len(evidence_sources),
        "cache_hits": cache_hits,
        "cache_misses": cache_misses,
        "outputs": {key: str(value) for key, value in outputs.items()},
    }


def _controlled_values() -> dict[str, set[str]]:
    values = {key: set(item or ()) for key, item in CONTROLLED_VALUES.items()}
    taxonomy = Path(__file__).resolve().parent.parent / "docs" / "IOC_TAXONOMY_BUCKETS.csv"
    if taxonomy.is_file():
        with taxonomy.open(newline="", encoding="utf-8") as handle:
            for row in csv.DictReader(handle):
                domain = _clean(row.get("domain"))
                key = _clean(row.get("bucket_key"))
                if domain == "body_location" and key:
                    values["body_location"].add(key)
                elif domain in {"tissue_pathology", "injury_type"} and key:
                    values["tissue_pathology"].add(key)
                    values["injury_type"].add(key)
    values["body_location"].update({"unknown", "unspecified"})
    values["tissue_pathology"].update({"unknown", "nonspecific"})
    values["injury_type"].update({"unknown", "nonspecific"})
    return values


def _add(errors: list[dict[str, str]], code: str, message: str) -> None:
    errors.append({"code": code, "message": message})


def _object_shape(
    value: Any, allowed: set[str], required: set[str], path: str, errors: list[dict[str, str]]
) -> bool:
    if not isinstance(value, dict):
        _add(errors, "invalid_schema", f"{path} must be an object")
        return False
    unknown = set(value) - allowed
    missing = required - set(value)
    if unknown:
        _add(errors, "unknown_schema_key", f"{path} contains {len(unknown)} unknown key(s)")
    if missing:
        _add(errors, "missing_schema_key", f"{path} is missing {len(missing)} required key(s)")
    return not unknown and not missing


def _list(value: Any, path: str, errors: list[dict[str, str]]) -> list[Any]:
    if not isinstance(value, list):
        _add(errors, "invalid_schema", f"{path} must be a list")
        return []
    return value


def _validate_shapes(
    evidence: dict[str, Any], profile: dict[str, Any], mapping: dict[str, Any], inventory: dict[str, Any]
) -> list[dict[str, str]]:
    errors: list[dict[str, str]] = []
    evidence_keys = {"evidence_version", "scanner_version", "plan_version", "team", "team_key", "season", "sources"}
    source_keys = {"id", "role", "kind", "sha256", "structure_contract", "sheets"}
    structure_keys = {
        "file_format", "sheets", "column_classes", "date_order", "duplicate_keys",
        "required_metrics", "exposure_grain_evidence",
    }
    sheet_keys = {
        "name", "dimensions", "physical_data_rows", "substantive_rows", "columns",
        "category_frequencies", "joint_category_frequencies", "identifier_stats", "reporting_window", "duplicate_groups",
        "exact_duplicate_groups", "exact_duplicate_rows",
        "exposure_grain_evidence", "missing_required_metrics", "blank_required_metrics",
        "anomalies", "privacy_redactions",
    }
    column_keys = {"name", "class", "populated", "blank", "observed_types"}
    _object_shape(evidence, evidence_keys, evidence_keys, "evidence", errors)
    if not all(isinstance(evidence.get(field), str) for field in evidence_keys - {"sources"}):
        _add(errors, "invalid_schema_type", "evidence identity/version fields must be strings")
    for source in _list(evidence.get("sources"), "evidence.sources", errors):
        if not _object_shape(source, source_keys, source_keys, "evidence.source", errors):
            continue
        if not all(
            isinstance(source.get(field), str)
            for field in source_keys - {"sheets", "structure_contract"}
        ):
            _add(errors, "invalid_schema_type", "evidence source fields must be strings")
        contract = source.get("structure_contract")
        if _object_shape(contract, structure_keys, structure_keys, "evidence.source.structure_contract", errors):
            if contract.get("file_format") != "xlsx" or not isinstance(contract.get("sheets"), list) \
                    or not isinstance(contract.get("column_classes"), dict) \
                    or not isinstance(contract.get("duplicate_keys"), list) \
                    or not isinstance(contract.get("required_metrics"), list) \
                    or not isinstance(contract.get("exposure_grain_evidence"), dict):
                _add(errors, "invalid_schema_type", "evidence structure contract has invalid types")
        for sheet in _list(source.get("sheets"), "evidence.source.sheets", errors):
            if not _object_shape(sheet, sheet_keys, sheet_keys, "evidence.sheet", errors):
                continue
            if not isinstance(sheet.get("name"), str) or not all(
                isinstance(sheet.get(field), int)
                for field in (
                    "physical_data_rows", "substantive_rows", "privacy_redactions",
                    "exact_duplicate_groups", "exact_duplicate_rows",
                )
            ):
                _add(errors, "invalid_schema_type", "evidence sheet names/counts have invalid types")
            _object_shape(sheet.get("dimensions"), {"max_row", "max_column"}, {"max_row", "max_column"}, "evidence.sheet.dimensions", errors)
            for column in _list(sheet.get("columns"), "evidence.sheet.columns", errors):
                _object_shape(column, column_keys, column_keys, "evidence.column", errors)
                if isinstance(column, dict):
                    if not isinstance(column.get("name"), str) or column.get("class") not in COLUMN_CLASSES or not all(
                        isinstance(column.get(field), int) for field in ("populated", "blank")
                    ):
                        _add(errors, "invalid_schema_type", "evidence column fields have invalid types")
                    _object_shape(column.get("observed_types"), {"boolean", "date", "number", "text"}, set(), "evidence.column.observed_types", errors)
            frequencies = sheet.get("category_frequencies")
            if not isinstance(frequencies, dict) or any(
                not isinstance(items, dict) or any(not isinstance(count, int) for count in items.values())
                for items in frequencies.values()
            ):
                _add(errors, "invalid_schema", "evidence category frequencies must contain integer counts")
            joint_keys = {"source_id", "sheet", "fields", "values", "count"}
            for joint in _list(sheet.get("joint_category_frequencies"), "evidence.joint_category_frequencies", errors):
                if _object_shape(joint, joint_keys, joint_keys, "evidence.joint_category_frequency", errors):
                    if joint["source_id"] != source["id"] or joint["sheet"] != sheet["name"]:
                        _add(errors, "invalid_schema", "joint category frequency scope does not match its source and sheet")
                    if not isinstance(joint["source_id"], str) or not isinstance(joint["sheet"], str) \
                            or not isinstance(joint["fields"], list) or not isinstance(joint["values"], list) \
                            or len(joint["fields"]) != len(joint["values"]) \
                            or not all(isinstance(value, str) for value in joint["fields"] + joint["values"]) \
                            or not isinstance(joint["count"], int):
                        _add(errors, "invalid_schema_type", "joint category frequency fields have invalid types")
            identifier_stats = sheet.get("identifier_stats")
            if not isinstance(identifier_stats, dict):
                _add(errors, "invalid_schema", "evidence identifier_stats must be an object")
            else:
                stats_keys = {"populated", "unique", "duplicate_value_count", "rows_in_duplicate_values"}
                for stats in identifier_stats.values():
                    _object_shape(stats, stats_keys, stats_keys, "evidence.identifier_stats", errors)
            window = sheet.get("reporting_window")
            if _object_shape(window, {"start", "end", "date_columns"}, {"start", "end", "date_columns"}, "evidence.reporting_window", errors):
                date_columns = window["date_columns"]
                if not isinstance(date_columns, dict):
                    _add(errors, "invalid_schema", "evidence date_columns must be an object")
                else:
                    date_keys = {"start", "end", "patterns", "invalid"}
                    for item in date_columns.values():
                        _object_shape(item, date_keys, date_keys, "evidence.date_column", errors)
            for duplicate in _list(sheet.get("duplicate_groups"), "evidence.duplicate_groups", errors):
                _object_shape(duplicate, {"columns", "groups", "rows"}, {"columns", "groups", "rows"}, "evidence.duplicate_group", errors)
            grain = sheet.get("exposure_grain_evidence")
            grain_keys = {"candidate", "confirmed", "weekly_columns", "session_columns"}
            if _object_shape(grain, grain_keys, grain_keys, "evidence.exposure_grain", errors):
                stat_keys = {"column", "populated", "distinct", "repeated_value_count", "rows_in_repeated_values"}
                for side in ("weekly_columns", "session_columns"):
                    for item in _list(grain[side], f"evidence.exposure_grain.{side}", errors):
                        if _object_shape(item, stat_keys, stat_keys, "evidence.exposure_grain.column", errors):
                            if not isinstance(item["column"], str) or not all(
                                isinstance(item[field], int) for field in stat_keys - {"column"}
                            ):
                                _add(errors, "invalid_schema_type", "exposure grain statistics have invalid types")
            for dynamic in ("anomalies",):
                if not isinstance(sheet.get(dynamic), dict) or any(not isinstance(count, int) for count in sheet[dynamic].values()):
                    _add(errors, "invalid_schema", f"evidence.{dynamic} must contain integer counts")

    inventory_keys = {"inventory_version", "team", "team_key", "season", "evidence_sha256", "sources"}
    inventory_source_keys = {"id", "role", "kind", "sha256", "sheets"}
    inventory_sheet_keys = {"name", "dimensions", "physical_data_rows", "substantive_rows", "columns"}
    _object_shape(inventory, inventory_keys, inventory_keys, "inventory", errors)
    if not all(isinstance(inventory.get(field), str) for field in inventory_keys - {"sources"}):
        _add(errors, "invalid_schema_type", "inventory identity/version fields must be strings")
    for source in _list(inventory.get("sources"), "inventory.sources", errors):
        if not _object_shape(source, inventory_source_keys, inventory_source_keys, "inventory.source", errors):
            continue
        for sheet in _list(source.get("sheets"), "inventory.source.sheets", errors):
            if not _object_shape(sheet, inventory_sheet_keys, inventory_sheet_keys, "inventory.sheet", errors):
                continue
            _object_shape(sheet.get("dimensions"), {"max_row", "max_column"}, {"max_row", "max_column"}, "inventory.sheet.dimensions", errors)
            for column in _list(sheet.get("columns"), "inventory.sheet.columns", errors):
                _object_shape(column, column_keys, column_keys, "inventory.column", errors)

    mapping_keys = {
        "mapping_version", "team", "team_key", "season", "status", "inventory_sha256",
        "evidence_sha256", "mappings",
    }
    mapping_entry_keys = {
        "canonical_field", "canonical_value", "evidence_class", "source_evidence",
        "specificity_change", "supporting_evidence", "evidence_source_id", "evidence_sheet",
        "rule", "protocol_rule_id", "adjudication_id",
    }
    _object_shape(mapping, mapping_keys, mapping_keys, "mapping", errors)
    if not all(isinstance(mapping.get(field), str) for field in mapping_keys - {"mappings"}):
        _add(errors, "invalid_schema_type", "mapping identity/version fields must be strings")
    for entry in _list(mapping.get("mappings"), "mapping.mappings", errors):
        if _object_shape(entry, mapping_entry_keys, mapping_entry_keys, "mapping.entry", errors):
            if not all(isinstance(entry.get(field), str) for field in (
                "canonical_field", "canonical_value", "evidence_class", "specificity_change", "rule",
                "evidence_source_id", "evidence_sheet",
            )) or not all(
                entry.get(field) is None or isinstance(entry.get(field), str)
                for field in ("protocol_rule_id", "adjudication_id")
            ) or not isinstance(entry.get("source_evidence"), dict) \
                    or not isinstance(entry.get("supporting_evidence"), dict):
                _add(errors, "invalid_schema_type", "mapping entry fields have invalid types")

    profile_keys = {
        "profile_version", "team", "team_key", "season", "decision", "evidence_path",
        "evidence_sha256", "column_inventory_path", "column_inventory_sha256", "mapping_path",
        "mapping_sha256", "mapping_version", "ai_review_status", "ai_reviewed_by",
        "ai_reviewed_at", "approval_status", "approved_by", "approved_at",
        "approved_input_sha256s", "unresolved_adjudication_ids", "provenance_review",
        "reporting_reviews", "taxonomy_review", "tests_and_reconciliation_samples", "ai_review",
        "canonical_field_assessments",
    }
    _object_shape(profile, profile_keys, profile_keys, "profile", errors)
    if not all(isinstance(profile.get(field), str) for field in {
        "profile_version", "team", "team_key", "season", "decision", "evidence_path",
        "evidence_sha256", "column_inventory_path", "column_inventory_sha256",
        "ai_review_status", "approval_status",
    }):
        _add(errors, "invalid_schema_type", "profile identity/version/status fields must be strings")
    mapping_bindings = [profile.get(field) for field in ("mapping_path", "mapping_sha256", "mapping_version")]
    if not (
        all(value is None for value in mapping_bindings)
        or all(isinstance(value, str) and value for value in mapping_bindings)
    ):
        _add(errors, "invalid_schema_type", "profile mapping bindings must be all strings or all null")
    evidence_field_keys = {"status", "value"}
    provenance_keys = {
        "source_id", "preparer", "preparation_timestamp", "codebook_version",
        "secure_original_locator", "secure_original_checksum", "pseudonymisation_status",
        "player_identifier_field", "player_identifier_status", "carried_locator_status",
        "row_reconciliation",
    }
    for item in _list(profile.get("provenance_review"), "profile.provenance_review", errors):
        if not _object_shape(item, provenance_keys, provenance_keys, "profile.provenance", errors):
            continue
        for field in provenance_keys - {"source_id", "row_reconciliation"}:
            evidence_field = item.get(field)
            if _object_shape(evidence_field, evidence_field_keys, evidence_field_keys, f"profile.provenance.{field}", errors):
                if not isinstance(evidence_field["status"], str) or not (
                    evidence_field["value"] is None or isinstance(evidence_field["value"], str)
                ):
                    _add(errors, "invalid_schema_type", "provenance evidence fields have invalid types")
        _object_shape(item.get("row_reconciliation"), {"status", "source_rows", "profiled_rows", "notes"}, {"status", "source_rows", "profiled_rows", "notes"}, "profile.row_reconciliation", errors)
    reporting = profile.get("reporting_reviews")
    if _object_shape(reporting, {"injury", "exposure"}, {"injury", "exposure"}, "profile.reporting_reviews", errors):
        review_keys = {
            "status", "units", "gaps", "repeated_measure_structure", "native_grain",
            "grain_conclusion", "grain_review_rationale", "anomalies_reviewed",
        }
        for kind in ("injury", "exposure"):
            _object_shape(reporting[kind], review_keys, review_keys, f"profile.reporting_reviews.{kind}", errors)
    taxonomy_keys = {"status", "body_location_inventory_complete", "tissue_pathology_inventory_complete", "notes"}
    _object_shape(profile.get("taxonomy_review"), taxonomy_keys, taxonomy_keys, "profile.taxonomy_review", errors)
    test_keys = {"id", "status", "evidence", "notes"}
    for item in _list(profile.get("tests_and_reconciliation_samples"), "profile.tests", errors):
        _object_shape(item, test_keys, test_keys, "profile.test", errors)
    ai_review = profile.get("ai_review")
    if _object_shape(ai_review, {"status", "findings"}, {"status", "findings"}, "profile.ai_review", errors):
        finding_keys = {"finding", "disposition", "status"}
        for item in _list(ai_review["findings"], "profile.ai_review.findings", errors):
            _object_shape(item, finding_keys, finding_keys, "profile.ai_review.finding", errors)
    assessment_keys = {
        "canonical_field", "status", "source_fields", "rule", "evidence_class", "origin_status",
        "coverage_before", "coverage_after", "conflicts", "review_required", "tests",
    }
    assessment_source_field_keys = {"source_id", "sheet", "field"}
    for item in _list(profile.get("canonical_field_assessments"), "profile.assessments", errors):
        if _object_shape(item, assessment_keys, assessment_keys, "profile.assessment", errors):
            for source_field in _list(item.get("source_fields"), "profile.assessment.source_fields", errors):
                if _object_shape(
                    source_field, assessment_source_field_keys, assessment_source_field_keys,
                    "profile.assessment.source_field", errors,
                ) and not all(isinstance(source_field.get(key), str) for key in assessment_source_field_keys):
                    _add(errors, "invalid_schema_type", "assessment source field locators must be strings")
    return errors


def validate_package(
    evidence_path: Path, profile_path: Path, mapping_path: Path, inventory_path: Path
) -> dict[str, Any]:
    """Validate a completed local draft without changing any artifact."""
    errors: list[dict[str, str]] = []
    try:
        evidence = _load_json(Path(evidence_path))
        profile = _load_json(Path(profile_path))
        mapping = _load_json(Path(mapping_path))
        inventory = _load_json(Path(inventory_path))
    except ProfileError as exc:
        return {"status": "FAIL", "errors": [{"code": "invalid_json", "message": str(exc)}]}
    documents = {"evidence": evidence, "profile": profile, "mapping": mapping, "inventory": inventory}
    if not all(isinstance(value, dict) for value in documents.values()):
        return {"status": "FAIL", "errors": [{"code": "invalid_document", "message": "all package files must be objects"}]}
    schema_errors = _validate_shapes(evidence, profile, mapping, inventory)
    if schema_errors:
        return {"status": "FAIL", "errors": schema_errors}

    versions = {
        "evidence": ("evidence_version", EVIDENCE_VERSION),
        "profile": ("profile_version", PROFILE_VERSION),
        "mapping": ("mapping_version", MAPPING_VERSION),
        "inventory": ("inventory_version", INVENTORY_VERSION),
    }
    for name, (field, expected) in versions.items():
        if documents[name].get(field) != expected:
            _add(errors, "version_mismatch", f"{name}.{field} must be {expected}")
    identity = (profile.get("team"), profile.get("team_key"), profile.get("season"))
    for name in ("evidence", "mapping", "inventory"):
        if (documents[name].get("team"), documents[name].get("team_key"), documents[name].get("season")) != identity:
            _add(errors, "identity_mismatch", f"{name} team/team_key/season does not match profile")
    expected_hashes = {
        "evidence_sha256": _digest_bytes(Path(evidence_path).read_bytes()),
        "column_inventory_sha256": _digest_bytes(Path(inventory_path).read_bytes()),
    }
    if profile.get("mapping_sha256") is not None:
        expected_hashes["mapping_sha256"] = _digest_bytes(Path(mapping_path).read_bytes())
    for field, expected in expected_hashes.items():
        if profile.get(field) != expected:
            _add(errors, "checksum_mismatch", f"profile.{field} does not match artifact")
    if mapping.get("evidence_sha256") != expected_hashes["evidence_sha256"]:
        _add(errors, "checksum_mismatch", "mapping evidence checksum does not match")
    if mapping.get("inventory_sha256") != expected_hashes["column_inventory_sha256"]:
        _add(errors, "checksum_mismatch", "mapping inventory checksum does not match")
    if inventory.get("evidence_sha256") != expected_hashes["evidence_sha256"]:
        _add(errors, "checksum_mismatch", "inventory evidence checksum does not match")
    expected_inventory = _build_inventory(
        {"team": evidence["team"], "team_key": evidence["team_key"], "season": evidence["season"]},
        evidence,
    )
    if inventory != expected_inventory:
        _add(errors, "inventory_content_mismatch", "inventory does not exactly match mechanical evidence")

    if profile.get("approval_status") != "pending" or profile.get("approved_by") is not None \
            or profile.get("approved_at") is not None or profile.get("approved_input_sha256s") != []:
        _add(errors, "unsafe_approval_state", "local draft must not claim approval")
    decision = profile.get("decision")
    unresolved = profile.get("unresolved_adjudication_ids")
    if decision not in DECISIONS:
        _add(errors, "invalid_decision", "profile decision is not controlled")
    if not isinstance(unresolved, list) or any(not _clean(item) for item in unresolved):
        _add(errors, "invalid_adjudications", "unresolved adjudications must be nonblank identifiers")
        unresolved = []
    if decision == "adjudication_required" and not unresolved:
        _add(errors, "adjudication_inconsistent", "adjudication_required needs unresolved identifiers")
    if decision in {"compatible", "adapter_required"} and unresolved:
        _add(errors, "adjudication_inconsistent", "accepted decisions cannot retain unresolved identifiers")
    profile_mapping_bindings = [profile.get(field) for field in ("mapping_path", "mapping_sha256", "mapping_version")]
    if decision == "compatible" and not mapping.get("mappings"):
        if any(value is not None for value in profile_mapping_bindings):
            _add(errors, "mapping_binding_inconsistent", "compatible without mappings requires null mapping bindings")
    elif not all(isinstance(value, str) and value for value in profile_mapping_bindings):
        _add(errors, "mapping_binding_inconsistent", "a mapping document requires bound path, checksum, and version")

    provenance = profile.get("provenance_review", [])
    if len(provenance) != len(evidence.get("sources", [])) or {item.get("source_id") for item in provenance} != {
        source.get("id") for source in evidence.get("sources", [])
    }:
        _add(errors, "incomplete_provenance", "provenance must cover every source exactly once")
    provenance_fields = {
        "preparer", "preparation_timestamp", "codebook_version", "secure_original_locator",
        "secure_original_checksum", "pseudonymisation_status", "player_identifier_field",
        "player_identifier_status", "carried_locator_status",
    }
    for item in provenance:
        for field in provenance_fields:
            evidence_field = item[field]
            status, value = evidence_field["status"], evidence_field["value"]
            if status not in {"available", "unavailable"} or (status == "available" and not _clean(value)) \
                    or (status == "unavailable" and value is not None):
                _add(errors, "incomplete_provenance", "provenance evidence must explicitly be available or unavailable")
        reconciliation = item["row_reconciliation"]
        evidence_source = next(
            (source for source in evidence.get("sources", []) if source.get("id") == item.get("source_id")),
            None,
        )
        expected_profiled_rows = sum(
            sheet.get("physical_data_rows", 0) for sheet in (evidence_source or {}).get("sheets", [])
        )
        if reconciliation["status"] != "completed" or not all(
            isinstance(reconciliation[field], int) and reconciliation[field] >= 0
            for field in ("source_rows", "profiled_rows")
        ) or reconciliation.get("profiled_rows") != expected_profiled_rows \
                or not _clean(reconciliation["notes"]):
            _add(errors, "incomplete_provenance", "row reconciliation must be completed with counts and notes")

    reporting = profile.get("reporting_reviews", {})
    for kind in ("injury", "exposure"):
        review = reporting.get(kind, {})
        if review.get("status") != "completed" or not isinstance(review.get("units"), dict) or not review["units"] \
                or any(not _clean(key) or not _clean(value) for key, value in review.get("units", {}).items()) \
                or not all(_clean(review.get(field)) for field in (
                    "gaps", "repeated_measure_structure", "native_grain", "grain_conclusion",
                    "grain_review_rationale",
                )) or review.get("anomalies_reviewed") is not True:
            _add(errors, "incomplete_reporting_review", f"{kind} reporting review is incomplete")
    exposure_review = reporting.get("exposure", {})
    has_exposure = any(
        source.get("role") == "proposed_intake" and source.get("kind") == "exposure"
        for source in evidence.get("sources", [])
    )
    if has_exposure:
        if exposure_review.get("grain_conclusion") not in {"reviewed_weekly", "reviewed_session"}:
            _add(errors, "invalid_grain_conclusion", "exposure grain requires a human-reviewed conclusion")
        elif exposure_review.get("native_grain") != exposure_review["grain_conclusion"].removeprefix("reviewed_"):
            _add(errors, "invalid_grain_conclusion", "exposure native grain and reviewed conclusion disagree")
    elif exposure_review.get("grain_conclusion") != "not_applicable" or exposure_review.get("native_grain") != "not_applicable":
        _add(errors, "invalid_grain_conclusion", "missing exposure source requires not_applicable grain review")
    if reporting.get("injury", {}).get("grain_conclusion") != "not_applicable":
        _add(errors, "invalid_grain_conclusion", "injury grain conclusion must be not_applicable")

    taxonomy = profile.get("taxonomy_review", {})
    if taxonomy.get("status") != "completed" or taxonomy.get("body_location_inventory_complete") is not True \
            or taxonomy.get("tissue_pathology_inventory_complete") is not True or not _clean(taxonomy.get("notes")):
        _add(errors, "incomplete_taxonomy_review", "taxonomy inventories and review must be complete")
    test_samples = profile.get("tests_and_reconciliation_samples", [])
    if not test_samples or any(
        item.get("status") != "passed" or not all(_clean(item.get(field)) for field in ("id", "evidence", "notes"))
        for item in test_samples
    ):
        _add(errors, "incomplete_reconciliation_tests", "at least one passed reconciliation test is required")
    ai_review = profile.get("ai_review", {})
    findings = ai_review.get("findings", [])
    if profile.get("ai_review_status") != "completed" or ai_review.get("status") != "completed" \
            or not _clean(profile.get("ai_reviewed_by")) or not _clean(profile.get("ai_reviewed_at")) \
            or not findings or any(
                item.get("status") != "resolved"
                or not _clean(item.get("finding")) or not _clean(item.get("disposition"))
                for item in findings
            ):
        _add(errors, "incomplete_ai_review", "completed AI review metadata and dispositions are required")

    assessments = profile.get("canonical_field_assessments")
    if not isinstance(assessments, list):
        _add(errors, "incomplete_assessments", "canonical_field_assessments must be a list")
        assessments = []
    counts = Counter(item.get("canonical_field") for item in assessments if isinstance(item, dict))
    if set(counts) != set(REQUIRED_ASSESSMENTS) or any(counts[field] != 1 for field in counts):
        _add(errors, "incomplete_assessments", "exactly one assessment is required for each canonical field")
    proposed_columns = {
        (source["id"], sheet["name"], column["name"])
        for source in evidence.get("sources", []) if source.get("role") == "proposed_intake"
        for sheet in source.get("sheets", []) for column in sheet.get("columns", [])
        if isinstance(column, dict) and isinstance(column.get("name"), str)
    }
    for item in assessments:
        source_fields = item.get("source_fields")
        coverage = (item.get("coverage_before"), item.get("coverage_after"))
        coverage_valid = all(
            isinstance(value, (int, float)) and not isinstance(value, bool) and 0 <= value <= 1
            for value in coverage
        ) and coverage[1] >= coverage[0]
        if item.get("status") != "complete" or not _clean(item.get("rule")) \
                or item.get("evidence_class") not in EVIDENCE_CLASSES or not _clean(item.get("origin_status")) \
                or not isinstance(source_fields, list) or not source_fields \
                or any(
                    not isinstance(field, dict)
                    or set(field) != {"source_id", "sheet", "field"}
                    or not all(isinstance(field.get(key), str) and field[key].strip()
                               for key in ("source_id", "sheet", "field"))
                    or (field["source_id"], field["sheet"], field["field"]) not in proposed_columns
                    for field in source_fields
                ) \
                or not coverage_valid or not isinstance(item.get("conflicts"), list) \
                or any(not isinstance(value, str) for value in item.get("conflicts", [])) \
                or not isinstance(item.get("review_required"), bool) \
                or not isinstance(item.get("tests"), list) or not item["tests"] \
                or any(not _clean(value) for value in item["tests"]):
            _add(errors, "incomplete_assessments", "every canonical assessment requires complete structured evidence")
            break
    safe_values: dict[tuple[str, str, str], set[str]] = {}
    joint_values: set[tuple[str, str, tuple[str, ...], tuple[str, ...]]] = set()
    for source in evidence.get("sources", []):
        if source.get("role") != "proposed_intake":
            continue
        for sheet in source.get("sheets", []):
            scope = (source.get("id"), sheet.get("name"))
            for field, frequencies in sheet.get("category_frequencies", {}).items():
                if isinstance(frequencies, dict):
                    safe_values.setdefault((*scope, field), set()).update(str(value) for value in frequencies)
            for joint in sheet.get("joint_category_frequencies", []):
                joint_values.add((
                    joint.get("source_id"), joint.get("sheet"),
                    tuple(joint.get("fields", [])), tuple(joint.get("values", [])),
                ))
    controlled = _controlled_values()
    mappings = mapping.get("mappings")
    if not isinstance(mappings, list):
        _add(errors, "invalid_mapping", "mapping.mappings must be a list")
        mappings = []
    if decision == "adapter_required" and not mappings:
        _add(errors, "adapter_mapping_required", "adapter_required needs at least one mapping")
    for index, entry in enumerate(mappings):
        if not isinstance(entry, dict):
            _add(errors, "invalid_mapping", f"mapping {index} must be an object")
            continue
        field = entry.get("canonical_field")
        value = entry.get("canonical_value")
        if field not in controlled:
            _add(errors, "invalid_canonical_field", f"mapping {index} canonical field is invalid")
        elif value not in controlled[field]:
            _add(errors, "invalid_canonical_value", f"mapping {index} canonical value is invalid")
        if entry.get("evidence_class") not in EVIDENCE_CLASSES:
            _add(errors, "invalid_evidence_class", f"mapping {index} evidence class is invalid")
        source_evidence = entry.get("source_evidence")
        if not isinstance(source_evidence, dict) or not source_evidence or any(
            not isinstance(key, str) or not key.strip() or not isinstance(item_value, str) or not item_value.strip()
            for key, item_value in source_evidence.items()
        ):
            _add(errors, "invalid_mapping", f"mapping {index} requires source_evidence")
            source_evidence = {}
        supporting_evidence = entry.get("supporting_evidence")
        if not isinstance(supporting_evidence, dict) or any(
            not isinstance(key, str) or not key.strip() or not isinstance(item_value, str) or not item_value.strip()
            for key, item_value in supporting_evidence.items()
        ):
            _add(errors, "invalid_mapping", f"mapping {index} supporting_evidence must be an object of strings")
            supporting_evidence = {}
        evidence_source_id = entry.get("evidence_source_id")
        evidence_sheet = entry.get("evidence_sheet")
        if not _clean(evidence_source_id) or not _clean(evidence_sheet):
            _add(errors, "invalid_mapping", f"mapping {index} requires nonblank evidence source and sheet locators")
        for source_field, source_value in {**source_evidence, **supporting_evidence}.items():
            scoped_field = (evidence_source_id, evidence_sheet, source_field)
            if scoped_field not in proposed_columns:
                _add(errors, "mapping_source_field_absent", f"mapping {index} evidence field is absent from the named proposed intake sheet")
            elif str(source_value) not in safe_values.get(scoped_field, set()):
                _add(errors, "mapping_source_value_unobserved", f"mapping {index} evidence value is not in the named sheet's safe observed inventory")
        specificity = entry.get("specificity_change")
        if specificity not in {"equivalent", "broader", "narrower"}:
            _add(errors, "invalid_mapping", f"mapping {index} specificity_change is invalid")
        if not _clean(entry.get("rule")):
            _add(errors, "invalid_mapping", f"mapping {index} requires a rule")
        evidence_class = entry.get("evidence_class")
        if evidence_class == "protocol_defined_inference" and entry.get("protocol_rule_id") not in PROTOCOL_RULE_IDS:
            _add(errors, "invalid_protocol_rule", f"mapping {index} protocol rule is not controlled")
        if evidence_class == "manual_adjudication" and not _clean(entry.get("adjudication_id")):
            _add(errors, "invalid_adjudication", f"mapping {index} manual mapping needs an adjudication")
        broad_terms = {
            "injury", "other", "upper limb", "lower limb", "unspecified", "unknown", "nonspecific",
            "non specific", "unspecified crossing",
        }
        evidence_values = [
            re.sub(r"[^a-z0-9]+", " ", value.casefold()).strip()
            for value in source_evidence.values() if isinstance(value, str)
        ]
        all_broad = bool(evidence_values) and all(
            value in broad_terms or any(term in value for term in ("unspecified", "unknown", "nonspecific"))
            for value in evidence_values
        )
        specific_target = value not in {"unknown", "unspecified", "nonspecific", "multiple"}
        clinical_narrowing = field in {"body_location", "tissue_pathology", "injury_type"} \
            and specific_target and (specificity == "narrower" or all_broad)
        if clinical_narrowing:
            combined_evidence = {**source_evidence, **supporting_evidence}
            disjoint = not (set(source_evidence) & set(supporting_evidence))
            observed_together = any(
                set(fields) == set(combined_evidence)
                and tuple(combined_evidence[field_] for field_ in fields) == values_
                and source_id == evidence_source_id
                and sheet_name == evidence_sheet
                for source_id, sheet_name, fields, values_ in joint_values
            )
            supports = bool(supporting_evidence) and disjoint and observed_together
            justified = (
                evidence_class == "protocol_defined_inference"
                and entry.get("protocol_rule_id") in PROTOCOL_RULE_IDS
            ) or (
                evidence_class == "manual_adjudication" and bool(_clean(entry.get("adjudication_id")))
            )
            if not supports or not justified:
                _add(errors, "unsupported_clinical_narrowing", f"mapping {index} clinical narrowing lacks observed support and controlled review evidence")

    # The serialized artifacts themselves must not carry protected aliases.
    serialized = "\n".join(json.dumps(item, sort_keys=True) for item in documents.values())
    if re.search(r'(?<![A-Za-z])Team [A-Z](?![A-Za-z])', serialized):
        _add(errors, "protected_alias_present", "package contains protected alias metadata")
    return {"status": "PASS" if not errors else "FAIL", "errors": errors}


def _family_structure(path: Path) -> dict[str, Any]:
    evidence = _load_json(path)
    if not isinstance(evidence, dict) or evidence.get("evidence_version") != EVIDENCE_VERSION \
            or evidence.get("scanner_version") != SCANNER_VERSION:
        raise ProfileError("family-check requires current mechanical evidence")
    if not all(isinstance(evidence.get(field), str) and evidence[field] for field in ("team_key", "season")):
        raise ProfileError("family-check evidence identity is incomplete")

    exact_sources: list[dict[str, Any]] = []
    shape_sources: list[dict[str, Any]] = []
    sheet_bindings: list[dict[str, Any]] = []
    blockers: set[str] = set()
    sources = evidence.get("sources")
    if not isinstance(sources, list) or not sources:
        raise ProfileError("family-check evidence has no sources")
    source_keys: set[tuple[str, str, str]] = set()
    for source in sorted(sources, key=lambda item: (item.get("id", ""), item.get("role", ""), item.get("kind", ""))):
        if not isinstance(source, dict) or not all(
            isinstance(source.get(field), str) and source[field] for field in ("id", "role", "kind")
        ):
            raise ProfileError("family-check evidence source is invalid")
        source_key = (source["id"], source["role"], source["kind"])
        if source_key in source_keys:
            raise ProfileError("family-check evidence contains duplicate source keys")
        source_keys.add(source_key)
        contract = source.get("structure_contract")
        sheets = source.get("sheets")
        if not isinstance(contract, dict) or not isinstance(sheets, list) or not sheets:
            raise ProfileError("family-check evidence lacks a structure contract")
        sheet_names = [sheet.get("name") for sheet in sheets if isinstance(sheet, dict)]
        if len(sheet_names) != len(sheets) or contract.get("sheets") != sheet_names:
            raise ProfileError("family-check sheet bindings disagree with the scan contract")

        contract_exact = dict(contract)
        contract_shape = {**contract, "sheets": list(range(len(sheet_names)))}
        exact_sheet_rows: list[dict[str, Any]] = []
        shape_sheet_rows: list[dict[str, Any]] = []
        for position, sheet in enumerate(sheets):
            columns = sheet.get("columns")
            window = sheet.get("reporting_window")
            grain = sheet.get("exposure_grain_evidence")
            if not isinstance(columns, list) or not isinstance(window, dict) or not isinstance(grain, dict):
                raise ProfileError("family-check sheet evidence is incomplete")
            column_contract: list[dict[str, Any]] = []
            for column in columns:
                if not isinstance(column, dict) or not isinstance(column.get("name"), str) \
                        or column.get("class") not in COLUMN_CLASSES or not isinstance(column.get("observed_types"), dict):
                    raise ProfileError("family-check column evidence is invalid")
                if column["class"] == "safe_category" and _sensitive_safe_header(column["name"]):
                    raise ProfileError("family-check refuses sensitive-looking safe-category evidence")
                column_contract.append({
                    "name": column["name"],
                    "class": column["class"],
                    "observed_types": sorted(column["observed_types"]),
                })
            dates: list[dict[str, Any]] = []
            date_columns = window.get("date_columns")
            if not isinstance(date_columns, dict):
                raise ProfileError("family-check date evidence is invalid")
            for name, item in sorted(date_columns.items()):
                if not isinstance(item, dict) or not isinstance(item.get("patterns"), dict) \
                        or not isinstance(item.get("invalid"), int):
                    raise ProfileError("family-check date-column evidence is invalid")
                dates.append({"name": name, "patterns": sorted(item["patterns"]), "has_invalid": item["invalid"] > 0})
                if item["invalid"] > 0:
                    blockers.add("invalid_date_values")
            if source["role"] == "proposed_intake" and source["kind"] == "injury" and not date_columns:
                blockers.add("date_evidence_not_configured")
            if sheet.get("missing_required_metrics"):
                blockers.add("missing_required_metrics")
            if sheet.get("blank_required_metrics"):
                blockers.add("blank_required_metrics")
            if sheet.get("privacy_redactions", 0):
                blockers.add("privacy_redactions_present")
            if source["role"] == "proposed_intake" and source["kind"] == "exposure":
                if not contract.get("required_metrics"):
                    blockers.add("required_metrics_not_configured")
                if grain.get("candidate") not in {"weekly", "session"}:
                    blockers.add("unresolved_exposure_grain")
                if grain.get("candidate") == "session" and not date_columns:
                    blockers.add("date_evidence_not_configured")
            sheet_contract = {
                "columns": column_contract,
                "dates": dates,
                "grain_candidate": grain.get("candidate"),
                "weekly_columns": [item.get("column") for item in grain.get("weekly_columns", [])],
                "session_columns": [item.get("column") for item in grain.get("session_columns", [])],
                "has_missing_required_metrics": bool(sheet.get("missing_required_metrics")),
                "has_blank_required_metrics": bool(sheet.get("blank_required_metrics")),
            }
            exact_sheet_rows.append({"name": sheet_names[position], **sheet_contract})
            shape_sheet_rows.append({"position": position, **sheet_contract})
        source_contract = {"id": source["id"], "role": source["role"], "kind": source["kind"]}
        exact_sources.append({**source_contract, "contract": contract_exact, "sheets": exact_sheet_rows})
        shape_sources.append({**source_contract, "contract": contract_shape, "sheets": shape_sheet_rows})
        sheet_bindings.append({"source_id": source["id"], "sheets": sheet_names})

    exact_sha = _digest_bytes(_json_bytes(exact_sources))
    shape_sha = _digest_bytes(_json_bytes(shape_sources))
    return {
        "team_key": evidence["team_key"],
        "season": evidence["season"],
        "evidence_sha256": _digest_bytes(path.read_bytes()),
        "structure_sha256": exact_sha,
        "shape_sha256": shape_sha,
        "sheet_bindings": sheet_bindings,
        "blockers": sorted(blockers),
    }


def family_check(evidence_paths: Sequence[Path], output: Path | None = None) -> dict[str, Any]:
    """Find optional reusable structural families before scientific interpretation."""
    members = [_family_structure(Path(path)) for path in evidence_paths]
    identities = [(member["team_key"], member["season"]) for member in members]
    if len(set(identities)) != len(identities):
        raise ProfileError("family-check requires unique team/season evidence")
    members.sort(key=lambda item: (item["season"], item["team_key"]))

    grouped: dict[str, list[dict[str, Any]]] = {}
    for member in members:
        if not member["blockers"]:
            grouped.setdefault(member["shape_sha256"], []).append(member)
    families: list[dict[str, Any]] = []
    grouped_ids: set[tuple[str, str]] = set()
    for shape_sha, group in sorted(grouped.items()):
        if len(group) < 2:
            continue
        classification = (
            "exact_structure_candidate"
            if len({member["structure_sha256"] for member in group}) == 1
            else "shape_equivalent_candidate"
        )
        for member in group:
            grouped_ids.add((member["team_key"], member["season"]))
        family = {
            "family_id": f"family_{shape_sha[:16]}",
            "classification": classification,
            "members": [
                {"team_key": member["team_key"], "season": member["season"]}
                for member in group
            ],
            "reusable_scope": [
                "source_roles_and_kinds", "sheet_bindings", "ordered_columns_and_privacy_classes",
                "date_configuration", "required_metric_questions", "exposure_grain_evidence_columns",
            ],
            "matched_dimensions": [
                "source_set", "sheet_positions", "ordered_columns", "privacy_classes",
                "observed_type_names", "date_patterns", "required_metrics", "grain_evidence",
            ],
            "must_remain_team_specific": [
                "provenance_and_row_reconciliation", "reporting_windows_gaps_and_units",
                "category_and_clinical_mappings", "anomaly_and_candidate_key_conclusions",
                "grain_confirmation", "canonical_assessments", "decision_ai_review_adjudication_and_approval",
            ],
            "blockers": [],
        }
        if classification == "shape_equivalent_candidate":
            family["sheet_bindings"] = [
                {"team_key": member["team_key"], "season": member["season"], "sources": member["sheet_bindings"]}
                for member in group
            ]
        families.append(family)

    result = {
        "family_check_version": FAMILY_CHECK_VERSION,
        "inputs": [
            {
                "team_key": member["team_key"], "season": member["season"],
                "evidence_sha256": member["evidence_sha256"],
                "structure_sha256": member["structure_sha256"],
            }
            for member in members
        ],
        "candidate_families": families,
        "ungrouped": [
            {
                "team_key": member["team_key"], "season": member["season"],
                "blockers": member["blockers"] or ["no_matching_structure"],
            }
            for member in members
            if (member["team_key"], member["season"]) not in grouped_ids
        ],
        "advisory_only": True,
    }
    if output is not None:
        _write_json(Path(output), result)
    return result


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="python3 -m pipeline.profiling")
    commands = parser.add_subparsers(dest="command", required=True)
    scan = commands.add_parser("scan")
    scan.add_argument("--plan", type=Path, required=True)
    scan.add_argument("--output-root", type=Path, required=True)
    scan.add_argument("--cache-root", type=Path, required=True)
    validate = commands.add_parser("validate")
    validate.add_argument("--evidence", type=Path, required=True)
    validate.add_argument("--profile", type=Path, required=True)
    validate.add_argument("--mapping", type=Path, required=True)
    validate.add_argument("--inventory", type=Path, required=True)
    family = commands.add_parser("family-check")
    family.add_argument("--evidence", type=Path, nargs="+", required=True)
    family.add_argument("--output", type=Path, required=True)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        if args.command == "scan":
            result = scan_plan(args.plan, args.output_root, args.cache_root)
        elif args.command == "validate":
            result = validate_package(args.evidence, args.profile, args.mapping, args.inventory)
        else:
            result = family_check(args.evidence, args.output)
    except ProfileError as exc:
        print(json.dumps({"status": "FAIL", "error": str(exc)}, sort_keys=True))
        return 1
    printable = result
    if args.command == "scan":
        printable = {key: value for key, value in result.items() if key != "sources"}
    print(json.dumps(printable, indent=2, sort_keys=True))
    return 1 if result.get("status") == "FAIL" else 0


if __name__ == "__main__":
    sys.exit(main())
