from __future__ import annotations

import argparse
import contextlib
import csv
import difflib
import getpass
import hashlib
import io
import json
import os
import platform
import re
import shlex
import shutil
import stat
import subprocess
import sys
import tempfile
import time
import uuid
from collections import defaultdict
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from pipeline.season_contracts import (
    YEAR2_2025_26_RELEASE_TUPLE,
    fixture_contract_for,
    fixture_provenance_rows,
    release_contract_for,
    validate_fixture_rows,
    validate_fixture_provenance_binding,
)


LOCATOR_FIELDS = [
    "source_archive_path",
    "source_file_sha256",
    "source_sheet",
    "source_row_number",
    "standardised_file_sha256",
    "standardised_row_number",
    "source_locator_status",
]

UID_FIELDS = ["player_uid", "injury_uid"]

EXPOSURE_LOCATOR_FIELDS = [
    "source_archive_path",
    "source_file_sha256",
    "source_sheet",
    "source_row_number",
    "source_row_sha256",
    "standardised_file_sha256",
    "standardised_row_number",
    "source_locator_status",
]

EXPOSURE_REPORTING_GRAINS = ("weekly", "session")
EXPOSURE_DECLARED_GRAIN_FIELD = "declared_exposure_grain"
EXPOSURE_ORIGIN_FIELDS = ["minutes_total_origin", "distance_total_origin"]

EDINBURGH_URC_OPPONENTS = [
    "benetton",
    "bulls",
    "cardiff",
    "connacht",
    "dragons",
    "glasgow",
    "leinster",
    "lions",
    "munster",
    "ospreys",
    "scarlets",
    "sharks",
    "stormers",
    "ulster",
    "zebre",
]
EDINBURGH_EXPOSURE_SCOPE_RULE_VERSION = "edinburgh_exposure_scope_2026-07-07_v2"
EXPOSURE_CANONICAL_SCHEMA_VERSION = "exposure_core_2026-07-07_v1"
INJURY_PROCESSING_RULE_VERSION = "injury_processing_2026-07-07_v2"
YEAR2_INJURY_ELIGIBILITY_BRIDGE_RULE_VERSION = "urc_2025_26_injury_eligibility_bridge_v1"
INPUT_REPRESENTATION_CORRECTION_RULE_VERSION = "input_representation_correction_2026-07-13_v1"
EXPOSURE_PROCESSING_RULE_VERSION = "exposure_processing_2026-07-07_v1"
URC_OPPONENT_FUZZY_CUTOFF = 0.78

# The team-name to league-alias map is protected research metadata. It lives in a
# Git-ignored local file (backed up in UCD-managed storage), never in code or Git.
TEAM_ALIAS_MAP_PATH = Path(__file__).resolve().parent.parent / "data" / "intake" / "team_alias_map.json"
TEAM_ALIAS_CODEBOOK_PATH = Path("/Users/abdelbabiker/Desktop/URC/codebook- Teams URC.csv")


def load_fixture_team_aliases() -> dict[str, str]:
    if not TEAM_ALIAS_MAP_PATH.exists():
        raise SystemExit(
            f"protected team alias map not found: {TEAM_ALIAS_MAP_PATH}; "
            "restore it from UCD-managed storage (it is intentionally excluded from Git)"
        )
    data = json.loads(TEAM_ALIAS_MAP_PATH.read_text())
    aliases = data.get("fixture_team_aliases")
    if not isinstance(aliases, dict) or not aliases:
        raise SystemExit(
            f"invalid team alias map (expected non-empty 'fixture_team_aliases' object): {TEAM_ALIAS_MAP_PATH}"
        )
    return {str(name): str(alias) for name, alias in aliases.items()}


# Phase 3.5 cohort-signal capture (Adjudication 4, 10 July 2026). Public
# team-name key to look up in load_fixture_team_aliases() for each
# reporting.teams.team_key, so process-intake can resolve "our own team's
# protected alias" without an operator having to type it in by hand. Every
# value here is a public club name (verified live against the real,
# Git-ignored team_alias_map.json 'fixture_team_aliases' keys, 10 July
# 2026); none of these values is itself the protected alias. Covers all 16
# reporting.teams.team_key values (20260709120100_reporting_teams_dimension.sql).
TEAM_KEY_ALIAS_LOOKUP_NAMES: dict[str, str] = {
    "connacht": "Connacht",
    "leinster": "Leinster",
    "munster": "Munster",
    "ulster": "Ulster",
    "cardiff": "Cardiff",
    "dragons": "Dragons RFC",
    "ospreys": "Ospreys",
    "scarlets": "Scarlets",
    "bulls": "Bulls",
    "lions": "Lions",
    "sharks": "Hollywoodbets Sharks",
    "stormers": "DHL Stormers",
    "benetton": "Benetton",
    "zebre": "Zebre Parma",
    "edinburgh": "Edinburgh",
    "glasgow": "Glasgow Warriors",
}


def own_team_alias_for(team_key: str, fixture_team_aliases: dict[str, str]) -> str:
    """Resolve team_key's own protected league alias from the alias map, for
    in-memory comparison only. Callers must never log, print, or persist the
    returned value -- pass it straight into received_in_team_status() and
    discard it.
    """
    lookup_name = TEAM_KEY_ALIAS_LOOKUP_NAMES.get(team_key)
    if not lookup_name:
        raise SystemExit(f"no TEAM_KEY_ALIAS_LOOKUP_NAMES entry configured for team_key {team_key!r}")
    alias = fixture_team_aliases.get(lookup_name)
    if not alias:
        raise SystemExit(
            f"protected team alias map has no fixture_team_aliases entry for {lookup_name!r} "
            f"(team_key {team_key!r})"
        )
    return alias


FIXTURE_DATE_CORRECTIONS = {
    "DHL Stormers v Vodacom Bulls": "2025-02-08",
    "Hollywoodbets Sharks v Emirates Lions": "2025-03-08",
    "Glasgow Warriors v Edinburgh Rugby": "2024-12-22",
    "Leinster v Connacht": "2024-12-21",
    "Ospreys v Scarlets": "2024-12-21",
    "Ulster v Munster": "2024-12-20",
    "Connacht v Ulster": "2024-12-28",
    "Munster v Leinster": "2024-12-27",
    "Vodacom Bulls v Emirates Lions": "2025-02-22",
    "Glasgow Warriors v Connacht": "2025-01-26",
    "Dragons RFC v Glasgow Warriors": "2025-02-16",
    "Zebre Parma v Edinburgh Rugby": "2025-04-25",
}

DUPLICATE_SIGNATURE_FIELDS = [
    "PlayerID",
    "Date Injured",
    "Body Part",
    "Orchard Code",
    "Side",
    "Description",
    "Nature of onset",
    "Illness Code",
    "Injury Tissue Type/s",
]

YEAR2_INJURY_DATE_BASES = frozenset(
    {
        "source_date_within_window",
        "season_attributed_undated",
        "source_date_unparseable",
        "source_date_outside_window",
    }
)
YEAR2_INJURY_BRIDGE_ELIGIBILITY = frozenset(
    {"included_pending_protocol", "review_required"}
)
YEAR2_ALLOWED_ANALYSIS_AUDIT_REASONS = frozenset({"explicit_source_exclusion"})

V13_INTAKE_PROFILE_SCHEMA = "urc_2025_26_v13_signed_intake_profile_v1"
V13_INTAKE_MANIFEST_SCHEMA = "urc_2025_26_v13_signed_intake_manifest_v1"
V13_DATABASE_AUTHORISATION = {
    "database_action_authorised": True,
    "basis": (
        "the exact approval line names the project, database, ingestion, "
        "processing, build and release"
    ),
    "project_ref": "eukkvswaxweenovqqgzr",
    "database": "postgres",
    "actions": ["ingestion", "processing", "build", "release"],
    "approval_line_sha256": (
        "49cd90905a27faf74b0f1d53d80ea2084964ca1b6e36bd7e4b795ee2e69eb542"
    ),
}
V13_V12_ROOT_MANIFEST_SHA256 = (
    "01dd17a82ab1835fd84f2c84048b9e15b4072a4f9bca3b3d3a348817a68d7241"
)
V13_V12_ROOT_FILE_SET_SHA256 = (
    "5ea322d4e246510ce82075f5690ea2ac5715dace31ead35bff9db3bacc6a7abd"
)
V13_V12_PREDECESSOR_OUTPUT_MAP_SHA256 = V13_V12_ROOT_FILE_SET_SHA256
V13_FRESH_REVIEW_SHA256 = (
    "61caebf232f0422f7bd5340609c113b0e0931ab01f15262f75a1e5da860ae1df"
)
V13_FRESH_REVIEWER = {
    "model": "gpt-5.6-sol",
    "reasoning_effort": "xhigh",
    "task": "/root/v13_signer_acceptance",
    "completed_at": "2026-08-22T17:55:08Z",
}
# Non-circular membership proof for the accepted V12 root above. Each tuple is
# display team, injury SHA, exposure SHA and team-manifest SHA, independently
# reconciled to the exact fresh review and V12 root output map.
V13_REVIEWED_V12_TEAMS = {
    "benetton": ("Benetton", "58602000b171e29d0db271eec95b4357508a484602a1e80d95b20d1d1cde4d9b", "6d9fdd02873a2b69c81cd2ce1e6bfe1bd4c82812ae710e52f1809c3ebcb40c61", "61ea454fb28bb98db8f6c6df2ddfef95873ed145c19a7542b442af88b73ee408"),
    "bulls": ("Bulls", "f2a069c5f235d5b82b135de03c38fb96cba645b34b7622853a19d1fef5f53717", "eb6a5f0e67b3b4d074bff58250849d29f88c52d2d5837a79b50aa4b2b737333c", "87278acd651ebd75099b8deade3a4ab761b36b77742ae9d11a96cd87a658a156"),
    "cardiff": ("Cardiff", "011e8c7c6cf84ab34e7425eb4e8d018b88717e9251410d5cffb9c2a438b5e0a7", "17f65d2bc16dca0401f8ff8cee602fa8cd65669ab32158b6d81e13f1a99879c3", "36553787fbe144c6ea002f86443274e70d344f7e4fa5cd97bad2e6225d60223b"),
    "connacht": ("Connacht", "6da2bbcfce008ecc7ccff30bfcc7d1b23fc262865f976d784ced4236bebd4d0d", "65baa074f67f0ce97e9b415079e7f9411831940d2a9dc825c1e1446cd5676cd3", "8870387f076ba9dc1ec95fd833187da7acbb270d19078b5d0fc310b7c8760474"),
    "dragons": ("Dragons", "2c9ee7fe0860f3d7a7063b860f7c5d60e5c088d8504e8d9d58d8c9bb7f410b22", "c71c2065db1b0d93b636e4ba75b4b6ade211e6fd393a9ebae7cefb11dfba089e", "b61f57377151ac6a19c59d96455c47ece140d43a644a59eda4d0b4e8f23d3365"),
    "edinburgh": ("Edinburgh", "214e0eec95972245363f9fd61e3f3cb04b335e1c6243091e2c3b7562ae91ad17", "dce20b6194f1d2500950f642a942cccbc88809b4d029d4644027cc2527c3d189", "8a06d569f722677cc6411e4313b8ad0cd298a03ea379be6f4c97736802f8e572"),
    "glasgow": ("Glasgow", "e1a63d8f7a896c09e7d555f1b9cf9faa704b011af0f50eb49223baf050177f1f", "a0b75e0af48fde6324fee98808ead2368740a7054884ae48e205a740e5f41c56", "3e9a7d08dab191c136032835f0fb595da4a0941e17ce855dd6745738081bd493"),
    "leinster": ("Leinster", "11e0aaba4abc77259d76008909cdd0b42bf2c121a75fd8ba622d5f8b0fc8fa52", "e1e8b4296e3cde67fd820ae36823a01213a73732f62a9a37a9692a2f13c9328a", "2934f73cd087a7f5104c624429e4e43956afb8b1de50945f1ad357207c5f385c"),
    "lions": ("Lions", "21a9288bed1767cd9f25044b166111d8759c5daf845fcc7a3a1e435f85c2acff", "c9b2fc02bdaeb395f04e2a72d2d5b66db9a7585227ec05753473c1ff2ef16a3f", "a6934738df0d63f0411eab05acbe54cbcb05d5be29a2d4bd1cb8a9e1a4ae02bd"),
    "munster": ("Munster", "b7206d5c75ff7a9cd2f15ae82bfe788d47a88ef4602d8ce21cb5255003d71c79", "acd8f2e153d35d985509c73b7fd3af822f875140d3dcfbedf83cd97e35edf43e", "977aead02b09881093f9bd239216a2c8ba8bcc1f604ba7ecdec0e41be3573fa3"),
    "ospreys": ("Ospreys", "870216b1d3570f869d684694c3256116bb31c5ad20ad50b1aa8035415ec529a4", "9d95f1538751a2d9fb9b4430c484c2c83889146c09b602a26fe837a257bcd756", "ce60d08da4f833bb9fd1a409dfc65712e1287008d0a1a5889f5d0ea62c2b2b42"),
    "scarlets": ("Scarlets", "eb21acd209a662782f7c9e1f3b92e77cf8b2fe418758c6e23ff44b59f81fa14b", "6c6ea646c8355f53b50916504a9d72cd3de129d91250757d6e2cab2834c32777", "acc660de0b0805d1255ec754575e63628f35060410e581c122879a4ef7c94005"),
    "sharks": ("Sharks", "5cc526aa901cd3eb167caeaf8c424070e50ff243f6c0c30b7d5e3c4ea8ec9cf8", "3d47447729164800bd776c28127b19b39b59067e90ba6c7dc84eda531f371bbb", "35212d97bafcdf0d5ca74dee188189c359e9976638e0e3d42a64030862b16c43"),
    "stormers": ("Stormers", "4f598ec18e230dae104b2b08e6fb19826629e77673169ac791665a0547bfd1dd", "db087678787e6431f9dbf2bcd78d6dc1913c84a144a35ba7113226f74c0b0df4", "bd96b786545ccddf57ea5f769031dda07d568d4e879dfbe1ab527595bdc417eb"),
    "ulster": ("Ulster", "cbb18d92e8253414ed3abfcf02161e11b80259b258fbb8572827ae40b97c6dc9", "e17efcd5732d1853ce9f2e00c71498748a0d98f4e2ca45433d779441edafde34", "03c221fcf0c26f4fda33065fa42b49b91aba9ea08a2291f564c0e75819565a63"),
    "zebre": ("Zebre", "318bd3303fe5c91e9480ab204f938fa17de77dc7fe4757e3a353e398cfaed95a", "deceea12ad1ab587ae17ba3a669b78d3bff20f07021dee7a266d8f256cacef7f", "53cacfd494f5713de228f3ec071908b5e01998343c428d6505c2d194a7829920"),
}
YEAR2_APPROVED_ROOTS_PATH = Path(__file__).with_name("approved_year2_roots.json")

MISSING_VALUES = {"", "na", "n/a", "null", "none", "unknown", "unspecified/crossing"}

BODY_LOCATION_MAP = {
    "ankle": "ankle",
    "buttock/pelvis": "lumbosacral",
    "chest": "chest",
    "elbow": "elbow",
    "foot": "foot",
    "head": "head",
    "hip/groin": "hip_groin",
    "knee": "knee",
    "lower leg": "lower_leg",
    "lumbar spine": "lumbosacral",
    "neck": "neck",
    "shoulder": "shoulder",
    "thigh": "thigh",
    "thoracic spine": "thoracic_spine",
    "trunk/abdominal": "abdomen",
    "wrist/hand": "unspecified",
}

INJURY_TYPE_MAP = {
    "bruising/ haematoma": "contusion_superficial",
    "concussion/ brain injury": "brain_spinal_cord_injury",
    "disc": "nonspecific",
    "dislocation": "joint_sprain",
    "fracture": "fracture",
    "instability": "chronic_instability",
    "laceration/ abrasion": "laceration",
    "ligament": "joint_sprain",
    "muscle strain/spasm": "muscle_injury",
    "nerve": "peripheral_nerve_injury",
    "organ damage": "internal_organ_trauma",
    "osteoarthritis": "arthritis",
    "osteochondral": "cartilage_injury",
    "other pain/ unspecified": "nonspecific",
    "post surgery": "nonspecific",
    "synovitis/ impingement/ bursitis": "synovitis_capsulitis",
    "tendon": "tendinopathy",
    "unspecified/crossing": "nonspecific",
}

ORCHARD_PATHOLOGY_TYPE_MAP = {
    "M": "muscle_injury",
    "T": "tendinopathy",
    "F": "fracture",
    "J": "joint_sprain",
    "N": "peripheral_nerve_injury",
    "H": "contusion_superficial",
    "K": "laceration",
    "O": "internal_organ_trauma",
    "G": "synovitis_capsulitis",
    "A": "arthritis",
    "U": "chronic_instability",
    "D": "joint_sprain",
}

INJURY_DIAGNOSIS_TEXT_PATTERNS = [
    ("brain_spinal_cord_injury", ["concussion"]),
    ("tendon_rupture", ["tendon rupture"]),
    ("bone_stress_injury", ["stress fracture", "stress injury", "stress reaction", "shin splints"]),
    ("bone_contusion", ["bone contusion"]),
    ("fracture", ["fractur"]),
    ("peripheral_nerve_injury", ["nerve", "brachial plexus", "burner/stinger"]),
    ("cartilage_injury", ["osteochondral", "cartilage", "labral", "labrum", "meniscal"]),
    ("arthritis", ["osteoarthritis", "arthritis"]),
    ("tendinopathy", ["tendinopathy", "tendon injury", "plantar fasci", "plantar heel pain", "tendon strain"]),
    ("bursitis", ["bursitis"]),
    ("synovitis_capsulitis", ["synovitis", "impingement"]),
    ("chronic_instability", ["instability"]),
    ("joint_sprain", ["sprain", "ligament", "dislocation", "subluxation", "plantar plate disruption", "hyperextension", "acl", "mcl", "pcl", "ucl"]),
    ("muscle_contusion", ["muscle haematoma", "muscle contusion"]),
    ("laceration", ["laceration"]),
    ("abrasion", ["abrasion"]),
    ("contusion_superficial", ["contusion", "haematoma", "bruis", "head impact", "head/neck impact"]),
    ("muscle_injury", ["muscle", "strain", "hamstring", "gastroc", "soleus", "rectus femoris", "quadriceps", "pectoralis"]),
    ("nonspecific", ["disc", "pain/injury", "pain undiagnosed", "not otherwise specified", "functional pain", "spinal injury", "soreness", "overuse injuries", "perforated ear drum"]),
]

IOC_BODY_CODE_MAP = {
    "H": "head",
    "N": "neck",
    "S": "shoulder",
    "U": "upper_arm",
    "E": "elbow",
    "R": "forearm",
    "W": "wrist",
    "P": "hand",
    "C": "chest",
    "D": "thoracic_spine",
    "L": "lumbosacral",
    "O": "abdomen",
    "G": "hip_groin",
    "T": "thigh",
    "K": "knee",
    "Q": "lower_leg",
    "A": "ankle",
    "F": "foot",
    "Z": "unspecified",
    "X": "multiple",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_json(value: object) -> str:
    return hashlib.sha256(
        json.dumps(value, sort_keys=True, separators=(",", ":")).encode()
    ).hexdigest()


def without_keys(value: object, keys: set[str]) -> object:
    if isinstance(value, dict):
        return {
            key: without_keys(item, keys)
            for key, item in value.items()
            if key not in keys
        }
    if isinstance(value, list):
        return [without_keys(item, keys) for item in value]
    return value


def is_protected_team_alias_value(value: str) -> bool:
    return bool(re.fullmatch(r"Team [A-Z]", clean_text(value)))


# SQL-side counterpart of is_protected_team_alias_value(), used both to redact
# live rows and to scan for regressions. Kept as a plain regex (no named
# groups/backrefs) so it is safe to splice directly into generated SQL text.
PROTECTED_ALIAS_SQL_PATTERN = "^Team [A-Z]$"
PROTECTED_ALIAS_REDACTED_MARKER = "[REDACTED_PROTECTED_METADATA]"
# Deliberately does NOT encode which letter A-Z was redacted (no length/shape
# clue either): the whole point of the redaction is to purge the pairing
# between a source row and its league alias, so the audit trail must not let
# that pairing be reconstructed from old_value even though old_value is
# normally a real prior value elsewhere in this pipeline.
PROTECTED_ALIAS_OLD_VALUE_MARKER = "[REDACTED_PRIOR_PROTECTED_ALIAS_VALUE]"
PROTECTED_ALIAS_REDACTION_RULE_VERSION = "protected_alias_redaction_2026-07-09_v1"
AGGREGATE_RELEASE_RETIREMENT_RULE_VERSION = "aggregate_release_retirement_2026-07-09_v1"


def protected_alias_scan_sql(check_label: str) -> str:
    """Read-only 'do' block: raises if any live protected Team A-Z alias value
    remains in ingestion.source_rows.source_values or
    processing.record_versions.record_state->>'team_alias'. Only counts are
    reported (never the offending value), so this is safe to run and log from
    self-check, the release gate, or a bare read-only connection.
    """
    return f"""
      do $$
      declare
        source_hits integer;
        record_hits integer;
      begin
        select count(*) into source_hits
        from ingestion.source_rows sr, jsonb_each(sr.source_values) e
        where jsonb_typeof(e.value) = 'string'
          and (e.value #>> '{{}}') ~ '{PROTECTED_ALIAS_SQL_PATTERN}';
        select count(*) into record_hits
        from processing.record_versions
        where record_state ->> 'team_alias' ~ '{PROTECTED_ALIAS_SQL_PATTERN}';
        if source_hits > 0 or record_hits > 0 then
          raise exception '{check_label}: % protected Team A-Z alias value(s) remain (% in ingestion.source_rows.source_values, % in processing.record_versions.record_state.team_alias)',
            source_hits + record_hits, source_hits, record_hits;
        end if;
      end $$;
    """


class SqlParams:
    # Multi-statement SQL batches read values from a transaction-local table
    # because pg bind parameters apply to a single prepared statement.
    def __init__(self) -> None:
        self.values: list[object] = []

    def add(self, value: object) -> int:
        self.values.append(value)
        return len(self.values)

    def text(self, value: object) -> str:
        return f"(select value #>> '{{}}' from _pipeline_params where idx = {self.add(value)})"

    def jsonb(self, value: object) -> str:
        return f"(select value from _pipeline_params where idx = {self.add(value)})"


def run_sql(sql: str, params: list[object] | None = None) -> None:
    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False) as handle:
        handle.write(sql)
        sql_path = handle.name
    params_path = None
    if params is not None:
        with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False) as handle:
            json.dump(params, handle, sort_keys=True)
            params_path = handle.name

    try:
        db_url = os.environ.get("SUPABASE_DB_URL")
        if not db_url:
            raise SystemExit("SUPABASE_DB_URL is required; load .env.local before DB writes")
        command = ["node", str(Path(__file__).with_name("sql_exec.mjs")), sql_path]
        if params_path:
            command.append(params_path)
        subprocess.run(command, check=True)
    finally:
        Path(sql_path).unlink(missing_ok=True)
        if params_path:
            Path(params_path).unlink(missing_ok=True)


def query_sql(sql: str, params: list[object] | None = None) -> list[dict[str, Any]]:
    """Read-only counterpart of run_sql(): runs exactly one query in a
    read-only transaction (pipeline/sql_query.mjs) and returns its rows as
    Python dicts. Used to decide what to do (idempotence checks, team_key
    resolution, dashboard-JSON reconciliation) before ever writing; never
    call this expecting side effects, the transaction is always rolled back.
    """
    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False) as handle:
        handle.write(sql)
        sql_path = handle.name
    params_path = None
    if params is not None:
        with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False) as handle:
            json.dump(params, handle, sort_keys=True)
            params_path = handle.name

    try:
        db_url = os.environ.get("SUPABASE_DB_URL")
        if not db_url:
            raise SystemExit("SUPABASE_DB_URL is required; load .env.local before DB reads")
        command = ["node", str(Path(__file__).with_name("sql_query.mjs")), sql_path]
        if params_path:
            command.append(params_path)
        try:
            result = subprocess.run(command, check=True, capture_output=True, text=True)
        except subprocess.CalledProcessError as exc:
            detail = clean_text(exc.stderr) or clean_text(exc.stdout) or str(exc)
            raise RuntimeError(f"read-only database query failed: {detail}") from exc
        return json.loads(result.stdout)
    finally:
        Path(sql_path).unlink(missing_ok=True)
        if params_path:
            Path(params_path).unlink(missing_ok=True)


def decimal_values_close(left: object, right: object) -> bool:
    if left is None or right is None:
        return left is None and right is None
    try:
        return abs(Decimal(str(left)) - Decimal(str(right))) <= Decimal("0.000000001")
    except InvalidOperation:
        return False


def integer_values_equal(left: object, right: object) -> bool:
    try:
        return int(left) == int(right)
    except (TypeError, ValueError):
        return False


def first_release_payload_mismatch(
    dashboard: dict[str, Any], semantic: dict[str, Any],
) -> tuple[str, object, object] | None:
    """Return the first safe aggregate mismatch in reading order."""
    headline_items = {
        item.get("key"): item
        for item in dashboard.get("headline", [])
        if isinstance(item, dict)
    }
    coverage = dashboard.get("coverage", {})
    coverage = coverage if isinstance(coverage, dict) else {}
    checks = (
        (
            "headline.recorded_injuries",
            headline_items.get("recorded_injuries", {}).get("value"),
            semantic["recorded_injuries"],
            integer_values_equal,
        ),
        (
            "headline.time_loss_injuries",
            headline_items.get("time_loss_injuries", {}).get("value"),
            semantic["time_loss_injuries"],
            integer_values_equal,
        ),
        (
            "monthly.time_loss_injuries",
            semantic["monthly_time_loss_injuries"],
            semantic["dated_time_loss_injuries"],
            integer_values_equal,
        ),
        (
            "monthly.exposure_hours",
            semantic["monthly_exposure_hours"],
            semantic["exposure_hours"],
            decimal_values_close,
        ),
        (
            "coverage.hours",
            coverage.get("hours"),
            semantic["exposure_hours"],
            decimal_values_close,
        ),
        (
            "headline.incidence_per_1000h.denominator",
            headline_items.get("incidence_per_1000h", {}).get("denominator"),
            coverage.get("hours"),
            decimal_values_close,
        ),
        (
            "headline.burden_per_1000h.denominator",
            headline_items.get("burden_per_1000h", {}).get("denominator"),
            coverage.get("hours"),
            decimal_values_close,
        ),
    )
    return next(
        (
            (label, actual, expected)
            for label, actual, expected, values_equal in checks
            if not values_equal(actual, expected)
        ),
        None,
    )


REPO_ROOT = Path(__file__).resolve().parent.parent
APPROVED_ADJUDICATION_14_WORKBOOK_SHA256 = "b258bd9ad13d1fa6ddb58f99fec1f6cf1dfa559cfcd01fa8787931b53b484f1d"
APPROVED_ADJUDICATION_14_EVIDENCE_SHA256 = "d3be9f4308f070951abc0e0f6fd2e33f4f8c670f3b514d1176dc0ebaf5cdbf7e"
APPROVED_ADJUDICATION_14_MANIFEST_SHA256 = "26237f484d2c3aac3a161caa89f61e0207611e2b83fcd78751616f30348c1d78"
APPROVED_ADJUDICATION_14_WORKBOOK_PATH = (
    REPO_ROOT / "data" / "reporting" /
    "adjudication_14_needs_abdel_2024-25_approved_b258bd9a.xlsx"
).resolve()
APPROVED_ADJUDICATION_14_MANIFEST_PATH = (
    REPO_ROOT / "data" / "reporting" / "adjudication_14_approved_batch.json"
).resolve()
APPROVED_ADJUDICATION_14_EVIDENCE_PATH = (
    REPO_ROOT / "data" / "reporting" / "adjudication_checklist_2024-25_evidence.json"
).resolve()


def run_provenance() -> dict[str, str]:
    """Run-identity metadata threaded into every audit.pipeline_runs insert:
    the exact code commit, dependency environment, and operator that
    produced the run. Cheap (a couple of subprocess calls and one file
    hash); call once per command invocation and reuse the result rather
    than recomputing per insert.

    - code_version: `git rev-parse HEAD`, suffixed `-dirty` when
      `git status --porcelain` is non-empty. `release()` refuses to run
      when dirty; every other command may run dirty (see AGENTS.md: the
      pipeline is human-in-the-loop, small verified steps, not a single
      always-clean command).
    - dependency_lock_hash: SHA-256 of `<python_version>\\x1f<sha256 of
      package-lock.json bytes>`, so it changes if either the interpreter
      or the locked JS dependency set changes.
    - operator: PIPELINE_OPERATOR env var, else the OS username.
    """
    commit = subprocess.run(
        ["git", "rev-parse", "HEAD"],
        cwd=REPO_ROOT,
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()
    dirty = bool(
        subprocess.run(
            ["git", "status", "--porcelain"],
            cwd=REPO_ROOT,
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()
    )
    code_version = f"{commit}-dirty" if dirty else commit

    package_lock_path = REPO_ROOT / "package-lock.json"
    if not package_lock_path.exists():
        raise SystemExit(f"dependency lock file not found: {package_lock_path}")
    dependency_lock_hash = hashlib.sha256(
        f"{platform.python_version()}\x1f{sha256_file(package_lock_path)}".encode()
    ).hexdigest()

    operator = os.environ.get("PIPELINE_OPERATOR") or getpass.getuser()

    return {
        "code_version": code_version,
        "dependency_lock_hash": dependency_lock_hash,
        "operator": operator,
    }


def dirty_worktree_paths() -> list[str]:
    """Return changed paths without reading file contents."""
    output = subprocess.run(
        ["git", "status", "--porcelain=v1", "-z", "--untracked-files=all"],
        cwd=REPO_ROOT,
        check=True,
        capture_output=True,
        text=True,
    ).stdout
    paths: list[str] = []
    records = output.split("\0")
    index = 0
    while index < len(records):
        record = records[index]
        if not record:
            index += 1
            continue
        status = record[:2]
        paths.append(record[3:])
        index += 1
        if ("R" in status or "C" in status) and index < len(records):
            paths.append(records[index])
            index += 1
    return paths


def release_owned_dirty_paths(paths: list[str]) -> list[str]:
    release_prefixes = (
        "config/",
        "docs/",
        "lib/",
        "pipeline/",
        "supabase/migrations/",
        "tools/",
        "tests/",
        "content/reporting/",
    )
    release_files = {
        "AGENTS.md",
        "package.json",
        "package-lock.json",
        "pyproject.toml",
        "tsconfig.json",
    }
    return sorted(
        path
        for path in paths
        if path in release_files or path.startswith(release_prefixes)
    )


def validate_dirty_release_override(
    current_dirty_paths: list[str], allowed_paths: list[str],
) -> None:
    blocked_dirty_paths = release_owned_dirty_paths(current_dirty_paths)
    if blocked_dirty_paths:
        raise SystemExit(
            "release-league dirty-tree override covers release-owned files; "
            "commit these paths before preflight or promotion: "
            + ", ".join(blocked_dirty_paths)
        )
    unapproved_dirty_paths = sorted(
        set(current_dirty_paths) - set(allowed_paths)
    )
    if unapproved_dirty_paths:
        raise SystemExit(
            "release-league dirty-tree override requires an exact path "
            "allowlist in PIPELINE_ALLOWED_DIRTY_RELEASE_LEAGUE_PATHS; "
            "unapproved paths: "
            + ", ".join(unapproved_dirty_paths)
        )


def read_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() != ".csv":
        return []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_rows(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_analysis_source_workbook(
    path: Path,
    rows: list[dict[str, str]],
    fieldnames: list[str],
    excluded_row_numbers: set[str],
    filled_columns_by_row_number: dict[str, set[str]],
) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("openpyxl is required to write analysis source workbooks") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        shutil.copy2(path, path.with_suffix(path.suffix + ".bak"))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "analysis_source"
    sheet.append(fieldnames)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(rows) + 1}"
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    red_font = Font(color="C00000")
    green_font = Font(color="008000")
    for output_row_number, row in enumerate(rows, start=2):
        sheet.append([row.get(field, "") for field in fieldnames])
        source_row_number = str(output_row_number)
        excluded = source_row_number in excluded_row_numbers
        filled_columns = filled_columns_by_row_number.get(source_row_number, set())
        for column_index, field in enumerate(fieldnames, start=1):
            cell = sheet.cell(row=output_row_number, column=column_index)
            if field in {"Date Injured", "Confirmed Return Date"}:
                cell.number_format = "@"
            if not clean_text(str(cell.value or "")):
                continue
            if excluded:
                cell.font = red_font
            elif field in filled_columns:
                cell.font = green_font

    for column_index, field in enumerate(fieldnames, start=1):
        width = 14 if field in {"Date Injured", "Confirmed Return Date", "Fit For Selection Date"} else 18
        if field in {"Description", "Mechanism Notes"}:
            width = 28
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    workbook.save(path)
    load_workbook(path, read_only=True).close()


def validate_alias_map(args: argparse.Namespace) -> None:
    alias_map = load_fixture_team_aliases()
    codebook_path = Path(args.codebook)
    codebook_rows = read_rows(codebook_path)
    if not codebook_rows:
        raise SystemExit(f"no team codebook rows found: {codebook_path}")
    missing_columns = [column for column in ["original_id", "new_id"] if column not in codebook_rows[0]]
    if missing_columns:
        raise SystemExit(f"team codebook missing column(s): {', '.join(missing_columns)}")

    codebook_aliases = {clean_text(row.get("new_id")) for row in codebook_rows if clean_text(row.get("new_id"))}
    local_aliases = {clean_text(alias) for alias in alias_map.values() if clean_text(alias)}
    missing_aliases = sorted(local_aliases - codebook_aliases)
    if missing_aliases:
        raise SystemExit(
            f"alias map uses {len(missing_aliases)} codebook alias code(s) absent from {codebook_path.name}"
        )

    print(
        json.dumps(
            {
                "alias_map": str(TEAM_ALIAS_MAP_PATH),
                "alias_map_sha256": sha256_file(TEAM_ALIAS_MAP_PATH),
                "codebook": str(codebook_path),
                "codebook_sha256": sha256_file(codebook_path),
                "fixture_names": len(alias_map),
                "fixture_alias_codes": len(local_aliases),
                "codebook_alias_codes": len(codebook_aliases),
                "status": "valid",
            },
            indent=2,
        )
    )


def stable_uid(prefix: str, *parts: object) -> str:
    text = "\x1f".join(str(part) for part in parts)
    return f"{prefix}_{hashlib.sha256(text.encode()).hexdigest()[:24]}"


def raw_record_id(team: str, season: str, file_hash: str, source_row_number: int) -> str:
    return f"{team}:{season}:{file_hash[:12]}:{source_row_number}"


def parse_uk_date(value: str) -> datetime | None:
    value = value.strip()
    if not value:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return None


def parse_flexible_date(value: object, date_order: str = "month-first") -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = clean_text(str(value) if value is not None else "")
    if not text:
        return None
    slash_formats = (
        ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%m/%d/%y")
        if date_order == "day-first"
        else ("%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%d/%m/%y")
    )
    for fmt in (*slash_formats, "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_int(value: str) -> int | None:
    value = value.strip()
    if not value:
        return None
    try:
        return int(float(value))
    except ValueError:
        return None


def parse_float(value: object) -> float | None:
    text = clean_text(str(value) if value is not None else "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_minutes(value: object) -> float | None:
    parsed = parse_float(value)
    if parsed is not None:
        return parsed
    text = clean_text(str(value) if value is not None else "")
    parts = text.split(":")
    if len(parts) != 3:
        return None
    try:
        hours, minutes, seconds = (float(part) for part in parts)
    except ValueError:
        return None
    return hours * 60 + minutes + seconds / 60


def parse_exposure_timestamp(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = clean_text(str(value) if value is not None else "")
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass
    for fmt in (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
        "%I:%M:%S %p", "%I:%M %p", "%H:%M:%S", "%H:%M",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def adapter_confirmed_return_date(row: dict[str, str]) -> tuple[datetime | None, str]:
    value = clean_text(row.get("Adapter Canonical Confirmed Return Date"))
    origin = clean_text(row.get("Adapter Canonical Confirmed Return Date Origin"))
    if not value and not origin:
        return None, ""
    if not value or not (
        origin.startswith("approved_mapping:") or origin.startswith("manual_adjudication:")
    ):
        raise SystemExit("invalid adapter canonical confirmed return date override")
    returned_at = parse_date_value(value)
    injured_at = parse_date_value(row.get("Date Injured", ""))
    if returned_at is None or (injured_at and returned_at < injured_at):
        raise SystemExit("invalid or unordered adapter canonical confirmed return date")
    return returned_at, origin


def effective_days_injured_with_origin(row: dict[str, str]) -> tuple[int | None, str]:
    days = parse_int(clean_text(row.get("Days Injured")))
    if days is not None and days >= 0:
        return days, "preserved_source_days_injured"

    injured_at = parse_date_value(row.get("Date Injured", ""))
    returned_at = parse_date_value(row.get("Confirmed Return Date", ""))
    if returned_at is None:
        returned_at, _ = adapter_confirmed_return_date(row)
    training_days = parse_int(clean_text(row.get("Training Days Missed")))
    if training_days == -1:
        return 0, "mapped_minus_one_training_days_to_same_day_zero"
    if training_days is not None and training_days >= 0:
        if injured_at and returned_at and returned_at >= injured_at:
            calculated_days = max(0, (returned_at - injured_at).days - 1)
            origin = (
                "mapped_from_training_days_missed_date_consistent"
                if calculated_days == training_days
                else "mapped_from_training_days_missed_date_conflict"
            )
            return training_days, origin
        return training_days, "mapped_from_training_days_missed_without_complete_dates"
    if injured_at and returned_at:
        try:
            corrected_return = returned_at.replace(year=returned_at.year + 1)
        except ValueError:
            corrected_return = None
        corrected_days = max(0, (corrected_return - injured_at).days - 1) if corrected_return else -1
        if training_days is not None and training_days < -1 and 0 <= corrected_days <= 90:
            return corrected_days, "corrected_return_year_then_derived_excluding_injury_day"
        if returned_at >= injured_at:
            return max(0, (returned_at - injured_at).days - 1), "derived_from_dates_excluding_injury_day"
    if clean_text(row.get("TimeLoss vs Medical Attention")).lower() == "medical attention":
        return 0, "inferred_zero_from_medical_attention_with_invalid_or_missing_duration"
    return None, "insufficient_duration_evidence"


def effective_days_injured(row: dict[str, str]) -> int | None:
    return effective_days_injured_with_origin(row)[0]


def effective_confirmed_return_date(
    row: dict[str, str], days: int | None, days_origin: str
) -> tuple[date | None, str]:
    injured_at = parse_date_value(row.get("Date Injured", ""))
    returned_at = parse_date_value(row.get("Confirmed Return Date", ""))
    adapter_returned_at, adapter_return_origin = adapter_confirmed_return_date(row)
    if (
        injured_at
        and days is not None
        and days_origin == "mapped_from_training_days_missed_date_conflict"
    ):
        if days == 0:
            return injured_at, "derived_same_day_return_from_zero_days_date_conflict"
        return injured_at + timedelta(days=days + 1), "derived_from_training_days_missed_date_conflict"
    if returned_at and (not injured_at or returned_at >= injured_at):
        return returned_at, "preserved_source_confirmed_return_date"
    if adapter_returned_at:
        return adapter_returned_at, adapter_return_origin
    closure_override = adapter_canonical_override(
        row,
        "Adapter Canonical Injury Closed",
        "Adapter Canonical Injury Closed Origin",
        {"closed", "open", "unknown"},
    )
    if closure_override and not returned_at:
        return None, "adapter_no_supported_return_date"
    if injured_at and returned_at and days_origin == "corrected_return_year_then_derived_excluding_injury_day":
        try:
            return returned_at.replace(year=returned_at.year + 1), "corrected_source_confirmed_return_year"
        except ValueError:
            pass
    if injured_at and days == 0 and days_origin in {
        "mapped_minus_one_training_days_to_same_day_zero",
        "inferred_zero_from_medical_attention_with_invalid_or_missing_duration",
    }:
        return injured_at, "inferred_same_day_return"
    if injured_at and days is not None:
        offset = days if days_origin == "preserved_source_days_injured" else days + 1
        return injured_at + timedelta(days=offset), "derived_from_date_injured_and_days_definition"
    return None, "insufficient_return_date_evidence"


def clean_text(value: str | None) -> str:
    return (value or "").strip()


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def clean_exposure_cell(value: object, *, preserve_time: bool = False) -> str:
    if isinstance(value, datetime) and preserve_time:
        return value.isoformat(sep=" ")
    return clean_cell(value)


def is_missing(value: str | None) -> bool:
    return clean_text(value).lower() in MISSING_VALUES


def adapter_canonical_override(
    row: dict[str, str], value_field: str, origin_field: str, allowed: set[str]
) -> tuple[str, str] | None:
    value = clean_text(row.get(value_field))
    origin = clean_text(row.get(origin_field))
    if not value and not origin:
        return None
    if value not in allowed or not origin:
        raise SystemExit(f"invalid adapter canonical override: {value_field}")
    if not (origin.startswith("approved_mapping:") or origin.startswith("manual_adjudication:")):
        raise SystemExit(f"invalid adapter canonical override origin: {origin_field}")
    return value, origin


# Single source of truth for the "explicit non-URC Match Type" marker scan.
# Shared by injury_cohort_exclusion_reasons() (the dashboard's live exclusion
# check) and urc_match_scope() (the Phase 3.5 curated-column reproduction of
# the same check, added so the signal is queryable without re-running the
# Python dashboard code). Keep this list in exactly one place.
NON_URC_MATCH_TYPE_MARKERS = (
    "academy",
    "club",
    "cup",
    "friendly",
    "international",
    # This addition requires a recorded adjudication before any team is reprocessed under it.
    "italian elite championship",
    "national",
    "premiership",
    "pro team a",
    "super rugby",
    "top 14",
    "u18",
    "u19",
    "u20",
    "u21",
    "under 18",
    "under 19",
    "under 20",
    "under 21",
)


def injury_cohort_exclusion_reasons(row: dict[str, str], expected_team: str = "") -> list[str]:
    reasons = []
    received_in_team = clean_text(row.get("Received/Injured In Team"))
    if expected_team and not is_missing(received_in_team) and received_in_team.casefold() != expected_team.casefold():
        reasons.append("received_or_injured_in_other_team")

    match_type = clean_text(row.get("Match Type")).casefold()
    if not is_missing(match_type) and any(marker in match_type for marker in NON_URC_MATCH_TYPE_MARKERS):
        reasons.append("explicit_non_urc_match_type")
    return reasons


# Phase 3.5 cohort-signal capture (Adjudication 4, 10 July 2026). Both
# functions below are pure and row-local: they never receive or return the
# protected alias itself beyond the single in-memory comparison in
# received_in_team_status(), and their output is one of a small controlled
# set of category strings, safe to store in curated.injuries and to log.
#
# received_in_team_status reproduces injury_cohort_exclusion_reasons()'s
# "received_or_injured_in_other_team" check as a stored category rather than
# a live re-scan: own_team only on an exact (casefold) match to the team's
# own protected alias; 'club' is broken out from the general 'other_team'
# bucket for audit readability even though (matching the current dashboard
# rule, which excludes anything that is not blank and not the exact own
# alias) both are treated as "not this team's own player" downstream.
def received_in_team_status(row: dict[str, str], own_team_alias: str) -> tuple[str, str]:
    override = adapter_canonical_override(
        row,
        "Adapter Canonical Received In Team Status",
        "Adapter Canonical Received In Team Status Origin",
        {"own_team", "other_team", "club", "missing"},
    )
    if override:
        return override
    value = clean_text(row.get("Received/Injured In Team"))
    if is_missing(value):
        return "missing", "source_missing_or_unknown"
    if value.casefold() == "club":
        return "club", "matched_club_marker"
    if value.casefold() == own_team_alias.casefold():
        return "own_team", "matched_own_team_alias"
    return "other_team", "did_not_match_own_team_alias_or_club_marker"


# urc_match_scope reproduces injury_cohort_exclusion_reasons()'s
# "explicit_non_urc_match_type" check (the NON_URC_MATCH_TYPE_MARKERS scan)
# as a stored category. 'training' and 'urc' are both "retained" outcomes
# under the current rule (as is any other non-missing, non-marker text, e.g.
# 'Other' -- the current rule does not positively verify URC competition
# text, it only excludes on an explicit non-URC marker hit; 'urc' here means
# "not excluded by the marker scan", not "confirmed URC").
def urc_match_scope(row: dict[str, str]) -> tuple[str, str]:
    match_type = clean_text(row.get("Match Type")).casefold()
    if is_missing(match_type):
        return "unknown", "source_missing_or_unknown"
    if match_type == "training":
        return "training", "mapped_from_match_type_training"
    if any(marker in match_type for marker in NON_URC_MATCH_TYPE_MARKERS):
        return "non_urc_marker", "matched_non_urc_match_type_marker"
    return "urc", "no_non_urc_marker_matched"


def activity_context(row: dict[str, str]) -> tuple[str, str]:
    override = adapter_canonical_override(
        row, "Adapter Canonical Activity Context", "Adapter Canonical Activity Context Origin",
        {"urc_match", "match", "training", "unknown"},
    )
    if override:
        return override
    occasion = clean_text(row.get("Occasion category")).lower()
    match_type = clean_text(row.get("Match Type")).lower()
    if occasion in {"game", "match"} and match_type in {"united rugby championship", "urc"}:
        return "urc_match", "mapped_from_occasion_and_match_type"
    if occasion == "training" or match_type == "training":
        return "training", "mapped_from_occasion_category"
    if occasion in {"game", "match"}:
        return "match", "mapped_from_occasion_category_non_urc_match"
    return "unknown", "insufficient_direct_evidence"


def contact_context(row: dict[str, str]) -> tuple[str, str]:
    override = adapter_canonical_override(
        row, "Adapter Canonical Contact Context", "Adapter Canonical Contact Context Origin",
        {"contact", "non_contact", "unknown"},
    )
    if override:
        return override
    value = clean_text(row.get("Is Contact")).lower()
    if value == "contact" or value.startswith("contact ("):
        return "contact", "mapped_from_is_contact"
    if value in {"non-contact", "non-contact trauma", "overuse (gradual onset)", "overuse (sudden onset)"}:
        return "non_contact", "mapped_from_is_contact"
    tissue = clean_text(row.get("Injury Tissue Type/s")).lower()
    onset = clean_text(row.get("Nature of onset")).lower()
    if is_missing(value) and tissue == "muscle strain/spasm" and onset == "acute":
        return "non_contact", "inferred_from_acute_muscle_strain"
    return "unknown", "source_missing_or_unknown"


def recurrence_status(row: dict[str, str]) -> tuple[str, str]:
    override = adapter_canonical_override(
        row, "Adapter Canonical Recurrence Status", "Adapter Canonical Recurrence Status Origin",
        {"first_episode", "recurrence", "unknown"},
    )
    if override:
        return override
    value = clean_text(row.get("Recurrence")).lower()
    if value in {"first episode", "new injury (not recurrent)"}:
        return "first_episode", "mapped_from_recurrence"
    if value == "recurrence" or value.endswith(" recurrence"):
        return "recurrence", "mapped_from_recurrence"
    return "unknown", "source_missing_or_unknown"


def severity_category(days_injured: int | None, is_closed: bool | None) -> tuple[str, str]:
    if days_injured is None or is_closed is False:
        return "unknown_or_censored", "missing_days_or_unclosed_injury"
    if days_injured == 0:
        return "zero_days_medical_attention_only", "derived_from_days_injured"
    if days_injured == 1:
        return "one_day", "derived_from_days_injured"
    if 2 <= days_injured <= 3:
        return "two_to_three_days", "derived_from_days_injured"
    if 4 <= days_injured <= 7:
        return "four_to_seven_days", "derived_from_days_injured"
    if 8 <= days_injured <= 28:
        return "eight_to_twenty_eight_days", "derived_from_days_injured"
    return "greater_than_twenty_eight_days", "derived_from_days_injured"


def injury_closed(row: dict[str, str]) -> tuple[bool | None, str]:
    override = adapter_canonical_override(
        row, "Adapter Canonical Injury Closed", "Adapter Canonical Injury Closed Origin",
        {"closed", "open", "unknown"},
    )
    if override:
        value, origin = override
        return {"closed": True, "open": False, "unknown": None}[value], origin
    value = clean_text(row.get("is_injury_closed"))
    if value == "1":
        return True, "mapped_from_is_injury_closed"
    if value == "0":
        return False, "mapped_from_is_injury_closed"
    if parse_date_value(clean_text(row.get("Confirmed Return Date"))) is not None:
        return True, "mapped_from_confirmed_return_date"
    adapter_return, adapter_return_origin = adapter_confirmed_return_date(row)
    if adapter_return is not None:
        return True, adapter_return_origin
    time_loss = clean_text(row.get("TimeLoss vs Medical Attention")).lower()
    if time_loss == "time loss":
        return False, "unclosed_time_loss_without_return_date"
    if time_loss == "medical attention":
        return True, "mapped_from_medical_attention_only"
    return None, "source_missing_or_unknown"


def effective_orchard_code(row: dict[str, str]) -> str:
    source = clean_text(row.get("Orchard Code"))
    if source:
        return source
    if clean_text(row.get("Problem type")).lower() == "injury":
        return clean_text(row.get("Illness Code"))
    return ""


def body_location(row: dict[str, str]) -> tuple[str, str]:
    override = adapter_canonical_override(
        row, "Adapter Canonical Body Location", "Adapter Canonical Body Location Origin",
        set(BODY_LOCATION_LABELS),
    )
    if override:
        return override
    orchard_code = effective_orchard_code(row)
    if orchard_code:
        mapped_from_code = IOC_BODY_CODE_MAP.get(orchard_code[0].upper())
        if mapped_from_code:
            return mapped_from_code, "mapped_from_orchard_code_ioc_body_area"
    source = clean_text(row.get("Body Part"))
    controlled = BODY_LOCATION_LABEL_TO_KEY.get(source.lower())
    if controlled:
        return controlled, "preserved_controlled_body_part"
    mapped = BODY_LOCATION_MAP.get(source.lower())
    if mapped:
        return mapped, "mapped_from_body_part_ioc_body_area"
    evidence = " ".join(
        clean_text(row.get(field)).lower()
        for field in ["Description", "Orchard Code", "Injury Tissue Type/s"]
    )
    if "concussion" in evidence or "brain injury" in evidence:
        return "head", "inferred_from_concussion_evidence"
    return "unknown", "source_missing_or_unknown"


def problem_type(row: dict[str, str]) -> tuple[str, str]:
    override = adapter_canonical_override(
        row, "Adapter Canonical Problem Type", "Adapter Canonical Problem Type Origin",
        {"injury", "illness", "unknown"},
    )
    if override:
        return override
    source = clean_text(row.get("Problem type")).lower()
    if source in {"injury", "illness"}:
        return source, "mapped_from_problem_type"
    if not is_missing(row.get("Orchard Code")) or not is_missing(row.get("Injury Tissue Type/s")):
        return "injury", "inferred_from_orchard_code_or_injury_type"
    if not is_missing(row.get("Illness Code")):
        return "illness", "inferred_from_illness_code"
    return "unknown", "source_missing_or_unknown"


def injury_type(row: dict[str, str]) -> tuple[str, str]:
    override = adapter_canonical_override(
        row, "Adapter Canonical Tissue Pathology", "Adapter Canonical Tissue Pathology Origin",
        set(INJURY_TYPE_LABELS),
    )
    if override:
        return override
    if clean_text(row.get("Problem type")).lower() == "illness":
        return "unknown", "not_applicable_to_illness"
    value = clean_text(row.get("Injury Tissue Type/s")).lower()
    controlled = INJURY_TYPE_LABEL_TO_KEY.get(value)
    if controlled:
        return controlled, "preserved_controlled_injury_tissue_type"
    orchard_code = effective_orchard_code(row).upper()
    if (is_missing(value) or value in {"other pain/ unspecified", "unspecified/crossing"}) and len(orchard_code) >= 2:
        mapped_from_code = ORCHARD_PATHOLOGY_TYPE_MAP.get(orchard_code[1])
        if mapped_from_code:
            return mapped_from_code, "mapped_from_orchard_code_ioc_pathology"
    mapped = INJURY_TYPE_MAP.get(value)
    if mapped:
        return mapped, "mapped_from_injury_tissue_type"
    if "not concussion" in value or "not diagnosed as concussion" in value:
        return "contusion_superficial", "mapped_from_explicit_diagnosis_text_ioc_pathology"
    if "lumbar pain" in value:
        return "nonspecific", "mapped_from_explicit_diagnosis_text_ioc_pathology"
    for injury_key, patterns in INJURY_DIAGNOSIS_TEXT_PATTERNS:
        if any(pattern in value for pattern in patterns):
            return injury_key, "mapped_from_explicit_diagnosis_text_ioc_pathology"
    if len(orchard_code) >= 2:
        mapped_from_code = ORCHARD_PATHOLOGY_TYPE_MAP.get(orchard_code[1])
        if mapped_from_code:
            return mapped_from_code, "mapped_from_orchard_code_ioc_pathology"
    return "unknown", "source_missing_or_unknown"


def prepare_intake(args: argparse.Namespace) -> None:
    standardised_path = Path(args.file)
    source_path = Path(args.source_file)
    output_path = Path(args.output)
    rows = read_rows(standardised_path)
    if source_path.suffix.lower() in {".xlsx", ".xlsm"}:
        _, source_rows = read_xlsx_rows(source_path, args.source_sheet)
        source_row_numbers = [int(row["_source_row_number"]) for row in source_rows]
    else:
        source_rows = read_rows(source_path)
        source_row_numbers = list(range(2, len(source_rows) + 2))
    if not rows:
        raise SystemExit(f"no standardised rows found: {standardised_path}")
    if args.player_id_column not in rows[0]:
        raise SystemExit(f"missing player ID column: {args.player_id_column}")
    if len(rows) != len(source_rows):
        raise SystemExit(
            f"row-count mismatch: standardised={len(rows)} source={len(source_rows)}"
        )

    original_hash = sha256_file(standardised_path)
    source_hash = sha256_file(source_path)
    prepared_rows = []
    for index, row in enumerate(rows):
        offset = index + 2
        prepared = dict(row)
        prepared.update(
            {
                "source_archive_path": str(source_path),
                "source_file_sha256": source_hash,
                "source_sheet": args.source_sheet,
                "source_row_number": str(source_row_numbers[index]),
                "standardised_file_sha256": original_hash,
                "standardised_row_number": str(offset),
                "source_locator_status": "provisional_reference_locator",
                "player_uid": stable_uid("ply", args.team, row[args.player_id_column]),
                "injury_uid": stable_uid("inj", args.team, args.season, original_hash, offset),
            }
        )
        prepared_rows.append(prepared)

    fieldnames = list(rows[0].keys()) + LOCATOR_FIELDS + UID_FIELDS
    write_rows(output_path, prepared_rows, fieldnames)

    output_hash = sha256_file(output_path)
    if args.manifest:
        manifest_path = Path(args.manifest)
        manifest = json.loads(manifest_path.read_text())
        manifest.update(
            {
                "locator_enriched_intake_file": str(output_path),
                "locator_enriched_file_sha256": output_hash,
                "locator_enriched_row_count": len(prepared_rows),
                "locator_fields": LOCATOR_FIELDS,
                "uid_fields": UID_FIELDS,
                "source_locator_status": "provisional_reference_locator",
            }
        )
        manifest_path.write_text(json.dumps(manifest, indent=2) + "\n")

    print(f"prepared {output_path} sha256={output_hash} rows={len(prepared_rows)}")


def export_xlsx_sheet(args: argparse.Namespace) -> None:
    headers, rows = read_xlsx_rows(Path(args.file), args.sheet)
    blank_columns = list(getattr(args, "blank_column", []) or [])
    missing_blank_columns = sorted(set(blank_columns) - set(headers))
    if missing_blank_columns:
        raise SystemExit(
            "cannot blank missing workbook column(s): " + ", ".join(missing_blank_columns)
        )
    output_path = Path(args.output)
    write_rows(
        output_path,
        [
            {
                header: "" if header in blank_columns else clean_cell(row.get(header))
                for header in headers if header
            }
            for row in rows
        ],
        [header for header in headers if header],
    )
    print(
        json.dumps(
            {
                "exported": str(output_path),
                "rows": len(rows),
                "columns": len([header for header in headers if header]),
                "blanked_columns": blank_columns,
                "blanked_value_counts": {
                    header: sum(1 for row in rows if clean_cell(row.get(header)))
                    for header in blank_columns
                },
                "sha256": sha256_file(output_path),
            },
            indent=2,
        )
    )


def adapt_injury_intake(args: argparse.Namespace) -> None:
    source_path = Path(args.file)
    headers, rows = read_xlsx_rows(source_path, args.sheet)
    output_path = Path(args.output)
    audit_path = Path(args.audit_output)
    team_key = clean_text(args.team).lower()
    injury_date_order = clean_text(getattr(args, "date_order", "day-first")) or "day-first"
    required = {"DOB", "Problem type", "Occasion category", "Is Contact", "Match Type"}
    missing = sorted(required - set(headers))
    if missing:
        raise SystemExit("injury adapter missing standard column(s): " + ", ".join(missing))

    fixture_dates: set[date] = set()
    fixture_path_text = clean_text(getattr(args, "fixture_file", ""))
    if fixture_path_text:
        public_team_name = TEAM_KEY_ALIAS_LOOKUP_NAMES.get(team_key)
        if not public_team_name:
            raise SystemExit(f"unknown canonical team key for fixture adapter: {team_key}")
        team_alias = load_fixture_team_aliases().get(public_team_name)
        if not team_alias:
            raise SystemExit(f"fixture alias is unavailable for canonical team key: {team_key}")
        for fixture in read_rows(Path(fixture_path_text)):
            if team_alias not in {
                clean_text(fixture.get("home_team_alias")),
                clean_text(fixture.get("away_team_alias")),
            }:
                continue
            parsed = parse_flexible_date(fixture.get("corrected_date"), "day-first")
            if parsed:
                fixture_dates.add(parsed.date())

    output_rows: list[dict[str, str]] = []
    events: list[dict[str, object]] = []
    dob_blanked = 0
    populated = 0
    dates_normalized = 0
    for row in rows:
        source_row_number = int(row["_source_row_number"])
        adapted = {header: clean_cell(row.get(header)) for header in headers if header}
        if adapted.get("DOB"):
            adapted["DOB"] = ""
            dob_blanked += 1

        for field in ("Date Injured", "Confirmed Return Date"):
            source_value = adapted.get(field, "")
            parsed_date = parse_flexible_date(source_value, injury_date_order)
            if not parsed_date:
                continue
            canonical_value = parsed_date.strftime("%d/%m/%Y")
            if canonical_value == source_value:
                continue
            adapted[field] = canonical_value
            dates_normalized += 1
            events.append(
                {
                    "source_row_number": source_row_number,
                    "field": field,
                    "old_value": source_value,
                    "new_value": canonical_value,
                    "action": "deterministic_normalization",
                    "rule": f"{injury_date_order}_date_to_canonical_dd_mm_yyyy",
                }
            )

        def fill(field: str, value: str, rule: str) -> None:
            nonlocal populated
            if adapted.get(field) or not value:
                return
            adapted[field] = value
            populated += 1
            events.append(
                {
                    "source_row_number": source_row_number,
                    "field": field,
                    "old_value": None,
                    "new_value": value,
                    "action": "deterministic_derivation",
                    "rule": rule,
                }
            )

        fallback_occasion = adapted.get("Match vs Training", "").upper()
        if fallback_occasion in {"GAME", "TRAINING"}:
            fill("Occasion category", fallback_occasion, "explicit_match_vs_training_value")

        if adapted.get("Orchard Code") or adapted.get("Injury Tissue Type/s"):
            fill("Problem type", "Injury", "injury_code_or_tissue_evidence")
        elif adapted.get("Illness Code"):
            fill("Problem type", "Illness", "illness_code_evidence")

        mechanism = adapted.get("Mechanism of Injury", "").casefold().strip()
        if re.search(r"\(non[ -]?contact\)\s*$", mechanism):
            fill("Is Contact", "Non-contact", "explicit_mechanism_suffix")
        elif re.search(r"\(contact\)\s*$", mechanism):
            fill("Is Contact", "Contact", "explicit_mechanism_suffix")

        injured_at = parse_flexible_date(adapted.get("Date Injured"), "day-first")
        if (
            fixture_dates
            and adapted.get("Occasion category", "").casefold() in {"game", "match"}
            and injured_at
            and injured_at.date() in fixture_dates
        ):
            fill("Match Type", "URC", "unique_team_fixture_date_link")
        output_rows.append(adapted)

    write_rows(output_path, output_rows, [header for header in headers if header])
    audit = {
        "rule_version": "sa_injury_boundary_adapter_2026-07-13_v2",
        "date_order": injury_date_order,
        "team": args.team,
        "season": args.season,
        "source_file": str(source_path),
        "source_file_sha256": sha256_file(source_path),
        "source_sheet": args.sheet,
        "source_rows": len(rows),
        "output_file": str(output_path),
        "output_file_sha256": sha256_file(output_path),
        "output_rows": len(output_rows),
        "fixture_file_sha256": sha256_file(Path(fixture_path_text)) if fixture_path_text else None,
        "fixture_dates_available": len(fixture_dates),
        "action_counts": {
            "blank_dob": dob_blanked,
            "normalize_date": dates_normalized,
            "populate_approved_blank_field": populated,
        },
        "events": events,
        "privacy": {
            "dob_values_remaining": sum(1 for row in output_rows if row.get("DOB")),
            "identifying_values_logged": False,
        },
        "row_reconciliation": "one output row per nonblank physical source row; original worksheet row retained by prepare-intake",
    }
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_path.write_text(json.dumps(audit, indent=2) + "\n")
    print(
        json.dumps(
            {
                "adapted": str(output_path),
                "rows": len(output_rows),
                "sha256": audit["output_file_sha256"],
                "dob_values_blanked": dob_blanked,
                "approved_blank_fields_populated": populated,
                "audit": str(audit_path),
            },
            indent=2,
        )
    )


def read_xlsx_rows(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required to read exposure workbooks") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"sheet not found: {sheet_name}")
    worksheet = workbook[sheet_name]
    values = worksheet.iter_rows(values_only=True)
    try:
        headers = [clean_cell(value) for value in next(values)]
    except StopIteration as exc:
        raise SystemExit(f"empty workbook sheet: {sheet_name}") from exc

    rows: list[dict[str, Any]] = []
    for source_row_number, raw_values in enumerate(values, start=2):
        row = {
            header: raw_values[index] if index < len(raw_values) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(clean_cell(value) for value in row.values()):
            row["_source_row_number"] = source_row_number
            rows.append(row)
    workbook.close()
    return headers, rows


def required_exposure_reporting_grain(args: argparse.Namespace) -> str:
    grain = clean_text(getattr(args, "reporting_grain", ""))
    if grain not in EXPOSURE_REPORTING_GRAINS:
        raise SystemExit(
            "exposure reporting grain must be supplied explicitly as one of: "
            + ", ".join(EXPOSURE_REPORTING_GRAINS)
        )
    return grain


def prepare_exposure(args: argparse.Namespace) -> None:
    reporting_grain = required_exposure_reporting_grain(args)
    workbook_path = Path(args.file)
    output_path = Path(args.output)
    qc_path = Path(args.qc_output)
    if workbook_path.suffix.casefold() == ".csv":
        rows = read_rows(workbook_path)
        headers = list(rows[0]) if rows else []
        rows = [{**row, "_source_row_number": index} for index, row in enumerate(rows, start=2)]
    else:
        headers, rows = read_xlsx_rows(workbook_path, args.sheet)
    if not rows:
        raise SystemExit(f"no exposure rows found: {workbook_path}")
    if args.player_column not in headers:
        raise SystemExit(f"missing player column: {args.player_column}")
    if EXPOSURE_DECLARED_GRAIN_FIELD in headers:
        raise SystemExit(
            f"source exposure workbook must not contain reserved pipeline column: {EXPOSURE_DECLARED_GRAIN_FIELD}"
        )

    source_hash = sha256_file(workbook_path)
    source_columns = list(headers)
    derive_minutes = bool(getattr(args, "derive_minutes_from_timestamps", False))
    start_timestamp_column = getattr(args, "start_timestamp_column", "session start date time")
    end_timestamp_column = getattr(args, "end_timestamp_column", "session end date time")
    if derive_minutes:
        missing_timestamp_columns = [
            column for column in (start_timestamp_column, end_timestamp_column)
            if column not in headers
        ]
        if missing_timestamp_columns:
            raise SystemExit(
                "cannot derive exposure minutes; missing timestamp column(s): "
                + ", ".join(missing_timestamp_columns)
            )

    distance_source_file = clean_text(getattr(args, "distance_source_file", ""))
    distance_source_sheet = clean_text(getattr(args, "distance_source_sheet", ""))
    distance_source_column = clean_text(getattr(args, "distance_source_column", ""))
    distance_source_hash = ""
    distance_by_source_row: dict[int, object] = {}
    if any((distance_source_file, distance_source_sheet, distance_source_column)):
        if not all((distance_source_file, distance_source_sheet, distance_source_column)):
            raise SystemExit(
                "distance restoration requires --distance-source-file, "
                "--distance-source-sheet, and --distance-source-column"
            )
        distance_source_path = Path(distance_source_file)
        distance_headers, distance_rows = read_xlsx_rows(
            distance_source_path, distance_source_sheet
        )
        if distance_source_column not in distance_headers:
            raise SystemExit(f"missing distance source column: {distance_source_column}")
        distance_source_hash = sha256_file(distance_source_path)
        distance_by_source_row = {
            int(row["_source_row_number"]): row.get(distance_source_column)
            for row in distance_rows
        }

    prepared_rows: list[dict[str, str]] = []
    dates: list[datetime] = []
    date_parse_failures: list[int] = []
    missing_player_rows: list[int] = []
    missing_date_rows: list[int] = []
    negative_minutes_rows: list[int] = []
    negative_distance_rows: list[int] = []
    exact_hashes: set[str] = set()
    exact_duplicate_rows = 0
    player_date_counts: dict[str, list[int]] = {}
    derived_minutes_rows = 0
    unparseable_timestamp_rows: list[int] = []
    restored_distance_rows = 0

    for output_row_number, row in enumerate(rows, start=2):
        source_row_number = int(row["_source_row_number"])
        source_payload = {
            header: clean_exposure_cell(
                row.get(header),
                preserve_time=header in {start_timestamp_column, end_timestamp_column},
            )
            for header in headers
            if header
        }
        source_row_hash = hashlib.sha256(
            json.dumps(source_payload, sort_keys=True).encode()
        ).hexdigest()
        if source_row_hash in exact_hashes:
            exact_duplicate_rows += 1
        exact_hashes.add(source_row_hash)

        prepared_values = {
            header: clean_exposure_cell(
                row.get(header),
                preserve_time=header in {start_timestamp_column, end_timestamp_column},
            )
            for header in source_columns
        }
        minutes_origin = "source_reported" if prepared_values.get(args.minutes_column) else "missing"
        if derive_minutes and not prepared_values.get(args.minutes_column):
            start = parse_exposure_timestamp(row.get(start_timestamp_column))
            end = parse_exposure_timestamp(row.get(end_timestamp_column))
            if start is None or end is None or end < start:
                unparseable_timestamp_rows.append(source_row_number)
            else:
                prepared_values[args.minutes_column] = f"{(end - start).total_seconds() / 60:.6f}"
                minutes_origin = "deterministic_end_minus_start"
                derived_minutes_rows += 1

        distance_origin = "source_reported" if prepared_values.get(args.distance_column) else "missing"
        if not prepared_values.get(args.distance_column) and distance_by_source_row:
            restored = clean_cell(distance_by_source_row.get(source_row_number))
            if restored:
                prepared_values[args.distance_column] = restored
                distance_origin = "row_aligned_reference_source"
                restored_distance_rows += 1

        player_value = prepared_values.get(args.player_column, "")
        if not player_value:
            missing_player_rows.append(source_row_number)
        player_uid = stable_uid("ply", args.team, player_value)

        parsed_date = parse_flexible_date(row.get(args.date_column), args.date_order)
        if parsed_date is None:
            missing_date_rows.append(source_row_number)
            if clean_cell(row.get(args.date_column)):
                date_parse_failures.append(source_row_number)
        else:
            dates.append(parsed_date)

        minutes = parse_minutes(prepared_values.get(args.minutes_column))
        distance = parse_float(prepared_values.get(args.distance_column))
        if minutes is not None and minutes < 0:
            negative_minutes_rows.append(source_row_number)
        if distance is not None and distance < 0:
            negative_distance_rows.append(source_row_number)

        if player_value and parsed_date:
            key = stable_uid("key", args.team, player_value, parsed_date.date().isoformat())
            player_date_counts.setdefault(key, []).append(source_row_number)

        prepared = dict(prepared_values)
        prepared.update(
            {
                "minutes_total_origin": minutes_origin,
                "distance_total_origin": distance_origin,
                "source_archive_path": str(workbook_path),
                "source_file_sha256": source_hash,
                "source_sheet": args.sheet,
                "source_row_number": str(source_row_number),
                "source_row_sha256": source_row_hash,
                "standardised_file_sha256": source_hash,
                "standardised_row_number": str(output_row_number),
                "source_locator_status": "provisional_reference_locator",
                "player_uid": player_uid,
                EXPOSURE_DECLARED_GRAIN_FIELD: reporting_grain,
            }
        )
        prepared_rows.append(prepared)

    fieldnames = (
        source_columns + EXPOSURE_ORIGIN_FIELDS + EXPOSURE_LOCATOR_FIELDS
        + ["player_uid", EXPOSURE_DECLARED_GRAIN_FIELD]
    )
    write_rows(output_path, prepared_rows, fieldnames)
    output_hash = sha256_file(output_path)

    blank_columns = [
        header for header in source_columns
        if all(not clean_cell(row.get(header)) for row in rows)
    ]
    populated_columns = [
        header for header in source_columns
        if header not in blank_columns
    ]
    expected_columns: list[str] = []
    if args.codebook:
        codebook_rows = read_rows(Path(args.codebook))
        expected_columns = [
            clean_text(row.get("Standard_Column_Name"))
            for row in codebook_rows
            if clean_text(row.get("Standard_Column_Name"))
        ]
    normalized_headers = {header.strip() for header in headers}
    missing_codebook_columns = [
        column for column in expected_columns
        if column.strip() not in normalized_headers
    ]
    duplicate_player_date_groups = [
        {"key_hash": key, "rows": row_numbers}
        for key, row_numbers in player_date_counts.items()
        if len(row_numbers) > 1
    ]
    exposure_reporting_grain = {
        "current_file_reporting_grain": reporting_grain,
        "selection_source": "required_cli_argument",
        "allowed_values": list(EXPOSURE_REPORTING_GRAINS),
        "note": "Reporting grain is an explicit intake attribute; source team aliases do not determine it.",
    }

    qc = {
        "file": str(output_path),
        "file_sha256": output_hash,
        "team": args.team,
        "season": args.season,
        "source_workbook": str(workbook_path),
        "source_workbook_sha256": source_hash,
        "source_sheet": args.sheet,
        "row_count": len(prepared_rows),
        "column_count": len(fieldnames),
        "source_column_count": len(headers),
        "deidentified_source_label_columns_preserved": [args.player_column],
        "exposure_reporting_grain": exposure_reporting_grain,
        "player_uid_count": len({row["player_uid"] for row in prepared_rows}),
        "date_column": args.date_column,
        "date_order": args.date_order,
        "date_parseable_rows": len(dates),
        "date_parse_failure_rows": date_parse_failures,
        "date_min": min(dates).date().isoformat() if dates else None,
        "date_max": max(dates).date().isoformat() if dates else None,
        "blank_columns": blank_columns,
        "populated_columns": populated_columns,
        "missing_codebook_columns": missing_codebook_columns,
        "exact_duplicate_source_rows": exact_duplicate_rows,
        "duplicate_player_date_groups": duplicate_player_date_groups,
        "anomalies": {
            "missing_player_rows": missing_player_rows,
            "missing_date_rows": missing_date_rows,
            "negative_minutes_rows": negative_minutes_rows,
            "negative_distance_rows": negative_distance_rows,
            "unparseable_timestamp_rows": unparseable_timestamp_rows,
        },
        "adapter": {
            "duration_rule": (
                "end_timestamp_minus_start_timestamp"
                if derive_minutes else "preserve_source_duration"
            ),
            "derived_minutes_rows": derived_minutes_rows,
            "distance_alignment": (
                "physical_source_row" if distance_by_source_row else "not_applicable"
            ),
            "distance_source_file_sha256": distance_source_hash or None,
            "distance_source_sheet": distance_source_sheet or None,
            "distance_source_column": distance_source_column or None,
            "restored_distance_rows": restored_distance_rows,
        },
        "notes": [
            "The source player label column is already de-identified, is preserved, and is also used to derive player_uid.",
            "Rows are locator-enriched for local review only; this command does not ingest exposure into Supabase.",
        ],
    }
    qc_path.parent.mkdir(parents=True, exist_ok=True)
    qc_path.write_text(json.dumps(qc, indent=2) + "\n")

    if args.manifest:
        manifest_path = Path(args.manifest)
        manifest = json.loads(manifest_path.read_text())
        manifest["exposure_intake"] = {
            "team": args.team,
            "season": args.season,
            "source_workbook": str(workbook_path),
            "source_workbook_sha256": source_hash,
            "source_sheet": args.sheet,
            "locator_enriched_intake_file": str(output_path),
            "locator_enriched_file_sha256": output_hash,
            "locator_enriched_row_count": len(prepared_rows),
            "qc_file": str(qc_path),
            "qc_file_sha256": sha256_file(qc_path),
            "locator_fields": EXPOSURE_LOCATOR_FIELDS,
            "uid_fields": ["player_uid"],
            "deidentified_source_label_columns_preserved": [args.player_column],
            "date_order": args.date_order,
            "exposure_reporting_grain": exposure_reporting_grain,
            "source_locator_status": "provisional_reference_locator",
            "adapter": qc["adapter"],
        }
        manifest_path.write_text(json.dumps(manifest, indent=2) + "\n")

    print(
        json.dumps(
            {
                "prepared": str(output_path),
                "team": args.team,
                "season": args.season,
                "rows": len(prepared_rows),
                "sha256": output_hash,
                "qc": str(qc_path),
                "date_min": qc["date_min"],
                "date_max": qc["date_max"],
                "player_uid_count": qc["player_uid_count"],
                "reporting_grain": reporting_grain,
            },
            indent=2,
        )
    )


def urc_game_opponent(description: str) -> str | None:
    bracket = re.search(r"\(([^()]*)\)", description)
    if bracket is None:
        return None
    candidate = re.sub(r"\b(home|away|h|a)\b", "", bracket.group(1).lower())
    candidate = re.sub(r"[^a-z]+", " ", candidate).strip()
    matches = difflib.get_close_matches(
        candidate,
        EDINBURGH_URC_OPPONENTS,
        n=1,
        cutoff=URC_OPPONENT_FUZZY_CUTOFF,
    )
    return matches[0] if matches else None


def exposure_scope_status(row: dict[str, str], team: str = "") -> tuple[str, str]:
    fields = ["Competition", "session type", "If match, surface?", "Training With", "Training Type"]
    text = " ".join(clean_text(row.get(field)).lower() for field in fields).strip()
    if not text:
        return "scope_unknown_included", "blank_scope_fields_retained"
    if any(term in text for term in ["rehab", "return to play", "rtp"]):
        return "out_of_scope_explicit", "explicit_rehab_or_rtp"
    if any(term in text for term in ["international", "national", "scotland"]):
        return "out_of_scope_explicit", "explicit_international_exposure"
    edinburgh = clean_text(team).lower() == "edinburgh"
    if "academy" in text and (not edinburgh or "game" in text):
        return "out_of_scope_explicit", "academy_game"
    if edinburgh and ("warm" in text or "top up" in text):
        return "in_scope_explicit", "warmup_or_topup_retained"
    if edinburgh and re.search(r"\bgame\b", text):
        if urc_game_opponent(text):
            return "in_scope_explicit", "urc_opponent_game"
        return "out_of_scope_explicit", "non_urc_or_unspecified_game"
    return "in_scope_explicit", "explicit_scope_context_not_excluded"


def clean_exposure(args: argparse.Namespace) -> None:
    path = Path(args.file)
    output_path = Path(args.output)
    qc_path = Path(args.qc_output)
    rows = read_rows(path)
    if not rows:
        raise SystemExit("no exposure rows found")

    reporting_grain = required_exposure_reporting_grain(args)
    declared_grains = {
        clean_text(row.get(EXPOSURE_DECLARED_GRAIN_FIELD))
        for row in rows
        if clean_text(row.get(EXPOSURE_DECLARED_GRAIN_FIELD))
    }
    declared_grain_rows = sum(
        1 for row in rows if clean_text(row.get(EXPOSURE_DECLARED_GRAIN_FIELD))
    )
    invalid_declared_grains = declared_grains - set(EXPOSURE_REPORTING_GRAINS)
    if invalid_declared_grains:
        raise SystemExit(
            "cleaned exposure intake contains invalid declared exposure grain value(s): "
            + ", ".join(sorted(invalid_declared_grains))
        )
    if declared_grains and declared_grains != {reporting_grain}:
        raise SystemExit(
            "--reporting-grain does not match the grain declared during prepare-exposure: "
            + ", ".join(sorted(declared_grains))
        )
    if declared_grain_rows != len(rows):
        raise SystemExit(
            "prepared exposure grain declaration is missing from one or more rows; "
            "rerun prepare-exposure before cleaning"
        )

    exact_seen: dict[str, int] = {}
    cleaned_rows: list[dict[str, str]] = []
    counts: dict[str, int] = {}
    reasons: dict[str, int] = {}
    grain_counts: dict[str, int] = {}
    included_minutes = 0.0
    included_distance = 0.0
    dates: list[datetime] = []
    window_start = parse_date_value(clean_text(getattr(args, "window_start", "")))
    window_end = parse_date_value(clean_text(getattr(args, "window_end", "")))
    if bool(window_start) != bool(window_end):
        raise SystemExit("exposure analysis window requires both --window-start and --window-end")
    if window_start and window_end and window_start > window_end:
        raise SystemExit("exposure analysis window start must not be after its end")

    for row in rows:
        grain = reporting_grain
        grain_counts[grain] = grain_counts.get(grain, 0) + 1
        minutes = parse_minutes(row.get("minutes total"))
        distance = parse_float(row.get("distance total"))
        parsed_date = parse_flexible_date(row.get("session date"), args.date_order)
        if parsed_date:
            dates.append(parsed_date)
        scope_status, scope_reason = exposure_scope_status(row, getattr(args, "team", ""))
        source_hash = row.get("source_row_sha256") or hashlib.sha256(
            json.dumps(
                {
                    key: value
                    for key, value in row.items()
                    if key not in EXPOSURE_LOCATOR_FIELDS + ["player_uid"]
                },
                sort_keys=True,
            ).encode()
        ).hexdigest()
        duplicate_copy = source_hash in exact_seen
        exact_seen.setdefault(source_hash, int(row.get("source_row_number", "0") or 0))

        exclusion_reasons: list[str] = []
        if duplicate_copy:
            exclusion_reasons.append("exact_duplicate_copy")
        if not clean_text(row.get("name")):
            exclusion_reasons.append("missing_player_label")
        if parsed_date is None:
            exclusion_reasons.append("missing_or_unparseable_date")
        if minutes is None:
            exclusion_reasons.append("missing_or_unparseable_minutes")
        if distance is None:
            exclusion_reasons.append("missing_or_unparseable_distance")
        if (minutes is not None and minutes < 0) or (distance is not None and distance < 0):
            exclusion_reasons.append("negative_minutes_or_distance")
        if minutes == 0 and distance == 0:
            exclusion_reasons.append("zero_minutes_and_zero_distance")
        if grain == "weekly":
            if minutes is not None and minutes < 5:
                exclusion_reasons.append("weekly_minutes_below_5")
            if minutes is not None and minutes > 1100:
                exclusion_reasons.append("weekly_minutes_above_1100")
            if distance is not None and distance > 40000:
                exclusion_reasons.append("weekly_distance_above_40000m")
        else:
            if minutes is not None and minutes < 5:
                exclusion_reasons.append("session_minutes_below_5")
            if distance is not None and distance < 200:
                exclusion_reasons.append("session_distance_below_200m")
            if minutes is not None and minutes > 220:
                exclusion_reasons.append("session_minutes_above_220")
            if distance is not None and distance > 20000:
                exclusion_reasons.append("session_distance_above_20000m")
            if minutes is not None and distance is not None and minutes > 0 and (distance / minutes) > 1000:
                exclusion_reasons.append("session_impossible_distance_per_minute")
        if scope_status == "out_of_scope_explicit":
            exclusion_reasons.append(scope_reason)
        if parsed_date and window_start and window_end:
            period_start = parsed_date.date()
            period_end = period_start + (timedelta(days=6) if grain == "weekly" else timedelta())
            if period_end < window_start or period_start > window_end:
                exclusion_reasons.append("outside_official_analysis_window")

        action = "exclude_from_primary" if exclusion_reasons else "include"
        counts[action] = counts.get(action, 0) + 1
        for reason in exclusion_reasons:
            reasons[reason] = reasons.get(reason, 0) + 1
        if action == "include":
            included_minutes += minutes or 0.0
            included_distance += distance or 0.0

        cleaned = dict(row)
        cleaned.update(
            {
                "exposure_grain": grain,
                "scope_status": scope_status,
                "scope_reason": scope_reason,
                "cleaned_date": parsed_date.date().isoformat() if parsed_date else "",
                "week_start_date": parsed_date.date().isoformat() if parsed_date and grain == "weekly" else "",
                "session_date_clean": parsed_date.date().isoformat() if parsed_date and grain == "session" else "",
                "minutes_total_clean": "" if minutes is None else f"{minutes:.6f}",
                "distance_total_m_clean": "" if distance is None else f"{distance:.6f}",
                "cleaning_action": action,
                "exclusion_reason": ";".join(exclusion_reasons),
            }
        )
        cleaned_rows.append(cleaned)

    added_fields = [
        "exposure_grain",
        "scope_status",
        "scope_reason",
        "cleaned_date",
        "week_start_date",
        "session_date_clean",
        "minutes_total_clean",
        "distance_total_m_clean",
        "cleaning_action",
        "exclusion_reason",
    ]
    fieldnames = list(rows[0].keys()) + added_fields
    write_rows(output_path, cleaned_rows, fieldnames)
    output_hash = sha256_file(output_path)
    qc = {
        "file": str(output_path),
        "file_sha256": output_hash,
        "source_file": str(path),
        "source_file_sha256": sha256_file(path),
        "row_count": len(cleaned_rows),
        "action_counts": counts,
        "exclusion_reason_counts": reasons,
        "exposure_grain_counts": grain_counts,
        "date_min": min(dates).date().isoformat() if dates else None,
        "date_max": max(dates).date().isoformat() if dates else None,
        "date_order": args.date_order,
        "team": getattr(args, "team", ""),
        "season": getattr(args, "season", ""),
        "reporting_grain": reporting_grain,
        "reporting_grain_evidence": {
            "selection_source": "required_cli_argument",
            "prepared_row_declaration": "matched_all_rows",
        },
        "analysis_window": {
            "start": window_start.isoformat() if window_start else None,
            "end": window_end.isoformat() if window_end else None,
            "weekly_rule": "retain a weekly row when its seven-day period overlaps the analysis window",
        },
        "included_minutes_total": round(included_minutes, 6),
        "included_distance_total_m": round(included_distance, 6),
        "rules": {
            "canonical_schema_version": EXPOSURE_CANONICAL_SCHEMA_VERSION,
            "canonical_columns": {
                "cleaned_date": "ISO date",
                "exposure_grain": "session or weekly",
                "minutes_total_clean": "minutes",
                "distance_total_m_clean": "metres",
                "scope_status": "scope eligibility",
                "scope_reason": "scope rule outcome",
                "cleaning_action": "primary-analysis eligibility",
                "exclusion_reason": "semicolon-delimited controlled reasons",
            },
            "scope_rule_version": (
                EDINBURGH_EXPOSURE_SCOPE_RULE_VERSION
                if clean_text(getattr(args, "team", "")).lower() == "edinburgh"
                else "global_exposure_scope_v0.1.0"
            ),
            "scope": "Exclude explicit international/national-team exposure, academy, and rehab/RTP. Edinburgh additionally excludes game-labelled sessions without a fuzzy-matched URC opponent, except warm-ups and top-ups are retained; blank scope fields are retained as scope_unknown_included.",
            "edinburgh_urc_opponents": EDINBURGH_URC_OPPONENTS if clean_text(getattr(args, "team", "")).lower() == "edinburgh" else [],
            "urc_opponent_fuzzy_cutoff": URC_OPPONENT_FUZZY_CUTOFF,
            "edinburgh_scope_decisions": [
                "exclude all international and national-team exposure",
                "exclude rehab and RTP exposure",
                "exclude all academy games",
                "retain warm-up and top-up exposure",
                "retain game-labelled exposure only when the bracketed opponent fuzzy-matches one of the other 15 URC teams",
            ] if clean_text(getattr(args, "team", "")).lower() == "edinburgh" else [],
            "reporting_grain_selection": "Required explicit weekly/session intake attribute; never inferred from a team alias.",
            "weekly_exclusions": ["minutes < 5", "minutes > 1100", "distance > 40000m"],
            "session_exclusions": ["minutes < 5", "distance < 200m", "minutes > 220", "distance > 20000m", "distance/minute > 1000"],
            "global_exclusions": ["exact duplicate copy", "missing player/date/minutes/distance", "negative minutes/distance", "minutes = 0 and distance = 0", "outside official analysis window"],
        },
    }
    qc_path.parent.mkdir(parents=True, exist_ok=True)
    qc_path.write_text(json.dumps(qc, indent=2) + "\n")
    if args.manifest:
        manifest_path = Path(args.manifest)
        manifest = json.loads(manifest_path.read_text())
        manifest["exposure_cleaning"] = {
            "team": args.team,
            "season": args.season,
            "cleaned_file": str(output_path),
            "cleaned_file_sha256": output_hash,
            "qc_file": str(qc_path),
            "qc_file_sha256": sha256_file(qc_path),
            "protocol_document": "docs/EXPOSURE_CLEANING_PROTOCOL.md",
            "row_count": len(cleaned_rows),
            "action_counts": counts,
            "exclusion_reason_counts": reasons,
            "exposure_grain_counts": grain_counts,
            "reporting_grain": reporting_grain,
            "reporting_grain_evidence": qc["reporting_grain_evidence"],
            "scope_rule_version": qc["rules"]["scope_rule_version"],
            "canonical_schema_version": qc["rules"]["canonical_schema_version"],
            "analysis_window": qc["analysis_window"],
        }
        manifest_path.write_text(json.dumps(manifest, indent=2) + "\n")
    print(
        json.dumps(
            {
                "cleaned": str(output_path),
                "rows": len(cleaned_rows),
                "action_counts": counts,
                "exclusion_reason_counts": reasons,
                "qc": str(qc_path),
                "sha256": output_hash,
                "reporting_grain": reporting_grain,
            },
            indent=2,
        )
    )


def fetch_standing_eligibility_adjudications(
    team: str, season: str, file_hash: str
) -> dict[int, dict[str, Any]]:
    """Read-only lookup of audit.adjudications rows that carry a standing
    analysis_eligibility_status override for source rows in this exact
    ingested file, keyed by source_row_number. Only field_name =
    'analysis_eligibility_status' decisions carry an eligibility override
    (e.g. the duplicate_adjudicated_exclusion decision written by
    adjudicate-duplicate-exclusion); other adjudication field_names such as
    'duplicate_review' record a reviewed decision without overriding
    eligibility and are intentionally excluded by the `decision ?
    'analysis_eligibility_status'` filter below, so they are correctly
    treated as no-ops here.

    Called by process_intake/process_exposure so a rerun reapplies a prior
    human adjudication instead of silently reverting
    processing.record_versions.eligibility_status to whatever the current
    build_processing_state/cleaning_action computes -- the audit trail
    contract requires manual corrections to be reapplied by the pipeline,
    never lost on rerun.
    """
    lookup_params = SqlParams()
    rows = query_sql(
        f"""
        select sr.source_row_number, a.id as adjudication_id, a.decision, a.rationale
        from audit.adjudications a
        join ingestion.source_rows sr on sr.id = a.source_row_id
        join ingestion.source_files sf on sf.id = sr.source_file_id
        where sf.team = {lookup_params.text(team)}
          and sf.season = {lookup_params.text(season)}
          and sf.file_sha256 = {lookup_params.text(file_hash)}
          and a.field_name = 'analysis_eligibility_status'
          and a.decision ? 'analysis_eligibility_status'
        order by a.decided_at, sr.source_row_number
        """,
        lookup_params.values,
    )
    # decided_at ascending + dict overwrite: the most recent standing decision
    # per source row wins if a row is ever re-adjudicated.
    return {int(row["source_row_number"]): row for row in rows}


def fetch_standing_source_field_adjudications(
    team: str, season: str, file_hash: str
) -> dict[int, list[dict[str, Any]]]:
    """Return approved source-field overlays for one immutable intake file.

    The source representation remains unchanged. process-intake applies these
    allowlisted decisions only in memory before deriving a new record version.
    """
    lookup_params = SqlParams()
    rows = query_sql(
        f"""
        select sr.source_row_number, sr.row_sha256, a.id as adjudication_id,
               a.field_name, a.decision, a.rationale
        from audit.adjudications a
        join ingestion.source_rows sr on sr.id = a.source_row_id
        join ingestion.source_files sf on sf.id = sr.source_file_id
        where sf.team = {lookup_params.text(team)}
          and sf.season = {lookup_params.text(season)}
          and sf.file_sha256 = {lookup_params.text(file_hash)}
          and a.field_name in ('Date Injured', 'Fit For Selection Date')
          and a.decision ->> 'decision_type' = 'source_field_correction'
        order by sr.source_row_number, a.decided_at, a.id
        """,
        lookup_params.values,
    )
    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[int(row["source_row_number"])].append(row)
    return grouped


def apply_source_field_adjudications(
    row: dict[str, Any], adjudications: list[dict[str, Any]]
) -> list[dict[str, object]]:
    """Apply validated manual source corrections to a processing copy only."""
    events: list[dict[str, object]] = []
    for adjudication in adjudications:
        field_name = clean_text(adjudication.get("field_name"))
        if field_name not in {"Date Injured", "Fit For Selection Date"}:
            raise SystemExit(f"source-field adjudication uses disallowed field {field_name!r}")
        decision = adjudication.get("decision")
        if not isinstance(decision, dict):
            raise SystemExit("source-field adjudication decision must be a JSON object")
        expected_row_sha256 = clean_text(decision.get("source_row_sha256"))
        actual_row_sha256 = clean_text(adjudication.get("row_sha256"))
        if not expected_row_sha256 or expected_row_sha256 != actual_row_sha256:
            raise SystemExit(
                f"source-field adjudication {adjudication.get('adjudication_id')} row fingerprint mismatch"
            )
        expected_old = clean_text(decision.get("old_value"))
        actual_old = clean_text(row.get(field_name))
        if actual_old != expected_old:
            raise SystemExit(
                f"source-field adjudication {adjudication.get('adjudication_id')} expected "
                f"{field_name}={expected_old!r}, found {actual_old!r}"
            )
        new_value = clean_text(decision.get("new_value"))
        if not new_value:
            raise SystemExit("source-field correction cannot replace a value with blank")
        row[field_name] = new_value
        if field_name == "Fit For Selection Date":
            # This source column is Cardiff's return-to-availability evidence.
            # Bind it explicitly to the existing allowlisted adapter seam so
            # the generic processing rule can derive duration without learning
            # a team-specific column name.
            row["Adapter Canonical Confirmed Return Date"] = new_value
            row["Adapter Canonical Confirmed Return Date Origin"] = (
                f"manual_adjudication:{clean_text(decision.get('item_id'))}"
            )
        events.append(
            {
                "field_name": field_name,
                "old_value": expected_old or None,
                "new_value": new_value,
                "action": "correct",
                "reason_code": "source_field_adjudicated_correction",
                "rationale": (
                    f"Approved workbook item {clean_text(decision.get('item_id'))}; "
                    f"immutable source value retained. Evidence sha256="
                    f"{clean_text(decision.get('evidence_sha256'))}."
                ),
                "review_status": "adjudicated",
            }
        )
    return events


def adjudicated_derived_change_events(
    old_state: dict[str, Any] | None,
    new_state: dict[str, Any],
    item_ids: list[str],
) -> list[dict[str, object]]:
    """Make every derived change caused by a source correction explicit."""
    if old_state is None:
        return []
    ignored = {"source_locator", "provisional_qc_window"}
    events: list[dict[str, object]] = []
    for field_name in sorted(set(old_state) | set(new_state)):
        if field_name in ignored:
            continue
        old_value = old_state.get(field_name)
        new_value = new_state.get(field_name)
        if old_value == new_value:
            continue
        events.append(
            {
                "field_name": field_name,
                "old_value": old_value,
                "new_value": new_value,
                "action": "rederive_after_correction",
                "reason_code": "source_field_adjudicated_correction",
                "rationale": (
                    f"Derived field changed after approved source correction(s) "
                    f"{', '.join(item_ids)}; prior record version remains immutable."
                ),
                "review_status": "adjudicated",
            }
        )
    return events


def apply_standing_adjudication(
    state: dict[str, Any],
    events: list[dict[str, Any]],
    adjudication: dict[str, Any] | None,
) -> bool:
    """Overlay a standing audit.adjudications decision onto a freshly
    computed record state, appending a 'reapply' record_events entry when
    the adjudicated status differs from what was just computed. Returns True
    if a reapplication happened. Shared by process_intake and
    process_exposure so a rerun never silently drops a decision made through
    the audit.adjudications table (see fetch_standing_eligibility_adjudications).
    Pure/offline so self_check can exercise it without a DB connection.
    """
    if adjudication is None:
        return False
    decision = adjudication["decision"]
    adjudicated_status = clean_text(str(decision.get("analysis_eligibility_status", "")))
    if not adjudicated_status:
        return False
    pre_status = clean_text(str(state.get("analysis_eligibility_status", "")))
    if adjudicated_status == pre_status:
        return False
    state.update(decision)
    state["analysis_eligibility_status"] = adjudicated_status
    decision_rationale = clean_text(str(decision.get("rationale") or adjudication.get("rationale") or ""))
    events.append(
        {
            "field_name": "analysis_eligibility_status",
            "old_value": pre_status or None,
            "new_value": adjudicated_status,
            "action": "reapply",
            "reason_code": "adjudication_reapplied",
            "rationale": (
                f"Standing audit.adjudications decision {adjudication['adjudication_id']} "
                f"reapplied on rerun: {decision_rationale}"
            ),
            "review_status": "adjudicated",
        }
    )
    return True


def reconcile_registered_intake_rows(
    corrected_rows: list[dict[str, str]],
    registered_rows: list[dict[str, Any]],
    *,
    source_date_order: str,
    adapter_qc_sha256: str,
) -> list[dict[str, str]]:
    """Bind a corrected processing artifact to one immutable registered source."""
    if len(corrected_rows) != len(registered_rows):
        raise SystemExit(
            "corrected processing artifact row count does not match the registered source file"
        )
    reconciled: list[dict[str, str]] = []
    identity_fields = (
        "source_file_sha256", "source_sheet", "source_row_number",
        "standardised_row_number", "source_locator_status", "player_uid",
    )
    correction_fields = {"Date Injured", "Confirmed Return Date", "Match Type"}
    regenerated_metadata = {"standardised_file_sha256", "injury_uid"}
    for corrected, registered in zip(corrected_rows, registered_rows, strict=True):
        source_values = registered.get("source_values")
        if not isinstance(source_values, dict):
            raise SystemExit("registered source row is missing its preserved source_values")
        row_number = clean_text(corrected.get("standardised_row_number")) or "<unknown>"
        for field in identity_fields:
            if clean_text(corrected.get(field)) != clean_text(source_values.get(field)):
                raise SystemExit(
                    f"corrected processing artifact does not match registered source row {row_number} "
                    f"on {field}"
                )
        for field, corrected_value in corrected.items():
            if field in correction_fields or field in regenerated_metadata or field in identity_fields:
                continue
            registered_value = clean_text(source_values.get(field))
            if registered_value == "[REDACTED_PROTECTED_METADATA]":
                continue
            if clean_text(corrected_value) != registered_value:
                raise SystemExit(
                    f"corrected processing artifact changes disallowed field {field} on source row {row_number}"
                )

        merged = {key: clean_text(value) for key, value in source_values.items()}
        correction_events: list[dict[str, object]] = []
        for field in ("Date Injured", "Confirmed Return Date"):
            registered_value = clean_text(source_values.get(field))
            parsed = parse_flexible_date(registered_value, source_date_order)
            expected = parsed.strftime("%d/%m/%Y") if parsed else registered_value
            corrected_value = clean_text(corrected.get(field))
            if corrected_value != expected:
                raise SystemExit(
                    f"corrected processing artifact has a non-deterministic {field} on source row {row_number}"
                )
            merged[field] = corrected_value
            if corrected_value != registered_value:
                correction_events.append(
                    {
                        "field_name": field,
                        "old_value": registered_value or None,
                        "new_value": corrected_value or None,
                        "action": "derive",
                        "reason_code": "deterministic_derivation",
                        "rationale": (
                            f"Approved {source_date_order} source-date normalization from adapter QC "
                            f"sha256={adapter_qc_sha256}; immutable source value retained upstream."
                        ),
                        "review_status": "not_required",
                    }
                )

        registered_match_type = clean_text(source_values.get("Match Type"))
        corrected_match_type = clean_text(corrected.get("Match Type"))
        if corrected_match_type != registered_match_type and not (
            not registered_match_type and corrected_match_type == "URC"
        ):
            raise SystemExit(
                f"corrected processing artifact has a disallowed Match Type change on source row {row_number}"
            )
        merged["Match Type"] = corrected_match_type
        if corrected_match_type != registered_match_type:
            correction_events.append(
                {
                    "field_name": "Match Type",
                    "old_value": registered_match_type or None,
                    "new_value": corrected_match_type,
                    "action": "map",
                    "reason_code": "canonical_mapping",
                    "rationale": (
                        "Fixture-derived URC match classification from the checksummed corrected fixture file; "
                        f"adapter QC sha256={adapter_qc_sha256}."
                    ),
                    "review_status": "not_required",
                }
            )
        merged["_registered_source_confirmed_return_date"] = clean_text(
            source_values.get("Confirmed Return Date")
        )
        merged["_bridge_correction_events"] = correction_events  # type: ignore[assignment]
        reconciled.append(merged)
    return reconciled


def duplicate_source_row_numbers(
    rows: list[dict[str, str]], columns: list[str], excluded_row_numbers: set[int] | None = None
) -> set[int]:
    """Return the legacy duplicate candidates for frozen processing paths."""
    excluded = excluded_row_numbers or set()
    counts: dict[str, list[int]] = {}
    for row in rows:
        row_number = int(row["standardised_row_number"])
        if row_number in excluded:
            continue
        key = "|".join(row.get(column, "").strip() for column in columns)
        counts.setdefault(key, []).append(row_number)
    return {
        row_number
        for key, row_numbers in counts.items()
        if key and len(row_numbers) > 1
        for row_number in row_numbers
    }


def year2_duplicate_source_row_numbers(
    rows: list[dict[str, str]], columns: list[str], excluded_row_numbers: set[int] | None = None
) -> set[int]:
    """Return only sufficiently evidenced Year 2 duplicate candidates.

    A separator-only collection of empty values is not evidence of a duplicate.
    Neither is an Unknown identity. The caller retains every source row and
    uses this result as a review flag only, never as an automatic exclusion.
    """
    excluded = excluded_row_numbers or set()
    counts: dict[str, list[int]] = {}
    for row in rows:
        row_number = int(row["standardised_row_number"])
        if row_number in excluded:
            continue
        player_uid = clean_text(row.get("player_uid"))
        injured_at = parse_uk_date(row.get("Date Injured", ""))
        clinical_values = [
            clean_text(row.get(column))
            for column in columns
            if column not in {"PlayerID", "Date Injured"}
            and clean_text(row.get(column)).casefold() not in MISSING_VALUES
        ]
        if not player_uid or player_uid.casefold() == "unknown" or injured_at is None or not clinical_values:
            continue
        key = "|".join([player_uid, injured_at.date().isoformat(), *clinical_values])
        counts.setdefault(key, []).append(row_number)
    return {
        row_number
        for row_numbers in counts.values()
        if len(row_numbers) > 1
        for row_number in row_numbers
    }


def year2_injury_eligibility_vector(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Return the private, deterministic status vector bound to a Year 2 CSV.

    The vector contains only opaque locators, opaque IDs and controlled status
    values. It is hashed in the protected bridge document and is not logged.
    """
    return [
        {
            "standardised_row_number": clean_text(row.get("standardised_row_number")),
            "source_file_sha256": clean_text(row.get("source_file_sha256")),
            "source_sheet": clean_text(row.get("source_sheet")),
            "source_row_number": clean_text(row.get("source_row_number")),
            "injury_uid": clean_text(row.get("injury_uid")),
            "player_uid": clean_text(row.get("player_uid")),
            "injury_date_basis": clean_text(row.get("injury_date_basis")),
            "injury_eligibility_status": clean_text(row.get("injury_eligibility_status")),
        }
        for row in rows
    ]


def validate_year2_injury_eligibility_bridge(
    *, args: argparse.Namespace, rows: list[dict[str, str]], file_hash: str,
) -> dict[int, dict[str, str]]:
    """Validate the explicit Year 2 undated-injury bridge without DB access."""
    bridge_arg = clean_text(getattr(args, "injury_eligibility_bridge_file", ""))
    if args.season != "2025-26":
        if bridge_arg:
            raise SystemExit("the Year 2 injury eligibility bridge is valid only for season 2025-26")
        return {}
    if not bridge_arg:
        raise SystemExit(
            "process-intake for 2025-26 requires a checksum-bound injury eligibility bridge"
        )
    bridge_path = Path(bridge_arg)
    try:
        bridge = json.loads(bridge_path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit("Year 2 injury eligibility bridge must be readable JSON") from exc
    expected_window = {"start": "2025-09-01", "end": "2026-06-30"}
    if not isinstance(bridge, dict) or any(
        bridge.get(key) != value
        for key, value in {
            "schema": "urc_2025_26_injury_eligibility_bridge_v1",
            "rule_version": YEAR2_INJURY_ELIGIBILITY_BRIDGE_RULE_VERSION,
            "team": args.team,
            "season": args.season,
            "injury_file_sha256": file_hash,
            "row_count": len(rows),
            "window": expected_window,
        }.items()
    ):
        raise SystemExit("Year 2 injury eligibility bridge does not bind this exact processing input")
    if args.window_start != expected_window["start"] or args.window_end != expected_window["end"]:
        raise SystemExit("Year 2 injury eligibility bridge requires the registered 2025-26 reporting window")
    expected_vector_hash = sha256_json(year2_injury_eligibility_vector(rows))
    if bridge.get("eligibility_vector_sha256") != expected_vector_hash:
        raise SystemExit("Year 2 injury eligibility bridge status vector checksum mismatch")
    audit_path = clean_text(getattr(args, "analysis_audit_file", ""))
    expected_audit = bridge.get("analysis_audit")
    if not isinstance(expected_audit, dict) or not audit_path:
        raise SystemExit("Year 2 injury eligibility bridge requires its checksum-bound analysis audit")
    try:
        if (
            Path(audit_path).resolve() != (bridge_path.parent / clean_text(expected_audit.get("path"))).resolve()
            or sha256_file(Path(audit_path)) != clean_text(expected_audit.get("sha256"))
        ):
            raise ValueError
    except (OSError, ValueError) as exc:
        raise SystemExit("Year 2 injury eligibility bridge analysis audit checksum mismatch") from exc

    allowed_audit_reasons = {
        clean_text(reason)
        for reason in expected_audit.get("allowed_reason_codes", [])
        if clean_text(reason)
    }
    if not allowed_audit_reasons or not allowed_audit_reasons <= YEAR2_ALLOWED_ANALYSIS_AUDIT_REASONS:
        raise SystemExit("Year 2 injury eligibility bridge has invalid audit reason permissions")
    for audit_row in read_rows(Path(audit_path)):
        if audit_row.get("field") != "analysis_eligibility" or audit_row.get("action") != "exclude":
            raise SystemExit("Year 2 injury eligibility bridge has an invalid analysis audit action")
        try:
            row_number = int(audit_row["standardised_row_number"])
        except (KeyError, ValueError) as exc:
            raise SystemExit("Year 2 injury eligibility bridge analysis audit has an invalid row locator") from exc
        reason = clean_text(audit_row.get("reason"))
        if row_number not in {int(row["standardised_row_number"]) for row in rows} or reason not in allowed_audit_reasons:
            raise SystemExit("Year 2 injury eligibility bridge analysis audit is not permitted")

    bridge_rows: dict[int, dict[str, str]] = {}
    for row in rows:
        try:
            row_number = int(row["standardised_row_number"])
        except (KeyError, ValueError) as exc:
            raise SystemExit("Year 2 injury eligibility bridge input has an invalid row locator") from exc
        date_basis = clean_text(row.get("injury_date_basis"))
        eligibility = clean_text(row.get("injury_eligibility_status"))
        injured_at = parse_uk_date(row.get("Date Injured", ""))
        if date_basis not in YEAR2_INJURY_DATE_BASES or eligibility not in YEAR2_INJURY_BRIDGE_ELIGIBILITY:
            raise SystemExit("Year 2 injury eligibility bridge has an invalid controlled status")
        valid = (
            (date_basis == "season_attributed_undated" and injured_at is None and not clean_text(row.get("Date Injured")) and eligibility == "included_pending_protocol")
            or (date_basis == "source_date_within_window" and injured_at is not None and expected_window["start"] <= injured_at.date().isoformat() <= expected_window["end"] and eligibility == "included_pending_protocol")
            or (date_basis == "source_date_unparseable" and bool(clean_text(row.get("Date Injured"))) and injured_at is None and eligibility == "review_required")
            or (date_basis == "source_date_outside_window" and injured_at is not None and not (expected_window["start"] <= injured_at.date().isoformat() <= expected_window["end"]) and eligibility == "review_required")
        )
        if not valid or row_number in bridge_rows:
            raise SystemExit("Year 2 injury eligibility bridge has an invalid row basis")
        bridge_rows[row_number] = {
            "date_basis": date_basis,
            "eligibility_status": eligibility,
        }
    return bridge_rows


def process_exposure(args: argparse.Namespace) -> None:
    path = Path(args.file)
    rows = read_rows(path)
    if not rows:
        raise SystemExit("no cleaned exposure rows found")
    reporting_grain = required_exposure_reporting_grain(args)
    observed_grains = {clean_text(row.get("exposure_grain")) for row in rows}
    if observed_grains != {reporting_grain}:
        observed_label = ", ".join(sorted(grain or "<blank>" for grain in observed_grains))
        raise SystemExit(
            f"--reporting-grain {reporting_grain!r} does not match cleaned exposure row grain(s): "
            f"{observed_label}"
        )
    file_hash = sha256_file(path)
    version_number = args.version_number
    included_rows = sum(1 for row in rows if row.get("cleaning_action") == "include")
    excluded_rows = sum(1 for row in rows if row.get("cleaning_action") == "exclude_from_primary")
    reason_counts: dict[str, int] = {}
    grain_counts: dict[str, int] = {}
    scope_counts: dict[str, int] = {}
    record_sql = []
    event_sql = []
    output_states: list[dict[str, object]] = []
    params = SqlParams()
    provenance = run_provenance()
    standing_adjudications = fetch_standing_eligibility_adjudications(args.team, args.season, file_hash)
    reapplied_adjudication_rows = 0

    for index, row in enumerate(rows, start=2):
        raw_id = raw_record_id(args.team, args.season, file_hash, index)
        exclusion_reasons = [
            reason for reason in row.get("exclusion_reason", "").split(";") if reason
        ]
        for reason in exclusion_reasons:
            reason_counts[reason] = reason_counts.get(reason, 0) + 1
        grain = row.get("exposure_grain", "")
        scope = row.get("scope_status", "")
        if grain:
            grain_counts[grain] = grain_counts.get(grain, 0) + 1
        if scope:
            scope_counts[scope] = scope_counts.get(scope, 0) + 1
        action = row.get("cleaning_action", "")
        eligibility = "included_pending_protocol" if action == "include" else "excluded_from_primary"
        state = {
            "player_uid": row.get("player_uid") or None,
            "exposure_grain": grain or None,
            "scope_status": scope or None,
            "scope_reason": row.get("scope_reason") or None,
            "cleaned_date": row.get("cleaned_date") or None,
            "week_start_date": row.get("week_start_date") or None,
            "session_date": row.get("session_date_clean") or None,
            "minutes_total_clean": parse_float(row.get("minutes_total_clean")),
            "distance_total_m_clean": parse_float(row.get("distance_total_m_clean")),
            "cleaning_action": action or None,
            "exclusion_reasons": exclusion_reasons,
            "analysis_eligibility_status": eligibility,
            "source_locator": {
                field: row.get(field)
                for field in EXPOSURE_LOCATOR_FIELDS
                if field in row
            },
        }
        reapply_events: list[dict[str, object]] = []
        if apply_standing_adjudication(state, reapply_events, standing_adjudications.get(index)):
            reapplied_adjudication_rows += 1
            eligibility = clean_text(str(state["analysis_eligibility_status"]))
        output_states.append(state)
        record_sql.append(
            f"""
            insert into processing.record_versions
              (source_row_id, step_run_id, version_number, record_state, eligibility_status)
            select sr.id, step.id, {version_number}, {params.jsonb(state)},
              {params.text(eligibility)}
            from ingestion.source_rows sr, current_step step
            where sr.raw_record_id = {params.text(raw_id)}
            ;
            """
        )
        audit_reasons = exclusion_reasons or ["exposure_cleaning_applied"]
        for reason in audit_reasons:
            event_sql.append(
                f"""
                insert into audit.record_events
                  (step_run_id, source_row_id, field_name, old_value, new_value, action, reason_code, rationale, rule_version, review_status)
                select step.id, sr.id, 'analysis_eligibility_status', null,
                  {params.jsonb(eligibility)}, {params.text('exclude' if exclusion_reasons else 'classify')},
                  {params.text(reason)},
                  {params.text(f'Exposure eligibility set by controlled reason: {reason}. Source values are preserved.')},
                  {params.text(args.step_version)}, 'not_required'
                from ingestion.source_rows sr, current_step step
                where sr.raw_record_id = {params.text(raw_id)};
                """
            )
        for event in reapply_events:
            event_sql.append(
                f"""
                insert into audit.record_events
                  (step_run_id, source_row_id, field_name, old_value, new_value, action, reason_code, rationale, rule_version, review_status)
                select step.id, sr.id, {params.text(event["field_name"])}, {params.jsonb(event["old_value"])},
                  {params.jsonb(event["new_value"])}, {params.text(event["action"])},
                  {params.text(event["reason_code"])}, {params.text(event["rationale"])},
                  {params.text(args.step_version)}, {params.text(event["review_status"])}
                from ingestion.source_rows sr, current_step step
                where sr.raw_record_id = {params.text(raw_id)};
                """
            )

    output_hash = sha256_json(output_states)
    exposure_reason_codes = sorted(reason_counts)
    reason_code_sql = "".join(
        f"insert into audit.reason_codes (code, description) values ({params.text(reason)}, {params.text('Exposure exclusion reason emitted by the versioned cleaning protocol.')}) on conflict (code) do update set description = excluded.description;"
        for reason in exposure_reason_codes
    )
    if reapplied_adjudication_rows and not query_sql(
        "select 1 from audit.reason_codes where code = 'adjudication_reapplied'"
    ):
        raise SystemExit(
            "process-exposure would reapply a standing audit.adjudications decision but reason code "
            "'adjudication_reapplied' is not seeded; run migration 20260710000137_adjudication_reapplication first"
        )

    sql = f"""
      insert into audit.reason_codes (code, description) values
        ('exposure_cleaning_applied', 'Exposure cleaning protocol applied and analysis eligibility recorded.'),
        ('exposure_no_exclusions', 'Exposure cleaning protocol produced zero exclusions for this file.'),
        ('exposure_exclusion', 'Exposure row excluded from the primary denominator by a protocol-defined rule.')
      on conflict (code) do update set description = excluded.description;

      {reason_code_sql}

      do $$
      begin
        if (select count(*) from ingestion.source_rows sr join ingestion.source_files sf on sf.id = sr.source_file_id where sf.team = {params.text(args.team)} and sf.season = {params.text(args.season)} and sf.file_sha256 = {params.text(file_hash)}) <> {len(rows)} then
          raise exception 'process-exposure requires every source row to be registered';
        end if;
        if exists (select 1 from processing.record_versions rv join ingestion.source_rows sr on sr.id = rv.source_row_id join ingestion.source_files sf on sf.id = sr.source_file_id where sf.team = {params.text(args.team)} and sf.season = {params.text(args.season)} and sf.file_sha256 = {params.text(file_hash)} and rv.version_number = {version_number}) then
          raise exception 'process-exposure version already exists';
        end if;
      end $$;

      create temp table current_step on commit drop as
      with run as (
        insert into audit.pipeline_runs
          (command, team, season, status, input_hash, output_hash, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values (
          'process-exposure', {params.text(args.team)}, {params.text(args.season)}, 'succeeded',
          {params.text(file_hash)}, {params.text(output_hash)},
          {params.jsonb({
            'file': path.name,
            'step': args.step_name,
            'step_version': args.step_version,
            'version_number': version_number,
            'reporting_grain': reporting_grain,
          })},
          now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      ),
      step as (
        insert into audit.step_runs
          (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count, counts_by_team, input_hash, output_hash, ended_at)
        select id, {params.text(args.step_name)}, {params.text(args.step_version)},
          {params.text('exposure_no_exclusions' if excluded_rows == 0 else 'exposure_exclusion')},
          {len(rows)}, {included_rows},
          {params.jsonb({
            args.team: {
              'rows': len(rows),
              'included_rows': included_rows,
              'excluded_from_primary_rows': excluded_rows,
              'exclusion_reason_counts': reason_counts,
              'exposure_grain_counts': grain_counts,
              'scope_status_counts': scope_counts,
              'reapplied_adjudication_rows': reapplied_adjudication_rows,
            }
          })},
          {params.text(file_hash)}, {params.text(output_hash)}, now()
        from run
        returning id
      )
      select id from step;

      {"".join(record_sql)}
      {"".join(event_sql)}
    """
    run_sql(sql, params.values)
    print(
        json.dumps(
            {
                "processed": str(path),
                "rows": len(rows),
                "included_rows": included_rows,
                "excluded_from_primary_rows": excluded_rows,
                "exclusion_reason_counts": reason_counts,
                "exposure_grain_counts": grain_counts,
                "reporting_grain": reporting_grain,
                "scope_status_counts": scope_counts,
                "record_events": len(event_sql),
                "reapplied_adjudication_rows": reapplied_adjudication_rows,
            },
            indent=2,
        )
    )


def split_fixture(fixture: str) -> tuple[str, str]:
    parts = fixture.split(" v ")
    if len(parts) != 2:
        raise SystemExit(f"fixture must contain one ' v ': {fixture}")
    return parts[0].strip(), parts[1].strip()


def read_total_exposure_hours(path: Path, team_column: str, hours_column: str) -> dict[str, float]:
    if not path.exists():
        return {}
    rows = read_rows(path)
    if not rows:
        return {}
    missing = [column for column in [team_column, hours_column] if column not in rows[0]]
    if missing:
        raise SystemExit(f"missing total exposure column(s) in {path}: {', '.join(missing)}")
    totals: dict[str, float] = {}
    for row in rows:
        team = clean_text(row.get(team_column))
        hours = parse_float(row.get(hours_column))
        if team and hours is not None:
            totals[team] = hours
    return totals


def build_fixture_exposure(args: argparse.Namespace) -> None:
    source_path = Path(args.file)
    preserved_path = Path(args.preserved_output)
    fixture_output = Path(args.fixture_output)
    exposure_output = Path(args.exposure_output)
    qc_output = Path(args.qc_output)
    rows = read_rows(source_path)
    if not rows:
        raise SystemExit(f"no fixture rows found: {source_path}")

    preserved_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source_path, preserved_path)

    source_hash = sha256_file(source_path)
    fixture_team_aliases = load_fixture_team_aliases()
    corrected_rows: list[dict[str, str]] = []
    team_matches: dict[str, int] = defaultdict(int)
    corrections_applied = 0
    unresolved_teams: list[str] = []

    for source_row_number, row in enumerate(rows, start=2):
        fixture = clean_text(row.get("fixture"))
        home, away = split_fixture(fixture)
        home_alias = fixture_team_aliases.get(home, "")
        away_alias = fixture_team_aliases.get(away, "")
        if not home_alias:
            unresolved_teams.append(home)
        if not away_alias:
            unresolved_teams.append(away)

        source_date = parse_flexible_date(row.get("date"), args.date_order)
        if source_date is None:
            raise SystemExit(f"unparseable fixture date on source row {source_row_number}: {row.get('date')}")
        corrected_date = FIXTURE_DATE_CORRECTIONS.get(fixture, source_date.date().isoformat())
        if corrected_date != source_date.date().isoformat():
            corrections_applied += 1
        if home_alias:
            team_matches[home_alias] += 1
        if away_alias:
            team_matches[away_alias] += 1

        corrected_rows.append(
            {
                **row,
                "source_date": source_date.date().isoformat(),
                "corrected_date": corrected_date,
                "date_status": "corrected" if fixture in FIXTURE_DATE_CORRECTIONS else "source_confirmed",
                "home_team": home,
                "away_team": away,
                "home_team_alias": home_alias,
                "away_team_alias": away_alias,
                "match_hours_per_team": f"{args.player_hours_per_team_match:.6f}",
                "source_file_sha256": source_hash,
                "source_row_number": str(source_row_number),
            }
        )

    unresolved_unique = sorted(set(unresolved_teams))
    if unresolved_unique:
        raise SystemExit(f"unmapped fixture team(s): {', '.join(unresolved_unique)}")

    total_hours = read_total_exposure_hours(
        Path(args.total_exposure_file), args.total_team_column, args.total_hours_column
    )
    exposure_rows = []
    for team_alias in sorted(team_matches):
        matches = team_matches[team_alias]
        match_hours = matches * args.player_hours_per_team_match
        total = total_hours.get(team_alias)
        exposure_rows.append(
            {
                "team_alias": team_alias,
                "matches": str(matches),
                "match_hours": f"{match_hours:.2f}",
                "total_hours": "" if total is None else f"{total:.2f}",
                "training_hours": "" if total is None else f"{total - match_hours:.2f}",
                "training_formula": "total_hours - match_hours",
            }
        )

    stage_counts: dict[str, int] = {}
    for row in corrected_rows:
        stage = clean_text(row.get("stage"))
        stage_counts[stage] = stage_counts.get(stage, 0) + 1

    fixture_fields = list(rows[0].keys()) + [
        "source_date",
        "corrected_date",
        "date_status",
        "home_team",
        "away_team",
        "home_team_alias",
        "away_team_alias",
        "match_hours_per_team",
        "source_file_sha256",
        "source_row_number",
    ]
    write_rows(fixture_output, corrected_rows, fixture_fields)
    write_rows(
        exposure_output,
        exposure_rows,
        ["team_alias", "matches", "match_hours", "total_hours", "training_hours", "training_formula"],
    )

    qc = {
        "source_file": str(source_path),
        "source_file_sha256": source_hash,
        "preserved_source_file": str(preserved_path),
        "preserved_source_file_sha256": sha256_file(preserved_path),
        "corrected_fixture_file": str(fixture_output),
        "corrected_fixture_file_sha256": sha256_file(fixture_output),
        "team_exposure_file": str(exposure_output),
        "team_exposure_file_sha256": sha256_file(exposure_output),
        "fixture_rows": len(rows),
        "stage_counts": stage_counts,
        "date_corrections_applied": corrections_applied,
        "date_corrections_expected": len(FIXTURE_DATE_CORRECTIONS),
        "date_order": args.date_order,
        "player_hours_per_team_match": args.player_hours_per_team_match,
        "training_exposure_rule": "always total_hours - match_hours",
        "total_exposure_file": str(args.total_exposure_file),
        "total_exposure_file_sha256": (
            sha256_file(Path(args.total_exposure_file))
            if Path(args.total_exposure_file).exists()
            else None
        ),
    }
    qc_output.parent.mkdir(parents=True, exist_ok=True)
    qc_output.write_text(json.dumps(qc, indent=2) + "\n")
    playoff_rows = len(rows) - stage_counts.get("Regular season", 0)
    if len(rows) != 151 or stage_counts.get("Regular season") != 144 or playoff_rows != 7:
        raise SystemExit(f"unexpected fixture structure: rows={len(rows)} stage_counts={stage_counts}")
    if corrections_applied != len(FIXTURE_DATE_CORRECTIONS):
        raise SystemExit(
            f"date corrections mismatch: applied={corrections_applied} expected={len(FIXTURE_DATE_CORRECTIONS)}"
        )
    print(
        json.dumps(
            {
                "preserved": str(preserved_path),
                "fixtures": str(fixture_output),
                "team_exposure": str(exposure_output),
                "qc": str(qc_output),
                "fixture_rows": len(rows),
                "stage_counts": stage_counts,
                "date_corrections_applied": corrections_applied,
            },
            indent=2,
        )
    )


def qa_intake(args: argparse.Namespace) -> None:
    path = Path(args.file)
    rows = read_rows(path)
    if not rows:
        raise SystemExit("no rows found")
    missing_locator_fields = [field for field in LOCATOR_FIELDS + UID_FIELDS if field not in rows[0]]
    if missing_locator_fields:
        raise SystemExit(f"missing locator fields: {', '.join(missing_locator_fields)}")

    def value(row: dict[str, str], key: str) -> str:
        return row.get(key, "").strip()

    exact_seen: set[str] = set()
    exact_duplicates = 0
    key_counts: dict[str, dict[str, list[int]]] = {"injury_signature": {}}
    outside_window: list[int] = []
    unparseable_dates: list[int] = []
    derived_return_dates = 0

    window_start = datetime.strptime(args.window_start, "%Y-%m-%d")
    window_end = datetime.strptime(args.window_end, "%Y-%m-%d")

    for index, row in enumerate(rows, start=2):
        row_hash = hashlib.sha256(json.dumps(row, sort_keys=True).encode()).hexdigest()
        if row_hash in exact_seen:
            exact_duplicates += 1
        exact_seen.add(row_hash)

        keys = {"injury_signature": "|".join(value(row, field) for field in DUPLICATE_SIGNATURE_FIELDS)}
        for name, key in keys.items():
            key_counts[name].setdefault(key, []).append(index)

        injured_at = parse_uk_date(value(row, "Date Injured"))
        if injured_at is None:
            unparseable_dates.append(index)
        elif injured_at < window_start or injured_at > window_end:
            outside_window.append(index)

        days = effective_days_injured(row)
        if injured_at and days is not None:
            _ = injured_at + timedelta(days=days)
            derived_return_dates += 1

    duplicate_keys = {
        name: [
            {"key_hash": stable_uid("key", name, key), "rows": indices}
            for key, indices in counts.items()
            if key and len(indices) > 1
        ]
        for name, counts in key_counts.items()
    }

    report = {
        "file": str(path),
        "file_sha256": sha256_file(path),
        "row_count": len(rows),
        "exact_duplicate_rows": exact_duplicates,
        "duplicate_key_groups": duplicate_keys,
        "date_injured_unparseable_rows": unparseable_dates,
        "outside_provisional_qc_window_rows": outside_window,
        "provisional_qc_window": {
            "start": args.window_start,
            "end": args.window_end,
        },
        "derived_return_date_rows": derived_return_dates,
        "return_date_rule": "For Scottish-team rows, Days Injured comes from Training Days Missed and excludes the injury day; source return dates are preserved or corrected only with recorded evidence.",
        "notes": [
            "Rows are flagged for review or analysis exclusion; source rows are not deleted.",
            "Duplicate keys are hashed in this report to avoid printing player or injury identifiers.",
        ],
    }
    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(report, indent=2) + "\n")
    print(json.dumps(report, indent=2))


def build_processing_state(
    row: dict[str, str],
    *,
    window_start: datetime,
    window_end: datetime,
    duplicate_signature_rows: set[int],
    own_team_alias: str | None = None,
    injury_eligibility_bridge: dict[str, str] | None = None,
) -> tuple[dict[str, object], list[dict[str, object]]]:
    source_row_number = int(row["standardised_row_number"])
    injured_at = parse_uk_date(row.get("Date Injured", ""))
    days_injured, days_injured_origin = effective_days_injured_with_origin(row)
    is_closed, is_closed_origin = injury_closed(row)
    effective_return_date, return_date_origin = effective_confirmed_return_date(
        row, days_injured, days_injured_origin
    )
    derived_return_date = effective_return_date.isoformat() if effective_return_date else None

    activity, activity_origin = activity_context(row)
    contact, contact_origin = contact_context(row)
    recurrence, recurrence_origin = recurrence_status(row)
    severity, severity_origin = severity_category(days_injured, is_closed)
    body, body_origin = body_location(row)
    problem, problem_origin = problem_type(row)
    injury, injury_origin = injury_type(row)
    # Phase 3.5 cohort-signal capture (Adjudication 4): own_team_alias is
    # only ever used for this one in-memory comparison and is discarded
    # immediately after; only the resulting category is stored below.
    received_status, received_status_origin = received_in_team_status(
        row, own_team_alias if own_team_alias else ""
    )
    match_scope, match_scope_origin = urc_match_scope(row)
    if derived_return_date and is_closed is False:
        return_date_origin = f"{return_date_origin}_unclosed_censored"

    if injury_eligibility_bridge is None:
        outside_window = injured_at is None or injured_at < window_start or injured_at > window_end
        injury_date_basis = None
        bridge_eligibility = None
    else:
        injury_date_basis = injury_eligibility_bridge["date_basis"]
        bridge_eligibility = injury_eligibility_bridge["eligibility_status"]
        outside_window = injury_date_basis in {
            "source_date_unparseable", "source_date_outside_window"
        }
    duplicate_flags = {
        "candidate_duplicate_injury_signature": source_row_number in duplicate_signature_rows,
    }
    review_required = (
        (injury_eligibility_bridge is None and any(duplicate_flags.values()))
        or outside_window
        or (injured_at is None and injury_eligibility_bridge is None)
        or bridge_eligibility == "review_required"
    )
    state = {
        "player_uid": row["player_uid"],
        "injury_uid": row["injury_uid"],
        "date_injured": injured_at.date().isoformat() if injured_at else None,
        "days_injured_source": days_injured,
        "source_confirmed_return_date": clean_text(
            row.get("_registered_source_confirmed_return_date")
            or row.get("Confirmed Return Date")
        ) or None,
        "derived_return_date": derived_return_date,
        "return_date_origin": return_date_origin,
        "is_closed": is_closed,
        "activity_context": activity,
        "contact_context": contact,
        "recurrence_status": recurrence,
        "severity_category": severity,
        "body_location": body,
        "problem_type": problem,
        "injury_type": injury,
        "received_in_team_status": received_status,
        "urc_match_scope": match_scope,
        "field_origins": {
            "is_closed": is_closed_origin,
            "days_injured": days_injured_origin,
            "activity_context": activity_origin,
            "contact_context": contact_origin,
            "recurrence_status": recurrence_origin,
            "severity_category": severity_origin,
            "body_location": body_origin,
            "problem_type": problem_origin,
            "injury_type": injury_origin,
            "received_in_team_status": received_status_origin,
            "urc_match_scope": match_scope_origin,
        },
        "provisional_qc_window": {
            "start": window_start.date().isoformat(),
            "end": window_end.date().isoformat(),
        },
        "season_window_status": (
            "season_attributed_undated"
            if injury_date_basis == "season_attributed_undated"
            else "outside_provisional_qc_window" if outside_window else "inside_provisional_qc_window"
        ),
        "analysis_eligibility_status": "review_required" if review_required else "included_pending_protocol",
        "duplicate_flags": duplicate_flags,
        "source_locator": {
            field: row[field]
            for field in LOCATOR_FIELDS
        },
    }
    if injury_date_basis is not None:
        state["injury_date_basis"] = injury_date_basis

    events: list[dict[str, object]] = list(row.get("_bridge_correction_events", []))
    if injury_date_basis == "season_attributed_undated":
        events.append(
            {
                "field_name": "injury_date_basis",
                "old_value": None,
                "new_value": injury_date_basis,
                "action": "derive",
                "reason_code": "season_attributed_undated_injury",
                "rationale": "The checksum-bound Year 2 package attributes this blank source injury date to the registered season; no date was fabricated.",
                "review_status": "not_required",
            }
        )
    if derived_return_date:
        events.append(
            {
                "field_name": "derived_return_date",
                "old_value": None,
                "new_value": derived_return_date,
                "action": "derive",
                "reason_code": "derived_return_date",
                "rationale": "Derived from Date Injured + Days Injured; source Confirmed Return Date is preserved separately.",
                "review_status": "not_required",
            }
        )
    if days_injured is not None:
        events.append(
            {
                "field_name": "days_injured",
                "old_value": row.get("Days Injured", "").strip() or None,
                "new_value": days_injured,
                "action": "infer" if days_injured_origin.startswith("inferred_") else "derive",
                "reason_code": "controlled_inference" if days_injured_origin.startswith("inferred_") else "deterministic_derivation",
                "rationale": f"Effective days injured set by {days_injured_origin}; source duration fields preserved.",
                "review_status": "needs_review" if days_injured_origin.startswith("inferred_") else "not_required",
            }
        )
    for field_name in [
        "activity_context",
        "contact_context",
        "recurrence_status",
        "severity_category",
        "body_location",
        "is_closed",
        "problem_type",
        "injury_type",
    ]:
        origin = state["field_origins"][field_name]
        action = "infer" if str(origin).startswith("inferred_") else "map"
        events.append(
            {
                "field_name": field_name,
                "old_value": None,
                "new_value": state[field_name],
                "action": action,
                "reason_code": "controlled_inference" if action == "infer" else "canonical_mapping",
                "rationale": f"{field_name} set by {origin}; source value preserved.",
                "review_status": "needs_review" if action == "infer" else "not_required",
            }
        )
    for cohort_signal_field, cohort_signal_value, cohort_signal_origin in (
        ("received_in_team_status", received_status, received_status_origin),
        ("urc_match_scope", match_scope, match_scope_origin),
    ):
        events.append(
            {
                "field_name": cohort_signal_field,
                "old_value": None,
                "new_value": cohort_signal_value,
                "action": "derive",
                "reason_code": "cohort_signal_derivation",
                "rationale": (
                    f"{cohort_signal_field} set by {cohort_signal_origin}; source column value preserved "
                    "upstream, the protected team alias itself is never stored here."
                ),
                "review_status": "not_required",
            }
        )
    if duplicate_flags["candidate_duplicate_injury_signature"]:
        events.append(
            {
                "field_name": "candidate_duplicate_injury_signature",
                "old_value": None,
                "new_value": True,
                "action": "flag",
                "reason_code": "candidate_duplicate",
                "rationale": "Same player, date, diagnostic evidence, body part, side, description, and onset appear on more than one ingested row; row retained for review.",
                "review_status": "needs_review",
            }
        )
    if outside_window:
        events.append(
            {
                "field_name": "season_window_status",
                "old_value": None,
                "new_value": state["season_window_status"],
                "action": "flag",
                "reason_code": "outside_provisional_window",
                "rationale": "Date Injured falls outside the provisional July-to-June QC window or is unparseable; row retained.",
                "review_status": "needs_review",
            }
        )
    return state, events


def process_intake(args: argparse.Namespace) -> None:
    path = Path(args.file)
    rows = read_rows(path)
    if not rows:
        raise SystemExit("no rows found")
    missing_locator_fields = [field for field in LOCATOR_FIELDS + UID_FIELDS if field not in rows[0]]
    if missing_locator_fields:
        raise SystemExit(f"missing locator fields: {', '.join(missing_locator_fields)}")

    window_start = datetime.strptime(args.window_start, "%Y-%m-%d")
    window_end = datetime.strptime(args.window_end, "%Y-%m-%d")
    file_hash = sha256_file(path)
    year2_eligibility_bridge = validate_year2_injury_eligibility_bridge(
        args=args, rows=rows, file_hash=file_hash
    )
    registered_source_file_sha256 = clean_text(
        getattr(args, "registered_source_file_sha256", "")
    )
    registered_file_hash = registered_source_file_sha256 or file_hash
    manifest_arg = clean_text(getattr(args, "manifest", ""))
    if registered_source_file_sha256:
        if not manifest_arg:
            raise SystemExit(
                "--manifest is required with --registered-source-file-sha256"
            )
        manifest_path = Path(manifest_arg)
        manifest = json.loads(manifest_path.read_text())
        validate_intake_profile_manifest(
            manifest, manifest_path, file_hash, args.team, args.season
        )
        adapter_qc_arg = clean_text(getattr(args, "adapter_qc_file", ""))
        if not adapter_qc_arg:
            raise SystemExit("--adapter-qc-file is required with --registered-source-file-sha256")
        adapter_qc_path = Path(adapter_qc_arg)
        adapter_qc = json.loads(adapter_qc_path.read_text())
        adapter_qc_sha256 = sha256_file(adapter_qc_path)
        source_date_order = clean_text(adapter_qc.get("date_order"))
        standardised_hashes = {
            clean_text(row.get("standardised_file_sha256")) for row in rows
        }
        correction = manifest.get("processing_correction")
        expected_correction = {
            "schema_version": "registered_source_correction_v1",
            "registered_source_file_sha256": registered_source_file_sha256,
            "processing_artifact_sha256": file_hash,
            "adapter_qc_sha256": adapter_qc_sha256,
            "adapter_rule_version": "sa_injury_boundary_adapter_2026-07-13_v2",
            "allowed_fields": ["Date Injured", "Confirmed Return Date", "Match Type"],
            "approved_by": "Abdel Babiker",
        }
        if not isinstance(correction, dict) or any(
            correction.get(key) != value for key, value in expected_correction.items()
        ):
            raise SystemExit("manifest processing_correction does not bind the exact approved bridge")
        if (
            correction.get("reason_code") != "input_representation_correction"
            or not clean_text(correction.get("rationale"))
        ):
            raise SystemExit("manifest processing_correction requires its controlled reason and rationale")
        try:
            correction_approved_at = datetime.fromisoformat(
                clean_text(correction.get("approved_at")).replace("Z", "+00:00")
            )
            if (
                correction_approved_at.tzinfo is None
                or correction_approved_at.astimezone(UTC) > datetime.now(UTC) + timedelta(minutes=5)
            ):
                raise ValueError
        except ValueError as exc:
            raise SystemExit(
                "manifest processing_correction approved_at must be a valid non-future timezone-aware value"
            ) from exc
        if (
            adapter_qc.get("rule_version") != "sa_injury_boundary_adapter_2026-07-13_v2"
            or clean_text(adapter_qc.get("team")).casefold() != clean_text(args.team).casefold()
            or clean_text(adapter_qc.get("season")) != args.season
            or source_date_order not in {"month-first", "day-first"}
            or standardised_hashes != {clean_text(adapter_qc.get("output_file_sha256"))}
            or clean_text(adapter_qc.get("fixture_file_sha256")) != URC_FIXTURES_2024_25_CORRECTED_SHA256
            or int(adapter_qc.get("output_rows", -1)) != len(rows)
        ):
            raise SystemExit("adapter QC does not match the corrected processing artifact")
        if args.step_version != INPUT_REPRESENTATION_CORRECTION_RULE_VERSION:
            raise SystemExit(
                "registered-source correction requires --step-version "
                f"{INPUT_REPRESENTATION_CORRECTION_RULE_VERSION}"
            )
        bridge_params = SqlParams()
        registered_rows = query_sql(
            f"""
            select sr.source_row_number, sr.source_values
            from ingestion.source_rows sr
            join ingestion.source_files sf on sf.id = sr.source_file_id
            where sf.team = {bridge_params.text(args.team)}
              and sf.season = {bridge_params.text(args.season)}
              and sf.file_sha256 = {bridge_params.text(registered_source_file_sha256)}
            order by sr.source_row_number
            """,
            bridge_params.values,
        )
        rows = reconcile_registered_intake_rows(
            rows,
            registered_rows,
            source_date_order=source_date_order,
            adapter_qc_sha256=adapter_qc_sha256,
        )
    else:
        adapter_qc_path = None
        adapter_qc_sha256 = None
    analysis_audit_file = clean_text(getattr(args, "analysis_audit_file", ""))
    analysis_audit_path = Path(analysis_audit_file) if analysis_audit_file else None
    analysis_audit_rows = read_rows(analysis_audit_path) if analysis_audit_path else []
    analysis_exclusions: dict[int, list[dict[str, str]]] = defaultdict(list)
    for event in analysis_audit_rows:
        if event.get("field") != "analysis_eligibility" or event.get("action") != "exclude":
            continue
        try:
            row_number = int(event["standardised_row_number"])
        except (KeyError, ValueError) as exc:
            raise SystemExit("analysis audit contains an invalid standardised row number") from exc
        analysis_exclusions[row_number].append(event)
    known_row_numbers = {int(row["standardised_row_number"]) for row in rows}
    unknown_audit_rows = sorted(set(analysis_exclusions) - known_row_numbers)
    if unknown_audit_rows:
        raise SystemExit(f"analysis audit references unknown rows: {unknown_audit_rows[:20]}")
    if year2_eligibility_bridge:
        unexpected_year2_reasons = {
            clean_text(event.get("reason"))
            for events in analysis_exclusions.values()
            for event in events
        } - YEAR2_ALLOWED_ANALYSIS_AUDIT_REASONS
        if unexpected_year2_reasons:
            raise SystemExit(
                "Year 2 injury eligibility bridge permits only the seeded "
                "explicit_source_exclusion analysis-audit reason"
            )

    duplicate_rows_fn = (
        year2_duplicate_source_row_numbers
        if year2_eligibility_bridge
        else duplicate_source_row_numbers
    )
    duplicate_signature_rows = duplicate_rows_fn(
        rows, DUPLICATE_SIGNATURE_FIELDS, set(analysis_exclusions)
    )
    standing_adjudications = fetch_standing_eligibility_adjudications(
        args.team, args.season, registered_file_hash
    )
    source_field_adjudications = fetch_standing_source_field_adjudications(
        args.team, args.season, registered_file_hash
    )
    source_field_adjudication_count = sum(
        len(decisions) for decisions in source_field_adjudications.values()
    )
    prior_state_params = SqlParams()
    prior_state_rows = query_sql(
        f"""
        select distinct on (sr.source_row_number)
          sr.source_row_number, rv.record_state
        from ingestion.source_rows sr
        join ingestion.source_files sf on sf.id = sr.source_file_id
        join processing.record_versions rv on rv.source_row_id = sr.id
        where sf.team = {prior_state_params.text(args.team)}
          and sf.season = {prior_state_params.text(args.season)}
          and sf.file_sha256 = {prior_state_params.text(registered_file_hash)}
        order by sr.source_row_number, rv.version_number desc
        """,
        prior_state_params.values,
    )
    prior_states = {
        int(prior["source_row_number"]): prior["record_state"]
        for prior in prior_state_rows
        if isinstance(prior.get("record_state"), dict)
    }

    # Phase 3.5 cohort-signal capture (Adjudication 4): resolved once per
    # process-intake run, then discarded after this point -- passed only
    # into the pure per-row classifier, never logged, never written to SQL
    # params, never entering record_state. Requires the migration that seeds
    # the 'cohort_signal_derivation' reason code (mirrors the
    # adjudication_reapplied precondition below).
    if not query_sql("select 1 from audit.reason_codes where code = 'cohort_signal_derivation'"):
        raise SystemExit(
            "process-intake requires reason code 'cohort_signal_derivation' to be seeded first; "
            "run the Phase 3.5 cohort-signal-columns migration before rerunning process-intake"
        )
    if year2_eligibility_bridge:
        required_year2_reason_codes = query_sql(
            "select code from audit.reason_codes where code in "
            "('season_attributed_undated_injury', 'explicit_source_exclusion')"
        )
        if {clean_text(item.get("code")) for item in required_year2_reason_codes} != {
            "season_attributed_undated_injury", "explicit_source_exclusion",
        }:
            raise SystemExit(
                "process-intake requires Year 2 injury eligibility reason codes seeded by migration 20260822030000"
            )
    cohort_signal_team_key = resolve_team_key(args.team)
    own_team_alias = own_team_alias_for(cohort_signal_team_key, load_fixture_team_aliases())

    record_rows: list[dict[str, object]] = []
    event_rows: list[dict[str, object]] = []
    output_states: list[dict[str, object]] = []
    params = SqlParams()
    provenance = run_provenance()
    changed_rows = 0
    event_count = 0
    review_required_rows = 0
    reapplied_adjudication_rows = 0
    for row in rows:
        source_row_number = int(row["standardised_row_number"])
        row_source_adjudications = source_field_adjudications.get(source_row_number, [])
        source_correction_events = apply_source_field_adjudications(
            row, row_source_adjudications
        )
        state, events = build_processing_state(
            row,
            window_start=window_start,
            window_end=window_end,
            duplicate_signature_rows=duplicate_signature_rows,
            own_team_alias=own_team_alias,
            injury_eligibility_bridge=year2_eligibility_bridge.get(source_row_number),
        )
        events.extend(source_correction_events)
        events.extend(
            adjudicated_derived_change_events(
                prior_states.get(source_row_number),
                state,
                [
                    clean_text(adjudication["decision"].get("item_id"))
                    for adjudication in row_source_adjudications
                ],
            )
        )
        for exclusion in analysis_exclusions.get(source_row_number, []):
            reason = clean_text(exclusion.get("reason"))
            if not reason:
                raise SystemExit(f"analysis audit exclusion on row {source_row_number} has no reason")
            state["analysis_eligibility_status"] = "excluded_from_analysis"
            events.append(
                {
                    "field_name": "analysis_eligibility_status",
                    "old_value": "included_pending_protocol",
                    "new_value": "excluded_from_analysis",
                    "action": "exclude",
                    "reason_code": reason,
                    "rationale": f"Final analysis cohort exclusion: {reason}.",
                    "review_status": exclusion.get("review_status") or "pipeline_decision",
                }
            )
        adjudication = standing_adjudications.get(source_row_number)
        if adjudication is not None and source_row_number in analysis_exclusions:
            raise SystemExit(
                f"row {source_row_number} has both a CSV analysis-audit exclusion and a standing "
                "audit.adjudications decision; resolve the conflict manually before rerunning"
            )
        if apply_standing_adjudication(state, events, adjudication):
            reapplied_adjudication_rows += 1
        output_states.append(state)
        if state["analysis_eligibility_status"] == "review_required":
            review_required_rows += 1
        if events:
            changed_rows += 1
        raw_id = raw_record_id(args.team, args.season, registered_file_hash, source_row_number)
        record_rows.append(
            {
                "raw_record_id": raw_id,
                "record_state": state,
                "eligibility_status": state["analysis_eligibility_status"],
            }
        )
        for event in events:
            event_count += 1
            event_rows.append(
                {
                    "raw_record_id": raw_id,
                    "field_name": event["field_name"],
                    "old_value": event["old_value"],
                    "new_value": event["new_value"],
                    "action": event["action"],
                    "reason_code": event["reason_code"],
                    "rationale": event["rationale"],
                    "review_status": event["review_status"],
                }
            )

    output_hash = sha256_json(output_states)
    analysis_reason_codes = sorted(
        {
            clean_text(event.get("reason"))
            for events in analysis_exclusions.values()
            for event in events
            if clean_text(event.get("reason"))
        }
    )
    reason_code_sql = "".join(
        f"insert into audit.reason_codes (code, description) values ({params.text(reason)}, {params.text('Final team analysis cohort exclusion emitted by the versioned dashboard pipeline.')}) on conflict (code) do update set description = excluded.description;"
        for reason in analysis_reason_codes
    )
    if reapplied_adjudication_rows and not query_sql(
        "select 1 from audit.reason_codes where code = 'adjudication_reapplied'"
    ):
        raise SystemExit(
            "process-intake would reapply a standing audit.adjudications decision but reason code "
            "'adjudication_reapplied' is not seeded; run migration 20260710000137_adjudication_reapplication first"
        )
    sql = f"""
      insert into audit.reason_codes (code, description) values
        ('locator_enriched_intake', 'Intake row includes provisional source row locator and stable opaque UIDs.'),
        ('derived_return_date', 'Return date derived from Date Injured plus Days Injured; source value preserved.'),
        ('deterministic_derivation', 'Canonical value derived deterministically from preserved source fields.'),
        ('canonical_mapping', 'Canonical analysis field mapped from source field without overwriting the source value.'),
        ('controlled_inference', 'Canonical analysis field inferred from explicit high-confidence source evidence and marked with origin metadata.'),
        ('candidate_duplicate', 'Candidate duplicate flagged for review; source row retained.'),
        ('outside_provisional_window', 'Row falls outside provisional QC season window or has unparseable injury date; source row retained.')
        ,('source_field_adjudicated_correction', 'Human-approved correction overlaid on an immutable source field before deterministic re-derivation.')
      on conflict (code) do update set description = excluded.description;

      {reason_code_sql}

      do $$
      begin
        if (select count(*) from ingestion.source_rows sr join ingestion.source_files sf on sf.id = sr.source_file_id where sf.team = {params.text(args.team)} and sf.season = {params.text(args.season)} and sf.file_sha256 = {params.text(registered_file_hash)}) <> {len(rows)} then
          raise exception 'process-intake requires every source row to be registered';
        end if;
        if exists (select 1 from processing.record_versions rv join ingestion.source_rows sr on sr.id = rv.source_row_id join ingestion.source_files sf on sf.id = sr.source_file_id where sf.team = {params.text(args.team)} and sf.season = {params.text(args.season)} and sf.file_sha256 = {params.text(registered_file_hash)} and rv.version_number = {args.version_number}) then
          raise exception 'process-intake version already exists';
        end if;
      end $$;

      create temp table current_step on commit drop as
      with run as (
        insert into audit.pipeline_runs
          (command, team, season, status, input_hash, output_hash, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values (
          'process-intake', {params.text(args.team)}, {params.text(args.season)}, 'succeeded',
          {params.text(file_hash)}, {params.text(output_hash)},
          {params.jsonb({
            'file': path.name,
            'analysis_audit_file': str(analysis_audit_path) if analysis_audit_path else None,
            'analysis_audit_hash': sha256_file(analysis_audit_path) if analysis_audit_path else None,
            'step': args.step_name,
            'step_version': args.step_version,
            'window_start': args.window_start,
            'window_end': args.window_end,
            'processing_artifact_sha256': file_hash,
            'registered_source_file_sha256': registered_file_hash,
            'profile_manifest': str(Path(manifest_arg)) if manifest_arg else None,
            'profile_manifest_sha256': sha256_file(Path(manifest_arg)) if manifest_arg else None,
            'adapter_qc_file': str(adapter_qc_path) if adapter_qc_path else None,
            'adapter_qc_sha256': adapter_qc_sha256,
            'injury_eligibility_bridge_file': (
                str(Path(getattr(args, 'injury_eligibility_bridge_file', '')))
                if clean_text(getattr(args, 'injury_eligibility_bridge_file', '')) else None
            ),
            'injury_eligibility_bridge_sha256': (
                sha256_file(Path(getattr(args, 'injury_eligibility_bridge_file', '')))
                if clean_text(getattr(args, 'injury_eligibility_bridge_file', '')) else None
            ),
          })},
          now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      ),
      step as (
        insert into audit.step_runs
          (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count, counts_by_team, input_hash, output_hash, ended_at)
        select id, {params.text(args.step_name)}, {params.text(args.step_version)}, 'locator_enriched_intake',
          {len(rows)}, {len(rows)},
          {params.jsonb({
            args.team: {
              'rows': len(rows),
              'changed_or_flagged_rows': changed_rows,
              'review_required_rows': review_required_rows,
              'record_events': event_count,
              'duplicate_signature_rows': len(duplicate_signature_rows),
              'reapplied_adjudication_rows': reapplied_adjudication_rows,
              'source_field_adjudications': source_field_adjudication_count,
            }
          })},
          {params.text(file_hash)}, {params.text(output_hash)}, now()
        from run
        returning id
      )
      select id from step;

      update audit.adjudications adjudication
      set consumed_by_step_run_id = step.id
      from ingestion.source_rows sr
      join ingestion.source_files sf on sf.id = sr.source_file_id
      cross join current_step step
      where adjudication.source_row_id = sr.id
        and sf.team = {params.text(args.team)}
        and sf.season = {params.text(args.season)}
        and sf.file_sha256 = {params.text(registered_file_hash)}
        and adjudication.field_name in ('Date Injured', 'Fit For Selection Date')
        and adjudication.decision ->> 'decision_type' = 'source_field_correction'
        and adjudication.consumed_by_step_run_id is null;

      insert into processing.record_versions
        (source_row_id, step_run_id, version_number, record_state, eligibility_status)
      select
        sr.id,
        step.id,
        {args.version_number},
        planned_record.item -> 'record_state',
        planned_record.item ->> 'eligibility_status'
      from jsonb_array_elements({params.jsonb(record_rows)}) with ordinality
        as planned_record(item, ordinal)
      join ingestion.source_rows sr
        on sr.raw_record_id = planned_record.item ->> 'raw_record_id'
      cross join current_step step
      order by planned_record.ordinal;

      insert into audit.record_events
        (step_run_id, source_row_id, field_name, old_value, new_value, action, reason_code, rationale, rule_version, review_status)
      select
        step.id,
        sr.id,
        planned_event.item ->> 'field_name',
        planned_event.item -> 'old_value',
        planned_event.item -> 'new_value',
        planned_event.item ->> 'action',
        planned_event.item ->> 'reason_code',
        planned_event.item ->> 'rationale',
        {params.text(args.step_version)},
        planned_event.item ->> 'review_status'
      from jsonb_array_elements({params.jsonb(event_rows)}) with ordinality
        as planned_event(item, ordinal)
      join ingestion.source_rows sr
        on sr.raw_record_id = planned_event.item ->> 'raw_record_id'
      cross join current_step step
      order by planned_event.ordinal;
    """
    run_sql(sql, params.values)
    print(
        json.dumps(
            {
                "processed": str(path),
                "rows": len(rows),
                "changed_or_flagged_rows": changed_rows,
                "review_required_rows": review_required_rows,
                "record_events": event_count,
                "duplicate_signature_rows": len(duplicate_signature_rows),
                "reapplied_adjudication_rows": reapplied_adjudication_rows,
                "source_field_adjudications": source_field_adjudication_count,
                "registered_source_file_sha256": registered_file_hash,
            },
            indent=2,
        )
    )


def trace_row(args: argparse.Namespace) -> None:
    intake_rows = read_rows(Path(args.file))
    row = next(
        (item for item in intake_rows if item.get("standardised_row_number") == str(args.row_number)),
        None,
    )
    if row is None:
        raise SystemExit(f"standardised row {args.row_number} not found")

    result = {"standardised_row": row}
    if args.include_source:
        source_rows = read_rows(Path(row["source_archive_path"]))
        source_index = int(row["source_row_number"]) - 2
        if source_index < 0 or source_index >= len(source_rows):
            raise SystemExit(f"source row {row['source_row_number']} not found")
        result["source_row"] = source_rows[source_index]
    print(json.dumps(result, indent=2))


def validate_v13_reviewed_package_binding(
    manifest: dict[str, Any],
    manifest_path: Path,
    profile: dict[str, Any],
    profile_document: dict[str, Any],
    input_sha256: str,
    team: str,
    input_path: Path | None,
) -> None:
    team_key = clean_text(manifest.get("team_key"))
    reviewed_team = V13_REVIEWED_V12_TEAMS.get(team_key)
    if reviewed_team is None:
        raise SystemExit("V13 reviewed V12 package binding is missing or inconsistent")
    expected_team, injury_sha256, exposure_sha256, team_manifest_sha256 = reviewed_team
    expected_inputs = {
        "injury": {
            "path": "injury_intake_locator_enriched_v10.csv",
            "sha256": injury_sha256,
        },
        "exposure": {
            "path": "exposure_intake_final_clean_v10.csv",
            "sha256": exposure_sha256,
        },
    }
    expected_team_manifest = {
        "path": "intake_manifest_v12.json",
        "sha256": team_manifest_sha256,
    }
    expected_review = {
        "path": "provenance/v12_fresh_ai_review_evidence.json",
        "sha256": V13_FRESH_REVIEW_SHA256,
        "reviewer": V13_FRESH_REVIEWER,
        "decision": "COMPLETED_WITH_RECORDED_LIMITATIONS",
        "v12_root_manifest_sha256": V13_V12_ROOT_MANIFEST_SHA256,
        "v12_root_file_set_sha256": V13_V12_ROOT_FILE_SET_SHA256,
    }
    expected_ai_reviewer = (
        "gpt-5.6-sol/xhigh /root/v13_signer_acceptance"
    )
    if (
        team != expected_team
        or manifest.get("team") != expected_team
        or profile_document.get("team") != expected_team
        or profile_document.get("team_key") != team_key
        or manifest.get("season") != "2025-26"
        or profile_document.get("season") != "2025-26"
        or profile.get("decision") != "adapter_required"
        or profile_document.get("decision") != "adapter_required"
        or profile.get("ai_reviewed_by") != expected_ai_reviewer
        or profile_document.get("ai_reviewed_by") != expected_ai_reviewer
        or profile.get("ai_reviewed_at") != V13_FRESH_REVIEWER["completed_at"]
        or profile_document.get("ai_reviewed_at") != V13_FRESH_REVIEWER["completed_at"]
        or profile.get("approved_input_sha256s") != [injury_sha256, exposure_sha256]
        or profile_document.get("approved_input_sha256s") != [injury_sha256, exposure_sha256]
        or manifest.get("v12_input_bindings") != expected_inputs
        or profile_document.get("v12_input_bindings") != expected_inputs
        or manifest.get("source_v12_manifest") != expected_team_manifest
        or profile_document.get("v12_manifest") != expected_team_manifest
        or manifest.get("fresh_ai_review_evidence") != expected_review
        or profile_document.get("fresh_ai_review_evidence") != expected_review
        or input_sha256 not in {injury_sha256, exposure_sha256}
    ):
        raise SystemExit("V13 reviewed V12 package binding is missing or inconsistent")

    current_binding = next(
        binding for binding in expected_inputs.values()
        if binding["sha256"] == input_sha256
    )
    if input_path is not None and input_path.name != current_binding["path"]:
        raise SystemExit("V13 current input does not match its reviewed V12 binding")

    harness_provenance = manifest.get("harness_provenance")
    if (
        not isinstance(harness_provenance, dict)
        or profile_document.get("harness_provenance") != harness_provenance
        or set(harness_provenance) != {
            "script", "config", "fresh_ai_review_evidence"
        }
    ):
        raise SystemExit("V13 package provenance is missing or inconsistent")
    expected_provenance_paths = {
        "script": "provenance/v13_signing_harness.py",
        "config": "provenance/v13_signing_harness_config.json",
        "fresh_ai_review_evidence": (
            "provenance/v12_fresh_ai_review_evidence.json"
        ),
    }
    package_root = manifest_path.parent.parent
    for key, expected_path in expected_provenance_paths.items():
        binding = harness_provenance.get(key)
        if (
            not isinstance(binding, dict)
            or set(binding) != {"path", "sha256"}
            or binding.get("path") != expected_path
            or not isinstance(binding.get("sha256"), str)
            or not re.fullmatch(r"[0-9a-f]{64}", binding["sha256"])
            or (
                key == "fresh_ai_review_evidence"
                and binding["sha256"] != V13_FRESH_REVIEW_SHA256
            )
        ):
            raise SystemExit("V13 package provenance is missing or inconsistent")
        evidence_path = package_root / expected_path
        if not evidence_path.is_file() or sha256_file(evidence_path) != binding["sha256"]:
            raise SystemExit("V13 package provenance checksum mismatch")

    source_manifest_path = manifest_path.parent / expected_team_manifest["path"]
    if (
        manifest_path.parent.name != team_key
        or not source_manifest_path.is_file()
        or sha256_file(source_manifest_path) != team_manifest_sha256
    ):
        raise SystemExit("V13 source V12 team-manifest checksum mismatch")


def validate_intake_profile_manifest(
    manifest: dict[str, Any],
    manifest_path: Path,
    input_sha256: str,
    team: str,
    season: str,
    input_path: Path | None = None,
) -> None:
    if not isinstance(manifest, dict):
        raise SystemExit("ingest manifest must be a JSON object")
    profile = manifest.get("intake_profile")
    if not isinstance(profile, dict):
        raise SystemExit("ingest manifest requires an intake_profile object")

    def required_text(field: str) -> str:
        value = profile.get(field)
        if not isinstance(value, str) or not value.strip():
            raise SystemExit(f"intake_profile.{field} is required")
        return value.strip()

    decision = required_text("decision")
    if required_text("team") != team or required_text("season") != season:
        raise SystemExit("intake profile team/season does not match ingest target")
    if decision not in {"compatible", "adapter_required"}:
        raise SystemExit(f"intake profile decision blocks ingest: {decision}")
    if required_text("ai_review_status") != "completed":
        raise SystemExit("intake_profile.ai_review_status must be completed")
    required_text("ai_reviewed_by")
    required_text("profile_version")
    if required_text("approved_by") != "Abdel Babiker":
        raise SystemExit("intake profile must be approved by Abdel Babiker")
    if profile.get("unresolved_adjudication_ids") != []:
        raise SystemExit("intake profile has unresolved adjudications")

    reviewed_at = required_text("ai_reviewed_at")
    approved_at = required_text("approved_at")
    try:
        reviewed_time = datetime.fromisoformat(reviewed_at.replace("Z", "+00:00"))
        approved_time = datetime.fromisoformat(approved_at.replace("Z", "+00:00"))
        if reviewed_time.tzinfo is None or approved_time.tzinfo is None:
            raise ValueError
        if approved_time < reviewed_time:
            raise ValueError
        latest_allowed = datetime.now(UTC) + timedelta(minutes=5)
        if reviewed_time.astimezone(UTC) > latest_allowed or approved_time.astimezone(UTC) > latest_allowed:
            raise ValueError
    except ValueError as exc:
        raise SystemExit(
            "intake profile review/approval timestamps must be timezone-aware ISO values "
            "with approval at or after review and neither value in the future"
        ) from exc

    approved_inputs = profile.get("approved_input_sha256s")
    if not isinstance(approved_inputs, list) or input_sha256 not in approved_inputs:
        raise SystemExit("current intake checksum is not covered by profile approval")

    def verify_file(path_field: str, sha_field: str) -> Path:
        evidence_path = Path(required_text(path_field))
        if not evidence_path.is_absolute():
            evidence_path = manifest_path.parent / evidence_path
        expected_sha = required_text(sha_field)
        if not evidence_path.is_file():
            raise SystemExit(f"intake profile evidence file not found: {evidence_path}")
        if sha256_file(evidence_path) != expected_sha:
            raise SystemExit(f"intake profile evidence checksum mismatch: {evidence_path}")
        return evidence_path

    profile_path = verify_file("profile_path", "profile_sha256")
    try:
        profile_document = json.loads(profile_path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit("intake profile evidence must be valid JSON") from exc
    v13_authorisation_required = (
        season == "2025-26"
        or profile.get("profile_version") == "urc_2025_26_v13_signed_profile_v1"
        or manifest.get("schema") == V13_INTAKE_MANIFEST_SCHEMA
        or (
            isinstance(profile_document, dict)
            and profile_document.get("schema") == V13_INTAKE_PROFILE_SCHEMA
        )
    )
    if v13_authorisation_required:
        if (
            not isinstance(profile_document, dict)
            or manifest.get("schema") != V13_INTAKE_MANIFEST_SCHEMA
            or profile_document.get("schema") != V13_INTAKE_PROFILE_SCHEMA
            or profile.get("database_action_authorised") is not True
            or profile_document.get("database_action_authorised") is not True
            or manifest.get("database_action_authorised") is not True
            or profile.get("authorisation") != V13_DATABASE_AUTHORISATION
            or profile_document.get("authorisation") != V13_DATABASE_AUTHORISATION
            or manifest.get("authorisation") != V13_DATABASE_AUTHORISATION
            or profile.get("approval_line_sha256")
            != V13_DATABASE_AUTHORISATION["approval_line_sha256"]
            or profile_document.get("approval_line_sha256")
            != V13_DATABASE_AUTHORISATION["approval_line_sha256"]
            or manifest.get("approval_line_sha256")
            != V13_DATABASE_AUTHORISATION["approval_line_sha256"]
        ):
            raise SystemExit(
                "V13 database action authorisation is missing, false, inconsistent, "
                "or outside the approved target/action scope"
            )
        validate_v13_reviewed_package_binding(
            manifest,
            manifest_path,
            profile,
            profile_document,
            input_sha256,
            team,
            input_path,
        )
    bound_fields = (
        "team", "season", "profile_version", "decision", "mapping_path", "mapping_sha256",
        "mapping_version", "ai_review_status", "ai_reviewed_by", "ai_reviewed_at", "approved_by",
        "approved_at", "approval_line_sha256", "unresolved_adjudication_ids",
        "approved_input_sha256s",
    )
    if not isinstance(profile_document, dict) or any(
        profile_document.get(field) != profile.get(field) for field in bound_fields
    ):
        raise SystemExit("intake profile approval fields do not match checksummed profile JSON")
    mapping_values = [
        profile.get(field) for field in ("mapping_path", "mapping_sha256", "mapping_version")
    ]
    if any(value is not None for value in mapping_values):
        if not all(isinstance(value, str) and value.strip() for value in mapping_values):
            raise SystemExit("intake profile mapping fields must all be set or all be null")
        mapping_path = verify_file("mapping_path", "mapping_sha256")
        try:
            mapping_document = json.loads(mapping_path.read_text())
        except (OSError, json.JSONDecodeError) as exc:
            raise SystemExit("intake mapping evidence must be valid JSON") from exc
        mappings = mapping_document.get("mappings") if isinstance(mapping_document, dict) else None
        if (
            not isinstance(mapping_document, dict)
            or mapping_document.get("mapping_version") != profile["mapping_version"]
            or not isinstance(mappings, list)
            or not mappings
        ):
            raise SystemExit(
                "intake mapping must match mapping_version and contain non-empty mapping objects"
            )
        evidence_classes = {
            "source_reported", "deterministic_derivation", "protocol_defined_inference",
            "manual_adjudication",
        }
        for entry in mappings:
            source_evidence = entry.get("source_evidence") if isinstance(entry, dict) else None
            if (
                not isinstance(entry, dict)
                or not all(
                    isinstance(entry.get(field), str) and entry[field].strip()
                    for field in ("canonical_field", "canonical_value")
                )
                or entry.get("evidence_class") not in evidence_classes
                or not isinstance(source_evidence, dict)
                or not source_evidence
                or not all(
                    isinstance(key, str) and key.strip()
                    and isinstance(value, str) and value.strip()
                    for key, value in source_evidence.items()
                )
            ):
                raise SystemExit(
                    "each intake mapping requires canonical_field, canonical_value, "
                    "source_evidence, and a controlled evidence_class"
                )
    elif decision == "adapter_required":
        raise SystemExit("adapter_required intake profile requires a versioned mapping file")


def accepted_repository_source_sha256s(
    acceptance_commit: str, relative_paths: set[str]
) -> dict[str, str]:
    repository = Path(__file__).resolve().parent.parent
    result: dict[str, str] = {}
    for relative_path in sorted(relative_paths):
        try:
            content = subprocess.run(
                ["git", "show", f"{acceptance_commit}:{relative_path}"],
                cwd=repository,
                check=True,
                capture_output=True,
            ).stdout
        except (OSError, subprocess.CalledProcessError) as exc:
            raise SystemExit("V13 accepted repository validator source is unavailable") from exc
        result[relative_path] = hashlib.sha256(content).hexdigest()
    return result


def current_repository_source_sha256s(relative_paths: set[str]) -> dict[str, str]:
    repository = Path(__file__).resolve().parent.parent
    return {
        relative_path: sha256_file(repository / relative_path)
        for relative_path in sorted(relative_paths)
    }


def repository_file_matches_head(path: Path) -> bool:
    repository = Path(__file__).resolve().parent.parent
    try:
        relative_path = path.resolve().relative_to(repository).as_posix()
        committed = subprocess.run(
            ["git", "show", f"HEAD:{relative_path}"],
            cwd=repository,
            check=True,
            capture_output=True,
        ).stdout
    except (OSError, subprocess.CalledProcessError, ValueError):
        return False
    return path.is_file() and path.read_bytes() == committed


def validate_v13_signed_root_candidate(
    signed_root_manifest_path: Path,
    manifest_path: Path,
    input_path: Path,
    input_sha256: str,
    team: str,
    season: str,
) -> str:
    """Validate a complete signed candidate without consulting the release ledger.

    This is the protected signer's post-finalisation check. It has no CLI flag or
    production-ingest route; ingest calls the allow-listed wrapper below.
    """
    if season != "2025-26":
        raise SystemExit("V13 signed-root validation is only defined for 2025-26")
    supplied_root_path = signed_root_manifest_path.absolute()
    if supplied_root_path.is_symlink():
        raise SystemExit("signed V13 root manifest must not be a symlink")
    root_path = supplied_root_path.resolve()
    if root_path.name != "v13_signed_root_manifest.json" or not root_path.is_file():
        raise SystemExit("2025-26 ingest requires the physical signed V13 root manifest")
    try:
        root = json.loads(root_path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit("signed V13 root manifest must be valid JSON") from exc
    if not isinstance(root, dict):
        raise SystemExit("signed V13 root manifest must be a JSON object")
    package_root = root_path.parent
    if (
        not stat.S_ISDIR(package_root.lstat().st_mode)
        or stat.S_IMODE(package_root.lstat().st_mode) != 0o700
    ):
        raise SystemExit("signed V13 package root directory must have exact mode 0700")
    actual_outputs: dict[str, str] = {}
    for path in sorted(package_root.rglob("*")):
        relative_path = path.relative_to(package_root)
        if "__pycache__" in relative_path.parts:
            raise SystemExit("signed V13 package must not contain __pycache__")
        path_mode = path.lstat().st_mode
        if stat.S_ISLNK(path_mode):
            raise SystemExit("signed V13 package must not contain symlinks")
        if stat.S_ISDIR(path_mode):
            if stat.S_IMODE(path_mode) != 0o700:
                raise SystemExit("signed V13 package directories must have exact mode 0700")
            continue
        if not stat.S_ISREG(path_mode):
            raise SystemExit("signed V13 package must contain only regular files")
        if stat.S_IMODE(path_mode) != 0o600:
            raise SystemExit("signed V13 package files must have exact mode 0600")
        if path.resolve() != root_path:
            actual_outputs[relative_path.as_posix()] = sha256_file(path)
    output_sha256s = root.get("output_sha256s")
    if output_sha256s != actual_outputs:
        raise SystemExit("signed V13 root output map does not close the physical package")
    actual_file_set_sha256 = hashlib.sha256(
        json.dumps(actual_outputs, sort_keys=True, separators=(",", ":")).encode()
    ).hexdigest()
    if root.get("root_file_set_sha256") != actual_file_set_sha256:
        raise SystemExit("signed V13 root file-set checksum mismatch")

    expected_predecessor = {
        "package": "all_16_intake_envelopes_20260822_v12_duplicate_safe_candidate",
        "root_manifest_sha256": V13_V12_ROOT_MANIFEST_SHA256,
        "root_file_set_sha256": V13_V12_ROOT_FILE_SET_SHA256,
    }
    if (
        root.get("schema") != "urc_2025_26_v13_signed_root_manifest_v1"
        or root.get("season") != "2025-26"
        or root.get("predecessor") != expected_predecessor
        or root.get("approved_by") != "Abdel Babiker"
        or root.get("approval_ready") is not True
        or root.get("ingest_ready") is not True
        or root.get("database_action_authorised") is not True
        or root.get("authorisation") != V13_DATABASE_AUTHORISATION
        or root.get("approval_line_sha256")
        != V13_DATABASE_AUTHORISATION["approval_line_sha256"]
    ):
        raise SystemExit("signed V13 root approval or predecessor binding is invalid")
    predecessor_outputs = root.get("predecessor_output_sha256s")
    if (
        root.get("predecessor_output_count") != 196
        or not isinstance(predecessor_outputs, dict)
        or len(predecessor_outputs) != 196
        or any(
            not isinstance(path, str)
            or not path
            or "\\" in path
            or Path(path).is_absolute()
            or any(part in {"", ".", ".."} for part in path.split("/"))
            or not isinstance(digest, str)
            or not re.fullmatch(r"[0-9a-f]{64}", digest)
            for path, digest in predecessor_outputs.items()
        )
    ):
        raise SystemExit("signed V13 root predecessor output map must contain exactly 196 safe members")
    predecessor_map_sha256 = hashlib.sha256(
        json.dumps(
            predecessor_outputs, sort_keys=True, separators=(",", ":")
        ).encode()
    ).hexdigest()
    if (
        predecessor_map_sha256 != V13_V12_PREDECESSOR_OUTPUT_MAP_SHA256
        or root.get("predecessor_output_file_set_sha256")
        != V13_V12_PREDECESSOR_OUTPUT_MAP_SHA256
        or any(
            output_sha256s.get(path) != digest
            for path, digest in predecessor_outputs.items()
        )
    ):
        raise SystemExit("signed V13 root predecessor output map differs from exact V12 membership")

    validator_results = root.get("validator_results")
    if not isinstance(validator_results, list) or len(validator_results) != 16:
        raise SystemExit("signed V13 root requires 16 validator results")
    results_by_team: dict[str, dict[str, Any]] = {}
    for result in validator_results:
        if not isinstance(result, dict):
            raise SystemExit("signed V13 root validator result is invalid")
        result_team_key = clean_text(result.get("team_key"))
        if result_team_key in results_by_team or result_team_key not in V13_REVIEWED_V12_TEAMS:
            raise SystemExit("signed V13 root validator team membership is invalid")
        expected_profile_path = f"{result_team_key}/v13_approved_intake_profile.json"
        expected_manifest_path = f"{result_team_key}/v13_approved_intake_manifest.json"
        if (
            result.get("status") != "pass"
            or result.get("validated_inputs") != ["injury", "exposure"]
            or result.get("profile_sha256") != output_sha256s.get(expected_profile_path)
            or result.get("manifest_sha256") != output_sha256s.get(expected_manifest_path)
        ):
            raise SystemExit("signed V13 root validator result does not bind team outputs")
        results_by_team[result_team_key] = result
    if set(results_by_team) != set(V13_REVIEWED_V12_TEAMS):
        raise SystemExit("signed V13 root does not cover all 16 reviewed teams")
    for team_key, (_, injury_sha, exposure_sha, team_manifest_sha) in (
        V13_REVIEWED_V12_TEAMS.items()
    ):
        expected_members = {
            f"{team_key}/injury_intake_locator_enriched_v10.csv": injury_sha,
            f"{team_key}/exposure_intake_final_clean_v10.csv": exposure_sha,
            f"{team_key}/intake_manifest_v12.json": team_manifest_sha,
        }
        if any(output_sha256s.get(path) != digest for path, digest in expected_members.items()):
            raise SystemExit("signed V13 root V12 membership differs from the reviewed root")

    expected_candidate_results: dict[tuple[str, str], dict[str, str]] = {}
    for team_key, (_, injury_sha, exposure_sha, _) in V13_REVIEWED_V12_TEAMS.items():
        for input_kind, filename, digest in (
            ("injury", "injury_intake_locator_enriched_v10.csv", injury_sha),
            ("exposure", "exposure_intake_final_clean_v10.csv", exposure_sha),
        ):
            expected_candidate_results[(team_key, input_kind)] = {
                "team_key": team_key,
                "input_kind": input_kind,
                "path": f"{team_key}/{filename}",
                "sha256": digest,
                "status": "pass",
            }
    candidate_results = root.get("root_candidate_validator_results")
    if not isinstance(candidate_results, list) or len(candidate_results) != 32:
        raise SystemExit("signed V13 root requires 32 final candidate-validator results")
    actual_candidate_results: dict[tuple[str, str], dict[str, Any]] = {}
    for result in candidate_results:
        if not isinstance(result, dict) or set(result) != {
            "team_key", "input_kind", "path", "sha256", "status"
        }:
            raise SystemExit("signed V13 root candidate-validator result is invalid")
        key = (clean_text(result.get("team_key")), clean_text(result.get("input_kind")))
        if key in actual_candidate_results or result != expected_candidate_results.get(key):
            raise SystemExit("signed V13 root candidate-validator results do not bind all inputs")
        if output_sha256s.get(result["path"]) != result["sha256"]:
            raise SystemExit("signed V13 root candidate-validator input is not a package member")
        actual_candidate_results[key] = result
    if actual_candidate_results != expected_candidate_results:
        raise SystemExit("signed V13 root candidate-validator results do not bind all inputs")

    signing_binding = root.get("signing_record")
    if signing_binding != {
        "path": "v13_signing_record.json",
        "sha256": output_sha256s.get("v13_signing_record.json"),
    }:
        raise SystemExit("signed V13 root signing-record binding is invalid")
    signing_path = package_root / "v13_signing_record.json"
    try:
        signing = json.loads(signing_path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit("V13 signing record must be valid JSON") from exc
    if not isinstance(signing, dict):
        raise SystemExit("V13 signing record must be a JSON object")
    common_signing_fields = (
        signing.get("schema") == "urc_2025_26_v13_signing_record_v1"
        and signing.get("approved_by") == root.get("approved_by")
        and signing.get("approved_at") == root.get("approved_at")
        and signing.get("approval_line_sha256") == root.get("approval_line_sha256")
        and signing.get("v12_root_manifest_sha256") == V13_V12_ROOT_MANIFEST_SHA256
        and signing.get("v12_root_file_set_sha256") == V13_V12_ROOT_FILE_SET_SHA256
        and signing.get("predecessor_output_sha256s") == predecessor_outputs
        and signing.get("predecessor_output_count") == 196
        and signing.get("predecessor_output_file_set_sha256")
        == V13_V12_PREDECESSOR_OUTPUT_MAP_SHA256
        and signing.get("database_action_authorised") is True
        and signing.get("authorisation") == root.get("authorisation")
        and signing.get("fresh_ai_review_evidence")
        == root.get("fresh_ai_review_evidence")
        and signing.get("harness_provenance") == root.get("harness_provenance")
    )
    if not common_signing_fields:
        raise SystemExit("V13 signing record differs from the signed root")
    if signing.get("candidate_preservation") != {
        "v12_non_root_file_count": 196,
        "all_v12_non_root_bytes_preserved": True,
        "physical_v12_root_manifest_copied": False,
        "coverage_limitations_preserved": True,
    }:
        raise SystemExit("V13 signing record predecessor preservation is invalid")

    privacy_binding = root.get("privacy_evidence")
    privacy_relative = "privacy_scan_v13.json"
    if privacy_binding != {
        "path": privacy_relative,
        "sha256": output_sha256s.get(privacy_relative),
        "status": "pass",
    }:
        raise SystemExit("signed V13 root privacy-evidence binding is invalid")
    privacy_path = package_root / privacy_relative
    try:
        privacy = json.loads(privacy_path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit("V13 privacy evidence must be valid JSON") from exc
    covered_outputs = {
        path: digest
        for path, digest in output_sha256s.items()
        if path != privacy_relative
    }
    covered_file_set_sha256 = hashlib.sha256(
        json.dumps(covered_outputs, sort_keys=True, separators=(",", ":")).encode()
    ).hexdigest()
    if (
        not isinstance(privacy, dict)
        or privacy.get("schema")
        != "urc_2025_26_v13_signing_privacy_evidence_v1"
        or privacy.get("status") != "pass"
        or privacy.get("direct_identifier_match_count") != 0
        or privacy.get("forbidden_key_match_count") != 0
        or privacy.get("scanned_file_count") != len(covered_outputs)
        or privacy.get("covered_output_sha256s") != covered_outputs
        or privacy.get("covered_file_set_sha256") != covered_file_set_sha256
        or privacy.get("excluded_paths")
        != [privacy_relative, "v13_signed_root_manifest.json"]
        or privacy.get("final_closed_regular_file_count")
        != len(output_sha256s) + 1
    ):
        raise SystemExit("V13 privacy evidence does not cover the complete closed package")

    expected_review = {
        "path": "provenance/v12_fresh_ai_review_evidence.json",
        "sha256": V13_FRESH_REVIEW_SHA256,
        "reviewer": V13_FRESH_REVIEWER,
        "decision": "COMPLETED_WITH_RECORDED_LIMITATIONS",
        "v12_root_manifest_sha256": V13_V12_ROOT_MANIFEST_SHA256,
        "v12_root_file_set_sha256": V13_V12_ROOT_FILE_SET_SHA256,
    }
    if root.get("fresh_ai_review_evidence") != expected_review:
        raise SystemExit("signed V13 root fresh-review binding is invalid")
    harness_provenance = root.get("harness_provenance")
    expected_provenance_paths = {
        "script": "provenance/v13_signing_harness.py",
        "config": "provenance/v13_signing_harness_config.json",
        "fresh_ai_review_evidence": "provenance/v12_fresh_ai_review_evidence.json",
    }
    if not isinstance(harness_provenance, dict) or set(harness_provenance) != set(
        expected_provenance_paths
    ):
        raise SystemExit("signed V13 root package provenance is invalid")
    for key, relative_path in expected_provenance_paths.items():
        binding = harness_provenance.get(key)
        if (
            not isinstance(binding, dict)
            or binding.get("path") != relative_path
            or binding.get("sha256") != output_sha256s.get(relative_path)
            or (
                key == "fresh_ai_review_evidence"
                and binding.get("sha256") != V13_FRESH_REVIEW_SHA256
            )
        ):
            raise SystemExit("signed V13 root package provenance is invalid")

    config_path = package_root / expected_provenance_paths["config"]
    try:
        config = json.loads(config_path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit("signed V13 harness config must be valid JSON") from exc
    repository_validator = signing.get("repository_validator")
    source_sha256s = (
        repository_validator.get("source_sha256s")
        if isinstance(repository_validator, dict)
        else None
    )
    acceptance_commit = (
        clean_text(repository_validator.get("acceptance_commit"))
        if isinstance(repository_validator, dict)
        else ""
    )
    expected_source_paths = {
        "pipeline/__init__.py", "pipeline/__main__.py", "pipeline/season_contracts.py"
    }
    if (
        not isinstance(config, dict)
        or config.get("schema") != "urc_2025_26_v13_signing_harness_config_v3"
        or config.get("source_package") != expected_predecessor["package"]
        or config.get("source_root_manifest") != "v12_duplicate_safe_root_manifest.json"
        or config.get("successor_package") != package_root.name
        or config.get("successor_root_manifest") != root_path.name
        or config.get("expected_v12_root_manifest_sha256")
        != V13_V12_ROOT_MANIFEST_SHA256
        or config.get("expected_v12_root_file_set_sha256")
        != V13_V12_ROOT_FILE_SET_SHA256
        or config.get("expected_v12_non_root_file_count") != 196
        or config.get("required_team_count") != 16
        or config.get("required_approver") != "Abdel Babiker"
        or config.get("approval_bearing_files")
        != {
            "profile": "v13_approved_intake_profile.json",
            "manifest": "v13_approved_intake_manifest.json",
        }
        or Path(clean_text(config.get("repository_root"))).resolve()
        != Path(__file__).resolve().parent.parent
        or hashlib.sha256(clean_text(config.get("required_approval_line")).encode()).hexdigest()
        != root.get("approval_line_sha256")
        or config.get("fresh_ai_review")
        != {
            "filename": "v12_fresh_ai_review_evidence.json",
            "sha256": V13_FRESH_REVIEW_SHA256,
            **V13_FRESH_REVIEWER,
        }
        or config.get("database_authorisation")
        != {
            "project_ref": "eukkvswaxweenovqqgzr",
            "database": "postgres",
            "actions": ["ingestion", "processing", "build", "release"],
        }
        or not re.fullmatch(r"[0-9a-f]{40}", acceptance_commit)
        or config.get("repository_acceptance_commit") != acceptance_commit
        or not isinstance(source_sha256s, dict)
        or set(source_sha256s) != expected_source_paths
        or config.get("repository_validator_main_sha256")
        != source_sha256s.get("pipeline/__main__.py")
        or repository_validator.get("entry_point")
        != "pipeline.__main__.validate_intake_profile_manifest"
        or repository_validator.get("passed_team_count") != 16
        or repository_validator.get("validated_input_count") != 32
        or repository_validator.get("status") != "pass"
    ):
        raise SystemExit("signed V13 config or repository-validator binding is invalid")
    accepted_source_sha256s = accepted_repository_source_sha256s(
        acceptance_commit, expected_source_paths
    )
    if (
        accepted_source_sha256s != source_sha256s
        or current_repository_source_sha256s(expected_source_paths) != source_sha256s
    ):
        raise SystemExit("signed V13 repository-validator checksums do not match the accepted commit")

    manifest_resolved = manifest_path.resolve()
    input_resolved = input_path.resolve()
    if (
        manifest_resolved.parent.parent != package_root
        or input_resolved.parent != manifest_resolved.parent
    ):
        raise SystemExit("current V13 team inputs must be inside the signed package")
    try:
        current_manifest = json.loads(manifest_resolved.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit("current V13 team manifest must be valid JSON") from exc
    if not isinstance(current_manifest, dict):
        raise SystemExit("current V13 team manifest must be a JSON object")
    team_key = clean_text(current_manifest.get("team_key"))
    envelope = current_manifest.get("intake_profile")
    if not isinstance(envelope, dict):
        raise SystemExit("current V13 team manifest lacks its profile envelope")
    profile_relative = f"{team_key}/{clean_text(envelope.get('profile_path'))}"
    manifest_relative = manifest_resolved.relative_to(package_root).as_posix()
    input_relative = input_resolved.relative_to(package_root).as_posix()
    if (
        manifest_relative != f"{team_key}/v13_approved_intake_manifest.json"
        or output_sha256s.get(manifest_relative) != sha256_file(manifest_resolved)
        or output_sha256s.get(profile_relative) != envelope.get("profile_sha256")
        or output_sha256s.get(input_relative) != input_sha256
    ):
        raise SystemExit("current team is not an exact member of the signed V13 root")
    profile_path = package_root / profile_relative
    try:
        profile_document = json.loads(profile_path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit("current V13 profile must be valid JSON") from exc
    if not isinstance(profile_document, dict):
        raise SystemExit("current V13 profile must be a JSON object")
    approved_at = clean_text(root.get("approved_at"))
    if (
        current_manifest.get("approved_by") != root.get("approved_by")
        or current_manifest.get("approved_at") != approved_at
        or current_manifest.get("approval_line_sha256")
        != root.get("approval_line_sha256")
        or profile_document.get("approved_by") != root.get("approved_by")
        or profile_document.get("approved_at") != approved_at
        or profile_document.get("approval_line_sha256")
        != root.get("approval_line_sha256")
        or envelope.get("approved_by") != root.get("approved_by")
        or envelope.get("approved_at") != approved_at
        or envelope.get("approval_line_sha256")
        != root.get("approval_line_sha256")
        or any(
            document.get("database_action_authorised") is not True
            or document.get("authorisation") != root.get("authorisation")
            for document in (current_manifest, profile_document, envelope)
        )
        or any(
            document.get("approval_ready") is not True
            or document.get("ingest_ready") is not True
            for document in (current_manifest, profile_document, envelope)
        )
        or current_manifest.get("fresh_ai_review_evidence")
        != root.get("fresh_ai_review_evidence")
        or profile_document.get("fresh_ai_review_evidence")
        != root.get("fresh_ai_review_evidence")
        or current_manifest.get("harness_provenance") != root.get("harness_provenance")
        or profile_document.get("harness_provenance") != root.get("harness_provenance")
    ):
        raise SystemExit("V13 root, signing record and team approval fields differ")
    validate_intake_profile_manifest(
        current_manifest,
        manifest_resolved,
        input_sha256,
        team,
        season,
        input_path=input_resolved,
    )
    return sha256_file(root_path)


def validate_v13_signed_root_for_ingest(
    signed_root_manifest_path: Path,
    manifest_path: Path,
    input_path: Path,
    input_sha256: str,
    team: str,
    season: str,
) -> str:
    root_sha256 = validate_v13_signed_root_candidate(
        signed_root_manifest_path,
        manifest_path,
        input_path,
        input_sha256,
        team,
        season,
    )
    if not repository_file_matches_head(YEAR2_APPROVED_ROOTS_PATH):
        raise SystemExit("approved Year2 root ledger must match its tracked HEAD bytes")
    try:
        ledger = json.loads(YEAR2_APPROVED_ROOTS_PATH.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit("approved Year2 root ledger is unavailable or invalid") from exc
    approved_roots = ledger.get("approved_root_sha256s") if isinstance(ledger, dict) else None
    if (
        not isinstance(ledger, dict)
        or ledger.get("schema") != "urc_2025_26_approved_roots_v1"
        or not isinstance(approved_roots, list)
        or any(
            not isinstance(value, str) or not re.fullmatch(r"[0-9a-f]{64}", value)
            for value in approved_roots
        )
        or len(set(approved_roots)) != len(approved_roots)
    ):
        raise SystemExit("approved Year2 root ledger is unavailable or invalid")
    if root_sha256 not in approved_roots:
        raise SystemExit("signed V13 root is not present in the approved Year2 root ledger")
    return root_sha256


def ingest(args: argparse.Namespace) -> None:
    path = Path(args.file)
    file_hash = sha256_file(path)
    manifest_path = Path(args.manifest)
    source_manifest = json.loads(manifest_path.read_text())
    if args.season == "2025-26":
        signed_root_manifest = clean_text(
            getattr(args, "signed_root_manifest", "")
        )
        if not signed_root_manifest:
            raise SystemExit("--signed-root-manifest is required for 2025-26 ingest")
        validate_v13_signed_root_for_ingest(
            Path(signed_root_manifest),
            manifest_path,
            path,
            file_hash,
            args.team,
            args.season,
        )
    validate_intake_profile_manifest(
        source_manifest,
        manifest_path,
        file_hash,
        args.team,
        args.season,
        input_path=path,
    )
    rows = read_rows(path)
    excluded_source_fields = {
        clean_text(field)
        for field in clean_text(getattr(args, "exclude_source_fields", "")).split(",")
        if clean_text(field)
    }
    redacted_manifest_keys = {
        clean_text(key)
        for key in clean_text(getattr(args, "redact_manifest_keys", "")).split(",")
        if clean_text(key)
    } | {"current_file_team_aliases", "weekly_team_aliases"}
    redacted_source_value_keys = {
        clean_text(value).casefold()
        for value in clean_text(getattr(args, "redact_source_values", "")).split(",")
        if clean_text(value)
    }
    redacted_source_value_keys.update(
        clean_text(value).casefold() for value in load_fixture_team_aliases().values()
    )
    manifest = {
        "manifest": without_keys(source_manifest, redacted_manifest_keys),
        "original_path": str(path),
        "database_redactions": {
            "excluded_source_fields": sorted(excluded_source_fields),
            "redacted_manifest_keys": sorted(redacted_manifest_keys),
        },
    }

    row_sql = []
    redacted_source_value_count = 0
    params = SqlParams()
    provenance = run_provenance()
    for index, row in enumerate(rows, start=2):
        row_hash = hashlib.sha256(json.dumps(row, sort_keys=True).encode()).hexdigest()
        database_values = {}
        for field, value in row.items():
            if field in excluded_source_fields:
                continue
            if (
                clean_text(value).casefold() in redacted_source_value_keys
                or is_protected_team_alias_value(value)
            ):
                database_values[field] = "[REDACTED_PROTECTED_METADATA]"
                redacted_source_value_count += 1
            else:
                database_values[field] = value
        raw_record_id = f"{args.team}:{args.season}:{file_hash[:12]}:{index}"
        row_sql.append(
            f"""
            insert into ingestion.source_rows
              (source_file_id, source_row_number, raw_record_id, row_sha256, source_values)
            select id, {index}, {params.text(raw_record_id)}, {params.text(row_hash)}, {params.jsonb(database_values)}
            from ingestion.source_files
            where team = {params.text(args.team)}
              and season = {params.text(args.season)}
              and file_sha256 = {params.text(file_hash)}
            ;
            """
        )

    manifest["database_redactions"]["redacted_source_value_count"] = redacted_source_value_count
    sql = f"""
      do $$
      begin
        if exists (select 1 from ingestion.source_files where team = {params.text(args.team)} and season = {params.text(args.season)} and file_sha256 = {params.text(file_hash)}) then
          raise exception 'source file is already registered';
        end if;
      end $$;

      with source_file as (
        insert into ingestion.source_files
          (team, season, file_name, file_sha256, file_size_bytes, intake_manifest, row_count)
        values (
          {params.text(args.team)}, {params.text(args.season)}, {params.text(path.name)},
          {params.text(file_hash)}, {path.stat().st_size}, {params.jsonb(manifest)}, {len(rows) if rows else 'null'}
        )
        returning id
      )
      insert into audit.pipeline_runs (command, team, season, status, input_hash, parameters, ended_at, code_version, dependency_lock_hash, operator)
      values ('ingest', {params.text(args.team)}, {params.text(args.season)}, 'succeeded', {params.text(file_hash)}, {params.jsonb({'file': path.name, 'excluded_source_fields': sorted(excluded_source_fields), 'redacted_manifest_keys': sorted(redacted_manifest_keys), 'redacted_source_value_count': redacted_source_value_count})}, now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])});
      {"".join(row_sql)}
      do $$
      begin
        if (select count(*) from ingestion.source_rows sr join ingestion.source_files sf on sf.id = sr.source_file_id where sf.team = {params.text(args.team)} and sf.season = {params.text(args.season)} and sf.file_sha256 = {params.text(file_hash)}) <> {len(rows)} then
          raise exception 'ingest row cardinality mismatch';
        end if;
      end $$;
    """
    run_sql(sql, params.values)
    print(f"registered {path.name} sha256={file_hash} rows={len(rows) if rows else 'not-loaded'}")


def run_step(args: argparse.Namespace) -> None:
    params = SqlParams()
    provenance = run_provenance()
    sql = f"""
      with run as (
        insert into audit.pipeline_runs (command, team, season, status, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values ('run', {params.text(args.team)}, {params.text(args.season)}, 'succeeded', {params.jsonb({'step': args.step})}, now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])})
        returning id
      )
      insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count, ended_at)
      select id, {params.text(args.step)}, '0.1.0', 'placeholder_step', 0, 0, now()
      from run;
    """
    run_sql(sql, params.values)
    print(f"recorded step {args.step}")


RELEASE_DASHBOARDS_MIGRATION_VERSION = "20260710130000"
INJURY_COHORT_V1_AMENDMENT_MIGRATION_VERSION = "20260710120000"
FULL_DASHBOARD_RELEASE_RULE_VERSION = "full_dashboard_release_2026-07-10_v1"
LEAGUE_DASHBOARD_V2_MIGRATION_VERSION = "20260714130000"
ADJUDICATED_REPORTING_CLASSIFICATION_MIGRATION_VERSION = "20260720150000"
SEASON_BOUND_REPORTING_MIGRATION_VERSION = "20260720170000"
REVIEWED_BUNDLE_PAYLOAD_VALIDATION_MIGRATION_VERSION = "20260720180000"
OSIICS_EXACT_REPORTING_CLASSIFICATION_MIGRATION_VERSION = "20260722140000"
INCREMENTAL_CLASSIFICATION_BUNDLE_MIGRATION_VERSION = "20260722150000"
LEAGUE_DASHBOARD_RELEASE_RULE_VERSION = "league_dashboard_release_2026-07-14_v2"
SEASON_BOUND_LEAGUE_DASHBOARD_RELEASE_RULE_VERSION = "league_dashboard_release_2026-07-20_v3"
LINEAGE_LEAGUE_DASHBOARD_RELEASE_RULE_VERSION = "league_dashboard_release_2026-07-24_v4"
ANALYSIS_WINDOW_LEAGUE_DASHBOARD_RELEASE_RULE_VERSION = "league_dashboard_release_2026-07-25_v5"
INJURY_MASTER_LINEAGE_MIGRATION_VERSION = "20260724180000"
LINEAGE_RESTATED_REPORTING_MIGRATION_VERSION = "20260724181000"
LINEAGE_V4_CANDIDATE_FAST_PATH_MIGRATION_VERSION = "20260724190000"
ANALYSIS_WINDOW_REPORTING_V5_MIGRATION_VERSION = "20260725190000"
ANALYSIS_WINDOW_V5_CANDIDATE_OPTIMIZATION_MIGRATION_VERSION = "20260726010000"
ANALYSIS_WINDOW_V5_SHARED_COHORT_SNAPSHOT_MIGRATION_VERSION = "20260726015000"
ANALYSIS_WINDOW_V5_CANDIDATE_SNAPSHOT_MIGRATION_VERSION = "20260726020000"
ANALYSIS_WINDOW_V5_COVERAGE_SNAPSHOT_MIGRATION_VERSION = "20260726120000"
CONTACT_DISTRIBUTION_V5_MIGRATION_VERSION = "20260726160000"
CONTACT_DISTRIBUTION_READER_V4_MIGRATION_VERSION = "20260726161000"
ANALYSIS_WINDOW_V5_COHORT_VIEW_VERSION = "analysis_window_2024-25_2026-07-25_v1"
ANALYSIS_WINDOW_V5_EVIDENCE_LOCATOR = "docs/evidence/analysis_window_2024-25_v5.json"
ANALYSIS_WINDOW_V5_EVIDENCE_SHA256 = "c9530c949c60ff4abe91753571dfed6dd9d1146f33cc466dfbbc7fdeddb8443d"
ANALYSIS_WINDOW_V5_INJURY_AUDIT_LOCATOR = (
    "docs/evidence/analysis_window_2024-25_v5_injury_cohort_audit.csv"
)
ANALYSIS_WINDOW_V5_EXPOSURE_EVIDENCE_LOCATOR = (
    "docs/evidence/analysis_window_2024-25_v5_exposure_cohort_evidence.csv"
)
ANALYSIS_WINDOW_V5_SQL_RECONCILIATION_LOCATOR = (
    "docs/evidence/analysis_window_2024-25_v5_sql_reconciliation.json"
)
ANALYSIS_WINDOW_V5_CANDIDATE_PERFORMANCE_LOCATOR = (
    "docs/evidence/analysis_window_2024-25_v5_candidate_performance.json"
)
DASHBOARD_EXPORT_GRAIN_LABELS = {"weekly": "weekly", "session": "session-level", "mixed": "mixed-grain"}
# The five dashboard cohort-exclusion reason codes analysis.coverage_v1
# cannot reproduce under its curated-only read rule (see
# 20260710100000_analysis_views_v1.sql header). release_cohort_filter_flags()
# tallies these from audit.record_events (field_name='analysis_eligibility_status',
# action='exclude') instead -- audit evidence, outside any analysis.*_v1 view.
DASHBOARD_COHORT_FILTER_REASON_CODES = [
    "received_or_injured_in_other_team",
    "explicit_non_urc_match_type",
    "non_injury_problem_type",
    "injury_date_missing_or_outside_exposure_coverage",
    "adjudicated_duplicate",
]
RELEASE_TABLE_ROWS_SECTIONS = (
    "headline",
    "setting_split",
    "monthly",
    "body_locations",
    "injury_types",
    "severity_distribution",
)


def slug_key(text: str) -> str:
    """Stable, ascii-safe reporting.release_table_rows.row_key for a display
    label (setting_split / body_locations / injury_types row identity)."""
    slug = re.sub(r"[^a-z0-9]+", "_", clean_text(text).lower()).strip("_")
    return slug or "unknown"


def strip_none_keys(mapping: dict[str, Any], keys: set[str]) -> dict[str, Any]:
    return {key: value for key, value in mapping.items() if not (key in keys and value is None)}


def fetch_team_season_rows(sql_template: str, team_key: str, season: str) -> list[dict[str, Any]]:
    params = SqlParams()
    return query_sql(
        sql_template.format(team_key=params.text(team_key), season=params.text(season)),
        params.values,
    )


def team_display_name_for(team_key: str) -> str:
    params = SqlParams()
    rows = query_sql(
        f"select display_name from reporting.teams where team_key = {params.text(team_key)}",
        params.values,
    )
    if not rows:
        raise SystemExit(f"no reporting.teams row for team_key={team_key!r}")
    return rows[0]["display_name"]


def adjudicated_duplicate_row_numbers(team_key: str, season: str) -> list[int]:
    params = SqlParams()
    rows = query_sql(
        f"""
        select sr.source_row_number
        from audit.adjudications adj
        join ingestion.source_rows sr on sr.id = adj.source_row_id
        join ingestion.source_files sf on sf.id = sr.source_file_id
        join reporting.team_key_aliases a on a.alias = sf.team
        where a.team_key = {params.text(team_key)} and sf.season = {params.text(season)}
          and adj.decision ->> 'decision' = 'exclude_duplicate'
        order by sr.source_row_number
        """,
        params.values,
    )
    return sorted(int(row["source_row_number"]) for row in rows)


def combine_cohort_filter_reason_counts(
    audit_reason_rows: list[dict[str, Any]],
    curated_counts: dict[str, int],
) -> tuple[dict[str, int], dict[str, int]]:
    """Deduplicate current audit evidence and merge it with curated counts.

    Repeated events for one source row/reason count once. Audit evidence is
    authoritative for a reason it contains; curated evidence fills reasons
    absent from audit so a hybrid team never loses a current curated-only
    exclusion. Returns (merged_counts, audit_counts) so callers can preserve
    the existing injured_in_team_applied evidence rule.
    """
    unique_audit_evidence = {
        (clean_text(str(row.get("source_row_id", ""))), clean_text(str(row.get("reason_code", ""))))
        for row in audit_reason_rows
        if clean_text(str(row.get("source_row_id", "")))
        and clean_text(str(row.get("reason_code", ""))) in DASHBOARD_COHORT_FILTER_REASON_CODES
    }
    audit_counts: dict[str, int] = {}
    for _, reason_code in unique_audit_evidence:
        audit_counts[reason_code] = audit_counts.get(reason_code, 0) + 1

    merged_counts = {
        reason_code: int(count)
        for reason_code, count in curated_counts.items()
        if reason_code in DASHBOARD_COHORT_FILTER_REASON_CODES and int(count) > 0
    }
    merged_counts.update(audit_counts)
    return dict(sorted(merged_counts.items())), dict(sorted(audit_counts.items()))


def release_cohort_filter_flags(team_key: str, season: str) -> dict[str, Any]:
    """Adjudication 2 (data/reporting/analysis_parity_adjudications_2026-07-10.json):
    reproduces the dashboard's coverage.injury_cohort_filters block from
    audit/curated evidence at release time, outside any analysis.*_v1 view.

    Two evidence paths are merged per reason, with audit evidence overriding
    the same reason and curated cohort signals filling reasons absent from
    audit:

    - Audit path (glasgow): its old dashboard-run exclusions were folded
      into eligibility_status by its analysis-audit-file reapplication, and
      the per-reason detail lives only in audit.record_events
      (field_name='analysis_eligibility_status', action='exclude'). Only
      the latest record-version step carrying dashboard exclusion reasons is
      eligible while the active record remains excluded, and repeated
      source-row/reason events count once. This drops stale reasons after a
      later inclusion/review status, selects a newer reason set after a later
      re-exclusion, and preserves multi-filter lineage through an
      adjudication-only excluded version.
    - Curated path (edinburgh and every team processed after Phase 3.5,
      plus the four Irish teams, which have no such audit exclude events):
      the cohort exclusions actually applied by
      analysis.injury_cohort_v1's WHERE clause are counted directly from
      the active curated build, per view filter, using the view's own
      exposure-coverage-window formula. A row failing several filters
      counts once under each, matching the old dashboard's
      injury_scope_exclusion_counts semantics.

    injured_in_team_applied is true when either path shows the
    received/injured-in-team check operating: audit exclusions recorded,
    or any non-null received_in_team_status in the active build (the
    Phase 3.5 signal is computed at process-intake, so non-null values
    mean the view's filter is live for this team even if it excluded
    zero rows).
    """
    reason_list_sql = ", ".join(f"'{code}'" for code in DASHBOARD_COHORT_FILTER_REASON_CODES)
    reason_params = SqlParams()
    reason_rows = query_sql(
        f"""
        select distinct ev.source_row_id, ev.reason_code
        from curated.injuries i
        join curated.builds b on b.id = i.curated_build_id
        join processing.record_versions current_rv on current_rv.id = i.record_version_id
        join lateral (
          select historical_rv.step_run_id, historical_rv.eligibility_status
          from processing.record_versions historical_rv
          where historical_rv.source_row_id = i.source_row_id
            and historical_rv.version_number <= current_rv.version_number
            and historical_rv.eligibility_status in ('excluded_from_analysis', 'excluded_duplicate_adjudicated')
            and exists (
              select 1
              from audit.record_events historical_ev
              where historical_ev.step_run_id = historical_rv.step_run_id
                and historical_ev.source_row_id = historical_rv.source_row_id
                and historical_ev.field_name = 'analysis_eligibility_status'
                and historical_ev.action = 'exclude'
                and historical_ev.reason_code in ({reason_list_sql})
                and historical_ev.new_value #>> '{{}}' = historical_rv.eligibility_status
            )
          order by historical_rv.version_number desc, historical_rv.created_at desc, historical_rv.id desc
          limit 1
        ) applicable on true
        join audit.record_events ev
          on ev.step_run_id = applicable.step_run_id
          and ev.source_row_id = i.source_row_id
        where b.team_key = {reason_params.text(team_key)} and b.season = {reason_params.text(season)}
          and b.status = 'active'
          and current_rv.eligibility_status in ('excluded_from_analysis', 'excluded_duplicate_adjudicated')
          and ev.field_name = 'analysis_eligibility_status' and ev.action = 'exclude'
          and ev.reason_code in ({reason_list_sql})
          and ev.new_value #>> '{{}}' = applicable.eligibility_status
        """,
        reason_params.values,
    )

    curated_params = SqlParams()
    curated_rows = query_sql(
        f"""
        with active as (
          select id from curated.builds
          where team_key = {curated_params.text(team_key)} and season = {curated_params.text(season)}
            and status = 'active'
        ),
        win as (
          select
            min(coalesce(e.session_date, e.week_start_date)) as coverage_start,
            max(coalesce(e.session_date, e.week_start_date))
              + case when count(distinct e.grain) = 1 and min(e.grain) = 'weekly' then 6 else 0 end
              as coverage_end
          from curated.exposure e
          join active a on a.id = e.curated_build_id
          where e.eligibility_status = 'included_pending_protocol'
            and coalesce(e.session_date, e.week_start_date) is not null
        )
        select
          count(*) filter (where i.received_in_team_status in ('other_team', 'club')) as received_or_injured_in_other_team,
          count(*) filter (where i.urc_match_scope = 'non_urc_marker') as explicit_non_urc_match_type,
          count(*) filter (where i.problem_type is distinct from 'injury') as non_injury_problem_type,
          count(*) filter (
            where i.date_injured is null or i.date_injured < win.coverage_start or i.date_injured > win.coverage_end
          ) as injury_date_missing_or_outside_exposure_coverage,
          count(*) filter (where i.eligibility_status = 'excluded_duplicate_adjudicated') as adjudicated_duplicate,
          count(*) filter (where i.received_in_team_status is not null) as received_in_team_status_computed
        from curated.injuries i
        join active a on a.id = i.curated_build_id
        cross join win
        """,
        curated_params.values,
    )
    curated_counts = {
        code: int(curated_rows[0][code])
        for code in DASHBOARD_COHORT_FILTER_REASON_CODES
        if int(curated_rows[0][code]) > 0
    }
    received_signal_computed = int(curated_rows[0]["received_in_team_status_computed"]) > 0

    exclusion_reason_counts, audit_counts = combine_cohort_filter_reason_counts(
        reason_rows, curated_counts
    )
    return {
        "injured_in_team_applied": (
            "received_or_injured_in_other_team" in audit_counts or received_signal_computed
        ),
        "explicit_non_urc_match_type_applied": True,
        "exclusion_reason_counts": exclusion_reason_counts,
    }


def render_release_table_rows(rendered: dict[str, Any], monthly_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Flattens render_analysis_dashboard_sections() output (the Phase 3.2
    parity renderer, reused unmodified here) into reporting.release_table_rows
    records: one dict per row, each tagged with its section/row_key/ordinal.
    """
    table_rows: list[dict[str, Any]] = []
    for ordinal, metric in enumerate(rendered["headline"]):
        table_rows.append(
            {
                "section": "headline", "row_key": metric["key"], "ordinal": ordinal,
                "label": metric["label"], "value": metric["value"], "unit": metric["unit"],
                "numerator": metric.get("numerator"), "denominator": metric.get("denominator"),
                "formula": metric["formula"],
            }
        )
    for ordinal, row in enumerate(rendered["setting_split"]):
        table_rows.append(
            {
                "section": "setting_split", "row_key": slug_key(row["label"]), "ordinal": ordinal,
                "label": row["label"], "time_loss_injuries": row["time_loss_injuries"],
                "days_lost": row["days_lost"], "mean_severity_days": row["mean_severity_days"],
            }
        )
    monthly_sorted_raw = sorted(monthly_rows, key=lambda item: str(item["month_start_text"]))
    if len(monthly_sorted_raw) != len(rendered["monthly"]):
        raise SystemExit("monthly row count mismatch between raw view rows and rendered rows")
    for ordinal, (row, raw) in enumerate(zip(rendered["monthly"], monthly_sorted_raw)):
        table_rows.append(
            {
                "section": "monthly", "row_key": raw["month_start_text"], "ordinal": ordinal,
                "month": row["month"], "exposure_hours": row["exposure_hours"], "distance_km": row["distance_km"],
                "time_loss_injuries": row["time_loss_injuries"], "days_lost": row["days_lost"],
                "incidence_per_1000h": row["incidence_per_1000h"], "burden_per_1000h": row["burden_per_1000h"],
            }
        )
    for section in ("body_locations", "injury_types"):
        for ordinal, row in enumerate(rendered[section]):
            table_rows.append(
                {
                    "section": section, "row_key": slug_key(row["label"]), "ordinal": ordinal,
                    "label": row["label"], "time_loss_injuries": row["time_loss_injuries"],
                    "days_lost": row["days_lost"], "incidence_per_1000h": row["incidence_per_1000h"],
                    "burden_per_1000h": row["burden_per_1000h"], "mean_severity_days": row["mean_severity_days"],
                }
            )
    for ordinal, row in enumerate(rendered["severity_distribution"]):
        table_rows.append(
            {
                "section": "severity_distribution", "row_key": row["key"], "ordinal": ordinal,
                "label": row["label"], "recorded_injuries": row["recorded_injuries"],
                "time_loss_injuries": row["time_loss_injuries"], "days_lost": row["days_lost"],
            }
        )
    return table_rows


def assemble_release_dashboard(ctx: dict[str, Any], table_rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Build the public TeamDashboardData document from one release context
    and its flattened rows. Both read-only preflight and post-write export
    use this exact serializer so their strict diff can differ only at the
    generated_at timestamp.
    """
    sections: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in table_rows:
        section = row["section"]
        if section == "headline":
            doc = strip_none_keys(
                {
                    "key": row["row_key"], "label": row["label"], "value": as_number(row["value"]),
                    "unit": row["unit"], "numerator": as_number(row["numerator"]),
                    "denominator": as_number(row["denominator"]), "formula": row["formula"],
                },
                {"numerator", "denominator"},
            )
        elif section == "setting_split":
            doc = {
                "label": row["label"], "time_loss_injuries": as_number(row["time_loss_injuries"]),
                "days_lost": as_number(row["days_lost"]), "mean_severity_days": as_number(row["mean_severity_days"]),
            }
        elif section == "monthly":
            doc = {
                "month": row["month"], "exposure_hours": as_number(row["exposure_hours"]),
                "distance_km": as_number(row["distance_km"]),
                "time_loss_injuries": as_number(row["time_loss_injuries"]),
                "days_lost": as_number(row["days_lost"]),
                "incidence_per_1000h": as_number(row["incidence_per_1000h"]),
                "burden_per_1000h": as_number(row["burden_per_1000h"]),
            }
        elif section in ("body_locations", "injury_types"):
            doc = {
                "label": row["label"], "time_loss_injuries": as_number(row["time_loss_injuries"]),
                "days_lost": as_number(row["days_lost"]),
                "incidence_per_1000h": as_number(row["incidence_per_1000h"]),
                "burden_per_1000h": as_number(row["burden_per_1000h"]),
                "mean_severity_days": as_number(row["mean_severity_days"]),
            }
        elif section == "severity_distribution":
            doc = {
                "key": row["row_key"], "label": row["label"],
                "recorded_injuries": as_number(row["recorded_injuries"]),
                "time_loss_injuries": as_number(row["time_loss_injuries"]),
                "days_lost": as_number(row["days_lost"]),
            }
        else:
            raise SystemExit(f"unknown release_table_rows section {section!r}")
        sections[section].append(doc)

    dashboard = {
        "generated_at": ctx["generated_at"],
        "team": ctx["team_display_name"],
        "season": ctx["season"],
        "analysis_window": {
            "start": ctx["analysis_window_start"], "end": ctx["analysis_window_end"],
            "basis": ctx["analysis_window_basis"],
        },
        "method": ctx["method"],
        "coverage": ctx["coverage"],
        "headline": sections["headline"],
        "setting_split": sections["setting_split"],
        "monthly": sections["monthly"],
        "body_locations": sections["body_locations"],
        "injury_types": sections["injury_types"],
        "severity_distribution": sections["severity_distribution"],
        "prior_season": ctx["prior_season"],
        "limitations": ctx["limitations"],
    }
    return without_keys(dashboard, {"source_files", "pipeline_evidence"})


def write_text_atomic(path: Path, value: str) -> None:
    """Write text by same-directory replace so a failed write keeps `path`."""
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            "w",
            encoding="utf-8",
            dir=path.parent,
            prefix=f".{path.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            temporary_path = Path(handle.name)
            handle.write(value)
            handle.flush()
            os.fsync(handle.fileno())
        mode = (path.stat().st_mode & 0o777) if path.exists() else 0o644
        os.chmod(temporary_path, mode)
        os.replace(temporary_path, path)
        temporary_path = None
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def write_json_atomic(path: Path, value: object) -> None:
    write_text_atomic(path, json.dumps(value, indent=2) + "\n")


def export_release_dashboard_json(release_label: str) -> dict[str, Any]:
    """Read the just-written release snapshot back from the DB and assemble
    the public dashboard through the same serializer as release preflight.
    """
    ctx_params = SqlParams()
    ctx_rows = query_sql(
        f"""
        select
          rc.team_display_name, rc.season,
          to_char(rc.generated_at at time zone 'utc', 'YYYY-MM-DD"T"HH24:MI:SS"Z"') as generated_at,
          rc.analysis_window_start::text as analysis_window_start,
          rc.analysis_window_end::text as analysis_window_end,
          rc.analysis_window_basis, rc.method, rc.coverage, rc.prior_season, rc.limitations
        from reporting.release_context rc
        join reporting.aggregate_releases r on r.id = rc.release_id
        where r.release_label = {ctx_params.text(release_label)}
        """,
        ctx_params.values,
    )
    if len(ctx_rows) != 1:
        raise SystemExit(f"expected exactly one release_context row for release_label={release_label!r}")

    rows_params = SqlParams()
    table_rows = query_sql(
        f"""
        select r.section, r.row_key, r.label, r.month, r.value, r.numerator, r.denominator,
          r.unit, r.formula, r.exposure_hours, r.distance_km, r.time_loss_injuries, r.recorded_injuries,
          r.days_lost, r.incidence_per_1000h, r.burden_per_1000h, r.mean_severity_days
        from reporting.release_table_rows r
        join reporting.aggregate_releases rel on rel.id = r.release_id
        where rel.release_label = {rows_params.text(release_label)}
        order by r.section, r.ordinal
        """,
        rows_params.values,
    )
    return assemble_release_dashboard(ctx_rows[0], table_rows)


# ---------------------------------------------------------------------------
# Phase 4.3 planning: old-vs-new dashboard JSON diff check. Pure file diff,
# no DB access -- meant to gate the five-team re-release, one team at a
# time: diff the previously-committed content/reporting/<team>_dashboard.json
# against the freshly-exported one and confirm every difference is a
# whitelisted Phase 4 shape/text change (never a numeric change to an
# already-published value).
# ---------------------------------------------------------------------------

DASHBOARD_JSON_DIFF_WHITELIST_NOTES = {
    "generated_at": "release timestamp is always regenerated at release time",
}


def classify_dashboard_json_diff(path: str, kind: str) -> str | None:
    """Returns a whitelist reason string if this old-vs-new dashboard JSON
    diff at `path` is an expected, approved Phase 4 shape/text change; None
    if the diff must be treated as a blocking, unexplained change.

    Whitelisted categories (all recorded in
    data/reporting/analysis_parity_adjudications_2026-07-10.json, or this
    executor's own documented, flagged design decisions -- see the Phase 4
    gate report):
      - generated_at always differs (fresh release timestamp).
      - source_files / pipeline_evidence disappearing (Phase 4 internal-key
        stripping fix).
      - coverage.scope_status -> coverage.scope_status_counts, and newly
        appearing coverage.exposure_periods / coverage.exposure_grain
        (Adjudication 1: coverage-shape regeneration).
      - coverage.injury_cohort_filters appearing or changing (Adjudication 2:
        carried from audit evidence at release time, outside analysis views).
      - method[*] / limitations[*] narrative lines (including their array
        lengths): the committed JSONs carry heterogeneous text vintages
        (munster 7 method lines, connacht/leinster/ulster 8, edinburgh 9,
        glasgow 10; Irish limitations still say 'local and draft'), while
        the release command regenerates the current frozen narrative for
        every team from DB evidence (grain, cohort-filter state, and
        audit.adjudications-derived duplicate rows). These lines are
        descriptive method text, never published metric values; the
        embedded adjudicated-row numbers are re-derived from
        audit.adjudications at release time, so a wrong number here would
        mean wrong audit evidence, which the release gates check
        separately.
    Note: the contact mechanism ring (2026-07-26) deliberately did NOT add an
    entry here. release-league never calls this classifier -- its parity export
    writes from the promoted bundle and treats any bundle diff as fatal -- so
    whitelisting contact_distribution would have widened a frozen gate that
    only `release` and `diff-dashboard-json` use, for no benefit. If either of
    those paths ever meets the new section, blocking is the correct outcome and
    the diff belongs in a recorded adjudication.
    Anything else -- any numeric headline/monthly/body_locations/
    injury_types/severity_distribution/setting_split value, any label,
    team name, or analysis-window value -- is BLOCKED.
    """
    if path in DASHBOARD_JSON_DIFF_WHITELIST_NOTES:
        return DASHBOARD_JSON_DIFF_WHITELIST_NOTES[path]
    if path == "method" or path.startswith("method[") or path == "method.length":
        return "regenerated method narrative (current frozen wording, derived from DB evidence at release time)"
    if path == "limitations" or path.startswith("limitations[") or path == "limitations.length":
        return "regenerated limitations narrative (current frozen wording, derived from DB evidence at release time)"
    if path == "source_files" or path.startswith("source_files."):
        return "Phase 4 internal-key stripping (source_files is never exported)"
    if path == "pipeline_evidence" or path.startswith("pipeline_evidence."):
        return "Phase 4 internal-key stripping (pipeline_evidence is never exported)"
    if path == "coverage.scope_status" and kind == "missing_in_new":
        return "Adjudication 1: coverage-shape regeneration (scope_status -> scope_status_counts)"
    if path == "coverage.scope_status_counts" and kind == "extra_in_new":
        return "Adjudication 1: coverage-shape regeneration (scope_status -> scope_status_counts)"
    if path in {"coverage.exposure_periods", "coverage.exposure_grain"} and kind in {
        "missing_in_new", "extra_in_new",
    }:
        return "Adjudication 1: coverage-shape regeneration (added exposure_periods/exposure_grain)"
    if path == "coverage.injury_cohort_filters" or path.startswith("coverage.injury_cohort_filters."):
        return "Adjudication 2: injury_cohort_filters carried from audit evidence at release time"
    return None


def diff_json_documents(old: object, new: object) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []

    def record(path: str, old_value: object, new_value: object, kind: str) -> None:
        results.append({"path": path, "kind": kind, "old": old_value, "new": new_value})

    def walk(path: str, old_value: object, new_value: object) -> None:
        if isinstance(old_value, dict) and isinstance(new_value, dict):
            for key in sorted(set(old_value) | set(new_value)):
                child_path = f"{path}.{key}" if path else key
                if key not in new_value:
                    record(child_path, old_value[key], None, "missing_in_new")
                elif key not in old_value:
                    record(child_path, None, new_value[key], "extra_in_new")
                else:
                    walk(child_path, old_value[key], new_value[key])
            return
        if isinstance(old_value, list) and isinstance(new_value, list):
            if len(old_value) != len(new_value):
                record(f"{path}.length", len(old_value), len(new_value), "row_count")
            for index in range(max(len(old_value), len(new_value))):
                child_path = f"{path}[{index}]"
                if index >= len(new_value):
                    record(child_path, old_value[index], None, "missing_in_new")
                elif index >= len(old_value):
                    record(child_path, None, new_value[index], "extra_in_new")
                else:
                    walk(child_path, old_value[index], new_value[index])
            return
        if isinstance(old_value, (dict, list)) != isinstance(new_value, (dict, list)):
            record(path, old_value, new_value, "type_mismatch")
            return
        if not parity_values_equal(old_value, new_value):
            record(path, old_value, new_value, "value_mismatch")

    walk("", old, new)
    return results


def classify_preflight_release_diffs(diffs: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Strict first-release comparison: generated_at is the only permitted
    difference between the reviewed preflight candidate and release output.
    """
    allowed = []
    blocked = []
    for diff in diffs:
        if diff["path"] == "generated_at" and diff["kind"] == "value_mismatch":
            allowed.append(
                {
                    **diff,
                    "whitelist_reason": "release timestamp is regenerated after preflight approval",
                }
            )
        else:
            blocked.append(diff)
    return allowed, blocked


def classify_historical_release_diffs(
    diffs: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Classify a re-release against the frozen historical whitelist."""
    allowed = []
    blocked = []
    for diff in diffs:
        reason = classify_dashboard_json_diff(diff["path"], diff["kind"])
        if reason is None:
            blocked.append(diff)
        else:
            allowed.append({**diff, "whitelist_reason": reason})
    return allowed, blocked


def validate_release_restatement_envelope(
    envelope: dict[str, Any],
    *,
    team_key: str,
    season: str,
    previous_dashboard_sha256: str,
    release_content_hash: str,
    blocked_diffs: list[dict[str, Any]],
) -> dict[str, Any]:
    """Validate one exact, reviewer-approved exception to historical drift."""
    expected = {
        "schema_version": "release_restatement_v1",
        "team_key": team_key,
        "season": season,
        "previous_dashboard_sha256": previous_dashboard_sha256,
        "release_content_hash": release_content_hash,
        "blocked_diffs_sha256": sha256_json(blocked_diffs),
        "reason_code": "input_representation_correction",
        "approved_by": "Abdel Babiker",
    }
    for field, value in expected.items():
        if envelope.get(field) != value:
            label = "blocked diff checksum" if field == "blocked_diffs_sha256" else field
            raise SystemExit(f"release restatement {label} does not match the current correction")
    if not clean_text(envelope.get("rationale")):
        raise SystemExit("release restatement rationale is required")
    try:
        approved_at = datetime.fromisoformat(
            clean_text(envelope.get("approved_at")).replace("Z", "+00:00")
        )
        if approved_at.tzinfo is None or approved_at.astimezone(UTC) > datetime.now(UTC) + timedelta(minutes=5):
            raise ValueError
    except ValueError as exc:
        raise SystemExit("release restatement approved_at must be a valid non-future timezone-aware value") from exc
    return envelope


def validate_release_restatement(
    path: Path,
    **kwargs: Any,
) -> dict[str, Any]:
    try:
        envelope = json.loads(path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit(f"release restatement must be valid JSON: {path}") from exc
    if not isinstance(envelope, dict):
        raise SystemExit("release restatement must be a JSON object")
    return validate_release_restatement_envelope(envelope, **kwargs)


def diff_dashboard_json(args: argparse.Namespace) -> None:
    """Phase 4.3 planning tool: pure file diff (no DB access) between an old
    (previously-committed) and new (just-exported) dashboard JSON,
    classifying every difference as ALLOWED (a whitelisted Phase 4
    shape/text change, per classify_dashboard_json_diff) or BLOCKED (an
    unexplained change that must be investigated before the release is kept).
    Exit code is nonzero if any BLOCKED diff exists.
    """
    old_path = Path(args.old)
    new_path = Path(args.new)
    if not old_path.exists():
        raise SystemExit(f"old dashboard JSON not found: {old_path}")
    if not new_path.exists():
        raise SystemExit(f"new dashboard JSON not found: {new_path}")
    old_doc = json.loads(old_path.read_text())
    new_doc = json.loads(new_path.read_text())

    diffs = diff_json_documents(old_doc, new_doc)
    strict_preflight_release = bool(getattr(args, "preflight_release", False))
    if strict_preflight_release:
        allowed, blocked = classify_preflight_release_diffs(diffs)
    else:
        allowed, blocked = classify_historical_release_diffs(diffs)

    result = {
        "mode": "preflight_release" if strict_preflight_release else "historical_release",
        "old": str(old_path),
        "new": str(new_path),
        "total_diffs": len(diffs),
        "allowed": len(allowed),
        "blocked": len(blocked),
        "overall": "BLOCKED" if blocked else "ALLOWED_ONLY",
        "blocked_diffs": blocked,
        "allowed_diffs": allowed,
    }
    print(json.dumps(result, indent=2))
    if blocked:
        raise SystemExit(1)


def release_promotion_statement(
    release_label: str,
    curated_build_id: str,
    team_key: str,
    season: str,
    verified_candidate: dict[str, Any],
    expected_previous_release_id: str | None,
) -> tuple[str, list[object]]:
    """Build the atomic approval statement.

    The public consumer payload is reconstructed through
    reporting.latest_team_dashboard *after* the guarded draft-to-approved
    update, then compared with the already verified draft candidate before
    either approval or audit success can commit.
    """
    params = SqlParams()
    label_sql = params.text(release_label)
    curated_build_sql = params.text(curated_build_id)
    candidate_sql = params.jsonb(verified_candidate)
    team_key_sql = params.text(team_key)
    season_sql = params.text(season)
    expected_previous_release_sql = params.text(expected_previous_release_id)
    sql = f"""
      do $$
      declare
        changed integer;
        run_id uuid;
        target_release_id uuid;
        actual_previous_release_id uuid;
        consumer_dashboard jsonb;
      begin
        -- Serialize promotions per team. This lock plus the predecessor
        -- recheck prevents two processes that read the same history from
        -- both approving a successor.
        perform 1
        from reporting.teams t
        where t.team_key = {team_key_sql}
        for update;
        if not found then
          raise exception 'release promotion requires an existing reporting team';
        end if;

        select r.id, r.pipeline_run_id into target_release_id, run_id
        from reporting.aggregate_releases r
        where r.release_label = {label_sql} and r.status = 'draft'
        for update;
        if target_release_id is null or run_id is null then
          raise exception 'release promotion requires exactly one draft release';
        end if;

        select rc.release_id into actual_previous_release_id
        from reporting.release_context rc
        join reporting.aggregate_releases r on r.id = rc.release_id
        join audit.pipeline_runs pr on pr.id = r.pipeline_run_id
        where rc.team_key = {team_key_sql}
          and rc.season = {season_sql}
          and (
            r.status = 'approved'
            or (r.status = 'retired' and pr.status = 'succeeded')
          )
        order by
          case when r.status = 'approved' then 0 else 1 end,
          r.approved_at desc nulls last,
          r.created_at desc,
          r.id desc
        limit 1;
        if actual_previous_release_id is distinct from {expected_previous_release_sql}::uuid then
          raise exception 'release promotion refused: accepted predecessor changed; rerun release';
        end if;

        if not exists (
          select 1 from curated.builds
          where id = {curated_build_sql}::uuid and status = 'active'
        ) then
          raise exception 'release promotion refused: curated build is no longer active';
        end if;

        update reporting.aggregate_releases
        set status = 'approved', approved_at = clock_timestamp()
        where id = target_release_id
          and status = 'draft'
          and pipeline_run_id = run_id;
        get diagnostics changed = row_count;
        if changed <> 1 then
          raise exception 'release promotion failed to approve exactly one draft';
        end if;

        update reporting.aggregate_releases r
        set status = 'retired'
        where r.id <> target_release_id
          and r.status = 'approved'
          and exists (
            select 1
            from reporting.release_context rc
            where rc.release_id = r.id
              and rc.team_key = {team_key_sql}
              and rc.season = {season_sql}
          );

        select jsonb_build_object(
          'generated_at', to_char(d.generated_at at time zone 'utc', 'YYYY-MM-DD"T"HH24:MI:SS"Z"'),
          'team', d.team,
          'season', d.season,
          'analysis_window', d.analysis_window,
          'method', d.method,
          'coverage', d.coverage,
          'headline', d.headline,
          'setting_split', d.setting_split,
          'monthly', d.monthly,
          'body_locations', d.body_locations,
          'injury_types', d.injury_types,
          'severity_distribution', d.severity_distribution,
          'prior_season', d.prior_season,
          'limitations', d.limitations
        ) into consumer_dashboard
        from reporting.latest_team_dashboard d
        where d.release_id = target_release_id
          and d.team_key = {team_key_sql}
          and d.season = {season_sql};

        if consumer_dashboard is null then
          raise exception 'release promotion failed: consumer view did not resolve the approved release';
        end if;
        if consumer_dashboard is distinct from {candidate_sql} then
          raise exception 'release promotion failed: consumer view payload differs from verified draft candidate';
        end if;

        update audit.pipeline_runs
        set status = 'succeeded', ended_at = now()
        where id = run_id and status = 'started';
        get diagnostics changed = row_count;
        if changed <> 1 then
          raise exception 'release promotion requires one started pipeline run';
        end if;

        update audit.step_runs
        set ended_at = now()
        where pipeline_run_id = run_id
          and step_name = 'release_full_dashboard'
          and ended_at is null;
        get diagnostics changed = row_count;
        if changed <> 1 then
          raise exception 'release promotion requires one open release step';
        end if;
      end $$;
    """
    return sql, params.values


def execute_release_promotion(
    release_label: str,
    curated_build_id: str,
    team_key: str,
    season: str,
    verified_candidate: dict[str, Any],
    expected_previous_release_id: str | None,
    runner: Any = None,
) -> None:
    """Execute the atomic promotion; runner injection keeps tests offline."""
    sql, params = release_promotion_statement(
        release_label,
        curated_build_id,
        team_key,
        season,
        verified_candidate,
        expected_previous_release_id,
    )
    (runner or run_sql)(sql, params)


def release_failure_cleanup_statement(
    release_label: str,
    team_key: str,
    season: str,
    previous_release_id: str | None,
    previous_release_status: str | None,
    failure_stage: str,
) -> tuple[str, list[object]]:
    """Retire a failed attempt and restore the prior public release safely.

    If promotion committed but local artifact export failed, the exact prior
    approved predecessor is re-approved under the same team lock. A failed
    draft never displaced it, and a later concurrent successor prevents the
    restoration branch from firing.
    """
    params = SqlParams()
    label_sql = params.text(release_label)
    team_key_sql = params.text(team_key)
    season_sql = params.text(season)
    previous_release_id_sql = params.text(previous_release_id)
    previous_release_status_sql = params.text(previous_release_status)
    failure_parameters_sql = params.jsonb({"failure_stage": failure_stage})
    sql = f"""
      do $$
      declare
        failed_release_id uuid;
        latest_approved_id uuid;
      begin
        perform 1
        from reporting.teams t
        where t.team_key = {team_key_sql}
        for update;

        select r.id into failed_release_id
        from reporting.aggregate_releases r
        where r.release_label = {label_sql};

        select rc.release_id into latest_approved_id
        from reporting.release_context rc
        join reporting.aggregate_releases r on r.id = rc.release_id
        where rc.team_key = {team_key_sql}
          and rc.season = {season_sql}
          and r.status = 'approved'
        order by r.approved_at desc nulls last, r.created_at desc, r.id desc
        limit 1;

        update reporting.aggregate_releases r
        set status = 'retired'
        where r.id = failed_release_id
          and r.status in ('draft', 'approved');

        if latest_approved_id = failed_release_id
           and {previous_release_status_sql} = 'approved'
           and {previous_release_id_sql}::uuid is not null then
          update reporting.aggregate_releases r
          set status = 'approved'
          where r.id = {previous_release_id_sql}::uuid
            and r.status = 'retired';
        end if;
      end $$;

      update audit.pipeline_runs pr
      set status = 'failed', ended_at = now(),
          parameters = pr.parameters || {failure_parameters_sql}
      from reporting.aggregate_releases r
      where r.release_label = {label_sql}
        and pr.id = r.pipeline_run_id
        and pr.status in ('started', 'succeeded');

      update audit.step_runs sr
      set ended_at = coalesce(sr.ended_at, now())
      from reporting.aggregate_releases r
      where r.release_label = {label_sql}
        and sr.pipeline_run_id = r.pipeline_run_id
        and sr.step_name = 'release_full_dashboard';
    """
    return sql, params.values


def assert_checksum_bound_migrations(
    contracts: tuple[Any, ...], operation: str,
    *, local_evidence_records: list[dict[str, str]] | None = None,
) -> None:
    """Fail closed unless exact local migration bytes equal registered rows."""
    if not contracts or len({item.version for item in contracts}) != len(contracts):
        raise SystemExit(f"{operation} lacks unique checksum-bound migration contracts")
    root = Path(__file__).resolve().parents[1]
    for item in contracts:
        migration_path = root / "supabase" / "migrations" / f"{item.version}_{item.name}.sql"
        if not migration_path.is_file() or sha256_file(migration_path) != item.sha256:
            raise SystemExit(
                f"{operation} local migration bytes do not match the approved contract: {item.version}"
            )
    if local_evidence_records is not None:
        assert_local_evidence_bytes(local_evidence_records, operation)
    migration_params = SqlParams()
    registered = query_sql(
        f"select version, name, statements from supabase_migrations.schema_migrations "
        f"where version = any(array[{', '.join(migration_params.text(item.version) for item in contracts)}])",
        migration_params.values,
    )
    expected = {
        item.version: (item.name, [item.statement])
        for item in contracts
    }
    actual = {
        clean_text(row.get("version")): (clean_text(row.get("name")), row.get("statements"))
        for row in registered
    }
    if actual != expected:
        raise SystemExit(
            f"{operation} requires exact registered migration checksums for "
            + ", ".join(item.version for item in contracts)
        )


def assert_checksum_bound_release_migrations(contract: Any, operation: str) -> None:
    """Fail closed unless local V6 migration bytes equal the registered rows."""
    contracts = tuple(contract.required_migration_contracts)
    versions = tuple(contract.required_migrations)
    if not contracts or {item.version for item in contracts} != set(versions):
        raise SystemExit(f"{operation} lacks complete checksum-bound migration contracts")
    assert_checksum_bound_migrations(
        contracts,
        operation,
        local_evidence_records=year2_release_local_evidence_records(contract),
    )


def year2_release_local_evidence_records(contract: Any) -> list[dict[str, str]]:
    fixture_contract = fixture_contract_for("2025-26")
    if fixture_contract is None:
        raise SystemExit("V6 release contract lacks the Year 2 fixture evidence identity")
    records = [
        {
            "role": "fixture_preparation",
            "locator": fixture_contract.evidence_locator,
            "sha256": fixture_contract.evidence_sha256,
        },
        {
            "role": "cohort_rule",
            "locator": clean_text(contract.cohort_evidence_locator),
            "sha256": clean_text(contract.cohort_evidence_sha256),
        },
        {
            "role": "classification_rule",
            "locator": clean_text(contract.classification_rule_evidence_locator),
            "sha256": clean_text(contract.classification_rule_evidence_sha256),
        },
        {
            "role": "incomplete_exposure_reporting",
            "locator": clean_text(contract.exposure_coverage_evidence_locator),
            "sha256": clean_text(contract.exposure_coverage_evidence_sha256),
        },
        {
            "role": "injury_eligibility_bridge",
            "locator": clean_text(contract.injury_eligibility_evidence_locator),
            "sha256": clean_text(contract.injury_eligibility_evidence_sha256),
        },
    ]
    if any(not item["locator"] or not re.fullmatch(r"[0-9a-f]{64}", item["sha256"]) for item in records):
        raise SystemExit("V6 release contract lacks exact local evidence identities")
    return records


def assert_local_evidence_bytes(records: list[dict[str, str]], operation: str) -> None:
    """Bind claimed safe evidence identities to the repository bytes in use."""
    root = Path(__file__).resolve().parents[1]
    for item in records:
        locator = clean_text(item.get("locator"))
        expected_sha256 = clean_text(item.get("sha256"))
        path = (root / locator).resolve()
        if root not in path.parents or not path.is_file() or sha256_file(path) != expected_sha256:
            raise SystemExit(
                f"{operation} local evidence bytes do not match the approved contract: {item.get('role')}"
            )


def v6_team_preflight_manifest(
    *, team_key: str, contract: Any, candidate: dict[str, Any],
    predecessor: dict[str, Any] | None, preflight_file_sha256: str,
    provenance: dict[str, str],
) -> dict[str, object]:
    """Build the exact, checksum-bound local review record for one V6 team."""
    return {
        "schema_version": "urc_team_release_v6_preflight_v1",
        "season": "2025-26",
        "team_key": team_key,
        "release_tuple": {
            "analysis_version": "v6",
            "classification_view_version": contract.classification_view_version,
            "cohort_view_version": contract.cohort_view_version,
        },
        "candidate_view": contract.team_candidate_view,
        "curated_build_id": candidate["curated_build_id"],
        "payload_sha256": candidate["payload_sha256"],
        "classification_evidence_sha256": candidate["classification_evidence_sha256"],
        "cohort_evidence_sha256": candidate["cohort_evidence_sha256"],
        "local_evidence_files": year2_release_local_evidence_records(contract),
        "required_migrations": [
            {
                "version": item.version,
                "name": item.name,
                "sha256": item.sha256,
            }
            for item in contract.required_migration_contracts
        ],
        "provenance": {
            "code_version": provenance["code_version"],
            "dependency_lock_hash": provenance["dependency_lock_hash"],
        },
        "predecessor_release_id": predecessor["release_id"] if predecessor else None,
        "preflight_file_sha256": preflight_file_sha256,
    }


def read_v6_team_reviewed_preflight(
    *, reviewed_path: Path, team_key: str, contract: Any,
    candidate: dict[str, Any], predecessor: dict[str, Any] | None,
    provenance: dict[str, str],
) -> tuple[dict[str, Any], dict[str, Any], str, str]:
    """Read only a full, exact manifest bound to the reviewed payload bytes."""
    manifest_path = Path(f"{reviewed_path}.manifest.json")
    if not reviewed_path.is_file() or not manifest_path.is_file():
        raise SystemExit("reviewed V6 preflight and its manifest are required")
    try:
        reviewed_bytes = reviewed_path.read_bytes()
        manifest_bytes = manifest_path.read_bytes()
        reviewed = json.loads(reviewed_bytes)
        reviewed_exact = json.loads(
            reviewed_bytes,
            parse_float=Decimal,
            parse_int=Decimal,
        )
        manifest = json.loads(manifest_bytes)
    except (OSError, json.JSONDecodeError) as error:
        raise SystemExit("reviewed V6 preflight artefacts must be valid JSON") from error
    if (
        not isinstance(reviewed, dict)
        or not isinstance(reviewed_exact, dict)
        or not isinstance(manifest, dict)
    ):
        raise SystemExit("reviewed V6 preflight manifest and payload must be JSON objects")
    reviewed_sha256 = hashlib.sha256(reviewed_bytes).hexdigest()
    expected_manifest = v6_team_preflight_manifest(
        team_key=team_key,
        contract=contract,
        candidate=candidate,
        predecessor=predecessor,
        preflight_file_sha256=reviewed_sha256,
        provenance=provenance,
    )
    if manifest != expected_manifest:
        raise SystemExit("reviewed V6 preflight manifest does not bind the exact active candidate")
    return (
        reviewed,
        reviewed_exact,
        reviewed_sha256,
        hashlib.sha256(manifest_bytes).hexdigest(),
    )


def v6_team_release_label(
    *, team_key: str, curated_build_id: str, payload_sha256: str,
    attempt_id: str | None = None,
) -> str:
    """Return an append-only V6 identity, independent of public payload bytes."""
    try:
        build_token = str(uuid.UUID(curated_build_id))
    except ValueError as error:
        raise SystemExit("V6 team candidate has an invalid curated build ID") from error
    if not re.fullmatch(r"[0-9a-f]{64}", payload_sha256):
        raise SystemExit("V6 team candidate lacks its canonical database payload hash")
    release_attempt = clean_text(attempt_id or uuid.uuid4().hex)
    if not re.fullmatch(r"[0-9a-f]{12,32}", release_attempt):
        raise SystemExit("V6 team release attempt ID is invalid")
    return (
        f"urc-2025-26-v6-{team_key}-{build_token}-"
        f"{payload_sha256[:12]}-{release_attempt}"
    )


def release_team_v6(args: argparse.Namespace) -> None:
    """Promote one reviewed Year 2 team payload without legacy release tables.

    The candidate is deliberately build-derived and has no accepted-release
    dependency.  A later league release can only see this immutable payload
    through ``analysis.league_member_releases_v6`` once all sixteen teams are
    approved.
    """
    team = clean_text(args.team)
    season = clean_text(args.season)
    if not team or season != "2025-26":
        raise SystemExit("release_team_v6 requires --team and --season 2025-26")
    preflight = bool(getattr(args, "preflight", False))
    preflight_file = clean_text(getattr(args, "preflight_file", "") or "")
    reviewer = clean_text(getattr(args, "preflight_reviewer", "") or "")
    if preflight and preflight_file:
        raise SystemExit("--preflight cannot be combined with --preflight-file")
    if bool(getattr(args, "previous_dashboard_file", "") or getattr(args, "restatement_file", "")):
        raise SystemExit("V6 team successors are selected from the immutable prior release automatically; dashboard-restatement files are not accepted here")
    if preflight_file and not reviewer:
        raise SystemExit("--preflight-reviewer is required with --preflight-file")
    if reviewer and not preflight_file:
        raise SystemExit("--preflight-reviewer requires --preflight-file")
    if preflight_file and reviewer != "Abdel Babiker":
        raise SystemExit("V6 team promotion requires --preflight-reviewer 'Abdel Babiker'")
    if not preflight and not preflight_file:
        raise SystemExit("V6 team release requires --preflight or a reviewed --preflight-file")
    provenance = run_provenance()
    if provenance["code_version"].endswith("-dirty"):
        raise SystemExit("V6 team release refuses an uncommitted working tree")

    team_key = resolve_team_key(team)
    contract = release_contract_for("2025-26", YEAR2_2025_26_RELEASE_TUPLE)
    if not contract.team_candidate_view:
        raise SystemExit("V6 release contract lacks a team candidate view")
    assert_checksum_bound_release_migrations(contract, "V6 team release")

    candidate_params = SqlParams()
    rows = query_sql(
        f"""
        select candidate.team_key, candidate.season,
          candidate.curated_build_id::text, candidate.analysis_version,
          candidate.classification_view_version,
          candidate.classification_evidence_sha256,
          candidate.cohort_view_version, candidate.cohort_evidence_sha256,
          candidate.dashboard::text as dashboard_json,
          reporting.canonical_jsonb_sha256_v1(candidate.dashboard) as payload_sha256,
          coalesce((
            select jsonb_agg(
              jsonb_build_object(
                'release_id', payload.release_id::text,
                'release_label', release.release_label,
                'payload_sha256', payload.payload_sha256
              )
              order by release.approved_at desc nulls last,
                release.created_at desc, payload.release_id desc
            )
            from reporting.team_release_payloads_v6 payload
            join reporting.aggregate_releases release on release.id = payload.release_id
            where payload.team_key = candidate.team_key
              and payload.season = candidate.season
              and release.status = 'approved'
          ), '[]'::jsonb) as approved_predecessors
        from {contract.team_candidate_view} candidate
        where candidate.team_key = {candidate_params.text(team_key)}
          and candidate.season = '2025-26' and candidate.analysis_version = 'v6'
          and candidate.classification_view_version = 'reporting_classification_2026-07-22_v2'
          and candidate.cohort_view_version = 'analysis_window_2025-26_2026-08-15_v1'
        """,
        candidate_params.values,
    )
    if len(rows) != 1 or not isinstance(rows[0].get("dashboard_json"), str):
        raise SystemExit("V6 team release requires exactly one complete active-build candidate")
    candidate = rows[0]
    if not all(
        isinstance(candidate.get(field), str)
        and re.fullmatch(r"[0-9a-f]{64}", candidate[field])
        for field in ("classification_evidence_sha256", "cohort_evidence_sha256")
    ):
        raise SystemExit("V6 team candidate lacks exact classification or cohort evidence")
    candidate_dashboard_json = candidate["dashboard_json"]
    try:
        dashboard = json.loads(candidate_dashboard_json)
        dashboard_exact = json.loads(
            candidate_dashboard_json,
            parse_float=Decimal,
            parse_int=Decimal,
        )
    except json.JSONDecodeError as error:
        raise SystemExit("V6 team candidate dashboard is not valid PostgreSQL JSON text") from error
    if not isinstance(dashboard, dict) or not isinstance(dashboard_exact, dict):
        raise SystemExit("V6 team candidate dashboard must be a JSON object")
    assert_v6_public_dashboard_contract(dashboard, "team dashboard")
    payload_sha256 = clean_text(candidate.get("payload_sha256"))
    if not re.fullmatch(r"[0-9a-f]{64}", payload_sha256):
        raise SystemExit("V6 team candidate lacks its canonical database payload hash")
    label = v6_team_release_label(
        team_key=team_key,
        curated_build_id=clean_text(candidate.get("curated_build_id")),
        payload_sha256=payload_sha256,
    )

    predecessors = candidate.get("approved_predecessors")
    if not isinstance(predecessors, list) or any(
        not isinstance(predecessor, dict) for predecessor in predecessors
    ):
        raise SystemExit("V6 team candidate has an invalid approved predecessor set")
    if len(predecessors) > 1:
        raise SystemExit("V6 team successor requires at most one approved predecessor")
    predecessor = predecessors[0] if predecessors else None

    if preflight:
        build_path_token = clean_text(candidate["curated_build_id"]).replace("-", "")
        output = Path(clean_text(getattr(args, "output", "") or "") or (
            f"data/reporting/urc_2025_26_{team_key}_{build_path_token}_{payload_sha256[:16]}_preflight.json"
        ))
        if Path("content/reporting").resolve() in output.resolve().parents:
            raise SystemExit("V6 team preflight output must stay outside content/reporting")
        write_text_atomic(output, candidate_dashboard_json + "\n")
        manifest = v6_team_preflight_manifest(
            team_key=team_key, contract=contract, candidate=candidate,
            predecessor=predecessor, preflight_file_sha256=sha256_file(output),
            provenance=provenance,
        )
        write_text_atomic(Path(f"{output}.manifest.json"), json.dumps(manifest, indent=2, sort_keys=True) + "\n")
        print(json.dumps({"status": "preflight", "output_path": str(output), "payload_sha256": payload_sha256}, indent=2))
        return

    reviewed_path = Path(preflight_file)
    reviewed, reviewed_exact, reviewed_sha256, reviewed_manifest_sha256 = read_v6_team_reviewed_preflight(
        reviewed_path=reviewed_path, team_key=team_key, contract=contract,
        candidate=candidate, predecessor=predecessor, provenance=provenance,
    )
    assert_v6_public_dashboard_contract(reviewed, "reviewed team dashboard")
    _, reviewed_diffs = compare_complete_public_payloads(
        reviewed_exact,
        dashboard_exact,
        reviewed_display=reviewed,
        candidate_display=dashboard,
    )
    if reviewed_diffs:
        raise SystemExit("reviewed V6 preflight does not match the exact active-build candidate")

    params = SqlParams()
    sql = f"""
      do $$ begin
        perform 1 from reporting.teams
        where team_key = {params.text(team_key)}
        for update;
        if not found then
          raise exception 'V6 team is no longer in the reporting roster';
        end if;
        perform 1 from curated.builds
        where id = {params.text(candidate['curated_build_id'])}::uuid
          and team_key = {params.text(team_key)} and season = '2025-26'
          and status = 'active'
        for update;
        if not found then
          raise exception 'active V6 build changed after review';
        end if;
        if exists (select 1 from reporting.aggregate_releases where release_label = {params.text(label)}) then
          raise exception 'immutable V6 team release already exists';
        end if;
        if not exists (
          select 1 from {contract.team_candidate_view}
          where team_key = {params.text(team_key)} and season = '2025-26'
            and curated_build_id = {params.text(candidate['curated_build_id'])}::uuid
            and analysis_version = 'v6'
            and classification_view_version = 'reporting_classification_2026-07-22_v2'
            and classification_evidence_sha256 = {params.text(candidate['classification_evidence_sha256'])}
            and cohort_view_version = 'analysis_window_2025-26_2026-08-15_v1'
            and cohort_evidence_sha256 = {params.text(candidate['cohort_evidence_sha256'])}
            and dashboard = {params.text(candidate_dashboard_json)}::jsonb
            and reporting.canonical_jsonb_sha256_v1(dashboard) = {params.text(payload_sha256)}
        ) then raise exception 'active V6 candidate changed after review'; end if;
        if (
          select count(*)
          from reporting.team_release_payloads_v6 payload
          join reporting.aggregate_releases release on release.id = payload.release_id
          where payload.team_key = {params.text(team_key)} and payload.season = '2025-26'
            and release.status = 'approved'
        ) <> (
          -- Keep CASE's THEN nested so PL/pgSQL does not terminate this IF early.
          case
            when {params.text(predecessor['release_id'] if predecessor else None)} is null then 0
            else 1
          end
        )
        or (
          {params.text(predecessor['release_id'] if predecessor else None)} is not null
          and not exists (
            select 1 from reporting.team_release_payloads_v6 payload
            join reporting.aggregate_releases release on release.id = payload.release_id
            where payload.release_id = {params.text(predecessor['release_id'] if predecessor else None)}::uuid
              and payload.team_key = {params.text(team_key)} and payload.season = '2025-26'
              and release.status = 'approved'
          )
        )
        then raise exception 'approved V6 team predecessor set changed after review'; end if;
      end $$;
      create temp table current_v6_team_release on commit drop as
      with run as (
        insert into audit.pipeline_runs (command, team, season, status, input_hash, output_hash, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values ('release_team_v6', {params.text(team)}, '2025-26', 'started', {params.text(candidate['curated_build_id'])}, {params.text(payload_sha256)},
          {params.jsonb({'team_key': team_key, 'release_tuple': contract.release_tuple, 'reviewer': reviewer, 'preflight_sha256': reviewed_sha256, 'reviewed_preflight_manifest_sha256': reviewed_manifest_sha256, 'predecessor_release_id': predecessor['release_id'] if predecessor else None, 'payload_hash_algorithm': 'postgres_jsonb_text_sha256'})}, null,
          {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}) returning id
      ), step as (
        insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count, counts_by_team)
        select id, 'release_team_v6', {params.text(contract.release_rule_version or 'league_dashboard_release_2026-08-15_v6')},
          'team_dashboard_release_v6', 1, 1, {params.jsonb({team_key: 1})} from run
        returning pipeline_run_id
      ), release as (
        insert into reporting.aggregate_releases (release_label, status, pipeline_run_id)
        select {params.text(label)}, 'draft', id from run returning id
      ) select id from release;
      insert into reporting.team_release_payloads_v6
        (release_id, team_key, season, curated_build_id, analysis_version,
         classification_view_version, classification_evidence_sha256,
         cohort_view_version, cohort_evidence_sha256, dashboard_payload, payload_sha256)
      select current.id, {params.text(team_key)}, '2025-26', {params.text(candidate['curated_build_id'])}::uuid, 'v6',
        'reporting_classification_2026-07-22_v2', {params.text(candidate['classification_evidence_sha256'])},
        'analysis_window_2025-26_2026-08-15_v1', {params.text(candidate['cohort_evidence_sha256'])},
        {params.text(candidate_dashboard_json)}::jsonb, {params.text(payload_sha256)}
      from current_v6_team_release current;
      update reporting.aggregate_releases release set status = 'retired'
      where release.id = {params.text(predecessor['release_id'] if predecessor else None)}::uuid
        and release.status = 'approved';
      update reporting.aggregate_releases release set status = 'approved', approved_at = now()
      from current_v6_team_release current where release.id = current.id and release.status = 'draft';
      update audit.pipeline_runs run set status = 'succeeded', ended_at = now()
      from current_v6_team_release current join reporting.aggregate_releases release on release.id = current.id
      where run.id = release.pipeline_run_id and release.status = 'approved';
      update audit.step_runs step set ended_at = now()
      from current_v6_team_release current join reporting.aggregate_releases release on release.id = current.id
      where step.pipeline_run_id = release.pipeline_run_id and step.step_name = 'release_team_v6';
      do $$ begin
        if not exists (
          select 1 from reporting.team_release_payloads_v6 payload
          join current_v6_team_release current on current.id = payload.release_id
          where payload.payload_sha256 = {params.text(payload_sha256)}
            and payload.payload_sha256 = reporting.canonical_jsonb_sha256_v1(payload.dashboard_payload)
        ) then raise exception 'stored V6 team payload hash differs from canonical reviewed bytes'; end if;
      end $$;
    """
    run_sql(sql, params.values)
    print(json.dumps({"status": "approved", "release_label": label, "payload_sha256": payload_sha256}, indent=2))


def release(args: argparse.Namespace) -> None:
    """Phase 4 release: reads analysis.*_v1 views directly (no JSON/CSV
    input), snapshots every dashboard section into reporting.release_context
    / reporting.release_table_rows in one transaction as a recorded pipeline
    run, and exports content/reporting/<team_key>_dashboard_<season>.json
    from that DB snapshot with internal keys stripped (never populated in
    the first place, unlike the old JSON-file-driven release()).
    """
    if clean_text(getattr(args, "season", "")) == "2025-26":
        release_team_v6(args)
        return
    team = clean_text(args.team)
    season = clean_text(args.season)
    if not team or not season:
        raise SystemExit("--team and --season are required")
    preflight = bool(getattr(args, "preflight", False))
    preflight_file_arg = clean_text(getattr(args, "preflight_file", "") or "")
    preflight_reviewer = clean_text(getattr(args, "preflight_reviewer", "") or "")
    previous_dashboard_file_arg = clean_text(getattr(args, "previous_dashboard_file", "") or "")
    restatement_file_arg = clean_text(getattr(args, "restatement_file", "") or "")
    if preflight and preflight_file_arg:
        raise SystemExit("--preflight cannot be combined with --preflight-file")
    if preflight_file_arg and previous_dashboard_file_arg:
        raise SystemExit("--preflight-file and --previous-dashboard-file cannot be used together")
    if restatement_file_arg and (preflight or not previous_dashboard_file_arg):
        raise SystemExit("--restatement-file requires a non-preflight re-release with --previous-dashboard-file")
    if preflight_file_arg and not preflight_reviewer:
        raise SystemExit("--preflight-reviewer is required with --preflight-file")
    if preflight_reviewer and not preflight_file_arg:
        raise SystemExit("--preflight-reviewer requires --preflight-file")

    restatement_path: Path | None = None
    cached_restatement: dict[str, Any] | None = None
    cached_restatement_sha256: str | None = None
    if restatement_file_arg:
        restatement_path = Path(restatement_file_arg)
        if not restatement_path.exists():
            raise SystemExit(f"release restatement not found: {restatement_path}")
        restatement_bytes = restatement_path.read_bytes()
        cached_restatement_sha256 = hashlib.sha256(restatement_bytes).hexdigest()
        try:
            cached_restatement = json.loads(restatement_bytes)
        except json.JSONDecodeError as exc:
            raise SystemExit(f"release restatement must be valid JSON: {restatement_path}") from exc
        if not isinstance(cached_restatement, dict):
            raise SystemExit("release restatement must be a JSON object")

    # Cache and parse a supplied historical artifact before the command makes
    # even its first database query. Whether it is required is decided from
    # accepted full-release history below; no file read can race with a later
    # overwrite once release evaluation has started.
    previous_dashboard_path: Path | None = None
    previous_dashboard: dict[str, Any] | None = None
    previous_dashboard_sha256: str | None = None
    if previous_dashboard_file_arg:
        previous_dashboard_path = Path(previous_dashboard_file_arg)
        if not previous_dashboard_path.exists():
            raise SystemExit(f"previous dashboard JSON not found: {previous_dashboard_path}")
        previous_dashboard_bytes = previous_dashboard_path.read_bytes()
        try:
            previous_dashboard = json.loads(previous_dashboard_bytes)
        except json.JSONDecodeError as exc:
            raise SystemExit(
                f"previous dashboard JSON is invalid: {previous_dashboard_path}: {exc}"
            ) from exc
        if not isinstance(previous_dashboard, dict):
            raise SystemExit(f"previous dashboard JSON must be an object: {previous_dashboard_path}")
        previous_dashboard_sha256 = hashlib.sha256(previous_dashboard_bytes).hexdigest()

    provenance = run_provenance()
    if provenance["code_version"].endswith("-dirty"):
        raise SystemExit(
            "release refuses to run from an uncommitted working tree "
            f"(code_version={provenance['code_version']}); commit or stash local "
            "changes before releasing so the release row records an exact, "
            "reproducible commit"
        )

    team_key = resolve_team_key(team)

    required_migrations = [
        CURATED_LAYER_MIGRATION_VERSION,
        ANALYSIS_VIEWS_MIGRATION_VERSION,
        INJURY_COHORT_V1_AMENDMENT_MIGRATION_VERSION,
        RELEASE_DASHBOARDS_MIGRATION_VERSION,
    ]
    migration_in_list = ", ".join(f"'{version}'" for version in required_migrations)
    migration_rows = query_sql(
        f"select version from supabase_migrations.schema_migrations where version in ({migration_in_list})"
    )
    applied = {row["version"] for row in migration_rows}
    missing = sorted(set(required_migrations) - applied)
    if missing:
        raise SystemExit(f"release requires migrations {', '.join(missing)} to be applied and tracked first")

    history_params = SqlParams()
    history_rows = query_sql(
        f"""
        select
          r.id as release_id,
          r.release_label,
          r.status as release_status,
          count(*) over ()::int as release_count
        from reporting.release_context rc
        join reporting.aggregate_releases r on r.id = rc.release_id
        join audit.pipeline_runs pr on pr.id = r.pipeline_run_id
        where rc.team_key = {history_params.text(team_key)}
          and rc.season = {history_params.text(season)}
          and (
            r.status = 'approved'
            or (r.status = 'retired' and pr.status = 'succeeded')
          )
        order by
          case when r.status = 'approved' then 0 else 1 end,
          r.approved_at desc nulls last,
          r.created_at desc,
          r.id desc
        limit 1
        """,
        history_params.values,
    )
    prior_release_count = int(history_rows[0]["release_count"]) if history_rows else 0
    is_first_release = prior_release_count == 0
    previous_release_label = history_rows[0]["release_label"] if history_rows else None
    previous_release_id = history_rows[0]["release_id"] if history_rows else None
    previous_release_status = history_rows[0]["release_status"] if history_rows else None
    if is_first_release and previous_dashboard is not None:
        raise SystemExit(
            "--previous-dashboard-file is only valid for a re-release; first releases require "
            "--preflight-file and --preflight-reviewer"
        )
    if not is_first_release and preflight_file_arg:
        raise SystemExit(
            "--preflight-file is only valid for a first release; re-releases require "
            "--previous-dashboard-file"
        )
    if not is_first_release:
        if previous_dashboard is None or previous_release_label is None:
            raise SystemExit(
                "re-release for this team/season requires --previous-dashboard-file pointing to the "
                "latest accepted dashboard snapshot"
            )
        approved_previous_dashboard = export_release_dashboard_json(previous_release_label)
        approved_binding_diffs = diff_json_documents(previous_dashboard, approved_previous_dashboard)
        if approved_binding_diffs:
            paths = ", ".join(diff["path"] for diff in approved_binding_diffs[:10])
            raise SystemExit(
                "release refuses to run: --previous-dashboard-file does not exactly match the latest "
                f"accepted full release ({paths})"
            )

    # --- Active curated build, and a freshness gate (strengthens the old
    # file-hash gate: refuses to release from a curated build that is stale
    # against the latest processing.record_versions). ----------------------
    build_params = SqlParams()
    build_rows = query_sql(
        f"""
        select id, source_version_set_hash
        from curated.builds
        where team_key = {build_params.text(team_key)} and season = {build_params.text(season)} and status = 'active'
        """,
        build_params.values,
    )
    if len(build_rows) != 1:
        raise SystemExit(
            f"release requires exactly one active curated.builds row for team_key={team_key!r} "
            f"season={season!r}, found {len(build_rows)}; run build-curated first"
        )
    curated_build_id = build_rows[0]["id"]
    injury_ids = latest_curated_source_version_ids(team, season, "%injury%")
    exposure_ids = latest_curated_source_version_ids(team, season, "%exposure%")
    fresh_hash = curated_source_version_set_hash(injury_ids, exposure_ids)
    if fresh_hash != build_rows[0]["source_version_set_hash"]:
        raise SystemExit(
            "release refuses to run: the active curated build is stale against the latest "
            "processing.record_versions for this team/season; run build-curated --rebuild first"
        )

    # --- Duplicate-adjudication consistency (strengthens the old file-based
    # gate: every adjudicated exclude_duplicate decision for this team/season
    # must already be reflected in the active curated build's eligibility_status). --
    dup_params = SqlParams()
    unreflected_dups = query_sql(
        f"""
        select sr.source_row_number
        from audit.adjudications adj
        join ingestion.source_rows sr on sr.id = adj.source_row_id
        join ingestion.source_files sf on sf.id = sr.source_file_id
        join reporting.team_key_aliases a on a.alias = sf.team
        where a.team_key = {dup_params.text(team_key)}
          and sf.season = {dup_params.text(season)}
          and adj.decision ->> 'decision' = 'exclude_duplicate'
          and not exists (
            select 1 from curated.injuries i
            where i.source_row_id = adj.source_row_id
              and i.curated_build_id = {dup_params.text(curated_build_id)}::uuid
              and i.eligibility_status = 'excluded_duplicate_adjudicated'
          )
        order by sr.source_row_number
        """,
        dup_params.values,
    )
    if unreflected_dups:
        rows = ", ".join(str(row["source_row_number"]) for row in unreflected_dups)
        raise SystemExit(
            f"release refuses to run: adjudicated duplicate exclusion(s) for source row(s) {rows} "
            "are not reflected in the active curated build; rebuild curated data first"
        )

    # --- Read analysis.*_v1 views (read-only; the same seven queries
    # verify-analysis-parity uses, plus the exposure coverage window). -----
    headline_rows = fetch_team_season_rows(
        "select * from analysis.headline_metrics_v1 where team_key = {team_key} and season = {season}",
        team_key, season,
    )
    if len(headline_rows) != 1:
        raise SystemExit(
            f"expected exactly one analysis.headline_metrics_v1 row for team_key={team_key!r} "
            f"season={season!r}, found {len(headline_rows)}"
        )
    coverage_rows = fetch_team_season_rows(
        "select * from analysis.coverage_v1 where team_key = {team_key} and season = {season}",
        team_key, season,
    )
    if len(coverage_rows) != 1:
        raise SystemExit(
            f"expected exactly one analysis.coverage_v1 row for team_key={team_key!r} season={season!r}, "
            f"found {len(coverage_rows)}"
        )
    setting_rows = fetch_team_season_rows(
        "select * from analysis.setting_split_v1 where team_key = {team_key} and season = {season} "
        "order by time_loss_injuries desc, days_lost desc, label asc",
        team_key, season,
    )
    monthly_rows = fetch_team_season_rows(
        "select *, month_start::text as month_start_text from analysis.monthly_v1 "
        "where team_key = {team_key} and season = {season} order by month_start",
        team_key, season,
    )
    body_location_rows = fetch_team_season_rows(
        "select * from analysis.body_locations_v1 where team_key = {team_key} and season = {season} order by rank",
        team_key, season,
    )
    injury_type_rows = fetch_team_season_rows(
        "select * from analysis.injury_types_v1 where team_key = {team_key} and season = {season} order by rank",
        team_key, season,
    )
    severity_rows = fetch_team_season_rows(
        "select * from analysis.severity_distribution_v1 where team_key = {team_key} and season = {season} "
        "order by band_order",
        team_key, season,
    )
    window_rows = fetch_team_season_rows(
        "select distinct coverage_start::text as coverage_start, coverage_end::text as coverage_end "
        "from analysis.injury_cohort_v1 where team_key = {team_key} and season = {season}",
        team_key, season,
    )
    if len(window_rows) != 1:
        raise SystemExit(
            f"expected exactly one distinct exposure coverage window for team_key={team_key!r} "
            f"season={season!r} in analysis.injury_cohort_v1, found {len(window_rows)}; "
            "is there at least one injury in the cohort?"
        )

    rendered = render_analysis_dashboard_sections(
        headline_row=headline_rows[0],
        setting_rows=setting_rows,
        monthly_rows=monthly_rows,
        body_location_rows=body_location_rows,
        injury_type_rows=injury_type_rows,
        severity_rows=severity_rows,
        coverage_row=coverage_rows[0],
    )

    # --- Released-rows-vs-views consistency check: every rendered section
    # must reconcile against the headline totals it is supposed to sum to.
    # This exercises the SAME rendered rows that will be written, so it also
    # proves render_analysis_dashboard_sections() (Phase 3.2's own parity
    # renderer, reused unmodified here) did not drop or duplicate rows
    # before they are ever written. -----------------------------------------
    headline_time_loss = next(m["value"] for m in rendered["headline"] if m["key"] == "time_loss_injuries")
    headline_recorded = next(m["value"] for m in rendered["headline"] if m["key"] == "recorded_injuries")
    monthly_time_loss_sum = sum(row["time_loss_injuries"] for row in rendered["monthly"])
    severity_time_loss_sum = sum(row["time_loss_injuries"] for row in rendered["severity_distribution"])
    severity_recorded_sum = sum(row["recorded_injuries"] for row in rendered["severity_distribution"])
    if monthly_time_loss_sum != headline_time_loss:
        raise SystemExit(
            f"release refuses to run: analysis.monthly_v1 time_loss_injuries sum ({monthly_time_loss_sum}) "
            f"does not reconcile with analysis.headline_metrics_v1 time_loss_injuries ({headline_time_loss})"
        )
    if severity_time_loss_sum != headline_time_loss:
        raise SystemExit(
            f"release refuses to run: analysis.severity_distribution_v1 time_loss_injuries sum "
            f"({severity_time_loss_sum}) does not reconcile with headline time_loss_injuries ({headline_time_loss})"
        )
    if severity_recorded_sum != headline_recorded:
        raise SystemExit(
            f"release refuses to run: analysis.severity_distribution_v1 recorded_injuries sum "
            f"({severity_recorded_sum}) does not reconcile with headline recorded_injuries ({headline_recorded})"
        )
    # setting_split is time-loss-only by construction but is not required to
    # sum to the full time-loss total (a row can carry the "unknown" setting
    # label), so it is intentionally not cross-checked against the headline
    # total here -- matches the dashboard's own documented behaviour.

    # --- Cohort filter block (Adjudication 2: from audit evidence, outside
    # analysis views). -------------------------------------------------------
    cohort_filters = release_cohort_filter_flags(team_key, season)
    team_display_name = team_display_name_for(team_key)

    grain = clean_text(coverage_rows[0].get("exposure_grain")) or "unknown"
    grain_label = DASHBOARD_EXPORT_GRAIN_LABELS.get(grain, "unknown-grain")
    monthly_basis = (
        f"Monthly exposure is assigned to the week-start month because {team_display_name} reports weekly exposure."
        if grain == "weekly"
        else "Monthly exposure is assigned to the cleaned exposure-date month."
    )
    setting_limitation = (
        f"{team_display_name} exposure is weekly, so match and training incidence cannot be split until "
        "setting-specific exposure denominators are approved."
        if grain == "weekly"
        else "Setting-specific rates are not split until setting-specific exposure denominators are approved."
    )
    adjudicated_duplicate_rows = adjudicated_duplicate_row_numbers(team_key, season)
    adjudicated_duplicate_text = ", ".join(str(number) for number in adjudicated_duplicate_rows)

    method = [
        "Headline injury metrics use time-loss injuries only: Days Injured > 0.",
        "Incidence = time-loss injuries / exposure hours * 1000.",
        "Severity = mean and median days lost per time-loss injury.",
        "Burden = days lost / exposure hours * 1000.",
        f"Exposure hours = sum(minutes_total_clean) / 60 for included {grain_label} exposure rows.",
        monthly_basis,
        "IOC-aligned body-location, tissue/pathology, and severity labels come from the accepted V2 mapping.",
        (
            "Received/Injured In Team retains the approved team plus blank/N/A values; explicit other-team "
            "or Club values are excluded."
            if cohort_filters["injured_in_team_applied"]
            else "No Received/Injured In Team cohort filter applied."
        ),
        "Match Type retains URC, training, Other, blank/N/A, and generic match/game values; explicit non-URC "
        "competitions and teams are excluded.",
        (
            f"Adjudicated duplicate standardised rows excluded from this aggregate: {adjudicated_duplicate_text}."
            if adjudicated_duplicate_rows
            else "No adjudicated duplicate injury rows were excluded from this aggregate."
        ),
    ]
    limitations = [
        setting_limitation,
        (
            f"Adjudicated duplicate standardised row exclusions applied: {adjudicated_duplicate_text}."
            if adjudicated_duplicate_rows
            else "No adjudicated duplicate injury row exclusions applied."
        ),
        "Aggregate release is approved only after explicit confirmation of the live Supabase target; the "
        "dashboard reports aggregate values only.",
    ]
    prior_season = {
        "season": "2023-24",
        "status": "not_loaded_in_v2",
        "note": (
            f"No {team_display_name} prior-season injury and exposure denominator pair exists in this V2 "
            "workspace, so the dashboard leaves comparison blank rather than mixing in legacy report figures."
        ),
    }
    coverage = {**rendered["coverage"], "injury_cohort_filters": cohort_filters}

    generated_at_iso = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    analysis_window_start = window_rows[0]["coverage_start"]
    analysis_window_end = window_rows[0]["coverage_end"]
    analysis_window_basis = (
        f"{team_display_name} exposure coverage window from included {grain_label} exposure rows"
    )

    table_rows = render_release_table_rows(rendered, monthly_rows)
    counts_by_section: dict[str, int] = defaultdict(int)
    for row in table_rows:
        counts_by_section[row["section"]] += 1

    release_content_hash = sha256_json(
        {
            "context": {
                "team_key": team_key, "season": season, "method": method, "coverage": coverage,
                "prior_season": prior_season, "limitations": limitations,
                "analysis_window": {
                    "start": analysis_window_start, "end": analysis_window_end, "basis": analysis_window_basis,
                },
            },
            "rows": table_rows,
        }
    )
    # Each attempt gets an immutable label. A failed draft is retired rather
    # than deleted, so a nonce is required to make an identical retry safe.
    label = f"{team_key}-{season}-{release_content_hash[:12]}-{uuid.uuid4().hex[:12]}"

    context_record = {
        "team_key": team_key, "season": season, "team_display_name": team_display_name,
        "curated_build_id": curated_build_id, "analysis_view_version": ANALYSIS_VIEW_VERSION_SUFFIX,
        "generated_at": generated_at_iso,
        "analysis_window_start": analysis_window_start, "analysis_window_end": analysis_window_end,
        "analysis_window_basis": analysis_window_basis,
        "method": method, "coverage": coverage, "injury_cohort_filters": cohort_filters,
        "prior_season": prior_season, "limitations": limitations,
    }

    candidate_dashboard = assemble_release_dashboard(context_record, table_rows)
    historical_diff = None
    historical_blocked: list[dict[str, Any]] = []
    if previous_dashboard is not None:
        historical_diffs = diff_json_documents(previous_dashboard, candidate_dashboard)
        historical_allowed, historical_blocked = classify_historical_release_diffs(historical_diffs)
        historical_diff = {
            "overall": "BLOCKED" if historical_blocked else "ALLOWED_ONLY",
            "allowed": len(historical_allowed),
            "blocked": len(historical_blocked),
            "allowed_paths": [diff["path"] for diff in historical_allowed],
            "blocked_paths": [diff["path"] for diff in historical_blocked],
            "blocked_diffs_sha256": sha256_json(historical_blocked),
        }
    output_arg = clean_text(getattr(args, "output", "") or "")
    if preflight:
        # Run the final protected-alias gate under the read-only query runner.
        # Every operation above is also query-only; this branch exits before
        # SqlParams for inserts are built and before run_sql() is reachable.
        query_sql(protected_alias_scan_sql("release preflight gate"))
        preflight_path = (
            Path(output_arg)
            if output_arg
            else REPO_ROOT / "data" / "reporting" /
            f"{team_key}_dashboard_{season}_{release_content_hash[:12]}_preflight.json"
        )
        resolved_preflight_path = preflight_path.resolve()
        public_reporting_dir = (REPO_ROOT / "content" / "reporting").resolve()
        if resolved_preflight_path.is_relative_to(public_reporting_dir):
            raise SystemExit(
                "release preflight refuses to write under content/reporting; "
                "use the default Git-ignored data/reporting candidate path"
            )
        write_json_atomic(preflight_path, candidate_dashboard)
        print(
            json.dumps(
                {
                    "mode": "preflight",
                    "database_writes": 0,
                    "release_content_hash": release_content_hash,
                    "team_key": team_key,
                    "season": season,
                    "first_release": is_first_release,
                    "prior_release_count": prior_release_count,
                    "curated_build_id": curated_build_id,
                    "rows_candidate": len(table_rows),
                    "counts_by_section": dict(counts_by_section),
                    "historical_diff": historical_diff,
                    "candidate_path": str(preflight_path),
                    "candidate_sha256": sha256_file(preflight_path),
                    "next": (
                        "bind the exact blocked diff to an approved restatement, then rerun release with "
                        "--previous-dashboard-file and --restatement-file"
                        if historical_blocked
                        else "review and sign off this candidate before running release"
                    ),
                },
                indent=2,
            )
        )
        return

    export_path = (
        Path(output_arg)
        if output_arg
        else Path("content") / "reporting" / f"{team_key}_dashboard_{season}.json"
    )
    reviewed_preflight_path: Path | None = None
    reviewed_candidate: dict[str, Any] | None = None
    reviewed_preflight_sha256: str | None = None
    release_restatement: dict[str, Any] | None = None
    if preflight_file_arg:
        reviewed_preflight_path = Path(preflight_file_arg)
        if not reviewed_preflight_path.exists():
            raise SystemExit(f"reviewed preflight candidate not found: {reviewed_preflight_path}")
        if reviewed_preflight_path.resolve() == export_path.resolve():
            raise SystemExit("--output and --preflight-file must resolve to different files")
        reviewed_preflight_bytes = reviewed_preflight_path.read_bytes()
        try:
            reviewed_candidate = json.loads(reviewed_preflight_bytes)
        except json.JSONDecodeError as exc:
            raise SystemExit(
                f"reviewed preflight candidate is invalid JSON: {reviewed_preflight_path}: {exc}"
            ) from exc
        if not isinstance(reviewed_candidate, dict):
            raise SystemExit(
                f"reviewed preflight candidate must be a JSON object: {reviewed_preflight_path}"
            )
        reviewed_preflight_sha256 = hashlib.sha256(reviewed_preflight_bytes).hexdigest()
        prewrite_diffs = diff_json_documents(reviewed_candidate, candidate_dashboard)
        _, prewrite_blocked = classify_preflight_release_diffs(prewrite_diffs)
        if prewrite_blocked:
            paths = ", ".join(diff["path"] for diff in prewrite_blocked[:10])
            raise SystemExit(
                "release refuses to run: current analysis/audit evidence differs from the reviewed "
                f"preflight candidate outside generated_at ({paths}); generate and review a new preflight"
            )
    elif is_first_release:
        raise SystemExit(
                "first release for this team/season requires a reviewed preflight candidate; run release "
                "--preflight, review/sign off the Git-ignored candidate, then rerun release with --preflight-file"
            )
    else:
        if previous_dashboard is None or previous_dashboard_path is None:
            raise AssertionError("re-release previous-dashboard gate was not established before analysis")
        if previous_dashboard_path.resolve() == export_path.resolve():
            raise SystemExit(
                "--output and --previous-dashboard-file must resolve to different files; snapshot the "
                "previously approved dashboard before re-release"
            )
        if historical_blocked:
            if not restatement_file_arg or previous_dashboard_sha256 is None:
                paths = ", ".join(diff["path"] for diff in historical_blocked[:10])
                raise SystemExit(
                    "release refuses to run: current candidate differs from the previous approved dashboard "
                    f"outside the historical whitelist ({paths}); an exact approved --restatement-file is required"
                )
            if cached_restatement is None:
                raise AssertionError("restatement bytes were not cached before release evaluation")
            release_restatement = validate_release_restatement_envelope(
                cached_restatement,
                team_key=team_key,
                season=season,
                previous_dashboard_sha256=previous_dashboard_sha256,
                release_content_hash=release_content_hash,
                blocked_diffs=historical_blocked,
            )
            historical_diff["overall"] = "APPROVED_RESTATEMENT"
        elif restatement_file_arg:
            raise SystemExit("--restatement-file is unnecessary because the re-release has no blocked drift")

    release_parameters = {
        "release": label,
        "team_key": team_key,
        "curated_build_id": curated_build_id,
        "analysis_view_version": ANALYSIS_VIEW_VERSION_SUFFIX,
        "first_release": is_first_release,
    }
    if reviewed_preflight_sha256 is not None:
        release_parameters["reviewed_preflight_sha256"] = reviewed_preflight_sha256
        release_parameters["preflight_reviewer"] = preflight_reviewer
    if previous_dashboard_sha256 is not None:
        release_parameters["previous_dashboard_sha256"] = previous_dashboard_sha256
        release_parameters["previous_release_id"] = previous_release_id
        release_parameters["previous_release_label"] = previous_release_label
        release_parameters["previous_release_status"] = previous_release_status
    if release_restatement is not None:
        release_parameters["release_restatement"] = release_restatement
        release_parameters["restatement_file_sha256"] = cached_restatement_sha256

    params = SqlParams()
    context_insert = f"""
      insert into reporting.release_context
        (release_id, team_key, season, team_display_name, curated_build_id, analysis_view_version,
         generated_at, analysis_window_start, analysis_window_end, analysis_window_basis,
         method, coverage, injury_cohort_filters, prior_season, limitations)
      select
        current_release.id, ctx.team_key, ctx.season, ctx.team_display_name, ctx.curated_build_id::uuid,
        ctx.analysis_view_version, ctx.generated_at::timestamptz, ctx.analysis_window_start::date,
        ctx.analysis_window_end::date, ctx.analysis_window_basis, ctx.method, ctx.coverage,
        ctx.injury_cohort_filters, ctx.prior_season, ctx.limitations
      from current_release,
        jsonb_to_recordset({params.jsonb([context_record])}) as ctx(
          team_key text, season text, team_display_name text, curated_build_id text,
          analysis_view_version text, generated_at text, analysis_window_start text,
          analysis_window_end text, analysis_window_basis text, method jsonb, coverage jsonb,
          injury_cohort_filters jsonb, prior_season jsonb, limitations jsonb
        );
    """
    rows_insert = f"""
      insert into reporting.release_table_rows
        (release_id, team_key, season, section, row_key, ordinal, label, month, value, numerator,
         denominator, unit, formula, exposure_hours, distance_km, time_loss_injuries, recorded_injuries,
         days_lost, incidence_per_1000h, burden_per_1000h, mean_severity_days)
      select
        current_release.id, {params.text(team_key)}, {params.text(season)},
        row.section, row.row_key, row.ordinal, row.label, row.month, row.value, row.numerator,
        row.denominator, row.unit, row.formula, row.exposure_hours, row.distance_km,
        row.time_loss_injuries, row.recorded_injuries, row.days_lost, row.incidence_per_1000h,
        row.burden_per_1000h, row.mean_severity_days
      from current_release,
        jsonb_to_recordset({params.jsonb(table_rows)}) as row(
          section text, row_key text, ordinal int, label text, month text, value numeric,
          numerator numeric, denominator numeric, unit text, formula text, exposure_hours numeric,
          distance_km numeric, time_loss_injuries numeric, recorded_injuries numeric, days_lost numeric,
          incidence_per_1000h numeric, burden_per_1000h numeric, mean_severity_days numeric
        );
    """
    sql = f"""
      {protected_alias_scan_sql('release gate')}

      do $$
      begin
        if not exists (select 1 from curated.builds where id = {params.text(curated_build_id)}::uuid and status = 'active') then
          raise exception 'release refuses to run: the active curated build changed since it was read; rerun release';
        end if;
        if exists (select 1 from reporting.aggregate_releases where release_label = {params.text(label)}) then
          raise exception 'immutable release already exists';
        end if;
      end $$;

      create temp table current_release on commit drop as
      with run as (
        insert into audit.pipeline_runs (command, team, season, status, input_hash, output_hash, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values ('release', {params.text(team)}, {params.text(season)}, 'started',
          {params.text(curated_build_id)}, {params.text(release_content_hash)},
          {params.jsonb(release_parameters)}, null,
          {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])})
        returning id
      ),
      step as (
        insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count, counts_by_team)
        select id, 'release_full_dashboard', {params.text(FULL_DASHBOARD_RELEASE_RULE_VERSION)}, 'full_dashboard_release',
          {len(table_rows)}, {len(table_rows)}, {params.jsonb(dict(counts_by_section))}
        from run
        returning pipeline_run_id
      ),
      release as (
        insert into reporting.aggregate_releases (release_label, status, pipeline_run_id, approved_at)
        select {params.text(label)}, 'draft', id, null
        from run
        returning id
      )
      select id from release;

      {context_insert}

      {rows_insert}
    """
    expected_counts = {
        "headline_n": counts_by_section.get("headline", 0),
        "setting_split_n": counts_by_section.get("setting_split", 0),
        "monthly_n": counts_by_section.get("monthly", 0),
        "body_locations_n": counts_by_section.get("body_locations", 0),
        "injury_types_n": counts_by_section.get("injury_types", 0),
        "severity_distribution_n": counts_by_section.get("severity_distribution", 0),
    }
    preflight_diff = None
    draft_created = False
    release_stage = "draft_insert"
    try:
        # Transaction 1 creates an invisible draft snapshot and leaves its
        # pipeline run/step open. Draft rows cannot appear in the consumer
        # view, which filters aggregate_releases to status='approved'.
        run_sql(sql, params.values)
        draft_created = True
        release_stage = "draft_verify"

        draft_params = SqlParams()
        draft_rows = query_sql(
            f"""
            select
              r.status as release_status,
              pr.status as pipeline_run_status,
              count(distinct rc.id)::int as context_n,
              count(rt.id) filter (where rt.section = 'headline')::int as headline_n,
              count(rt.id) filter (where rt.section = 'setting_split')::int as setting_split_n,
              count(rt.id) filter (where rt.section = 'monthly')::int as monthly_n,
              count(rt.id) filter (where rt.section = 'body_locations')::int as body_locations_n,
              count(rt.id) filter (where rt.section = 'injury_types')::int as injury_types_n,
              count(rt.id) filter (where rt.section = 'severity_distribution')::int as severity_distribution_n
            from reporting.aggregate_releases r
            join audit.pipeline_runs pr on pr.id = r.pipeline_run_id
            left join reporting.release_context rc on rc.release_id = r.id
            left join reporting.release_table_rows rt on rt.release_id = r.id
            where r.release_label = {draft_params.text(label)}
            group by r.status, pr.status
            """,
            draft_params.values,
        )
        if len(draft_rows) != 1:
            raise SystemExit(f"draft verification found {len(draft_rows)} release rows, expected 1")
        draft_verify = draft_rows[0]
        if draft_verify["release_status"] != "draft" or draft_verify["pipeline_run_status"] != "started":
            raise SystemExit(
                "draft verification expected release status=draft and pipeline run status=started, "
                f"found {draft_verify['release_status']!r}/{draft_verify['pipeline_run_status']!r}"
            )
        if int(draft_verify["context_n"]) != 1:
            raise SystemExit(
                f"draft verification expected one release_context row, found {draft_verify['context_n']!r}"
            )
        for key, expected in expected_counts.items():
            if int(draft_verify[key]) != expected:
                raise SystemExit(
                    f"draft verification failed: release_table_rows {key}={draft_verify[key]!r}, expected {expected}"
                )

        dashboard_json = export_release_dashboard_json(label)
        serialization_diffs = diff_json_documents(candidate_dashboard, dashboard_json)
        if serialization_diffs:
            paths = ", ".join(diff["path"] for diff in serialization_diffs[:10])
            raise SystemExit(f"draft serialization differs from the assembled candidate ({paths})")

        if reviewed_candidate is not None:
            reviewed_diffs = diff_json_documents(reviewed_candidate, dashboard_json)
            allowed_reviewed, blocked_reviewed = classify_preflight_release_diffs(reviewed_diffs)
            preflight_diff = {
                "overall": "BLOCKED" if blocked_reviewed else "ALLOWED_ONLY",
                "allowed": len(allowed_reviewed),
                "blocked": len(blocked_reviewed),
                "allowed_paths": [diff["path"] for diff in allowed_reviewed],
                "blocked_paths": [diff["path"] for diff in blocked_reviewed],
            }
            if blocked_reviewed:
                paths = ", ".join(diff["path"] for diff in blocked_reviewed[:10])
                raise SystemExit(
                    "draft release differs from the reviewed preflight candidate outside generated_at "
                    f"({paths})"
                )

        if previous_dashboard is not None:
            historical_diffs = diff_json_documents(previous_dashboard, dashboard_json)
            historical_allowed, historical_blocked = classify_historical_release_diffs(historical_diffs)
            historical_diff = {
                "overall": "BLOCKED" if historical_blocked else "ALLOWED_ONLY",
                "allowed": len(historical_allowed),
                "blocked": len(historical_blocked),
                "allowed_paths": [diff["path"] for diff in historical_allowed],
                "blocked_paths": [diff["path"] for diff in historical_blocked],
            }
            if historical_blocked and release_restatement is not None and previous_dashboard_sha256 is not None:
                if cached_restatement is None:
                    raise AssertionError("restatement bytes were not cached before draft verification")
                validate_release_restatement_envelope(
                    cached_restatement,
                    team_key=team_key,
                    season=season,
                    previous_dashboard_sha256=previous_dashboard_sha256,
                    release_content_hash=release_content_hash,
                    blocked_diffs=historical_blocked,
                )
                historical_diff["overall"] = "APPROVED_RESTATEMENT"
            elif historical_blocked:
                paths = ", ".join(diff["path"] for diff in historical_blocked[:10])
                raise SystemExit(
                    "draft release differs from the previous approved dashboard outside the historical "
                    f"whitelist ({paths})"
                )

        # Transaction 2 promotes only this exact, still-draft snapshot and
        # reconstructs its exact public JSON through the consumer view inside
        # that same transaction. Approval and audit success cannot commit
        # unless the consumer payload equals this verified draft candidate.
        release_stage = "promote"
        execute_release_promotion(
            label,
            curated_build_id,
            team_key,
            season,
            dashboard_json,
            previous_release_id,
        )

        release_stage = "export"
        write_json_atomic(export_path, dashboard_json)
    except BaseException:
        if draft_created:
            cleanup_sql, cleanup_values = release_failure_cleanup_statement(
                label,
                team_key,
                season,
                previous_release_id,
                previous_release_status,
                release_stage,
            )
            try:
                run_sql(cleanup_sql, cleanup_values)
            except BaseException as cleanup_error:
                print(
                    f"WARNING: failed to retire release attempt {label!r} after {release_stage}: {cleanup_error}",
                    file=sys.stderr,
                )
        raise

    print(
        json.dumps(
            {
                "release_label": label,
                "team_key": team_key,
                "season": season,
                "curated_build_id": curated_build_id,
                "rows_written": len(table_rows),
                "counts_by_section": dict(counts_by_section),
                "export_path": str(export_path),
                "preflight_diff": preflight_diff,
                "historical_diff": historical_diff,
            },
            indent=2,
        )
    )


def current_league_bundle_snapshot(season: str) -> tuple[dict[str, Any], dict[str, Any]]:
    """Read the exact latest approved immutable bundle using one SQL statement."""
    if season == "2025-26":
        bundle_view = "reporting.latest_approved_league_bundle_v6"
        league_payload_view = "reporting.league_release_payloads_v6"
        team_payload_view = "reporting.team_dashboard_payloads_v2"
        team_release_column = "bundle_release_id"
    else:
        bundle_view = "reporting.latest_approved_dashboard_bundle_v4"
        league_payload_view = "reporting.dashboard_bundle_league_payloads_v1"
        team_payload_view = "reporting.dashboard_bundle_team_payloads_v1"
        team_release_column = "bundle_release_id"
    params = SqlParams()
    rows = query_sql(
        f"""
        with current_bundle as (
          -- The fixed season-specific reader is the same completeness boundary
          -- the website uses, never a looser "newest approved" rule.
          select b.release_id, b.season, r.release_label, r.approved_at
          from {bundle_view} b
          join reporting.aggregate_releases r on r.id = b.release_id
          where b.season = {params.text(season)}
        )
        select b.release_id::text, b.release_label, b.approved_at,
          league.dashboard_payload as league,
          coalesce(jsonb_agg(jsonb_build_object(
            'team_key', team.team_key, 'dashboard', team.dashboard_payload
          ) order by team.team_key), '[]'::jsonb) as teams
        from current_bundle b
        join {league_payload_view} league
          on league.release_id = b.release_id
        join {team_payload_view} team
          on team.{team_release_column} = b.release_id
        group by b.release_id, b.release_label, b.approved_at, league.dashboard_payload
        """,
        params.values,
    )
    if len(rows) != 1 or not isinstance(rows[0].get("league"), dict):
        raise SystemExit(f"no complete approved dashboard bundle exists for {season}")
    if not isinstance(rows[0].get("teams"), list) or len(rows[0]["teams"]) != 16:
        raise SystemExit("latest approved dashboard bundle is not a complete 16-team snapshot")
    bundle = {
        "schema_version": "urc_dashboard_bundle_v2",
        "season": season,
        "league": rows[0]["league"],
        "teams": rows[0]["teams"],
    }
    metadata = {
        "release_id": rows[0]["release_id"],
        "release_label": rows[0]["release_label"],
        "approved_at": rows[0]["approved_at"],
        "bundle_sha256": sha256_json(bundle),
    }
    return bundle, metadata


def snapshot_current_league_bundle(args: argparse.Namespace) -> None:
    season = clean_text(args.season)
    output_arg = clean_text(args.output or "")
    if not output_arg:
        raise SystemExit("--snapshot-current requires --output")
    output_path = Path(output_arg)
    if Path("content/reporting").resolve() in output_path.resolve().parents:
        raise SystemExit("current bundle snapshot must stay outside content/reporting")
    bundle, metadata = current_league_bundle_snapshot(season)
    write_json_atomic(output_path, bundle)
    print(json.dumps({"status": "current_snapshot", "season": season,
        "output_path": str(output_path), **metadata}, indent=2))


def assert_public_payload_is_publishable(payload: object, label: str) -> None:
    """Refuse to write a payload carrying protected metadata into Git.

    AGENTS.md names content/reporting/*.json as a surface that must never
    carry a protected club-alias placeholder ("Team A" .. "Team Z") or a
    player pseudonym. The release paths enforce that with a live SQL scan;
    this is the local equivalent for the export path, checked against the
    exact bytes about to be written.
    """
    serialized = json.dumps(payload, indent=2, sort_keys=True)
    if re.search(r"\bTeam [A-Z]\b", serialized):
        raise SystemExit(
            f"refusing to export {label}: payload carries a protected club-alias "
            "placeholder string"
        )
    if "Ath_" in serialized:
        raise SystemExit(
            f"refusing to export {label}: payload carries a player pseudonym"
        )


V6_PUBLIC_DASHBOARD_KEYS = frozenset({
    "generated_at", "team", "season", "analysis_window", "method", "coverage",
    "headline", "monthly", "body_locations", "injury_types", "injury_profiles",
    "injury_type_families", "severity_distribution", "setting_split",
    "setting_metrics", "contact_distribution", "prior_season", "limitations",
})
V6_HEADLINE_KEYS = (
    "recorded_injuries", "time_loss_injuries", "incidence_per_1000h",
    "severity_mean_days", "severity_median_days", "burden_per_1000h",
)
V6_CONTACT_GRID = tuple(
    (setting, key, label)
    for setting in ("all", "match", "training", "unknown")
    for key, label in (
        ("contact", "Contact"),
        ("non_contact", "Non-contact"),
        ("unknown", "Unknown"),
    )
)
V6_SETTING_GRID = (
    ("all", "All"),
    ("match", "Match"),
    ("training", "Training"),
    ("unknown", "Unknown"),
)


def assert_v6_public_dashboard_contract(payload: object, label: str) -> None:
    """Fail closed unless a Year 2 payload is exactly safe for the reader.

    This is deliberately a narrow release-boundary validator rather than a
    best-effort normaliser.  It prevents an added SQL field, a partial contact
    grid, or a malformed injury-type-family payload from reaching a reviewed
    preflight or immutable release snapshot.
    """
    if not isinstance(payload, dict):
        raise SystemExit(f"V6 {label} must be a dashboard object")
    actual_keys = set(payload)
    if actual_keys != V6_PUBLIC_DASHBOARD_KEYS:
        unexpected = sorted(actual_keys - V6_PUBLIC_DASHBOARD_KEYS)
        missing = sorted(V6_PUBLIC_DASHBOARD_KEYS - actual_keys)
        raise SystemExit(
            f"V6 {label} has unexpected top-level payload fields={unexpected} "
            f"or missing fields={missing}"
        )
    if payload.get("season") != "2025-26":
        raise SystemExit(f"V6 {label} must be bound to season 2025-26")
    for field in ("generated_at", "team"):
        if not isinstance(payload.get(field), str) or not payload[field]:
            raise SystemExit(f"V6 {label} has invalid {field}")
    analysis_window = payload.get("analysis_window")
    if not isinstance(analysis_window, dict) or set(analysis_window) != {"start", "end", "basis"}:
        raise SystemExit(f"V6 {label} has invalid analysis_window")
    if analysis_window.get("start") != "2025-09-01" or analysis_window.get("end") != "2026-06-30":
        raise SystemExit(f"V6 {label} has an unapproved analysis window")
    coverage = payload.get("coverage")
    coverage_required = {
        "hours", "match_hours", "training_hours", "distance_km", "exposure_rows",
        "exposed_players", "weeks", "included_exposure_status",
        "analysis_window_start", "analysis_window_end",
    }
    coverage_optional = {"exposure_grain", "teams_included"}
    if (
        not isinstance(coverage, dict)
        or not coverage_required <= set(coverage)
        or not set(coverage) <= coverage_required | coverage_optional
    ):
        raise SystemExit(f"V6 {label} has incomplete coverage")
    expected_coverage_keys = coverage_required | (
        {"teams_included"} if payload["team"] == "URC Overall" else {"exposure_grain"}
    )
    if set(coverage) != expected_coverage_keys:
        raise SystemExit(f"V6 {label} has an invalid coverage shape")
    if coverage.get("analysis_window_start") != "2025-09-01" or coverage.get("analysis_window_end") != "2026-06-30":
        raise SystemExit(f"V6 {label} coverage window differs from the release contract")
    headline = payload.get("headline")
    if not isinstance(headline, list) or tuple(
        item.get("key") if isinstance(item, dict) else None for item in headline
    ) != V6_HEADLINE_KEYS:
        raise SystemExit(f"V6 {label} has an invalid headline metric sequence")
    headline_contract = {
        "recorded_injuries": (
            {"key", "label", "value", "unit", "formula"},
            "count(eligible injury rows in the immutable reporting window, including season-attributed undated rows)",
        ),
        "time_loss_injuries": (
            {"key", "label", "value", "unit", "formula"},
            "count(eligible injury rows where days lost > 0)",
        ),
        "incidence_per_1000h": (
            {"key", "label", "value", "unit", "numerator", "denominator", "formula"},
            "pooled time-loss injuries / pooled exposure hours * 1000",
        ),
        "severity_mean_days": (
            {"key", "label", "value", "unit", "numerator", "denominator", "formula"},
            "pooled days lost / pooled time-loss injuries",
        ),
        "severity_median_days": (
            {"key", "label", "value", "unit", "formula"},
            "median(days lost) across pooled time-loss injuries",
        ),
        "burden_per_1000h": (
            {"key", "label", "value", "unit", "numerator", "denominator", "formula"},
            "pooled days lost / pooled exposure hours * 1000",
        ),
    }
    for item in headline:
        if not isinstance(item, dict):
            raise SystemExit(f"V6 {label} headline metrics require formulas")
        expected_keys, expected_formula = headline_contract[item["key"]]
        if set(item) != expected_keys or item.get("formula") != expected_formula:
            raise SystemExit(f"V6 {label} headline metrics have an invalid formula contract")
    contact = payload.get("contact_distribution")
    if not isinstance(contact, list) or len(contact) != len(V6_CONTACT_GRID):
        raise SystemExit(f"V6 {label} requires the complete 12-cell contact grid")
    for item, expected in zip(contact, V6_CONTACT_GRID, strict=True):
        setting, key, contact_label = expected
        if not isinstance(item, dict) or set(item) != {
            "key", "label", "setting", "recorded_injuries", "time_loss_injuries",
        } or (item.get("setting"), item.get("key"), item.get("label")) != expected:
            raise SystemExit(f"V6 {label} contact grid is not ordered and labelled exactly")

    section_contracts = {
        "monthly": {
            "month", "exposure_hours", "distance_km", "time_loss_injuries",
            "days_lost", "incidence_per_1000h", "burden_per_1000h",
        },
        "body_locations": {
            "key", "label", "time_loss_injuries", "days_lost", "exposure_hours",
            "incidence_per_1000h", "burden_per_1000h", "mean_severity_days",
        },
        "injury_types": {
            "key", "label", "time_loss_injuries", "days_lost", "exposure_hours",
            "incidence_per_1000h", "burden_per_1000h", "mean_severity_days",
        },
        "severity_distribution": {
            "key", "label", "recorded_injuries", "time_loss_injuries", "days_lost",
        },
        "setting_split": {
            "key", "label", "time_loss_injuries", "days_lost", "exposure_hours",
        },
        "setting_metrics": {
            "setting", "label", "time_loss_injuries", "days_lost", "exposure_hours",
            "incidence_per_1000h", "burden_per_1000h", "mean_severity_days",
        },
    }
    for section, expected_keys in section_contracts.items():
        rows = payload.get(section)
        if not isinstance(rows, list) or any(
            not isinstance(row, dict) or set(row) != expected_keys
            for row in rows
        ):
            raise SystemExit(f"V6 {label} {section} has an invalid public shape")
    for section, key_field in (("setting_split", "key"), ("setting_metrics", "setting")):
        rows = payload[section]
        actual_grid = tuple((row.get(key_field), row.get("label")) for row in rows)
        if actual_grid != V6_SETTING_GRID:
            raise SystemExit(
                f"V6 {label} {section} must be the ordered all/match/training/unknown grid"
            )
    profiles = payload.get("injury_profiles")
    if not isinstance(profiles, list):
        raise SystemExit(f"V6 {label} injury_profiles must be an array")
    profile_keys = {
        "dimension", "code", "label", "setting", "time_loss_injuries", "days_lost",
        "exposure_hours", "incidence_per_1000h", "burden_per_1000h", "mean_severity_days",
    }
    for profile in profiles:
        if not isinstance(profile, dict) or set(profile) != profile_keys:
            raise SystemExit(f"V6 {label} injury_profiles have an invalid public shape")
        if profile.get("setting") not in {"all", "match", "training", "unknown"}:
            raise SystemExit(f"V6 {label} injury_profiles contain an invalid setting")
    if profiles and not any(profile.get("dimension") == "diagnosis" for profile in profiles):
        raise SystemExit(f"V6 {label} injury_profiles omit the accepted diagnosis dimension")
    families = payload.get("injury_type_families")
    if not isinstance(families, list):
        raise SystemExit(f"V6 {label} injury_type_families must be an array")
    family_keys = profile_keys | {"mapping_version", "subtypes"}
    for family in families:
        if not isinstance(family, dict) or set(family) != family_keys:
            raise SystemExit(f"V6 {label} injury_type_families have an invalid public shape")
        if family.get("dimension") != "injury_type_family" or family.get("mapping_version") != "injury_type_family_2026-07-21_v1":
            raise SystemExit(f"V6 {label} injury_type_families lost their versioned mapping")
        subtypes = family.get("subtypes")
        if not isinstance(subtypes, list) or not subtypes:
            raise SystemExit(f"V6 {label} injury_type_families require subtype evidence")
        for subtype in subtypes:
            if not isinstance(subtype, dict) or set(subtype) != profile_keys:
                raise SystemExit(f"V6 {label} injury type family subtype has an invalid public shape")
            if subtype.get("dimension") != "injury_type" or subtype.get("setting") != family.get("setting"):
                raise SystemExit(f"V6 {label} injury type family subtype does not match its family")
    prior_season = payload.get("prior_season")
    if (
        not isinstance(prior_season, dict)
        or set(prior_season) != {"season", "status", "note"}
        or prior_season.get("season") != "2024-25"
        or prior_season.get("status") != "frozen"
    ):
        raise SystemExit(f"V6 {label} must retain exactly the frozen prior-season marker")
    for section in ("method", "limitations"):
        if not isinstance(payload.get(section), list) or not all(
            isinstance(item, str) and item for item in payload[section]
        ):
            raise SystemExit(f"V6 {label} {section} must be a non-empty public string array")
    assert_public_payload_is_publishable(payload, f"V6 {label}")


def write_parity_export_set(
    planned: list[tuple[Path, object]],
) -> None:
    """Replace the full parity set, restoring every prior file on failure."""
    reporting_dir = Path("content/reporting")
    reporting_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        prefix=".parity-export-", dir=reporting_dir
    ) as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        staged: list[tuple[Path, Path]] = []
        backups: dict[Path, Path] = {}
        for index, (target, public) in enumerate(planned):
            staged_path = temp_dir / f"{index:02d}-{target.name}"
            write_json_atomic(staged_path, public)
            staged.append((target, staged_path))
            if target.exists():
                backup_path = temp_dir / f"{index:02d}-{target.name}.backup"
                shutil.copy2(target, backup_path)
                backups[target] = backup_path

        attempted: list[Path] = []
        try:
            for target, staged_path in staged:
                target.parent.mkdir(parents=True, exist_ok=True)
                attempted.append(target)
                os.replace(staged_path, target)
        except BaseException as exc:
            recovery_errors: list[str] = []
            for target in reversed(attempted):
                backup_path = backups.get(target)
                try:
                    if backup_path is not None:
                        os.replace(backup_path, target)
                    elif target.exists():
                        recovery_path = (
                            Path("data/reporting")
                            / f"{target.stem}_failed_parity_{uuid.uuid4().hex}.json"
                        )
                        recovery_path.parent.mkdir(parents=True, exist_ok=True)
                        os.replace(target, recovery_path)
                except BaseException as recovery_exc:
                    recovery_errors.append(f"{target}: {recovery_exc}")
            if recovery_errors:
                raise RuntimeError(
                    "parity export failed and rollback was incomplete: "
                    + "; ".join(recovery_errors)
                ) from exc
            raise


def write_team_dashboard_parity_exports(
    season: str, *, expected_release_label: str | None = None,
    expected_release_id: str | None = None,
    expected_bundle: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Refresh committed per-team parity exports from the approved bundle.

    release-league rewrites only content/reporting/urc_dashboard_<season>.json,
    so the 16 per-team files go stale after every league release while the
    website keeps serving the bundle through reporting.latest_team_dashboard_v5.
    Those files are a parity and emergency export, never an application input,
    so this rewrites each one from the same approved bundle the site serves and
    records the release identity it came from. Run it after every accepted
    release-league promotion.

    Values are runtime-equivalent to the stored bundle, not character-identical
    to its numeric text: a Postgres numeric such as 2.42234971686951152000
    round-trips through an IEEE double and reserializes as 2.4223497168695114.
    That is far beyond dashboard precision and these files are not application
    inputs, so it is accepted rather than corrected, but "verbatim" would be
    the wrong word for it.

    This is the only writer of the committed per-team payloads, so it applies
    the same two guards every other dashboard emitter applies: internal keys
    are stripped, and the serialized bytes are refused if they carry a
    protected club-alias placeholder or a player pseudonym.
    """
    bundle, metadata = current_league_bundle_snapshot(season)
    if (
        expected_release_label is not None
        and metadata["release_label"] != expected_release_label
    ):
        raise SystemExit(
            "approved bundle changed before parity export: "
            f"expected_release_label={expected_release_label!r}, "
            f"actual_release_label={metadata['release_label']!r}"
        )
    if expected_release_id is not None and metadata["release_id"] != expected_release_id:
        raise SystemExit(
            "approved bundle changed before parity export: "
            f"expected_release_id={expected_release_id!r}, "
            f"actual_release_id={metadata['release_id']!r}"
        )
    if expected_bundle is not None:
        bundle_diffs = diff_json_documents(expected_bundle, bundle)
        if bundle_diffs:
            paths = ", ".join(diff["path"] for diff in bundle_diffs[:10])
            raise SystemExit(
                "approved bundle differs from the just-promoted candidate "
                f"before parity export ({paths})"
            )
    # Resolve every target path and refuse an unaccounted stale export BEFORE
    # writing anything, so a failure cannot leave a partial refresh alongside
    # the stale file it was meant to catch.
    planned: list[tuple[Path, object]] = []
    for team in bundle["teams"]:
        team_key = clean_text(team.get("team_key"))
        dashboard = team.get("dashboard")
        if not team_key or not isinstance(dashboard, dict):
            raise SystemExit("approved bundle contains an incomplete team payload")
        public = without_keys(dashboard, {"source_files", "pipeline_evidence"})
        assert_public_payload_is_publishable(public, team_key)
        planned.append(
            (
                Path("content") / "reporting" / f"{team_key}_dashboard_{season}.json",
                public,
            )
        )
    league_public = without_keys(bundle["league"], {"source_files", "pipeline_evidence"})
    assert_public_payload_is_publishable(league_public, f"urc_dashboard_{season}")
    league_path = Path("content") / "reporting" / f"urc_dashboard_{season}.json"
    planned.append((league_path, league_public))
    expected_paths = {str(path) for path, _ in planned}
    stale = {
        str(path)
        for path in Path("content/reporting").glob(f"*_dashboard_{season}.json")
        if str(path) not in expected_paths
        and path.name != f"urc_dashboard_{season}.json"
    }
    if stale:
        raise SystemExit(
            "content/reporting holds per-team exports the approved bundle does "
            f"not account for: {sorted(stale)}"
        )
    written = [str(path) for path, _ in planned]
    write_parity_export_set(planned)
    export_hashes = {str(path): sha256_file(path) for path, _ in planned}
    return {
        "status": "team_parity_exported", "season": season,
        "team_count": len(bundle["teams"]), "paths": written,
        "league_path": str(league_path),
        "export_set_sha256": sha256_json(export_hashes), **metadata,
    }


def export_team_dashboard_parity_json(args: argparse.Namespace) -> None:
    season = clean_text(args.season)
    if not season:
        raise SystemExit("--season is required")
    print(
        json.dumps(
            write_team_dashboard_parity_exports(season),
            indent=2,
        )
    )


def league_release_manifest_document(
    *, release_label: str, season: str, release_tuple: dict[str, str],
    required_migrations: list[object], member_count: int,
    member_input_hash: str, league_payload_sha256: str,
    bundle_payload_sha256: str, team_payload_sha256s: dict[str, str],
    reviewed_preflight_sha256: str,
    reviewed_preflight_manifest_sha256: str,
    provenance: dict[str, str], dirty_worktree_paths: list[str],
    dirty_worktree_allowed_paths: list[str], parity_export: dict[str, Any],
    rollback: dict[str, Any], rollback_of_release_id: str | None,
    rollback_replaces_release_id: str | None,
    timings_ms: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Build the deterministic local closeout record for one promoted bundle."""
    manifest = {
        "schema_version": "urc_league_release_manifest_v1",
        "status": "promoted_and_exported",
        "release_label": release_label,
        "season": season,
        "release_tuple": release_tuple,
        "required_migrations": required_migrations,
        "member_count": member_count,
        "member_input_hash": member_input_hash,
        "league_payload_sha256": league_payload_sha256,
        "bundle_payload_sha256": bundle_payload_sha256,
        "team_payload_sha256s": team_payload_sha256s,
        "reviewed_preflight_sha256": reviewed_preflight_sha256,
        "reviewed_preflight_manifest_sha256": reviewed_preflight_manifest_sha256,
        "provenance": provenance,
        "dirty_worktree_paths": dirty_worktree_paths,
        "dirty_worktree_allowed_paths": dirty_worktree_allowed_paths,
        "candidate_assembly_reads": 1,
        "promotion_candidate_validation_reads": 1,
        "reconciliation": {
            "migration_prerequisites": "passed",
            "semantic_sections": "passed",
            "member_roster": "passed",
            "reviewed_candidate_equality": "passed",
            "database_readback": "passed",
            "parity_export": "passed",
        },
        "parity_export": {
            "team_count": parity_export["team_count"],
            "export_set_sha256": parity_export["export_set_sha256"],
            "bundle_sha256": parity_export.get("bundle_sha256"),
        },
    }
    if season == "2025-26" and release_tuple.get("analysis_version") == "v6":
        manifest["rollback"] = {
            "rollback_of_release_id": rollback_of_release_id,
            "replaces_release_id": rollback_replaces_release_id,
            "retained_predecessor": rollback,
        }
    else:
        # The frozen Year 1 manifest shape is a public audit contract. Keep its
        # direct predecessor object and timing block byte-for-byte compatible;
        # the append-only rollback envelope is V6-only.
        manifest["timings_ms"] = timings_ms or {}
        manifest["rollback"] = rollback
    return manifest


def v6_local_finalizer_command(
    *, release_id: str, release_label: str, preflight_file: Path,
) -> str:
    return shlex.join([
        "python3", "-m", "pipeline", "finalize-v6-league-release-local",
        "--release-id", release_id,
        "--release-label", release_label,
        "--preflight-file", str(preflight_file),
    ])


def finalize_v6_league_release_local(args: argparse.Namespace) -> None:
    """Resume only the local exports/evidence for one approved V6 release."""
    release_id = clean_text(args.release_id)
    release_label = clean_text(args.release_label)
    preflight_path = Path(clean_text(args.preflight_file))
    try:
        uuid.UUID(release_id)
    except ValueError as error:
        raise SystemExit("--release-id must be a UUID") from error
    if not release_label or not preflight_path.is_file():
        raise SystemExit("exact --release-label and existing --preflight-file are required")
    reviewed_manifest_path = Path(f"{preflight_path}.manifest.json")
    if not reviewed_manifest_path.is_file():
        raise SystemExit("reviewed V6 preflight manifest is required for local finalisation")
    try:
        reviewed_bundle = json.loads(preflight_path.read_bytes())
        reviewed_manifest = json.loads(reviewed_manifest_path.read_bytes())
    except json.JSONDecodeError as error:
        raise SystemExit("reviewed V6 preflight artefacts must be valid JSON") from error
    if not isinstance(reviewed_bundle, dict) or not isinstance(reviewed_manifest, dict):
        raise SystemExit("reviewed V6 preflight artefacts must be JSON objects")
    expected_contract = release_contract_for("2025-26", YEAR2_2025_26_RELEASE_TUPLE)
    assert_checksum_bound_release_migrations(expected_contract, "V6 local finalisation")

    params = SqlParams()
    rows = query_sql(
        f"""
        select release.id::text as release_id, release.release_label,
          context.season, context.analysis_version,
          context.classification_view_version, context.cohort_view_version,
          run.status as pipeline_run_status, run.parameters,
          run.code_version, run.dependency_lock_hash, run.operator,
          league.payload_sha256 as league_payload_sha256,
          teams.team_payload_sha256s, teams.member_count,
          reporting.canonical_jsonb_sha256_v1(jsonb_build_object(
            'schema_version','urc_dashboard_bundle_v2', 'season',context.season,
            'league',league.dashboard_payload, 'teams',teams.dashboards
          )) as bundle_payload_sha256
        from reporting.aggregate_releases release
        join reporting.league_release_context_v2 context on context.release_id=release.id
        join audit.pipeline_runs run on run.id=release.pipeline_run_id
        join reporting.league_release_payloads_v6 league on league.release_id=release.id
        cross join lateral (
          select count(*)::integer as member_count,
            jsonb_object_agg(payload.team_key,payload.payload_sha256 order by payload.team_key)
              as team_payload_sha256s,
            jsonb_agg(jsonb_build_object('team_key',payload.team_key,'dashboard',payload.dashboard_payload)
              order by payload.team_key) as dashboards
          from reporting.team_dashboard_payloads_v2 payload
          where payload.bundle_release_id=release.id
        ) teams
        where release.id={params.text(release_id)}::uuid
          and release.release_label={params.text(release_label)}
          and release.status='approved' and context.season='2025-26'
          and context.analysis_version='v6'
        """,
        params.values,
    )
    if len(rows) != 1:
        raise SystemExit("exact approved V6 release identity was not found")
    row = rows[0]
    parameters = row.get("parameters")
    team_payload_sha256s = row.get("team_payload_sha256s")
    if (
        row.get("pipeline_run_status") != "succeeded"
        or row.get("member_count") != 16
        or not isinstance(parameters, dict)
        or not isinstance(team_payload_sha256s, dict)
        or len(team_payload_sha256s) != 16
    ):
        raise SystemExit("approved V6 release audit or member snapshot is incomplete")

    reviewed_sha256 = sha256_file(preflight_path)
    reviewed_manifest_sha256 = sha256_file(reviewed_manifest_path)
    if (
        parameters.get("reviewed_preflight_sha256") != reviewed_sha256
        or parameters.get("reviewed_preflight_manifest_sha256") != reviewed_manifest_sha256
        or reviewed_manifest.get("preflight_file_sha256") != reviewed_sha256
    ):
        raise SystemExit("reviewed V6 preflight checksums do not match the approved release audit")
    expected_migrations = [
        {"version": item.version, "name": item.name, "sha256": item.sha256}
        for item in expected_contract.required_migration_contracts
    ]
    if reviewed_manifest.get("required_migrations") != expected_migrations:
        raise SystemExit("reviewed V6 preflight does not bind the exact migration contract")
    if reviewed_manifest.get("local_evidence_files") != year2_release_local_evidence_records(expected_contract):
        raise SystemExit("reviewed V6 preflight does not bind the exact local evidence bytes")
    if (
        reviewed_manifest.get("bundle_payload_sha256") != row.get("bundle_payload_sha256")
        or parameters.get("bundle_payload_sha256") != row.get("bundle_payload_sha256")
        or reviewed_manifest.get("league_payload_sha256") != row.get("league_payload_sha256")
        or parameters.get("league_dashboard_payload_sha256") != row.get("league_payload_sha256")
        or reviewed_manifest.get("team_payload_sha256s") != team_payload_sha256s
        or parameters.get("team_dashboard_payload_sha256s") != team_payload_sha256s
        or reviewed_manifest.get("member_input_hash") != parameters.get("member_input_hash")
    ):
        raise SystemExit("approved V6 database payloads differ from the reviewed preflight evidence")

    approved_bundle, metadata = current_league_bundle_snapshot("2025-26")
    if metadata["release_id"] != release_id or metadata["release_label"] != release_label:
        raise SystemExit("requested V6 release is no longer the currently approved complete bundle")
    diffs = diff_json_documents(reviewed_bundle, approved_bundle)
    if diffs:
        raise SystemExit(
            "approved V6 bundle differs from the reviewed preflight: "
            + ", ".join(diff["path"] for diff in diffs[:10])
        )
    parity_export = write_team_dashboard_parity_exports(
        "2025-26", expected_release_id=release_id,
        expected_release_label=release_label, expected_bundle=reviewed_bundle,
    )

    predecessor_id = clean_text(parameters.get("predecessor_release_id"))
    if predecessor_id:
        predecessor_params = SqlParams()
        predecessor_rows = query_sql(
            f"select release_label,approved_at from reporting.aggregate_releases "
            f"where id={predecessor_params.text(predecessor_id)}::uuid and status in ('approved','retired')",
            predecessor_params.values,
        )
        if len(predecessor_rows) != 1:
            raise SystemExit("retained V6 predecessor audit is incomplete")
        rollback = {
            "release_id": predecessor_id,
            "release_label": predecessor_rows[0]["release_label"],
            "approved_at": predecessor_rows[0]["approved_at"],
            "bundle_sha256": parameters.get("predecessor_bundle_sha256"),
            "snapshot_sha256": parameters.get("predecessor_snapshot_sha256"),
        }
    else:
        rollback = {"status": "no predecessor existed"}

    release_manifest = league_release_manifest_document(
        release_label=release_label,
        season="2025-26",
        release_tuple=reviewed_manifest["release_tuple"],
        required_migrations=expected_migrations,
        member_count=16,
        member_input_hash=parameters["member_input_hash"],
        league_payload_sha256=row["league_payload_sha256"],
        bundle_payload_sha256=row["bundle_payload_sha256"],
        team_payload_sha256s=team_payload_sha256s,
        reviewed_preflight_sha256=reviewed_sha256,
        reviewed_preflight_manifest_sha256=reviewed_manifest_sha256,
        provenance={
            "code_version": row["code_version"],
            "dependency_lock_hash": row["dependency_lock_hash"],
            "operator": row["operator"],
        },
        dirty_worktree_paths=reviewed_manifest.get("dirty_worktree_paths", []),
        dirty_worktree_allowed_paths=reviewed_manifest.get("dirty_worktree_allowed_paths", []),
        parity_export=parity_export,
        rollback=rollback,
        rollback_of_release_id=reviewed_manifest.get("rollback_of_release_id"),
        rollback_replaces_release_id=reviewed_manifest.get("rollback_replaces_release_id"),
    )
    release_manifest_path = Path("data/reporting") / f"{release_label}_release_manifest.json"
    if release_manifest_path.exists():
        try:
            existing_manifest = json.loads(release_manifest_path.read_bytes())
        except json.JSONDecodeError as error:
            raise SystemExit("existing V6 release manifest is invalid JSON") from error
        if existing_manifest != release_manifest:
            raise SystemExit("existing V6 release manifest differs from deterministic finalisation")
    else:
        write_json_atomic(release_manifest_path, release_manifest)
    print(json.dumps({
        "status": "v6_local_finalisation_complete",
        "release_id": release_id,
        "release_label": release_label,
        "release_manifest_path": str(release_manifest_path),
        "parity_export": release_manifest["parity_export"],
    }, indent=2))


def record_failed_league_release_attempt(
    *, label: str, season: str, input_hash: str, output_hash: str,
    parameters: dict[str, Any], provenance: dict[str, str], failure_stage: str,
) -> None:
    """Best-effort immutable failure marker after a rolled-back promotion."""
    params = SqlParams()
    failed_parameters = {**parameters, "failure_stage": failure_stage}
    run_sql(
        f"""
        do $$
        begin
          if not exists (
            select 1 from reporting.aggregate_releases
            where release_label = {params.text(label)}
          ) then
            with failed_run as (
              insert into audit.pipeline_runs
                (command, team, season, status, input_hash, output_hash, parameters,
                 ended_at, code_version, dependency_lock_hash, operator)
              values ('release-league', 'URC Overall', {params.text(season)}, 'failed',
                {params.text(input_hash)}, {params.text(output_hash)},
                {params.jsonb(failed_parameters)}, now(),
                {params.text(provenance['code_version'])},
                {params.text(provenance['dependency_lock_hash'])},
                {params.text(provenance['operator'])})
              returning id
            )
            insert into reporting.aggregate_releases
              (release_label, status, pipeline_run_id)
            select {params.text(label)}, 'retired', id from failed_run;
          end if;
        end $$;
        """,
        params.values,
    )


def league_release_plan(
    *, season: str, analysis_version: str,
    classification_view_version: str, cohort_view_version: str,
) -> dict[str, Any]:
    """Return the approval-separated operator plan without touching the database."""
    tuple_args = [
        "--analysis-version", analysis_version,
        "--classification-view-version", classification_view_version,
        "--cohort-view-version", cohort_view_version,
    ]
    preflight_path = (
        f"data/reporting/urc_dashboard_{season}_{analysis_version}_preflight.json"
    )
    previous_path = f"data/reporting/urc_dashboard_{season}_previous.json"
    reviewer_arg = "Abdel Babiker" if analysis_version == "v6" else "<reviewer>"
    runner = ["node", "pipeline/run_with_pooler.mjs", "python3", "-m", "pipeline"]
    steps: list[dict[str, Any]] = [
        {
            "stage": "local",
            "approval": "none",
            "action": "verify evidence files, Git state, and focused tests",
        },
    ]
    if analysis_version == "v5":
        steps.append(
            {
                "stage": "live_write",
                "approval": "exact hosted target and snapshot refresh required",
                "action": shlex.join(
                    [
                        "node", "pipeline/run_with_pooler.mjs", "node",
                        "pipeline/sql_exec.mjs",
                        "tools/sql/refresh_analysis_window_v5_candidate_snapshots.sql",
                    ]
                ),
            }
        )
    first_promotion = shlex.join(
        runner + ["release-league", "--season", season]
        + tuple_args
        + ["--preflight-file", preflight_path, "--preflight-reviewer", reviewer_arg]
    )
    successor_promotion = shlex.join(
        runner + ["release-league", "--season", season]
        + tuple_args
        + [
            "--previous-bundle-file", previous_path,
            "--preflight-file", preflight_path,
            "--preflight-reviewer", reviewer_arg,
        ]
    )
    if analysis_version == "v6":
        steps.extend([
            {
                "stage": "read_only_live",
                "approval": "database read access only",
                "condition": "only when an approved predecessor exists",
                "action": shlex.join(
                    runner + ["release-league", "--season", season]
                    + tuple_args
                    + ["--snapshot-current", "--output", previous_path]
                ),
            },
            {
                "stage": "read_only_live",
                "approval": "database read access only",
                "action": shlex.join(
                    runner + ["release-league", "--season", season]
                    + tuple_args
                    + ["--preflight", "--output", preflight_path]
                ),
            },
            {
                "stage": "human_review",
                "approval": "record review of the exact preflight file",
                "action": preflight_path,
            },
            {
                "stage": "live_write",
                "approval": "exact hosted promotion required",
                "action": first_promotion,
                "action_if_predecessor_exists": successor_promotion,
                "includes": "promotion and 16-team parity export",
            },
        ])
    else:
        # Preserve the frozen Year 1 operator plan exactly. Every accepted
        # Year 1 release already has a predecessor, so its snapshot and
        # predecessor-bound promotion are unconditional legacy steps.
        steps.extend([
            {
                "stage": "read_only_live",
                "approval": "database read access only",
                "action": shlex.join(
                    runner + ["release-league", "--season", season]
                    + tuple_args
                    + ["--snapshot-current", "--output", previous_path]
                ),
            },
            {
                "stage": "read_only_live",
                "approval": "database read access only",
                "action": shlex.join(
                    runner + ["release-league", "--season", season]
                    + tuple_args
                    + ["--preflight", "--output", preflight_path]
                ),
            },
            {
                "stage": "human_review",
                "approval": "record review of the exact preflight file",
                "action": preflight_path,
            },
            {
                "stage": "live_write",
                "approval": "exact hosted promotion required",
                "action": successor_promotion,
                "includes": "promotion and 16-team parity export",
            },
        ])
    return {
        "status": "plan",
        "season": season,
        "release_tuple": {
            "analysis_version": analysis_version,
            "classification_view_version": classification_view_version,
            "cohort_view_version": cohort_view_version,
        },
        "database_access": "none",
        "steps": steps,
        "rollback": (
            {
                "mode": "append_only_retained_bundle_successor",
                "target": "<exact-predecessor-release-uuid>",
                "preflight": shlex.join(
                    runner + ["release-league", "--season", season]
                    + tuple_args
                    + [
                        "--rollback-of-release-id", "<exact-predecessor-release-uuid>",
                        "--preflight", "--output",
                        f"data/reporting/urc_dashboard_{season}_v6_rollback_preflight.json",
                    ]
                ),
                "promotion": "run the same command with --previous-bundle-file, "
                "the reviewed --preflight-file, and --preflight-reviewer 'Abdel Babiker'",
                "invariant": "creates a new approved release and never re-approves history",
            }
            if analysis_version == "v6"
            else (
                "re-promote the retained predecessor tuple, then regenerate the "
                "16-team parity exports"
            )
        ),
    }


def load_league_release_candidate(
    *, season: str, analysis_version: str,
    classification_view_version: str, cohort_view_version: str,
    league_candidate_view: str, team_candidate_view: str,
) -> dict[str, Any]:
    """Load and hash the build-pinned league and 16-team candidate once."""
    dashboard_hash_sql = (
        "reporting.canonical_jsonb_sha256_v1(dashboard)"
        if analysis_version == "v6"
        else "encode(digest(convert_to(dashboard::text, 'UTF8'), 'sha256'), 'hex')"
    )
    bundle_hash_sql = (
        "reporting.canonical_jsonb_sha256_v1(document)"
        if analysis_version == "v6"
        else "encode(digest(convert_to(document::text, 'UTF8'), 'sha256'), 'hex')"
    )
    candidate_params = SqlParams()
    rows = query_sql(
        f"""
        with league as (
          select dashboard, classification_evidence_sha256, cohort_evidence_sha256
          from {league_candidate_view}
          where season = {candidate_params.text(season)}
            and analysis_version = {candidate_params.text(analysis_version)}
            and classification_view_version = {candidate_params.text(classification_view_version)}
            and cohort_view_version = {candidate_params.text(cohort_view_version)}
        ), team_rows as (
          select candidate.team_key, candidate.team_release_id::text,
                 candidate.curated_build_id::text, candidate.dashboard
          from {team_candidate_view} candidate
          cross join league
          where candidate.season = {candidate_params.text(season)}
            and candidate.analysis_version = {candidate_params.text(analysis_version)}
            and candidate.classification_view_version = {candidate_params.text(classification_view_version)}
            and candidate.classification_evidence_sha256 is not distinct from
                league.classification_evidence_sha256
            and candidate.cohort_view_version = {candidate_params.text(cohort_view_version)}
            and candidate.cohort_evidence_sha256 is not distinct from
                league.cohort_evidence_sha256
        ), teams as (
          select
            jsonb_agg(jsonb_build_object(
              'team_key', team_key,
              'team_release_id', team_release_id,
              'curated_build_id', curated_build_id,
              'dashboard', dashboard
            ) order by team_key) as candidates,
            jsonb_agg(jsonb_build_object(
              'team_key', team_key, 'dashboard', dashboard
            ) order by team_key) as dashboards,
            jsonb_object_agg(
              team_key,
              {dashboard_hash_sql}
              order by team_key
            ) as hashes
          from team_rows
        ), bundle as (
          select league.dashboard, league.classification_evidence_sha256,
                 league.cohort_evidence_sha256, teams.candidates, teams.hashes,
                 jsonb_build_object(
                   'schema_version', 'urc_dashboard_bundle_v2',
                   'season', {candidate_params.text(season)},
                   'league', league.dashboard,
                   'teams', teams.dashboards
                 ) as document
          from league cross join teams
        )
        select dashboard, classification_evidence_sha256, cohort_evidence_sha256,
               candidates as team_payloads, hashes as team_payload_sha256s,
               {dashboard_hash_sql}
                 as league_payload_sha256,
               {bundle_hash_sql}
                 as bundle_payload_sha256,
               document::text as bundle_payload_json
        from bundle
        """,
        candidate_params.values,
    )
    if len(rows) != 1:
        raise SystemExit(
            f"release-league expected one complete league payload for {season!r}, "
            f"found {len(rows)}"
        )
    return rows[0]


def load_v6_retained_league_rollback_candidate(
    *, season: str, rollback_of_release_id: str,
) -> dict[str, Any]:
    """Load one retained V6 bundle as an append-only rollback successor source."""
    params = SqlParams()
    rows = query_sql(
        f"""
        with current_approved as (
          select context.release_id,
            run.parameters ->> 'predecessor_release_id' as predecessor_release_id
          from reporting.league_release_context_v2 context
          join reporting.aggregate_releases release on release.id = context.release_id
          join audit.pipeline_runs run on run.id = release.pipeline_run_id
          where context.season = {params.text(season)}
            and context.analysis_version = 'v6' and release.status = 'approved'
          order by release.approved_at desc nulls last, release.created_at desc,
            context.release_id desc
          limit 1
        ), prior as (
          select context.release_id, context.season, context.analysis_version,
            context.classification_view_version, context.classification_evidence_sha256,
            context.cohort_view_version, context.cohort_evidence_sha256,
            league.dashboard_payload as dashboard,
            current_approved.release_id as replaces_release_id
          from reporting.league_release_context_v2 context
          join reporting.aggregate_releases release on release.id = context.release_id
          join reporting.league_release_payloads_v6 league on league.release_id = context.release_id
          join current_approved
            on current_approved.predecessor_release_id = context.release_id::text
          where context.release_id = {params.text(rollback_of_release_id)}::uuid
            and context.season = {params.text(season)} and context.analysis_version = 'v6'
            and release.status in ('approved', 'retired')
        ), team_rows as (
          select payload.team_key, payload.team_release_id::text, payload.curated_build_id::text,
            payload.dashboard_payload as dashboard
          from reporting.team_dashboard_payloads_v2 payload
          join prior on prior.release_id = payload.bundle_release_id
        ), teams as (
          select jsonb_agg(jsonb_build_object(
              'team_key', team_key, 'team_release_id', team_release_id,
              'curated_build_id', curated_build_id, 'dashboard', dashboard
            ) order by team_key) as candidates,
            jsonb_agg(jsonb_build_object('team_key', team_key, 'dashboard', dashboard)
              order by team_key) as dashboards,
            jsonb_object_agg(team_key, reporting.canonical_jsonb_sha256_v1(dashboard)
              order by team_key) as hashes
          from team_rows
        ), bundle as (
          select prior.*, teams.candidates, teams.hashes,
            jsonb_build_object('schema_version', 'urc_dashboard_bundle_v2',
              'season', prior.season, 'league', prior.dashboard, 'teams', teams.dashboards) as document
          from prior cross join teams
        )
        select dashboard, classification_evidence_sha256, cohort_evidence_sha256,
          replaces_release_id::text,
          candidates as team_payloads, hashes as team_payload_sha256s,
          reporting.canonical_jsonb_sha256_v1(dashboard) as league_payload_sha256,
          reporting.canonical_jsonb_sha256_v1(document) as bundle_payload_sha256,
          document::text as bundle_payload_json
        from bundle
        """,
        params.values,
    )
    if len(rows) != 1:
        raise SystemExit("V6 rollback requires exactly one retained immutable 16-team bundle")
    candidate = rows[0]
    team_payloads = candidate.get("team_payloads")
    if (
        not isinstance(team_payloads, list)
        or len(team_payloads) != 16
        or len({row.get("team_key") for row in team_payloads if isinstance(row, dict)}) != 16
    ):
        raise SystemExit("V6 rollback retained bundle must contain exactly 16 team payloads")
    return candidate


def release_league(args: argparse.Namespace) -> None:
    """Preflight and atomically publish the build-pinned 16-team V2 league dashboard."""
    season = clean_text(args.season)
    if not season:
        raise SystemExit("--season is required")
    preflight = bool(args.preflight)
    preflight_file_arg = clean_text(args.preflight_file or "")
    reviewer = clean_text(args.preflight_reviewer or "")
    allow_legacy_preflight_without_manifest = bool(
        getattr(args, "allow_legacy_preflight_without_manifest", False)
    )
    previous_bundle_file_arg = clean_text(getattr(args, "previous_bundle_file", ""))
    rollback_of_release_id = clean_text(
        getattr(args, "rollback_of_release_id", "")
    )
    default_release_tuple = (
        YEAR2_2025_26_RELEASE_TUPLE if season == "2025-26" else ("v2", "v2", "v2")
    )
    analysis_version = clean_text(getattr(args, "analysis_version", "")) or default_release_tuple[0]
    classification_view_version = (
        clean_text(getattr(args, "classification_view_version", ""))
        or default_release_tuple[1]
    )
    cohort_view_version = (
        clean_text(getattr(args, "cohort_view_version", ""))
        or default_release_tuple[2]
    )
    release_tuple = (
        analysis_version,
        classification_view_version,
        cohort_view_version,
    )
    year2_release_contract = None
    if season == "2025-26":
        try:
            year2_release_contract = release_contract_for(season, release_tuple)
        except ValueError as error:
            raise SystemExit(str(error)) from error
    if rollback_of_release_id:
        if season != "2025-26" or analysis_version != "v6":
            raise SystemExit("--rollback-of-release-id is available only for the 2025-26 V6 release")
        try:
            rollback_of_release_id = str(uuid.UUID(rollback_of_release_id))
        except ValueError as error:
            raise SystemExit("--rollback-of-release-id must be a UUID") from error
    supported_release_variants = {
        ("v2", "v2", "v2"),
        ("v2", "reporting_classification_2026-07-20_v1", "v2"),
        ("v3", "reporting_classification_2026-07-20_v1", "season_bound_2026-07-20_v1"),
        ("v3", "reporting_classification_2026-07-22_v2", "season_bound_2026-07-20_v1"),
        ("v4", "reporting_classification_2026-07-22_v2", "lineage_2024-25_2026-07-24_v1"),
        ("v5", "reporting_classification_2026-07-22_v2", ANALYSIS_WINDOW_V5_COHORT_VIEW_VERSION),
    }
    if season == "2025-26":
        supported_release_variants.add(YEAR2_2025_26_RELEASE_TUPLE)
    if release_tuple not in supported_release_variants:
        raise SystemExit(
            "unsupported analysis/classification/cohort version combination; "
            "V3 requires an accepted reporting classification and the season-bound cohort; "
            "V4 requires the accepted OSIICS classification and lineage cohort; "
            "V5 requires the accepted OSIICS classification and analysis-window cohort"
        )
    if bool(getattr(args, "plan", False)):
        print(
            json.dumps(
                league_release_plan(
                    season=season,
                    analysis_version=analysis_version,
                    classification_view_version=classification_view_version,
                    cohort_view_version=cohort_view_version,
                ),
                indent=2,
            )
        )
        return
    if bool(getattr(args, "snapshot_current", False)):
        snapshot_current_league_bundle(args)
        return
    uses_osiics_successor = (
        analysis_version == "v3"
        and classification_view_version == "reporting_classification_2026-07-22_v2"
    )
    # A classification-only successor inherits every non-classification field
    # from the approved immutable bundle. Recomputing the full dashboard is both
    # unnecessary and much slower than replacing the three affected sections.
    # V4 and V5 read their respective candidate branches directly rather than
    # analysis.*_dashboard_release_candidates_v6. Those v6 views are UNION ALL
    # chains over every historical candidate generation, and a
    # `analysis_version = 'v4'` filter does not prune the legacy branches, so
    # each read plans and evaluates the whole stack (measured 2026-07-24: the
    # lineage branch alone answers in ~53s, the same projection through v6 did
    # not return in over 7 minutes and overran the pooler). The lineage
    # candidate views added by 20260724190000 contain exactly the rows v6
    # contributes for 'v4'; see that migration's header for the equivalence.
    league_candidate_view = (
        year2_release_contract.league_candidate_view
        if analysis_version == "v6" and year2_release_contract is not None
        else (
            "analysis.league_dashboard_release_candidates_analysis_window_v5"
            if analysis_version == "v5"
        else (
            "analysis.league_dashboard_release_candidates_lineage_v4"
            if analysis_version == "v4"
            else (
                "analysis.league_dashboard_classification_incremental_20260722_v1"
                if uses_osiics_successor else "analysis.league_dashboard_release_candidates_v4"
            )
        )
        )
    )
    team_candidate_view = (
        year2_release_contract.league_team_candidate_view
        if analysis_version == "v6" and year2_release_contract is not None
        else (
            "analysis.team_dashboard_release_candidates_analysis_window_v5"
            if analysis_version == "v5"
        else (
            "analysis.team_dashboard_release_candidates_lineage_v4"
            if analysis_version == "v4"
            else (
                "analysis.team_dashboard_classification_incremental_20260722_v1"
                if uses_osiics_successor else "analysis.team_dashboard_release_candidates_v4"
            )
        )
        )
    )
    if analysis_version == "v6" and not team_candidate_view:
        raise SystemExit("V6 release contract lacks an immutable league team candidate view")
    member_view = (
        year2_release_contract.member_view
        if analysis_version == "v6" and year2_release_contract is not None
        else "analysis.league_member_releases_v2"
    )
    if not member_view:
        raise SystemExit("release contract lacks a league member relation")
    release_rule_version = (
        year2_release_contract.release_rule_version
        if analysis_version == "v6" and year2_release_contract is not None
        else (
            ANALYSIS_WINDOW_LEAGUE_DASHBOARD_RELEASE_RULE_VERSION
            if analysis_version == "v5"
        else (
            LINEAGE_LEAGUE_DASHBOARD_RELEASE_RULE_VERSION
            if analysis_version == "v4"
            else (
                SEASON_BOUND_LEAGUE_DASHBOARD_RELEASE_RULE_VERSION
                if analysis_version == "v3"
                else LEAGUE_DASHBOARD_RELEASE_RULE_VERSION
            )
        )
        )
    )
    release_reason_code = (
        year2_release_contract.release_reason_code
        if analysis_version == "v6" and year2_release_contract is not None
        else (
            "league_dashboard_release_v5"
            if analysis_version == "v5"
        else (
            "league_dashboard_release_v4"
            if analysis_version == "v4"
            else (
                "league_dashboard_release_v3"
                if analysis_version == "v3"
                else "league_dashboard_release_v2"
            )
        )
        )
    )
    decision_recorded_at = (
        year2_release_contract.decision_recorded_at
        if analysis_version == "v6" and year2_release_contract is not None
        else (
            "2026-07-25"
            if analysis_version == "v5"
        else (
            "2026-07-24"
            if analysis_version == "v4"
            else "2026-07-19" if analysis_version == "v3" else "2026-07-14"
        )
        )
    )
    if preflight and preflight_file_arg:
        raise SystemExit("--preflight cannot be combined with --preflight-file")
    if preflight_file_arg and not reviewer:
        raise SystemExit("--preflight-reviewer is required with --preflight-file")
    if reviewer and not preflight_file_arg:
        raise SystemExit("--preflight-reviewer requires --preflight-file")
    if analysis_version == "v6" and preflight_file_arg and reviewer != "Abdel Babiker":
        raise SystemExit("V6 league promotion requires --preflight-reviewer 'Abdel Babiker'")
    if allow_legacy_preflight_without_manifest and (
        preflight or not preflight_file_arg
    ):
        raise SystemExit(
            "--allow-legacy-preflight-without-manifest requires promotion "
            "with --preflight-file"
        )
    if not preflight and not preflight_file_arg:
        raise SystemExit("league release requires --preflight or a reviewed --preflight-file")

    v5_evidence_sha256s: dict[str, str] = {}
    if analysis_version == "v5":
        for locator in (
            ANALYSIS_WINDOW_V5_INJURY_AUDIT_LOCATOR,
            ANALYSIS_WINDOW_V5_EXPOSURE_EVIDENCE_LOCATOR,
            ANALYSIS_WINDOW_V5_SQL_RECONCILIATION_LOCATOR,
            ANALYSIS_WINDOW_V5_CANDIDATE_PERFORMANCE_LOCATOR,
        ):
            evidence_path = Path(locator)
            if not evidence_path.is_file():
                raise SystemExit(f"V5 row-level evidence is missing: {locator}")
            v5_evidence_sha256s[locator] = hashlib.sha256(
                evidence_path.read_bytes()
            ).hexdigest()

    provenance = run_provenance()
    dirty_release_override = (
        os.environ.get("PIPELINE_ALLOW_DIRTY_RELEASE_LEAGUE", "").strip() == "1"
    )
    dirty_override_allowed_paths = sorted(
        {
            path.strip()
            for path in os.environ.get(
                "PIPELINE_ALLOWED_DIRTY_RELEASE_LEAGUE_PATHS", ""
            ).split(",")
            if path.strip()
        }
    )
    current_dirty_paths: list[str] = []
    if provenance["code_version"].endswith("-dirty") and not dirty_release_override:
        raise SystemExit(
            "release-league refuses to run from an uncommitted working tree "
            f"(code_version={provenance['code_version']}); set "
            "PIPELINE_ALLOW_DIRTY_RELEASE_LEAGUE=1 only for an explicitly "
            "authorised concurrent-work override"
        )
    if provenance["code_version"].endswith("-dirty"):
        current_dirty_paths = sorted(set(dirty_worktree_paths()))
        validate_dirty_release_override(
            current_dirty_paths,
            dirty_override_allowed_paths,
        )
        print(
            "WARNING: release-league is using the explicitly authorised "
            "concurrent-work dirty-tree override",
            file=sys.stderr,
        )

    if analysis_version == "v6" and year2_release_contract is not None:
        required_migrations = list(year2_release_contract.required_migrations)
    elif analysis_version == "v5":
        required_migrations = [
            INJURY_MASTER_LINEAGE_MIGRATION_VERSION,
            LINEAGE_RESTATED_REPORTING_MIGRATION_VERSION,
            LINEAGE_V4_CANDIDATE_FAST_PATH_MIGRATION_VERSION,
            OSIICS_EXACT_REPORTING_CLASSIFICATION_MIGRATION_VERSION,
            INCREMENTAL_CLASSIFICATION_BUNDLE_MIGRATION_VERSION,
            ANALYSIS_WINDOW_REPORTING_V5_MIGRATION_VERSION,
            ANALYSIS_WINDOW_V5_CANDIDATE_OPTIMIZATION_MIGRATION_VERSION,
            ANALYSIS_WINDOW_V5_SHARED_COHORT_SNAPSHOT_MIGRATION_VERSION,
            ANALYSIS_WINDOW_V5_CANDIDATE_SNAPSHOT_MIGRATION_VERSION,
            ANALYSIS_WINDOW_V5_COVERAGE_SNAPSHOT_MIGRATION_VERSION,
            # The candidate views now read from the contact snapshot layer, and
            # the reader pair is what makes the new section visible to
            # web_reader. Requiring both refuses a half-applied change that
            # would either release without the section or fail every dashboard
            # closed.
            CONTACT_DISTRIBUTION_V5_MIGRATION_VERSION,
            CONTACT_DISTRIBUTION_READER_V4_MIGRATION_VERSION,
        ]
    elif analysis_version == "v4":
        required_migrations = [
            INJURY_MASTER_LINEAGE_MIGRATION_VERSION,
            LINEAGE_RESTATED_REPORTING_MIGRATION_VERSION,
            LINEAGE_V4_CANDIDATE_FAST_PATH_MIGRATION_VERSION,
            OSIICS_EXACT_REPORTING_CLASSIFICATION_MIGRATION_VERSION,
            INCREMENTAL_CLASSIFICATION_BUNDLE_MIGRATION_VERSION,
        ]
    else:
        required_migration = (
            SEASON_BOUND_REPORTING_MIGRATION_VERSION
            if analysis_version == "v3"
            else ADJUDICATED_REPORTING_CLASSIFICATION_MIGRATION_VERSION
        )
        required_migrations = [required_migration]
    if uses_osiics_successor:
        required_migrations.extend([
            OSIICS_EXACT_REPORTING_CLASSIFICATION_MIGRATION_VERSION,
            INCREMENTAL_CLASSIFICATION_BUNDLE_MIGRATION_VERSION,
        ])
    if not preflight:
        required_migrations.append(REVIEWED_BUNDLE_PAYLOAD_VALIDATION_MIGRATION_VERSION)
    if analysis_version == "v6" and year2_release_contract is not None:
        assert_checksum_bound_release_migrations(year2_release_contract, "V6 league release")
    else:
        migration_rows = query_sql(
            "select version from supabase_migrations.schema_migrations "
            "where version = any(array["
            + ", ".join(f"'{version}'" for version in required_migrations)
            + "])"
        )
        if {row["version"] for row in migration_rows} != set(required_migrations):
            raise SystemExit(
                "release-league requires tracked migrations "
                + ", ".join(required_migrations)
            )
    if analysis_version == "v6" and year2_release_contract is not None:
        availability_params = SqlParams()
        availability_rows = query_sql(
            f"select analysis.release_contract_candidates_available_v1("
            f"{availability_params.text(season)}::text, "
            f"{availability_params.text(analysis_version)}::text, "
            f"{availability_params.text(classification_view_version)}::text, "
            f"{availability_params.text(cohort_view_version)}::text"
            f") as is_available",
            availability_params.values,
        )
        if len(availability_rows) != 1 or availability_rows[0].get("is_available") is not True:
            raise SystemExit(
                "V6 release contract is unavailable: the exact tuple and all "
                "registered candidate and semantic relations must exist"
            )

    workflow_started = time.perf_counter()
    candidate_started = time.perf_counter()
    candidate = (
        load_v6_retained_league_rollback_candidate(
            season=season,
            rollback_of_release_id=rollback_of_release_id,
        )
        if rollback_of_release_id
        else load_league_release_candidate(
            season=season,
            analysis_version=analysis_version,
            classification_view_version=classification_view_version,
            cohort_view_version=cohort_view_version,
            league_candidate_view=league_candidate_view,
            team_candidate_view=team_candidate_view,
        )
    )
    candidate_query_ms = round((time.perf_counter() - candidate_started) * 1000, 3)
    if not isinstance(candidate.get("dashboard"), dict):
        raise SystemExit(
            f"release-league expected one complete league payload for {season!r}, "
            "found an incomplete payload"
        )
    dashboard = candidate["dashboard"]
    if analysis_version == "v6":
        assert_v6_public_dashboard_contract(dashboard, "league dashboard")
    classification_evidence_sha256 = clean_text(
        candidate.get("classification_evidence_sha256")
    ) or None
    cohort_evidence_sha256 = clean_text(
        candidate.get("cohort_evidence_sha256")
    ) or None
    if classification_view_version != "v2" and not classification_evidence_sha256:
        raise SystemExit("accepted reporting classification evidence is missing")
    if cohort_view_version != "v2" and not cohort_evidence_sha256:
        raise SystemExit("accepted cohort evidence is missing")
    if analysis_version in {"v3", "v4", "v5", "v6"} and not rollback_of_release_id:
        if analysis_version == "v6" and year2_release_contract is not None:
            semantic_cohort_view = year2_release_contract.injury_cohort_view
            semantic_monthly_view = year2_release_contract.league_monthly_view
            semantic_summary_view = year2_release_contract.league_summary_view
            if not all((semantic_cohort_view, semantic_monthly_view, semantic_summary_view)):
                raise SystemExit("V6 release contract lacks required semantic relations")
            semantic_missing_error = "V6 semantic reconciliation returned no row"
            semantic_mismatch_error = (
                "V6 cohort, headline, or monthly reconciliation failed"
            )
        elif analysis_version == "v5":
            semantic_cohort_view = "analysis.analysis_window_injury_cohort_v5"
            semantic_monthly_view = "analysis.analysis_window_league_monthly_v5"
            semantic_summary_view = "analysis.analysis_window_league_summary_v5"
            semantic_missing_error = "analysis-window semantic reconciliation returned no row"
            semantic_mismatch_error = (
                "analysis-window cohort, headline, or monthly reconciliation failed"
            )
        elif analysis_version == "v4":
            semantic_cohort_view = "analysis.lineage_injury_cohort_v1"
            semantic_monthly_view = "analysis.lineage_league_monthly_v1"
            semantic_summary_view = "analysis.lineage_league_summary_v1"
            semantic_missing_error = "lineage semantic reconciliation returned no row"
            semantic_mismatch_error = (
                "lineage cohort, headline, or monthly reconciliation failed"
            )
        else:
            semantic_cohort_view = "analysis.injury_cohort_by_build_season_bound_v3"
            semantic_monthly_view = "analysis.season_bound_league_monthly_v3"
            semantic_summary_view = "analysis.season_bound_league_summary_v3"
            semantic_missing_error = "season-bound semantic reconciliation returned no row"
            semantic_mismatch_error = (
                "season-bound cohort, headline, or monthly reconciliation failed"
            )
        semantic_params = SqlParams()
        semantic_rows = query_sql(
            f"""
            with cohort as (
              select c.*
              from {semantic_cohort_view} c
              join {member_view} m
                using (curated_build_id, team_key, season)
              where c.season = {semantic_params.text(season)}
            ), monthly as (
              select sum(exposure_hours) as exposure_hours,
                     coalesce(sum(time_loss_injuries), 0) as time_loss_injuries
              from {semantic_monthly_view}
              where season = {semantic_params.text(season)}
            ), denominator as (
              select exposure_hours
              from {semantic_summary_view}
              where season = {semantic_params.text(season)}
            )
            select
              (select count(*) from cohort) as recorded_injuries,
              (select count(*) from cohort where is_time_loss) as time_loss_injuries,
              (select count(*) from cohort where is_undated) as undated_injuries,
              (select count(*) from cohort where is_time_loss and not is_undated)
                as dated_time_loss_injuries,
              monthly.time_loss_injuries as monthly_time_loss_injuries,
              denominator.exposure_hours,
              monthly.exposure_hours as monthly_exposure_hours
            from monthly cross join denominator
            """,
            semantic_params.values,
        )
        if len(semantic_rows) != 1:
            raise SystemExit(semantic_missing_error)
        semantic = semantic_rows[0]
        first_mismatch = first_release_payload_mismatch(dashboard, semantic)
        if first_mismatch is not None:
            label, actual, expected = first_mismatch
            raise SystemExit(
                f"{semantic_mismatch_error}: first_mismatch={label}, "
                f"actual={actual!r}, expected={expected!r}"
            )
    classification_adjudications: list[dict[str, Any]] = []
    if classification_view_version != "v2":
        if analysis_version == "v6" and year2_release_contract is not None:
            # Year 2 carries only the approved catalogue and conservative
            # inference rule, never the prior season's source-row ledger.
            classification_adjudications = [{
                "rule_version": "reporting_classification_2026-07-22_v2",
                "rule_evidence_locator": "docs/evidence/urc_2025_26_classification_rule.json",
                "rule_evidence_sha256": "e898320fc5fa8cdfbf4fde4382d1ade62c87fe2dbef820ecf72b557bfb07cd5f",
                "mapping_catalogue_projection_sha256": "79767a9fc4212309c8fa01749ddf47541a251e467897268a9c8edeb4265553ff",
                "mapping_catalogue_row_count": 52,
                "multi_type_catalogue_projection_sha256": "d7aa844af7a4e6a53072f90e129da5357dfd4523aef415ecf84fd447702db55a",
                "multi_type_catalogue_row_count": 1,
                "application_scope": "catalogue_and_conservative_inference_only",
                "year1_row_adjudications": "not_carried_forward",
            }]
        else:
            rule_params = SqlParams()
            if classification_view_version == "reporting_classification_2026-07-22_v2":
                adjudication_filter = """
                  (rule_version = 'reporting_classification_2026-07-20_v1'
                    and adjudication_ref in ('IA-02', 'ACL-01'))
                  or (rule_version = 'reporting_classification_2026-07-22_v2'
                    and adjudication_ref = 'OSIICS-01')
                """
                expected_refs = {"IA-02", "ACL-01", "OSIICS-01"}
            else:
                adjudication_filter = """
                  rule_version = 'reporting_classification_2026-07-20_v1'
                  and adjudication_ref in ('IA-02', 'ACL-01')
                """
                expected_refs = {"IA-02", "ACL-01"}
            classification_adjudications = query_sql(
                f"""
                select adjudication_ref, rule_version, evidence_sha256,
                       workbook_sha256, evidence_manifest_sha256, reviewer,
                       workbook_snapshot_locator, migration_version, migration_sha256,
                       rationale
                from audit.rule_adjudications
                where {adjudication_filter}
                order by adjudication_ref
                """,
                rule_params.values,
            )
            if len(classification_adjudications) != len(expected_refs) or {
                row["adjudication_ref"] for row in classification_adjudications
            } != expected_refs:
                raise SystemExit("exact reporting-classification evidence is incomplete")
    cohort_adjudications: list[dict[str, Any]] = []
    if cohort_view_version != "v2":
        cohort_params = SqlParams()
        if analysis_version == "v6" and year2_release_contract is not None:
            if not all((
                year2_release_contract.cohort_adjudication_ref,
                year2_release_contract.cohort_evidence_locator,
                year2_release_contract.cohort_evidence_sha256,
                year2_release_contract.required_migrations,
            )):
                raise SystemExit("V6 release contract lacks immutable cohort evidence")
            cohort_adjudication_filter = f"""
              and adjudication_ref = {cohort_params.text(year2_release_contract.cohort_adjudication_ref)}
              and evidence_sha256 = {cohort_params.text(year2_release_contract.cohort_evidence_sha256)}
              and evidence_locator = {cohort_params.text(year2_release_contract.cohort_evidence_locator)}
              and reviewer = 'Abdel Babiker'
              and migration_version = {cohort_params.text(year2_release_contract.required_migrations[0])}
            """
        elif analysis_version == "v5":
            cohort_adjudication_filter = f"""
              and adjudication_ref = 'ANALYSIS-WINDOW-01'
              and evidence_sha256 = {cohort_params.text(ANALYSIS_WINDOW_V5_EVIDENCE_SHA256)}
              and evidence_locator = {cohort_params.text(ANALYSIS_WINDOW_V5_EVIDENCE_LOCATOR)}
              and reviewer = 'Abdel Babiker'
              and migration_version = {cohort_params.text(ANALYSIS_WINDOW_REPORTING_V5_MIGRATION_VERSION)}
            """
        elif analysis_version == "v4":
            cohort_adjudication_filter = f"""
              and adjudication_ref = 'LINEAGE-01'
              and evidence_locator = {cohort_params.text('docs/evidence/lineage_cohort_2024-25.json')}
              and reviewer = 'Abdel Babiker'
              and migration_version = {cohort_params.text(LINEAGE_RESTATED_REPORTING_MIGRATION_VERSION)}
            """
        else:
            cohort_adjudication_filter = """
              and adjudication_ref = 'COHORT-01'
            """
        cohort_adjudications = query_sql(
            f"""
            select adjudication_ref, cohort_view_version, season, decision,
                   evidence_sha256, evidence_locator, reviewer, migration_version
            from audit.reporting_cohort_rule_adjudications_v3
            where cohort_view_version = {cohort_params.text(cohort_view_version)}
              and season = {cohort_params.text(season)}
              {cohort_adjudication_filter}
            """,
            cohort_params.values,
        )
        if len(cohort_adjudications) != 1:
            raise SystemExit("exact cohort adjudication evidence is incomplete")
    team_payloads = candidate.get("team_payloads")
    if not isinstance(team_payloads, list):
        team_payloads = []
    if len(team_payloads) != 16 or any(
        not isinstance(row, dict) or not isinstance(row.get("dashboard"), dict)
        for row in team_payloads
    ):
        raise SystemExit(
            f"release-league requires 16 complete team dashboard payloads, found {len(team_payloads)}"
        )
    if analysis_version == "v6":
        for row in team_payloads:
            if not isinstance(row.get("team_key"), str) or not row["team_key"]:
                raise SystemExit("V6 league candidate team payload lacks a team key")
            assert_v6_public_dashboard_contract(row["dashboard"], row["team_key"])
    if analysis_version in {"v3", "v4", "v5", "v6"} and any(
        "injury_cohort_filters" in row["dashboard"].get("coverage", {})
        for row in team_payloads
    ):
        raise SystemExit("season-bound team payload retained stale V2 cohort filters")

    if rollback_of_release_id:
        members = [
            {
                "team_key": row["team_key"],
                "team_release_id": row["team_release_id"],
                "curated_build_id": row["curated_build_id"],
            }
            for row in sorted(team_payloads, key=lambda item: item["team_key"])
        ]
    else:
        member_params = SqlParams()
        members = query_sql(
            f"select team_key, team_release_id::text, curated_build_id::text "
            f"from {member_view} "
            f"where season = {member_params.text(season)} order by team_key",
            member_params.values,
        )
    if len(members) != 16:
        raise SystemExit(f"release-league requires 16 approved member releases, found {len(members)}")
    member_by_team = {member["team_key"]: member for member in members}
    if set(member_by_team) != {row["team_key"] for row in team_payloads}:
        raise SystemExit("league member roster and team dashboard payload roster differ")
    for row in team_payloads:
        member = member_by_team[row["team_key"]]
        if (
            row["team_release_id"] != member["team_release_id"]
            or row["curated_build_id"] != member["curated_build_id"]
        ):
            raise SystemExit(
                f"team dashboard payload identity differs from league member for {row['team_key']}"
            )
    member_input_hash = sha256_json(members)
    public_bundle = {
        "schema_version": "urc_dashboard_bundle_v2",
        "season": season,
        "league": dashboard,
        "teams": [
            {"team_key": row["team_key"], "dashboard": row["dashboard"]}
            for row in team_payloads
        ],
    }
    bundle_payload_sha256 = clean_text(candidate.get("bundle_payload_sha256"))
    league_payload_sha256 = clean_text(candidate.get("league_payload_sha256"))
    team_payload_sha256s = candidate.get("team_payload_sha256s")
    canonical_bundle_json = candidate.get("bundle_payload_json")
    if (
        len(bundle_payload_sha256) != 64
        or len(league_payload_sha256) != 64
        or not isinstance(team_payload_sha256s, dict)
        or len(team_payload_sha256s) != 16
        or not isinstance(canonical_bundle_json, str)
    ):
        raise SystemExit("release-league canonical database payload hashes are incomplete")
    preflight_league_sha256 = league_payload_sha256
    preflight_bundle_sha256 = bundle_payload_sha256
    reconciliation_ms = round(
        (time.perf_counter() - workflow_started) * 1000 - candidate_query_ms,
        3,
    )
    timings_ms = {
        "candidate_query": candidate_query_ms,
        "reconciliation": max(reconciliation_ms, 0),
        "preflight_total": round((time.perf_counter() - workflow_started) * 1000, 3),
    }
    manifest_candidate_views = (
        {
            "league": "reporting.league_release_payloads_v6",
            "teams": "reporting.team_dashboard_payloads_v2",
        }
        if rollback_of_release_id
        else {"league": league_candidate_view, "teams": team_candidate_view}
    )
    manifest_required_migrations: list[object] = (
        [
            {"version": item.version, "name": item.name, "sha256": item.sha256}
            for item in year2_release_contract.required_migration_contracts
        ]
        if analysis_version == "v6" and year2_release_contract is not None
        else list(required_migrations)
    )

    if preflight:
        output_arg = clean_text(args.output or "")
        output_path = (
            Path(output_arg)
            if output_arg
            else Path("data/reporting")
            / f"urc_dashboard_bundle_{season}_{preflight_bundle_sha256[:16]}_preflight.json"
        )
        if Path("content/reporting").resolve() in output_path.resolve().parents:
            raise SystemExit("league preflight output must stay outside content/reporting")
        write_text_atomic(output_path, canonical_bundle_json + "\n")
        manifest_path = Path(f"{output_path}.manifest.json")
        preflight_manifest = {
            "schema_version": "urc_league_release_preflight_manifest_v1",
            "status": "preflight",
            "season": season,
            "release_tuple": {
                "analysis_version": analysis_version,
                "classification_view_version": classification_view_version,
                "cohort_view_version": cohort_view_version,
            },
            "candidate_views": manifest_candidate_views,
            "required_migrations": manifest_required_migrations,
            "member_count": len(members),
            "member_input_hash": member_input_hash,
            "league_payload_sha256": league_payload_sha256,
            "bundle_payload_sha256": bundle_payload_sha256,
            "preflight_file_sha256": sha256_file(output_path),
            "team_payload_sha256s": team_payload_sha256s,
            "evidence_sha256s": v5_evidence_sha256s,
            "classification_evidence_sha256": classification_evidence_sha256,
            "cohort_evidence_sha256": cohort_evidence_sha256,
            "classification_adjudications": classification_adjudications,
            "classification_adjudications_sha256": sha256_json(
                classification_adjudications
            ),
            "cohort_adjudications": cohort_adjudications,
            "cohort_adjudications_sha256": sha256_json(cohort_adjudications),
            "provenance": {
                "code_version": provenance["code_version"],
                "dependency_lock_hash": provenance["dependency_lock_hash"],
            },
            "dirty_worktree_paths": current_dirty_paths,
            "dirty_worktree_allowed_paths": dirty_override_allowed_paths,
            "candidate_assembly_reads": 1,
            "reconciliation": {
                "migration_prerequisites": "passed",
                "semantic_sections": "passed",
                "member_roster": "passed",
                "candidate_hashes": "passed",
            },
        }
        if analysis_version == "v6" and year2_release_contract is not None:
            preflight_manifest.update({
                "rollback_of_release_id": rollback_of_release_id or None,
                "rollback_replaces_release_id": (
                    candidate.get("replaces_release_id") if rollback_of_release_id else None
                ),
                "local_evidence_files": year2_release_local_evidence_records(
                    year2_release_contract
                ),
            })
        else:
            # Frozen Year 1 preflights retain their exact historical key set.
            preflight_manifest["timings_ms"] = timings_ms
        write_json_atomic(manifest_path, preflight_manifest)
        print(
            json.dumps(
                {
                    "status": "preflight",
                    "season": season,
                    "member_count": len(members),
                    "member_input_hash": member_input_hash,
                    "preflight_league_sha256": preflight_league_sha256,
                    "preflight_bundle_sha256": preflight_bundle_sha256,
                    "database_league_payload_sha256": league_payload_sha256,
                    "database_bundle_payload_sha256": bundle_payload_sha256,
                    "release_tuple": {
                        "analysis_version": analysis_version,
                        "classification_view_version": classification_view_version,
                        "cohort_view_version": cohort_view_version,
                    },
                    "candidate_assembly_reads": 1,
                    "reconciliation": {
                        "migration_prerequisites": "passed",
                        "semantic_sections": "passed",
                        "member_roster": "passed",
                        "candidate_hashes": "passed",
                    },
                    "timings_ms": timings_ms,
                    "output_path": str(output_path),
                    "manifest_path": str(manifest_path),
                },
                indent=2,
            )
        )
        return

    preflight_path = Path(preflight_file_arg)
    if not preflight_path.exists():
        raise SystemExit(f"league preflight file not found: {preflight_path}")
    reviewed_bytes = preflight_path.read_bytes()
    reviewed_sha256 = hashlib.sha256(reviewed_bytes).hexdigest()
    try:
        reviewed_bundle = json.loads(reviewed_bytes)
    except json.JSONDecodeError as exc:
        raise SystemExit(f"league preflight file is invalid JSON: {preflight_path}") from exc
    reviewed_hash_params = SqlParams()
    reviewed_bundle_hash_sql = (
        f"reporting.canonical_jsonb_sha256_v1(({reviewed_hash_params.text(reviewed_bytes.decode('utf-8'))})::jsonb)"
        if analysis_version == "v6"
        else f"encode(digest(convert_to(({reviewed_hash_params.text(reviewed_bytes.decode('utf-8'))})::jsonb::text, 'UTF8'), 'sha256'), 'hex')"
    )
    reviewed_hash_rows = query_sql(
        f"select {reviewed_bundle_hash_sql} as bundle_payload_sha256",
        reviewed_hash_params.values,
    )
    reviewed_diffs = diff_json_documents(reviewed_bundle, public_bundle)
    reviewed_canonical_sha256 = (
        clean_text(reviewed_hash_rows[0].get("bundle_payload_sha256"))
        if len(reviewed_hash_rows) == 1 else ""
    )
    if reviewed_canonical_sha256 != bundle_payload_sha256:
        paths = ", ".join(diff["path"] for diff in reviewed_diffs[:10]) or "$"
        raise SystemExit(
            f"dashboard bundle differs from reviewed preflight canonical hash ({paths})"
        )
    reviewed_manifest_path = Path(f"{preflight_path}.manifest.json")
    reviewed_manifest_sha256: str | None = None
    if reviewed_manifest_path.exists():
        if allow_legacy_preflight_without_manifest:
            raise SystemExit(
                "--allow-legacy-preflight-without-manifest cannot be used "
                "when the preflight manifest exists"
            )
        try:
            reviewed_manifest = json.loads(reviewed_manifest_path.read_text())
        except json.JSONDecodeError as exc:
            raise SystemExit(
                f"league preflight manifest is invalid JSON: {reviewed_manifest_path}"
            ) from exc
        expected_tuple = {
            "analysis_version": analysis_version,
            "classification_view_version": classification_view_version,
            "cohort_view_version": cohort_view_version,
        }
        expected_candidate_views = manifest_candidate_views
        preflight_required_migrations: list[object] = (
            manifest_required_migrations
            if analysis_version == "v6"
            else [
                version
                for version in required_migrations
                if version != REVIEWED_BUNDLE_PAYLOAD_VALIDATION_MIGRATION_VERSION
            ]
        )
        manifest_checks = [
            (
                "schema_version",
                reviewed_manifest.get("schema_version"),
                "urc_league_release_preflight_manifest_v1",
            ),
            ("status", reviewed_manifest.get("status"), "preflight"),
            ("season", reviewed_manifest.get("season"), season),
            (
                "release_tuple",
                reviewed_manifest.get("release_tuple"),
                expected_tuple,
            ),
            (
                "candidate_views",
                reviewed_manifest.get("candidate_views"),
                expected_candidate_views,
            ),
            (
                "required_migrations",
                reviewed_manifest.get("required_migrations"),
                preflight_required_migrations,
            ),
            (
                "member_count",
                reviewed_manifest.get("member_count"),
                len(members),
            ),
            (
                "member_input_hash",
                reviewed_manifest.get("member_input_hash"),
                member_input_hash,
            ),
            (
                "league_payload_sha256",
                clean_text(reviewed_manifest.get("league_payload_sha256")),
                league_payload_sha256,
            ),
            (
                "bundle_payload_sha256",
                clean_text(reviewed_manifest.get("bundle_payload_sha256")),
                bundle_payload_sha256,
            ),
            (
                "preflight_file_sha256",
                clean_text(reviewed_manifest.get("preflight_file_sha256")),
                reviewed_sha256,
            ),
            (
                "team_payload_sha256s",
                reviewed_manifest.get("team_payload_sha256s"),
                team_payload_sha256s,
            ),
            (
                "evidence_sha256s",
                reviewed_manifest.get("evidence_sha256s"),
                v5_evidence_sha256s,
            ),
            (
                "classification_evidence_sha256",
                reviewed_manifest.get("classification_evidence_sha256"),
                classification_evidence_sha256,
            ),
            (
                "cohort_evidence_sha256",
                reviewed_manifest.get("cohort_evidence_sha256"),
                cohort_evidence_sha256,
            ),
            (
                "classification_adjudications",
                reviewed_manifest.get("classification_adjudications"),
                classification_adjudications,
            ),
            (
                "classification_adjudications_sha256",
                reviewed_manifest.get("classification_adjudications_sha256"),
                sha256_json(classification_adjudications),
            ),
            (
                "cohort_adjudications",
                reviewed_manifest.get("cohort_adjudications"),
                cohort_adjudications,
            ),
            (
                "cohort_adjudications_sha256",
                reviewed_manifest.get("cohort_adjudications_sha256"),
                sha256_json(cohort_adjudications),
            ),
            (
                "provenance",
                reviewed_manifest.get("provenance"),
                {
                    "code_version": provenance["code_version"],
                    "dependency_lock_hash": provenance["dependency_lock_hash"],
                },
            ),
            (
                "dirty_worktree_paths",
                reviewed_manifest.get("dirty_worktree_paths"),
                current_dirty_paths,
            ),
            (
                "dirty_worktree_allowed_paths",
                reviewed_manifest.get("dirty_worktree_allowed_paths"),
                dirty_override_allowed_paths,
            ),
            (
                "candidate_assembly_reads",
                reviewed_manifest.get("candidate_assembly_reads"),
                1,
            ),
            (
                "reconciliation",
                reviewed_manifest.get("reconciliation"),
                {
                    "migration_prerequisites": "passed",
                    "semantic_sections": "passed",
                    "member_roster": "passed",
                    "candidate_hashes": "passed",
                },
            ),
        ]
        if analysis_version == "v6" and year2_release_contract is not None:
            manifest_checks.extend([
                (
                    "rollback_of_release_id",
                    reviewed_manifest.get("rollback_of_release_id"),
                    rollback_of_release_id or None,
                ),
                (
                    "rollback_replaces_release_id",
                    reviewed_manifest.get("rollback_replaces_release_id"),
                    candidate.get("replaces_release_id") if rollback_of_release_id else None,
                ),
                (
                    "local_evidence_files",
                    reviewed_manifest.get("local_evidence_files"),
                    year2_release_local_evidence_records(year2_release_contract),
                ),
            ])
        else:
            manifest_checks.append(
                ("timings_ms", reviewed_manifest.get("timings_ms"), timings_ms)
            )
        expected_manifest_keys = {field for field, _actual, _expected in manifest_checks}
        if set(reviewed_manifest) != expected_manifest_keys:
            raise SystemExit(
                "league preflight manifest has missing or unknown fields: "
                f"actual={sorted(reviewed_manifest)}, expected={sorted(expected_manifest_keys)}"
            )
        first_manifest_mismatch = next(
            (
                (field, actual, expected)
                for field, actual, expected in manifest_checks
                if actual != expected
            ),
            None,
        )
        if first_manifest_mismatch is not None:
            field, actual, expected = first_manifest_mismatch
            raise SystemExit(
                "league preflight manifest differs from the reviewed candidate: "
                f"first_mismatch={field}, actual={actual!r}, expected={expected!r}"
            )
        reviewed_manifest_sha256 = sha256_file(reviewed_manifest_path)
    elif not allow_legacy_preflight_without_manifest:
        raise SystemExit(
            "league preflight manifest is required; regenerate the preflight "
            "with this workflow or use --allow-legacy-preflight-without-manifest "
            "only for an explicitly approved historical candidate"
        )

    existing_params = SqlParams()
    existing = query_sql(
        f"select c.release_id::text, r.release_label from reporting.league_release_context_v2 c "
        f"join reporting.aggregate_releases r on r.id = c.release_id "
        f"where c.season = {existing_params.text(season)} and r.status = 'approved'",
        existing_params.values,
    )
    predecessor: dict[str, Any] | None = None
    if existing:
        if len(existing) != 1:
            raise SystemExit("expected exactly one approved predecessor bundle")
        if not previous_bundle_file_arg:
            raise SystemExit(
                "an approved bundle already exists; --previous-bundle-file is required for re-release"
            )
        previous_path = Path(previous_bundle_file_arg)
        if not previous_path.exists():
            raise SystemExit(f"previous bundle snapshot not found: {previous_path}")
        previous_bytes = previous_path.read_bytes()
        try:
            previous_document = json.loads(previous_bytes)
        except json.JSONDecodeError as exc:
            raise SystemExit("previous bundle snapshot is invalid JSON") from exc
        approved_previous, predecessor = current_league_bundle_snapshot(season)
        previous_diffs = diff_json_documents(previous_document, approved_previous)
        if previous_diffs:
            paths = ", ".join(diff["path"] for diff in previous_diffs[:10])
            raise SystemExit(
                f"--previous-bundle-file does not match the approved predecessor ({paths})"
            )
        predecessor["snapshot_sha256"] = hashlib.sha256(previous_bytes).hexdigest()
    elif previous_bundle_file_arg:
        raise SystemExit(
            "--previous-bundle-file is only valid when an approved predecessor exists"
        )
    if rollback_of_release_id:
        if predecessor is None or predecessor["release_id"] != candidate.get("replaces_release_id"):
            raise SystemExit(
                "V6 rollback target is no longer the exact predecessor of the current approved bundle"
            )

    label_prefix = f"urc-{season}-{analysis_version}-{bundle_payload_sha256[:12]}"
    attempt_params = SqlParams()
    attempt_rows = query_sql(
        f"select count(*)::integer as attempts from reporting.aggregate_releases "
        f"where release_label like {attempt_params.text(label_prefix + '-a%')}",
        attempt_params.values,
    )
    attempt_number = int(attempt_rows[0]["attempts"]) + 1
    label = f"{label_prefix}-a{attempt_number}"
    generated_at = clean_text(dashboard.get("generated_at"))
    if not generated_at:
        raise SystemExit("league payload generated_at is missing")
    v6_validation_migration: dict[str, str] | None = None
    if analysis_version == "v6" and year2_release_contract is not None:
        validation_contract = year2_release_contract.required_migration_contracts[-1]
        v6_validation_migration = {
            "version": validation_contract.version,
            "name": validation_contract.name,
            "sha256": validation_contract.sha256,
        }
    release_parameters = {
        "analysis_version": analysis_version,
        "classification_view_version": classification_view_version,
        "classification_evidence_sha256": classification_evidence_sha256,
        "classification_adjudications": classification_adjudications,
        "cohort_view_version": cohort_view_version,
        "cohort_evidence_sha256": cohort_evidence_sha256,
        "cohort_adjudications": cohort_adjudications,
        "member_count": len(members),
        "member_input_hash": member_input_hash,
        "member_release_ids": [member["team_release_id"] for member in members],
        "member_curated_build_ids": [member["curated_build_id"] for member in members],
        "hash_algorithm": "postgres_jsonb_text_sha256",
        "team_dashboard_payload_sha256s": team_payload_sha256s,
        "league_dashboard_payload_sha256": league_payload_sha256,
        "bundle_payload_sha256": bundle_payload_sha256,
        "preflight_json_sha256": preflight_bundle_sha256,
        "reviewed_preflight_sha256": reviewed_sha256,
        "reviewed_preflight_manifest_sha256": reviewed_manifest_sha256,
        "legacy_preflight_without_manifest": (
            allow_legacy_preflight_without_manifest
        ),
        "preflight_reviewer": reviewer,
        "payload_hash_validation_migration": (
            v6_validation_migration
            if v6_validation_migration is not None
            else REVIEWED_BUNDLE_PAYLOAD_VALIDATION_MIGRATION_VERSION
        ),
        "payload_candidate_validation_migration": (
            v6_validation_migration
            if v6_validation_migration is not None
            else (
                INCREMENTAL_CLASSIFICATION_BUNDLE_MIGRATION_VERSION
                if uses_osiics_successor
                else (
                    CONTACT_DISTRIBUTION_V5_MIGRATION_VERSION
                    if analysis_version == "v5"
                    else (
                        LINEAGE_V4_CANDIDATE_FAST_PATH_MIGRATION_VERSION
                        if analysis_version == "v4"
                        else REVIEWED_BUNDLE_PAYLOAD_VALIDATION_MIGRATION_VERSION
                    )
                )
            )
        ),
        "analysis_window_v5_evidence_sha256s": v5_evidence_sha256s,
        "dirty_worktree_override": dirty_release_override,
        "dirty_worktree_allowed_paths": dirty_override_allowed_paths,
        "candidate_assembly_reads": 1,
        "promotion_candidate_validation_reads": 1,
        "preflight_timings_ms": timings_ms,
        "match_exposure_decision": "all_registered_season_fixtures_15_players_x_80_minutes_div_60",
    }
    if analysis_version == "v6":
        release_parameters["rollback_of_release_id"] = rollback_of_release_id or None
    if predecessor is not None:
        release_parameters.update({
            "predecessor_release_id": predecessor["release_id"],
            "predecessor_release_label": predecessor["release_label"],
            "predecessor_bundle_sha256": predecessor["bundle_sha256"],
            "predecessor_snapshot_sha256": predecessor["snapshot_sha256"],
        })

    params = SqlParams()
    predecessor_lock_sql = ""
    predecessor_retire_sql = ""
    if predecessor is not None:
        predecessor_lock_sql = f"""
          if not exists (
            select 1 from reporting.aggregate_releases
            where id = {params.text(predecessor['release_id'])}::uuid
              and release_label = {params.text(predecessor['release_label'])}
              and status = 'approved'
          ) then
            raise exception 'approved predecessor changed after preflight validation';
          end if;
        """
        predecessor_retire_sql = f"""
          update reporting.aggregate_releases
          set status = 'retired'
          where id = {params.text(predecessor['release_id'])}::uuid
            and status = 'approved';
        """
    if rollback_of_release_id:
        reviewed_member_guard_sql = f"""
        perform 1
        from curated.builds build
          join reviewed_league_members member on member.curated_build_id = build.id
        where build.season = {params.text(season)}
          and build.team_key = member.team_key
        order by member.team_key, build.id
        for update;
        if (
          select count(*)
          from curated.builds build
            join reviewed_league_members member on member.curated_build_id = build.id
          where build.season = {params.text(season)}
            and build.team_key = member.team_key
        ) <> 16 then
          raise exception 'retained rollback curated-build identities are incomplete';
        end if;
        """
        rollback_context_insert_sql = f"""
      insert into reporting.v6_league_rollback_context
        (release_id, rollback_of_release_id, replaces_release_id)
      select id, {params.text(rollback_of_release_id)}::uuid,
        {params.text(candidate['replaces_release_id'])}::uuid
      from current_league_release;
        """
        league_payload_insert_sql = f"""
      insert into reporting.league_release_payloads_v2
        (release_id, dashboard_payload)
      select id, {params.jsonb(dashboard)}
      from current_league_release;
        """
        team_payload_insert_sql = f"""
      insert into reporting.team_dashboard_payloads_v2
        (bundle_release_id, team_key, team_release_id, curated_build_id, dashboard_payload)
      select current_league_release.id, retained.team_key,
        retained.team_release_id::uuid, retained.curated_build_id::uuid,
        retained.dashboard
      from current_league_release,
        jsonb_to_recordset({params.jsonb(team_payloads)}) as retained(
          team_key text, team_release_id text, curated_build_id text, dashboard jsonb
        );
        """
    else:
        reviewed_member_guard_sql = f"""
        -- Build-curated supersedes the previous active row before inserting a
        -- replacement. Lock each reviewed row in the same roster order before
        -- the post-lock member-view comparison, so a concurrent rebuild must
        -- either finish first and make this promotion fail closed, or wait
        -- until the immutable bundle has been committed.
        perform 1
        from curated.builds build
          join reviewed_league_members member on member.curated_build_id = build.id
        where build.season = {params.text(season)}
          and build.team_key = member.team_key
        order by member.team_key, build.id
        for update;
        if (
          select count(*)
          from curated.builds build
            join reviewed_league_members member on member.curated_build_id = build.id
          where build.season = {params.text(season)}
            and build.team_key = member.team_key
            and build.status = 'active'
        ) <> 16 then
          raise exception 'reviewed active curated builds changed after preflight validation';
        end if;
        if exists (
          select 1 from reviewed_league_members member
          full join (
            select team_key, team_release_id, curated_build_id
            from {member_view}
            where season = {params.text(season)}
          ) live
            on live.team_key = member.team_key
           and live.team_release_id = member.team_release_id
           and live.curated_build_id = member.curated_build_id
          where member.team_key is null or live.team_key is null
        ) then
          raise exception 'reviewed bundle member identities changed after preflight validation';
        end if;
        """
        rollback_context_insert_sql = ""
        league_payload_insert_sql = f"""
      insert into reporting.league_release_payloads_v2
        (release_id, dashboard_payload)
      select current_league_release.id, candidate.dashboard
      from current_league_release
      join {league_candidate_view} candidate
        on candidate.season = {params.text(season)}
       and candidate.analysis_version = {params.text(analysis_version)}
       and candidate.classification_view_version = {params.text(classification_view_version)}
       and candidate.classification_evidence_sha256 is not distinct from
         {params.text(classification_evidence_sha256)}
       and candidate.cohort_view_version = {params.text(cohort_view_version)}
       and candidate.cohort_evidence_sha256 is not distinct from {params.text(cohort_evidence_sha256)};
        """
        team_payload_insert_sql = f"""
      insert into reporting.team_dashboard_payloads_v2
        (bundle_release_id, team_key, team_release_id, curated_build_id, dashboard_payload)
      select current_league_release.id, candidate.team_key, candidate.team_release_id,
        candidate.curated_build_id, candidate.dashboard
      from current_league_release
      join {team_candidate_view} candidate
        on candidate.season = {params.text(season)}
       and candidate.analysis_version = {params.text(analysis_version)}
       and candidate.classification_view_version = {params.text(classification_view_version)}
       and candidate.classification_evidence_sha256 is not distinct from
         {params.text(classification_evidence_sha256)}
       and candidate.cohort_view_version = {params.text(cohort_view_version)}
       and candidate.cohort_evidence_sha256 is not distinct from {params.text(cohort_evidence_sha256)}
      join reviewed_league_members expected
        on expected.team_key = candidate.team_key
       and expected.team_release_id = candidate.team_release_id
       and expected.curated_build_id = candidate.curated_build_id;
        """
    sql = f"""
      {protected_alias_scan_sql('league release gate')}

      do $$
      begin
        perform 1 from reporting.teams order by team_key for update;
        {predecessor_lock_sql}
        if exists (select 1 from reporting.aggregate_releases where release_label = {params.text(label)}) then
          raise exception 'immutable league release label already exists';
        end if;
      end $$;

      create temp table reviewed_league_members on commit drop as
      select team_key, team_release_id::uuid as team_release_id,
        curated_build_id::uuid as curated_build_id
      from jsonb_to_recordset({params.jsonb(members)}) as member(
        team_key text, team_release_id text, curated_build_id text
      );

      do $$
      declare
        member_count integer;
      begin
        select count(*) into member_count from reviewed_league_members;
        if member_count <> 16
           or (select count(distinct team_key) from reviewed_league_members) <> 16 then
          raise exception 'reviewed bundle must contain exactly 16 distinct members';
        end if;
        {reviewed_member_guard_sql}
      end $$;

      create temp table current_league_release on commit drop as
      with run as (
        insert into audit.pipeline_runs
          (command, team, season, status, input_hash, output_hash, parameters,
           code_version, dependency_lock_hash, operator)
        values (
          'release-league', 'URC Overall', {params.text(season)}, 'started',
          {params.text(member_input_hash)}, {params.text(bundle_payload_sha256)},
          {params.jsonb(release_parameters)}, {params.text(provenance['code_version'])},
          {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      ), step as (
        insert into audit.step_runs
          (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count,
           counts_by_team, input_hash, output_hash)
        select id, 'release_league_dashboard', {params.text(release_rule_version)},
          {params.text(release_reason_code)}, 16, 17,
          {params.jsonb({member['team_key']: 1 for member in members})},
          {params.text(member_input_hash)}, {params.text(bundle_payload_sha256)}
        from run
        returning pipeline_run_id
      ), release as (
        insert into reporting.aggregate_releases (release_label, status, pipeline_run_id)
        select {params.text(label)}, 'draft', id from run
        returning id, pipeline_run_id
      )
      select * from release;

      insert into reporting.league_release_context_v2
        (release_id, season, analysis_version, generated_at,
         expected_member_count, match_exposure_decision, decision_reviewer, decision_recorded_at,
         classification_view_version, classification_evidence_sha256,
         cohort_view_version, cohort_evidence_sha256)
      select id, {params.text(season)}, {params.text(analysis_version)}, {params.text(generated_at)}::timestamptz,
        16, 'all_registered_season_fixtures_15_players_x_80_minutes_div_60',
        'Abdel Babiker', {params.text(decision_recorded_at)}::date,
        {params.text(classification_view_version)}, {params.text(classification_evidence_sha256)},
        {params.text(cohort_view_version)}, {params.text(cohort_evidence_sha256)}
      from current_league_release;

      {rollback_context_insert_sql}

      insert into reporting.league_release_members_v2
        (release_id, team_key, team_release_id, curated_build_id)
      select current_league_release.id, member.team_key, member.team_release_id::uuid,
        member.curated_build_id::uuid
      from current_league_release,
        jsonb_to_recordset({params.jsonb(members)}) as member(
          team_key text, team_release_id text, curated_build_id text
        );

      {league_payload_insert_sql}

      {team_payload_insert_sql}

      {predecessor_retire_sql}

      update reporting.aggregate_releases r
      set status = 'approved', approved_at = now()
      from current_league_release current
      where r.id = current.id and r.status = 'draft';

      do $$
      declare
        published_league_count integer;
        published_team_count integer;
        published_v6_league_count integer;
        published_v6_team_count integer;
        published_v6_token_count integer;
        stored_league_hash text;
        stored_team_hashes jsonb;
        stored_bundle_hash text;
      begin
        if {params.text(analysis_version)} = 'v6' then
          select count(*) into published_v6_league_count
          from reporting.latest_league_dashboard_v6 d
          where d.season = {params.text(season)};
          if published_v6_league_count <> 1 then
            raise exception 'V6 public league reader must expose exactly one completed bundle';
          end if;
          select count(*) into published_v6_team_count
          from reporting.latest_team_dashboard_v6 d
          where d.season = {params.text(season)};
          if published_v6_team_count <> 16 then
            raise exception 'V6 public team reader must expose exactly sixteen completed members';
          end if;
          select count(*) into published_v6_token_count
          from reporting.latest_dashboard_cache_token_v2 token
          where token.season = {params.text(season)};
          if published_v6_token_count <> 1 then
            raise exception 'V6 public cache token must expose exactly one completed bundle';
          end if;
        else
          select count(*) into published_league_count
          from reporting.latest_league_dashboard_v2 d
          where d.season = {params.text(season)};
          if published_league_count <> 1 then
            raise exception 'published league dashboard bundle must expose exactly one league row';
          end if;

          select count(*) into published_team_count
          from reporting.latest_team_dashboard_v2 d
          where d.season = {params.text(season)};
          if published_team_count <> 16 then
            raise exception 'published team dashboard bundle must expose exactly 16 teams';
          end if;
        end if;

        select p.payload_sha256 into stored_league_hash
        from reporting.league_release_payloads_v2 p
        join current_league_release current on current.id = p.release_id;
        if stored_league_hash is distinct from {params.text(league_payload_sha256)} then
          raise exception 'stored league payload hash differs from the canonical candidate hash';
        end if;

        select jsonb_object_agg(p.team_key, p.payload_sha256 order by p.team_key)
        into stored_team_hashes
        from reporting.team_dashboard_payloads_v2 p
        join current_league_release current on current.id = p.bundle_release_id;
        if stored_team_hashes is distinct from {params.jsonb(team_payload_sha256s)} then
          raise exception 'stored team payload hashes differ from the canonical candidate hashes';
        end if;

        select encode(digest(convert_to(jsonb_build_object(
          'schema_version', 'urc_dashboard_bundle_v2',
          'season', context.season,
          'league', league.dashboard_payload,
          'teams', teams.dashboards
        )::text, 'UTF8'), 'sha256'), 'hex')
        into stored_bundle_hash
        from current_league_release current
        join reporting.league_release_context_v2 context on context.release_id = current.id
        join reporting.league_release_payloads_v2 league on league.release_id = current.id
        cross join lateral (
          select jsonb_agg(jsonb_build_object(
            'team_key', payload.team_key, 'dashboard', payload.dashboard_payload
          ) order by payload.team_key) as dashboards
          from reporting.team_dashboard_payloads_v2 payload
          where payload.bundle_release_id = current.id
        ) teams;
        if stored_bundle_hash is distinct from {params.text(bundle_payload_sha256)} then
          raise exception 'stored bundle payload hash differs from the canonical candidate hash';
        end if;
      end $$;

      update audit.pipeline_runs pr
      set status = 'succeeded', ended_at = now()
      from current_league_release current
      where pr.id = current.pipeline_run_id and pr.status = 'started';

      update audit.step_runs sr
      set ended_at = now()
      from current_league_release current
      where sr.pipeline_run_id = current.pipeline_run_id
        and sr.step_name = 'release_league_dashboard';
    """
    output_arg = clean_text(args.output or "")
    export_path = Path(output_arg) if output_arg else Path("content/reporting") / f"urc_dashboard_{season}.json"
    export_backup_path: Path | None = None
    if export_path.exists():
        export_backup_path = Path("data/reporting") / (
            f"{export_path.stem}_{label}_backup.json"
        )
        if export_backup_path.exists():
            raise SystemExit(f"league export backup already exists: {export_backup_path}")
        export_backup_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(export_path, export_backup_path)
    staged_export_path = Path("data/reporting") / f"{export_path.stem}_{label}_candidate.json"
    if staged_export_path.exists():
        raise SystemExit(f"staged league export already exists: {staged_export_path}")
    write_json_atomic(staged_export_path, dashboard)
    try:
        run_sql(sql, params.values)
    except BaseException:
        with contextlib.suppress(BaseException):
            record_failed_league_release_attempt(
                label=label, season=season, input_hash=member_input_hash,
                output_hash=bundle_payload_sha256, parameters=release_parameters,
                provenance=provenance, failure_stage="database_promotion_rolled_back",
            )
        raise
    v6_finalizer_command = ""
    if analysis_version == "v6":
        identity_params = SqlParams()
        identity_rows = query_sql(
            f"select release.id::text as release_id "
            f"from reporting.aggregate_releases release "
            f"join reporting.league_release_context_v2 context "
            f"on context.release_id = release.id "
            f"where release.release_label = {identity_params.text(label)} "
            f"and release.status = 'approved' "
            f"and context.season = {identity_params.text(season)} "
            f"and context.analysis_version = 'v6'",
            identity_params.values,
        )
        if len(identity_rows) != 1 or not clean_text(identity_rows[0].get("release_id")):
            raise SystemExit(
                "V6 league promotion succeeded, but the exact approved release identity "
                "could not be read back. Do not rerun promotion; recover the approved UUID "
                f"for label {label!r}, then use finalize-v6-league-release-local with "
                f"--release-label {label!r} and --preflight-file {str(preflight_path)!r}."
            )
        v6_finalizer_command = v6_local_finalizer_command(
            release_id=clean_text(identity_rows[0]["release_id"]),
            release_label=label,
            preflight_file=preflight_path,
        )
    try:
        export_path.parent.mkdir(parents=True, exist_ok=True)
        os.replace(staged_export_path, export_path)
    except BaseException as export_error:
        if analysis_version == "v6":
            raise SystemExit(
                "V6 league promotion succeeded, but the local league export "
                f"replacement failed. The approved append-only release {label!r} "
                f"was retained; repair the local league export path, then run exactly: "
                f"{v6_finalizer_command}. Historical releases were not re-approved."
            ) from export_error
        recovery_params = SqlParams()
        predecessor_id = predecessor["release_id"] if predecessor is not None else None
        run_sql(
            f"""
            do $$
            declare
              failed_release_id uuid;
              latest_approved_id uuid;
            begin
              perform 1 from reporting.teams order by team_key for update;

              select id into failed_release_id
              from reporting.aggregate_releases
              where release_label = {recovery_params.text(label)};

              select c.release_id into latest_approved_id
              from reporting.league_release_context_v2 c
              join reporting.aggregate_releases r on r.id = c.release_id
              where c.season = {recovery_params.text(season)} and r.status = 'approved'
              order by r.approved_at desc nulls last, r.created_at desc, r.id desc
              limit 1;

              if failed_release_id is not null and latest_approved_id = failed_release_id then
                update reporting.aggregate_releases set status = 'retired'
                where id = failed_release_id and status = 'approved';

                if {recovery_params.text(predecessor_id)} is not null
                  and not exists (
                    select 1 from reporting.league_release_context_v2 c
                    join reporting.aggregate_releases r on r.id = c.release_id
                    where c.season = {recovery_params.text(season)} and r.status = 'approved'
                  )
                then
                  update reporting.aggregate_releases set status = 'approved'
                  where id = {recovery_params.text(predecessor_id)}::uuid
                    and status = 'retired';
                end if;
              elsif failed_release_id is not null then
                update reporting.aggregate_releases set status = 'retired'
                where id = failed_release_id and status = 'approved';
              end if;

              update audit.pipeline_runs set status = 'failed', ended_at = now(),
                parameters = parameters || jsonb_build_object(
                  'recovery', case
                    when latest_approved_id = failed_release_id
                      then 'local_export_failed_predecessor_restored'
                    else 'local_export_failed_later_successor_preserved'
                  end)
              where id = (
                select pipeline_run_id from reporting.aggregate_releases
                where id = failed_release_id
              );
            end $$;
            """,
            recovery_params.values,
        )
        raise

    parity_started = time.perf_counter()
    try:
        parity_export = write_team_dashboard_parity_exports(
            season,
            expected_release_label=label,
            expected_bundle=public_bundle,
        )
    except BaseException as exc:
        if analysis_version == "v6":
            raise SystemExit(
                "V6 league promotion succeeded, but the mandatory 16-team parity "
                "export failed. The approved release was retained; repair the parity "
                f"export path, then run exactly: {v6_finalizer_command}. Error: {exc}"
            ) from exc
        raise SystemExit(
            "league promotion succeeded, but the mandatory 16-team parity "
            "export failed; rerun export-team-dashboards for the approved "
            f"{season} bundle: {exc}"
        ) from exc
    parity_export_ms = round((time.perf_counter() - parity_started) * 1000, 3)
    workflow_total_ms = round((time.perf_counter() - workflow_started) * 1000, 3)
    release_manifest_path = (
        Path("data/reporting") / f"{label}_release_manifest.json"
    )
    release_manifest = league_release_manifest_document(
        release_label=label,
        season=season,
        release_tuple={
            "analysis_version": analysis_version,
            "classification_view_version": classification_view_version,
            "cohort_view_version": cohort_view_version,
        },
        required_migrations=manifest_required_migrations,
        member_count=len(members),
        member_input_hash=member_input_hash,
        league_payload_sha256=league_payload_sha256,
        bundle_payload_sha256=bundle_payload_sha256,
        team_payload_sha256s=team_payload_sha256s,
        reviewed_preflight_sha256=reviewed_sha256,
        reviewed_preflight_manifest_sha256=reviewed_manifest_sha256,
        provenance={
            "code_version": provenance["code_version"],
            "dependency_lock_hash": provenance["dependency_lock_hash"],
            "operator": provenance["operator"],
        },
        dirty_worktree_paths=current_dirty_paths,
        dirty_worktree_allowed_paths=dirty_override_allowed_paths,
        parity_export=parity_export,
        rollback=(predecessor if predecessor is not None else {"status": "no predecessor existed"}),
        rollback_of_release_id=rollback_of_release_id or None,
        rollback_replaces_release_id=(candidate.get("replaces_release_id") if rollback_of_release_id else None),
        timings_ms=(
            {
                **timings_ms,
                "parity_export": parity_export_ms,
                "workflow_total": workflow_total_ms,
            }
            if analysis_version != "v6"
            else None
        ),
    )
    try:
        write_json_atomic(release_manifest_path, release_manifest)
    except BaseException as exc:
        if analysis_version == "v6":
            raise SystemExit(
                "V6 league promotion and parity export succeeded, but the local "
                "release manifest write failed. The approved release was retained; "
                f"repair the release manifest path, then run exactly: "
                f"{v6_finalizer_command}. Error: {exc}"
            ) from exc
        raise SystemExit(
            "league promotion and parity export succeeded, but the local "
            "release manifest write failed; do not rerun promotion, preserve "
            f"the approved release and repair {release_manifest_path}: {exc}"
        ) from exc
    print(
        json.dumps(
            {
                "release_label": label,
                "season": season,
                "member_count": len(members),
                "member_input_hash": member_input_hash,
                "league_payload_sha256": league_payload_sha256,
                "bundle_payload_sha256": bundle_payload_sha256,
                "preflight_json_sha256": preflight_bundle_sha256,
                "export_path": str(export_path),
                "export_backup_path": str(export_backup_path) if export_backup_path else None,
                "parity_export": {
                    "team_count": parity_export["team_count"],
                    "export_set_sha256": parity_export["export_set_sha256"],
                },
                "release_manifest_path": str(release_manifest_path),
                "timings_ms": {
                    **timings_ms,
                    "parity_export": parity_export_ms,
                    "workflow_total": workflow_total_ms,
                },
            },
            indent=2,
        )
    )


def adjudicate_duplicate_exclusion(args: argparse.Namespace) -> None:
    path = Path(args.file)
    file_hash = sha256_file(path)
    raw_id = raw_record_id(args.team, args.season, file_hash, args.row_number)
    params = SqlParams()
    provenance = run_provenance()
    state = {
        "analysis_eligibility_status": "excluded_duplicate_adjudicated",
        "excluded_standardised_row_number": args.row_number,
        "duplicate_of_standardised_row_number": args.duplicate_of,
        "decision": "exclude_duplicate",
        "rationale": args.rationale,
    }
    output_hash = sha256_json(state)
    sql = f"""
      insert into audit.reason_codes (code, description) values
        ('duplicate_adjudicated_exclusion', 'Manual review adjudicated a candidate duplicate and excluded it from analysis.')
      on conflict (code) do update set description = excluded.description;

      do $$
      begin
        if not exists (select 1 from ingestion.source_rows where raw_record_id = {params.text(raw_id)}) then
          raise exception 'adjudication source row is not registered';
        end if;
        if not exists (select 1 from processing.record_versions rv join ingestion.source_rows sr on sr.id = rv.source_row_id where sr.raw_record_id = {params.text(raw_id)} and rv.version_number < {args.version_number}) then
          raise exception 'adjudication requires a prior processed record version';
        end if;
        if exists (select 1 from processing.record_versions rv join ingestion.source_rows sr on sr.id = rv.source_row_id where sr.raw_record_id = {params.text(raw_id)} and rv.version_number = {args.version_number}) then
          raise exception 'adjudication version already exists';
        end if;
      end $$;

      create temp table current_step on commit drop as
      with run as (
        insert into audit.pipeline_runs
          (command, team, season, status, input_hash, output_hash, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values (
          'adjudicate-duplicate-exclusion', {params.text(args.team)}, {params.text(args.season)}, 'succeeded',
          {params.text(file_hash)}, {params.text(output_hash)},
          {params.jsonb({
            'file': path.name,
            'row_number': args.row_number,
            'duplicate_of': args.duplicate_of,
            'step_version': args.step_version,
            'version_number': args.version_number,
          })},
          now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      ),
      step as (
        insert into audit.step_runs
          (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count, counts_by_team, input_hash, output_hash, ended_at)
        select id, 'duplicate_adjudication', {params.text(args.step_version)},
          'duplicate_adjudicated_exclusion', 1, 1,
          {params.jsonb({args.team: {'excluded_duplicate_rows': [args.row_number]}})},
          {params.text(file_hash)}, {params.text(output_hash)}, now()
        from run
        returning id
      )
      select id from step;

      insert into processing.record_versions
        (source_row_id, step_run_id, version_number, record_state, eligibility_status)
      select sr.id, step.id, {args.version_number},
        coalesce(previous.record_state, '{{}}'::jsonb) || {params.jsonb(state)},
        'excluded_duplicate_adjudicated'
      from ingestion.source_rows sr
      cross join current_step step
      left join lateral (
        select rv.record_state
        from processing.record_versions rv
        where rv.source_row_id = sr.id and rv.version_number < {args.version_number}
        order by rv.version_number desc
        limit 1
      ) previous on true
      where sr.raw_record_id = {params.text(raw_id)}
      ;

      insert into audit.record_events
        (step_run_id, source_row_id, field_name, old_value, new_value, action, reason_code, rationale, rule_version, review_status)
      select step.id, sr.id, 'analysis_eligibility_status', null,
        {params.jsonb('excluded_duplicate_adjudicated')}, 'exclude',
        'duplicate_adjudicated_exclusion', {params.text(args.rationale)},
        {params.text(args.step_version)}, 'adjudicated'
      from ingestion.source_rows sr, current_step step
      where sr.raw_record_id = {params.text(raw_id)};

      insert into audit.adjudications
        (source_row_id, field_name, decision, rationale, reviewer, consumed_by_step_run_id)
      select sr.id, 'analysis_eligibility_status', {params.jsonb(state)},
        {params.text(args.rationale)}, {params.text(args.reviewer)}, step.id
      from ingestion.source_rows sr, current_step step
      where sr.raw_record_id = {params.text(raw_id)}
        and not exists (
          select 1 from audit.adjudications existing
          where existing.source_row_id = sr.id
            and existing.field_name = 'analysis_eligibility_status'
            and existing.decision = {params.jsonb(state)}
        );
    """
    run_sql(sql, params.values)
    print(json.dumps({"team": args.team, "excluded_row": args.row_number, "raw_record_id": raw_id}, indent=2))


def expected_adjudication_batch_records(
    *, source_corrections: list[dict[str, Any]], duplicate_reviews: list[dict[str, Any]],
    rule_decisions: list[dict[str, Any]], workbook_sha256: str,
    evidence_manifest_sha256: str, workbook_path: Path,
    migration_version: str, migration_sha256: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    row_records: list[dict[str, Any]] = []
    for item in source_corrections:
        row_records.append({
            "source_row_id": item["source_row_id"], "field_name": item["field_name"],
            "decision": {
                "decision_type": "source_field_correction", "item_id": item["item_id"],
                "old_value": item["old_value"], "new_value": item["new_value"],
                "source_row_sha256": item["source_row_sha256"],
                "evidence_sha256": item["evidence_sha256"],
                "workbook_sha256": workbook_sha256,
                "evidence_manifest_sha256": evidence_manifest_sha256,
            },
            "rationale": item["rationale"], "reviewer": "Abdel Babiker",
        })
    for item in duplicate_reviews:
        row_records.append({
            "source_row_id": item["source_row_ids"][0], "field_name": "duplicate_review",
            "decision": {
                "decision_type": "duplicate_review", "item_id": item["item_id"],
                "decision": item["decision"], "source_row_ids": item["source_row_ids"],
                "source_row_sha256s": item["source_row_sha256s"],
                "differing_source_key": "training_day", "evidence_sha256": item["evidence_sha256"],
                "workbook_sha256": workbook_sha256,
                "evidence_manifest_sha256": evidence_manifest_sha256,
            },
            "rationale": item["rationale"], "reviewer": "Abdel Babiker",
        })
    rule_records = [{
        "adjudication_ref": item["item_id"], "rule_version": item["rule_version"],
        "decision": item["decision"], "evidence_sha256": item["evidence_sha256"],
        "workbook_sha256": workbook_sha256,
        "evidence_manifest_sha256": evidence_manifest_sha256,
        "reviewer": "Abdel Babiker", "workbook_snapshot_locator": str(workbook_path),
        "migration_version": migration_version, "migration_sha256": migration_sha256,
        "rationale": item["rationale"],
    } for item in rule_decisions]
    return (
        sorted(row_records, key=lambda row: clean_text(row["decision"]["item_id"])),
        sorted(rule_records, key=lambda row: clean_text(row["adjudication_ref"])),
    )


def apply_adjudication_batch(args: argparse.Namespace) -> None:
    """Record the exact 14-item workbook decision batch without editing sources."""
    manifest_path = Path(args.file)
    workbook_path = Path(args.workbook)
    evidence_path = Path(args.evidence_file)
    for required in (manifest_path, workbook_path, evidence_path):
        if not required.exists():
            raise SystemExit(f"adjudication input not found: {required}")
    if manifest_path.resolve() != APPROVED_ADJUDICATION_14_MANIFEST_PATH:
        raise SystemExit("adjudication batch must use the durable Git-ignored approved manifest")
    if workbook_path.resolve() != APPROVED_ADJUDICATION_14_WORKBOOK_PATH:
        raise SystemExit("adjudication batch must use the durable Git-ignored approved workbook snapshot")
    if evidence_path.resolve() != APPROVED_ADJUDICATION_14_EVIDENCE_PATH:
        raise SystemExit("adjudication batch must use the durable Git-ignored evidence manifest")
    manifest_sha256 = sha256_file(manifest_path)
    if manifest_sha256 != APPROVED_ADJUDICATION_14_MANIFEST_SHA256:
        raise SystemExit("adjudication decision envelope differs from the approved manifest")
    manifest = json.loads(manifest_path.read_text())
    if manifest.get("schema_version") != "urc_adjudication_batch_v1":
        raise SystemExit("unsupported adjudication batch schema")
    if clean_text(manifest.get("season")) != "2024-25":
        raise SystemExit("this adjudication batch must be bound to season 2024-25")
    if clean_text(manifest.get("reviewer")) != "Abdel Babiker":
        raise SystemExit("adjudication reviewer must be Abdel Babiker")
    workbook_sha256 = sha256_file(workbook_path)
    evidence_manifest_sha256 = sha256_file(evidence_path)
    if workbook_sha256 != APPROVED_ADJUDICATION_14_WORKBOOK_SHA256:
        raise SystemExit("workbook does not match the binding approved workbook checksum")
    if evidence_manifest_sha256 != APPROVED_ADJUDICATION_14_EVIDENCE_SHA256:
        raise SystemExit("evidence does not match the binding approved evidence checksum")
    migration_version = clean_text(manifest.get("classification_migration_version"))
    migration_path = Path("supabase/migrations") / f"{migration_version}_adjudicated_reporting_classification.sql"
    if migration_version != ADJUDICATED_REPORTING_CLASSIFICATION_MIGRATION_VERSION or not migration_path.exists():
        raise SystemExit("adjudication batch classification migration is missing or untracked")
    migration_sha256 = sha256_file(migration_path)
    if workbook_sha256 != clean_text(manifest.get("workbook_sha256")):
        raise SystemExit("adjudication workbook checksum differs from the approved batch")
    if evidence_manifest_sha256 != clean_text(manifest.get("evidence_manifest_sha256")):
        raise SystemExit("adjudication evidence checksum differs from the approved batch")
    if migration_sha256 != clean_text(manifest.get("classification_migration_sha256")):
        raise SystemExit("classification migration checksum differs from the approved batch")

    source_corrections = manifest.get("source_corrections")
    duplicate_reviews = manifest.get("duplicate_reviews")
    rule_decisions = manifest.get("rule_decisions")
    if not isinstance(source_corrections, list) or len(source_corrections) != 3:
        raise SystemExit("adjudication batch requires exactly three source corrections")
    if not isinstance(duplicate_reviews, list) or len(duplicate_reviews) != 9:
        raise SystemExit("adjudication batch requires exactly nine duplicate reviews")
    if not isinstance(rule_decisions, list) or len(rule_decisions) != 2:
        raise SystemExit("adjudication batch requires exactly two reporting-rule decisions")
    classification_rule_version = "reporting_classification_2026-07-20_v1"
    if {clean_text(item.get("rule_version")) for item in rule_decisions} != {classification_rule_version}:
        raise SystemExit("reporting-rule decisions use an unexpected classification version")
    item_ids = {
        *(clean_text(item.get("item_id")) for item in source_corrections),
        *(clean_text(item.get("item_id")) for item in duplicate_reviews),
        *(clean_text(item.get("item_id")) for item in rule_decisions),
    }
    expected_item_ids = {
        "ID-01", "ID-02", "ID-03", "DX-02", "DX-03", "DX-12", "DX-13",
        "DX-14", "DX-15", "DX-16", "DX-17", "DX-18", "IA-02", "ACL-01",
    }
    if item_ids != expected_item_ids:
        raise SystemExit(f"adjudication item set differs: {sorted(item_ids ^ expected_item_ids)}")

    source_row_ids: set[str] = set()
    for item in source_corrections:
        if item.get("field_name") not in {"Date Injured", "Fit For Selection Date"}:
            raise SystemExit(f"{item.get('item_id')} uses a disallowed source field")
        source_row_ids.add(str(uuid.UUID(clean_text(item.get("source_row_id")))))
    for item in duplicate_reviews:
        ids = item.get("source_row_ids")
        if not isinstance(ids, list) or len(ids) != 2:
            raise SystemExit(f"{item.get('item_id')} must bind exactly two source rows")
        source_row_ids.update(str(uuid.UUID(clean_text(value))) for value in ids)
        if item.get("decision") != "retain_distinct_training_day":
            raise SystemExit(f"{item.get('item_id')} is not an approved retain-distinct decision")

    lookup_params = SqlParams()
    live_rows = query_sql(
        f"""
        select sr.id::text as source_row_id, sr.row_sha256, sr.source_values,
               sf.team, sf.season
        from ingestion.source_rows sr
        join ingestion.source_files sf on sf.id = sr.source_file_id
        where sr.id in (
          select value::uuid
          from jsonb_array_elements_text({lookup_params.jsonb(sorted(source_row_ids))}) value
        )
        order by sr.id
        """,
        lookup_params.values,
    )
    if len(live_rows) != len(source_row_ids):
        raise SystemExit("one or more adjudication source rows are absent from the hosted database")
    live_by_id = {row["source_row_id"]: row for row in live_rows}
    for item in source_corrections:
        row = live_by_id[clean_text(item["source_row_id"])]
        source_values = row.get("source_values")
        if row["season"] != "2024-25" or row["row_sha256"] != item["source_row_sha256"]:
            raise SystemExit(f"{item['item_id']} source evidence fingerprint changed")
        if not isinstance(source_values, dict) or clean_text(source_values.get(item["field_name"])) != clean_text(item["old_value"]):
            raise SystemExit(f"{item['item_id']} expected old source value changed")
    for item in duplicate_reviews:
        for row_id, expected_hash in zip(item["source_row_ids"], item["source_row_sha256s"], strict=True):
            row = live_by_id[row_id]
            if row["season"] != "2024-25" or row["row_sha256"] != expected_hash:
                raise SystemExit(f"{item['item_id']} duplicate evidence fingerprint changed")
        first = live_by_id[item["source_row_ids"][0]]["source_values"]
        second = live_by_id[item["source_row_ids"][1]]["source_values"]
        if clean_text(first.get("training_day")) == clean_text(second.get("training_day")):
            raise SystemExit(f"{item['item_id']} no longer has the reviewed training_day distinction")

    expected_row_records, expected_rule_records = expected_adjudication_batch_records(
        source_corrections=source_corrections,
        duplicate_reviews=duplicate_reviews,
        rule_decisions=rule_decisions,
        workbook_sha256=workbook_sha256,
        evidence_manifest_sha256=evidence_manifest_sha256,
        workbook_path=workbook_path.resolve(),
        migration_version=migration_version,
        migration_sha256=migration_sha256,
    )

    if bool(getattr(args, "plan", False)):
        print(json.dumps({
            "status": "validated_read_only",
            "items": sorted(item_ids),
            "source_rows_verified": len(source_row_ids),
            "workbook_sha256": workbook_sha256,
            "evidence_manifest_sha256": evidence_manifest_sha256,
            "classification_migration_version": migration_version,
            "classification_migration_sha256": migration_sha256,
            "manifest_sha256": manifest_sha256,
        }, indent=2))
        return

    existing_rows = query_sql(
        """
        select source_row_id::text as source_row_id, field_name, decision,
               rationale, reviewer
        from audit.adjudications
        where decision ->> 'item_id' in (
          'ID-01', 'ID-02', 'ID-03', 'DX-02', 'DX-03', 'DX-12',
          'DX-13', 'DX-14', 'DX-15', 'DX-16', 'DX-17', 'DX-18'
        )
        order by decision ->> 'item_id'
        """
    )
    existing_rules = query_sql(
        """
        select adjudication_ref, rule_version, decision, evidence_sha256,
               workbook_sha256, evidence_manifest_sha256, reviewer,
               workbook_snapshot_locator, migration_version, migration_sha256,
               rationale
        from audit.rule_adjudications
        where adjudication_ref in ('IA-02', 'ACL-01')
          and rule_version = 'reporting_classification_2026-07-20_v1'
        order by adjudication_ref
        """
    )
    if existing_rows == expected_row_records and existing_rules == expected_rule_records:
        print(json.dumps({"status": "already_recorded", "items": sorted(item_ids),
            "workbook_sha256": workbook_sha256}, indent=2))
        return
    recorded_count = len(existing_rows) + len(existing_rules)
    if recorded_count:
        raise SystemExit(
            f"conflicting or partial adjudication batch already exists ({recorded_count} rows); "
            "refusing mixed replay"
        )

    _write_adjudication_batch(
        manifest=manifest,
        manifest_path=manifest_path,
        workbook_path=workbook_path,
        evidence_path=evidence_path,
        workbook_sha256=workbook_sha256,
        evidence_manifest_sha256=evidence_manifest_sha256,
        migration_version=migration_version,
        migration_sha256=migration_sha256,
        classification_rule_version=classification_rule_version,
        manifest_sha256=manifest_sha256,
        expected_row_records=expected_row_records,
        expected_rule_records=expected_rule_records,
        item_ids=item_ids,
    )


def _write_adjudication_batch(
    *, manifest: dict[str, Any], manifest_path: Path, workbook_path: Path,
    evidence_path: Path, workbook_sha256: str, evidence_manifest_sha256: str,
    migration_version: str, migration_sha256: str,
    classification_rule_version: str,
    manifest_sha256: str,
    item_ids: set[str],
    expected_row_records: list[dict[str, Any]],
    expected_rule_records: list[dict[str, Any]],
) -> None:
    provenance = run_provenance()
    input_hash = manifest_sha256
    params = SqlParams()
    adjudication_values: list[str] = []
    for record in expected_row_records:
        adjudication_values.append(
            f"({params.text(record['source_row_id'])}::uuid, {params.text(record['field_name'])}, "
            f"{params.jsonb(record['decision'])}, {params.text(record['rationale'])}, "
            f"{params.text(record['reviewer'])})"
        )
    rule_values = [
        f"({params.text(record['adjudication_ref'])}, {params.text(record['rule_version'])}, "
        f"{params.jsonb(record['decision'])}, {params.text(record['evidence_sha256'])}, "
        f"{params.text(record['workbook_sha256'])}, "
        f"{params.text(record['evidence_manifest_sha256'])}, "
        f"{params.text(record['reviewer'])}, {params.text(record['workbook_snapshot_locator'])}, "
        f"{params.text(record['migration_version'])}, {params.text(record['migration_sha256'])}, "
        f"{params.text(record['rationale'])})"
        for record in expected_rule_records
    ]
    run_parameters = {
        "schema_version": manifest["schema_version"], "manifest": str(manifest_path.resolve()),
        "workbook_sha256": workbook_sha256,
        "workbook_snapshot_locator": str(workbook_path.resolve()),
        "evidence_manifest": str(evidence_path.resolve()),
        "evidence_manifest_sha256": evidence_manifest_sha256,
        "classification_migration_version": migration_version,
        "classification_migration_sha256": migration_sha256,
        "reviewer": "Abdel Babiker", "item_ids": sorted(item_ids),
    }
    sql = f"""
      insert into audit.reason_codes (code, description) values
        ('source_field_adjudicated_correction', 'Human-approved correction overlaid on an immutable source field before deterministic re-derivation.'),
        ('duplicate_review_retain_distinct', 'Human-reviewed exposure pair retained because a substantive source field differs.'),
        ('reporting_classification_adjudication', 'Human-approved reporting classification rule bound to exact evidence.')
      on conflict (code) do update set description = excluded.description;

      with run as (
        insert into audit.pipeline_runs
          (command, team, season, status, input_hash, output_hash, parameters,
           ended_at, code_version, dependency_lock_hash, operator)
        values ('apply-adjudication-batch', 'URC adjudication batch', '2024-25',
          'succeeded', {params.text(input_hash)}, {params.text(input_hash)},
          {params.jsonb(run_parameters)}, now(), {params.text(provenance['code_version'])},
          {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])})
        returning id
      ) insert into audit.step_runs
          (pipeline_run_id, step_name, step_version, reason_code, input_count,
           output_count, input_hash, output_hash, ended_at)
        select id, 'record_adjudication_batch', 'urc_adjudication_batch_2026-07-20_v1',
          'reporting_classification_adjudication', 14, 14,
          {params.text(input_hash)}, {params.text(input_hash)}, now() from run;

      insert into audit.adjudications
        (source_row_id, field_name, decision, rationale, reviewer)
      select v.source_row_id, v.field_name, v.decision, v.rationale, v.reviewer
      from (values {', '.join(adjudication_values)})
        v(source_row_id, field_name, decision, rationale, reviewer)
      where not exists (
        select 1 from audit.adjudications existing
        where existing.source_row_id = v.source_row_id
          and existing.field_name = v.field_name and existing.decision = v.decision
          and existing.reviewer = v.reviewer
      );

      insert into audit.rule_adjudications
        (adjudication_ref, rule_version, decision, evidence_sha256,
         workbook_sha256, evidence_manifest_sha256, reviewer,
         workbook_snapshot_locator, migration_version, migration_sha256, rationale)
      values {', '.join(rule_values)}
      on conflict (adjudication_ref, rule_version) do nothing;

      do $$
      begin
        if (
          select count(*)
          from jsonb_to_recordset({params.jsonb(expected_row_records)}) as expected(
            source_row_id text, field_name text, decision jsonb, rationale text, reviewer text
          )
          join audit.adjudications actual
            on actual.source_row_id = expected.source_row_id::uuid
           and actual.field_name = expected.field_name
           and actual.decision = expected.decision
           and actual.rationale = expected.rationale
           and actual.reviewer = expected.reviewer
        ) <> 12 or (
          select count(*)
          from audit.adjudications
          where decision ->> 'item_id' in (
            'ID-01', 'ID-02', 'ID-03', 'DX-02', 'DX-03', 'DX-12',
            'DX-13', 'DX-14', 'DX-15', 'DX-16', 'DX-17', 'DX-18'
          )
        ) <> 12 then
          raise exception 'stored row adjudications differ from the approved batch';
        end if;

        if (
          select count(*)
          from jsonb_to_recordset({params.jsonb(expected_rule_records)}) as expected(
            adjudication_ref text, rule_version text, decision jsonb,
            evidence_sha256 text, workbook_sha256 text,
            evidence_manifest_sha256 text, reviewer text,
            workbook_snapshot_locator text, migration_version text,
            migration_sha256 text, rationale text
          )
          join audit.rule_adjudications actual
            on actual.adjudication_ref = expected.adjudication_ref
           and actual.rule_version = expected.rule_version
           and actual.decision = expected.decision
           and actual.evidence_sha256 = expected.evidence_sha256
           and actual.workbook_sha256 = expected.workbook_sha256
           and actual.evidence_manifest_sha256 = expected.evidence_manifest_sha256
           and actual.reviewer = expected.reviewer
           and actual.workbook_snapshot_locator = expected.workbook_snapshot_locator
           and actual.migration_version = expected.migration_version
           and actual.migration_sha256 = expected.migration_sha256
           and actual.rationale = expected.rationale
        ) <> 2 or (
          select count(*) from audit.rule_adjudications
          where rule_version = {params.text(classification_rule_version)}
            and adjudication_ref in ('IA-02', 'ACL-01')
        ) <> 2 then
          raise exception 'stored classification adjudications differ from the approved batch';
        end if;
      end $$;
    """
    run_sql(sql, params.values)
    print(json.dumps({"status": "recorded", "items": sorted(item_ids),
        "workbook_sha256": workbook_sha256,
        "evidence_manifest_sha256": evidence_manifest_sha256,
        "input_hash": input_hash}, indent=2))


def apply_osiics_mapping_adjudication(args: argparse.Namespace) -> None:
    """Validate and record the exact OSIICS-01 reporting-only rule decision."""
    evidence_path = Path(args.evidence_file).resolve()
    ledger_path = Path(args.row_ledger).resolve()
    workbook_path = Path(args.workbook).resolve()
    migration_path = Path(args.migration_file).resolve()
    for path in (evidence_path, ledger_path, workbook_path, migration_path):
        if not path.is_file():
            raise SystemExit(f"required OSIICS evidence file is missing: {path}")

    evidence = json.loads(evidence_path.read_text())
    ledger = json.loads(ledger_path.read_text())
    rows = ledger.get("rows")
    if not isinstance(rows, list) or len(rows) != 121:
        raise SystemExit("OSIICS row ledger must contain exactly 121 reviewed rows")
    row_ids = [clean_text(row.get("source_row_id")) for row in rows]
    if len(set(row_ids)) != 121 or any(not row_id for row_id in row_ids):
        raise SystemExit("OSIICS row ledger source-row identities are incomplete or duplicated")

    expected_workbook_sha = evidence["official_reference"]["official_workbook_sha256"]
    expected_ledger_sha = evidence["row_ledger"]["sha256"]
    expected_mapping_sha = evidence["mapping_catalogue"]["sha256"]
    expected_multi_sha = evidence["multi_type_catalogue"]["sha256"]
    mapping_path = evidence_path.parents[2] / evidence["mapping_catalogue"]["path"]
    multi_path = evidence_path.parents[2] / evidence["multi_type_catalogue"]["path"]
    actuals = {
        "workbook": sha256_file(workbook_path),
        "row_ledger": sha256_file(ledger_path),
        "mapping_catalogue": sha256_file(mapping_path),
        "multi_type_catalogue": sha256_file(multi_path),
    }
    expected = {
        "workbook": expected_workbook_sha,
        "row_ledger": expected_ledger_sha,
        "mapping_catalogue": expected_mapping_sha,
        "multi_type_catalogue": expected_multi_sha,
    }
    if actuals != expected:
        raise SystemExit(f"OSIICS evidence checksum mismatch: expected={expected} actual={actuals}")

    counts = evidence["expected_live_time_loss_counts"]
    if counts != {
        "cohort": 1120, "unknown_before": 245, "exact_code_candidates": 111,
        "explicit_text_candidates": 9, "multi_type_diagnosis_candidates": 1,
        "newly_classified": 121, "unknown_after": 124,
    }:
        raise SystemExit("OSIICS evidence counts differ from the approved decision envelope")

    live_params = SqlParams()
    live_rows = query_sql(
        f"""
        select sr.id::text as source_row_id, sr.row_sha256 as source_row_sha256,
               i.id::text as injury_id, i.team_key, i.season,
               case i.activity_context when 'urc_match' then 'match'
                 when 'match' then 'match' when 'training' then 'training'
                 else 'unknown' end as setting,
               i.days_injured as days_lost,
               i.body_location as original_body_location_code,
               i.injury_type as original_injury_type_code
        from ingestion.source_rows sr
        join curated.injuries i on i.source_row_id=sr.id
        join analysis.league_member_releases_v2 m
          on m.curated_build_id=i.curated_build_id and m.team_key=i.team_key
         and m.season=i.season
        where i.season='2024-25' and i.days_injured > 0
          and sr.id in (select value::uuid from jsonb_array_elements_text(
            {live_params.jsonb(row_ids)}) value)
        order by sr.id
        """,
        live_params.values,
    )
    if len(live_rows) != 121:
        raise SystemExit(f"hosted OSIICS cohort contains {len(live_rows)} of 121 reviewed rows")
    live_by_id = {row["source_row_id"]: row for row in live_rows}
    for reviewed in rows:
        live = live_by_id.get(reviewed["source_row_id"])
        for key in (
            "injury_id", "source_row_sha256", "team_key", "season", "setting",
            "days_lost", "original_body_location_code", "original_injury_type_code",
        ):
            if live is None or live.get(key) != reviewed.get(key):
                raise SystemExit(
                    f"OSIICS reviewed row fingerprint changed: {reviewed['source_row_id']} {key}"
                )
    unknown_rows = query_sql("""
      select count(*) as count
      from analysis.season_bound_reporting_classification_v3 p
      join analysis.league_member_releases_v2 m using (curated_build_id, team_key, season)
      where p.season='2024-25' and p.is_time_loss and p.diagnosis_code='unknown'
    """)
    if len(unknown_rows) != 1 or int(unknown_rows[0]["count"]) != 245:
        raise SystemExit("hosted predecessor Unknown count is no longer 245")

    evidence_manifest_sha = sha256_file(evidence_path)
    migration_sha = sha256_file(migration_path)
    approved_migration_sha = clean_text(args.expected_migration_sha256).lower()
    if not re.fullmatch(r"[0-9a-f]{64}", approved_migration_sha):
        raise SystemExit("--expected-migration-sha256 must be the exact approved 64-character hash")
    if migration_sha != approved_migration_sha:
        raise SystemExit(
            f"migration checksum differs from the approved hash: {migration_sha}"
        )
    decision = {
        "mapping_catalogue_sha256": expected_mapping_sha,
        "multi_type_catalogue_sha256": expected_multi_sha,
        "exact_code_candidate_count": 111,
        "explicit_text_candidate_count": 9,
        "multi_type_diagnosis_candidate_count": 1,
        "unknown_before": 245,
        "unknown_after": 124,
        "preserve_original_values": True,
        "conflicts_remain_unknown": True,
    }
    result = {
        "status": "validated_read_only" if args.plan else "recorded",
        "database_host": evidence["target"]["database_host"],
        "rule_version": evidence["rule_version"],
        "reviewed_rows": 121,
        "unknown_before": 245,
        "unknown_after": 124,
        "evidence_manifest_sha256": evidence_manifest_sha,
        "migration_sha256": migration_sha,
    }
    if args.plan:
        print(json.dumps(result, indent=2))
        return

    existing_params = SqlParams()
    existing = query_sql(
        f"""
        select adjudication_ref, rule_version, decision, evidence_sha256,
               workbook_sha256, evidence_manifest_sha256, reviewer,
               workbook_snapshot_locator, migration_version, migration_sha256, rationale
        from audit.rule_adjudications
        where adjudication_ref='OSIICS-01'
          and rule_version='reporting_classification_2026-07-22_v2'
        """,
        existing_params.values,
    )
    if existing:
        raise SystemExit(
            "OSIICS-01 already exists; refusing to append a second pipeline run or mixed replay"
        )

    params = SqlParams()
    provenance = run_provenance()
    rationale = (
        "Abdel Babiker approved exact OSIICS/OSICS mappings, uniquely explicit body/type "
        "descriptions, and NPM as Neck with candidate types muscle_injury;tendinopathy; "
        "the reporting primary type is nonspecific so the injury contributes once."
    )
    run_sql(
        f"""
        insert into audit.reason_codes (code, description) values
          ('reporting_classification_adjudication',
           'Human-approved reporting classification rule bound to exact evidence.')
        on conflict (code) do update set description=excluded.description;
        create temp table current_osiics_run on commit drop as
        with run as (
          insert into audit.pipeline_runs
            (command, team, season, status, input_hash, output_hash, parameters,
             code_version, dependency_lock_hash, operator)
          values ('apply-osiics-mapping-adjudication','URC Overall','2024-25','started',
            {params.text(evidence_manifest_sha)}, {params.text(expected_ledger_sha)},
            {params.jsonb(result)}, {params.text(provenance['code_version'])},
            {params.text(provenance['dependency_lock_hash'])},
            {params.text(provenance['operator'])}) returning id
        ) select id from run;
        insert into audit.step_runs
          (pipeline_run_id, step_name, step_version, reason_code, input_count,
           output_count, input_hash, output_hash)
        select id,'record_osiics_mapping_adjudication',
          'reporting_classification_2026-07-22_v2','reporting_classification_adjudication',
          245,121,{params.text(evidence_manifest_sha)},{params.text(expected_ledger_sha)}
        from current_osiics_run;
        insert into audit.rule_adjudications
          (adjudication_ref, rule_version, decision, evidence_sha256,
           workbook_sha256, evidence_manifest_sha256, reviewer,
           workbook_snapshot_locator, migration_version, migration_sha256, rationale)
        values ('OSIICS-01','reporting_classification_2026-07-22_v2',
          {params.jsonb(decision)},{params.text(expected_ledger_sha)},
          {params.text(expected_workbook_sha)},{params.text(evidence_manifest_sha)},
          'Abdel Babiker',{params.text(str(workbook_path))},'20260722140000',
          {params.text(migration_sha)},{params.text(rationale)});

        do $$
        declare
          reviewed_match_count integer;
          changed_count integer;
          unknown_count integer;
          cohort_count integer;
          mismatch_detail jsonb;
        begin
          select count(*) into reviewed_match_count
          from jsonb_to_recordset({params.jsonb(rows)}) as expected(
            injury_id text, mapped_body_location_code text,
            mapped_injury_type_code text, mapped_diagnosis_code text,
            mapped_diagnosis_label text, origin text, candidate_injury_types text
          )
          join analysis.season_bound_reporting_classification_v4 actual
            on actual.injury_id=expected.injury_id::uuid
          join analysis.league_member_releases_v2 member
            using (curated_build_id,team_key,season)
          where actual.season='2024-25' and actual.is_time_loss
            and actual.effective_body_location_code=expected.mapped_body_location_code
            and actual.effective_injury_type_code=expected.mapped_injury_type_code
            and actual.diagnosis_code=expected.mapped_diagnosis_code
            and (expected.mapped_diagnosis_label is null
              or actual.diagnosis_label=expected.mapped_diagnosis_label)
            and actual.diagnosis_origin=expected.origin
            and actual.candidate_injury_types is not distinct from expected.candidate_injury_types;
          if reviewed_match_count <> 121 then
            select jsonb_agg(to_jsonb(mismatch)) into mismatch_detail
            from (
              select expected.injury_id, expected.mapped_body_location_code as expected_body,
                actual.effective_body_location_code as actual_body,
                expected.mapped_injury_type_code as expected_type,
                actual.effective_injury_type_code as actual_type,
                expected.mapped_diagnosis_code as expected_diagnosis,
                actual.diagnosis_code as actual_diagnosis,
                expected.origin as expected_origin, actual.diagnosis_origin as actual_origin,
                expected.candidate_injury_types as expected_candidate_types,
                actual.candidate_injury_types as actual_candidate_types
              from jsonb_to_recordset({params.jsonb(rows)}) as expected(
                injury_id text, mapped_body_location_code text,
                mapped_injury_type_code text, mapped_diagnosis_code text,
                mapped_diagnosis_label text, origin text, candidate_injury_types text
              )
              left join analysis.season_bound_reporting_classification_v4 actual
                on actual.injury_id=expected.injury_id::uuid
              left join analysis.league_member_releases_v2 member
                on member.curated_build_id=actual.curated_build_id
               and member.team_key=actual.team_key and member.season=actual.season
              where member.team_key is null
                or actual.season<>'2024-25' or not actual.is_time_loss
                or actual.effective_body_location_code<>expected.mapped_body_location_code
                or actual.effective_injury_type_code<>expected.mapped_injury_type_code
                or actual.diagnosis_code<>expected.mapped_diagnosis_code
                or (expected.mapped_diagnosis_label is not null
                  and actual.diagnosis_label<>expected.mapped_diagnosis_label)
                or actual.diagnosis_origin<>expected.origin
                or actual.candidate_injury_types is distinct from expected.candidate_injury_types
            ) mismatch;
            raise exception 'OSIICS successor matches % of 121 reviewed row outcomes: %',
              reviewed_match_count, mismatch_detail;
          end if;

          select count(*), count(*) filter (where successor.diagnosis_code='unknown'),
                 count(*) filter (where successor.diagnosis_code is distinct from predecessor.diagnosis_code)
          into cohort_count, unknown_count, changed_count
          from analysis.season_bound_reporting_classification_v4 successor
          join analysis.season_bound_reporting_classification_v3 predecessor
            using (injury_id,curated_build_id,team_key,season,setting_code,is_time_loss,days_lost)
          join analysis.league_member_releases_v2 member
            using (curated_build_id,team_key,season)
          where successor.season='2024-25' and successor.is_time_loss;
          if cohort_count <> 1120 or unknown_count <> 124 or changed_count <> 121 then
            raise exception 'OSIICS successor reconciliation failed: cohort %, unknown %, changed %',
              cohort_count, unknown_count, changed_count;
          end if;
        end $$;

        update audit.pipeline_runs run
        set status='succeeded', ended_at=now()
        from current_osiics_run current
        where run.id=current.id and run.status='started';
        update audit.step_runs step set ended_at=now()
        from current_osiics_run current
        where step.pipeline_run_id=current.id
          and step.step_name='record_osiics_mapping_adjudication';
        """,
        params.values,
    )
    stored = query_sql("""
      select adjudication_ref, rule_version, decision, evidence_sha256,
             workbook_sha256, evidence_manifest_sha256, reviewer,
             workbook_snapshot_locator, migration_version, migration_sha256, rationale
      from audit.rule_adjudications
      where adjudication_ref='OSIICS-01'
        and rule_version='reporting_classification_2026-07-22_v2'
    """)
    if len(stored) != 1 or stored[0]["decision"] != decision:
        raise SystemExit("stored OSIICS-01 adjudication failed exact post-write verification")
    print(json.dumps(result, indent=2))


def reapply_adjudications(args: argparse.Namespace) -> None:
    """Backfill command for the reapplication gap fixed in process_intake/
    process_exposure (see fetch_standing_eligibility_adjudications and
    apply_standing_adjudication): for every standing audit.adjudications
    decision (field_name = 'analysis_eligibility_status') belonging to
    team/season, compares the adjudicated status against the *current*
    latest processing.record_versions.eligibility_status for that source
    row. Rows already matching are left untouched -- no write, no new
    pipeline_run, this command is a true no-op when nothing has drifted.
    Only rows where a prior rerun silently reverted the adjudicated status
    get one new record_versions row each (version_number = current max + 1,
    never editing the existing row in place) plus a record_events row
    (action 'reapply', reason_code 'adjudication_reapplied'), all under one
    recorded pipeline run.
    """
    team = clean_text(args.team)
    season = clean_text(args.season)
    if not team or not season:
        raise SystemExit("--team and --season are required")

    lookup_params = SqlParams()
    drifted = query_sql(
        f"""
        select distinct on (sr.id)
          sr.id as source_row_id,
          sr.source_row_number,
          sf.file_name,
          a.id as adjudication_id,
          a.decision,
          a.rationale,
          latest.version_number as latest_version_number,
          latest.eligibility_status as latest_eligibility_status
        from audit.adjudications a
        join ingestion.source_rows sr on sr.id = a.source_row_id
        join ingestion.source_files sf on sf.id = sr.source_file_id
        join lateral (
          select rv.version_number, rv.eligibility_status
          from processing.record_versions rv
          where rv.source_row_id = sr.id
          order by rv.version_number desc
          limit 1
        ) latest on true
        where sf.team = {lookup_params.text(team)}
          and sf.season = {lookup_params.text(season)}
          and a.field_name = 'analysis_eligibility_status'
          and a.decision ? 'analysis_eligibility_status'
          and (a.decision ->> 'analysis_eligibility_status') <> latest.eligibility_status
        order by sr.id, a.decided_at desc
        """,
        lookup_params.values,
    )

    if not drifted:
        print(
            json.dumps(
                {
                    "team": team,
                    "season": season,
                    "status": "no_op",
                    "reason": (
                        "no standing audit.adjudications decision (field_name = "
                        "analysis_eligibility_status) differs from the current latest "
                        "processing.record_versions.eligibility_status for its source row; "
                        "nothing to reapply"
                    ),
                    "drifted_rows": 0,
                },
                indent=2,
            )
        )
        return

    reason_code_seeded = query_sql(
        "select 1 from audit.reason_codes where code = 'adjudication_reapplied'"
    )
    if not reason_code_seeded:
        raise SystemExit(
            "reapply-adjudications found drifted rows but reason code 'adjudication_reapplied' is "
            "not seeded; run migration 20260710000137_adjudication_reapplication first"
        )

    params = SqlParams()
    provenance = run_provenance()
    output_hash = sha256_json(
        [
            {
                "source_row_id": row["source_row_id"],
                "new_version_number": int(row["latest_version_number"]) + 1,
                "decision": row["decision"],
            }
            for row in drifted
        ]
    )

    record_sql = []
    event_sql = []
    for row in drifted:
        source_row_id_p = params.text(row["source_row_id"])
        new_version = int(row["latest_version_number"]) + 1
        decision = row["decision"]
        adjudicated_status = decision.get("analysis_eligibility_status")
        rationale = clean_text(str(decision.get("rationale") or row["rationale"] or ""))
        record_sql.append(
            f"""
            insert into processing.record_versions
              (source_row_id, step_run_id, version_number, record_state, eligibility_status)
            select {source_row_id_p}::uuid, step.id, {new_version},
              coalesce(previous.record_state, '{{}}'::jsonb) || {params.jsonb(decision)},
              {params.text(adjudicated_status)}
            from current_step step
            left join lateral (
              select rv.record_state
              from processing.record_versions rv
              where rv.source_row_id = {source_row_id_p}::uuid
              order by rv.version_number desc
              limit 1
            ) previous on true
            ;
            """
        )
        event_sql.append(
            f"""
            insert into audit.record_events
              (step_run_id, source_row_id, field_name, old_value, new_value, action, reason_code, rationale, rule_version, review_status)
            select step.id, {source_row_id_p}::uuid, 'analysis_eligibility_status',
              {params.jsonb(row["latest_eligibility_status"])}, {params.jsonb(adjudicated_status)}, 'reapply',
              'adjudication_reapplied',
              {params.text(f'Standing audit.adjudications decision {row["adjudication_id"]} reapplied by reapply-adjudications backfill: {rationale}')},
              {params.text(args.step_version)}, 'adjudicated'
            from current_step step;
            """
        )

    sql = f"""
      do $$
      begin
        if not exists (select 1 from audit.reason_codes where code = 'adjudication_reapplied') then
          raise exception 'reapply-adjudications requires reason code adjudication_reapplied to be seeded by migration first';
        end if;
      end $$;

      create temp table current_run on commit drop as
      with run as (
        insert into audit.pipeline_runs
          (command, team, season, status, output_hash, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values (
          'reapply-adjudications', {params.text(team)}, {params.text(season)}, 'succeeded',
          {params.text(output_hash)},
          {params.jsonb({'drifted_rows': len(drifted), 'step_version': args.step_version})},
          now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      )
      select id from run;

      create temp table current_step on commit drop as
      with step as (
        insert into audit.step_runs
          (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count, counts_by_team, ended_at)
        select id, 'adjudication_reapplication', {params.text(args.step_version)}, 'adjudication_reapplied',
          {len(drifted)}, {len(drifted)},
          {params.jsonb({team: {'reapplied_rows': len(drifted)}})}, now()
        from current_run
        returning id
      )
      select id from step;

      {"".join(record_sql)}
      {"".join(event_sql)}
    """
    run_sql(sql, params.values)
    print(
        json.dumps(
            {
                "team": team,
                "season": season,
                "status": "reapplied",
                "drifted_rows": len(drifted),
                "rows": [
                    {
                        "source_row_number": row["source_row_number"],
                        "file": row["file_name"],
                        "new_version_number": int(row["latest_version_number"]) + 1,
                        "new_eligibility_status": row["decision"].get("analysis_eligibility_status"),
                    }
                    for row in drifted
                ],
            },
            indent=2,
        )
    )


def find_missing_standing_adjudications(season: str) -> list[dict[str, Any]]:
    """Read-only: source rows whose latest processing.record_versions row is an
    adjudicated duplicate exclusion written by a duplicate_adjudication step,
    but which have no standing eligibility decision in audit.adjudications.
    These are adjudication runs that predate the audit.adjudications insert in
    adjudicate-duplicate-exclusion (added in commit 6e4c50f); the reapplication
    safety net cannot protect them on a future rerun until backfilled. Every
    reported value is copied from recorded evidence: the record_version state
    (which for these historical rows is exactly the decision JSON the
    adjudicate command wrote), the step_run, the pipeline run parameters, and
    the record_events rationale.
    """
    lookup_params = SqlParams()
    return query_sql(
        f"""
        select
          sf.team, sf.season, sf.file_name, sr.source_row_number,
          sr.id as source_row_id,
          rv.version_number, rv.record_state, rv.step_run_id,
          pr.id as pipeline_run_id, pr.parameters as run_parameters,
          ev.rationale as original_rationale
        from processing.record_versions rv
        join ingestion.source_rows sr on sr.id = rv.source_row_id
        join ingestion.source_files sf on sf.id = sr.source_file_id
        join audit.step_runs st on st.id = rv.step_run_id
        join audit.pipeline_runs pr on pr.id = st.pipeline_run_id
        left join audit.record_events ev on ev.step_run_id = st.id
          and ev.source_row_id = sr.id
          and ev.field_name = 'analysis_eligibility_status'
        where sf.season = {lookup_params.text(season)}
          and rv.eligibility_status = 'excluded_duplicate_adjudicated'
          and st.step_name = 'duplicate_adjudication'
          and pr.command = 'adjudicate-duplicate-exclusion'
          and rv.version_number = (
            select max(rv2.version_number)
            from processing.record_versions rv2
            where rv2.source_row_id = rv.source_row_id
          )
          and not exists (
            select 1 from audit.adjudications a
            where a.source_row_id = sr.id
              and a.field_name = 'analysis_eligibility_status'
          )
        order by sf.team, sr.source_row_number
        """,
        lookup_params.values,
    )


def backfill_standing_adjudications(args: argparse.Namespace) -> None:
    """Backfill audit.adjudications rows for adjudicated duplicate exclusions
    whose runs predate the adjudications-table insert in
    adjudicate-duplicate-exclusion (commit 6e4c50f). Copies each decision
    verbatim from the recorded evidence (the historical record_version state
    is exactly the decision JSON the adjudicate command wrote), points
    consumed_by_step_run_id at the existing duplicate_adjudication step_run,
    and changes no record state -- processing.record_versions and curated
    rows are untouched. With --plan, prints the exact rows that would be
    inserted and exits without writing. The reviewer of the original runs was
    not recorded in their parameters, so the reviewer value is backfill
    attribution, marked as such in the rationale; do not treat it as a
    recorded fact about the original run.
    """
    season = clean_text(args.season)
    if not season:
        raise SystemExit("--season is required")
    reviewer = clean_text(args.reviewer)
    if not reviewer:
        raise SystemExit("--reviewer is required")

    candidates = find_missing_standing_adjudications(season)
    planned = []
    for row in candidates:
        decision = row["record_state"]
        adjudicated_status = clean_text(str(decision.get("analysis_eligibility_status", "")))
        if adjudicated_status != "excluded_duplicate_adjudicated":
            raise SystemExit(
                f"backfill candidate {row['team']} row {row['source_row_number']} has unexpected "
                f"record_state (analysis_eligibility_status={adjudicated_status!r}); refusing to backfill"
            )
        flat = json.dumps(row)
        if re.search(r"\bTeam [A-Z]\b", flat):
            raise SystemExit(
                f"backfill candidate {row['team']} row {row['source_row_number']} contains a protected "
                "Team A-Z alias value; redact before backfilling"
            )
        original_rationale = clean_text(
            str(row["original_rationale"] or decision.get("rationale") or "")
        )
        planned.append(
            {
                "team": row["team"],
                "season": row["season"],
                "file_name": row["file_name"],
                "source_row_number": row["source_row_number"],
                "source_row_id": row["source_row_id"],
                "field_name": "analysis_eligibility_status",
                "decision": decision,
                "rationale": (
                    "Evidence-based backfill of a standing decision that predates the "
                    "audit.adjudications insert in adjudicate-duplicate-exclusion: original run "
                    f"{row['pipeline_run_id']} (adjudicate-duplicate-exclusion, version "
                    f"{row['version_number']}) recorded this decision in pipeline run parameters, "
                    f"record events, and record version state. Original rationale: {original_rationale} "
                    "Reviewer value is backfill attribution; the original run predates reviewer recording."
                ),
                "reviewer": reviewer,
                "consumed_by_step_run_id": row["step_run_id"],
            }
        )

    if args.plan or not planned:
        print(
            json.dumps(
                {
                    "season": season,
                    "status": "plan" if planned else "no_op",
                    "missing_standing_adjudications": len(planned),
                    "rows": planned,
                },
                indent=2,
            )
        )
        return

    params = SqlParams()
    provenance = run_provenance()
    output_hash = sha256_json(planned)
    teams = sorted({row["team"] for row in planned})
    counts_by_team = {team: {"backfilled_adjudications": sum(1 for row in planned if row["team"] == team)} for team in teams}

    insert_sql = []
    for row in planned:
        insert_sql.append(
            f"""
            insert into audit.adjudications
              (source_row_id, field_name, decision, rationale, reviewer, consumed_by_step_run_id)
            select {params.text(row["source_row_id"])}::uuid, 'analysis_eligibility_status',
              {params.jsonb(row["decision"])}, {params.text(row["rationale"])},
              {params.text(row["reviewer"])}, {params.text(row["consumed_by_step_run_id"])}::uuid
            where not exists (
              select 1 from audit.adjudications existing
              where existing.source_row_id = {params.text(row["source_row_id"])}::uuid
                and existing.field_name = 'analysis_eligibility_status'
            );
            """
        )

    sql = f"""
      do $$
      begin
        if not exists (select 1 from audit.reason_codes where code = 'standing_adjudication_backfill') then
          raise exception 'backfill-standing-adjudications requires reason code standing_adjudication_backfill to be seeded by migration first';
        end if;
      end $$;

      create temp table current_run on commit drop as
      with run as (
        insert into audit.pipeline_runs
          (command, season, status, output_hash, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values (
          'backfill-standing-adjudications', {params.text(season)}, 'succeeded',
          {params.text(output_hash)},
          {params.jsonb({'teams': teams, 'backfilled_rows': len(planned), 'reviewer': args.reviewer})},
          now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      )
      select id from run;

      insert into audit.step_runs
        (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count, counts_by_team, output_hash, ended_at)
      select id, 'standing_adjudication_backfill', {params.text(args.step_version)}, 'standing_adjudication_backfill',
        {len(planned)}, {len(planned)}, {params.jsonb(counts_by_team)}, {params.text(output_hash)}, now()
      from current_run;

      {"".join(insert_sql)}
    """
    run_sql(sql, params.values)
    print(
        json.dumps(
            {
                "season": season,
                "status": "backfilled",
                "rows": [
                    {
                        "team": row["team"],
                        "source_row_number": row["source_row_number"],
                        "consumed_by_step_run_id": row["consumed_by_step_run_id"],
                    }
                    for row in planned
                ],
            },
            indent=2,
        )
    )


def redact_protected_team_aliases(args: argparse.Namespace) -> None:
    """Retroactively purge exact 'Team A'-'Team Z' league-alias values left in
    live rows (a gap predating alias redaction at ingest time). Replaces
    values only -- no key or row is ever deleted -- and records one pipeline
    run with two step_runs (one per table) plus one audit.record_events row
    per changed row. See PROTECTED_ALIAS_OLD_VALUE_MARKER for why old_value is
    a marker rather than the real prior string.
    """
    if args.scope != "all":
        raise SystemExit("redact-protected-team-aliases currently only supports --scope all")

    params = SqlParams()
    provenance = run_provenance()
    pattern = params.text(PROTECTED_ALIAS_SQL_PATTERN)
    marker = params.text(PROTECTED_ALIAS_REDACTED_MARKER)
    old_marker = params.text(PROTECTED_ALIAS_OLD_VALUE_MARKER)
    rule_version = params.text(PROTECTED_ALIAS_REDACTION_RULE_VERSION)
    field_name_team_alias = params.text("team_alias")
    step_source_rows = params.text("redact_source_row_values")
    step_record_versions = params.text("redact_record_version_team_alias")

    sql = f"""
      do $$
      begin
        if not exists (select 1 from audit.reason_codes where code = 'protected_metadata_redaction') then
          raise exception 'redact-protected-team-aliases requires reason code protected_metadata_redaction to be seeded by migration first';
        end if;
      end $$;

      create temp table current_run on commit drop as
      with run as (
        insert into audit.pipeline_runs (command, status, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values (
          'redact-protected-team-aliases', 'succeeded',
          {params.jsonb({"scope": args.scope, "pattern": PROTECTED_ALIAS_SQL_PATTERN, "rule_version": PROTECTED_ALIAS_REDACTION_RULE_VERSION})},
          now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      )
      select id from run;

      -- Step 1: ingestion.source_rows.source_values (any matching value, any key)
      create temp table source_rows_step on commit drop as
      with step as (
        insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, ended_at)
        select id, {step_source_rows}, {rule_version}, 'protected_metadata_redaction', now()
        from current_run
        returning id
      )
      select id from step;

      create temp table source_rows_changes on commit drop as
      select
        sr.id as source_row_id,
        sf.team as team,
        (
          select array_agg(e.key order by e.key)
          from jsonb_each(sr.source_values) e
          where jsonb_typeof(e.value) = 'string' and (e.value #>> '{{}}') ~ {pattern}
        ) as changed_keys,
        (
          select jsonb_object_agg(
            e.key,
            case when jsonb_typeof(e.value) = 'string' and (e.value #>> '{{}}') ~ {pattern}
                 then to_jsonb({marker}::text)
                 else e.value
            end
          )
          from jsonb_each(sr.source_values) e
        ) as redacted_values
      from ingestion.source_rows sr
      join ingestion.source_files sf on sf.id = sr.source_file_id
      where exists (
        select 1 from jsonb_each(sr.source_values) e
        where jsonb_typeof(e.value) = 'string' and (e.value #>> '{{}}') ~ {pattern}
      );

      update ingestion.source_rows sr
      set source_values = c.redacted_values
      from source_rows_changes c
      where sr.id = c.source_row_id;

      insert into audit.record_events
        (step_run_id, source_row_id, field_name, old_value, new_value, action, reason_code, rationale, rule_version, review_status)
      select
        (select id from source_rows_step),
        c.source_row_id,
        array_to_string(c.changed_keys, ','),
        jsonb_build_object('redacted_keys', to_jsonb(c.changed_keys), 'value', to_jsonb({old_marker}::text)),
        jsonb_build_object('redacted_keys', to_jsonb(c.changed_keys), 'value', to_jsonb({marker}::text)),
        'redact', 'protected_metadata_redaction',
        'Purged protected Team A-Z league-alias placeholder value(s) from stored source_values; keys retained, only the value(s) replaced.',
        {rule_version}, 'not_required'
      from source_rows_changes c;

      update audit.step_runs
      set input_count = (select count(*) from ingestion.source_rows),
          output_count = (select count(*) from source_rows_changes),
          counts_by_team = (
            select coalesce(jsonb_object_agg(team, cnt), '{{}}'::jsonb)
            from (select team, count(*) as cnt from source_rows_changes group by team) t
          ),
          ended_at = now()
      where id = (select id from source_rows_step);

      -- Step 2: processing.record_versions.record_state ->> 'team_alias'
      create temp table record_versions_step on commit drop as
      with step as (
        insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, ended_at)
        select id, {step_record_versions}, {rule_version}, 'protected_metadata_redaction', now()
        from current_run
        returning id
      )
      select id from step;

      create temp table record_versions_changes on commit drop as
      select rv.id as record_version_id, sr.id as source_row_id, sf.team as team
      from processing.record_versions rv
      join ingestion.source_rows sr on sr.id = rv.source_row_id
      join ingestion.source_files sf on sf.id = sr.source_file_id
      where rv.record_state ->> 'team_alias' ~ {pattern};

      update processing.record_versions rv
      set record_state = jsonb_set(rv.record_state, array[{field_name_team_alias}], to_jsonb({marker}::text))
      from record_versions_changes c
      where rv.id = c.record_version_id;

      insert into audit.record_events
        (step_run_id, source_row_id, field_name, old_value, new_value, action, reason_code, rationale, rule_version, review_status)
      select
        (select id from record_versions_step),
        c.source_row_id,
        {field_name_team_alias},
        to_jsonb({old_marker}::text),
        to_jsonb({marker}::text),
        'redact', 'protected_metadata_redaction',
        'Purged protected Team A-Z league-alias placeholder value from processing.record_versions.record_state.team_alias; key retained, only the value replaced.',
        {rule_version}, 'not_required'
      from record_versions_changes c;

      update audit.step_runs
      set input_count = (select count(*) from processing.record_versions),
          output_count = (select count(*) from record_versions_changes),
          counts_by_team = (
            select coalesce(jsonb_object_agg(team, cnt), '{{}}'::jsonb)
            from (select team, count(*) as cnt from record_versions_changes group by team) t
          ),
          ended_at = now()
      where id = (select id from record_versions_step);

      update audit.pipeline_runs
      set output_hash = md5(
        coalesce((select count(*) from source_rows_changes), 0)::text || ':' ||
        coalesce((select count(*) from record_versions_changes), 0)::text
      )
      where id = (select id from current_run);

      {protected_alias_scan_sql('redact-protected-team-aliases post-check')}
    """
    run_sql(sql, params.values)
    print(
        "redact-protected-team-aliases: pipeline run recorded (source_rows + "
        "record_versions steps); re-query row counts read-only to confirm."
    )


def retire_releases(args: argparse.Namespace) -> None:
    """Flip named reporting.aggregate_releases rows from 'approved' to
    'retired' in one audited run. Rows are kept (never deleted); only
    exact, explicitly named labels are accepted -- no implicit globbing of
    what counts as a 'smoke' release, so this stays a reviewed, bounded
    action rather than a standing cleanup rule.
    """
    labels = [clean_text(label) for label in args.labels.split(",") if clean_text(label)]
    if not labels:
        raise SystemExit("--labels must name at least one release_label to retire")

    params = SqlParams()
    provenance = run_provenance()
    labels_json = params.jsonb(labels)
    rule_version = params.text(AGGREGATE_RELEASE_RETIREMENT_RULE_VERSION)
    step_name = params.text("retire_aggregate_releases")

    sql = f"""
      do $$
      begin
        if not exists (select 1 from audit.reason_codes where code = 'aggregate_release_retired') then
          raise exception 'retire-releases requires reason code aggregate_release_retired to be seeded by migration first';
        end if;
        if exists (
          select 1
          from jsonb_array_elements_text({labels_json}) expected(label)
          where not exists (
            select 1 from reporting.aggregate_releases r
            where r.release_label = expected.label and r.status = 'approved'
          )
        ) then
          raise exception 'retire-releases requires every named label to currently exist with status approved';
        end if;
      end $$;

      create temp table current_run on commit drop as
      with run as (
        insert into audit.pipeline_runs (command, status, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values (
          'retire-releases', 'succeeded',
          {params.jsonb({"labels": labels, "rationale": args.rationale, "reviewer": args.reviewer})},
          now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      )
      select id from run;

      create temp table retire_step on commit drop as
      with step as (
        insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, ended_at)
        select id, {step_name}, {rule_version}, 'aggregate_release_retired', now()
        from current_run
        returning id
      )
      select id from step;

      update reporting.aggregate_releases r
      set status = 'retired'
      where r.release_label in (select jsonb_array_elements_text({labels_json}))
        and r.status = 'approved';

      update audit.step_runs
      set input_count = jsonb_array_length({labels_json}),
          output_count = (
            select count(*) from reporting.aggregate_releases
            where release_label in (select jsonb_array_elements_text({labels_json}))
              and status = 'retired'
          ),
          counts_by_team = '{{}}'::jsonb,
          ended_at = now()
      where id = (select id from retire_step);
    """
    run_sql(sql, params.values)
    print(f"retired {len(labels)} release(s): {', '.join(labels)}")


CURATED_LAYER_MIGRATION_VERSION = "20260709233356"
CURATED_BUILD_RULE_VERSION = "curated_build_2026-08-22_v2"
CURATED_FIXTURE_LOAD_RULE_VERSION = "curated_fixture_load_2026-07-10_v1"
ANALYSIS_VIEWS_MIGRATION_VERSION = "20260710100000"
ANALYSIS_VIEW_VERSION_SUFFIX = "v1"
# Fixed by docs/EXPOSURE_CLEANING_PROTOCOL.md: "Calculate fixture match
# exposure as 20 player-hours per team per match." Not a tunable parameter
# for the curated layer (unlike build_fixture_exposure's CLI-overridable
# --player-hours-per-team-match, which produces a separate file-based
# report rather than curated rows).
PLAYER_HOURS_PER_TEAM_MATCH = 20.0
URC_FIXTURES_2024_25_CORRECTED_PATH = "data/intake/2024-25/fixtures/urc_fixtures_2024_25.corrected.csv"
URC_FIXTURES_2024_25_CORRECTED_SHA256 = "9608ff7e932cf76743eeb6de7d3bce6f5746ab1dfa4cec80a01f001ec2e9c39c"
YEAR2_FIXTURE_PROVENANCE_MIGRATION_VERSION = "20260815010000"
YEAR2_FIXTURE_ALIAS_MIGRATION_VERSION = "20260822010000"

CURATED_EXPOSURE_SCOPE_PROJECTION = {
    "excluded": "scope_unknown_included",
    "outside_protocol_window": "scope_unknown_included",
    "within_protocol_window_scope_unknown": "scope_unknown_included",
}
CURATED_EXPOSURE_SCOPE_STATUSES = frozenset(
    {"in_scope_explicit", "scope_unknown_included", "out_of_scope_explicit"}
)


def canonical_curated_exposure_scope_status(value: str | None) -> str | None:
    if value is None or not value.strip():
        return None
    if value in CURATED_EXPOSURE_SCOPE_STATUSES:
        return value
    projected = CURATED_EXPOSURE_SCOPE_PROJECTION.get(value)
    if projected is None:
        raise ValueError(f"unsupported exposure scope_status for curated projection: {value!r}")
    return projected


def resolve_team_key(team: str) -> str:
    """Resolve a legacy team value exactly as stored in
    ingestion.source_files.team (e.g. 'Munster', 'glasgow') to its public
    reporting.teams.team_key via reporting.team_key_aliases. Never guesses:
    an unmapped or excluded alias is a hard error naming the exact value.
    """
    params = SqlParams()
    rows = query_sql(
        f"select team_key, excluded from reporting.team_key_aliases where alias = {params.text(team)}",
        params.values,
    )
    if not rows:
        raise SystemExit(
            f"no reporting.team_key_aliases entry for team {team!r}; add it by migration before building curated data"
        )
    if rows[0]["excluded"] or not rows[0]["team_key"]:
        raise SystemExit(
            f"team {team!r} is an excluded alias (e.g. a smoke-test artifact); refusing to build curated data for it"
        )
    return rows[0]["team_key"]


def latest_curated_source_version_ids(team: str, season: str, file_name_pattern: str) -> list[str]:
    """record_version ids for the latest version_number of every accepted
    source row in the source file matching file_name_pattern
    ('%injury%' / '%exposure%', verified against every live
    ingestion.source_files.file_name) for team/season. Read-only.
    """
    params = SqlParams()
    rows = query_sql(
        f"""
        select rv.id
        from processing.record_versions rv
        join ingestion.source_rows sr on sr.id = rv.source_row_id
        join ingestion.source_files sf on sf.id = sr.source_file_id
        where sf.team = {params.text(team)} and sf.season = {params.text(season)}
          and sf.file_name like {params.text(file_name_pattern)}
          and rv.version_number = (
            select max(rv2.version_number)
            from processing.record_versions rv2
            where rv2.source_row_id = rv.source_row_id
          )
        order by rv.id
        """,
        params.values,
    )
    return [row["id"] for row in rows]


def curated_source_version_set_hash(injury_ids: list[str], exposure_ids: list[str]) -> str:
    return sha256_json({"injury": sorted(injury_ids), "exposure": sorted(exposure_ids)})


def build_curated(args: argparse.Namespace) -> None:
    team = clean_text(args.team)
    season = clean_text(args.season)
    if not team or not season:
        raise SystemExit("--team and --season are required")
    team_key = resolve_team_key(team)

    injury_ids = latest_curated_source_version_ids(team, season, "%injury%")
    exposure_ids = latest_curated_source_version_ids(team, season, "%exposure%")
    if not injury_ids or not exposure_ids:
        raise SystemExit(
            f"no processing.record_versions found for team={team!r} season={season!r} "
            f"(injury rows: {len(injury_ids)}, exposure rows: {len(exposure_ids)}); "
            "run process-intake and process-exposure for this team before build-curated"
        )
    source_version_set_hash = curated_source_version_set_hash(injury_ids, exposure_ids)
    source_version_set_count = len(injury_ids) + len(exposure_ids)

    lookup_params = SqlParams()
    active_builds = query_sql(
        f"""
        select id, source_version_set_hash, status
        from curated.builds
        where team_key = {lookup_params.text(team_key)} and season = {lookup_params.text(season)} and status = 'active'
        """,
        lookup_params.values,
    )
    if active_builds:
        active = active_builds[0]
        if active["source_version_set_hash"] == source_version_set_hash and not args.rebuild:
            print(
                json.dumps(
                    {
                        "status": "no_op",
                        "reason": "an active curated build already matches the current source version set",
                        "team_key": team_key,
                        "season": season,
                        "build_id": active["id"],
                    },
                    indent=2,
                )
            )
            return
        if active["source_version_set_hash"] != source_version_set_hash and not args.rebuild:
            raise SystemExit(
                f"an active curated build exists for team_key={team_key!r} season={season!r} but the live "
                "source version set has changed since it was built; pass --rebuild to explicitly supersede it"
            )

    fixture_params = SqlParams()
    fixture_rows = query_sql(
        f"select count(*) as n, min(source_file_sha256) as sha from curated.fixtures where season = {fixture_params.text(season)}",
        fixture_params.values,
    )
    if not fixture_rows or int(fixture_rows[0]["n"]) == 0:
        raise SystemExit(f"curated.fixtures has no rows for season={season!r}; run load-curated-fixtures first")

    params = SqlParams()
    provenance = run_provenance()
    team_p = params.text(team)
    season_p = params.text(season)
    team_key_p = params.text(team_key)
    injury_pattern_p = params.text("%injury%")
    exposure_pattern_p = params.text("%exposure%")
    rebuild_flag = bool(getattr(args, "rebuild", False))

    supersede_sql = (
        f"update curated.builds set status = 'superseded' "
        f"where team_key = {team_key_p} and season = {season_p} and status = 'active';"
        if rebuild_flag
        else ""
    )

    sql = f"""
      do $$
      begin
        if not exists (select 1 from supabase_migrations.schema_migrations where version = '{CURATED_LAYER_MIGRATION_VERSION}') then
          raise exception 'build-curated requires migration {CURATED_LAYER_MIGRATION_VERSION}_curated_layer';
        end if;
        if (select count(*) from audit.reason_codes where code in ('curated_projection', 'curated_denominator_derivation')) <> 2 then
          raise exception 'build-curated requires reason codes curated_projection and curated_denominator_derivation to be seeded by migration first';
        end if;
        -- Count only PROCESSED source files: files with at least one
        -- processing.record_versions row. A superseded mis-ingested
        -- registration with zero record_versions (kept immutably, per the
        -- never-delete rule) must not block the build; two PROCESSED
        -- injury/exposure files still fail as genuinely ambiguous. The
        -- projection inserts and the version-set hash join through
        -- record_versions, so an unprocessed registration cannot reach
        -- curated rows either way; this guard is the only place that
        -- counted raw registrations.
        if (select count(distinct sf.id) from ingestion.source_files sf where sf.team = {team_p} and sf.season = {season_p} and sf.file_name like {injury_pattern_p}
            and exists (select 1 from ingestion.source_rows sr join processing.record_versions rv on rv.source_row_id = sr.id where sr.source_file_id = sf.id)) <> 1 then
          raise exception 'build-curated requires exactly one processed injury source file for team=% season=%', {team_p}, {season_p};
        end if;
        if (select count(distinct sf.id) from ingestion.source_files sf where sf.team = {team_p} and sf.season = {season_p} and sf.file_name like {exposure_pattern_p}
            and exists (select 1 from ingestion.source_rows sr join processing.record_versions rv on rv.source_row_id = sr.id where sr.source_file_id = sf.id)) <> 1 then
          raise exception 'build-curated requires exactly one processed exposure source file for team=% season=%', {team_p}, {season_p};
        end if;
        if exists (
          select 1
          from processing.record_versions rv
          join ingestion.source_rows sr on sr.id = rv.source_row_id
          join ingestion.source_files sf on sf.id = sr.source_file_id
          where sf.team = {team_p} and sf.season = {season_p} and sf.file_name like {exposure_pattern_p}
            and rv.version_number = (
              select max(rv2.version_number) from processing.record_versions rv2 where rv2.source_row_id = rv.source_row_id
            )
            and nullif(btrim(rv.record_state ->> 'scope_status'), '') is not null
            and rv.record_state ->> 'scope_status' not in (
              'in_scope_explicit', 'scope_unknown_included', 'out_of_scope_explicit',
              'excluded', 'outside_protocol_window', 'within_protocol_window_scope_unknown'
            )
        ) then
          raise exception 'build-curated found unsupported exposure scope_status in latest processed exposure rows';
        end if;
        if not exists (select 1 from curated.fixtures where season = {season_p}) then
          raise exception 'curated.fixtures has no rows for season=%; run load-curated-fixtures first', {season_p};
        end if;
      end $$;

      {supersede_sql}

      create temp table current_run on commit drop as
      with run as (
        insert into audit.pipeline_runs (command, team, season, status, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values (
          'build-curated', {team_p}, {season_p}, 'succeeded',
          {params.jsonb({
            "team_key": team_key,
            "rebuild": rebuild_flag,
            "rule_version": CURATED_BUILD_RULE_VERSION,
            "source_version_set_hash": source_version_set_hash,
            "source_version_set_count": source_version_set_count,
          })},
          now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      )
      select id from run;

      create temp table new_build on commit drop as
      with build as (
        insert into curated.builds (pipeline_run_id, team_key, season, source_version_set_count, source_version_set_hash, status)
        select id, {team_key_p}, {season_p}, {source_version_set_count}, {params.text(source_version_set_hash)}, 'active'
        from current_run
        returning id
      )
      select id from build;

      create temp table injuries_step on commit drop as
      with step as (
        insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, ended_at)
        select id, 'curated_injuries', {params.text(CURATED_BUILD_RULE_VERSION)}, 'curated_projection', now()
        from current_run
        returning id
      )
      select id from step;

      insert into curated.injuries (
        source_row_id, record_version_id, team_key, season, player_uid, injury_uid,
        date_injured, days_injured, derived_return_date, is_closed,
        activity_context, contact_context, recurrence_status, severity_category,
        body_location, injury_type, problem_type, eligibility_status,
        received_in_team_status, urc_match_scope,
        field_origins, source_locator, curated_build_id
      )
      select
        rv.source_row_id, rv.id, {team_key_p}, {season_p},
        rv.record_state ->> 'player_uid',
        rv.record_state ->> 'injury_uid',
        nullif(rv.record_state ->> 'date_injured', '')::date,
        nullif(rv.record_state ->> 'days_injured_source', '')::numeric::int,
        nullif(rv.record_state ->> 'derived_return_date', '')::date,
        nullif(rv.record_state ->> 'is_closed', '')::boolean,
        rv.record_state ->> 'activity_context',
        rv.record_state ->> 'contact_context',
        rv.record_state ->> 'recurrence_status',
        rv.record_state ->> 'severity_category',
        rv.record_state ->> 'body_location',
        rv.record_state ->> 'injury_type',
        rv.record_state ->> 'problem_type',
        rv.eligibility_status,
        -- Phase 3.5 cohort-signal capture (Adjudication 4): absent for any
        -- record_version written before this migration (jsonb ->> on a
        -- missing key returns NULL), which is exactly the NULL-safe
        -- non-excluding default analysis.injury_cohort_v1 relies on for
        -- teams that have not been reprocessed.
        rv.record_state ->> 'received_in_team_status',
        rv.record_state ->> 'urc_match_scope',
        coalesce(rv.record_state -> 'field_origins', '{{}}'::jsonb),
        coalesce(rv.record_state -> 'source_locator', '{{}}'::jsonb),
        (select id from new_build)
      from processing.record_versions rv
      join ingestion.source_rows sr on sr.id = rv.source_row_id
      join ingestion.source_files sf on sf.id = sr.source_file_id
      where sf.team = {team_p} and sf.season = {season_p} and sf.file_name like {injury_pattern_p}
        and rv.version_number = (
          select max(rv2.version_number) from processing.record_versions rv2 where rv2.source_row_id = rv.source_row_id
        );

      update audit.step_runs
      set input_count = {len(injury_ids)},
          output_count = (select count(*) from curated.injuries where curated_build_id = (select id from new_build)),
          counts_by_team = jsonb_build_object({team_key_p}, jsonb_build_object('injuries', (select count(*) from curated.injuries where curated_build_id = (select id from new_build)))),
          ended_at = now()
      where id = (select id from injuries_step);

      create temp table exposure_step on commit drop as
      with step as (
        insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, ended_at)
        select id, 'curated_exposure', {params.text(CURATED_BUILD_RULE_VERSION)}, 'curated_projection', now()
        from current_run
        returning id
      )
      select id from step;

      insert into curated.exposure (
        source_row_id, record_version_id, team_key, season, player_uid, grain,
        session_date, week_start_date, minutes_clean, distance_m_clean,
        scope_status, exclusion_reasons, eligibility_status, source_locator, curated_build_id
      )
      select
        rv.source_row_id, rv.id, {team_key_p}, {season_p},
        rv.record_state ->> 'player_uid',
        rv.record_state ->> 'exposure_grain',
        nullif(rv.record_state ->> 'session_date', '')::date,
        nullif(rv.record_state ->> 'week_start_date', '')::date,
        nullif(rv.record_state ->> 'minutes_total_clean', '')::numeric,
        nullif(rv.record_state ->> 'distance_total_m_clean', '')::numeric,
        case
          when nullif(btrim(rv.record_state ->> 'scope_status'), '') is null then null
          when rv.record_state ->> 'scope_status' = 'excluded' then 'scope_unknown_included'
          when rv.record_state ->> 'scope_status' = 'outside_protocol_window' then 'scope_unknown_included'
          when rv.record_state ->> 'scope_status' = 'within_protocol_window_scope_unknown' then 'scope_unknown_included'
          else rv.record_state ->> 'scope_status'
        end,
        coalesce(
          (select array_agg(x) from jsonb_array_elements_text(coalesce(rv.record_state -> 'exclusion_reasons', '[]'::jsonb)) x),
          '{{}}'::text[]
        ),
        rv.eligibility_status,
        coalesce(rv.record_state -> 'source_locator', '{{}}'::jsonb),
        (select id from new_build)
      from processing.record_versions rv
      join ingestion.source_rows sr on sr.id = rv.source_row_id
      join ingestion.source_files sf on sf.id = sr.source_file_id
      where sf.team = {team_p} and sf.season = {season_p} and sf.file_name like {exposure_pattern_p}
        and rv.version_number = (
          select max(rv2.version_number) from processing.record_versions rv2 where rv2.source_row_id = rv.source_row_id
        );

      update audit.step_runs
      set input_count = {len(exposure_ids)},
          output_count = (select count(*) from curated.exposure where curated_build_id = (select id from new_build)),
          counts_by_team = jsonb_build_object({team_key_p}, jsonb_build_object('exposure', (select count(*) from curated.exposure where curated_build_id = (select id from new_build)))),
          ended_at = now()
      where id = (select id from exposure_step);

      create temp table denominator_step on commit drop as
      with step as (
        insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, ended_at)
        select id, 'curated_team_exposure_denominator', {params.text(CURATED_BUILD_RULE_VERSION)}, 'curated_denominator_derivation', now()
        from current_run
        returning id
      )
      select id from step;

      create temp table denom_calc on commit drop as
      select
        (
          select count(*) from curated.fixtures f
          where f.season = {season_p} and (f.home_team_key = {team_key_p} or f.away_team_key = {team_key_p})
        ) as matches_played,
        (
          select coalesce(sum(e.minutes_clean), 0) / 60
          from curated.exposure e
          where e.curated_build_id = (select id from new_build) and e.eligibility_status = 'included_pending_protocol'
        ) as total_hours;

      insert into curated.team_exposure_denominators (
        team_key, season, matches_played, match_hours, training_hours, total_hours,
        method_note, fixture_source_sha256, curated_build_id
      )
      select
        {team_key_p}, {season_p}, matches_played,
        matches_played * {PLAYER_HOURS_PER_TEAM_MATCH},
        total_hours - matches_played * {PLAYER_HOURS_PER_TEAM_MATCH},
        total_hours,
        {params.text(
          'training_hours = total_hours - match_hours; match_hours = matches_played * '
          f'{PLAYER_HOURS_PER_TEAM_MATCH} player-hours per team per match '
          '(docs/EXPOSURE_CLEANING_PROTOCOL.md); total_hours = sum(curated.exposure.minutes_clean) / 60 '
          "where eligibility_status = 'included_pending_protocol'."
        )},
        (select min(source_file_sha256) from curated.fixtures where season = {season_p}),
        (select id from new_build)
      from denom_calc;

      update audit.step_runs
      set input_count = (select matches_played from denom_calc),
          output_count = 1,
          counts_by_team = jsonb_build_object({team_key_p}, (select to_jsonb(denom_calc.*) from denom_calc)),
          ended_at = now()
      where id = (select id from denominator_step);

      update curated.builds
      set row_counts = jsonb_build_object(
            'injuries', (select count(*) from curated.injuries where curated_build_id = (select id from new_build)),
            'exposure', (select count(*) from curated.exposure where curated_build_id = (select id from new_build)),
            'team_exposure_denominators', (select count(*) from curated.team_exposure_denominators where curated_build_id = (select id from new_build))
          ),
          output_hash = md5(
            coalesce((select count(*) from curated.injuries where curated_build_id = (select id from new_build)), 0)::text || ':' ||
            coalesce((select count(*) from curated.exposure where curated_build_id = (select id from new_build)), 0)::text
          )
      where id = (select id from new_build);

      update audit.pipeline_runs
      set output_hash = (select output_hash from curated.builds where id = (select id from new_build))
      where id = (select id from current_run);
    """
    run_sql(sql, params.values)

    summary_params = SqlParams()
    summary = query_sql(
        f"""
        select b.id as build_id, b.row_counts, d.matches_played, d.match_hours, d.training_hours, d.total_hours
        from curated.builds b
        left join curated.team_exposure_denominators d on d.curated_build_id = b.id
        where b.team_key = {summary_params.text(team_key)} and b.season = {summary_params.text(season)} and b.status = 'active'
        """,
        summary_params.values,
    )
    print(json.dumps({"status": "built", "team": team, "team_key": team_key, "season": season, "result": summary[0] if summary else None}, indent=2))


def load_curated_fixtures(args: argparse.Namespace) -> None:
    season = clean_text(args.season)
    path = Path(args.file)
    rows = read_rows(path)
    if not rows:
        raise SystemExit(f"no fixture rows found: {path}")
    missing_columns = [
        column
        for column in ["source_row_number", "stage", "round", "corrected_date", "date_status", "home_team", "away_team"]
        if column not in rows[0]
    ]
    if missing_columns:
        raise SystemExit(f"fixture file missing column(s): {', '.join(missing_columns)}")

    # New-season fixture files must prove their complete official schedule
    # structure before a checksum lookup or any database read. The frozen
    # 2024-25 correction file deliberately stays on its existing path.
    fixture_provenance: list[dict[str, object]] = []
    fixture_contract = fixture_contract_for(season)
    has_fixture_contract = fixture_contract is not None
    if has_fixture_contract:
        assert fixture_contract is not None
        assert_local_evidence_bytes([{
            "role": "fixture_preparation",
            "locator": fixture_contract.evidence_locator,
            "sha256": fixture_contract.evidence_sha256,
        }], "Year 2 fixture load")
        try:
            validate_fixture_rows(season, rows)
        except ValueError as error:
            raise SystemExit(str(error)) from error

    file_hash = sha256_file(path)
    if has_fixture_contract:
        # The first validation above deliberately happens before any checksum
        # or database access.  This second pure construction binds the actual
        # prepared CSV bytes to the already-validated upstream response proof.
        try:
            fixture_provenance = fixture_provenance_rows(
                season, rows, prepared_file_sha256=file_hash,
            )
            validate_fixture_provenance_binding(season, rows, fixture_provenance)
        except ValueError as error:
            raise SystemExit(str(error)) from error

    if str(path) == URC_FIXTURES_2024_25_CORRECTED_PATH and file_hash != URC_FIXTURES_2024_25_CORRECTED_SHA256:
        raise SystemExit(
            "fixture file checksum mismatch for the documented 2024-25 corrected fixture file "
            f"(docs/EXPOSURE_CLEANING_PROTOCOL.md): expected {URC_FIXTURES_2024_25_CORRECTED_SHA256}, got {file_hash}"
        )

    if fixture_provenance:
        year2_contract = release_contract_for(
            "2025-26", YEAR2_2025_26_RELEASE_TUPLE,
        )
        fixture_migrations = tuple(
            item for item in year2_contract.required_migration_contracts
            if item.version in {
                YEAR2_FIXTURE_PROVENANCE_MIGRATION_VERSION,
                YEAR2_FIXTURE_ALIAS_MIGRATION_VERSION,
            }
        )
        if len(fixture_migrations) != 2:
            raise SystemExit("Year 2 fixture load migration contract is incomplete")
        assert_checksum_bound_migrations(
            fixture_migrations, "Year 2 fixture load",
        )

    existing_params = SqlParams()
    existing = query_sql(
        f"select count(*) as n from curated.fixtures where season = {existing_params.text(season)}",
        existing_params.values,
    )
    existing_count = int(existing[0]["n"]) if existing else 0
    if existing_count > 0 and not fixture_provenance:
        print(
            json.dumps(
                {
                    "status": "no_op",
                    "reason": "curated.fixtures already loaded for this season",
                    "season": season,
                    "existing_rows": existing_count,
                },
                indent=2,
            )
        )
        return
    if fixture_provenance and existing_count not in {0, len(rows)}:
        raise SystemExit(
            "curated.fixtures contains a partial Year 2 fixture set; "
            f"expected 0 or {len(rows)} rows, found {existing_count}"
        )

    # Resolve every home/away team name through reporting.team_key_aliases.
    # These are public club names (e.g. 'Cardiff Rugby', 'Glasgow
    # Warriors') read from the fixture-list home_team/away_team columns;
    # the file's home_team_alias/away_team_alias columns (the protected
    # Team A-Z league alias, resolved elsewhere via the Git-ignored
    # protected alias map) are deliberately never read here.
    alias_rows = query_sql("select alias, team_key, excluded from reporting.team_key_aliases")
    alias_map = {row["alias"]: row for row in alias_rows}

    fixture_rows = []
    unresolved: list[str] = []
    for row in rows:
        home = clean_text(row.get("home_team"))
        away = clean_text(row.get("away_team"))
        home_entry = alias_map.get(home)
        away_entry = alias_map.get(away)
        if not home_entry or home_entry["excluded"] or not home_entry["team_key"]:
            unresolved.append(home)
        if not away_entry or away_entry["excluded"] or not away_entry["team_key"]:
            unresolved.append(away)
        fixture_rows.append(
            {
                "source_row_number": int(row["source_row_number"]),
                "stage": clean_text(row.get("stage")),
                "round": clean_text(row.get("round")),
                "match_date": clean_text(row.get("corrected_date")),
                "date_status": clean_text(row.get("date_status")),
                "home_team_key": home_entry["team_key"] if home_entry else None,
                "away_team_key": away_entry["team_key"] if away_entry else None,
            }
        )
    unresolved_unique = sorted({name for name in unresolved if name})
    if unresolved_unique:
        raise SystemExit(
            f"unmapped fixture team name(s) in reporting.team_key_aliases: {', '.join(unresolved_unique)}; "
            "add them by migration before loading curated fixtures"
        )

    params = SqlParams()
    provenance = run_provenance()
    values_sql = ",".join(
        f"""(
          {params.text(season)}, {params.text(r['stage'])}, {params.text(r['round'])},
          {params.text(r['match_date'])}::date, {params.text(r['date_status'])},
          {params.text(r['home_team_key'])}, {params.text(r['away_team_key'])},
          {r['source_row_number']}, {params.text(file_hash)}, (select id from current_run)
        )"""
        for r in fixture_rows
    )
    expected_fixture_values_sql = ",".join(
        f"""(
          {params.text(season)}, {r['source_row_number']},
          {params.text(r['stage'])}, {params.text(r['round'])},
          {params.text(r['match_date'])}::date, {params.text(r['date_status'])},
          {params.text(r['home_team_key'])}, {params.text(r['away_team_key'])},
          {params.text(file_hash)}
        )"""
        for r in fixture_rows
    )
    fixture_exact_binding_sql = (
        f"""
      create temp table expected_curated_fixtures (
        season text not null,
        source_row_number integer not null,
        stage text not null,
        round text not null,
        match_date date not null,
        date_status text not null,
        home_team_key text not null,
        away_team_key text not null,
        source_file_sha256 text not null
      ) on commit drop;

      insert into expected_curated_fixtures values {expected_fixture_values_sql};

      do $$
      begin
        if exists (
          select 1
          from expected_curated_fixtures expected
          full join (
            select season, source_row_number, stage, round, match_date,
              date_status, home_team_key, away_team_key, source_file_sha256
            from curated.fixtures where season = {params.text(season)}
          ) actual using (season, source_row_number)
          where expected.season is null or actual.season is null
             or actual.stage is distinct from expected.stage
             or actual.round is distinct from expected.round
             or actual.match_date is distinct from expected.match_date
             or actual.date_status is distinct from expected.date_status
             or actual.home_team_key is distinct from expected.home_team_key
             or actual.away_team_key is distinct from expected.away_team_key
             or actual.source_file_sha256 is distinct from expected.source_file_sha256
        ) then
          raise exception 'curated fixtures differ from the exact prepared schedule';
        end if;
      end $$;
        """
        if fixture_provenance else ""
    )
    provenance_values_sql = ",".join(
        f"""(
          {params.text(season)}, {item['source_row_number']},
          {params.text(item['upstream_match_id'])}, {params.text(item['source_locator'])},
          {params.text(item['prepared_file_sha256'])},
          {params.text(item['source_request_sha256'])}, {params.text(item['upstream_response_sha256'])},
          {params.text(item['retrieved_at'])}::timestamptz
        )"""
        for item in fixture_provenance
    )
    fixture_provenance_insert_sql = (
        f"""
      create temp table expected_fixture_provenance (
        season text not null,
        source_row_number integer not null,
        upstream_match_id text not null,
        source_locator text not null,
        prepared_file_sha256 text not null,
        source_request_sha256 text not null,
        upstream_response_sha256 text not null,
        retrieved_at timestamptz not null
      ) on commit drop;

      insert into expected_fixture_provenance values {provenance_values_sql};

      insert into curated.fixture_provenance_v1 (
        season, source_row_number, upstream_match_id, source_locator,
        prepared_file_sha256, source_request_sha256, upstream_response_sha256,
        retrieved_at, registered_by_run_id
      )
      select expected.*, (select id from current_run)
      from expected_fixture_provenance expected
      on conflict (season, source_row_number) do nothing;

      do $$
      begin
        if exists (
          select 1
          from expected_fixture_provenance expected
          left join curated.fixture_provenance_v1 actual
            using (season, source_row_number)
          where actual.upstream_match_id is distinct from expected.upstream_match_id
             or actual.source_locator is distinct from expected.source_locator
             or actual.prepared_file_sha256 is distinct from expected.prepared_file_sha256
             or actual.source_request_sha256 is distinct from expected.source_request_sha256
             or actual.upstream_response_sha256 is distinct from expected.upstream_response_sha256
             or actual.retrieved_at is distinct from expected.retrieved_at
        ) then
          raise exception 'fixture provenance conflicts with existing immutable evidence';
        end if;
        if exists (
          select 1
          from expected_fixture_provenance expected
          join curated.fixtures fixture using (season, source_row_number)
          where fixture.source_file_sha256 is distinct from expected.prepared_file_sha256
        ) then
          raise exception 'fixture provenance is not bound to the accepted curated fixture bytes';
        end if;
      end $$;
        """
        if fixture_provenance else ""
    )
    year2_fixture_migration_guard = (
        "\n".join(
            f"""
        if not exists (
          select 1 from supabase_migrations.schema_migrations
          where version = {params.text(migration.version)}
            and name = {params.text(migration.name)}
            and statements = array[{params.text(migration.statement)}]::text[]
        ) then
          raise exception 'load-curated-fixtures requires every exact checksum-bound Year 2 fixture migration';
        end if;
        """
            for migration in fixture_migrations
        )
        if fixture_provenance else ""
    )
    fixture_evidence_record = (
        {
            "role": "fixture_preparation",
            "locator": fixture_contract.evidence_locator,
            "sha256": fixture_contract.evidence_sha256,
        }
        if fixture_contract is not None else None
    )
    fixture_run_parameters: dict[str, object] = {
        "file": str(path),
        "rows": len(fixture_rows),
        "rule_version": CURATED_FIXTURE_LOAD_RULE_VERSION,
    }
    if fixture_evidence_record is not None:
        fixture_run_parameters["local_evidence_file"] = fixture_evidence_record
    sql = f"""
      do $$
      begin
        if not exists (select 1 from supabase_migrations.schema_migrations where version = '{CURATED_LAYER_MIGRATION_VERSION}') then
          raise exception 'load-curated-fixtures requires migration {CURATED_LAYER_MIGRATION_VERSION}_curated_layer';
        end if;
        if not exists (select 1 from audit.reason_codes where code = 'curated_fixture_load') then
          raise exception 'load-curated-fixtures requires reason code curated_fixture_load to be seeded by migration first';
        end if;
        {year2_fixture_migration_guard}
      end $$;

      create temp table current_run on commit drop as
      with run as (
        insert into audit.pipeline_runs (command, season, status, input_hash, parameters, ended_at, code_version, dependency_lock_hash, operator)
        values (
          'load-curated-fixtures', {params.text(season)}, 'succeeded', {params.text(file_hash)},
          {params.jsonb(fixture_run_parameters)},
          now(), {params.text(provenance['code_version'])}, {params.text(provenance['dependency_lock_hash'])}, {params.text(provenance['operator'])}
        )
        returning id
      )
      select id from run;

      create temp table current_step on commit drop as
      with step as (
        insert into audit.step_runs (pipeline_run_id, step_name, step_version, reason_code, input_count, output_count, input_hash, output_hash, ended_at)
        select id, 'curated_fixture_load', {params.text(CURATED_FIXTURE_LOAD_RULE_VERSION)}, 'curated_fixture_load',
          {len(fixture_rows)}, {len(fixture_rows)}, {params.text(file_hash)}, {params.text(file_hash)}, now()
        from current_run
        returning id
      )
      select id from step;

      insert into curated.fixtures (season, stage, round, match_date, date_status, home_team_key, away_team_key, source_row_number, source_file_sha256, loaded_by_run_id)
      values {values_sql}
      on conflict (season, source_row_number) do nothing;
      {fixture_exact_binding_sql}
      {fixture_provenance_insert_sql}
    """
    run_sql(sql, params.values)
    print(json.dumps({"status": "loaded", "season": season, "rows": len(fixture_rows), "file": str(path), "file_sha256": file_hash}, indent=2))


def dashboard_file_for_gate(args: argparse.Namespace, team_key: str, season: str) -> Path:
    """Resolve an optional dashboard candidate for read-only gates.

    Existing/re-release checks default to the committed public artifact. A
    first-release check can instead pass its Git-ignored preflight candidate.
    Relative explicit paths are repository-root-relative so the commands do
    not depend on the caller's current working directory.
    """
    dashboard_file = clean_text(getattr(args, "dashboard_file", "") or "")
    path = (
        Path(dashboard_file).expanduser()
        if dashboard_file
        else REPO_ROOT / "content" / "reporting" / f"{team_key}_dashboard_{season}.json"
    )
    if not path.is_absolute():
        path = REPO_ROOT / path
    return path.resolve()


def reconcile_curated(args: argparse.Namespace) -> None:
    """Read-only reconciliation gate: compares curated.injuries/exposure/
    team_exposure_denominators for the active build against (a)
    processing.record_versions and (b) either --dashboard-file or the default
    committed dashboard JSON. Never writes and never auto-fixes; every
    mismatch is printed with enough detail to adjudicate.
    """
    team = clean_text(args.team)
    season = clean_text(args.season)
    team_key = resolve_team_key(team)

    checks: list[dict[str, Any]] = []

    def check(name: str, passed: bool, detail: dict[str, Any]) -> None:
        checks.append({"check": name, "status": "PASS" if passed else "FAIL", **detail})

    def close_enough(a: float | None, b: float | None, tolerance: float = 0.05) -> bool:
        if a is None or b is None:
            return a is None and b is None
        return abs(a - b) <= tolerance

    build_params = SqlParams()
    builds = query_sql(
        f"""
        select id, source_version_set_hash, source_version_set_count, row_counts, created_at
        from curated.builds
        where team_key = {build_params.text(team_key)} and season = {build_params.text(season)} and status = 'active'
        """,
        build_params.values,
    )
    if not builds:
        check(
            "active_curated_build_exists",
            False,
            {"detail": f"no active curated.builds row for team_key={team_key!r} season={season!r}; run build-curated first"},
        )
        print(json.dumps({"team": team, "team_key": team_key, "season": season, "checks": checks}, indent=2))
        raise SystemExit(1)
    build = builds[0]
    check("active_curated_build_exists", True, {"build_id": build["id"], "created_at": build["created_at"]})

    injury_ids = latest_curated_source_version_ids(team, season, "%injury%")
    exposure_ids = latest_curated_source_version_ids(team, season, "%exposure%")
    live_hash = curated_source_version_set_hash(injury_ids, exposure_ids)
    check(
        "curated_build_matches_live_record_versions",
        build["source_version_set_hash"] == live_hash,
        {
            "built_hash": build["source_version_set_hash"],
            "live_hash": live_hash,
            "detail": "if FAIL, record_versions changed since this build; run build-curated --rebuild",
        },
    )

    counts_params = SqlParams()
    counts = query_sql(
        f"""
        select
          (select count(*) from curated.injuries where curated_build_id = {counts_params.text(build['id'])}::uuid) as curated_injuries,
          (select count(*) from curated.exposure where curated_build_id = {counts_params.text(build['id'])}::uuid) as curated_exposure,
          (select count(*) from curated.exposure where curated_build_id = {counts_params.text(build['id'])}::uuid and eligibility_status = 'included_pending_protocol') as curated_exposure_included,
          (select count(distinct player_uid) from curated.exposure where curated_build_id = {counts_params.text(build['id'])}::uuid and eligibility_status = 'included_pending_protocol') as curated_exposed_players,
          (select coalesce(sum(minutes_clean), 0) / 60 from curated.exposure where curated_build_id = {counts_params.text(build['id'])}::uuid and eligibility_status = 'included_pending_protocol') as curated_included_hours
        """,
        counts_params.values,
    )[0]
    check(
        "curated_injuries_count_matches_record_versions",
        int(counts["curated_injuries"]) == len(injury_ids),
        {"curated_injuries": int(counts["curated_injuries"]), "record_versions_injury_rows": len(injury_ids)},
    )
    check(
        "curated_exposure_count_matches_record_versions",
        int(counts["curated_exposure"]) == len(exposure_ids),
        {"curated_exposure": int(counts["curated_exposure"]), "record_versions_exposure_rows": len(exposure_ids)},
    )

    denom_params = SqlParams()
    denom_rows = query_sql(
        f"select matches_played, match_hours, training_hours, total_hours, fixture_source_sha256 "
        f"from curated.team_exposure_denominators where curated_build_id = {denom_params.text(build['id'])}::uuid",
        denom_params.values,
    )
    if not denom_rows:
        check("team_exposure_denominators_row_exists", False, {"detail": "no denominator row for this build"})
    else:
        denom = denom_rows[0]
        check(
            "denominator_total_hours_matches_curated_exposure",
            close_enough(float(denom["total_hours"]), float(counts["curated_included_hours"])),
            {"denominator_total_hours": float(denom["total_hours"]), "curated_included_hours": float(counts["curated_included_hours"])},
        )
        check(
            "denominator_match_hours_formula",
            close_enough(float(denom["match_hours"]), float(denom["matches_played"]) * PLAYER_HOURS_PER_TEAM_MATCH),
            {"match_hours": float(denom["match_hours"]), "matches_played": denom["matches_played"]},
        )
        check(
            "denominator_training_hours_formula",
            close_enough(float(denom["training_hours"]), float(denom["total_hours"]) - float(denom["match_hours"])),
            {"training_hours": float(denom["training_hours"])},
        )

    dashboard_path = dashboard_file_for_gate(args, team_key, season)
    if not dashboard_path.exists():
        check("dashboard_json_exists", False, {"detail": f"missing {dashboard_path}"})
    else:
        dashboard = json.loads(dashboard_path.read_text())
        coverage = dashboard.get("coverage", {})
        check(
            "curated_exposure_included_hours_matches_dashboard",
            close_enough(float(counts["curated_included_hours"]), float(coverage.get("hours", -1))),
            {"curated_hours": round(float(counts["curated_included_hours"]), 1), "dashboard_hours": coverage.get("hours")},
        )
        check(
            "curated_exposure_included_rows_matches_dashboard",
            int(counts["curated_exposure_included"]) == coverage.get("exposure_rows"),
            {"curated_exposure_included": int(counts["curated_exposure_included"]), "dashboard_exposure_rows": coverage.get("exposure_rows")},
        )
        check(
            "curated_exposed_players_matches_dashboard",
            int(counts["curated_exposed_players"]) == coverage.get("exposed_players"),
            {"curated_exposed_players": int(counts["curated_exposed_players"]), "dashboard_exposed_players": coverage.get("exposed_players")},
        )
        headline = {item["key"]: item["value"] for item in dashboard.get("headline", [])}
        included_injury_params = SqlParams()
        included_injuries = query_sql(
            f"select count(*) as n from curated.injuries where curated_build_id = {included_injury_params.text(build['id'])}::uuid and eligibility_status = 'included_pending_protocol'",
            included_injury_params.values,
        )[0]["n"]
        checks.append(
            {
                "check": "curated_included_injuries_vs_dashboard_recorded_injuries",
                "status": "INFO",
                "detail": (
                    "informational only, not a Phase 2 pass/fail gate: curated.injuries carries the DB "
                    "eligibility_status verbatim, while the dashboard JSON's recorded_injuries/time_loss_injuries "
                    "additionally apply cohort filters (received/injured-in-team, non-URC match type, non-injury "
                    "problem type, exposure-coverage-window) that live only in build_team_dashboard today. "
                    "Formalising those filters into one reusable definition is Phase 3 (analysis.injury_cohort_v1 "
                    "+ verify-analysis-parity), not this gate."
                ),
                "curated_included_pending_protocol_injuries": int(included_injuries),
                "dashboard_recorded_injuries": headline.get("recorded_injuries"),
                "dashboard_time_loss_injuries": headline.get("time_loss_injuries"),
            }
        )

    hard_failures = [c for c in checks if c["status"] == "FAIL"]
    print(
        json.dumps(
            {
                "team": team,
                "team_key": team_key,
                "season": season,
                "build_id": build["id"],
                "overall": "PASS" if not hard_failures else "FAIL",
                "checks": checks,
            },
            indent=2,
        )
    )
    if hard_failures:
        raise SystemExit(1)


def parse_date_value(value: str) -> date | None:
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def month_label(value: date) -> str:
    return value.strftime("%b %Y")


def severity_band(days: int | None, closed: bool) -> tuple[str, str]:
    if days is None or not closed:
        return "unknown_or_censored", "Unknown or censored"
    if days == 0:
        return "zero_days_medical_attention_only", "Medical attention"
    if days == 1:
        return "one_day", "1 day"
    if 2 <= days <= 3:
        return "two_to_three_days", "2-3 days"
    if 4 <= days <= 7:
        return "four_to_seven_days", "4-7 days"
    if 8 <= days <= 28:
        return "eight_to_twenty_eight_days", "8-28 days"
    return "greater_than_twenty_eight_days", ">28 days"


BODY_LOCATION_LABELS = {
    "abdomen": "Abdomen",
    "ankle": "Ankle",
    "chest": "Chest",
    "elbow": "Elbow",
    "forearm": "Forearm",
    "foot": "Foot",
    "hand": "Hand",
    "head": "Head",
    "hip_groin": "Hip/Groin",
    "knee": "Knee",
    "lower_leg": "Lower leg",
    "lumbosacral": "Lumbosacral",
    "multiple": "Multiple",
    "neck": "Neck",
    "shoulder": "Shoulder",
    "thigh": "Thigh",
    "thoracic_spine": "Thoracic spine",
    "unspecified": "Unspecified",
    "upper_arm": "Upper arm",
    "wrist": "Wrist",
    "unknown": "Unknown",
}

INJURY_TYPE_LABELS = {
    "abrasion": "Abrasion",
    "arthritis": "Arthritis",
    "avascular_necrosis": "Avascular necrosis",
    "bone_contusion": "Bone contusion",
    "bone_stress_injury": "Bone stress injury",
    "brain_spinal_cord_injury": "Brain/spinal cord injury",
    "bursitis": "Bursitis",
    "cartilage_injury": "Cartilage injury",
    "chronic_instability": "Chronic instability",
    "contusion_superficial": "Contusion (superficial)",
    "fracture": "Fracture",
    "internal_organ_trauma": "Internal organs (organ trauma)",
    "joint_sprain": "Joint sprain",
    "laceration": "Laceration",
    "muscle_compartment_syndrome": "Muscle compartment syndrome",
    "muscle_contusion": "Muscle contusion",
    "muscle_injury": "Muscle injury",
    "nonspecific": "Nonspecific",
    "peripheral_nerve_injury": "Peripheral nerve injury",
    "physis_injury": "Physis injury",
    "synovitis_capsulitis": "Synovitis/capsulitis",
    "tendon_rupture": "Tendon rupture",
    "tendinopathy": "Tendinopathy",
    "unknown": "Unknown",
    "vascular_trauma": "Vessels (vascular trauma)",
    "stump_injury": "Stump injury",
}

CONTROLLED_BODY_LOCATION_LABELS = set(BODY_LOCATION_LABELS.values())
CONTROLLED_INJURY_TYPE_LABELS = set(INJURY_TYPE_LABELS.values())
BODY_LOCATION_LABEL_TO_KEY = {label.lower(): key for key, label in BODY_LOCATION_LABELS.items()}
INJURY_TYPE_LABEL_TO_KEY = {label.lower(): key for key, label in INJURY_TYPE_LABELS.items()}

SEVERITY_LABELS = {
    "zero_days_medical_attention_only": "Medical Attention",
    "one_day": "1 day",
    "two_to_three_days": "2-3 days",
    "four_to_seven_days": "4-7 days",
    "eight_to_twenty_eight_days": "8-28 days",
    "greater_than_twenty_eight_days": ">28 days",
    "unknown_or_censored": "Unknown",
}

def format_uk_date(value: date | None) -> str:
    return value.strftime("%d/%m/%Y") if value else ""


def fit_for_selection_status(row: dict[str, str], is_closed: bool | None, is_closed_origin: str) -> tuple[str, str]:
    override = adapter_canonical_override(
        row,
        "Adapter Canonical Fit For Selection Status",
        "Adapter Canonical Fit For Selection Status Origin",
        {"fit", "not_fit", "unknown"},
    )
    if override:
        return override
    return {True: "fit", False: "not_fit"}.get(is_closed, "unknown"), is_closed_origin


def filled_injury_export_row(row: dict[str, str]) -> dict[str, str]:
    output = dict(row)
    injured_at = parse_date_value(row.get("Date Injured", ""))
    days, days_origin = effective_days_injured_with_origin(row)
    is_closed, is_closed_origin = injury_closed(row)
    activity, activity_origin = activity_context(row)
    contact, contact_origin = contact_context(row)
    recurrence, recurrence_origin = recurrence_status(row)
    severity, severity_origin = severity_category(days, is_closed)
    body, body_origin = body_location(row)
    problem, problem_origin = problem_type(row)
    injury, injury_origin = injury_type(row)
    fit_status, fit_status_origin = fit_for_selection_status(row, is_closed, is_closed_origin)
    return_date, return_date_origin = effective_confirmed_return_date(row, days, days_origin)

    output["Date Injured"] = format_uk_date(injured_at) or clean_text(row.get("Date Injured", ""))
    if days is not None and not clean_text(row.get("Days Injured")):
        output["Days Injured"] = str(days)
    output["Days Injured origin"] = days_origin
    if clean_text(row.get("Training Days Missed")):
        output["Training Days Missed"] = ""
        output["Training Days Missed origin"] = "moved_to_days_injured_source_preserved_in_intake"
    source_diagnosis = clean_text(row.get("Injury Tissue Type/s"))
    source_diagnosis_key = source_diagnosis.lower()
    if (
        source_diagnosis
        and not clean_text(row.get("Description"))
        and source_diagnosis_key not in INJURY_TYPE_MAP
        and source_diagnosis_key not in INJURY_TYPE_LABEL_TO_KEY
    ):
        output["Description"] = source_diagnosis
        output["Description origin"] = "copied_from_source_diagnosis_before_ioc_bucketing"
    source_recurrence = clean_text(row.get("Recurrence"))
    if "recurrence" in source_recurrence.lower() and not clean_text(row.get("Recurrence Stage")):
        output["Recurrence Stage"] = source_recurrence
        output["Recurrence Stage origin"] = "copied_from_source_recurrence_subtype"
    orchard_code = effective_orchard_code(row)
    if orchard_code and not clean_text(row.get("Orchard Code")):
        output["Orchard Code"] = orchard_code
    output["Problem type"] = {"injury": "Injury", "illness": "Illness"}.get(problem, "Unknown")
    output["Injury Status"] = {True: "Closed", False: "Open/Ongoing"}.get(is_closed, "Unknown")
    output["Fit for selection"] = {"fit": "Yes", "not_fit": "No"}.get(fit_status, "Unknown")
    output["Fit For Selection Date"] = ""
    output["Confirmed Return Date"] = format_uk_date(return_date) if is_closed is True else ""
    output["Occasion category"] = {"urc_match": "match", "match": "match", "training": "training"}.get(activity, "unknown")
    output["Match Type"] = {"urc_match": "URC", "training": "training"}.get(activity, "unknown")
    output["Body Part"] = BODY_LOCATION_LABELS.get(body, "Unknown")
    output["Injury Tissue Type/s"] = INJURY_TYPE_LABELS.get(injury, "Unknown")
    output["Injury Grade"] = SEVERITY_LABELS.get(severity, "Unknown")
    output["Recurrence"] = {"first_episode": "First episode", "recurrence": "Recurrence"}.get(
        recurrence, "Unknown"
    )
    output["Is Contact"] = {"contact": "Contact", "non_contact": "Non-Contact"}.get(contact, "Unknown")
    source_time_loss = clean_text(row.get("TimeLoss vs Medical Attention"))
    output["TimeLoss vs Medical Attention"] = (
        source_time_loss
        if days is None and source_time_loss in {"Time Loss", "Medical Attention"}
        else "Unknown"
        if days is None
        else "Time Loss"
        if days > 0
        else "Medical Attention"
    )
    output["Problem type origin"] = problem_origin
    output["Injury Status origin"] = is_closed_origin
    output["Fit for selection origin"] = fit_status_origin
    output["Confirmed Return Date origin"] = return_date_origin if is_closed is True and return_date else ""
    output["Occasion category origin"] = activity_origin
    output["Match Type origin"] = activity_origin
    output["Body Part origin"] = body_origin
    output["Injury Tissue Type/s origin"] = injury_origin
    output["Injury Grade origin"] = severity_origin
    output["Recurrence origin"] = recurrence_origin
    output["Is Contact origin"] = contact_origin
    output["TimeLoss vs Medical Attention origin"] = (
        "preserved_source_classification_for_censored_injury"
        if days is None and source_time_loss in {"Time Loss", "Medical Attention"}
        else "derived_from_days_injured"
    )
    return output


def rate_per_1000(count: float, hours: float) -> float | None:
    if hours <= 0:
        return None
    return count / hours * 1000


def rounded(value: float | None, digits: int = 1) -> float | None:
    if value is None:
        return None
    return round(value, digits)


def build_group_rows(
    rows: list[dict[str, Any]], field: str, hours: float, limit: int | None = None
) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"label": "Unknown", "time_loss_injuries": 0, "days_lost": 0}
    )
    for row in rows:
        label = clean_text(row.get(field)) or "Unknown"
        grouped[label]["label"] = label
        grouped[label]["time_loss_injuries"] += 1
        grouped[label]["days_lost"] += row["days_lost"]

    result = sorted(
        (
            {
                **item,
                "incidence_per_1000h": rounded(rate_per_1000(item["time_loss_injuries"], hours)),
                "burden_per_1000h": rounded(rate_per_1000(item["days_lost"], hours)),
                "mean_severity_days": rounded(
                    item["days_lost"] / item["time_loss_injuries"]
                    if item["time_loss_injuries"]
                    else None
                ),
            }
            for item in grouped.values()
        ),
        key=lambda item: (-item["time_loss_injuries"], -item["days_lost"], item["label"]),
    )
    return result[:limit] if limit else result


def exposure_analysis_date(row: dict[str, str], grain: str) -> date | None:
    if grain == "weekly":
        return parse_date_value(
            row.get("week_start_date", "")
            or row.get("cleaned_date", "")
            or row.get("session_date_clean", "")
        )
    return parse_date_value(
        row.get("session_date_clean", "")
        or row.get("cleaned_date", "")
        or row.get("week_start_date", "")
    )


# ---------------------------------------------------------------------------
# Phase 3.2 parity harness (verify-analysis-parity)
#
# Renders the analysis.*_v1 views into the exact dashboard JSON shape that
# build_team_dashboard() writes (the TeamDashboardData sections consumed by
# lib/reporting.ts), adds only the audit/curated-evidence-backed
# coverage.injury_cohort_filters block used by release, and diffs every field
# against the committed content/reporting/<team_key>_dashboard_<season>.json.
# Read-only: queries views/evidence, writes nothing to the DB, and writes its
# diff log only to the Git-ignored data/reporting/ path. Display rounding is
# applied HERE with the pipeline's own rounded() helper, matching the
# analysis_views_v1 migration contract that views return raw numerics.
# ---------------------------------------------------------------------------

DASHBOARD_PARITY_SECTIONS = [
    "headline",
    "setting_split",
    "monthly",
    "body_locations",
    "injury_types",
    "severity_distribution",
    "coverage",
]


def as_number(value: object) -> int | float | None:
    """Coerce a query_sql() value (pg numeric/bigint arrive as JSON strings,
    float8 as numbers) into a Python number; integral floats become ints so
    rendered counts serialize like build_team_dashboard()'s ints."""
    if value is None:
        return None
    if isinstance(value, bool):
        raise SystemExit(f"as_number called on boolean {value!r}")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    text = str(value).strip()
    if not text:
        return None
    number = float(text)
    return int(number) if number.is_integer() else number


def as_float(value: object) -> float | None:
    number = as_number(value)
    return None if number is None else float(number)


def render_analysis_dashboard_sections(
    *,
    headline_row: dict[str, Any],
    setting_rows: list[dict[str, Any]],
    monthly_rows: list[dict[str, Any]],
    body_location_rows: list[dict[str, Any]],
    injury_type_rows: list[dict[str, Any]],
    severity_rows: list[dict[str, Any]],
    coverage_row: dict[str, Any],
    group_limit: int = 10,
) -> dict[str, Any]:
    """Render analysis.*_v1 view rows into the dashboard JSON section shape.

    Static label/unit/formula strings are copied verbatim from
    build_team_dashboard() so full-shape parity is checked; every number
    comes from the views, with rounded() applied only here.
    """
    time_loss = as_number(headline_row["time_loss_injuries"]) or 0
    recorded = as_number(headline_row["recorded_injuries"]) or 0
    days_lost_total = as_number(headline_row["days_lost_total"]) or 0
    hours_rounded = rounded(as_float(headline_row["exposure_hours"]))
    headline = [
        {
            "key": "recorded_injuries",
            "label": "Recorded injuries in coverage window",
            "value": recorded,
            "unit": "injuries",
            "formula": "count(injury rows with Date Injured inside exposure coverage window)",
        },
        {
            "key": "time_loss_injuries",
            "label": "Time-loss injuries",
            "value": time_loss,
            "unit": "injuries",
            "formula": "count(injury rows where Days Injured > 0)",
        },
        {
            "key": "incidence_per_1000h",
            "label": "Incidence",
            "value": rounded(as_float(headline_row["incidence_per_1000h"])),
            "unit": "per 1,000 player-hours",
            "numerator": time_loss,
            "denominator": hours_rounded,
            "formula": "time-loss injuries / exposure hours * 1000",
        },
        {
            "key": "severity_mean_days",
            "label": "Mean severity",
            "value": rounded(as_float(headline_row["mean_severity_days"])),
            "unit": "days lost per injury",
            "numerator": days_lost_total,
            "denominator": time_loss,
            "formula": "days lost / time-loss injuries",
        },
        {
            "key": "severity_median_days",
            "label": "Median severity",
            "value": rounded(as_float(headline_row["median_severity_days"])),
            "unit": "days lost per injury",
            "formula": "median(Days Injured) for time-loss injuries",
        },
        {
            "key": "burden_per_1000h",
            "label": "Burden",
            "value": rounded(as_float(headline_row["burden_per_1000h"])),
            "unit": "days lost per 1,000 player-hours",
            "numerator": days_lost_total,
            "denominator": hours_rounded,
            "formula": "days lost / exposure hours * 1000",
        },
    ]

    setting_split = sorted(
        (
            {
                "label": row["label"],
                "time_loss_injuries": as_number(row["time_loss_injuries"]),
                "days_lost": as_number(row["days_lost"]),
                "mean_severity_days": rounded(as_float(row["mean_severity_days"])),
            }
            for row in setting_rows
        ),
        key=lambda item: (-item["time_loss_injuries"], -item["days_lost"], item["label"]),
    )

    monthly = [
        {
            "month": row["month_label"],
            "exposure_hours": rounded(as_float(row["exposure_hours"])),
            "distance_km": rounded(as_float(row["distance_km"])),
            "time_loss_injuries": as_number(row["time_loss_injuries"]),
            "days_lost": as_number(row["days_lost"]),
            "incidence_per_1000h": rounded(as_float(row["incidence_per_1000h"])),
            "burden_per_1000h": rounded(as_float(row["burden_per_1000h"])),
        }
        for row in sorted(monthly_rows, key=lambda item: str(item["month_start"]))
    ]

    def render_group_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        ranked = sorted(rows, key=lambda item: int(item["rank"]))
        return [
            {
                "label": row["label"],
                "time_loss_injuries": as_number(row["time_loss_injuries"]),
                "days_lost": as_number(row["days_lost"]),
                "incidence_per_1000h": rounded(as_float(row["incidence_per_1000h"])),
                "burden_per_1000h": rounded(as_float(row["burden_per_1000h"])),
                "mean_severity_days": rounded(as_float(row["mean_severity_days"])),
            }
            for row in ranked[:group_limit]
        ]

    severity_distribution = [
        {
            "key": row["key"],
            "label": row["label"],
            "recorded_injuries": as_number(row["recorded_injuries"]),
            "time_loss_injuries": as_number(row["time_loss_injuries"]),
            "days_lost": as_number(row["days_lost"]),
        }
        for row in sorted(severity_rows, key=lambda item: int(item["band_order"]))
    ]

    coverage = {
        "exposure_rows": as_number(coverage_row["exposure_rows"]),
        "exposed_players": as_number(coverage_row["exposed_players"]),
        "weeks": as_number(coverage_row["weeks"]),
        "exposure_periods": as_number(coverage_row["exposure_periods"]),
        "exposure_grain": coverage_row["exposure_grain"],
        "hours": rounded(as_float(coverage_row["hours"])),
        "distance_km": rounded(as_float(coverage_row["distance_km"])),
        "included_exposure_status": coverage_row["included_exposure_status"],
        "scope_status_counts": {
            key: as_number(value)
            for key, value in (coverage_row["scope_status_counts"] or {}).items()
        },
    }

    return {
        "headline": headline,
        "setting_split": setting_split,
        "monthly": monthly,
        "body_locations": render_group_rows(body_location_rows),
        "injury_types": render_group_rows(injury_type_rows),
        "severity_distribution": severity_distribution,
        "coverage": coverage,
    }


def parity_values_equal(committed: object, rendered: object) -> bool:
    if committed is None or rendered is None:
        return committed is None and rendered is None
    committed_numeric = isinstance(committed, (int, float)) and not isinstance(committed, bool)
    rendered_numeric = isinstance(rendered, (int, float)) and not isinstance(rendered, bool)
    if committed_numeric and rendered_numeric:
        return abs(float(committed) - float(rendered)) <= 1e-9
    return committed == rendered


def v6_public_scalars_equal(reviewed: object, candidate: object) -> bool:
    """Compare V6 scalar values exactly, allowing numeric scale alone."""
    if isinstance(reviewed, bool) or isinstance(candidate, bool):
        return (
            isinstance(reviewed, bool)
            and isinstance(candidate, bool)
            and reviewed == candidate
        )
    reviewed_numeric = isinstance(reviewed, (int, float, Decimal))
    candidate_numeric = isinstance(candidate, (int, float, Decimal))
    if reviewed_numeric or candidate_numeric:
        return (
            reviewed_numeric
            and candidate_numeric
            and Decimal(str(reviewed)) == Decimal(str(candidate))
        )
    return reviewed == candidate


def diff_dashboard_sections(committed: dict[str, Any], rendered: dict[str, Any]) -> list[dict[str, Any]]:
    """Field-by-field diff of the seven dashboard parity sections. Returns
    one result row per compared field: status PASS or DIFF, with a `kind`
    on every DIFF (value_mismatch / missing_in_rendered / extra_in_rendered
    / row_count / type_mismatch). Values here are approved aggregates and
    IOC bucket labels only; never player-level or protected values.
    """
    results: list[dict[str, Any]] = []

    def record(section: str, path: str, committed_value: object, rendered_value: object, status: str, kind: str | None = None) -> None:
        row: dict[str, Any] = {
            "section": section,
            "path": path,
            "status": status,
            "committed": committed_value,
            "rendered": rendered_value,
        }
        if kind:
            row["kind"] = kind
        results.append(row)

    def walk(section: str, path: str, committed_value: object, rendered_value: object) -> None:
        if isinstance(committed_value, dict) and isinstance(rendered_value, dict):
            for key in sorted(set(committed_value) | set(rendered_value)):
                child_path = f"{path}.{key}"
                if key not in rendered_value:
                    record(section, child_path, committed_value[key], None, "DIFF", "missing_in_rendered")
                elif key not in committed_value:
                    record(section, child_path, None, rendered_value[key], "DIFF", "extra_in_rendered")
                else:
                    walk(section, child_path, committed_value[key], rendered_value[key])
            return
        if isinstance(committed_value, list) and isinstance(rendered_value, list):
            if len(committed_value) != len(rendered_value):
                record(section, f"{path}.length", len(committed_value), len(rendered_value), "DIFF", "row_count")
            for index in range(max(len(committed_value), len(rendered_value))):
                child_path = f"{path}[{index}]"
                if index >= len(rendered_value):
                    record(section, child_path, committed_value[index], None, "DIFF", "missing_in_rendered")
                elif index >= len(committed_value):
                    record(section, child_path, None, rendered_value[index], "DIFF", "extra_in_rendered")
                else:
                    walk(section, child_path, committed_value[index], rendered_value[index])
            return
        if isinstance(committed_value, (dict, list)) != isinstance(rendered_value, (dict, list)):
            record(section, path, committed_value, rendered_value, "DIFF", "type_mismatch")
            return
        if parity_values_equal(committed_value, rendered_value):
            record(section, path, committed_value, rendered_value, "PASS")
        else:
            record(section, path, committed_value, rendered_value, "DIFF", "value_mismatch")

    for section in DASHBOARD_PARITY_SECTIONS:
        walk(section, section, committed.get(section), rendered.get(section))
    return results


def compare_complete_public_payloads(
    reviewed: object,
    candidate: object,
    *,
    reviewed_display: object | None = None,
    candidate_display: object | None = None,
) -> tuple[int, list[dict[str, Any]]]:
    """Compare every public JSON leaf without inventing a projection.

    V6 review is over the immutable candidate itself, so a legacy subset
    renderer would weaken the gate. Both inputs have already passed the strict
    V6 public contract before this helper is called; mismatch values are
    therefore aggregate/public evidence only.
    """
    compared = 0
    diffs: list[dict[str, Any]] = []
    reviewed_display = reviewed if reviewed_display is None else reviewed_display
    candidate_display = candidate if candidate_display is None else candidate_display

    def record(path: str, reviewed_value: object, candidate_value: object, kind: str) -> None:
        diffs.append({
            "path": path,
            "kind": kind,
            "reviewed": reviewed_value,
            "candidate": candidate_value,
        })

    def walk(
        path: str,
        reviewed_value: object,
        candidate_value: object,
        reviewed_display_value: object,
        candidate_display_value: object,
    ) -> None:
        nonlocal compared
        if isinstance(reviewed_value, dict) and isinstance(candidate_value, dict):
            for key in sorted(set(reviewed_value) | set(candidate_value)):
                child = f"{path}.{key}" if path else key
                if key not in candidate_value:
                    display = (
                        reviewed_display_value.get(key)
                        if isinstance(reviewed_display_value, dict)
                        else reviewed_value[key]
                    )
                    record(child, display, None, "missing_in_candidate")
                elif key not in reviewed_value:
                    display = (
                        candidate_display_value.get(key)
                        if isinstance(candidate_display_value, dict)
                        else candidate_value[key]
                    )
                    record(child, None, display, "extra_in_candidate")
                else:
                    walk(
                        child,
                        reviewed_value[key],
                        candidate_value[key],
                        reviewed_display_value[key],
                        candidate_display_value[key],
                    )
            return
        if isinstance(reviewed_value, list) and isinstance(candidate_value, list):
            if len(reviewed_value) != len(candidate_value):
                record(
                    f"{path}.length", len(reviewed_value), len(candidate_value),
                    "row_count",
                )
            for index in range(max(len(reviewed_value), len(candidate_value))):
                child = f"{path}[{index}]"
                if index >= len(candidate_value):
                    record(
                        child,
                        reviewed_display_value[index],
                        None,
                        "missing_in_candidate",
                    )
                elif index >= len(reviewed_value):
                    record(
                        child,
                        None,
                        candidate_display_value[index],
                        "extra_in_candidate",
                    )
                else:
                    walk(
                        child,
                        reviewed_value[index],
                        candidate_value[index],
                        reviewed_display_value[index],
                        candidate_display_value[index],
                    )
            return
        compared += 1
        if isinstance(reviewed_value, (dict, list)) != isinstance(candidate_value, (dict, list)):
            record(
                path,
                reviewed_display_value,
                candidate_display_value,
                "type_mismatch",
            )
        elif not v6_public_scalars_equal(reviewed_value, candidate_value):
            record(
                path,
                reviewed_display_value,
                candidate_display_value,
                "value_mismatch",
            )

    walk("", reviewed, candidate, reviewed_display, candidate_display)
    return compared, diffs


def verify_analysis_parity_v6(args: argparse.Namespace) -> None:
    """Compare a reviewed Year 2 preflight with its exact V6 DB candidate.

    The reviewed preflight manifest binds its exact bytes to the database
    candidate's canonical hash and release identity. Every public field is
    also compared, while the legacy ``analysis.*_v1`` path stays entirely out
    of Year 2 verification.
    """
    team = clean_text(args.team)
    team_key = resolve_team_key(team)
    contract = release_contract_for("2025-26", YEAR2_2025_26_RELEASE_TUPLE)
    if not contract.team_candidate_view:
        raise SystemExit("V6 analysis parity contract lacks a team candidate view")
    provenance = run_provenance()

    reviewed_path = dashboard_file_for_gate(args, team_key, "2025-26")
    assert_checksum_bound_release_migrations(contract, "V6 analysis parity")
    params = SqlParams()
    rows = query_sql(
        f"""
        select candidate.team_key, candidate.season,
          candidate.curated_build_id::text, candidate.analysis_version,
          candidate.classification_view_version,
          candidate.classification_evidence_sha256,
          candidate.cohort_view_version, candidate.cohort_evidence_sha256,
          candidate.dashboard::text as dashboard_json,
          reporting.canonical_jsonb_sha256_v1(candidate.dashboard) as payload_sha256,
          coalesce((
            select jsonb_agg(
              jsonb_build_object(
                'release_id', payload.release_id::text,
                'release_label', release.release_label,
                'payload_sha256', payload.payload_sha256
              )
              order by release.approved_at desc nulls last,
                release.created_at desc, payload.release_id desc
            )
            from reporting.team_release_payloads_v6 payload
            join reporting.aggregate_releases release on release.id = payload.release_id
            where payload.team_key = candidate.team_key
              and payload.season = candidate.season
              and release.status = 'approved'
          ), '[]'::jsonb) as approved_predecessors
        from {contract.team_candidate_view} candidate
        where candidate.team_key = {params.text(team_key)}
          and candidate.season = '2025-26'
          and candidate.analysis_version = 'v6'
          and candidate.classification_view_version = 'reporting_classification_2026-07-22_v2'
          and candidate.cohort_view_version = 'analysis_window_2025-26_2026-08-15_v1'
        """,
        params.values,
    )
    if len(rows) != 1 or not isinstance(rows[0].get("dashboard_json"), str):
        raise SystemExit(
            "V6 analysis parity requires exactly one complete active-build candidate"
        )
    candidate = rows[0]
    candidate_dashboard_json = candidate["dashboard_json"]
    try:
        candidate_dashboard = json.loads(candidate_dashboard_json)
        candidate_dashboard_exact = json.loads(
            candidate_dashboard_json,
            parse_float=Decimal,
            parse_int=Decimal,
        )
    except json.JSONDecodeError as error:
        raise SystemExit(
            "V6 analysis parity candidate dashboard is not valid PostgreSQL JSON text"
        ) from error
    if not isinstance(candidate_dashboard, dict) or not isinstance(candidate_dashboard_exact, dict):
        raise SystemExit("V6 analysis parity candidate dashboard must be a JSON object")
    assert_v6_public_dashboard_contract(candidate_dashboard, "candidate team dashboard")

    exact_fields = {
        "season": contract.season,
        "analysis_version": contract.analysis_version,
        "classification_view_version": contract.classification_view_version,
        "cohort_view_version": contract.cohort_view_version,
    }
    for field, expected in exact_fields.items():
        if clean_text(candidate.get(field)) != expected:
            raise SystemExit(f"V6 analysis parity candidate has invalid {field}")

    database_evidence_hashes = {
        field: candidate.get(field)
        for field in ("classification_evidence_sha256", "cohort_evidence_sha256")
    }
    for field, value in database_evidence_hashes.items():
        if not isinstance(value, str) or not re.fullmatch(r"[0-9a-f]{64}", value):
            raise SystemExit(f"V6 analysis parity candidate has invalid {field}")
    classification_evidence_sha256 = database_evidence_hashes[
        "classification_evidence_sha256"
    ]
    cohort_evidence_sha256 = database_evidence_hashes["cohort_evidence_sha256"]

    candidate_hash = clean_text(candidate.get("payload_sha256"))
    if not re.fullmatch(r"[0-9a-f]{64}", candidate_hash):
        raise SystemExit("V6 analysis parity lacks its canonical database payload hash")

    predecessors = candidate.get("approved_predecessors")
    if not isinstance(predecessors, list) or any(
        not isinstance(predecessor, dict) for predecessor in predecessors
    ):
        raise SystemExit("V6 analysis parity candidate has an invalid approved predecessor set")
    if len(predecessors) > 1:
        raise SystemExit("V6 analysis parity requires at most one approved predecessor")
    predecessor = predecessors[0] if predecessors else None
    reviewed, reviewed_exact, reviewed_sha256, reviewed_manifest_sha256 = read_v6_team_reviewed_preflight(
        reviewed_path=reviewed_path,
        team_key=team_key,
        contract=contract,
        candidate=candidate,
        predecessor=predecessor,
        provenance=provenance,
    )
    assert_v6_public_dashboard_contract(reviewed, "reviewed team dashboard")

    fields_compared, diffs = compare_complete_public_payloads(
        reviewed_exact,
        candidate_dashboard_exact,
        reviewed_display=reviewed,
        candidate_display=candidate_dashboard,
    )

    log = {
        "schema_version": "urc_v6_analysis_parity_v1",
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "command": "verify-analysis-parity",
        "team": team,
        "team_key": team_key,
        "season": "2025-26",
        "reviewed_dashboard": (
            str(reviewed_path.relative_to(REPO_ROOT))
            if reviewed_path.is_relative_to(REPO_ROOT)
            else str(reviewed_path)
        ),
        "reviewed_file_sha256": reviewed_sha256,
        "reviewed_preflight_manifest_sha256": reviewed_manifest_sha256,
        "candidate_view": contract.team_candidate_view,
        "canonical_payload_sha256": candidate_hash,
        "classification_evidence_sha256": classification_evidence_sha256,
        "cohort_evidence_sha256": cohort_evidence_sha256,
        "required_migrations": [
            {"version": item.version, "name": item.name, "sha256": item.sha256}
            for item in contract.required_migration_contracts
        ],
        "summary": {
            "overall": "PARITY" if not diffs else "DIFFS",
            "fields_compared": fields_compared,
            "diff": len(diffs),
        },
        "diffs": diffs,
    }
    log_path = REPO_ROOT / "data" / "reporting" / f"{team_key}_analysis_parity_2025-26.json"
    write_json_atomic(log_path, log)
    print(json.dumps({
        "team": team,
        "team_key": team_key,
        "season": "2025-26",
        "overall": log["summary"]["overall"],
        "fields_compared": fields_compared,
        "diff": len(diffs),
        "canonical_payload_sha256": candidate_hash,
        "classification_evidence_sha256": classification_evidence_sha256,
        "cohort_evidence_sha256": cohort_evidence_sha256,
        "diff_log": str(log_path.relative_to(REPO_ROOT)),
    }, indent=2))
    if diffs:
        raise SystemExit(1)


def verify_analysis_parity(args: argparse.Namespace) -> None:
    """Read-only candidate parity gate selected by the exact season contract.

    Year 2 compares the complete reviewed V6 candidate and PostgreSQL canonical
    hashes. Historical seasons retain the original ``analysis.*_v1`` renderer,
    including its release-time cohort-filter evidence. Both routes write only
    aggregate/public diff evidence under Git-ignored ``data/reporting/``.
    """
    team = clean_text(args.team)
    season = clean_text(args.season)
    if season == "2025-26":
        verify_analysis_parity_v6(args)
        return
    team_key = resolve_team_key(team)

    migration = query_sql(
        f"select 1 as ok from supabase_migrations.schema_migrations where version = '{ANALYSIS_VIEWS_MIGRATION_VERSION}'"
    )
    if not migration:
        raise SystemExit(
            f"verify-analysis-parity requires migration {ANALYSIS_VIEWS_MIGRATION_VERSION}_analysis_views_v1 "
            "to be applied to the live target first"
        )

    def team_rows(sql_template: str) -> list[dict[str, Any]]:
        params = SqlParams()
        return query_sql(
            sql_template.format(team_key=params.text(team_key), season=params.text(season)),
            params.values,
        )

    headline_rows = team_rows(
        "select * from analysis.headline_metrics_v1 where team_key = {team_key} and season = {season}"
    )
    if len(headline_rows) != 1:
        raise SystemExit(
            f"expected exactly one analysis.headline_metrics_v1 row for team_key={team_key!r} "
            f"season={season!r}, found {len(headline_rows)}; is there an active curated build?"
        )
    coverage_rows = team_rows(
        "select * from analysis.coverage_v1 where team_key = {team_key} and season = {season}"
    )
    if len(coverage_rows) != 1:
        raise SystemExit(
            f"expected exactly one analysis.coverage_v1 row for team_key={team_key!r} season={season!r}, "
            f"found {len(coverage_rows)}"
        )
    setting_rows = team_rows(
        "select * from analysis.setting_split_v1 where team_key = {team_key} and season = {season} "
        "order by time_loss_injuries desc, days_lost desc, label asc"
    )
    monthly_rows = team_rows(
        "select * from analysis.monthly_v1 where team_key = {team_key} and season = {season} order by month_start"
    )
    body_location_rows = team_rows(
        "select * from analysis.body_locations_v1 where team_key = {team_key} and season = {season} order by rank"
    )
    injury_type_rows = team_rows(
        "select * from analysis.injury_types_v1 where team_key = {team_key} and season = {season} order by rank"
    )
    severity_rows = team_rows(
        "select * from analysis.severity_distribution_v1 where team_key = {team_key} and season = {season} "
        "order by band_order"
    )

    rendered = render_analysis_dashboard_sections(
        headline_row=headline_rows[0],
        setting_rows=setting_rows,
        monthly_rows=monthly_rows,
        body_location_rows=body_location_rows,
        injury_type_rows=injury_type_rows,
        severity_rows=severity_rows,
        coverage_row=coverage_rows[0],
    )
    rendered["coverage"]["injury_cohort_filters"] = release_cohort_filter_flags(team_key, season)

    committed_path = dashboard_file_for_gate(args, team_key, season)
    if not committed_path.exists():
        raise SystemExit(f"no dashboard JSON at {committed_path}")
    committed = json.loads(committed_path.read_text())

    results = diff_dashboard_sections(committed, rendered)
    diffs = [row for row in results if row["status"] == "DIFF"]
    section_diff_counts: dict[str, int] = defaultdict(int)
    for row in diffs:
        section_diff_counts[row["section"]] += 1

    log = {
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "command": "verify-analysis-parity",
        "team": team,
        "team_key": team_key,
        "season": season,
        "committed_dashboard": (
            str(committed_path.relative_to(REPO_ROOT))
            if committed_path.is_relative_to(REPO_ROOT)
            else str(committed_path)
        ),
        "committed_dashboard_sha256": sha256_file(committed_path),
        "analysis_views_migration": ANALYSIS_VIEWS_MIGRATION_VERSION,
        "analysis_view_version": ANALYSIS_VIEW_VERSION_SUFFIX,
        "summary": {
            "overall": "PARITY" if not diffs else "DIFFS",
            "fields_compared": len(results),
            "pass": len(results) - len(diffs),
            "diff": len(diffs),
            "diffs_by_section": dict(sorted(section_diff_counts.items())),
        },
        "diffs": diffs,
        "results": results,
    }
    log_path = REPO_ROOT / "data" / "reporting" / f"{team_key}_analysis_parity_{season}.json"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_path.write_text(json.dumps(log, indent=2) + "\n")

    print(
        json.dumps(
            {
                "team": team,
                "team_key": team_key,
                "season": season,
                "overall": log["summary"]["overall"],
                "fields_compared": log["summary"]["fields_compared"],
                "pass": log["summary"]["pass"],
                "diff": log["summary"]["diff"],
                "diffs_by_section": log["summary"]["diffs_by_section"],
                "diff_log": str(log_path.relative_to(REPO_ROOT)),
            },
            indent=2,
        )
    )
    if diffs:
        raise SystemExit(1)


def build_team_dashboard(args: argparse.Namespace) -> None:
    excluded_injury_rows = {
        clean_text(item)
        for item in clean_text(getattr(args, "exclude_injury_rows", "")).split(",")
        if clean_text(item)
    }
    injured_in_team = clean_text(getattr(args, "injured_in_team", ""))
    exposure_rows = [
        row
        for row in read_rows(Path(args.exposure_file))
        if clean_text(row.get("cleaning_action")) == "include"
    ]
    if not exposure_rows:
        raise SystemExit("no included exposure rows found")
    scope_status_counts: dict[str, int] = {}
    for row in exposure_rows:
        scope_status = clean_text(row.get("scope_status")) or "unknown"
        scope_status_counts[scope_status] = scope_status_counts.get(scope_status, 0) + 1

    grains = {
        clean_text(row.get("exposure_grain")) or "unknown"
        for row in exposure_rows
    }
    primary_grain = grains.pop() if len(grains) == 1 else "mixed"
    exposure_dates = [
        parsed
        for row in exposure_rows
        if (parsed := exposure_analysis_date(row, primary_grain)) is not None
    ]
    if not exposure_dates:
        raise SystemExit("no valid exposure date values found")

    coverage_start = min(exposure_dates)
    coverage_end = max(exposure_dates) + (timedelta(days=6) if primary_grain == "weekly" else timedelta())
    hours = sum(parse_float(row.get("minutes_total_clean")) or 0 for row in exposure_rows) / 60
    distance_m = sum(parse_float(row.get("distance_total_m_clean")) or 0 for row in exposure_rows)
    players = {row.get("player_uid") for row in exposure_rows if clean_text(row.get("player_uid"))}
    exposure_periods = sorted({date.isoformat() for date in exposure_dates})

    exposure_by_month: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"month": "", "exposure_hours": 0.0, "distance_km": 0.0}
    )
    for row in exposure_rows:
        exposure_date = exposure_analysis_date(row, primary_grain)
        if exposure_date is None:
            continue
        key = exposure_date.strftime("%Y-%m")
        exposure_by_month[key]["month"] = month_label(exposure_date)
        exposure_by_month[key]["exposure_hours"] += (parse_float(row.get("minutes_total_clean")) or 0) / 60
        exposure_by_month[key]["distance_km"] += (parse_float(row.get("distance_total_m_clean")) or 0) / 1000

    injury_source_rows = read_rows(Path(args.injury_file))
    analysis_source_rows = []
    analysis_excluded_row_numbers: set[str] = set()
    analysis_filled_columns: dict[str, set[str]] = {}
    standardisation_events: list[dict[str, str]] = []
    standardisation_errors: list[str] = []
    injury_scope_exclusion_counts: dict[str, int] = defaultdict(int)
    injury_rows = []
    for index, row in enumerate(injury_source_rows, start=2):
        source_row_number = str(index)
        injured_at = parse_date_value(row.get("Date Injured", ""))
        filled_row = filled_injury_export_row(row)
        protected_alias_replaced_fields = set()
        if injured_in_team:
            for field, value in filled_row.items():
                if clean_text(value).casefold() == injured_in_team.casefold():
                    filled_row[field] = args.team
                    protected_alias_replaced_fields.add(field)
        source_tissue = clean_text(row.get("Injury Tissue Type/s"))
        source_contact = clean_text(row.get("Is Contact")).lower()
        source_recurrence = clean_text(row.get("Recurrence")).lower()
        source_occasion = clean_text(row.get("Occasion category")).lower()
        if (
            clean_text(row.get("Problem type")).lower() == "injury"
            and source_tissue
            and not is_missing(source_tissue)
            and filled_row.get("Injury Tissue Type/s") == "Unknown"
        ):
            standardisation_errors.append(f"row {source_row_number}: explicit injury diagnosis mapped to Unknown")
        if (
            source_contact not in {"", "na", "n/a", "unknown", "other - non-rugby"}
            and filled_row.get("Is Contact") == "Unknown"
        ):
            standardisation_errors.append(f"row {source_row_number}: explicit contact value mapped to Unknown")
        if (
            source_recurrence not in {"", "na", "n/a", "unknown"}
            and not any(marker in source_recurrence for marker in {"ã", "â"})
            and filled_row.get("Recurrence") == "Unknown"
        ):
            standardisation_errors.append(f"row {source_row_number}: explicit recurrence value mapped to Unknown")
        if source_occasion in {"match", "game", "training"} and filled_row.get("Occasion category") == "unknown":
            standardisation_errors.append(f"row {source_row_number}: explicit occasion mapped to unknown")
        analysis_source_rows.append(filled_row)
        analysis_filled_columns[source_row_number] = {
            field
            for field, value in filled_row.items()
            if clean_text(value) and not clean_text(row.get(field, ""))
        }
        origin_fields = {
            "Days Injured": "Days Injured origin",
            "Training Days Missed": "Training Days Missed origin",
            "Description": "Description origin",
            "Problem type": "Problem type origin",
            "Injury Status": "Injury Status origin",
            "Fit for selection": "Fit for selection origin",
            "Confirmed Return Date": "Confirmed Return Date origin",
            "Occasion category": "Occasion category origin",
            "Match Type": "Match Type origin",
            "Body Part": "Body Part origin",
            "Injury Tissue Type/s": "Injury Tissue Type/s origin",
            "Injury Grade": "Injury Grade origin",
            "Recurrence": "Recurrence origin",
            "Recurrence Stage": "Recurrence Stage origin",
            "Is Contact": "Is Contact origin",
            "TimeLoss vs Medical Attention": "TimeLoss vs Medical Attention origin",
        }
        for field, origin_field in origin_fields.items():
            old_value = clean_text(row.get(field, ""))
            new_value = clean_text(filled_row.get(field, ""))
            if old_value == new_value:
                continue
            standardisation_events.append(
                {
                    "standardised_row_number": source_row_number,
                    "field": field,
                    "old_value": old_value,
                    "new_value": new_value,
                    "action": "fill" if not old_value else "standardise",
                    "reason": (
                        "protected_team_alias_replaced_in_team_scoped_export"
                        if field in protected_alias_replaced_fields
                        else clean_text(filled_row.get(origin_field, ""))
                    ),
                    "review_status": "pipeline_decision",
                }
            )
        cohort_exclusion_reasons = injury_cohort_exclusion_reasons(row, injured_in_team)
        if source_row_number in excluded_injury_rows:
            cohort_exclusion_reasons.append("adjudicated_duplicate")
        if injured_at is None or injured_at < coverage_start or injured_at > coverage_end:
            cohort_exclusion_reasons.append("injury_date_missing_or_outside_exposure_coverage")
        if filled_row.get("Problem type") != "Injury":
            cohort_exclusion_reasons.append("non_injury_problem_type")
        for reason in cohort_exclusion_reasons:
            injury_scope_exclusion_counts[reason] += 1
            standardisation_events.append(
                {
                    "standardised_row_number": source_row_number,
                    "field": "analysis_eligibility",
                    "old_value": "",
                    "new_value": "excluded",
                    "action": "exclude",
                    "reason": reason,
                    "review_status": "user_adjudicated" if reason == "adjudicated_duplicate" else "pipeline_decision",
                }
            )
        if (
            cohort_exclusion_reasons
        ):
            analysis_excluded_row_numbers.add(source_row_number)
            continue
        days_lost = effective_days_injured(filled_row)
        closed, _ = injury_closed(filled_row)
        band_key, band_label = severity_band(days_lost, closed is True)
        injury_rows.append(
            {
                **row,
                "Occasion category": filled_row["Occasion category"],
                "Match Type": filled_row["Match Type"],
                "Body Part": filled_row["Body Part"],
                "Injury Tissue Type/s": filled_row["Injury Tissue Type/s"],
                "injured_at": injured_at,
                "days_lost": days_lost or 0,
                "is_time_loss": (days_lost or 0) > 0,
                "severity_band_key": band_key,
                "severity_band_label": band_label,
            }
        )

    if standardisation_errors:
        raise SystemExit("standardisation coverage failed: " + "; ".join(standardisation_errors[:20]))

    time_loss_rows = [row for row in injury_rows if row["is_time_loss"]]
    days_lost_total = sum(row["days_lost"] for row in time_loss_rows)
    mean_severity = days_lost_total / len(time_loss_rows) if time_loss_rows else None
    sorted_days = sorted(row["days_lost"] for row in time_loss_rows)
    median_severity = None
    if sorted_days:
        midpoint = len(sorted_days) // 2
        median_severity = (
            sorted_days[midpoint]
            if len(sorted_days) % 2
            else (sorted_days[midpoint - 1] + sorted_days[midpoint]) / 2
        )

    injuries_by_month: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"time_loss_injuries": 0, "days_lost": 0}
    )
    for row in time_loss_rows:
        key = row["injured_at"].strftime("%Y-%m")
        injuries_by_month[key]["time_loss_injuries"] += 1
        injuries_by_month[key]["days_lost"] += row["days_lost"]

    monthly = []
    for key in sorted(set(exposure_by_month) | set(injuries_by_month)):
        exposure = exposure_by_month[key]
        injuries = injuries_by_month[key]
        month_hours = exposure["exposure_hours"]
        monthly.append(
            {
                "month": exposure.get("month") or month_label(datetime.strptime(key, "%Y-%m").date()),
                "exposure_hours": rounded(month_hours),
                "distance_km": rounded(exposure["distance_km"]),
                "time_loss_injuries": injuries["time_loss_injuries"],
                "days_lost": injuries["days_lost"],
                "incidence_per_1000h": rounded(
                    rate_per_1000(injuries["time_loss_injuries"], month_hours)
                ),
                "burden_per_1000h": rounded(rate_per_1000(injuries["days_lost"], month_hours)),
            }
        )

    severity_groups: dict[str, dict[str, Any]] = {}
    for row in injury_rows:
        key = row["severity_band_key"]
        group = severity_groups.setdefault(
            key,
            {
                "key": key,
                "label": row["severity_band_label"],
                "recorded_injuries": 0,
                "time_loss_injuries": 0,
                "days_lost": 0,
            },
        )
        group["recorded_injuries"] += 1
        if row["is_time_loss"]:
            group["time_loss_injuries"] += 1
            group["days_lost"] += row["days_lost"]

    setting_split = [
        {
            "label": item["label"],
            "time_loss_injuries": item["time_loss_injuries"],
            "days_lost": item["days_lost"],
            "mean_severity_days": item["mean_severity_days"],
        }
        for item in build_group_rows(time_loss_rows, "Occasion category", hours)
    ]

    grain_label = {
        "weekly": "weekly",
        "session": "session-level",
        "mixed": "mixed-grain",
    }.get(primary_grain, "unknown-grain")
    monthly_basis = (
        f"Monthly exposure is assigned to the week-start month because {args.team} reports weekly exposure."
        if primary_grain == "weekly"
        else "Monthly exposure is assigned to the cleaned exposure-date month."
    )
    setting_limitation = (
        f"{args.team} exposure is weekly, so match and training incidence cannot be split until setting-specific exposure denominators are approved."
        if primary_grain == "weekly"
        else "Setting-specific rates are not split until setting-specific exposure denominators are approved."
    )

    dashboard = {
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "team": args.team,
        "season": args.season,
        "analysis_window": {
            "start": coverage_start.isoformat(),
            "end": coverage_end.isoformat(),
            "basis": f"{args.team} exposure coverage window from included {grain_label} exposure rows",
        },
        "source_files": {
            "injury": str(args.injury_file),
            "exposure": str(args.exposure_file),
        },
        "pipeline_evidence": {
            "injury_file_sha256": sha256_file(Path(args.injury_file)),
            "exposure_file_sha256": sha256_file(Path(args.exposure_file)),
            "injury_processing_rule_version": INJURY_PROCESSING_RULE_VERSION,
            "exposure_processing_rule_version": EXPOSURE_PROCESSING_RULE_VERSION,
            "adjudicated_duplicate_rows": sorted(int(row) for row in excluded_injury_rows),
            "standardisation_audit_file": None,
            "standardisation_audit_sha256": None,
        },
        "method": [
            "Headline injury metrics use time-loss injuries only: Days Injured > 0.",
            "Incidence = time-loss injuries / exposure hours * 1000.",
            "Severity = mean and median days lost per time-loss injury.",
            "Burden = days lost / exposure hours * 1000.",
            f"Exposure hours = sum(minutes_total_clean) / 60 for included {grain_label} exposure rows.",
            monthly_basis,
            "IOC-aligned body-location, tissue/pathology, and severity labels come from the accepted V2 mapping.",
            "Received/Injured In Team retains the approved team plus blank/N/A values; explicit other-team or Club values are excluded."
            if injured_in_team
            else "No Received/Injured In Team cohort filter applied.",
            "Match Type retains URC, training, Other, blank/N/A, and generic match/game values; explicit non-URC competitions and teams are excluded.",
            f"Adjudicated duplicate standardised rows excluded from this aggregate: {', '.join(sorted(excluded_injury_rows))}."
            if excluded_injury_rows
            else "No adjudicated duplicate injury rows were excluded from this aggregate.",
        ],
        "coverage": {
            "exposure_rows": len(exposure_rows),
            "exposed_players": len(players),
            "weeks": len(exposure_periods) if primary_grain == "weekly" else 0,
            "exposure_periods": len(exposure_periods),
            "exposure_grain": primary_grain,
            "hours": rounded(hours),
            "distance_km": rounded(distance_m / 1000),
            "included_exposure_status": "included",
            "scope_status_counts": scope_status_counts,
            "injury_cohort_filters": {
                "injured_in_team_applied": bool(injured_in_team),
                "explicit_non_urc_match_type_applied": True,
                "exclusion_reason_counts": dict(sorted(injury_scope_exclusion_counts.items())),
            },
        },
        "headline": [
            {
                "key": "recorded_injuries",
                "label": "Recorded injuries in coverage window",
                "value": len(injury_rows),
                "unit": "injuries",
                "formula": "count(injury rows with Date Injured inside exposure coverage window)",
            },
            {
                "key": "time_loss_injuries",
                "label": "Time-loss injuries",
                "value": len(time_loss_rows),
                "unit": "injuries",
                "formula": "count(injury rows where Days Injured > 0)",
            },
            {
                "key": "incidence_per_1000h",
                "label": "Incidence",
                "value": rounded(rate_per_1000(len(time_loss_rows), hours)),
                "unit": "per 1,000 player-hours",
                "numerator": len(time_loss_rows),
                "denominator": rounded(hours),
                "formula": "time-loss injuries / exposure hours * 1000",
            },
            {
                "key": "severity_mean_days",
                "label": "Mean severity",
                "value": rounded(mean_severity),
                "unit": "days lost per injury",
                "numerator": days_lost_total,
                "denominator": len(time_loss_rows),
                "formula": "days lost / time-loss injuries",
            },
            {
                "key": "severity_median_days",
                "label": "Median severity",
                "value": rounded(median_severity),
                "unit": "days lost per injury",
                "formula": "median(Days Injured) for time-loss injuries",
            },
            {
                "key": "burden_per_1000h",
                "label": "Burden",
                "value": rounded(rate_per_1000(days_lost_total, hours)),
                "unit": "days lost per 1,000 player-hours",
                "numerator": days_lost_total,
                "denominator": rounded(hours),
                "formula": "days lost / exposure hours * 1000",
            },
        ],
        "setting_split": setting_split,
        "monthly": monthly,
        "body_locations": build_group_rows(time_loss_rows, "Body Part", hours, limit=10),
        "injury_types": build_group_rows(time_loss_rows, "Injury Tissue Type/s", hours, limit=10),
        "severity_distribution": sorted(
            severity_groups.values(),
            key=lambda item: [
                "zero_days_medical_attention_only",
                "one_day",
                "two_to_three_days",
                "four_to_seven_days",
                "eight_to_twenty_eight_days",
                "greater_than_twenty_eight_days",
                "unknown_or_censored",
            ].index(item["key"]),
        ),
        "prior_season": {
            "season": "2023-24",
            "status": "not_loaded_in_v2",
            "note": f"No {args.team} prior-season injury and exposure denominator pair exists in this V2 workspace, so the dashboard leaves comparison blank rather than mixing in legacy report figures.",
        },
        "limitations": [
            setting_limitation,
            f"Adjudicated duplicate standardised row exclusions applied: {', '.join(sorted(excluded_injury_rows))}."
            if excluded_injury_rows
            else "No adjudicated duplicate injury row exclusions applied.",
            "Aggregate release is approved only after explicit confirmation of the live Supabase target; the dashboard reports aggregate values only.",
        ],
    }

    if dashboard["coverage"]["hours"] <= 0:
        raise SystemExit("exposure hours must be positive")
    if dashboard["headline"][2]["value"] is None or dashboard["headline"][5]["value"] is None:
        raise SystemExit("headline rates were not calculated")

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    analysis_source_output = clean_text(getattr(args, "analysis_source_output", ""))
    analysis_source_workbook = ""
    if analysis_source_output:
        export_excluded_fields = set(LOCATOR_FIELDS + UID_FIELDS)
        fieldnames = [
            field
            for field in (list(injury_source_rows[0]) if injury_source_rows else [])
            if field not in export_excluded_fields
        ]
        exported_rows = [{field: row.get(field, "") for field in fieldnames} for row in analysis_source_rows]
        analysis_source_path = Path(analysis_source_output)
        if analysis_source_path.suffix.lower() == ".xlsx":
            write_rows(analysis_source_path.with_suffix(".csv"), exported_rows, fieldnames)
            write_analysis_source_workbook(
                analysis_source_path,
                exported_rows,
                fieldnames,
                analysis_excluded_row_numbers,
                analysis_filled_columns,
            )
            analysis_source_workbook = str(analysis_source_path)
        else:
            write_rows(analysis_source_path, exported_rows, fieldnames)
            xlsx_path = analysis_source_path.with_suffix(".xlsx")
            write_analysis_source_workbook(
                xlsx_path,
                exported_rows,
                fieldnames,
                analysis_excluded_row_numbers,
                analysis_filled_columns,
            )
            analysis_source_workbook = str(xlsx_path)
    standardisation_audit_output = clean_text(getattr(args, "standardisation_audit_output", ""))
    if standardisation_audit_output:
        standardisation_audit_path = Path(standardisation_audit_output)
        write_rows(
            standardisation_audit_path,
            standardisation_events,
            [
                "standardised_row_number",
                "field",
                "old_value",
                "new_value",
                "action",
                "reason",
                "review_status",
            ],
        )
        dashboard["pipeline_evidence"]["standardisation_audit_file"] = str(
            standardisation_audit_path
        )
        dashboard["pipeline_evidence"]["standardisation_audit_sha256"] = sha256_file(
            standardisation_audit_path
        )
    output.write_text(json.dumps(dashboard, indent=2) + "\n")
    print(
        json.dumps(
            {
                "output": str(output),
                "analysis_source_output": analysis_source_output or None,
                "analysis_source_workbook": analysis_source_workbook or None,
                "standardisation_audit_output": standardisation_audit_output or None,
                "standardisation_events": len(standardisation_events),
                "team": args.team,
                "season": args.season,
            },
            indent=2,
        )
    )


def build_munster_dashboard(args: argparse.Namespace) -> None:
    build_team_dashboard(args)


def retired_legacy_dashboard_command(args: argparse.Namespace) -> None:
    raise SystemExit(
        "this local CSV dashboard command is retired because it bypasses the audited database "
        "pipeline. Use ingest -> process-intake/process-exposure -> build-curated -> release -> "
        "release-league; see docs/PIPELINE_RUNBOOK.md."
    )


def quiet_call(func: Any, args: argparse.Namespace) -> None:
    with contextlib.redirect_stdout(io.StringIO()):
        func(args)


def add_exposure_cli_parsers(subcommands: Any) -> None:
    exposure_parser = subcommands.add_parser("prepare-exposure")
    exposure_parser.add_argument("--team", required=True)
    exposure_parser.add_argument("--season", required=True)
    exposure_parser.add_argument("--file", required=True)
    exposure_parser.add_argument("--sheet", default="Standardized Data")
    exposure_parser.add_argument("--codebook", required=True)
    exposure_parser.add_argument("--output", required=True)
    exposure_parser.add_argument("--qc-output", required=True)
    exposure_parser.add_argument("--manifest", required=True)
    exposure_parser.add_argument("--reporting-grain", choices=EXPOSURE_REPORTING_GRAINS, required=True)
    exposure_parser.add_argument("--player-column", default="name")
    exposure_parser.add_argument("--date-column", default="session date")
    exposure_parser.add_argument("--minutes-column", default="minutes total")
    exposure_parser.add_argument("--distance-column", default="distance total")
    exposure_parser.add_argument("--derive-minutes-from-timestamps", action="store_true")
    exposure_parser.add_argument("--start-timestamp-column", default="session start date time")
    exposure_parser.add_argument("--end-timestamp-column", default="session end date time")
    exposure_parser.add_argument("--distance-source-file", default="")
    exposure_parser.add_argument("--distance-source-sheet", default="")
    exposure_parser.add_argument("--distance-source-column", default="")
    exposure_parser.add_argument("--date-order", choices=["month-first", "day-first"], default="month-first")
    exposure_parser.set_defaults(func=prepare_exposure)

    clean_exposure_parser = subcommands.add_parser("clean-exposure")
    clean_exposure_parser.add_argument("--file", required=True)
    clean_exposure_parser.add_argument("--team", required=True)
    clean_exposure_parser.add_argument("--season", required=True)
    clean_exposure_parser.add_argument("--output", required=True)
    clean_exposure_parser.add_argument("--qc-output", required=True)
    clean_exposure_parser.add_argument("--manifest", required=True)
    clean_exposure_parser.add_argument("--reporting-grain", choices=EXPOSURE_REPORTING_GRAINS, required=True)
    clean_exposure_parser.add_argument("--date-order", choices=["month-first", "day-first"], default="month-first")
    clean_exposure_parser.add_argument("--window-start", default="")
    clean_exposure_parser.add_argument("--window-end", default="")
    clean_exposure_parser.set_defaults(func=clean_exposure)

    process_exposure_parser = subcommands.add_parser("process-exposure")
    process_exposure_parser.add_argument("--team", required=True)
    process_exposure_parser.add_argument("--season", required=True)
    process_exposure_parser.add_argument("--file", required=True)
    process_exposure_parser.add_argument("--reporting-grain", choices=EXPOSURE_REPORTING_GRAINS, required=True)
    process_exposure_parser.add_argument("--step-name", default="exposure_cleaning")
    process_exposure_parser.add_argument("--step-version", default=EXPOSURE_PROCESSING_RULE_VERSION)
    process_exposure_parser.add_argument("--version-number", type=int, default=101)
    process_exposure_parser.set_defaults(func=process_exposure)


def self_check(args: argparse.Namespace) -> None:
    assert sha256_json({"b": 2, "a": 1}) == sha256_json({"a": 1, "b": 2})
    assert without_keys(
        {"keep": 1, "nested": {"protected": 2}, "items": [{"protected": 3}]},
        {"protected"},
    ) == {"keep": 1, "nested": {}, "items": [{}]}
    assert is_protected_team_alias_value("Team N")
    assert not is_protected_team_alias_value("Glasgow Warriors")
    params = SqlParams()
    assert params.text("value") == "(select value #>> '{}' from _pipeline_params where idx = 1)"
    assert params.jsonb({"key": "value"}) == "(select value from _pipeline_params where idx = 2)"
    assert params.values == ["value", {"key": "value"}]

    with tempfile.TemporaryDirectory() as profile_tmp:
        profile_dir = Path(profile_tmp)
        intake_file = profile_dir / "intake.csv"
        intake_file.write_text("player_uid\nply_example\n")
        input_sha = sha256_file(intake_file)
        reviewed_at = (datetime.now(UTC) - timedelta(minutes=2)).isoformat()
        approved_at = (datetime.now(UTC) - timedelta(minutes=1)).isoformat()
        profile_document = {
            "team": "Example Club", "season": "2024-25", "decision": "compatible",
            "profile_version": "1", "mapping_path": None, "mapping_sha256": None,
            "mapping_version": None, "ai_review_status": "completed", "ai_reviewed_by": "Codex",
            "ai_reviewed_at": reviewed_at, "approved_by": "Abdel Babiker",
            "approved_at": approved_at, "unresolved_adjudication_ids": [],
            "approved_input_sha256s": [input_sha],
        }
        profile_file = profile_dir / "team_intake_profile.json"
        profile_file.write_text(json.dumps(profile_document) + "\n")
        valid_profile = {
            **profile_document,
            "profile_path": str(profile_file), "profile_sha256": sha256_file(profile_file),
        }
        manifest_path = profile_dir / "manifest.json"

        def expect_profile_rejection(manifest: dict[str, Any], expected: str) -> None:
            try:
                validate_intake_profile_manifest(
                    manifest, manifest_path, input_sha, "Example Club", "2024-25"
                )
                raise AssertionError(f"profile gate should reject: {expected}")
            except SystemExit as exc:
                assert expected in str(exc)

        validate_intake_profile_manifest(
            {"intake_profile": valid_profile}, manifest_path, input_sha, "Example Club", "2024-25"
        )
        expect_profile_rejection({}, "requires an intake_profile object")
        for field, value, expected in [
            ("approved_input_sha256s", ["0" * 64], "current intake checksum is not covered"),
            ("profile_sha256", "0" * 64, "evidence checksum mismatch"),
            ("decision", "adjudication_required", "decision blocks ingest"),
            ("decision", "protocol_incompatible", "decision blocks ingest"),
            ("ai_review_status", "pending", "ai_review_status must be completed"),
            ("ai_reviewed_by", "Other model", "approval fields do not match"),
            ("approved_at", (datetime.now(UTC) + timedelta(days=1)).isoformat(), "neither value in the future"),
            ("unresolved_adjudication_ids", ["adj-1"], "has unresolved adjudications"),
            ("team", "Other Club", "team/season does not match ingest target"),
        ]:
            expect_profile_rejection({"intake_profile": {**valid_profile, field: value}}, expected)

        adapter_profile_file = profile_dir / "adapter_profile.json"
        adapter_document = {**profile_document, "decision": "adapter_required"}
        adapter_profile_file.write_text(json.dumps(adapter_document) + "\n")
        adapter_without_mapping = {
            **adapter_document, "profile_path": str(adapter_profile_file),
            "profile_sha256": sha256_file(adapter_profile_file),
        }
        expect_profile_rejection(
            {"intake_profile": adapter_without_mapping}, "requires a versioned mapping file"
        )
        mapping_file = profile_dir / "source_to_canonical_mapping.json"
        mapping_file.write_text(
            json.dumps({"mapping_version": "1", "mappings": [{"source": "Knee", "target": "knee"}]})
            + "\n"
        )
        mapped_adapter_profile_file = profile_dir / "mapped_adapter_profile.json"

        def mapped_adapter_manifest() -> dict[str, Any]:
            document = {
                **adapter_document, "mapping_path": str(mapping_file),
                "mapping_sha256": sha256_file(mapping_file), "mapping_version": "1",
            }
            mapped_adapter_profile_file.write_text(json.dumps(document) + "\n")
            return {
                "intake_profile": {
                    **document, "profile_path": str(mapped_adapter_profile_file),
                    "profile_sha256": sha256_file(mapped_adapter_profile_file),
                }
            }

        expect_profile_rejection(
            mapped_adapter_manifest(), "each intake mapping requires canonical_field"
        )
        mapping_file.write_text(json.dumps({
            "mapping_version": "1",
            "mappings": [{
                "canonical_field": "body_location", "canonical_value": "knee",
                "source_evidence": {"Body Part": "Knee"}, "evidence_class": "source_reported",
            }],
        }) + "\n")
        validate_intake_profile_manifest(
            mapped_adapter_manifest(),
            manifest_path,
            input_sha,
            "Example Club",
            "2024-25",
        )
        mismatched_profile_file = profile_dir / "mismatched_profile.json"
        mismatched_profile_file.write_text(
            json.dumps({**profile_document, "decision": "protocol_incompatible"}) + "\n"
        )
        expect_profile_rejection(
            {
                "intake_profile": {
                    **valid_profile,
                    "profile_path": str(mismatched_profile_file),
                    "profile_sha256": sha256_file(mismatched_profile_file),
                }
            },
            "approval fields do not match checksummed profile JSON",
        )
        invalid_manifest_path = profile_dir / "invalid_manifest.json"
        invalid_manifest_path.write_text("{}\n")
        original_read_rows = read_rows

        def fail_if_rows_are_loaded(_path: Path) -> list[dict[str, str]]:
            raise AssertionError("ingest loaded rows before validating the profile gate")

        globals()["read_rows"] = fail_if_rows_are_loaded
        try:
            quiet_call(
                ingest,
                argparse.Namespace(
                    file=str(intake_file),
                    manifest=str(invalid_manifest_path),
                    team="Example Club",
                    season="2024-25",
                ),
            )
            raise AssertionError("ingest should enforce the profile gate before database access")
        except SystemExit as exc:
            assert "requires an intake_profile object" in str(exc)
        finally:
            globals()["read_rows"] = original_read_rows

    # Generic exposure commands must identify their target and paths explicitly,
    # and must never infer reporting grain from protected source aliases.
    exposure_cli_parser = argparse.ArgumentParser(add_help=False)
    exposure_cli_subcommands = exposure_cli_parser.add_subparsers(required=True)
    add_exposure_cli_parsers(exposure_cli_subcommands)
    exposure_cli_cases = {
        "prepare-exposure": [
            ("--team", "Example Club"),
            ("--season", "2024-25"),
            ("--file", "source.xlsx"),
            ("--codebook", "codebook.csv"),
            ("--output", "prepared.csv"),
            ("--qc-output", "prepare-qc.json"),
            ("--manifest", "manifest.json"),
            ("--reporting-grain", "weekly"),
        ],
        "clean-exposure": [
            ("--team", "Example Club"),
            ("--season", "2024-25"),
            ("--file", "prepared.csv"),
            ("--output", "cleaned.csv"),
            ("--qc-output", "clean-qc.json"),
            ("--manifest", "manifest.json"),
            ("--reporting-grain", "session"),
        ],
        "process-exposure": [
            ("--team", "Example Club"),
            ("--season", "2024-25"),
            ("--file", "cleaned.csv"),
            ("--reporting-grain", "weekly"),
        ],
    }
    for command, option_pairs in exposure_cli_cases.items():
        valid_argv = [command] + [item for pair in option_pairs for item in pair]
        parsed = exposure_cli_parser.parse_args(valid_argv)
        assert parsed.reporting_grain in EXPOSURE_REPORTING_GRAINS
        for omitted_option, _ in option_pairs:
            omitted_argv = [command] + [
                item
                for option_pair in option_pairs
                if option_pair[0] != omitted_option
                for item in option_pair
            ]
            with contextlib.redirect_stderr(io.StringIO()):
                try:
                    exposure_cli_parser.parse_args(omitted_argv)
                    raise AssertionError(f"{command} should require {omitted_option}")
                except SystemExit as exc:
                    assert exc.code == 2

    provenance = run_provenance()
    assert re.fullmatch(r"[0-9a-f]{40}(-dirty)?", provenance["code_version"]), provenance["code_version"]
    assert re.fullmatch(r"[0-9a-f]{64}", provenance["dependency_lock_hash"])
    assert provenance["operator"].strip()
    assert set(provenance) == {"code_version", "dependency_lock_hash", "operator"}

    assert parse_flexible_date("10/7/24").date().isoformat() == "2024-10-07"
    assert parse_flexible_date("10/7/24", "day-first").date().isoformat() == "2024-07-10"
    assert parse_minutes("01:30:30") == 90.5
    assert effective_days_injured({"Days Injured": "", "Training Days Missed": "12"}) == 12
    assert effective_days_injured({"Date Injured": "01/07/2024", "Confirmed Return Date": "01/07/2024", "Training Days Missed": "-1"}) == 0
    assert effective_days_injured({"Date Injured": "13/12/2024", "Confirmed Return Date": "06/01/2024", "Training Days Missed": "-343"}) == 23
    assert effective_days_injured({"Date Injured": "30/12/2024", "Confirmed Return Date": "07/12/2024", "Training Days Missed": "-24", "TimeLoss vs Medical Attention": "Medical Attention"}) == 0
    assert effective_days_injured({"Training Days Missed": "-1", "TimeLoss vs Medical Attention": "Time Loss"}) == 0
    assert effective_confirmed_return_date(
        {"Date Injured": "02/10/2024", "Confirmed Return Date": "03/10/2024"},
        1,
        "mapped_from_training_days_missed_date_conflict",
    ) == (date(2024, 10, 4), "derived_from_training_days_missed_date_conflict")
    assert effective_confirmed_return_date(
        {"Date Injured": "10/05/2025", "Confirmed Return Date": "05/10/2025"},
        0,
        "mapped_from_training_days_missed_date_conflict",
    ) == (date(2025, 5, 10), "derived_same_day_return_from_zero_days_date_conflict")
    assert effective_orchard_code({"Problem type": "Injury", "Illness Code": "TMXX"}) == "TMXX"
    assert effective_orchard_code({"Problem type": "Illness", "Illness Code": "R05"}) == ""
    assert problem_type({"Problem type": "Illness", "Injury Tissue Type/s": "Respiratory"}) == (
        "illness",
        "mapped_from_problem_type",
    )
    assert exposure_scope_status({}) == ("scope_unknown_included", "blank_scope_fields_retained")
    assert exposure_scope_status({"Competition": "academy"}, "Edinburgh")[0] == "in_scope_explicit"
    assert exposure_scope_status({"session type": "Scotland U20"})[0] == "out_of_scope_explicit"
    assert exposure_scope_status({"session type": "National Academy"})[0] == "out_of_scope_explicit"
    assert exposure_scope_status({"session type": "Academy Training"}, "Edinburgh")[0] == "in_scope_explicit"
    assert exposure_scope_status({"session type": "Academy Game (Glasgow A)"}, "Edinburgh") == (
        "out_of_scope_explicit",
        "academy_game",
    )
    assert exposure_scope_status({"session type": "Game (Connacht A)"}, "Edinburgh") == (
        "in_scope_explicit",
        "urc_opponent_game",
    )
    assert exposure_scope_status({"session type": "Game (Benneton H)"}, "Edinburgh") == (
        "in_scope_explicit",
        "urc_opponent_game",
    )
    assert exposure_scope_status({"session type": "Game (Bath H)"}, "Edinburgh") == (
        "out_of_scope_explicit",
        "non_urc_or_unspecified_game",
    )
    assert exposure_scope_status({"session type": "Game Warm Up"}, "Edinburgh") == (
        "in_scope_explicit",
        "warmup_or_topup_retained",
    )
    assert exposure_scope_status({"session type": "Game Top Up"}, "Edinburgh") == (
        "in_scope_explicit",
        "warmup_or_topup_retained",
    )
    assert exposure_scope_status({"Training With": "Academy Squad"}) == (
        "out_of_scope_explicit",
        "academy_game",
    )
    assert exposure_scope_status({"Training Type": "RTP"}) == (
        "out_of_scope_explicit",
        "explicit_rehab_or_rtp",
    )
    assert exposure_scope_status({"session type": "Club Training"})[0] == "in_scope_explicit"
    assert injury_cohort_exclusion_reasons(
        {"Received/Injured In Team": "Team M", "Match Type": "URC"}, "Team M"
    ) == []
    assert injury_cohort_exclusion_reasons(
        {"Received/Injured In Team": "N/A", "Match Type": "Other"}, "Team M"
    ) == []
    assert injury_cohort_exclusion_reasons(
        {"Received/Injured In Team": "Club", "Match Type": "Champions Cup"}, "Team M"
    ) == ["received_or_injured_in_other_team", "explicit_non_urc_match_type"]
    assert injury_cohort_exclusion_reasons({"Match Type": "Pro team A game"}) == [
        "explicit_non_urc_match_type"
    ]
    assert injury_cohort_exclusion_reasons({"Match Type": "Unclear competition label"}) == []

    # Phase 3.5 cohort-signal capture (Adjudication 4): offline, pure-function
    # coverage. None of these assertions touches the real, Git-ignored
    # team_alias_map.json -- every "own_team_alias" below is a fake test value.
    assert set(TEAM_KEY_ALIAS_LOOKUP_NAMES) == {
        "connacht", "leinster", "munster", "ulster",
        "cardiff", "dragons", "ospreys", "scarlets",
        "bulls", "lions", "sharks", "stormers",
        "benetton", "zebre", "edinburgh", "glasgow",
    }
    assert own_team_alias_for("connacht", {"Connacht": "Team Q"}) == "Team Q"
    try:
        own_team_alias_for("connacht", {"Some Other Team": "Team Q"})
        raise AssertionError("own_team_alias_for should refuse a map with no entry for the lookup name")
    except SystemExit:
        pass
    try:
        own_team_alias_for("not-a-real-team-key", {"Connacht": "Team Q"})
        raise AssertionError("own_team_alias_for should refuse an unconfigured team_key")
    except SystemExit:
        pass

    assert received_in_team_status({"Received/Injured In Team": "Team Q"}, "Team Q") == (
        "own_team", "matched_own_team_alias"
    )
    assert received_in_team_status({"Received/Injured In Team": "team q"}, "Team Q")[0] == "own_team"
    assert received_in_team_status({"Received/Injured In Team": "Team R"}, "Team Q")[0] == "other_team"
    assert received_in_team_status({"Received/Injured In Team": "Club"}, "Team Q")[0] == "club"
    assert received_in_team_status({"Received/Injured In Team": "Edinburgh A"}, "Team Q")[0] == "other_team"
    assert received_in_team_status({"Received/Injured In Team": "N/A"}, "Team Q")[0] == "missing"
    assert received_in_team_status({"Received/Injured In Team": ""}, "Team Q")[0] == "missing"
    assert received_in_team_status({}, "Team Q")[0] == "missing"

    assert urc_match_scope({"Match Type": "URC"})[0] == "urc"
    assert urc_match_scope({"Match Type": "United Rugby Championship"})[0] == "urc"
    assert urc_match_scope({"Match Type": "Other"})[0] == "urc"
    assert urc_match_scope({"Match Type": "training"}) == ("training", "mapped_from_match_type_training")
    assert urc_match_scope({"Match Type": "Challenge Cup"})[0] == "non_urc_marker"
    assert urc_match_scope({"Match Type": "Italian Elite Championship"})[0] == "non_urc_marker"
    assert urc_match_scope({"Match Type": "Pro team A game"})[0] == "non_urc_marker"
    assert urc_match_scope({"Match Type": "NAG U20"})[0] == "non_urc_marker"
    assert urc_match_scope({"Match Type": "N/A"})[0] == "unknown"
    assert urc_match_scope({})[0] == "unknown"

    cohort_signal_state, cohort_signal_events = build_processing_state(
        {
            **{field: "fixture" for field in LOCATOR_FIELDS},
            "standardised_row_number": "2",
            "player_uid": "player_1",
            "injury_uid": "injury_1",
            "Date Injured": "02/07/2024",
            "Days Injured": "10",
            "Confirmed Return Date": "",
            "Occasion category": "Game",
            "Match Type": "URC",
            "Received/Injured In Team": "Team Q",
        },
        window_start=datetime(2024, 7, 1),
        window_end=datetime(2025, 6, 30),
        duplicate_signature_rows=set(),
        own_team_alias="Team Q",
    )
    assert cohort_signal_state["received_in_team_status"] == "own_team"
    assert cohort_signal_state["urc_match_scope"] == "urc"
    assert cohort_signal_state["field_origins"]["received_in_team_status"] == "matched_own_team_alias"
    assert any(
        event["field_name"] == "received_in_team_status" and event["reason_code"] == "cohort_signal_derivation"
        for event in cohort_signal_events
    )
    assert any(
        event["field_name"] == "urc_match_scope" and event["reason_code"] == "cohort_signal_derivation"
        for event in cohort_signal_events
    )
    # Backward-compatible default: an omitted own_team_alias must still
    # populate both new keys (conservatively, as other_team) rather than
    # raise or silently skip them.
    default_state, _ = build_processing_state(
        {
            **{field: "fixture" for field in LOCATOR_FIELDS},
            "standardised_row_number": "2",
            "player_uid": "player_1",
            "injury_uid": "injury_1",
            "Date Injured": "02/07/2024",
            "Received/Injured In Team": "Team Q",
        },
        window_start=datetime(2024, 7, 1),
        window_end=datetime(2025, 6, 30),
        duplicate_signature_rows=set(),
    )
    assert default_state["received_in_team_status"] == "other_team"

    assert severity_category(10, True)[0] == "eight_to_twenty_eight_days"
    assert contact_context({"Is Contact": "NA", "Injury Tissue Type/s": "Muscle Strain/Spasm", "Nature of onset": "Acute"}) == (
        "non_contact",
        "inferred_from_acute_muscle_strain",
    )
    assert activity_context({"Occasion category": "match", "Match Type": "URC"}) == (
        "urc_match",
        "mapped_from_occasion_and_match_type",
    )
    assert activity_context({"Occasion category": "Match", "Match Type": "Challenge Cup"}) == (
        "match",
        "mapped_from_occasion_category_non_urc_match",
    )
    assert contact_context({"Is Contact": "Contact (with other player)"})[0] == "contact"
    assert contact_context({"Is Contact": "Overuse (gradual onset)"})[0] == "non_contact"
    assert recurrence_status({"Recurrence": "New injury (not recurrent)"})[0] == "first_episode"
    assert recurrence_status({"Recurrence": "Delayed recurrence"})[0] == "recurrence"
    assert injury_type({"Injury Tissue Type/s": "Other Pain/ unspecified", "Orchard Code": "NJPX"}) == (
        "joint_sprain",
        "mapped_from_orchard_code_ioc_pathology",
    )
    assert body_location({"Body Part": "Lower leg"}) == ("lower_leg", "preserved_controlled_body_part")
    assert injury_type({"Injury Tissue Type/s": "Muscle injury"}) == (
        "muscle_injury",
        "preserved_controlled_injury_tissue_type",
    )
    assert injury_type({"Problem type": "Injury", "Injury Tissue Type/s": "Biceps femoris strain grade 1 - 2"})[0] == "muscle_injury"
    detailed_export = filled_injury_export_row(
        {
            "Problem type": "Injury",
            "Injury Tissue Type/s": "Anterior talofibular ligament sprain",
            "Recurrence": "Early recurrence",
            "Training Days Missed": "5",
        }
    )
    assert detailed_export["Description"] == "Anterior talofibular ligament sprain"
    assert detailed_export["Injury Tissue Type/s"] == "Joint sprain"
    assert detailed_export["Recurrence"] == "Recurrence"
    assert detailed_export["Recurrence Stage"] == "Early recurrence"
    assert detailed_export["Days Injured"] == "5"
    assert detailed_export["Training Days Missed"] == ""
    assert set(BODY_LOCATION_LABELS.values()) == CONTROLLED_BODY_LOCATION_LABELS
    assert set(INJURY_TYPE_LABELS.values()) == CONTROLLED_INJURY_TYPE_LABELS
    unknown_export = filled_injury_export_row({"Date Injured": "02/07/2024", "Days Injured": "0"})
    assert unknown_export["Body Part"] == "Unknown"
    assert unknown_export["Injury Tissue Type/s"] == "Unknown"
    assert unknown_export["Fit for selection"] == "Unknown"
    assert filled_injury_export_row({"TimeLoss vs Medical Attention": "Time Loss"})[
        "TimeLoss vs Medical Attention"
    ] == "Time Loss"
    assert rate_per_1000(1, 2) == 500

    locator = {field: "fixture" for field in LOCATOR_FIELDS}
    locator["standardised_row_number"] = "2"
    state, events = build_processing_state(
        {
            **locator,
            "player_uid": "player_1",
            "injury_uid": "injury_1",
            "Date Injured": "02/07/2024",
            "Days Injured": "10",
            "Confirmed Return Date": "",
            "is_injury_closed": "1",
            "Occasion category": "Game",
            "Match Type": "United Rugby Championship",
            "Is Contact": "Contact",
            "Recurrence": "First Episode",
            "Body Part": "Ankle",
            "Orchard Code": "",
            "Injury Tissue Type/s": "Muscle",
            "Illness Code": "",
            "Description": "",
        },
        window_start=datetime(2024, 7, 1),
        window_end=datetime(2025, 6, 30),
        duplicate_signature_rows=set(),
    )
    assert state["analysis_eligibility_status"] == "included_pending_protocol"
    assert state["derived_return_date"] == "2024-07-12"
    assert state["activity_context"] == "urc_match"
    assert state["body_location"] == "ankle"
    assert any(event["field_name"] == "derived_return_date" for event in events)

    # Standing-adjudication reapplication (offline, pure): a rerun must never
    # silently revert an adjudicated eligibility_status.
    reapply_state = {"analysis_eligibility_status": "included_pending_protocol"}
    reapply_events: list[dict[str, Any]] = []
    assert apply_standing_adjudication(reapply_state, reapply_events, None) is False
    assert reapply_state == {"analysis_eligibility_status": "included_pending_protocol"}
    assert reapply_events == []
    # duplicate_review-style decisions carry no eligibility override: no-op.
    assert (
        apply_standing_adjudication(
            reapply_state,
            reapply_events,
            {
                "adjudication_id": "fixture-review",
                "decision": {"decision": "not_duplicate_distinct_injury"},
                "rationale": "retain both rows",
            },
        )
        is False
    )
    assert reapply_state == {"analysis_eligibility_status": "included_pending_protocol"}
    assert reapply_events == []
    exclusion_adjudication = {
        "adjudication_id": "fixture-exclusion",
        "decision": {
            "analysis_eligibility_status": "excluded_duplicate_adjudicated",
            "decision": "exclude_duplicate",
            "excluded_standardised_row_number": 2,
            "duplicate_of_standardised_row_number": 3,
            "rationale": "adjudicated duplicate of row 3",
        },
        "rationale": "adjudicated duplicate of row 3",
    }
    assert apply_standing_adjudication(reapply_state, reapply_events, exclusion_adjudication) is True
    assert reapply_state["analysis_eligibility_status"] == "excluded_duplicate_adjudicated"
    assert reapply_state["decision"] == "exclude_duplicate"
    assert len(reapply_events) == 1
    assert reapply_events[0]["action"] == "reapply"
    assert reapply_events[0]["reason_code"] == "adjudication_reapplied"
    assert reapply_events[0]["review_status"] == "adjudicated"
    assert reapply_events[0]["old_value"] == "included_pending_protocol"
    assert reapply_events[0]["new_value"] == "excluded_duplicate_adjudicated"
    assert "fixture-exclusion" in reapply_events[0]["rationale"]
    # Idempotent: reapplying when the status already matches is a no-op.
    assert apply_standing_adjudication(reapply_state, reapply_events, exclusion_adjudication) is False
    assert len(reapply_events) == 1

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        from openpyxl import Workbook

        standardised_source = tmp_path / "standardised.xlsx"
        source_workbook = Workbook()
        source_sheet = source_workbook.active
        source_sheet.title = "Standardized Data"
        source_sheet.append(["PlayerID"])
        source_sheet.append(["player_1"])
        source_sheet.append([])
        source_sheet.append(["player_2"])
        source_workbook.save(standardised_source)
        standardised_csv = tmp_path / "standardised.csv"
        prepared_csv = tmp_path / "prepared.csv"
        write_rows(standardised_csv, [{"PlayerID": "player_1"}, {"PlayerID": "player_2"}], ["PlayerID"])
        quiet_call(
            prepare_intake,
            argparse.Namespace(
                team="Self Check",
                season="2024-25",
                file=str(standardised_csv),
                source_file=str(standardised_source),
                output=str(prepared_csv),
                manifest=None,
                source_sheet="Standardized Data",
                player_id_column="PlayerID",
            ),
        )
        assert [row["source_row_number"] for row in read_rows(prepared_csv)] == ["2", "4"]

        exposure_workbook_path = tmp_path / "exposure.xlsx"
        exposure_workbook = Workbook()
        exposure_sheet = exposure_workbook.active
        exposure_sheet.title = "Standardized Data"
        exposure_sheet.append(["Team", "name", "session date", "minutes total", "distance total"])
        exposure_sheet.append(["source-team-label", "player", "07/01/2024", "120", "10000"])
        exposure_workbook.save(exposure_workbook_path)
        exposure_workbook.close()
        exposure_codebook = tmp_path / "exposure_codebook.csv"
        write_rows(
            exposure_codebook,
            [{"Standard_Column_Name": column} for column in ["name", "session date", "minutes total", "distance total"]],
            ["Standard_Column_Name"],
        )
        exposure_manifest = tmp_path / "manifest.json"
        exposure_manifest.write_text("{}\n")
        prepared_exposure_file = tmp_path / "prepared_exposure.csv"
        prepared_exposure_qc = tmp_path / "prepared_exposure_qc.json"
        quiet_call(
            prepare_exposure,
            argparse.Namespace(
                team="Example Club",
                season="2024-25",
                file=str(exposure_workbook_path),
                sheet="Standardized Data",
                codebook=str(exposure_codebook),
                output=str(prepared_exposure_file),
                qc_output=str(prepared_exposure_qc),
                manifest=str(exposure_manifest),
                reporting_grain="weekly",
                player_column="name",
                date_column="session date",
                minutes_column="minutes total",
                distance_column="distance total",
                date_order="month-first",
            ),
        )
        prepared_exposure_rows = read_rows(prepared_exposure_file)
        assert {row[EXPOSURE_DECLARED_GRAIN_FIELD] for row in prepared_exposure_rows} == {"weekly"}
        prepare_exposure_qc = json.loads(prepared_exposure_qc.read_text())
        assert prepare_exposure_qc["team"] == "Example Club"
        assert prepare_exposure_qc["season"] == "2024-25"
        assert prepare_exposure_qc["exposure_reporting_grain"]["current_file_reporting_grain"] == "weekly"
        prepared_manifest = json.loads(exposure_manifest.read_text())
        assert prepared_manifest["exposure_intake"]["team"] == "Example Club"
        assert prepared_manifest["exposure_intake"]["season"] == "2024-25"
        assert prepared_manifest["exposure_intake"]["exposure_reporting_grain"]["selection_source"] == "required_cli_argument"

        exposure_source = tmp_path / "exposure.csv"
        exposure_clean = tmp_path / "exposure_clean.csv"
        exposure_qc = tmp_path / "exposure_qc.json"
        write_rows(
            exposure_source,
            [
                {
                    "Team": "source-team-label",
                    "Competition": "",
                    "session type": "",
                    "If match, surface?": "",
                    "name": "player",
                    "session date": "07/01/2024",
                    "minutes total": "300",
                    "distance total": "25000",
                    "player_uid": "player_1",
                    "source_row_number": "2",
                    "source_row_sha256": "row_a",
                    EXPOSURE_DECLARED_GRAIN_FIELD: "weekly",
                },
                {
                    "Team": "source-team-label",
                    "Competition": "academy",
                    "session type": "",
                    "If match, surface?": "",
                    "name": "player",
                    "session date": "07/08/2024",
                    "minutes total": "300",
                    "distance total": "25000",
                    "player_uid": "player_1",
                    "source_row_number": "3",
                    "source_row_sha256": "row_b",
                    EXPOSURE_DECLARED_GRAIN_FIELD: "weekly",
                },
                {
                    "Team": "source-team-label",
                    "Competition": "",
                    "session type": "",
                    "If match, surface?": "",
                    "name": "player",
                    "session date": "07/01/2024",
                    "minutes total": "300",
                    "distance total": "25000",
                    "player_uid": "player_1",
                    "source_row_number": "4",
                    "source_row_sha256": "row_a",
                    EXPOSURE_DECLARED_GRAIN_FIELD: "weekly",
                },
            ],
            [
                "Team",
                "Competition",
                "session type",
                "If match, surface?",
                "name",
                "session date",
                "minutes total",
                "distance total",
                "player_uid",
                "source_row_number",
                "source_row_sha256",
                EXPOSURE_DECLARED_GRAIN_FIELD,
            ],
        )
        quiet_call(
            clean_exposure,
            argparse.Namespace(
                file=str(exposure_source),
                output=str(exposure_clean),
                qc_output=str(exposure_qc),
                manifest=None,
                date_order="month-first",
                team="Edinburgh",
                season="2024-25",
                reporting_grain="weekly",
                window_start="2024-07-06",
                window_end="2024-07-10",
            ),
        )
        cleaned = read_rows(exposure_clean)
        cleaning_actions = [row["cleaning_action"] for row in cleaned]
        assert cleaning_actions == [
            "include",
            "include",
            "exclude_from_primary",
        ], cleaning_actions
        assert cleaned[2]["exclusion_reason"] == "exact_duplicate_copy"
        assert {row["exposure_grain"] for row in cleaned} == {"weekly"}
        assert cleaned[0]["week_start_date"] == "2024-07-01"
        weekly_qc = json.loads(exposure_qc.read_text())
        # These rows exceed the session caps but remain valid under the
        # explicitly selected weekly rule set.
        assert weekly_qc["included_minutes_total"] == 600.0
        assert weekly_qc["analysis_window"]["start"] == "2024-07-06"
        assert weekly_qc["reporting_grain"] == "weekly"
        assert weekly_qc["reporting_grain_evidence"]["prepared_row_declaration"] == "matched_all_rows"
        prepared_grain_rows = read_rows(exposure_source)
        missing_grain_cases = {
            "zero": [
                {key: value for key, value in row.items() if key != EXPOSURE_DECLARED_GRAIN_FIELD}
                for row in prepared_grain_rows
            ],
            "partial": [
                {
                    **row,
                    EXPOSURE_DECLARED_GRAIN_FIELD: (
                        row[EXPOSURE_DECLARED_GRAIN_FIELD] if index == 0 else ""
                    ),
                }
                for index, row in enumerate(prepared_grain_rows)
            ],
        }
        for case_name, case_rows in missing_grain_cases.items():
            missing_grain_source = tmp_path / f"{case_name}_declared_grain.csv"
            write_rows(missing_grain_source, case_rows, list(case_rows[0]))
            try:
                quiet_call(
                    clean_exposure,
                    argparse.Namespace(
                        file=str(missing_grain_source),
                        output=str(tmp_path / f"{case_name}_declared_grain_clean.csv"),
                        qc_output=str(tmp_path / f"{case_name}_declared_grain_qc.json"),
                        manifest=None,
                        date_order="month-first",
                        team="Edinburgh",
                        season="2024-25",
                        reporting_grain="weekly",
                        window_start="",
                        window_end="",
                    ),
                )
                raise AssertionError(f"clean-exposure should reject {case_name} prepared grain declarations")
            except SystemExit as exc:
                assert "rerun prepare-exposure before cleaning" in str(exc)
        try:
            quiet_call(
                clean_exposure,
                argparse.Namespace(
                    file=str(exposure_source),
                    output=str(tmp_path / "mismatched_clean.csv"),
                    qc_output=str(tmp_path / "mismatched_clean_qc.json"),
                    manifest=None,
                    date_order="month-first",
                    team="Edinburgh",
                    season="2024-25",
                    reporting_grain="session",
                    window_start="",
                    window_end="",
                ),
            )
            raise AssertionError("clean-exposure should reject a prepared-row grain mismatch")
        except SystemExit as exc:
            assert "does not match the grain declared during prepare-exposure" in str(exc)
        with contextlib.redirect_stderr(io.StringIO()):
            try:
                process_exposure(
                    argparse.Namespace(
                        file=str(exposure_clean),
                        team="Edinburgh",
                        season="2024-25",
                        reporting_grain="session",
                        version_number=101,
                    )
                )
                raise AssertionError("process-exposure should reject a reporting-grain mismatch")
            except SystemExit as exc:
                assert "does not match cleaned exposure row grain" in str(exc)

        injury_source = tmp_path / "injury.csv"
        dashboard_output = tmp_path / "dashboard.json"
        analysis_source_output = tmp_path / "analysis_source.csv"
        analysis_audit_output = tmp_path / "analysis_audit.csv"
        write_rows(
            injury_source,
            [
                {
                    "Date Injured": "2024-07-02",
                    "Confirmed Return Date": "",
                    "Days Injured": "10",
                    "is_injury_closed": "1",
                    "Occasion category": "Game",
                    "Body Part": "Ankle",
                    "Injury Tissue Type/s": "Muscle",
                    "Received/Injured In Team": "Team Z",
                },
                {
                    "Date Injured": "01/01/2026",
                    "Confirmed Return Date": "",
                    "Days Injured": "5",
                    "is_injury_closed": "1",
                    "Occasion category": "Game",
                    "Body Part": "Ankle",
                    "Injury Tissue Type/s": "Unknown",
                    "Received/Injured In Team": "",
                }
            ],
            [
                "Date Injured",
                "Confirmed Return Date",
                "Days Injured",
                "is_injury_closed",
                "Occasion category",
                "Body Part",
                "Injury Tissue Type/s",
                "Received/Injured In Team",
            ],
        )
        quiet_call(
            build_team_dashboard,
            argparse.Namespace(
                team="Self Check",
                season="2024-25",
                injury_file=str(injury_source),
                exposure_file=str(exposure_clean),
                output=str(dashboard_output),
                analysis_source_output=str(analysis_source_output),
                standardisation_audit_output=str(analysis_audit_output),
                injured_in_team="Team Z",
            ),
        )
        dashboard = json.loads(dashboard_output.read_text())
        assert dashboard["coverage"]["hours"] == 10.0
        assert dashboard["headline"][0]["value"] == 1
        assert dashboard["headline"][2]["value"] == 100.0
        assert dashboard["headline"][5]["value"] == 1000.0
        analysis_rows = read_rows(analysis_source_output)
        assert len(analysis_rows) == 2
        assert list(analysis_rows[0]) == [
            "Date Injured",
            "Confirmed Return Date",
            "Days Injured",
            "is_injury_closed",
            "Occasion category",
            "Body Part",
            "Injury Tissue Type/s",
            "Received/Injured In Team",
        ]
        assert analysis_rows[0]["Date Injured"] == "02/07/2024"
        assert analysis_rows[0]["Confirmed Return Date"] == "12/07/2024"
        assert analysis_rows[0]["Received/Injured In Team"] == "Self Check"
        assert all("Team Z" not in row.values() for row in analysis_rows)
        assert "Problem type origin" not in analysis_rows[0]
        assert any(
            row["standardised_row_number"] == "3"
            and row["reason"] == "injury_date_missing_or_outside_exposure_coverage"
            for row in read_rows(analysis_audit_output)
        )
        from openpyxl import load_workbook

        analysis_workbook = load_workbook(analysis_source_output.with_suffix(".xlsx"))
        analysis_sheet = analysis_workbook.active
        def font_rgb(cell: Any) -> str:
            color = cell.font.color
            return color.rgb if color and color.type == "rgb" else ""

        assert not font_rgb(analysis_sheet["A2"]).endswith(("008000", "C00000"))
        assert font_rgb(analysis_sheet["B2"]).endswith("008000")
        assert font_rgb(analysis_sheet["A3"]).endswith("C00000")
        analysis_workbook.close()
        xlsx_only_output = tmp_path / "analysis_source_direct.xlsx"
        quiet_call(
            build_team_dashboard,
            argparse.Namespace(
                team="Self Check",
                season="2024-25",
                injury_file=str(injury_source),
                exposure_file=str(exposure_clean),
                output=str(tmp_path / "dashboard_direct_xlsx.json"),
                analysis_source_output=str(xlsx_only_output),
                injured_in_team="Team Z",
            ),
        )
        assert xlsx_only_output.exists()
        assert read_rows(xlsx_only_output.with_suffix(".csv")) == analysis_rows
        session_exposure_source = tmp_path / "session_exposure.csv"
        session_exposure_clean = tmp_path / "session_exposure_clean.csv"
        session_exposure_qc = tmp_path / "session_exposure_qc.json"
        write_rows(
            session_exposure_source,
            [
                {
                    "Team": "source-team-label",
                    "Competition": "",
                    "session type": "",
                    "If match, surface?": "",
                    "name": "player",
                    "session date": "07/02/2024",
                    "minutes total": "60",
                    "distance total": "8000",
                    "player_uid": "player_1",
                    "source_row_number": "2",
                    EXPOSURE_DECLARED_GRAIN_FIELD: "session",
                },
                {
                    "Team": "source-team-label",
                    "Competition": "",
                    "session type": "",
                    "If match, surface?": "",
                    "name": "player_2",
                    "session date": "07/02/2024",
                    "minutes total": "300",
                    "distance total": "25000",
                    "player_uid": "player_2",
                    "source_row_number": "3",
                    EXPOSURE_DECLARED_GRAIN_FIELD: "session",
                },
            ],
            [
                "Team",
                "Competition",
                "session type",
                "If match, surface?",
                "name",
                "session date",
                "minutes total",
                "distance total",
                "player_uid",
                "source_row_number",
                EXPOSURE_DECLARED_GRAIN_FIELD,
            ],
        )
        quiet_call(
            clean_exposure,
            argparse.Namespace(
                file=str(session_exposure_source),
                output=str(session_exposure_clean),
                qc_output=str(session_exposure_qc),
                manifest=None,
                date_order="month-first",
                team="Edinburgh",
                season="2024-25",
                reporting_grain="session",
            ),
        )
        quiet_call(
            build_team_dashboard,
            argparse.Namespace(
                team="Session Check",
                season="2024-25",
                injury_file=str(injury_source),
                exposure_file=str(session_exposure_clean),
                output=str(dashboard_output),
            ),
        )
        session_dashboard = json.loads(dashboard_output.read_text())
        session_cleaned = read_rows(session_exposure_clean)
        assert {row["exposure_grain"] for row in session_cleaned} == {"session"}
        assert [row["cleaning_action"] for row in session_cleaned] == ["include", "exclude_from_primary"]
        assert session_cleaned[1]["exclusion_reason"] == "session_minutes_above_220;session_distance_above_20000m"
        session_qc = json.loads(session_exposure_qc.read_text())
        assert session_qc["reporting_grain"] == "session"
        assert session_qc["reporting_grain_evidence"]["prepared_row_declaration"] == "matched_all_rows"
        assert session_dashboard["coverage"]["exposure_grain"] == "session"
        assert session_dashboard["coverage"]["weeks"] == 0
        assert session_dashboard["analysis_window"]["end"] == "2024-07-02"

    assert PLAYER_HOURS_PER_TEAM_MATCH == 20.0
    assert curated_source_version_set_hash(["b", "a"], ["c"]) == curated_source_version_set_hash(["a", "b"], ["c"])
    assert curated_source_version_set_hash(["a"], ["c"]) != curated_source_version_set_hash(["a"], ["d"])
    assert curated_source_version_set_hash([], []) == curated_source_version_set_hash([], [])

    # The curated_layer migration's code_lists seed must stay in lockstep
    # with the pipeline's controlled vocabularies: a drift here (someone
    # adds a body-location bucket to BODY_LOCATION_LABELS but forgets the
    # migration, or vice versa) would either reject valid values at
    # build-curated time or silently accept a bucket the pipeline can never
    # emit. This is a text-level check (no DB connection required).
    curated_migration_path = REPO_ROOT / "supabase" / "migrations" / f"{CURATED_LAYER_MIGRATION_VERSION}_curated_layer.sql"
    curated_migration_text = curated_migration_path.read_text()
    seeded_pairs = set(re.findall(r"\('([a-z_]+)', '([a-z0-9_]+)',", curated_migration_text))

    def seeded_codes(list_name: str) -> set[str]:
        return {code for (seeded_list, code) in seeded_pairs if seeded_list == list_name}

    assert seeded_codes("body_location") == set(BODY_LOCATION_LABELS), seeded_codes("body_location") ^ set(BODY_LOCATION_LABELS)
    assert seeded_codes("injury_type") == set(INJURY_TYPE_LABELS), seeded_codes("injury_type") ^ set(INJURY_TYPE_LABELS)
    assert seeded_codes("activity_context") == {"urc_match", "training", "match", "unknown"}
    assert seeded_codes("contact_context") == {"contact", "non_contact", "unknown"}
    assert seeded_codes("recurrence_status") == {"first_episode", "recurrence", "unknown"}
    assert seeded_codes("problem_type") == {"injury", "illness", "unknown"}
    assert seeded_codes("severity_category") == {
        "zero_days_medical_attention_only",
        "one_day",
        "two_to_three_days",
        "four_to_seven_days",
        "eight_to_twenty_eight_days",
        "greater_than_twenty_eight_days",
        "unknown_or_censored",
    }

    # Phase 3.1 analysis_views_v1 migration contract (text-level, offline).
    analysis_migration_path = (
        REPO_ROOT / "supabase" / "migrations" / f"{ANALYSIS_VIEWS_MIGRATION_VERSION}_analysis_views_v1.sql"
    )
    analysis_migration_text = analysis_migration_path.read_text()
    analysis_sql = "\n".join(
        line for line in analysis_migration_text.splitlines() if not line.lstrip().startswith("--")
    )
    analysis_sql_no_comments = re.sub(r"comment on (view|function)[^;]*;", "", analysis_sql)
    expected_analysis_views = [
        "injury_cohort_v1",
        "exposure_hours_v1",
        "headline_metrics_v1",
        "monthly_v1",
        "setting_split_v1",
        "body_locations_v1",
        "injury_types_v1",
        "severity_distribution_v1",
        "coverage_v1",
    ]
    for view_name in expected_analysis_views:
        assert f"create view analysis.{view_name}" in analysis_sql, view_name
        assert view_name.endswith(f"_{ANALYSIS_VIEW_VERSION_SUFFIX}")
    # security_invoker on all 9 views, matching public.dashboard_team_metrics.
    assert analysis_sql.count("security_invoker = true") == len(expected_analysis_views)
    assert "create function analysis.rate_per_1000_v1" in analysis_sql
    # Curated-only read rule: analysis views must never read processing/
    # ingestion/audit schemas, and from reporting only the teams dimension.
    for forbidden_schema in ("processing.", "ingestion.", "audit."):
        assert forbidden_schema not in analysis_sql_no_comments, forbidden_schema
    assert set(re.findall(r"reporting\.(\w+)", analysis_sql_no_comments)) <= {"teams"}
    # Formula-once rule: the /1000h rate formula exists exactly once, in
    # analysis.rate_per_1000_v1 (comment-on statements stripped above).
    assert analysis_sql_no_comments.count("* 1000") == 1

    # Phase 3.5 migration contracts (text-level, offline).
    cohort_signal_columns_path = (
        REPO_ROOT / "supabase" / "migrations" / "20260710110000_cohort_signal_columns.sql"
    )
    cohort_signal_columns_text = cohort_signal_columns_path.read_text()
    assert "add column received_in_team_status text" in cohort_signal_columns_text
    assert "add column urc_match_scope text" in cohort_signal_columns_text
    assert "'own_team', 'other_team', 'club', 'missing'" in cohort_signal_columns_text
    assert "'urc', 'non_urc_marker', 'training', 'unknown'" in cohort_signal_columns_text
    assert "cohort_signal_derivation" in cohort_signal_columns_text

    cohort_amendment_path = (
        REPO_ROOT / "supabase" / "migrations" / "20260710120000_injury_cohort_v1_amendment.sql"
    )
    cohort_amendment_text = cohort_amendment_path.read_text()
    assert "create or replace view analysis.injury_cohort_v1" in cohort_amendment_text
    # NULL-safe: both new filters must let a NULL signal pass unconditionally
    # (a team not yet reprocessed keeps its old cohort output byte-identical).
    assert "i.received_in_team_status is null or i.received_in_team_status not in ('other_team', 'club')" in cohort_amendment_text
    assert "i.urc_match_scope is null or i.urc_match_scope <> 'non_urc_marker'" in cohort_amendment_text
    # Every pre-existing output column must still be present, in order,
    # ahead of the two new trailing columns -- create-or-replace on a view
    # can append columns but never reorder or drop one.
    original_injury_cohort_columns = [
        "i.id as injury_id",
        "i.team_key",
        "i.season",
        "i.source_row_id",
        "i.record_version_id",
        "i.curated_build_id",
        "i.player_uid",
        "i.injury_uid",
        "i.date_injured",
        "i.days_injured",
        "as days_lost",
        "as is_time_loss",
        "i.is_closed",
        "i.activity_context",
        "as setting_label",
        "i.contact_context",
        "i.recurrence_status",
        "as severity_category",
        "as severity_label",
        "as body_location",
        "as body_location_label",
        "as injury_type",
        "as injury_type_label",
        "i.problem_type",
        "i.eligibility_status",
        "w.coverage_start",
        "w.coverage_end",
    ]
    positions = [cohort_amendment_text.index(marker) for marker in original_injury_cohort_columns]
    assert positions == sorted(positions), "amendment must preserve the original column order"
    assert cohort_amendment_text.index("i.received_in_team_status") > positions[-1]
    assert cohort_amendment_text.index("i.urc_match_scope") > positions[-1]

    # Phase 3.2 parity harness: coercion, render, and diff round-trip.
    assert as_number("12") == 12 and isinstance(as_number("12"), int)
    assert as_number("12.5") == 12.5
    assert as_number(12.0) == 12 and isinstance(as_number(12.0), int)
    assert as_number(None) is None and as_number("") is None
    assert parity_values_equal(16, 16.0)
    assert parity_values_equal(None, None)
    assert not parity_values_equal(None, 0)
    assert not parity_values_equal(1, 2)
    assert parity_values_equal("match", "match")
    assert dashboard_file_for_gate(
        argparse.Namespace(dashboard_file=""), "munster", "2025-26"
    ) == REPO_ROOT / "content" / "reporting" / "munster_dashboard_2025-26.json"
    assert dashboard_file_for_gate(
        argparse.Namespace(dashboard_file="data/reporting/first_release_preflight.json"),
        "munster",
        "2024-25",
    ) == REPO_ROOT / "data" / "reporting" / "first_release_preflight.json"
    assert dashboard_file_for_gate(
        argparse.Namespace(dashboard_file="/tmp/first_release_preflight.json"),
        "munster",
        "2024-25",
    ) == Path("/tmp/first_release_preflight.json").resolve()

    parity_headline_row = {
        "recorded_injuries": "3",
        "time_loss_injuries": "2",
        "days_lost_total": "30",
        "mean_severity_days": "15",
        "median_severity_days": 15.0,
        "exposure_hours": "100",
        "incidence_per_1000h": "20",
        "burden_per_1000h": "300",
    }
    parity_group_row = {
        "label": "Thigh",
        "time_loss_injuries": "2",
        "days_lost": "30",
        "incidence_per_1000h": "20",
        "burden_per_1000h": "300",
        "mean_severity_days": "15",
        "rank": "1",
    }
    parity_rendered = render_analysis_dashboard_sections(
        headline_row=parity_headline_row,
        setting_rows=[
            {"label": "training", "time_loss_injuries": "1", "days_lost": "5", "mean_severity_days": "5"},
            {"label": "match", "time_loss_injuries": "1", "days_lost": "25", "mean_severity_days": "25"},
        ],
        monthly_rows=[
            {
                "month_start": "2024-10-01",
                "month_label": "Oct 2024",
                "exposure_hours": "40",
                "distance_km": "20",
                "time_loss_injuries": "1",
                "days_lost": "25",
                "incidence_per_1000h": "25",
                "burden_per_1000h": "625",
                }
            ,
            {
                "month_start": "2024-09-01",
                "month_label": "Sep 2024",
                "exposure_hours": "60",
                "distance_km": "30",
                "time_loss_injuries": "1",
                "days_lost": "5",
                "incidence_per_1000h": None,
                "burden_per_1000h": None,
            },
        ],
        body_location_rows=[parity_group_row],
        injury_type_rows=[{**parity_group_row, "label": "Muscle injury"}],
        severity_rows=[
            {
                "key": "eight_to_twenty_eight_days",
                "label": "8-28 days",
                "recorded_injuries": "2",
                "time_loss_injuries": "2",
                "days_lost": "30",
                "band_order": "4",
            },
            {
                "key": "zero_days_medical_attention_only",
                "label": "Medical attention",
                "recorded_injuries": "1",
                "time_loss_injuries": "0",
                "days_lost": "0",
                "band_order": "0",
            },
        ],
        coverage_row={
            "exposure_rows": "10",
            "exposed_players": "5",
            "weeks": "0",
            "exposure_periods": "10",
            "exposure_grain": "session",
            "hours": "100",
            "distance_km": "50",
            "included_exposure_status": "included",
            "scope_status_counts": {"in_scope_explicit": 10},
        },
    )
    assert [item["key"] for item in parity_rendered["headline"]] == [
        "recorded_injuries",
        "time_loss_injuries",
        "incidence_per_1000h",
        "severity_mean_days",
        "severity_median_days",
        "burden_per_1000h",
    ]
    assert parity_rendered["headline"][0]["value"] == 3
    assert parity_rendered["headline"][2]["value"] == 20.0
    assert parity_rendered["headline"][2]["numerator"] == 2
    assert parity_rendered["headline"][2]["denominator"] == 100.0
    # Months re-sorted chronologically; setting split sorted like
    # build_group_rows (-time_loss, -days_lost, label).
    assert [row["month"] for row in parity_rendered["monthly"]] == ["Sep 2024", "Oct 2024"]
    assert parity_rendered["monthly"][0]["incidence_per_1000h"] is None
    assert [row["label"] for row in parity_rendered["setting_split"]] == ["match", "training"]
    # Severity bands re-sorted into the dashboard's fixed band order.
    assert [row["key"] for row in parity_rendered["severity_distribution"]] == [
        "zero_days_medical_attention_only",
        "eight_to_twenty_eight_days",
    ]
    assert parity_rendered["coverage"]["scope_status_counts"] == {"in_scope_explicit": 10}
    # The view renderer remains curated-only. verify-analysis-parity adds the
    # release-time evidence block separately, then treats it like every other
    # field: an exact match passes and any changed flag/count fails.
    assert "injury_cohort_filters" not in parity_rendered["coverage"]
    parity_reference = json.loads(json.dumps(parity_rendered))
    parity_reference["coverage"]["injury_cohort_filters"] = {
        "injured_in_team_applied": True,
        "explicit_non_urc_match_type_applied": True,
        "exclusion_reason_counts": {"explicit_non_urc_match_type": 2},
    }

    parity_self_diff = diff_dashboard_sections(
        json.loads(json.dumps(parity_reference)), parity_reference
    )
    assert parity_self_diff and all(row["status"] == "PASS" for row in parity_self_diff)
    parity_mutated = json.loads(json.dumps(parity_reference))
    parity_mutated["headline"][1]["value"] = 99
    parity_mutated["coverage"].pop("exposure_grain")
    parity_mutated["coverage"]["scope_status"] = "scope_unknown_included"
    parity_mutated["coverage"]["injury_cohort_filters"]["injured_in_team_applied"] = False
    parity_diffs = [
        row for row in diff_dashboard_sections(parity_mutated, parity_reference) if row["status"] == "DIFF"
    ]
    assert {(row["path"], row["kind"]) for row in parity_diffs} == {
        ("headline[1].value", "value_mismatch"),
        ("coverage.exposure_grain", "extra_in_rendered"),
        ("coverage.injury_cohort_filters.injured_in_team_applied", "value_mismatch"),
        ("coverage.scope_status", "missing_in_rendered"),
    }, parity_diffs

    # Release-time cohort-filter evidence: repeated rerun events for the same
    # source row/reason count once; audit overrides only its own reason while
    # current curated evidence fills other reasons in a hybrid payload.
    cohort_rerun_rows = [
        {"source_row_id": "row-1", "reason_code": "received_or_injured_in_other_team"},
        {"source_row_id": "row-1", "reason_code": "received_or_injured_in_other_team"},
        {"source_row_id": "row-2", "reason_code": "received_or_injured_in_other_team"},
        {"source_row_id": "row-1", "reason_code": "explicit_non_urc_match_type"},
        {"source_row_id": "row-1", "reason_code": "not_a_dashboard_reason"},
    ]
    cohort_merged, cohort_audit = combine_cohort_filter_reason_counts(
        cohort_rerun_rows,
        {
            "received_or_injured_in_other_team": 99,
            "non_injury_problem_type": 3,
        },
    )
    assert cohort_audit == {
        "explicit_non_urc_match_type": 1,
        "received_or_injured_in_other_team": 2,
    }
    assert cohort_merged == {
        "explicit_non_urc_match_type": 1,
        "non_injury_problem_type": 3,
        "received_or_injured_in_other_team": 2,
    }
    assert combine_cohort_filter_reason_counts(
        [], {"adjudicated_duplicate": 1}
    ) == ({"adjudicated_duplicate": 1}, {})
    cohort_lineage_contract = "\n".join(
        value for value in release_cohort_filter_flags.__code__.co_consts if isinstance(value, str)
    )
    assert "join processing.record_versions current_rv on current_rv.id = i.record_version_id" in cohort_lineage_contract
    assert "historical_rv.version_number <= current_rv.version_number" in cohort_lineage_contract
    assert "historical_rv.eligibility_status = current_rv.eligibility_status" not in cohort_lineage_contract
    assert "current_rv.eligibility_status in ('excluded_from_analysis', 'excluded_duplicate_adjudicated')" in cohort_lineage_contract
    assert "historical_rv.eligibility_status in ('excluded_from_analysis', 'excluded_duplicate_adjudicated')" in cohort_lineage_contract
    assert "historical_ev.new_value #>> '{}' = historical_rv.eligibility_status" in cohort_lineage_contract
    assert "order by historical_rv.version_number desc" in cohort_lineage_contract
    assert "ev.step_run_id = applicable.step_run_id" in cohort_lineage_contract
    assert "ev.new_value #>> '{}' = applicable.eligibility_status" in cohort_lineage_contract

    # Phase 4 release/export helpers (pure functions only; DB-backed release
    # gates are exercised live, not here).
    assert slug_key("Lower Leg") == "lower_leg"
    assert slug_key("2-3 days") == "2_3_days"
    assert slug_key("") == "unknown"
    assert strip_none_keys({"a": 1, "b": None, "c": 2}, {"b"}) == {"a": 1, "c": 2}
    assert strip_none_keys({"a": 1, "b": None}, {"a"}) == {"a": 1, "b": None}

    release_table_rows_fixture = render_release_table_rows(
        {
            "headline": [
                {"key": "recorded_injuries", "label": "Recorded", "value": 10, "unit": "injuries", "formula": "f"},
                {
                    "key": "incidence_per_1000h", "label": "Incidence", "value": 5.0, "unit": "per 1,000h",
                    "numerator": 3, "denominator": 100.0, "formula": "f",
                },
            ],
            "setting_split": [{"label": "match", "time_loss_injuries": 2, "days_lost": 4, "mean_severity_days": 2.0}],
            "monthly": [
                {
                    "month": "Jul 2024", "exposure_hours": 10.0, "distance_km": 1.0, "time_loss_injuries": 1,
                    "days_lost": 2, "incidence_per_1000h": 1.0, "burden_per_1000h": 2.0,
                }
            ],
            "body_locations": [
                {
                    "label": "Knee", "time_loss_injuries": 1, "days_lost": 2, "incidence_per_1000h": 1.0,
                    "burden_per_1000h": 2.0, "mean_severity_days": 2.0,
                }
            ],
            "injury_types": [],
            "severity_distribution": [
                {"key": "one_day", "label": "1 day", "recorded_injuries": 1, "time_loss_injuries": 1, "days_lost": 1}
            ],
        },
        [{"month_start_text": "2024-07-01"}],
    )
    assert len(release_table_rows_fixture) == 6
    headline_fixture_rows = [row for row in release_table_rows_fixture if row["section"] == "headline"]
    assert headline_fixture_rows[0]["row_key"] == "recorded_injuries"
    assert headline_fixture_rows[0]["ordinal"] == 0
    assert headline_fixture_rows[1]["numerator"] == 3
    monthly_fixture_rows = [row for row in release_table_rows_fixture if row["section"] == "monthly"]
    assert monthly_fixture_rows[0]["row_key"] == "2024-07-01"
    setting_fixture_rows = [row for row in release_table_rows_fixture if row["section"] == "setting_split"]
    assert setting_fixture_rows[0]["row_key"] == "match"

    preflight_context_fixture = {
        "generated_at": "2026-07-11T10:00:00Z",
        "team_display_name": "Example RFC",
        "season": "2024-25",
        "analysis_window_start": "2024-07-01",
        "analysis_window_end": "2025-06-30",
        "analysis_window_basis": "fixture coverage",
        "method": ["fixture method"],
        "coverage": {"hours": 100},
        "prior_season": {"status": "not_loaded_in_v2"},
        "limitations": ["fixture limitation"],
        "source_files": {"must_not": "leak"},
    }
    preflight_dashboard_fixture = assemble_release_dashboard(
        preflight_context_fixture,
        release_table_rows_fixture,
    )
    assert preflight_dashboard_fixture["generated_at"] == "2026-07-11T10:00:00Z"
    assert preflight_dashboard_fixture["headline"][1]["numerator"] == 3
    assert "source_files" not in preflight_dashboard_fixture
    with tempfile.TemporaryDirectory() as atomic_tmp:
        atomic_target = Path(atomic_tmp) / "dashboard.json"
        atomic_target.write_text('{"version":"previous"}\n')
        write_json_atomic(atomic_target, {"version": "next"})
        assert json.loads(atomic_target.read_text()) == {"version": "next"}
        stable_bytes = atomic_target.read_bytes()
        try:
            write_json_atomic(atomic_target, {"not_json": object()})
            raise AssertionError("atomic JSON write should propagate serialization failure")
        except TypeError:
            pass
        assert atomic_target.read_bytes() == stable_bytes
        assert not list(atomic_target.parent.glob(f".{atomic_target.name}.*.tmp"))
    # Guard the release/preflight assembly seam: an earlier broad patch
    # accidentally placed these locals in process_exposure, which compiles
    # but fails only when that live path executes.
    assert "release_parameters" not in process_exposure.__code__.co_varnames
    assert {
        "is_first_release", "reviewed_preflight_path", "reviewed_candidate",
        "preflight_reviewer", "previous_dashboard", "previous_dashboard_sha256",
        "historical_diff", "release_parameters", "draft_created", "release_stage",
    } <= set(release.__code__.co_varnames)
    release_source_contract = "\n".join(
        value for value in release.__code__.co_consts if isinstance(value, str)
    )
    assert "or (r.status = 'retired' and pr.status = 'succeeded')" in release_source_contract
    assert "'draft', id, null" in release_source_contract
    assert "previous_dashboard_sha256" in release_source_contract
    assert "release_failure_cleanup_statement" in release.__code__.co_names
    released_dashboard_fixture = json.loads(json.dumps(preflight_dashboard_fixture))
    released_dashboard_fixture["generated_at"] = "2026-07-11T10:05:00Z"
    allowed_preflight, blocked_preflight = classify_preflight_release_diffs(
        diff_json_documents(preflight_dashboard_fixture, released_dashboard_fixture)
    )
    assert [diff["path"] for diff in allowed_preflight] == ["generated_at"]
    assert not blocked_preflight
    released_dashboard_fixture["method"] = ["changed after review"]
    _, blocked_preflight_changed = classify_preflight_release_diffs(
        diff_json_documents(preflight_dashboard_fixture, released_dashboard_fixture)
    )
    assert [diff["path"] for diff in blocked_preflight_changed] == ["method[0]"]
    assert "data/" in (REPO_ROOT / ".gitignore").read_text().splitlines()

    # Phase 4.3 old-vs-new dashboard JSON diff tool.
    old_dashboard_fixture = {
        "generated_at": "2026-01-01T00:00:00Z",
        "team": "Glasgow Warriors",
        "coverage": {"scope_status": "in_scope_explicit", "hours": 100},
        "source_files": {"injury": "x.csv"},
        "headline": [{"key": "recorded_injuries", "value": 10}],
    }
    new_dashboard_fixture = {
        "generated_at": "2026-07-10T00:00:00Z",
        "team": "Glasgow Warriors",
        "coverage": {
            "scope_status_counts": {"in_scope_explicit": 100}, "hours": 100,
            "exposure_grain": "session", "injury_cohort_filters": {"injured_in_team_applied": True},
        },
        "headline": [{"key": "recorded_injuries", "value": 10}],
    }
    fixture_diffs = diff_json_documents(old_dashboard_fixture, new_dashboard_fixture)
    fixture_allowed, fixture_blocked = classify_historical_release_diffs(fixture_diffs)
    assert fixture_allowed
    assert not fixture_blocked, fixture_blocked
    new_dashboard_fixture_bad = json.loads(json.dumps(new_dashboard_fixture))
    new_dashboard_fixture_bad["headline"][0]["value"] = 11
    fixture_diffs_bad = diff_json_documents(old_dashboard_fixture, new_dashboard_fixture_bad)
    _, fixture_blocked_bad = classify_historical_release_diffs(fixture_diffs_bad)
    assert [d["path"] for d in fixture_blocked_bad] == ["headline[0].value"], fixture_blocked_bad
    new_dashboard_fixture_renamed = json.loads(json.dumps(new_dashboard_fixture))
    new_dashboard_fixture_renamed["team"] = "Glasgow"
    fixture_diffs_renamed = diff_json_documents(old_dashboard_fixture, new_dashboard_fixture_renamed)
    _, fixture_blocked_renamed = classify_historical_release_diffs(fixture_diffs_renamed)
    assert [d["path"] for d in fixture_blocked_renamed] == ["team"], fixture_blocked_renamed

    # Promotion is one SQL transaction: approve the pinned draft, reconstruct
    # the complete public JSON through the consumer view, compare it with the
    # verified candidate parameter, and only then close the audit run/step.
    promotion_sql, promotion_params = release_promotion_statement(
        "example-2024-25-hash-attempt",
        "00000000-0000-0000-0000-000000000001",
        "example",
        "2024-25",
        preflight_dashboard_fixture,
        "00000000-0000-0000-0000-000000000000",
    )
    assert "from reporting.teams t" in promotion_sql
    assert "for update" in promotion_sql
    assert "accepted predecessor changed" in promotion_sql
    assert "set status = 'approved', approved_at = clock_timestamp()" in promotion_sql
    assert "set status = 'retired'" in promotion_sql
    assert "from reporting.latest_team_dashboard d" in promotion_sql
    assert "where d.release_id = target_release_id" in promotion_sql
    assert "consumer_dashboard is distinct from" in promotion_sql
    assert "consumer view payload differs from verified draft candidate" in promotion_sql
    assert promotion_sql.index("set status = 'approved'") < promotion_sql.index(
        "from reporting.latest_team_dashboard d"
    ) < promotion_sql.index("set status = 'succeeded'")
    assert preflight_dashboard_fixture in promotion_params

    mocked_promotion_calls: list[tuple[str, list[object]]] = []

    def successful_promotion_runner(sql_text: str, values: list[object]) -> None:
        mocked_promotion_calls.append((sql_text, values))

    execute_release_promotion(
        "example-2024-25-hash-attempt",
        "00000000-0000-0000-0000-000000000001",
        "example",
        "2024-25",
        preflight_dashboard_fixture,
        "00000000-0000-0000-0000-000000000000",
        runner=successful_promotion_runner,
    )
    assert len(mocked_promotion_calls) == 1
    assert mocked_promotion_calls[0][0] == promotion_sql
    assert mocked_promotion_calls[0][1] == promotion_params

    def failing_promotion_runner(_sql_text: str, _values: list[object]) -> None:
        raise RuntimeError("mocked transactional consumer mismatch")

    try:
        execute_release_promotion(
            "example-2024-25-hash-attempt",
            "00000000-0000-0000-0000-000000000001",
            "example",
            "2024-25",
            preflight_dashboard_fixture,
            "00000000-0000-0000-0000-000000000000",
            runner=failing_promotion_runner,
        )
        raise AssertionError("promotion runner failure must propagate to release cleanup")
    except RuntimeError as exc:
        assert str(exc) == "mocked transactional consumer mismatch"

    cleanup_sql, cleanup_params = release_failure_cleanup_statement(
        "example-2024-25-hash-attempt",
        "example",
        "2024-25",
        "00000000-0000-0000-0000-000000000000",
        "approved",
        "export",
    )
    assert "from reporting.teams t" in cleanup_sql
    assert "latest_approved_id = failed_release_id" in cleanup_sql
    assert "set status = 'retired'" in cleanup_sql
    assert "set status = 'approved'" in cleanup_sql
    assert "set status = 'failed', ended_at = now()" in cleanup_sql
    assert "00000000-0000-0000-0000-000000000000" in cleanup_params
    assert "approved" in cleanup_params
    assert {"failure_stage": "export"} in cleanup_params

    if os.environ.get("SUPABASE_DB_URL"):
        run_sql(protected_alias_scan_sql("self-check"))
        print("self-check: live protected-alias scan passed (0 hits)")
    else:
        print(
            "self-check: SUPABASE_DB_URL not set; skipping live protected-alias "
            "scan (this check is read-only, but requires DB connectivity)"
        )

    print("self-check passed")


def main() -> None:
    from pipeline.corrections import (
        capture_served_baseline,
        correction_apply,
        correction_batch_apply,
        correction_batch_propose,
        correction_batch_release,
        correction_propose,
        correction_release,
        correction_rollback,
        verify_served_baseline,
    )

    parser = argparse.ArgumentParser(prog="pipeline")
    subcommands = parser.add_subparsers(required=True)

    ingest_parser = subcommands.add_parser("ingest")
    ingest_parser.add_argument("--team", required=True)
    ingest_parser.add_argument("--season", required=True)
    ingest_parser.add_argument("--file", required=True)
    ingest_parser.add_argument("--manifest", required=True)
    ingest_parser.add_argument("--signed-root-manifest", default="")
    ingest_parser.add_argument("--exclude-source-fields", default="")
    ingest_parser.add_argument("--redact-manifest-keys", default="")
    ingest_parser.add_argument("--redact-source-values", default="")
    ingest_parser.set_defaults(func=ingest)

    prepare_parser = subcommands.add_parser("prepare-intake")
    prepare_parser.add_argument("--team", required=True)
    prepare_parser.add_argument("--season", required=True)
    prepare_parser.add_argument("--file", required=True)
    prepare_parser.add_argument("--source-file", required=True)
    prepare_parser.add_argument("--output", required=True)
    prepare_parser.add_argument("--manifest")
    prepare_parser.add_argument("--source-sheet", default="file")
    prepare_parser.add_argument("--player-id-column", default="PlayerID")
    prepare_parser.set_defaults(func=prepare_intake)

    export_parser = subcommands.add_parser("export-xlsx-sheet")
    export_parser.add_argument("--file", required=True)
    export_parser.add_argument("--sheet", default="Standardized Data")
    export_parser.add_argument("--output", required=True)
    export_parser.add_argument("--blank-column", action="append", default=[])
    export_parser.set_defaults(func=export_xlsx_sheet)

    injury_adapter_parser = subcommands.add_parser("adapt-injury-intake")
    injury_adapter_parser.add_argument("--team", required=True)
    injury_adapter_parser.add_argument("--season", required=True)
    injury_adapter_parser.add_argument("--file", required=True)
    injury_adapter_parser.add_argument("--sheet", default="Standardized Data")
    injury_adapter_parser.add_argument("--output", required=True)
    injury_adapter_parser.add_argument("--audit-output", required=True)
    injury_adapter_parser.add_argument("--fixture-file", default="")
    injury_adapter_parser.add_argument(
        "--date-order", choices=["month-first", "day-first"], default="day-first"
    )
    injury_adapter_parser.set_defaults(func=adapt_injury_intake)

    add_exposure_cli_parsers(subcommands)

    fixture_exposure_parser = subcommands.add_parser("build-fixture-exposure")
    fixture_exposure_parser.add_argument(
        "--file",
        default="/Users/abdelbabiker/Downloads/Injury & Exposure Data Master Sheet - Analysis - Fixtures.csv",
    )
    fixture_exposure_parser.add_argument(
        "--preserved-output",
        default="data/intake/2024-25/fixtures/urc_fixtures_2024_25.downloaded.csv",
    )
    fixture_exposure_parser.add_argument(
        "--fixture-output",
        default="data/intake/2024-25/fixtures/urc_fixtures_2024_25.corrected.csv",
    )
    fixture_exposure_parser.add_argument(
        "--exposure-output",
        default="data/reporting/urc_team_exposure_2024-25.csv",
    )
    fixture_exposure_parser.add_argument(
        "--qc-output",
        default="data/reporting/urc_fixture_exposure_qc_2024-25.json",
    )
    fixture_exposure_parser.add_argument(
        "--total-exposure-file",
        default="/Users/abdelbabiker/Downloads/Exposure_metrics___by_team.csv",
    )
    fixture_exposure_parser.add_argument("--total-team-column", default="Team")
    fixture_exposure_parser.add_argument("--total-hours-column", default="hours_total")
    fixture_exposure_parser.add_argument("--player-hours-per-team-match", type=float, default=20.0)
    fixture_exposure_parser.add_argument("--date-order", choices=["month-first", "day-first"], default="day-first")
    fixture_exposure_parser.set_defaults(func=build_fixture_exposure)

    qa_parser = subcommands.add_parser("qa-intake")
    qa_parser.add_argument("--file", required=True)
    qa_parser.add_argument("--output")
    qa_parser.add_argument("--window-start", default="2024-07-01")
    qa_parser.add_argument("--window-end", default="2025-06-30")
    qa_parser.set_defaults(func=qa_intake)

    process_parser = subcommands.add_parser("process-intake")
    process_parser.add_argument("--team", required=True)
    process_parser.add_argument("--season", required=True)
    process_parser.add_argument("--file", required=True)
    process_parser.add_argument("--window-start", default="2024-07-01")
    process_parser.add_argument("--window-end", default="2025-06-30")
    process_parser.add_argument("--step-name", default="intake_first_pass")
    process_parser.add_argument("--step-version", default=INJURY_PROCESSING_RULE_VERSION)
    process_parser.add_argument("--version-number", type=int, default=1)
    process_parser.add_argument("--analysis-audit-file", default="")
    process_parser.add_argument(
        "--injury-eligibility-bridge-file",
        default="",
        help=(
            "2025-26 only: protected checksum-bound status bridge for "
            "season-attributed undated injuries"
        ),
    )
    process_parser.add_argument("--manifest", default="")
    process_parser.add_argument(
        "--registered-source-file-sha256",
        default="",
        help="bind an approved corrected processing artifact to an already registered immutable source",
    )
    process_parser.add_argument("--adapter-qc-file", default="")
    process_parser.set_defaults(func=process_intake)

    trace_parser = subcommands.add_parser("trace-row")
    trace_parser.add_argument("--file", required=True)
    trace_parser.add_argument("--row-number", type=int, required=True)
    trace_parser.add_argument("--include-source", action="store_true")
    trace_parser.set_defaults(func=trace_row)

    run_parser = subcommands.add_parser("run")
    run_parser.add_argument("--team", required=True)
    run_parser.add_argument("--season", required=True)
    run_parser.add_argument("--step", required=True)
    run_parser.set_defaults(func=run_step)

    release_parser = subcommands.add_parser("release")
    release_parser.add_argument("--team", required=True)
    release_parser.add_argument("--season", default="2024-25")
    release_parser.add_argument("--output", default="")
    release_parser.add_argument(
        "--preflight",
        action="store_true",
        help="render the release candidate through all read-only gates and exit before any DB write",
    )
    release_parser.add_argument(
        "--preflight-file",
        default="",
        help="reviewed preflight candidate required before a team's first release",
    )
    release_parser.add_argument(
        "--preflight-reviewer",
        default="",
        help="reviewer name required whenever --preflight-file is supplied",
    )
    release_parser.add_argument(
        "--previous-dashboard-file",
        default="",
        help="latest accepted full-dashboard snapshot required for every re-release",
    )
    release_parser.add_argument(
        "--restatement-file",
        default="",
        help="exact checksummed approval envelope for a re-release with non-whitelisted corrected values",
    )
    release_parser.set_defaults(func=release)

    league_release_parser = subcommands.add_parser("release-league")
    league_release_parser.add_argument("--season", default="2024-25")
    league_release_parser.add_argument("--output", default="")
    league_release_parser.add_argument(
        "--plan",
        action="store_true",
        help=(
            "print the approval-separated local, read-only, and live-write "
            "workflow without reading the database"
        ),
    )
    league_release_parser.add_argument(
        "--snapshot-current", action="store_true",
        help="write the exact current approved immutable bundle outside content/reporting",
    )
    league_release_parser.add_argument(
        "--previous-bundle-file", default="",
        help="exact current approved bundle snapshot required for every bundle re-release",
    )
    league_release_parser.add_argument(
        "--rollback-of-release-id", default="",
        help=(
            "2025-26 V6 only: create an append-only successor from the exact "
            "retained immutable bundle identified by this release UUID"
        ),
    )
    league_release_parser.add_argument(
        "--analysis-version", default="", choices=["v2", "v3", "v4", "v5", "v6"],
        help=(
            "analytical candidate family; V3 requires the accepted season-bound cohort "
            "and V4 requires the accepted lineage cohort; V5 requires the accepted "
            "analysis-window cohort"
        ),
    )
    league_release_parser.add_argument(
        "--classification-view-version", default="",
        choices=[
            "v2",
            "reporting_classification_2026-07-20_v1",
            "reporting_classification_2026-07-22_v2",
        ],
    )
    league_release_parser.add_argument(
        "--cohort-view-version", default="",
        choices=[
            "v2",
            "season_bound_2026-07-20_v1",
            "lineage_2024-25_2026-07-24_v1",
            ANALYSIS_WINDOW_V5_COHORT_VIEW_VERSION,
            "analysis_window_2025-26_2026-08-15_v1",
        ],
    )
    league_release_parser.add_argument(
        "--preflight",
        action="store_true",
        help="render the exact 16-team league release candidate without a database write",
    )
    league_release_parser.add_argument(
        "--preflight-file",
        default="",
        help="reviewed league candidate required before first promotion",
    )
    league_release_parser.add_argument(
        "--preflight-reviewer",
        default="",
        help="named reviewer required with --preflight-file",
    )
    league_release_parser.add_argument(
        "--allow-legacy-preflight-without-manifest",
        action="store_true",
        help=(
            "explicit historical-candidate escape hatch; new preflights must "
            "use their generated manifest"
        ),
    )
    league_release_parser.set_defaults(func=release_league, command_name="release-league")

    team_parity_parser = subcommands.add_parser("export-team-dashboards")
    team_parity_parser.add_argument("--season", required=True)
    team_parity_parser.set_defaults(func=export_team_dashboard_parity_json)

    v6_finalize_parser = subcommands.add_parser("finalize-v6-league-release-local")
    v6_finalize_parser.add_argument("--release-id", required=True)
    v6_finalize_parser.add_argument("--release-label", required=True)
    v6_finalize_parser.add_argument("--preflight-file", required=True)
    v6_finalize_parser.set_defaults(func=finalize_v6_league_release_local)

    diff_dashboard_json_parser = subcommands.add_parser("diff-dashboard-json")
    diff_dashboard_json_parser.add_argument("--old", required=True)
    diff_dashboard_json_parser.add_argument("--new", required=True)
    diff_dashboard_json_parser.add_argument(
        "--preflight-release",
        action="store_true",
        help="allow only a generated_at change between reviewed preflight and release export",
    )
    diff_dashboard_json_parser.set_defaults(func=diff_dashboard_json)

    adjudicate_duplicate_parser = subcommands.add_parser("adjudicate-duplicate-exclusion")
    adjudicate_duplicate_parser.add_argument("--team", required=True)
    adjudicate_duplicate_parser.add_argument("--season", required=True)
    adjudicate_duplicate_parser.add_argument("--file", required=True)
    adjudicate_duplicate_parser.add_argument("--row-number", type=int, required=True)
    adjudicate_duplicate_parser.add_argument("--duplicate-of", type=int, required=True)
    adjudicate_duplicate_parser.add_argument("--rationale", required=True)
    adjudicate_duplicate_parser.add_argument("--reviewer", required=True)
    adjudicate_duplicate_parser.add_argument("--step-version", default=INJURY_PROCESSING_RULE_VERSION)
    adjudicate_duplicate_parser.add_argument("--version-number", type=int, default=2)
    adjudicate_duplicate_parser.set_defaults(func=adjudicate_duplicate_exclusion)

    adjudication_batch_parser = subcommands.add_parser("apply-adjudication-batch")
    adjudication_batch_parser.add_argument("--file", required=True)
    adjudication_batch_parser.add_argument("--workbook", required=True)
    adjudication_batch_parser.add_argument("--evidence-file", required=True)
    adjudication_batch_parser.add_argument("--plan", action="store_true")
    adjudication_batch_parser.set_defaults(func=apply_adjudication_batch)

    osiics_parser = subcommands.add_parser("apply-osiics-mapping-adjudication")
    osiics_parser.add_argument(
        "--evidence-file", default="docs/evidence/osiics_exact_mapping_2024-25.json"
    )
    osiics_parser.add_argument(
        "--row-ledger", default="data/reporting/osiics_exact_mapping_2024-25_rows.json"
    )
    osiics_parser.add_argument(
        "--workbook", default="data/reference/osiics/osiics-v15-8bfeab66.xlsx"
    )
    osiics_parser.add_argument(
        "--migration-file",
        default="supabase/migrations/20260722140000_osiics_source_body_pathology_mapping.sql",
    )
    osiics_parser.add_argument("--expected-migration-sha256", required=True)
    osiics_parser.add_argument("--plan", action="store_true")
    osiics_parser.set_defaults(func=apply_osiics_mapping_adjudication)

    reapply_adjudications_parser = subcommands.add_parser("reapply-adjudications")
    reapply_adjudications_parser.add_argument("--team", required=True)
    reapply_adjudications_parser.add_argument("--season", default="2024-25")
    reapply_adjudications_parser.add_argument("--step-version", default=INJURY_PROCESSING_RULE_VERSION)
    reapply_adjudications_parser.set_defaults(func=reapply_adjudications)

    backfill_adjudications_parser = subcommands.add_parser("backfill-standing-adjudications")
    backfill_adjudications_parser.add_argument("--season", default="2024-25")
    backfill_adjudications_parser.add_argument("--reviewer", default="Abdel Babiker")
    backfill_adjudications_parser.add_argument("--step-version", default="backfill_2026-07-10_v1")
    backfill_adjudications_parser.add_argument("--plan", action="store_true")
    backfill_adjudications_parser.set_defaults(func=backfill_standing_adjudications)

    def add_team_dashboard_args(command: argparse.ArgumentParser) -> None:
        command.add_argument("--team", default="Munster")
        command.add_argument("--season", default="2024-25")
        command.add_argument(
            "--injury-file",
            default="data/intake/2024-25/munster/munster_filled_standardised_2024-25.csv",
        )
        command.add_argument(
            "--exposure-file",
            default="data/intake/2024-25/munster/munster_exposure_cleaned_2024-25.csv",
        )
        command.add_argument(
            "--output",
            default="data/reporting/munster_dashboard_2024-25.json",
        )
        command.add_argument("--exclude-injury-rows", default="")
        command.add_argument("--injured-in-team", default="")
        command.add_argument("--analysis-source-output", default="")
        command.add_argument("--standardisation-audit-output", default="")

    team_dashboard_parser = subcommands.add_parser(
        "build-team-dashboard",
        help="retired and disabled; use the audited release pipeline",
        description="Retired local CSV calculator. This command always refuses to run.",
    )
    add_team_dashboard_args(team_dashboard_parser)
    team_dashboard_parser.set_defaults(func=retired_legacy_dashboard_command)

    dashboard_parser = subcommands.add_parser(
        "build-munster-dashboard",
        help="retired and disabled; use the audited release pipeline",
        description="Retired local CSV calculator. This command always refuses to run.",
    )
    add_team_dashboard_args(dashboard_parser)
    dashboard_parser.set_defaults(func=retired_legacy_dashboard_command)

    redact_alias_parser = subcommands.add_parser("redact-protected-team-aliases")
    redact_alias_parser.add_argument("--scope", default="all", choices=["all"])
    redact_alias_parser.set_defaults(func=redact_protected_team_aliases)

    retire_releases_parser = subcommands.add_parser("retire-releases")
    retire_releases_parser.add_argument("--labels", required=True)
    retire_releases_parser.add_argument("--rationale", required=True)
    retire_releases_parser.add_argument("--reviewer", required=True)
    retire_releases_parser.set_defaults(func=retire_releases)

    load_fixtures_parser = subcommands.add_parser("load-curated-fixtures")
    load_fixtures_parser.add_argument("--season", default="2024-25")
    load_fixtures_parser.add_argument("--file", default=URC_FIXTURES_2024_25_CORRECTED_PATH)
    load_fixtures_parser.set_defaults(func=load_curated_fixtures)

    build_curated_parser = subcommands.add_parser("build-curated")
    build_curated_parser.add_argument("--team", required=True)
    build_curated_parser.add_argument("--season", default="2024-25")
    build_curated_parser.add_argument("--rebuild", action="store_true")
    build_curated_parser.set_defaults(func=build_curated)

    reconcile_curated_parser = subcommands.add_parser("reconcile-curated")
    reconcile_curated_parser.add_argument("--team", required=True)
    reconcile_curated_parser.add_argument("--season", default="2024-25")
    reconcile_curated_parser.add_argument(
        "--dashboard-file",
        default="",
        help="dashboard JSON to reconcile; defaults to content/reporting/<team_key>_dashboard_<season>.json",
    )
    reconcile_curated_parser.set_defaults(func=reconcile_curated)

    verify_parity_parser = subcommands.add_parser("verify-analysis-parity")
    verify_parity_parser.add_argument("--team", required=True)
    verify_parity_parser.add_argument("--season", default="2024-25")
    verify_parity_parser.add_argument(
        "--dashboard-file",
        default="",
        help="dashboard JSON to compare; defaults to content/reporting/<team_key>_dashboard_<season>.json",
    )
    verify_parity_parser.set_defaults(func=verify_analysis_parity)

    check_parser = subcommands.add_parser("self-check")
    check_parser.set_defaults(func=self_check)

    alias_parser = subcommands.add_parser("validate-alias-map")
    alias_parser.add_argument("--codebook", default=str(TEAM_ALIAS_CODEBOOK_PATH))
    alias_parser.set_defaults(func=validate_alias_map)

    capture_correction_baseline_parser = subcommands.add_parser("capture-served-baseline")
    capture_correction_baseline_parser.add_argument("--season", required=True)
    capture_correction_baseline_parser.add_argument("--output", required=True)
    capture_correction_baseline_parser.set_defaults(func=capture_served_baseline)

    verify_correction_baseline_parser = subcommands.add_parser("verify-served-baseline")
    verify_correction_baseline_parser.add_argument("--season", required=True)
    verify_correction_baseline_parser.add_argument("--baseline-file", required=True)
    verify_correction_baseline_parser.set_defaults(func=verify_served_baseline)

    correction_proposal_parser = subcommands.add_parser("correction-propose")
    correction_proposal_parser.add_argument("--season", required=True)
    correction_proposal_parser.add_argument("--source-row-id", required=True)
    correction_proposal_parser.add_argument(
        "--field-name",
        required=True,
        choices=(
            "eligibility",
            "days_injured",
            "body_location_code",
            "injury_type_code",
            "diagnosis_code",
        ),
    )
    correction_proposal_parser.add_argument("--expected-value", required=True)
    correction_proposal_parser.add_argument("--new-value", required=True)
    correction_proposal_parser.add_argument("--reason", required=True)
    correction_proposal_parser.add_argument("--evidence-file", required=True)
    correction_proposal_parser.add_argument("--operator", required=True)
    correction_proposal_parser.add_argument("--rule-version", required=True)
    correction_proposal_parser.add_argument(
        "--supersedes-correction-id", default=""
    )
    correction_proposal_parser.add_argument("--output", required=True)
    correction_proposal_parser.set_defaults(func=correction_propose)

    correction_apply_parser = subcommands.add_parser("correction-apply")
    correction_apply_parser.add_argument("--proposal-file", required=True)
    correction_apply_parser.add_argument("--evidence-file", default="")
    correction_apply_parser.add_argument("--reviewer", required=True)
    correction_apply_parser.set_defaults(func=correction_apply)

    correction_release_parser = subcommands.add_parser("correction-release")
    correction_release_parser.add_argument("--proposal-file", default="")
    correction_release_parser.add_argument("--preflight", action="store_true")
    correction_release_parser.add_argument("--output", default="")
    correction_release_parser.add_argument("--preflight-file", default="")
    correction_release_parser.add_argument("--reviewer", default="")
    correction_release_parser.add_argument("--release-label", default="")
    correction_release_parser.add_argument("--rollback-release-label", default="")
    correction_release_parser.add_argument("--rollback-reviewer", default="")
    correction_release_parser.add_argument("--rollback-reason", default="")
    correction_release_parser.add_argument("--rollback-evidence-file", default="")
    correction_release_parser.add_argument("--rollback-operator", default="")
    correction_release_parser.set_defaults(func=correction_release)

    correction_batch_proposal_parser = subcommands.add_parser(
        "correction-batch-propose"
    )
    correction_batch_proposal_parser.add_argument("--season", required=True)
    correction_batch_proposal_parser.add_argument("--manifest", required=True)
    correction_batch_proposal_parser.add_argument("--operator", required=True)
    correction_batch_proposal_parser.add_argument("--output", required=True)
    correction_batch_proposal_parser.set_defaults(func=correction_batch_propose)

    correction_batch_apply_parser = subcommands.add_parser(
        "correction-batch-apply"
    )
    correction_batch_apply_parser.add_argument("--proposal-file", required=True)
    correction_batch_apply_parser.add_argument("--reviewer", required=True)
    correction_batch_apply_parser.set_defaults(func=correction_batch_apply)

    correction_batch_release_parser = subcommands.add_parser(
        "correction-batch-release"
    )
    correction_batch_release_parser.add_argument("--proposal-file", default="")
    correction_batch_release_parser.add_argument("--preflight", action="store_true")
    correction_batch_release_parser.add_argument("--output", default="")
    correction_batch_release_parser.add_argument("--preflight-file", default="")
    correction_batch_release_parser.add_argument("--reviewer", default="")
    correction_batch_release_parser.add_argument("--release-label", default="")
    correction_batch_release_parser.add_argument("--rollback-release-label", default="")
    correction_batch_release_parser.add_argument("--rollback-reviewer", default="")
    correction_batch_release_parser.add_argument("--rollback-reason", default="")
    correction_batch_release_parser.add_argument("--rollback-evidence-file", default="")
    correction_batch_release_parser.add_argument("--rollback-operator", default="")
    correction_batch_release_parser.set_defaults(func=correction_batch_release)

    correction_rollback_parser = subcommands.add_parser("correction-rollback")
    correction_rollback_parser.add_argument("--release-label", required=True)
    correction_rollback_parser.add_argument("--reviewer", required=True)
    correction_rollback_parser.add_argument("--rollback-release-label", required=True)
    correction_rollback_parser.add_argument("--reason", required=True)
    correction_rollback_parser.add_argument("--evidence-file", required=True)
    correction_rollback_parser.add_argument("--operator", required=True)
    correction_rollback_parser.set_defaults(func=correction_rollback)

    args = parser.parse_args()
    if getattr(args, "command_name", "") == "release-league" and not any(
        clean_text(getattr(args, field, ""))
        for field in (
            "analysis_version", "classification_view_version", "cohort_view_version",
        )
    ):
        default_tuple = (
            YEAR2_2025_26_RELEASE_TUPLE
            if clean_text(args.season) == "2025-26"
            else ("v2", "v2", "v2")
        )
        (
            args.analysis_version,
            args.classification_view_version,
            args.cohort_view_version,
        ) = default_tuple
    args.func(args)


if __name__ == "__main__":
    main()
