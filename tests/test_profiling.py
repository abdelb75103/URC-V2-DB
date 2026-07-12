from __future__ import annotations

import json
import hashlib
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from pipeline.profiling import ProfileError, family_check, scan_plan, validate_package


class ProfilingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.workbook = self.root / "source.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "Injuries"
        sheet.append([
            "PlayerID", "Date Injured", "Occasion", "Description", "Days", "Problem Type",
            "Injury Status", "Fit Status", "Return Date", "Recurrence", "Contact", "Body Part",
            "Tissue Type",
        ])
        sheet.append(["private-player-001", "01/08/2024", "MATCH", "private clinical note", 5, "Injury", "Closed", "Fit", "06/08/2024", "First", "Contact", "Knee", "Ligament"])
        sheet.append(["private-player-001", "01/08/2024", "MATCH", "private clinical note", 5, "Injury", "Closed", "Fit", "06/08/2024", "First", "Contact", "Knee", "Ligament"])
        sheet.append(["Team Q", "02/09/24", "TRAINING", "another private note", -1, "Injury", "Open", "Not fit", "", "Recurrent", "Non-contact", "Unknown", "Muscle"])
        sheet.append([None, None, None, None, None])
        book.save(self.workbook)
        book.close()
        self.plan = {
            "plan_version": "team_intake_profiling_plan_v1",
            "team": "Example Club",
            "team_key": "example",
            "season": "2024-25",
            "sources": [{
                "id": "injury",
                "role": "proposed_intake",
                "kind": "injury",
                "path": str(self.workbook),
                "sheets": ["Injuries"],
                "date_order": "day_first",
                "column_classes": {
                    "safe_category": [
                        "Occasion", "Problem Type", "Injury Status", "Fit Status", "Recurrence",
                        "Contact", "Body Part", "Tissue Type",
                    ],
                    "identifier": ["PlayerID"],
                    "free_text": ["Description"],
                    "opaque": ["Days", "Return Date"],
                    "date": ["Date Injured"],
                },
                "duplicate_keys": [["PlayerID", "Date Injured"]],
                "exact_row_duplicates": True,
                "joint_category_keys": [["Body Part", "Tissue Type"]],
                "required_metrics": ["Days", "Missing Metric"],
                "anomaly_rules": [
                    {"id": "negative_days", "column": "Days", "operator": "lt", "value": 0},
                    {"id": "blank_player_rows", "column": "PlayerID", "operator": "blank"},
                ],
            }],
        }

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def complete_profile(self, paths):
        profile = json.loads(paths["profile_draft"].read_text())
        profile.update({
            "decision": "compatible",
            "mapping_path": None,
            "mapping_sha256": None,
            "mapping_version": None,
            "ai_review_status": "completed",
            "ai_reviewed_by": "Fresh reviewer",
            "ai_reviewed_at": "2026-07-12T12:00:00+00:00",
            "unresolved_adjudication_ids": [],
        })
        source_for_field = {
            "occasion_category": "Occasion",
            "match_type": "Occasion",
            "problem_type": "Problem Type",
            "injury_status": "Injury Status",
            "fit_for_selection_status": "Fit Status",
            "confirmed_return_date": "Return Date",
            "days_injured": "Days",
            "severity_time_loss_category": "Days",
            "recurrence": "Recurrence",
            "contact_context": "Contact",
            "body_location": "Body Part",
            "tissue_pathology": "Tissue Type",
        }
        for assessment in profile["canonical_field_assessments"]:
            deterministic = assessment["canonical_field"] in {"days_injured", "severity_time_loss_category"}
            assessment.update({
                "status": "complete",
                "source_fields": [{
                    "source_id": "injury", "sheet": "Injuries",
                    "field": source_for_field[assessment["canonical_field"]],
                }],
                "rule": "Apply the named source evidence; retain source and use Unknown when unsupported",
                "evidence_class": "deterministic_derivation" if deterministic else "source_reported",
                "origin_status": "source_reported_or_derived",
                "coverage_before": 1.0,
                "coverage_after": 1.0,
                "conflicts": [],
                "review_required": False,
                "tests": ["synthetic_field_reconciliation"],
            })
        for provenance in profile["provenance_review"]:
            for field in (
                "preparer", "preparation_timestamp", "codebook_version", "pseudonymisation_status",
                "player_identifier_field", "player_identifier_status", "carried_locator_status",
            ):
                provenance[field] = {"status": "available", "value": f"synthetic {field}"}
            for field in ("secure_original_locator", "secure_original_checksum"):
                provenance[field] = {"status": "unavailable", "value": None}
            provenance["row_reconciliation"] = {
                "status": "completed", "source_rows": 4, "profiled_rows": 4,
                "notes": "Physical and profiled rows reconcile in the synthetic fixture",
            }
        profile["reporting_reviews"] = {
            "injury": {
                "status": "completed", "units": {"Days": "days"}, "gaps": "None in fixture",
                "repeated_measure_structure": "One injury row per event", "native_grain": "event",
                "grain_conclusion": "not_applicable", "grain_review_rationale": "Injury rows are event records",
                "anomalies_reviewed": True,
            },
            "exposure": {
                "status": "completed", "units": {"exposure": "not supplied"},
                "gaps": "No exposure source supplied in this fixture",
                "repeated_measure_structure": "not applicable", "native_grain": "not_applicable",
                "grain_conclusion": "not_applicable", "grain_review_rationale": "No exposure source supplied",
                "anomalies_reviewed": True,
            },
        }
        profile["taxonomy_review"] = {
            "status": "completed", "body_location_inventory_complete": True,
            "tissue_pathology_inventory_complete": True, "notes": "All synthetic categories reviewed",
        }
        profile["tests_and_reconciliation_samples"] = [{
            "id": "synthetic_reconciliation", "status": "passed",
            "evidence": "Aggregate counts match the workbook", "notes": "No raw values retained",
        }]
        profile["ai_review"] = {"status": "completed", "findings": [{
            "finding": "No unresolved synthetic-fixture finding",
            "disposition": "Reviewed and resolved", "status": "resolved",
        }]}
        paths["profile_draft"].write_text(json.dumps(profile, indent=2, sort_keys=True) + "\n")
        return profile

    @staticmethod
    def bind_mapping(profile, mapping_bytes):
        return {
            **profile,
            "mapping_path": "source_to_canonical_mapping.v2.draft.json",
            "mapping_sha256": hashlib.sha256(mapping_bytes).hexdigest(),
            "mapping_version": "source_to_canonical_mapping_v2",
        }

    def test_scan_is_safe_deterministic_and_uses_cache(self) -> None:
        calls = 0

        def counted_loader(*args, **kwargs):
            nonlocal calls
            calls += 1
            return load_workbook(*args, **kwargs)

        first = scan_plan(self.plan, self.root / "out-a", self.root / "cache", counted_loader)
        self.assertEqual(1, calls)
        self.assertEqual(1, first["cache_misses"])
        evidence_path = Path(first["outputs"]["mechanical_evidence"])
        evidence = json.loads(evidence_path.read_text())
        source = evidence["sources"][0]
        self.assertEqual(4, source["sheets"][0]["physical_data_rows"])
        self.assertEqual(3, source["sheets"][0]["substantive_rows"])
        self.assertEqual(1, source["sheets"][0]["duplicate_groups"][0]["groups"])
        self.assertEqual(1, source["sheets"][0]["exact_duplicate_groups"])
        self.assertEqual(2, source["sheets"][0]["exact_duplicate_rows"])
        self.assertEqual({"MATCH": 2, "TRAINING": 1}, source["sheets"][0]["category_frequencies"]["Occasion"])
        self.assertEqual(
            {"source_id": "injury", "sheet": "Injuries"},
            {key: source["sheets"][0]["joint_category_frequencies"][0][key]
             for key in ("source_id", "sheet")},
        )
        self.assertEqual(["Missing Metric"], source["sheets"][0]["missing_required_metrics"])
        self.assertEqual(1, source["sheets"][0]["anomalies"]["negative_days"])
        self.assertEqual(1, source["sheets"][0]["anomalies"]["blank_player_rows"])
        self.assertEqual(2, source["sheets"][0]["identifier_stats"]["PlayerID"]["unique"])
        self.assertEqual("2024-09-02", source["sheets"][0]["reporting_window"]["end"])
        self.assertEqual(
            1,
            source["sheets"][0]["reporting_window"]["date_columns"]["Date Injured"]["patterns"]["dd/mm/yy"],
        )

        combined = "\n".join(Path(path).read_text() for path in first["outputs"].values())
        for forbidden in ("private-player-001", "private clinical note", "another private note", "Team Q"):
            self.assertNotIn(forbidden, combined)

        second = scan_plan(
            self.plan,
            self.root / "out-b",
            self.root / "cache",
            lambda *_args, **_kwargs: self.fail("warm cache reopened workbook"),
        )
        self.assertEqual(1, second["cache_hits"])
        for key in first["outputs"]:
            self.assertEqual(
                Path(first["outputs"][key]).read_bytes(),
                Path(second["outputs"][key]).read_bytes(),
            )

    def test_validator_accepts_complete_safe_package_and_rejects_regressions(self) -> None:
        result = scan_plan(self.plan, self.root / "out", self.root / "cache")
        paths = {key: Path(value) for key, value in result["outputs"].items()}
        draft_report = validate_package(
            paths["mechanical_evidence"], paths["profile_draft"],
            paths["mapping_draft"], paths["column_inventory"],
        )
        self.assertEqual("FAIL", draft_report["status"])
        self.complete_profile(paths)
        report = validate_package(
            paths["mechanical_evidence"], paths["profile_draft"],
            paths["mapping_draft"], paths["column_inventory"],
        )
        self.assertEqual("PASS", report["status"])

        mapping = json.loads(paths["mapping_draft"].read_text())
        mapping["mappings"] = [{
            "canonical_field": "occasion_category",
            "canonical_value": "match",
            "evidence_class": "source_reported",
            "source_evidence": {"Absent Column": "anything"},
            "specificity_change": "equivalent",
            "supporting_evidence": {},
            "evidence_source_id": "injury",
            "evidence_sheet": "Injuries",
            "rule": "Direct source category",
            "protocol_rule_id": None,
            "adjudication_id": None,
        }]
        paths["mapping_draft"].write_text(json.dumps(mapping, indent=2, sort_keys=True) + "\n")
        failed = validate_package(
            paths["mechanical_evidence"], paths["profile_draft"],
            paths["mapping_draft"], paths["column_inventory"],
        )
        self.assertEqual("FAIL", failed["status"])
        self.assertIn("mapping_source_field_absent", {error["code"] for error in failed["errors"]})

    def test_validator_catches_gate_regression_classes(self) -> None:
        result = scan_plan(self.plan, self.root / "out", self.root / "cache")
        paths = {key: Path(value) for key, value in result["outputs"].items()}
        profile = self.complete_profile(paths)
        mapping = json.loads(paths["mapping_draft"].read_text())

        def validate_with(entry=None, profile_change=None):
            candidate_mapping = {**mapping, "mappings": [] if entry is None else [entry]}
            mapping_bytes = (json.dumps(candidate_mapping, indent=2, sort_keys=True) + "\n").encode()
            paths["mapping_draft"].write_bytes(mapping_bytes)
            candidate_profile = json.loads(json.dumps(profile))
            candidate_profile = self.bind_mapping(candidate_profile, mapping_bytes)
            if profile_change:
                profile_change(candidate_profile)
            paths["profile_draft"].write_text(json.dumps(candidate_profile, indent=2, sort_keys=True) + "\n")
            return validate_package(
                paths["mechanical_evidence"], paths["profile_draft"],
                paths["mapping_draft"], paths["column_inventory"],
            )

        valid_entry = {
            "canonical_field": "occasion_category",
            "canonical_value": "match",
            "evidence_class": "source_reported",
            "source_evidence": {"Occasion": "MATCH"},
            "specificity_change": "equivalent",
            "supporting_evidence": {},
            "evidence_source_id": "injury",
            "evidence_sheet": "Injuries",
            "rule": "Direct controlled source value",
            "protocol_rule_id": None,
            "adjudication_id": None,
        }
        self.assertEqual("PASS", validate_with(valid_entry)["status"])
        cases = [
            ({**valid_entry, "canonical_value": "invented"}, "invalid_canonical_value"),
            ({**valid_entry, "evidence_class": "guess"}, "invalid_evidence_class"),
            ({**valid_entry, "source_evidence": {"Occasion": "NEVER OBSERVED"}}, "mapping_source_value_unobserved"),
            ({**valid_entry, "evidence_sheet": "Missing"}, "mapping_source_field_absent"),
            ({
                **valid_entry,
                "canonical_field": "body_location",
                "canonical_value": "knee",
                "specificity_change": "narrower",
                "supporting_evidence": {},
                "evidence_class": "protocol_defined_inference",
                "protocol_rule_id": "team_specific_cross_field_v1",
            }, "unsupported_clinical_narrowing"),
        ]
        for entry, expected in cases:
            with self.subTest(expected=expected):
                report = validate_with(entry)
                self.assertIn(expected, {error["code"] for error in report["errors"]})

        incomplete = validate_with(
            valid_entry,
            lambda candidate: candidate["canonical_field_assessments"].pop(),
        )
        self.assertIn("incomplete_assessments", {error["code"] for error in incomplete["errors"]})
        wrong_scope = validate_with(
            valid_entry,
            lambda candidate: candidate["canonical_field_assessments"][0]["source_fields"][0].update(
                {"sheet": "Missing"}
            ),
        )
        self.assertIn("incomplete_assessments", {error["code"] for error in wrong_scope["errors"]})
        unsafe = validate_with(
            valid_entry,
            lambda candidate: candidate.update({"approval_status": "approved", "approved_by": "Someone"}),
        )
        self.assertIn("unsafe_approval_state", {error["code"] for error in unsafe["errors"]})
        mismatched = validate_with(valid_entry, lambda candidate: candidate.update({"team": "Other Club"}))
        self.assertIn("identity_mismatch", {error["code"] for error in mismatched["errors"]})

    def test_completed_assessments_require_ordered_numeric_coverage(self) -> None:
        result = scan_plan(self.plan, self.root / "out", self.root / "cache")
        paths = {key: Path(value) for key, value in result["outputs"].items()}
        complete = self.complete_profile(paths)
        for before, after in ((None, 1.0), (0.0, None), (-0.1, 1.0), (0.0, 1.1), (0.8, 0.7), (False, 1.0)):
            with self.subTest(before=before, after=after):
                profile = json.loads(json.dumps(complete))
                profile["canonical_field_assessments"][0].update({
                    "coverage_before": before, "coverage_after": after,
                })
                paths["profile_draft"].write_text(json.dumps(profile, indent=2, sort_keys=True) + "\n")
                report = validate_package(
                    paths["mechanical_evidence"], paths["profile_draft"],
                    paths["mapping_draft"], paths["column_inventory"],
                )
                self.assertIn("incomplete_assessments", {error["code"] for error in report["errors"]})

    def test_plan_rejects_unknown_keys_and_ambiguous_date_order(self) -> None:
        bad_plan = {**self.plan, "future_option": True}
        with self.assertRaises(ProfileError):
            scan_plan(bad_plan, self.root / "out-a", self.root / "cache-a")
        bad_source = {**self.plan["sources"][0]}
        bad_source.pop("date_order")
        with self.assertRaises(ProfileError):
            scan_plan(
                {**self.plan, "sources": [bad_source]},
                self.root / "out-b", self.root / "cache-b",
            )

    def test_sensitive_headers_cannot_be_declared_safe_before_outputs_exist(self) -> None:
        for header, original_class in (
            ("PlayerID", "identifier"), ("Description", "free_text"),
            ("Name", None), ("ID", None), ("UID", None), ("Clinical Notes", None),
            ("Surname", None), ("Forename", None), ("Initials", None),
            ("Player Code", None), ("Athlete Number", None), ("Email", None),
            ("playerid", None), ("playeruid", None), ("athleteid", None),
            ("dateofbirth", None), ("clinicalnotes", None), ("freetext", None),
            ("playerno", None), ("athleteno", None), ("participantno", None),
        ):
            with self.subTest(header=header):
                plan = json.loads(json.dumps(self.plan))
                classes = plan["sources"][0]["column_classes"]
                if original_class:
                    classes[original_class].remove(header)
                classes["safe_category"].append(header)
                output = self.root / f"blocked-{original_class}"
                cache = self.root / f"blocked-cache-{original_class}"
                with self.assertRaises(ProfileError):
                    scan_plan(plan, output, cache)
                self.assertFalse(output.exists())
                self.assertFalse(cache.exists())

    def test_anomaly_rule_shapes_fail_closed(self) -> None:
        invalid_rules = [
            {"id": "blank", "operator": "blank", "column": "Days", "value": 1},
            {"id": "lt", "operator": "lt", "column": "Days"},
            {"id": "elapsed", "operator": "elapsed_minutes_gt", "start_column": "Days", "value": 1},
            {"id": "duration", "operator": "duration_minutes_gt", "column": "Days"},
        ]
        for index, rule in enumerate(invalid_rules):
            with self.subTest(rule=rule["id"]):
                plan = json.loads(json.dumps(self.plan))
                plan["sources"][0]["anomaly_rules"] = [rule]
                with self.assertRaises(ProfileError):
                    scan_plan(plan, self.root / f"rule-out-{index}", self.root / f"rule-cache-{index}")

    def test_adapter_requires_mapping_but_compatible_allows_empty(self) -> None:
        result = scan_plan(self.plan, self.root / "out", self.root / "cache")
        paths = {key: Path(value) for key, value in result["outputs"].items()}
        profile = self.complete_profile(paths)
        compatible = validate_package(
            paths["mechanical_evidence"], paths["profile_draft"],
            paths["mapping_draft"], paths["column_inventory"],
        )
        self.assertEqual("PASS", compatible["status"])
        from pipeline.__main__ import validate_intake_profile_manifest
        from datetime import UTC, datetime, timedelta

        input_sha = "a" * 64
        now = datetime.now(UTC)
        approved_profile = {
            **profile,
            "ai_reviewed_at": (now - timedelta(minutes=2)).isoformat(),
            "approved_by": "Abdel Babiker",
            "approved_at": (now - timedelta(minutes=1)).isoformat(),
            "approved_input_sha256s": [input_sha],
        }
        paths["profile_draft"].write_text(json.dumps(approved_profile, indent=2, sort_keys=True) + "\n")
        bound_fields = (
            "team", "season", "profile_version", "decision", "mapping_path", "mapping_sha256",
            "mapping_version", "ai_review_status", "ai_reviewed_by", "ai_reviewed_at", "approved_by",
            "approved_at", "unresolved_adjudication_ids", "approved_input_sha256s",
        )
        envelope = {field: approved_profile[field] for field in bound_fields}
        envelope.update({
            "profile_path": str(paths["profile_draft"]),
            "profile_sha256": hashlib.sha256(paths["profile_draft"].read_bytes()).hexdigest(),
        })
        validate_intake_profile_manifest(
            {"intake_profile": envelope}, self.root / "manifest.json", input_sha,
            approved_profile["team"], approved_profile["season"],
        )
        paths["profile_draft"].write_text(json.dumps(profile, indent=2, sort_keys=True) + "\n")
        profile["decision"] = "adapter_required"
        paths["profile_draft"].write_text(json.dumps(profile, indent=2, sort_keys=True) + "\n")
        adapter = validate_package(
            paths["mechanical_evidence"], paths["profile_draft"],
            paths["mapping_draft"], paths["column_inventory"],
        )
        self.assertIn("adapter_mapping_required", {error["code"] for error in adapter["errors"]})

    def test_inventory_content_and_profiled_row_reconciliation_are_bound(self) -> None:
        result = scan_plan(self.plan, self.root / "out", self.root / "cache")
        paths = {key: Path(value) for key, value in result["outputs"].items()}
        profile = self.complete_profile(paths)
        profile["provenance_review"][0]["row_reconciliation"]["profiled_rows"] = 3
        paths["profile_draft"].write_text(json.dumps(profile, indent=2, sort_keys=True) + "\n")
        row_report = validate_package(
            paths["mechanical_evidence"], paths["profile_draft"],
            paths["mapping_draft"], paths["column_inventory"],
        )
        self.assertIn("incomplete_provenance", {error["code"] for error in row_report["errors"]})

        profile["provenance_review"][0]["row_reconciliation"]["profiled_rows"] = 4
        inventory = json.loads(paths["column_inventory"].read_text())
        inventory["sources"][0]["sheets"][0]["columns"][0]["populated"] -= 1
        inventory_bytes = (json.dumps(inventory, indent=2, sort_keys=True) + "\n").encode()
        paths["column_inventory"].write_bytes(inventory_bytes)
        inventory_sha = hashlib.sha256(inventory_bytes).hexdigest()
        mapping = json.loads(paths["mapping_draft"].read_text())
        mapping["inventory_sha256"] = inventory_sha
        mapping_bytes = (json.dumps(mapping, indent=2, sort_keys=True) + "\n").encode()
        paths["mapping_draft"].write_bytes(mapping_bytes)
        profile["column_inventory_sha256"] = inventory_sha
        profile["mapping_sha256"] = None
        paths["profile_draft"].write_text(json.dumps(profile, indent=2, sort_keys=True) + "\n")
        inventory_report = validate_package(
            paths["mechanical_evidence"], paths["profile_draft"],
            paths["mapping_draft"], paths["column_inventory"],
        )
        self.assertIn("inventory_content_mismatch", {error["code"] for error in inventory_report["errors"]})

    def test_recursive_schema_rejects_nested_raw_values_without_echoing_them(self) -> None:
        result = scan_plan(self.plan, self.root / "out", self.root / "cache")
        paths = {key: Path(value) for key, value in result["outputs"].items()}
        self.complete_profile(paths)
        evidence = json.loads(paths["mechanical_evidence"].read_text())
        evidence["sources"][0]["sheets"][0]["raw_values"] = ["do-not-echo-this-value"]
        paths["mechanical_evidence"].write_text(json.dumps(evidence, indent=2, sort_keys=True) + "\n")
        report = validate_package(
            paths["mechanical_evidence"], paths["profile_draft"],
            paths["mapping_draft"], paths["column_inventory"],
        )
        self.assertIn("unknown_schema_key", {error["code"] for error in report["errors"]})
        self.assertNotIn("do-not-echo-this-value", json.dumps(report))

    def test_broad_clinical_evidence_cannot_hide_as_equivalent(self) -> None:
        result = scan_plan(self.plan, self.root / "out", self.root / "cache")
        paths = {key: Path(value) for key, value in result["outputs"].items()}
        profile = self.complete_profile(paths)
        mapping = json.loads(paths["mapping_draft"].read_text())

        def check(entry):
            candidate = {**mapping, "mappings": [entry]}
            mapping_bytes = (json.dumps(candidate, indent=2, sort_keys=True) + "\n").encode()
            paths["mapping_draft"].write_bytes(mapping_bytes)
            candidate_profile = self.bind_mapping(profile, mapping_bytes)
            paths["profile_draft"].write_text(json.dumps(candidate_profile, indent=2, sort_keys=True) + "\n")
            return validate_package(
                paths["mechanical_evidence"], paths["profile_draft"],
                paths["mapping_draft"], paths["column_inventory"],
            )

        broad = {
            "canonical_field": "body_location", "canonical_value": "knee",
            "evidence_class": "source_reported", "source_evidence": {"Body Part": "Unknown"},
            "specificity_change": "equivalent", "supporting_evidence": {},
            "evidence_source_id": "injury", "evidence_sheet": "Injuries",
            "rule": "Direct label", "protocol_rule_id": None, "adjudication_id": None,
        }
        report = check(broad)
        self.assertIn("unsupported_clinical_narrowing", {error["code"] for error in report["errors"]})
        supported = {
            **broad,
            "evidence_class": "protocol_defined_inference",
            "supporting_evidence": {"Tissue Type": "Ligament"},
            "rule": "Use the named cross-field evidence",
            "protocol_rule_id": "team_specific_cross_field_v1",
        }
        marginal_only = check(supported)
        self.assertIn("unsupported_clinical_narrowing", {error["code"] for error in marginal_only["errors"]})
        observed_together = {**supported, "supporting_evidence": {"Tissue Type": "Muscle"}}
        self.assertEqual("PASS", check(observed_together)["status"])

    def test_joint_mapping_evidence_cannot_cross_sheet_boundaries(self) -> None:
        book = load_workbook(self.workbook)
        other = book.create_sheet("Other")
        other.append([cell.value for cell in book["Injuries"][1]])
        other.append([
            "private-player-002", "03/09/2024", "TRAINING", "private note", 2,
            "Injury", "Closed", "Fit", "05/09/2024", "First", "Contact", "Unknown", "Ligament",
        ])
        book.save(self.workbook)
        book.close()
        plan = json.loads(json.dumps(self.plan))
        plan["sources"][0]["sheets"].append("Other")
        result = scan_plan(plan, self.root / "scoped-out", self.root / "scoped-cache")
        paths = {key: Path(value) for key, value in result["outputs"].items()}
        profile = self.complete_profile(paths)
        evidence = json.loads(paths["mechanical_evidence"].read_text())
        profiled_rows = sum(sheet["physical_data_rows"] for sheet in evidence["sources"][0]["sheets"])
        profile["provenance_review"][0]["row_reconciliation"].update({
            "source_rows": profiled_rows, "profiled_rows": profiled_rows,
        })
        mapping = json.loads(paths["mapping_draft"].read_text())

        def check(evidence_sheet):
            entry = {
                "canonical_field": "body_location", "canonical_value": "knee",
                "evidence_class": "protocol_defined_inference",
                "source_evidence": {"Body Part": "Unknown"},
                "supporting_evidence": {"Tissue Type": "Muscle"},
                "evidence_source_id": "injury", "evidence_sheet": evidence_sheet,
                "specificity_change": "narrower", "rule": "Use observed cross-field evidence",
                "protocol_rule_id": "team_specific_cross_field_v1", "adjudication_id": None,
            }
            candidate = {**mapping, "mappings": [entry]}
            mapping_bytes = (json.dumps(candidate, indent=2, sort_keys=True) + "\n").encode()
            paths["mapping_draft"].write_bytes(mapping_bytes)
            candidate_profile = self.bind_mapping(profile, mapping_bytes)
            paths["profile_draft"].write_text(json.dumps(candidate_profile, indent=2, sort_keys=True) + "\n")
            return validate_package(
                paths["mechanical_evidence"], paths["profile_draft"],
                paths["mapping_draft"], paths["column_inventory"],
            )

        self.assertIn(
            "unsupported_clinical_narrowing",
            {error["code"] for error in check("Other")["errors"]},
        )
        same_sheet = check("Injuries")
        self.assertEqual("PASS", same_sheet["status"], same_sheet["errors"])

    def test_time_only_and_duration_anomalies_are_aggregate_only(self) -> None:
        book_path = self.root / "durations.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "Exposure"
        sheet.append(["Start", "End", "Duration"])
        sheet.append(["10:00:00 PM", "2:30:00 AM", "4:30:00"])
        sheet.append(["1:00:00 PM", "2:00:00 PM", "1:00:00"])
        book.save(book_path)
        book.close()
        plan = {
            "plan_version": "team_intake_profiling_plan_v1", "team": "Example Club",
            "team_key": "duration-example", "season": "2024-25", "sources": [{
                "id": "exposure", "role": "proposed_intake", "kind": "exposure",
                "path": str(book_path), "sheets": ["Exposure"],
                "column_classes": {"safe_category": [], "identifier": [], "free_text": [], "opaque": ["Start", "End", "Duration"], "date": []},
                "duplicate_keys": [], "required_metrics": [],
                "anomaly_rules": [
                    {"id": "elapsed_over_240", "operator": "elapsed_minutes_gt", "start_column": "Start", "end_column": "End", "value": 240},
                    {"id": "duration_over_240", "operator": "duration_minutes_gt", "column": "Duration", "value": 240},
                ],
            }],
        }
        result = scan_plan(plan, self.root / "duration-out", self.root / "duration-cache")
        anomalies = result["sources"][0]["sheets"][0]["anomalies"]
        self.assertEqual({"duration_over_240": 1, "elapsed_over_240": 1}, anomalies)

    def test_exposure_grain_is_only_a_populated_evidence_candidate(self) -> None:
        book_path = self.root / "grain.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "Exposure"
        sheet.append(["Week", "Session Date"])
        book.save(book_path)
        book.close()
        plan = {
            "plan_version": "team_intake_profiling_plan_v1", "team": "Example Club",
            "team_key": "grain-example", "season": "2024-25", "sources": [{
                "id": "exposure", "role": "proposed_intake", "kind": "exposure",
                "path": str(book_path), "sheets": ["Exposure"],
                "column_classes": {"safe_category": [], "identifier": [], "free_text": [], "opaque": ["Week", "Session Date"], "date": []},
                "duplicate_keys": [], "required_metrics": [], "anomaly_rules": [],
                "exposure_grain_evidence": {"weekly_columns": ["Week"], "session_columns": ["Session Date"]},
            }],
        }
        result = scan_plan(plan, self.root / "grain-out", self.root / "grain-cache")
        grain = result["sources"][0]["sheets"][0]["exposure_grain_evidence"]
        self.assertEqual("unknown", grain["candidate"])
        self.assertFalse(grain["confirmed"])
        self.assertEqual(0, grain["weekly_columns"][0]["populated"])
        self.assertEqual(0, grain["session_columns"][0]["distinct"])

    def test_family_check_is_early_optional_deterministic_and_aggregate_only(self) -> None:
        plans = []
        for team_key in ("club-a", "club-b"):
            plan = json.loads(json.dumps(self.plan))
            plan.update({"team": team_key.title(), "team_key": team_key})
            plan["sources"][0]["required_metrics"] = ["Days"]
            plans.append(plan)
        results = [
            scan_plan(plan, self.root / f"out-{index}", self.root / "family-cache")
            for index, plan in enumerate(plans)
        ]
        evidence = [Path(result["outputs"]["mechanical_evidence"]) for result in results]
        first_path = self.root / "family-a.json"
        second_path = self.root / "family-b.json"
        first = family_check(evidence, first_path)
        family_check(list(reversed(evidence)), second_path)
        self.assertEqual(first_path.read_bytes(), second_path.read_bytes())
        self.assertEqual("exact_structure_candidate", first["candidate_families"][0]["classification"])
        self.assertTrue(first["advisory_only"])
        single = family_check(evidence[:1])
        self.assertEqual([], single["candidate_families"])
        self.assertEqual(["no_matching_structure"], single["ungrouped"][0]["blockers"])
        redacted = json.loads(evidence[1].read_text())
        redacted["sources"][0]["sheets"][0]["privacy_redactions"] = 1
        redacted_path = self.root / "redacted-evidence.json"
        redacted_path.write_text(json.dumps(redacted, indent=2, sort_keys=True) + "\n")
        privacy_blocked = family_check([evidence[0], redacted_path])
        self.assertEqual([], privacy_blocked["candidate_families"])
        self.assertIn("privacy_redactions_present", privacy_blocked["ungrouped"][1]["blockers"])
        text = first_path.read_text()
        for forbidden in ("Occasion", "MATCH", "private-player", "Description", "Knee", "Ligament"):
            self.assertNotIn(forbidden, text)

    def test_family_check_allows_sheet_binding_only_but_rejects_structural_drift(self) -> None:
        base_plan = json.loads(json.dumps(self.plan))
        base_plan.update({"team": "Club A", "team_key": "club-a"})
        base_plan["sources"][0]["required_metrics"] = ["Days"]
        first = scan_plan(base_plan, self.root / "family-a", self.root / "family-cache")

        renamed_book = self.root / "renamed.xlsx"
        book = load_workbook(self.workbook)
        book.active.title = "Other Injuries"
        book.save(renamed_book)
        book.close()
        renamed_plan = json.loads(json.dumps(base_plan))
        renamed_plan.update({"team": "Club B", "team_key": "club-b"})
        renamed_plan["sources"][0].update({"path": str(renamed_book), "sheets": ["Other Injuries"]})
        second = scan_plan(renamed_plan, self.root / "family-b", self.root / "family-cache")
        evidence = [Path(item["outputs"]["mechanical_evidence"]) for item in (first, second)]
        shape = family_check(evidence)
        self.assertEqual("shape_equivalent_candidate", shape["candidate_families"][0]["classification"])
        self.assertIn("sheet_bindings", shape["candidate_families"][0])

        drift_plan = json.loads(json.dumps(base_plan))
        drift_plan.update({"team": "Club C", "team_key": "club-c"})
        drift_plan["sources"][0]["column_classes"]["safe_category"].remove("Occasion")
        drift_plan["sources"][0]["column_classes"]["opaque"].append("Occasion")
        drift = scan_plan(drift_plan, self.root / "family-c", self.root / "family-cache")
        independent = family_check([
            evidence[0], Path(drift["outputs"]["mechanical_evidence"]),
        ])
        self.assertEqual([], independent["candidate_families"])

    def test_corrupt_cache_is_rebuilt(self) -> None:
        first = scan_plan(self.plan, self.root / "out-a", self.root / "cache")
        cache_file = next((self.root / "cache").glob("*.json"))
        cache_file.write_text('{"payload": {}}\n')
        calls = 0

        def counted_loader(*args, **kwargs):
            nonlocal calls
            calls += 1
            return load_workbook(*args, **kwargs)

        second = scan_plan(self.plan, self.root / "out-b", self.root / "cache", counted_loader)
        self.assertEqual(1, calls)
        self.assertEqual(1, second["cache_misses"])
        self.assertEqual(
            Path(first["outputs"]["mechanical_evidence"]).read_bytes(),
            Path(second["outputs"]["mechanical_evidence"]).read_bytes(),
        )

    def test_cache_does_not_reuse_a_different_source_role(self) -> None:
        scan_plan(self.plan, self.root / "out-a", self.root / "cache")
        changed_source = {**self.plan["sources"][0], "role": "reference_only"}
        result = scan_plan(
            {**self.plan, "sources": [changed_source]},
            self.root / "out-b", self.root / "cache",
        )
        self.assertEqual(1, result["cache_misses"])
        self.assertEqual("reference_only", result["sources"][0]["role"])


if __name__ == "__main__":
    unittest.main()
