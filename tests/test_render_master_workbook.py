from __future__ import annotations

import csv
import importlib.util
import json
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "tools/render.py"
SPEC = importlib.util.spec_from_file_location("render_master_workbook", SCRIPT)
assert SPEC and SPEC.loader
RENDER = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(RENDER)


class RenderMasterWorkbookTests(unittest.TestCase):
    def synthetic_workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Injury Master"
        headers = [
            "Problem type",
            "Occasion category",
            "Value",
            "Date",
            "Exclusion Reason",
        ]
        sheet.append(headers)
        sheet.append(["Injury", "Match", 1.25, datetime(2024, 9, 7), "Non-URC match"])
        sheet.append(["Injury", "Non-Rugby", 2, date(2024, 9, 8), None])
        sheet.append(["Illness", "Gym-Based", 3, None, None])
        sheet["A2"].font = Font(name="Calibri", size=11, color="FF8F1D24")
        sheet["B2"].font = Font(name="Calibri", size=11, color="FF8F1D24")
        sheet["C2"].font = Font(name="Calibri", size=11, color="FF8F1D24")
        sheet["D2"].font = Font(name="Calibri", size=11, color="FF8F1D24")
        sheet["E2"].font = Font(name="Calibri", size=11, color="FF8F1D24")
        for cell in sheet[2]:
            cell.fill = PatternFill(
                patternType="solid",
                fgColor="FFFDEBEC",
                bgColor=RENDER.Color(indexed=64),
            )
        sheet["A3"].font = Font(name="Calibri", size=11, color="FF008000")
        sheet["B3"].font = Font(name="Calibri", size=11, color=RENDER.Color(indexed=10))
        sheet["C3"].fill = PatternFill(patternType="solid", fgColor="FFE2F0D9")
        sheet["C3"].number_format = "0.00"
        sheet["D2"].number_format = "dd/mm/yyyy"
        sheet["D3"].number_format = "yyyy-mm-dd"
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["D"].width = 14
        sheet.row_dimensions[3].height = 24
        table = Table(name="InjuryMaster", displayName="InjuryMaster", ref="A1:E4")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        second = workbook.create_sheet("Merged")
        second["A1"] = "Merged title"
        second.merge_cells("A1:C1")
        second.column_dimensions["A"].width = 18
        workbook.save(path)

    def test_extract_render_round_trip(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "source.xlsx"
            rendered = root / "rendered.xlsx"
            self.synthetic_workbook(source)
            data = RENDER.extract_workbook(source)
            RENDER.render_workbook(data, rendered)
            report = RENDER.compare_workbooks(source, rendered)
            self.assertEqual(report["difference_count"], 0, report["differences"])
            encoded_date = data["sheets"][0]["values"][1][3]
            self.assertEqual(encoded_date["$type"], "datetime")
            self.assertEqual(
                data["sheets"][0]["styles"]["B3"]["font"]["color"],
                {"type": "indexed", "value": 10, "tint": 0.0},
            )
            self.assertEqual(
                data["sheets"][0]["tables"],
                [
                    {
                        "name": "InjuryMaster",
                        "display_name": "InjuryMaster",
                        "ref": "A1:E4",
                        "header_row_count": 1,
                        "style": {
                            "name": "TableStyleMedium2",
                            "show_first_column": False,
                            "show_last_column": False,
                            "show_row_stripes": False,
                            "show_column_stripes": False,
                        },
                    }
                ],
            )
            rendered_tables = RENDER.extract_workbook(rendered)["sheets"][0]["tables"]
            self.assertEqual(rendered_tables, data["sheets"][0]["tables"])

    def test_compare_summarizes_nonzero_cell_differences(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            old = root / "old.xlsx"
            new = root / "new.xlsx"
            self.synthetic_workbook(old)
            workbook = load_workbook(old)
            workbook["Injury Master"]["C3"] = 99
            workbook.save(new)
            report = RENDER.compare_workbooks(old, new)
            self.assertEqual(report["difference_count"], 1)
            self.assertEqual(report["cell_property_counts"], {"value": 1})

    def test_compare_lists_cells_from_a_missing_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            old = root / "old.xlsx"
            new = root / "new.xlsx"
            self.synthetic_workbook(old)
            workbook = load_workbook(old)
            workbook.remove(workbook["Merged"])
            workbook.save(new)
            report = RENDER.compare_workbooks(old, new)
            missing_cells = [
                item
                for item in report["differences"]
                if item["kind"] == "cell" and item["sheet"] == "Merged"
            ]
            self.assertEqual(
                [item["cell"] for item in missing_cells],
                ["A1", "B1", "C1"],
            )
            self.assertIsNotNone(
                missing_cells[0]["differences"]["cell_presence"]["old"]
            )
            self.assertIsNone(
                missing_cells[0]["differences"]["cell_presence"]["new"]
            )

    def test_mark_excluded_changes_only_value_font_and_fill(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "source.xlsx"
            self.synthetic_workbook(source)
            data = RENDER.extract_workbook(source)
            before = json.loads(json.dumps(data))
            updated, summary = RENDER.mark_excluded_data(
                data,
                [{"source_workbook_row": "3"}],
                "Non-rugby or gym-based occasion",
            )
            sheet = updated["sheets"][0]
            self.assertEqual(
                sheet["values"][2][4],
                "Non-rugby or gym-based occasion",
            )
            self.assertEqual(summary["rows_marked"], 1)
            for column in "ABCDE":
                style = sheet["styles"][f"{column}3"]
                self.assertEqual(
                    style["font"]["color"],
                    {"type": "rgb", "value": "FF8F1D24", "tint": 0.0},
                )
                self.assertEqual(style["fill"]["fg_color"]["value"], "FFFDEBEC")
            self.assertEqual(
                before["sheets"][0]["styles"]["C3"]["number_format"],
                sheet["styles"]["C3"]["number_format"],
            )

    def test_mark_excluded_aborts_before_mutation_on_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "source.xlsx"
            self.synthetic_workbook(source)
            data = RENDER.extract_workbook(source)
            before = json.loads(json.dumps(data))
            with self.assertRaisesRegex(
                RENDER.RenderError,
                "Problem type is 'Illness', expected 'Injury'",
            ):
                RENDER.mark_excluded_data(
                    data,
                    [{"source_workbook_row": "4"}],
                    "Non-rugby or gym-based occasion",
                )
            self.assertEqual(data, before)

    def test_production_count_guards_are_atomic(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Injury Master"
            sheet.append(
                ["Problem type", "Occasion category", "Exclusion Reason"]
            )
            sheet.append(["Injury", "Match", "Non-URC match"])
            for cell in sheet[2]:
                cell.font = Font(name="Calibri", size=11, color="FF8F1D24")
                cell.fill = PatternFill(
                    patternType="solid",
                    fgColor="FFFDEBEC",
                    bgColor=RENDER.Color(indexed=64),
                )
            for index in range(RENDER.EXPECTED_MARK_COUNT):
                occasion = "Non-Rugby" if index < 93 else "Gym-Based"
                sheet.append(["Injury", occasion, None])
            workbook.save(source)
            base = RENDER.extract_workbook(source)
            valid_audit = [
                {"source_workbook_row": str(row)}
                for row in range(3, 3 + RENDER.EXPECTED_MARK_COUNT)
            ]
            updated, summary = RENDER.mark_excluded_data(
                json.loads(json.dumps(base)),
                valid_audit,
                "Non-rugby or gym-based occasion",
                require_production_counts=True,
            )
            self.assertEqual(summary["rows_marked"], 120)
            self.assertEqual(
                updated["sheets"][0]["values"][2][2],
                "Non-rugby or gym-based occasion",
            )
            invalid_cases = {
                "wrong total": valid_audit[:-1],
                "duplicate": valid_audit[:-1] + [valid_audit[0]],
                "out of range": valid_audit[:-1]
                + [{"source_workbook_row": "9999"}],
            }
            nonblank = json.loads(json.dumps(base))
            nonblank["sheets"][0]["values"][2][2] = "Already excluded"
            invalid_cases["nonblank reason"] = valid_audit
            for label, audit in invalid_cases.items():
                with self.subTest(label=label):
                    candidate = (
                        nonblank
                        if label == "nonblank reason"
                        else json.loads(json.dumps(base))
                    )
                    before = json.loads(json.dumps(candidate))
                    with self.assertRaises(RENDER.RenderError):
                        RENDER.mark_excluded_data(
                            candidate,
                            audit,
                            "Non-rugby or gym-based occasion",
                            require_production_counts=True,
                        )
                    self.assertEqual(candidate, before)


if __name__ == "__main__":
    unittest.main()
