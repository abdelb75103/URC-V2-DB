#!/usr/bin/env python3
"""Extract, render, compare, and mark the canonical URC master workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections import Counter
from copy import copy
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Color, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


FORMAT_VERSION = 1
MASTER_SHEET = "Injury Master"
EXCLUSION_REASON_HEADER = "Exclusion Reason"
EXPECTED_MARK_COUNT = 120
EXPECTED_OCCASION_COUNTS = {"Non-Rugby": 93, "Gym-Based": 27}
FONT_DEFAULTS = {
    "color": {"type": "theme", "value": 1, "tint": 0.0},
    "bold": False,
    "italic": False,
    "name": "Calibri",
    "size": 11.0,
}
FILL_DEFAULTS = {
    "type": None,
    "fg_color": {"type": "rgb", "value": "00000000", "tint": 0.0},
    "bg_color": {"type": "rgb", "value": "00000000", "tint": 0.0},
}
FONT_FIELDS = ("color", "bold", "italic", "name", "size")
SIDE_NAMES = (
    "left",
    "right",
    "top",
    "bottom",
    "diagonal",
    "vertical",
    "horizontal",
    "start",
    "end",
)
ALIGNMENT_FIELDS = (
    "horizontal",
    "vertical",
    "text_rotation",
    "wrap_text",
    "shrink_to_fit",
    "indent",
    "relativeIndent",
    "justifyLastLine",
    "readingOrder",
)


class RenderError(ValueError):
    """Raised when source data fails a required workbook contract."""


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def encode_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return {"$type": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"$type": "date", "value": value.isoformat()}
    if isinstance(value, time):
        return {"$type": "time", "value": value.isoformat()}
    if isinstance(value, timedelta):
        return {"$type": "timedelta", "seconds": value.total_seconds()}
    return value


def decode_value(value: Any) -> Any:
    if not isinstance(value, dict) or "$type" not in value:
        return value
    value_type = value["$type"]
    if value_type == "datetime":
        return datetime.fromisoformat(value["value"])
    if value_type == "date":
        return date.fromisoformat(value["value"])
    if value_type == "time":
        return time.fromisoformat(value["value"])
    if value_type == "timedelta":
        return timedelta(seconds=value["seconds"])
    raise RenderError(f"Unsupported typed cell value: {value_type!r}")


def color_record(color: Color | None) -> dict[str, Any] | None:
    if color is None:
        return None
    color_type = color.type
    if color_type == "rgb":
        value = color.rgb
    elif color_type == "indexed":
        value = color.indexed
    elif color_type == "theme":
        value = color.theme
    elif color_type == "auto":
        value = color.auto
    else:
        raise RenderError(f"Unsupported color type: {color_type!r}")
    return {"type": color_type, "value": value, "tint": color.tint}


def color_from_record(record: dict[str, Any] | None) -> Color | None:
    if record is None:
        return None
    kwargs = {record["type"]: record["value"], "tint": record.get("tint", 0.0)}
    return Color(**kwargs)


def font_projection(cell: Any) -> dict[str, Any]:
    return {
        "color": color_record(cell.font.color),
        "bold": bool(cell.font.bold),
        "italic": bool(cell.font.italic),
        "name": cell.font.name,
        "size": float(cell.font.sz) if cell.font.sz is not None else None,
    }


def fill_projection(cell: Any) -> dict[str, Any]:
    return {
        "type": cell.fill.fill_type,
        "fg_color": color_record(cell.fill.fgColor),
        "bg_color": color_record(cell.fill.bgColor),
    }


def side_record(side: Side | None) -> dict[str, Any] | None:
    if side is None:
        return None
    return {"style": side.style, "color": color_record(side.color)}


def border_projection(cell: Any) -> dict[str, Any]:
    border = cell.border
    record = {name: side_record(getattr(border, name)) for name in SIDE_NAMES}
    record.update(
        {
            "diagonalUp": bool(border.diagonalUp),
            "diagonalDown": bool(border.diagonalDown),
            "outline": border.outline,
        }
    )
    return record


def _default_border_projection() -> dict[str, Any]:
    border = Border()
    record = {name: side_record(getattr(border, name)) for name in SIDE_NAMES}
    record.update(
        {
            "diagonalUp": bool(border.diagonalUp),
            "diagonalDown": bool(border.diagonalDown),
            "outline": border.outline,
        }
    )
    return record


DEFAULT_BORDER = _default_border_projection()


def alignment_projection(cell: Any) -> dict[str, Any]:
    alignment = cell.alignment
    return {field: getattr(alignment, field) for field in ALIGNMENT_FIELDS}


def _default_alignment_projection() -> dict[str, Any]:
    alignment = Alignment()
    return {field: getattr(alignment, field) for field in ALIGNMENT_FIELDS}


DEFAULT_ALIGNMENT = _default_alignment_projection()


def sparse_style_record(cell: Any) -> dict[str, Any]:
    record: dict[str, Any] = {}
    font = font_projection(cell)
    sparse_font = {
        field: font[field]
        for field in FONT_FIELDS
        if font[field] != FONT_DEFAULTS[field]
    }
    if sparse_font:
        record["font"] = sparse_font
    fill = fill_projection(cell)
    if fill != FILL_DEFAULTS:
        record["fill"] = fill
    border = border_projection(cell)
    if border != DEFAULT_BORDER:
        record["border"] = border
    alignment = alignment_projection(cell)
    if alignment != DEFAULT_ALIGNMENT:
        record["alignment"] = alignment
    if cell.number_format != "General":
        record["number_format"] = cell.number_format
    return record


def _side_from_record(record: dict[str, Any] | None) -> Side | None:
    if record is None:
        return None
    return Side(
        style=record.get("style"),
        color=color_from_record(record.get("color")),
    )


def apply_style_record(cell: Any, record: dict[str, Any]) -> None:
    if "font" in record:
        font = copy(cell.font)
        source = record["font"]
        if "color" in source:
            font.color = color_from_record(source["color"])
        if "bold" in source:
            font.bold = source["bold"]
        if "italic" in source:
            font.italic = source["italic"]
        if "name" in source:
            font.name = source["name"]
        if "size" in source:
            font.sz = source["size"]
        cell.font = font
    if "fill" in record:
        fill = record["fill"]
        cell.fill = PatternFill(
            patternType=fill["type"],
            fgColor=color_from_record(fill["fg_color"]),
            bgColor=color_from_record(fill["bg_color"]),
        )
    if "border" in record:
        border = record["border"]
        kwargs = {name: _side_from_record(border[name]) for name in SIDE_NAMES}
        kwargs.update(
            {
                "diagonalUp": border["diagonalUp"],
                "diagonalDown": border["diagonalDown"],
                "outline": border["outline"],
            }
        )
        cell.border = Border(**kwargs)
    if "alignment" in record:
        cell.alignment = Alignment(**record["alignment"])
    if "number_format" in record:
        cell.number_format = record["number_format"]


def extract_workbook(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=False)
    sheets = []
    for worksheet in workbook.worksheets:
        values = []
        styles: dict[str, Any] = {}
        for row_number in range(1, worksheet.max_row + 1):
            row_values = []
            for column_number in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row_number, column_number)
                row_values.append(encode_value(cell.value))
                style = sparse_style_record(cell)
                if style:
                    styles[cell.coordinate] = style
            values.append(row_values)
        sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "values": values,
                "styles": styles,
                "column_widths": {
                    key: dimension.width
                    for key, dimension in worksheet.column_dimensions.items()
                    if dimension.width is not None
                },
                "row_heights": {
                    str(key): dimension.height
                    for key, dimension in worksheet.row_dimensions.items()
                    if dimension.height is not None
                },
                "merged_ranges": [
                    str(merged_range) for merged_range in worksheet.merged_cells.ranges
                ],
                "tables": _table_records(worksheet),
            }
        )
    return {
        "format": "urc-master-workbook",
        "format_version": FORMAT_VERSION,
        "source": {
            "path": str(workbook_path),
            "sha256": sha256_file(workbook_path),
        },
        "sheets": sheets,
    }


def render_workbook(data: dict[str, Any], output_path: Path) -> None:
    if data.get("format") != "urc-master-workbook":
        raise RenderError("Input JSON is not a URC master workbook extraction")
    if data.get("format_version") != FORMAT_VERSION:
        raise RenderError(
            f"Unsupported format version: {data.get('format_version')!r}"
        )
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet in data["sheets"]:
        worksheet = workbook.create_sheet(sheet["name"])
        for row_number, row_values in enumerate(sheet["values"], start=1):
            for column_number, encoded in enumerate(row_values, start=1):
                worksheet.cell(row_number, column_number).value = decode_value(encoded)
        for coordinate, style in sheet["styles"].items():
            apply_style_record(worksheet[coordinate], style)
        for column, width in sheet["column_widths"].items():
            worksheet.column_dimensions[column].width = width
        for row, height in sheet["row_heights"].items():
            worksheet.row_dimensions[int(row)].height = height
        for merged_range in sheet["merged_ranges"]:
            worksheet.merge_cells(merged_range)
        for record in sheet.get("tables", []):
            table = Table(
                name=record["name"],
                displayName=record["display_name"],
                ref=record["ref"],
                headerRowCount=record["header_row_count"],
            )
            style = record.get("style")
            if style is not None:
                table.tableStyleInfo = TableStyleInfo(
                    name=style["name"],
                    showFirstColumn=style["show_first_column"],
                    showLastColumn=style["show_last_column"],
                    showRowStripes=style["show_row_stripes"],
                    showColumnStripes=style["show_column_stripes"],
                )
            worksheet.add_table(table)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _table_records(worksheet: Any) -> list[dict[str, Any]]:
    records = []
    for table in worksheet.tables.values():
        style = table.tableStyleInfo
        records.append(
            {
                "name": table.name,
                "display_name": table.displayName,
                "ref": table.ref,
                "header_row_count": table.headerRowCount,
                "style": (
                    {
                        "name": style.name,
                        "show_first_column": bool(style.showFirstColumn),
                        "show_last_column": bool(style.showLastColumn),
                        "show_row_stripes": bool(style.showRowStripes),
                        "show_column_stripes": bool(style.showColumnStripes),
                    }
                    if style is not None
                    else None
                ),
            }
        )
    return sorted(records, key=lambda record: record["name"])


def _column_widths(worksheet: Any) -> dict[str, float]:
    return {
        key: dimension.width
        for key, dimension in worksheet.column_dimensions.items()
        if dimension.width is not None
    }


def _row_heights(worksheet: Any) -> dict[str, float]:
    return {
        str(key): dimension.height
        for key, dimension in worksheet.row_dimensions.items()
        if dimension.height is not None
    }


def cell_projection(cell: Any) -> dict[str, Any]:
    return {
        "value": encode_value(cell.value),
        "font": font_projection(cell),
        "fill": fill_projection(cell),
        "number_format": cell.number_format,
        "border": border_projection(cell),
        "alignment": alignment_projection(cell),
    }


def compare_workbooks(old_path: Path, new_path: Path) -> dict[str, Any]:
    old = load_workbook(old_path, data_only=False)
    new = load_workbook(new_path, data_only=False)
    differences: list[dict[str, Any]] = []
    old_names = old.sheetnames
    new_names = new.sheetnames
    if old_names != new_names:
        differences.append(
            {
                "kind": "sheet_order",
                "location": "workbook",
                "old": old_names,
                "new": new_names,
            }
        )
    for sheet_name in dict.fromkeys(old_names + new_names):
        if sheet_name not in old_names or sheet_name not in new_names:
            old_sheet = old[sheet_name] if sheet_name in old_names else None
            new_sheet = new[sheet_name] if sheet_name in new_names else None
            differences.append(
                {
                    "kind": "sheet_presence",
                    "location": sheet_name,
                    "old": old_sheet is not None,
                    "new": new_sheet is not None,
                }
            )
            present_sheet = old_sheet or new_sheet
            assert present_sheet is not None
            for row_number in range(1, present_sheet.max_row + 1):
                for column_number in range(1, present_sheet.max_column + 1):
                    projection = cell_projection(
                        present_sheet.cell(row_number, column_number)
                    )
                    differences.append(
                        {
                            "kind": "cell",
                            "sheet": sheet_name,
                            "cell": (
                                f"{get_column_letter(column_number)}{row_number}"
                            ),
                            "differences": {
                                "cell_presence": {
                                    "old": projection if old_sheet else None,
                                    "new": projection if new_sheet else None,
                                }
                            },
                        }
                    )
            continue
        old_sheet = old[sheet_name]
        new_sheet = new[sheet_name]
        structural = {
            "merged_ranges": (
                [str(item) for item in old_sheet.merged_cells.ranges],
                [str(item) for item in new_sheet.merged_cells.ranges],
            ),
            "column_widths": (_column_widths(old_sheet), _column_widths(new_sheet)),
            "row_heights": (_row_heights(old_sheet), _row_heights(new_sheet)),
            "tables": (_table_records(old_sheet), _table_records(new_sheet)),
        }
        for kind, (old_value, new_value) in structural.items():
            if old_value != new_value:
                differences.append(
                    {
                        "kind": kind,
                        "location": sheet_name,
                        "old": old_value,
                        "new": new_value,
                    }
                )
        max_row = max(old_sheet.max_row, new_sheet.max_row)
        max_column = max(old_sheet.max_column, new_sheet.max_column)
        for row_number in range(1, max_row + 1):
            for column_number in range(1, max_column + 1):
                old_cell = old_sheet.cell(row_number, column_number)
                new_cell = new_sheet.cell(row_number, column_number)
                old_projection = cell_projection(old_cell)
                new_projection = cell_projection(new_cell)
                if old_projection == new_projection:
                    continue
                property_differences = {
                    key: {
                        "old": old_projection[key],
                        "new": new_projection[key],
                    }
                    for key in old_projection
                    if old_projection[key] != new_projection[key]
                }
                differences.append(
                    {
                        "kind": "cell",
                        "sheet": sheet_name,
                        "cell": f"{get_column_letter(column_number)}{row_number}",
                        "differences": property_differences,
                    }
                )
    by_kind = Counter(item["kind"] for item in differences)
    property_counts: Counter[str] = Counter()
    for item in differences:
        if item["kind"] == "cell":
            property_counts.update(item["differences"].keys())
    return {
        "old": str(old_path),
        "new": str(new_path),
        "difference_count": len(differences),
        "cell_difference_count": by_kind["cell"],
        "structural_difference_count": len(differences) - by_kind["cell"],
        "cell_property_counts": dict(sorted(property_counts.items())),
        "differences": differences,
    }


def _sheet_by_name(data: dict[str, Any], name: str) -> dict[str, Any]:
    for sheet in data["sheets"]:
        if sheet["name"] == name:
            return sheet
    raise RenderError(f"Required sheet not found: {name!r}")


def _header_map(sheet: dict[str, Any]) -> dict[str, int]:
    if not sheet["values"]:
        raise RenderError(f"Sheet {sheet['name']!r} has no header row")
    return {
        value: index
        for index, value in enumerate(sheet["values"][0])
        if isinstance(value, str)
    }


def _full_font(style: dict[str, Any]) -> dict[str, Any]:
    value = dict(FONT_DEFAULTS)
    value.update(style.get("font", {}))
    return value


def _full_fill(style: dict[str, Any]) -> dict[str, Any]:
    value = dict(FILL_DEFAULTS)
    value.update(style.get("fill", {}))
    return value


def _sparse_font(full_font: dict[str, Any]) -> dict[str, Any]:
    return {
        key: full_font[key]
        for key in FONT_FIELDS
        if full_font[key] != FONT_DEFAULTS[key]
    }


def _style_for_coordinate(sheet: dict[str, Any], coordinate: str) -> dict[str, Any]:
    return sheet["styles"].get(coordinate, {})


def find_excluded_row_pattern(
    sheet: dict[str, Any], exclusion_column_index: int
) -> tuple[list[dict[str, Any]], int]:
    patterns: Counter[str] = Counter()
    decoded_patterns: dict[str, list[dict[str, Any]]] = {}
    for row_index, row in enumerate(sheet["values"], start=1):
        if row[exclusion_column_index] != "Non-URC match":
            continue
        pattern = []
        for column_index in range(sheet["max_column"]):
            coordinate = f"{get_column_letter(column_index + 1)}{row_index}"
            style = _style_for_coordinate(sheet, coordinate)
            pattern.append(
                {"font": _full_font(style), "fill": _full_fill(style)}
            )
        key = json.dumps(pattern, sort_keys=True)
        patterns[key] += 1
        decoded_patterns[key] = pattern
    if not patterns:
        raise RenderError("No existing 'Non-URC match' row was found")
    key, count = patterns.most_common(1)[0]
    return decoded_patterns[key], count


def mark_excluded_data(
    data: dict[str, Any],
    audit_rows: Iterable[dict[str, str]],
    reason: str,
    *,
    require_production_counts: bool = False,
) -> tuple[dict[str, Any], dict[str, Any]]:
    sheet = _sheet_by_name(data, MASTER_SHEET)
    headers = _header_map(sheet)
    required_headers = ("Problem type", "Occasion category", EXCLUSION_REASON_HEADER)
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        raise RenderError(f"Missing required headers: {missing_headers}")
    problem_column = headers["Problem type"]
    occasion_column = headers["Occasion category"]
    exclusion_column = headers[EXCLUSION_REASON_HEADER]
    pattern, pattern_count = find_excluded_row_pattern(sheet, exclusion_column)
    rows = list(audit_rows)
    source_rows: list[int] = []
    occasion_counts: Counter[str] = Counter()
    errors: list[str] = []
    for audit_index, audit_row in enumerate(rows, start=2):
        raw_source_row = audit_row.get("source_workbook_row")
        try:
            source_row = int(raw_source_row or "")
        except ValueError:
            errors.append(
                f"audit CSV row {audit_index}: invalid source_workbook_row "
                f"{raw_source_row!r}"
            )
            continue
        source_rows.append(source_row)
        if source_row < 2 or source_row > len(sheet["values"]):
            errors.append(
                f"audit CSV row {audit_index}: source workbook row "
                f"{source_row} is out of range"
            )
            continue
        values = sheet["values"][source_row - 1]
        problem = values[problem_column]
        occasion = values[occasion_column]
        exclusion = values[exclusion_column]
        if problem != "Injury":
            errors.append(
                f"source workbook row {source_row}: Problem type is "
                f"{problem!r}, expected 'Injury'"
            )
        if occasion not in EXPECTED_OCCASION_COUNTS:
            errors.append(
                f"source workbook row {source_row}: Occasion category is "
                f"{occasion!r}, expected 'Non-Rugby' or 'Gym-Based'"
            )
        else:
            occasion_counts[occasion] += 1
        if exclusion not in (None, ""):
            errors.append(
                f"source workbook row {source_row}: Exclusion Reason is "
                f"{exclusion!r}, expected blank"
            )
    duplicate_rows = sorted(
        row for row, count in Counter(source_rows).items() if count > 1
    )
    if duplicate_rows:
        errors.append(f"duplicate source_workbook_row values: {duplicate_rows}")
    if require_production_counts:
        if len(rows) != EXPECTED_MARK_COUNT:
            errors.append(
                f"audit row count is {len(rows)}, expected {EXPECTED_MARK_COUNT}"
            )
        if dict(occasion_counts) != EXPECTED_OCCASION_COUNTS:
            errors.append(
                "occasion counts are "
                f"{dict(occasion_counts)}, expected {EXPECTED_OCCASION_COUNTS}"
            )
    if errors:
        raise RenderError(
            "Audit precondition check failed; no rows were changed:\n"
            + "\n".join(f"- {error}" for error in errors)
        )
    for source_row in source_rows:
        sheet["values"][source_row - 1][exclusion_column] = reason
        for column_index, treatment in enumerate(pattern, start=1):
            coordinate = f"{get_column_letter(column_index)}{source_row}"
            style = sheet["styles"].setdefault(coordinate, {})
            sparse_font = _sparse_font(treatment["font"])
            if sparse_font:
                style["font"] = sparse_font
            else:
                style.pop("font", None)
            if treatment["fill"] != FILL_DEFAULTS:
                style["fill"] = treatment["fill"]
            else:
                style.pop("fill", None)
            if not style:
                sheet["styles"].pop(coordinate, None)
    unique_treatments = {
        json.dumps(item, sort_keys=True): item for item in pattern
    }
    applied_pattern: dict[str, Any]
    if len(unique_treatments) == 1:
        applied_pattern = {
            "scope": "columns A:AB",
            **next(iter(unique_treatments.values())),
        }
    else:
        applied_pattern = {
            "scope": "per-column A:AB",
            "columns": {
                get_column_letter(index): treatment
                for index, treatment in enumerate(pattern, start=1)
            },
        }
    summary = {
        "rows_marked": len(source_rows),
        "occasion_counts": dict(sorted(occasion_counts.items())),
        "source_workbook_rows": source_rows,
        "dominant_non_urc_pattern_rows": pattern_count,
        "excluded_row_style_pattern": applied_pattern,
    }
    return data, summary


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def command_extract(args: argparse.Namespace) -> int:
    payload = extract_workbook(args.workbook)
    _write_json(args.out, payload)
    print(
        f"Extracted {len(payload['sheets'])} sheets to {args.out} "
        f"(sha256 {sha256_file(args.out)})"
    )
    return 0


def command_render(args: argparse.Namespace) -> int:
    render_workbook(_read_json(args.data), args.out)
    print(f"Rendered {args.out} (sha256 {sha256_file(args.out)})")
    return 0


def command_compare(args: argparse.Namespace) -> int:
    report = compare_workbooks(args.old, args.new)
    if args.report:
        _write_json(args.report, report)
    print(
        "Comparison summary: "
        f"{report['difference_count']} differences, "
        f"{report['cell_difference_count']} cells, "
        f"{report['structural_difference_count']} structural"
    )
    print(
        "Cell property differences: "
        + json.dumps(report["cell_property_counts"], sort_keys=True)
    )
    if args.report:
        print(f"Detailed report: {args.report}")
    return 1 if report["difference_count"] else 0


def command_mark_excluded(args: argparse.Namespace) -> int:
    data = _read_json(args.data)
    with args.audit.open(newline="", encoding="utf-8-sig") as handle:
        audit_rows = list(csv.DictReader(handle))
    updated, summary = mark_excluded_data(
        data,
        audit_rows,
        args.reason,
        require_production_counts=True,
    )
    _write_json(args.out, updated)
    print(f"Marked {summary['rows_marked']} rows excluded")
    print(f"Occasion counts: {json.dumps(summary['occasion_counts'], sort_keys=True)}")
    print(
        "Excluded-row style pattern: "
        + json.dumps(summary["excluded_row_style_pattern"], sort_keys=True)
    )
    print(f"Wrote {args.out} (sha256 {sha256_file(args.out)})")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Canonical URC master workbook extraction and rendering"
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    extract_parser = subparsers.add_parser("extract")
    extract_parser.add_argument("--workbook", type=Path, required=True)
    extract_parser.add_argument("--out", type=Path, required=True)
    extract_parser.set_defaults(handler=command_extract)

    render_parser = subparsers.add_parser("render")
    render_parser.add_argument("--data", type=Path, required=True)
    render_parser.add_argument("--out", type=Path, required=True)
    render_parser.set_defaults(handler=command_render)

    compare_parser = subparsers.add_parser("compare")
    compare_parser.add_argument("--old", type=Path, required=True)
    compare_parser.add_argument("--new", type=Path, required=True)
    compare_parser.add_argument("--report", type=Path)
    compare_parser.set_defaults(handler=command_compare)

    mark_parser = subparsers.add_parser("mark-excluded")
    mark_parser.add_argument("--data", type=Path, required=True)
    mark_parser.add_argument("--audit", type=Path, required=True)
    mark_parser.add_argument("--reason", required=True)
    mark_parser.add_argument("--out", type=Path, required=True)
    mark_parser.set_defaults(handler=command_mark_excluded)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return args.handler(args)
    except (OSError, KeyError, RenderError, json.JSONDecodeError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
