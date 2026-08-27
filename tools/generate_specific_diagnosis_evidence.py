#!/usr/bin/env python3
"""Build deterministic, pseudonymised evidence for reviewed 2024-25 diagnoses.

The review workbook supplies the final Specific Diagnosis value. The grouped
summary supplies the mutually exclusive reporting group. The immutable v5
master supplies the stable source-row identity and its existing 28-field row
hash. No database access or direct athlete identifier is used.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import tempfile
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SEASON = "2024-25"
SCHEMA_VERSION = "urc_2024-25_specific_diagnosis_evidence_v1"

DEFAULT_WORKBOOK = ROOT / "data/2024-25/review/urc_injury_master_review_2024-25.xlsx"
DEFAULT_SUMMARY = ROOT / "data/2024-25/review/urc_specific_diagnosis_summary_2024-25.xlsx"
DEFAULT_MASTER = ROOT / "data/2024-25/master/master_2024-25_v5.json"
DEFAULT_OUTPUT = ROOT / "docs/evidence/urc_2024-25_specific_diagnosis_evidence.json"

EXPECTED_WORKBOOK_SHA256 = "4f1db130f9f5aff23c3473eb2ab64a467f739a0b6ac7e4f170ca0383d9072b73"
EXPECTED_SUMMARY_SHA256 = "158cc822298c7478360c2a2f7c39fd85712398ad47a24e8848653ba686ad3c00"
EXPECTED_MASTER_SHA256 = "15b9af0da05aa57698487f4c8ebacf9923cec4e66846ac00b76fa3c2b75f2f63"

MASTER_FIELDS = [
    "Team",
    "PlayerID",
    "Received At Club",
    "Received/Injured In Team",
    "Problem type",
    "Date Injured",
    "Fit For Selection Date",
    "Confirmed Return Date",
    "Days Injured",
    "Occasion category",
    "Body Part",
    "Orchard Code",
    "Illness Code",
    "Description",
    "Injury Tissue Type/s",
    "Side",
    "Nature of onset",
    "Recurrence",
    "Is Contact",
    "Mechanism of Injury",
    "Mechanism Notes",
    "Injury Surface Type",
    "Match Type",
    "Received At Position",
    "Required Surgery",
    "TimeLoss vs Medical Attention",
    "Diagnosis",
    "Exclusion Reason",
]

REVIEW_FIELDS = [
    *MASTER_FIELDS[:2],
    "Reporting At Club",
    *MASTER_FIELDS[3:],
    "Specific Diagnosis",
]
SUMMARY_MAPPING_FIELDS = ["Specific diagnosis", "Diagnosis group", "Mapping source"]
SUMMARY_GROUPED_FIELDS = [
    "Team",
    "Problem type",
    "Specific diagnosis",
    "Diagnosis group",
    "Body part",
    "Tissue type",
    "Workbook row",
]


class EvidenceError(ValueError):
    """Raised when a source artefact cannot support the evidence contract."""


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def compact_json_sha256(value: Any) -> str:
    payload = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _master_value(value: Any) -> Any:
    if isinstance(value, dict) and value.get("$type") == "datetime":
        raw = str(value.get("value", ""))[:10]
        return f"{raw[8:10]}/{raw[5:7]}/{raw[:4]}"
    return "" if value is None else value


def _workbook_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_blank(value: Any) -> bool:
    return not _text(value)


def load_master(path: Path) -> tuple[dict[int, dict[str, Any]], str]:
    actual_hash = sha256_file(path)
    if actual_hash != EXPECTED_MASTER_SHA256:
        raise EvidenceError(
            f"immutable master fingerprint changed: {actual_hash} != {EXPECTED_MASTER_SHA256}"
        )
    document = json.loads(path.read_text(encoding="utf-8"))
    sheets = [sheet for sheet in document.get("sheets", []) if sheet.get("name") == "Injury Master"]
    if len(sheets) != 1 or not isinstance(sheets[0].get("values"), list):
        raise EvidenceError("immutable master lacks one Injury Master value table")
    values = sheets[0]["values"]
    if values[0] != MASTER_FIELDS:
        raise EvidenceError("immutable master 28-field header order changed")
    if len(values) != 3061:
        raise EvidenceError(f"immutable master row count changed: {len(values) - 1}")
    rows: dict[int, dict[str, Any]] = {}
    for source_row, raw in enumerate(values[1:], start=2):
        padded = list(raw) + [None] * (len(MASTER_FIELDS) - len(raw))
        if len(padded) != len(MASTER_FIELDS):
            raise EvidenceError(f"master source row {source_row} has the wrong width")
        row_values = {field: _master_value(value) for field, value in zip(MASTER_FIELDS, padded)}
        row_hash = compact_json_sha256(row_values)
        rows[source_row] = {
            "values": row_values,
            "source_row_sha256": row_hash,
            "source_identity": f"{SEASON}:{source_row}:{row_hash}",
        }
    return rows, actual_hash


def load_review_workbook(
    path: Path, master_rows: dict[int, dict[str, Any]]
) -> tuple[dict[int, dict[str, Any]], str, Counter[str], int, int]:
    actual_hash = sha256_file(path)
    if actual_hash != EXPECTED_WORKBOOK_SHA256:
        raise EvidenceError(
            f"authoritative review workbook fingerprint changed: {actual_hash} != {EXPECTED_WORKBOOK_SHA256}"
        )
    workbook = load_workbook(path, data_only=False, read_only=True)
    if workbook.sheetnames != ["Injury Master", "Fixtures"]:
        raise EvidenceError("authoritative review workbook sheet order changed")
    sheet = workbook["Injury Master"]
    iterator = sheet.iter_rows(values_only=False)
    try:
        header_cells = next(iterator)
    except StopIteration as error:
        raise EvidenceError("authoritative review workbook is empty") from error
    headers = [_text(cell.value) for cell in header_cells]
    if headers != REVIEW_FIELDS:
        raise EvidenceError("authoritative review workbook header order changed")

    rows: dict[int, dict[str, Any]] = {}
    problem_types: Counter[str] = Counter()
    workbook_row_count = 0
    master_hash_matches = 0
    for workbook_row, cells in enumerate(iterator, start=2):
        workbook_row_count += 1
        values = [cell.value for cell in cells]
        if len(values) != len(REVIEW_FIELDS):
            raise EvidenceError(f"review workbook row {workbook_row} has the wrong width")
        if any(cell.data_type == "f" for cell in cells):
            raise EvidenceError(f"formula found in review workbook row {workbook_row}")
        row = dict(zip(REVIEW_FIELDS, values))
        if workbook_row not in master_rows:
            raise EvidenceError(f"review workbook has an unexpected row {workbook_row}")
        master_values = master_rows[workbook_row]["values"]
        if _text(row["Team"]) != _text(master_values["Team"]) or _text(row["PlayerID"]) != _text(master_values["PlayerID"]):
            raise EvidenceError(f"review workbook row identity moved at row {workbook_row}")
        problem_type = _text(row["Problem type"])
        problem_types[problem_type] += 1
        review_values = {
            field: _workbook_value(
                row["Reporting At Club"] if field == "Received At Club" else row[field]
            )
            for field in MASTER_FIELDS
        }
        canonical_values = {field: _text(value) for field, value in master_values.items()}
        if review_values == canonical_values:
            master_hash_matches += 1
        rows[workbook_row] = row
    if workbook_row_count != len(master_rows):
        raise EvidenceError(
            f"review workbook row count changed: {workbook_row_count} != {len(master_rows)}"
        )
    return rows, actual_hash, problem_types, workbook_row_count, master_hash_matches


def _diagnosis_group_code(label: str) -> str:
    if label.casefold() in {"diagnosis not specified", "unknown", "unknown diagnosis"}:
        return "unknown"
    folded = unicodedata.normalize("NFKD", label).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "_", folded.casefold()).strip("_") or "unknown"
    return f"dx_{slug}_{hashlib.sha256(label.encode('utf-8')).hexdigest()[:10]}"


def load_summary(path: Path) -> tuple[dict[str, dict[str, str]], dict[int, dict[str, Any]], str]:
    actual_hash = sha256_file(path)
    if actual_hash != EXPECTED_SUMMARY_SHA256:
        raise EvidenceError(
            f"grouped summary workbook fingerprint changed: {actual_hash} != {EXPECTED_SUMMARY_SHA256}"
        )
    workbook = load_workbook(path, data_only=True, read_only=True)
    expected_sheets = [
        "League Summary",
        "Team Summary",
        "Group Mapping",
        "Grouped Data",
        "Unresolved Review",
        "Method",
    ]
    if workbook.sheetnames != expected_sheets:
        raise EvidenceError("grouped summary sheet order changed")

    mapping_sheet = workbook["Group Mapping"]
    mapping_iterator = mapping_sheet.iter_rows(values_only=True)
    mapping_header = list(next(mapping_iterator))
    if mapping_header != SUMMARY_MAPPING_FIELDS:
        raise EvidenceError("grouped summary mapping header changed")
    mapping: dict[str, dict[str, str]] = {}
    for raw_label, group_label, mapping_source in mapping_iterator:
        raw = _text(raw_label)
        group = _text(group_label)
        source = _text(mapping_source)
        if not raw or not group or not source:
            raise EvidenceError("grouped summary contains a blank mapping value")
        if raw in mapping:
            raise EvidenceError(f"duplicate Specific Diagnosis mapping: {raw!r}")
        mapping[raw] = {
            "specific_diagnosis_source_label": raw,
            "diagnosis_group_label": group,
            "mapping_source": source,
            "diagnosis_group_code": _diagnosis_group_code(group),
        }

    grouped_sheet = workbook["Grouped Data"]
    grouped_iterator = grouped_sheet.iter_rows(values_only=True)
    grouped_header = list(next(grouped_iterator))
    if grouped_header != SUMMARY_GROUPED_FIELDS:
        raise EvidenceError("grouped summary data header changed")
    grouped: dict[int, dict[str, Any]] = {}
    for values in grouped_iterator:
        row = dict(zip(SUMMARY_GROUPED_FIELDS, values))
        source_row = row["Workbook row"]
        if not isinstance(source_row, int) or source_row <= 1:
            raise EvidenceError(f"invalid grouped summary workbook row: {source_row!r}")
        if source_row in grouped:
            raise EvidenceError(f"duplicate grouped summary workbook row: {source_row}")
        raw = _text(row["Specific diagnosis"])
        mapped = mapping.get(raw)
        if mapped is None:
            raise EvidenceError(f"grouped summary row {source_row} has no source-label mapping")
        if mapped["diagnosis_group_label"] != _text(row["Diagnosis group"]):
            raise EvidenceError(f"diagnosis group mismatch at summary row {source_row}")
        grouped[source_row] = {
            "team": _text(row["Team"]),
            "problem_type": _text(row["Problem type"]),
            "specific_diagnosis_source_label": raw,
            "diagnosis_group_label": _text(row["Diagnosis group"]),
            "body_part": _text(row["Body part"]),
            "tissue_type": _text(row["Tissue type"]),
            "mapping_source": mapped["mapping_source"],
            "diagnosis_group_code": mapped["diagnosis_group_code"],
        }
    if len(mapping) != 554:
        raise EvidenceError(f"grouped summary mapping row count changed: {len(mapping)}")
    if len(grouped) != 2052:
        raise EvidenceError(f"grouped summary data row count changed: {len(grouped)}")
    if len({row["diagnosis_group_label"] for row in mapping.values()}) != 310:
        raise EvidenceError("diagnosis group label count changed")
    if len({row["diagnosis_group_code"] for row in mapping.values()}) != 310:
        raise EvidenceError("diagnosis group code collision or unknown-code contract changed")
    return mapping, grouped, actual_hash


def _atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        newline="\n",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temporary_path = Path(handle.name)
        json.dump(payload, handle, ensure_ascii=False, indent=2, sort_keys=True)
        handle.write("\n")
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temporary_path, path)


def build_evidence(
    workbook_path: Path = DEFAULT_WORKBOOK,
    summary_path: Path = DEFAULT_SUMMARY,
    master_path: Path = DEFAULT_MASTER,
) -> dict[str, Any]:
    master_rows, master_hash = load_master(master_path)
    (
        review_rows,
        workbook_hash,
        workbook_problem_types,
        workbook_row_count,
        review_hash_matches,
    ) = load_review_workbook(workbook_path, master_rows)
    mapping, grouped, summary_hash = load_summary(summary_path)

    evidence_rows: list[dict[str, Any]] = []
    summary_problem_types: Counter[str] = Counter()
    team_counts: Counter[str] = Counter()
    group_counts: Counter[str] = Counter()
    summary_hash_matches = 0
    for source_row in sorted(grouped):
        grouped_row = grouped[source_row]
        review_row = review_rows[source_row]
        master = master_rows[source_row]
        current_problem_type = _text(review_row["Problem type"])
        if current_problem_type != grouped_row["problem_type"]:
            raise EvidenceError(f"problem type mismatch at workbook row {source_row}")
        if _text(review_row["Specific Diagnosis"]) != grouped_row["specific_diagnosis_source_label"]:
            raise EvidenceError(f"Specific Diagnosis mismatch at workbook row {source_row}")
        if _text(review_row["Team"]) != grouped_row["team"]:
            raise EvidenceError(f"team mismatch at workbook row {source_row}")
        if _text(review_row["Exclusion Reason"]):
            raise EvidenceError(f"grouped summary includes excluded workbook row {source_row}")
        canonical_problem_type = _text(master["values"]["Problem type"])
        if canonical_problem_type != current_problem_type:
            raise EvidenceError(f"canonical problem type mismatch at workbook row {source_row}")

        review_values = {
            field: _workbook_value(review_row["Reporting At Club"] if field == "Received At Club" else review_row[field])
            for field in MASTER_FIELDS
        }
        master_values = {field: _text(value) for field, value in master["values"].items()}
        summary_hash_matches += review_values == master_values
        is_injury = current_problem_type.casefold() == "injury"
        problem_type_code = current_problem_type.casefold().replace(" ", "_")
        row = {
            "season": SEASON,
            "master_source_row": source_row,
            "master_source_identity": master["source_identity"],
            "source_row_sha256": master["source_row_sha256"],
            "team": grouped_row["team"],
            "problem_type": current_problem_type,
            "problem_type_code": problem_type_code,
            "injury_metric_eligible": is_injury,
            "summary_included": True,
            "specific_diagnosis_source_label": grouped_row["specific_diagnosis_source_label"],
            "diagnosis_group_code": grouped_row["diagnosis_group_code"],
            "diagnosis_group_label": grouped_row["diagnosis_group_label"],
            "mapping_source": grouped_row["mapping_source"],
            "body_part": grouped_row["body_part"],
            "tissue_type": grouped_row["tissue_type"],
            "master_source_row_hash_matches_review_workbook": review_values == master_values,
        }
        evidence_rows.append(row)
        summary_problem_types[current_problem_type] += 1
        if is_injury:
            team_counts[grouped_row["team"]] += 1
            group_counts[grouped_row["diagnosis_group_code"]] += 1

    expected_team_names = 16
    if len(team_counts) != expected_team_names:
        raise EvidenceError(f"summary team count changed: {len(team_counts)}")
    if sum(team_counts.values()) != 1660:
        raise EvidenceError(f"injury row reconciliation failed: {sum(team_counts.values())} != 1660")
    if sum(group_counts.values()) != 1660:
        raise EvidenceError("injury group reconciliation failed")
    if summary_problem_types["Illness"] != 392:
        raise EvidenceError("illness row reconciliation failed")
    if any(row["injury_metric_eligible"] is False and row["problem_type"] == "Injury" for row in evidence_rows):
        raise EvidenceError("injury problem type was marked ineligible")
    if any(row["injury_metric_eligible"] and row["problem_type"] != "Injury" for row in evidence_rows):
        raise EvidenceError("non-injury problem type entered injury metrics")

    mapping_rows = [mapping[label] for label in sorted(mapping)]
    rows_hash = compact_json_sha256(evidence_rows)
    mapping_hash = compact_json_sha256(mapping_rows)
    return {
        "schema_version": SCHEMA_VERSION,
        "season": SEASON,
        "status": "prepared_not_executed",
        "scope": {
            "row_source": "Grouped Data from the authoritative specific-diagnosis summary workbook",
            "included_rows": 2052,
            "metric_problem_type": "Injury",
            "excluded_problem_type": "Illness",
            "exclusion_rule": "Only rows whose canonical Problem type is exactly Injury may enter diagnosis metrics. Illness rows remain evidence-visible and are marked injury_metric_eligible=false.",
            "source_label_visibility": "specific_diagnosis_source_label is retained in this private evidence artefact; dashboards should use diagnosis_group_code and diagnosis_group_label.",
            "unmapped_effective_cohort_fallback": {
                "rows": 4,
                "source_rows": [603, 1120, 1121, 1122],
                "diagnosis_group_code": "unknown",
                "diagnosis_group_label": "Unknown diagnosis",
                "note": "Four injury rows in the correction-aware effective cohort are outside the grouped summary scope and remain unknown until later row-level mappings bind them.",
            },
        },
        "source_artifacts": {
            "review_workbook": {
                "path": str(workbook_path.relative_to(ROOT)) if workbook_path.is_relative_to(ROOT) else str(workbook_path),
                "sha256": workbook_hash,
                "sheet": "Injury Master",
                "data_rows": workbook_row_count,
                "columns": 29,
                "problem_type_counts": dict(sorted(workbook_problem_types.items())),
            },
            "grouped_summary_workbook": {
                "path": str(summary_path.relative_to(ROOT)) if summary_path.is_relative_to(ROOT) else str(summary_path),
                "sha256": summary_hash,
                "group_mapping_rows": len(mapping_rows),
                "grouped_data_rows": len(grouped),
                "diagnosis_group_count": len({row["diagnosis_group_label"] for row in mapping_rows}),
            },
            "immutable_master": {
                "path": str(master_path.relative_to(ROOT)) if master_path.is_relative_to(ROOT) else str(master_path),
                "sha256": master_hash,
                "row_hash_fields": MASTER_FIELDS,
                "row_hash_method": "SHA-256 of compact UTF-8 JSON with sorted keys over the 28 replay-serialised A:AB fields",
            },
        },
        "mapping": {
            "diagnosis_group_code_method": "dx_ plus an ASCII slug of the exact group label plus the first 10 hexadecimal characters of its SHA-256",
            "source_label_mapping_rows": len(mapping_rows),
            "source_label_mapping_sha256": mapping_hash,
            "rows_sha256": rows_hash,
        },
        "aggregate_reconciliation": {
            "master_rows": len(master_rows),
            "master_problem_type_counts": dict(sorted(workbook_problem_types.items())),
            "summary_rows": len(evidence_rows),
            "summary_problem_type_counts": dict(sorted(summary_problem_types.items())),
            "injury_metric_rows": sum(team_counts.values()),
            "illness_rows_excluded_from_injury_metrics": summary_problem_types["Illness"],
            "injury_rows_with_illness_flag": 0,
            "team_injury_rows": dict(sorted(team_counts.items())),
            "diagnosis_groups_with_injury_rows": len(group_counts),
            "unknown_group_injury_rows": group_counts.get("unknown", 0),
            "diagnosis_groups_in_mapping": len({row["diagnosis_group_label"] for row in mapping_rows}),
            "review_rows_matching_immutable_master_28_fields": review_hash_matches,
            "review_rows_with_expected_source_value_differences": workbook_row_count - review_hash_matches,
            "summary_rows_matching_immutable_master_28_fields": summary_hash_matches,
            "summary_rows_with_source_value_differences": len(evidence_rows) - summary_hash_matches,
        },
        "source_label_mapping": mapping_rows,
        "rows": evidence_rows,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate 2024-25 specific-diagnosis evidence")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--summary", type=Path, default=DEFAULT_SUMMARY)
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    payload = build_evidence(args.workbook, args.summary, args.master)
    _atomic_write_json(args.output, payload)
    print(
        json.dumps(
            {
                "output": str(args.output),
                "rows": len(payload["rows"]),
                "injury_metric_rows": payload["aggregate_reconciliation"]["injury_metric_rows"],
                "illness_rows_excluded_from_injury_metrics": payload["aggregate_reconciliation"]["illness_rows_excluded_from_injury_metrics"],
                "diagnosis_groups": payload["aggregate_reconciliation"]["diagnosis_groups_in_mapping"],
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
