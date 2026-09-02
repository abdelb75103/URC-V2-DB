#!/usr/bin/env python3
"""Generate row-preserving high-speed-running exposure artefacts.

The two season mapping files are the only production configuration.  A mapping
entry identifies the accepted exposure CSV, its checksum, and either an HSR
field already present on that row or a checksum-bound external source row.
The generator never filters exposure rows and never changes an existing field
except the canonical HSR field it is responsible for.

The supported mapping shape is intentionally small::

    {
      "schema_version": "urc_hsr_mapping_v1",
      "season": "2024-25",
      "teams": [{
        "team_key": "cardiff",
        "accepted_exposure": {"path": "...", "sha256": "..."},
        "hsr": {
          "mode": "accepted_row",
          "source_field": "high speed running distance",
          "units": "m",
          "threshold_or_zone": "vendor-defined"
        }
      }]
    }

``external_row`` uses one checksum-bound source. ``locator_files`` resolves
monthly or multi-file sources from accepted file, sheet and physical-row
locators. ``observation_group`` uses explicit join pairs, while
``long_variable_files`` validates the duration and distance components of a
long-format observation before taking its HSR distance row. All source-backed
modes require the source field to be declared distance-like.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping, Sequence


ROOT = Path(__file__).resolve().parents[1]
SEASONS = ("2024-25", "2025-26")
EXPECTED_TEAMS = frozenset(
    {
        "benetton",
        "cardiff",
        "connacht",
        "dragons",
        "edinburgh",
        "glasgow",
        "leinster",
        "lions",
        "munster",
        "ospreys",
        "scarlets",
        "sharks",
        "stormers",
        "ulster",
        "bulls",
        "zebre",
    }
)
EXPECTED_AVAILABILITY = {"2024-25": 16, "2025-26": 14}
EXPECTED_GAPS = {"2024-25": frozenset(), "2025-26": frozenset({"benetton", "edinburgh"})}
CANONICAL_HSR_FIELD = "high speed running distance"
DATABASE_PARAMETER_SCHEMA = "urc_hsr_database_parameters_v1"
DATABASE_PARAMETER_FILENAME = "hsr_database_parameters.json"
DEFAULT_BLANK_REASON_FIELD = "high speed running distance blank reason"
DEFAULT_LOCATOR_FIELDS = (
    "source_file_sha256",
    "source_sheet",
    "source_row_number",
    "source_row_sha256",
)
HSR_PROVENANCE_FIELDS = (
    "hsr_source_file_sha256",
    "hsr_source_sheet",
    "hsr_source_row_number",
    "hsr_source_row_sha256",
)
MISSING_REASON_GAP = "hsr_source_unavailable_for_accepted_period"
MISSING_REASON_NO_ROW = "hsr_source_row_not_found"
MISSING_REASON_BLANK = "hsr_source_value_blank"
MISSING_REASON_TOKEN = "hsr_source_value_missing_token"
MISSING_REASON_EXCEEDS_TOTAL = "hsr_source_value_exceeds_total_distance"


class GenerationError(ValueError):
    """Raised when a mapping or source cannot be proved safe to use."""


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as handle:
            for block in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(block)
    except FileNotFoundError as exc:
        raise GenerationError(f"input file is missing: {path}") from exc
    return digest.hexdigest()


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _nonblank(value: Any) -> bool:
    return bool(_text(value).strip())


def _as_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    token = _text(value).strip().casefold()
    if token in {"true", "yes", "y", "1", "available"}:
        return True
    if token in {"false", "no", "n", "0", "unavailable", "gap"}:
        return False
    raise GenerationError(f"expected a boolean value, got {value!r}")


def _first(mapping: Mapping[str, Any], *names: str, default: Any = None) -> Any:
    for name in names:
        if name in mapping and mapping[name] is not None:
            return mapping[name]
    return default


def _as_path(value: Any, mapping_path: Path) -> Path:
    if not _nonblank(value):
        raise GenerationError(f"missing input path in mapping {mapping_path}")
    path = Path(_text(value))
    if not path.is_absolute():
        path = ROOT / path
    return path


def _display_path(path: Path) -> str:
    resolved = path.resolve()
    try:
        return resolved.relative_to(ROOT.resolve()).as_posix()
    except ValueError:
        try:
            return (Path("..") / resolved.relative_to(ROOT.parent.resolve())).as_posix()
        except ValueError:
            return f"external://{resolved.name}"


def _normalise_team(value: Any) -> str:
    token = re.sub(r"[^a-z0-9]+", "_", _text(value).casefold()).strip("_")
    aliases = {
        "cardiff_rugby": "cardiff",
        "cardiff_blues": "cardiff",
        "dragons_rfc": "dragons",
        "benetton_rugby": "benetton",
        "edinburgh_rugby": "edinburgh",
        "glasgow_warriors": "glasgow",
        "lions_rugby": "lions",
        "munster_rugby": "munster",
        "ospreys_rugby": "ospreys",
        "scarlets_rugby": "scarlets",
        "sharks_rugby": "sharks",
        "stormers_rugby": "stormers",
        "ulster_rugby": "ulster",
        "zebre_parma": "zebre",
    }
    return aliases.get(token, token)


def _entries(raw: Any, path: Path) -> list[dict[str, Any]]:
    if isinstance(raw, list):
        values = raw
    elif isinstance(raw, dict):
        value = _first(raw, "teams", "team_mappings", "team_seasons", "mappings")
        if isinstance(value, list):
            values = value
        elif isinstance(value, dict):
            values = [{**entry, "team_key": key} for key, entry in value.items()]
        else:
            raise GenerationError(f"mapping has no team entries: {path}")
    else:
        raise GenerationError(f"mapping must contain a JSON object: {path}")
    if not all(isinstance(item, dict) for item in values):
        raise GenerationError(f"mapping team entries must be objects: {path}")
    return [dict(item) for item in values]


def _expected_hash(value: Any, label: str) -> str:
    result = _text(value).strip().casefold()
    if not re.fullmatch(r"[0-9a-f]{64}", result):
        raise GenerationError(f"{label} must be a 64-character SHA-256")
    return result


def _normalise_locator_fields(value: Any, label: str) -> tuple[str, ...]:
    if value is None:
        return DEFAULT_LOCATOR_FIELDS
    if isinstance(value, str):
        values = [value]
    elif isinstance(value, list):
        values = value
    else:
        raise GenerationError(f"{label} must be a list of field names")
    fields = tuple(_text(item).strip() for item in values)
    if not fields or any(not field for field in fields) or len(set(fields)) != len(fields):
        raise GenerationError(f"{label} must contain unique nonblank field names")
    return fields


def _normalise_join_pairs(value: Any, accepted: Sequence[str], source: Sequence[str]) -> tuple[tuple[str, str], ...]:
    if value is None:
        if len(accepted) != len(source):
            raise GenerationError("accepted and source locator fields must have the same length")
        return tuple(zip(accepted, source))
    if isinstance(value, dict):
        accepted_value = _first(value, "accepted", "accepted_fields", "left")
        source_value = _first(value, "source", "source_fields", "right")
        if isinstance(accepted_value, list) and isinstance(source_value, list):
            if len(accepted_value) != len(source_value):
                raise GenerationError("accepted and source join field lists must have the same length")
            pairs = [(_text(a), _text(b)) for a, b in zip(accepted_value, source_value)]
        else:
            pairs = [(str(a), str(b)) for a, b in value.items()]
    elif isinstance(value, list):
        if all(isinstance(item, str) for item in value):
            if len(value) != len(accepted) or len(value) != len(source):
                raise GenerationError("string join_fields must match locator field lengths")
            return tuple(zip(accepted, value))
        pairs = []
        for item in value:
            if isinstance(item, dict):
                left = _first(item, "accepted", "accepted_field", "left")
                right = _first(item, "source", "source_field", "right")
                if not _nonblank(left) or not _nonblank(right):
                    raise GenerationError("join field objects require accepted and source fields")
                pairs.append((_text(left), _text(right)))
            elif isinstance(item, (list, tuple)) and len(item) == 2:
                pairs.append((_text(item[0]), _text(item[1])))
            else:
                raise GenerationError("join_fields must contain accepted/source pairs")
    else:
        raise GenerationError("join_fields must be an object or list")
    if not pairs or len({a for a, _ in pairs}) != len(pairs) or len({b for _, b in pairs}) != len(pairs):
        raise GenerationError("join_fields must contain unique fields on both sides")
    return tuple(pairs)


_FORBIDDEN_DISTANCE_WORDS = re.compile(
    r"(?:^|[^a-z])(?:count|counts|number|numbers|effort|efforts|rate|rates|ratio|percent|percentage|pct|"
    r"duration|minute|minutes|time|sprint|sprints|acceleration|accelerations|deceleration|decelerations|"
    r"tackle|tackles|carry|carries|impact|impacts|frequency)(?:$|[^a-z])"
)
_POSITIVE_DISTANCE_WORDS = re.compile(
    r"(?:distance|dist|metre|metres|meter|meters|kilometre|kilometres|kilometer|kilometers|\(m\)|\bkm\b)"
)
_HSR_NAME = re.compile(r"(?:^|[^a-z])h(?:igh)?\s*speed\s*running(?:$|[^a-z])|\bhsr(?:m)?\b")
_HSR_DISTANCE_NAME = re.compile(
    r"(?:high\s*speed|\bhsr(?:m)?\b|\bhs\s*distance\b|speed\s*zone|distance\s*z\s*\d)"
)


def is_distance_like_field(field: str, config: Mapping[str, Any]) -> bool:
    """Return whether explicit field metadata and the name identify distance."""
    name = re.sub(r"[^a-z0-9]+", " ", _text(field).casefold()).strip()
    if _FORBIDDEN_DISTANCE_WORDS.search(name):
        return False
    kind = _first(config, "metric_kind", "source_metric_kind", "value_kind", "field_kind", "type")
    if kind is not None and _text(kind).casefold() not in {"distance", "distance_like", "hsr_distance"}:
        return False
    if config.get("distance_like") is False:
        return False
    kind_token = _text(kind).casefold()
    units = _text(_first(config, "units", "unit", default="")).casefold().strip()
    distance_units = {"m", "metre", "metres", "meter", "meters", "km", "kilometre", "kilometres"}
    hsr_named = bool(_HSR_DISTANCE_NAME.search(name) or _HSR_NAME.search(name))
    distance_evidence = bool(_POSITIVE_DISTANCE_WORDS.search(name)) or units in distance_units or kind_token == "hsr_distance"
    return hsr_named and distance_evidence


def _source_config(entry: Mapping[str, Any]) -> dict[str, Any]:
    hsr = _first(entry, "hsr", "hsr_source", "high_speed_running", "high_speed_running_distance", default={})
    if not isinstance(hsr, dict):
        raise GenerationError("hsr configuration must be an object")
    merged = dict(hsr)
    source = _first(hsr, "source", "external_source", default={})
    if source is not None and not isinstance(source, dict):
        raise GenerationError("hsr source configuration must be an object")
    if isinstance(source, dict):
        for key, value in source.items():
            merged.setdefault(key, value)
    return merged


@dataclass(frozen=True)
class TeamMapping:
    season: str
    team_key: str
    exposure_path: Path
    exposure_sha256: str
    mode: str
    source_path: Path | None
    source_sha256: str | None
    source_sheet: str
    source_field: str
    source_metric_name: str
    units: str | None
    threshold_or_zone: str | None
    comparability_status: str
    source_available: bool
    gap_reason: str | None
    accepted_locator_fields: tuple[str, ...]
    source_locator_fields: tuple[str, ...]
    join_fields: tuple[tuple[str, str], ...]
    source_filter: dict[str, str]
    date_field: str | None
    blank_reason_field: str
    output_provenance_fields: tuple[str, ...]
    source_hash_fields: tuple[str, ...]
    source_hash_includes_physical_row: bool
    source_roots: tuple[Path, ...]
    local_source_available: bool
    declared_source_location: dict[str, str]
    source_file_aliases: dict[str, tuple[Path, str]]
    require_component_hash_match: bool
    source_date_order: str
    source_sheet_aliases: dict[str, str]
    allow_month_day_swap_repair: bool
    month_day_swap_repair_rationale: str | None


def _accepted_config(entry: Mapping[str, Any], path: Path) -> Mapping[str, Any]:
    value = _first(entry, "accepted_exposure", "exposure_input", "accepted_input", "input")
    if isinstance(value, str):
        return {
            "path": value,
            "sha256": _first(entry, "accepted_exposure_sha256", "exposure_sha256", "input_sha256", "exposure_checksum"),
        }
    if isinstance(value, dict):
        return value
    value = _first(entry, "accepted_exposure_path", "accepted_file", "exposure_path", "input_path", "exposure_csv")
    if value is not None:
        return {"path": value, "sha256": _first(entry, "accepted_exposure_sha256", "exposure_sha256", "input_sha256", "exposure_checksum")}
    raise GenerationError(f"team entry has no accepted exposure input: {path}")


def parse_team_mapping(entry: Mapping[str, Any], season: str, mapping_path: Path) -> TeamMapping:
    team_key = _normalise_team(_first(entry, "team_key", "team", "name"))
    if team_key not in EXPECTED_TEAMS:
        raise GenerationError(f"unexpected team in {mapping_path}: {team_key or '<blank>'}")
    accepted = _accepted_config(entry, mapping_path)
    exposure_path = _as_path(_first(accepted, "path", "file", "input_path"), mapping_path)
    exposure_sha = _expected_hash(_first(accepted, "sha256", "checksum", "file_sha256"), f"{team_key} exposure checksum")
    hsr = _source_config(entry)
    # Keep the contract readable for mapping authors who put the HSR source
    # properties beside ``exposure_path`` rather than under ``hsr``.
    for key in (
        "mode", "retrieval_mode", "source_field", "distance_source_field", "source_column", "field", "column", "hsr_source_field",
        "source_file", "source_file_path", "source_path", "source_sheet", "sheet", "source_file_sha256",
        "source_file_checksum", "source_sha256", "source_locator_fields", "source_row_locator_fields", "join_fields", "join", "source_filter",
        "source_row_filter", "distance_like", "metric_kind", "source_metric_kind", "value_kind", "field_kind",
        "units", "unit", "threshold_or_zone", "threshold", "zone", "definition",
        "provenance_fields", "output_provenance_fields", "row_provenance_fields", "source_row_hash_fields", "hash_fields",
        "source_row_hash_includes_physical_row", "blank_reason_field", "missing_reason_field",
        "comparability_status", "comparability", "gap_reason", "source_gap_reason",
        "unavailable_reason", "source_available", "available", "availability_status",
        "source_date_order", "source_sheet_aliases", "allow_month_day_swap_repair",
        "month_day_swap_repair_rationale",
    ):
        if key in entry and key not in hsr:
            hsr[key] = entry[key]
    location = _first(hsr, "source_location", default=_first(entry, "source_location", default=None))
    if isinstance(location, dict):
        for key, value in location.items():
            hsr.setdefault(key, value)
    configured_mode = _first(hsr, "mode", "retrieval_mode", default=_first(entry, "source_mode", "hsr_mode", default=None))
    if configured_mode is None:
        configured_mode = "external_row" if _first(hsr, "source_file", "source_file_path", "source_path", "path") is not None else "accepted_row"
    mode = _text(configured_mode).strip().casefold()
    mode_aliases = {
        "accepted": "accepted_row",
        "accepted_source_row": "accepted_row",
        "same_row": "external_row",
        "external": "external_row",
        "external_source_row": "external_row",
        "group": "observation_group",
        "same_observation": "observation_group",
        "same_observation_group": "observation_group",
        "source_row": "external_row",
        "unavailable": "gap",
        "not_available": "gap",
    }
    mode = mode_aliases.get(mode, mode)
    if mode not in {
        "accepted_row",
        "external_row",
        "observation_group",
        "locator_files",
        "long_variable_files",
        "declared_source_only",
        "gap",
    }:
        raise GenerationError(f"{team_key} has unsupported HSR mode: {mode}")
    availability_status = _text(_first(entry, "availability_status", default=_first(hsr, "availability_status", default=""))).strip().casefold()
    source_available_value = _first(entry, "source_available", default=_first(hsr, "source_available", "available", default=(False if availability_status in {"gap", "unavailable", "not_available"} else mode != "gap")))
    source_available = _as_bool(source_available_value, default=mode != "gap")
    gap_reason_value = _first(entry, "gap_reason", "source_gap_reason", "unavailable_reason", default=_first(hsr, "gap_reason", "source_gap_reason", "unavailable_reason"))
    gap_reason = _text(gap_reason_value).strip() or None
    if not source_available:
        if not gap_reason:
            raise GenerationError(f"{team_key} gap requires an explicit gap_reason")
        mode = "gap"
    if source_available and mode == "gap":
        raise GenerationError(f"{team_key} cannot use gap mode while source_available is true")
    local_source_available = source_available and mode != "declared_source_only"
    source_field = _text(_first(hsr, "source_field", "distance_source_field", "source_column", "field", "column", "hsr_source_field", default="")).strip()
    if source_available and not source_field:
        raise GenerationError(f"{team_key} HSR source_field is missing")
    source_filter_raw = _first(hsr, "source_filter", "source_row_filter", default={})
    if not isinstance(source_filter_raw, dict):
        raise GenerationError(f"{team_key} source_filter must be an object")
    source_filter = {_text(key): _text(value) for key, value in source_filter_raw.items()}
    metric_field = source_filter.get("variable", "") if mode == "long_variable_files" else source_field
    if source_available and not metric_field:
        raise GenerationError(f"{team_key} long-format HSR source requires a variable filter")
    if source_available and not is_distance_like_field(metric_field, hsr):
        raise GenerationError(f"{team_key} HSR source is not a distance-like field: {metric_field}")

    source_path_value = _first(hsr, "source_file", "source_file_path", "source_path", "file", "path")
    source_path: Path | None = None
    source_sha: str | None = None
    source_sheet = _text(_first(hsr, "source_sheet", "sheet", default="")).strip()
    if source_available and mode in {"external_row", "observation_group"}:
        source_path = _as_path(source_path_value, mapping_path)
        source_sha = _expected_hash(_first(hsr, "source_file_sha256", "source_file_checksum", "source_sha256", "sha256", "checksum"), f"{team_key} HSR source checksum")
        if source_path.suffix.casefold() in {".xlsx", ".xlsm", ".xls"} and not source_sheet:
            raise GenerationError(f"{team_key} XLSX HSR source requires source_sheet")
    elif source_available and mode == "accepted_row":
        # An accepted-row source is the immutable exposure file.  An optional
        # source path is metadata only and must not cause a second join.
        source_path = None
        source_sha = None

    source_roots_value = _first(hsr, "source_roots", "source_directories", default=[])
    if isinstance(source_roots_value, str):
        source_roots_value = [source_roots_value]
    if not isinstance(source_roots_value, list):
        raise GenerationError(f"{team_key} source_roots must be a list")
    source_roots = tuple(_as_path(value, mapping_path) for value in source_roots_value)
    if mode in {"locator_files", "long_variable_files"} and not source_roots and not _first(hsr, "source_files", default=[]):
        raise GenerationError(f"{team_key} {mode} requires source_roots")
    declared_source_value = _first(hsr, "declared_source_location", default={})
    if not isinstance(declared_source_value, dict):
        raise GenerationError(f"{team_key} declared_source_location must be an object")
    declared_source_location = {
        _text(key): _text(value) for key, value in declared_source_value.items() if _nonblank(value)
    }
    if mode == "declared_source_only" and not declared_source_location:
        raise GenerationError(f"{team_key} declared_source_only requires declared_source_location")
    aliases_value = _first(hsr, "source_files", default=[])
    if not isinstance(aliases_value, list):
        raise GenerationError(f"{team_key} source_files must be a list")
    source_file_aliases: dict[str, tuple[Path, str]] = {}
    for alias in aliases_value:
        if not isinstance(alias, dict):
            raise GenerationError(f"{team_key} source_files entries must be objects")
        accepted_sha = _expected_hash(
            _first(alias, "accepted_source_sha256", "content_sha256"),
            f"{team_key} accepted source checksum",
        )
        file_sha = _expected_hash(_first(alias, "sha256", "file_sha256"), f"{team_key} local source checksum")
        source_file_aliases[accepted_sha] = (_as_path(_first(alias, "path", "file"), mapping_path), file_sha)
    source_date_order = _text(_first(hsr, "source_date_order", default="month-first")).strip().casefold()
    if source_date_order not in {"month-first", "day-first"}:
        raise GenerationError(f"{team_key} source_date_order must be month-first or day-first")
    allow_month_day_swap_repair = _as_bool(
        _first(hsr, "allow_month_day_swap_repair", default=False), default=False
    )
    repair_rationale = _text(_first(hsr, "month_day_swap_repair_rationale", default="")).strip() or None
    if allow_month_day_swap_repair and not repair_rationale:
        raise GenerationError(f"{team_key} month/day swap repair requires an explicit rationale")
    sheet_aliases_value = _first(hsr, "source_sheet_aliases", default={})
    if not isinstance(sheet_aliases_value, dict):
        raise GenerationError(f"{team_key} source_sheet_aliases must be an object")

    locator_value = _first(entry, "accepted_locator_fields", "stable_locator_fields", "locator_fields", default=_first(accepted, "locator_fields", "row_locator_fields", default=None))
    stable_locator = _first(entry, "stable_row_locator", default=None)
    if locator_value is None and isinstance(stable_locator, dict):
        locator_value = _first(stable_locator, "accepted", "accepted_fields", "fields", default=None)
    accepted_locator = _normalise_locator_fields(locator_value, f"{team_key} accepted locator fields")
    source_locator_value = _first(hsr, "source_locator_fields", "source_row_locator_fields", default=_first(entry, "source_locator_fields", "source_row_locator_fields", default=None))
    source_locator = _normalise_locator_fields(source_locator_value, f"{team_key} source locator fields") if source_locator_value is not None else accepted_locator
    join_value = _first(hsr, "join_fields", "join", default=_first(entry, "join_fields", "join", default=None))
    if join_value is None:
        accepted_observation_fields = _first(hsr, "accepted_observation_fields", default=_first(entry, "accepted_observation_fields", default=None))
        source_observation_fields = _first(hsr, "source_observation_fields", default=_first(entry, "source_observation_fields", default=None))
        if accepted_observation_fields is not None or source_observation_fields is not None:
            join_value = {"accepted": accepted_observation_fields, "source": source_observation_fields}
    join = _normalise_join_pairs(join_value, accepted_locator, source_locator)
    if mode == "observation_group" and join_value is None:
        raise GenerationError(f"{team_key} observation_group requires explicit join_fields")
    if mode == "external_row":
        required_left = set(DEFAULT_LOCATOR_FIELDS)
        joined_left = {left for left, _ in join}
        if not required_left.issubset(joined_left):
            raise GenerationError(
                f"{team_key} external HSR rows require the full stable locator join: {list(DEFAULT_LOCATOR_FIELDS)}"
            )
    source_metric_name = _text(
        _first(hsr, "source_metric_name", "original_field_name", default=source_filter.get("variable", source_field))
    ).strip()
    output_fields = _first(
        hsr,
        "provenance_fields",
        "output_provenance_fields",
        "row_provenance_fields",
        default=HSR_PROVENANCE_FIELDS
        if mode in {"external_row", "observation_group", "locator_files", "long_variable_files"}
        else [],
    )
    if isinstance(output_fields, str):
        output_fields = [output_fields]
    elif isinstance(output_fields, tuple):
        output_fields = list(output_fields)
    if not isinstance(output_fields, list) or any(not _nonblank(item) for item in output_fields):
        raise GenerationError(f"{team_key} provenance_fields must be a list")
    blank_field = _text(_first(hsr, "blank_reason_field", "missing_reason_field", default=_first(entry, "blank_reason_field", default=DEFAULT_BLANK_REASON_FIELD))).strip()
    if not blank_field:
        raise GenerationError(f"{team_key} blank_reason_field cannot be blank")
    source_hash_fields_raw = _first(hsr, "source_row_hash_fields", "hash_fields", default=[])
    hash_config = _first(hsr, "source_row_hash", default=None)
    if isinstance(hash_config, dict):
        source_hash_fields_raw = _first(hash_config, "fields", "source_fields", default=source_hash_fields_raw)
    if isinstance(source_hash_fields_raw, str):
        source_hash_fields_raw = [source_hash_fields_raw]
    if not isinstance(source_hash_fields_raw, list):
        raise GenerationError(f"{team_key} source_row_hash_fields must be a list")
    source_hash_includes_physical_row = _as_bool(
        _first(hsr, "source_row_hash_includes_physical_row", default=_first(hash_config, "includes_physical_row", default=False) if isinstance(hash_config, dict) else False),
        default=False,
    )
    comparability = _text(_first(entry, "comparability_status", "comparability", default=_first(hsr, "comparability_status", "comparability", default="team_defined_hsr_distance"))).strip()
    if not comparability:
        raise GenerationError(f"{team_key} comparability_status cannot be blank")
    units = _first(hsr, "units", "unit", default=None)
    threshold = _first(hsr, "threshold_or_zone", "threshold", "zone", "definition", default=None)
    return TeamMapping(
        season=season,
        team_key=team_key,
        exposure_path=exposure_path,
        exposure_sha256=exposure_sha,
        mode=mode,
        source_path=source_path,
        source_sha256=source_sha,
        source_sheet=source_sheet,
        source_field=source_field,
        source_metric_name=source_metric_name,
        units=_text(units).strip() if units is not None and _text(units).strip() else "unknown",
        threshold_or_zone=_text(threshold).strip() if threshold is not None and _text(threshold).strip() else "unknown",
        comparability_status=comparability,
        source_available=source_available,
        gap_reason=gap_reason,
        accepted_locator_fields=accepted_locator,
        source_locator_fields=source_locator,
        join_fields=join,
        source_filter=source_filter,
        date_field=_text(_first(entry, "date_field", default=_first(accepted, "date_field", default=None))).strip() or None,
        blank_reason_field=blank_field,
        output_provenance_fields=tuple(_text(item).strip() for item in output_fields),
        source_hash_fields=tuple(_text(item).strip() for item in source_hash_fields_raw),
        source_hash_includes_physical_row=source_hash_includes_physical_row,
        source_roots=source_roots,
        local_source_available=local_source_available,
        declared_source_location=declared_source_location,
        source_file_aliases=source_file_aliases,
        require_component_hash_match=_as_bool(
            _first(hsr, "require_component_hash_match", default=True), default=True
        ),
        source_date_order=source_date_order,
        source_sheet_aliases={_text(key): _text(value) for key, value in sheet_aliases_value.items()},
        allow_month_day_swap_repair=allow_month_day_swap_repair,
        month_day_swap_repair_rationale=repair_rationale,
    )


def load_mapping(path: Path, expected_season: str) -> tuple[list[TeamMapping], str]:
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise GenerationError(f"required HSR mapping is missing: {path}") from exc
    except json.JSONDecodeError as exc:
        raise GenerationError(f"invalid HSR mapping JSON {path}: {exc}") from exc
    declared_season = _text(raw.get("season", "")) if isinstance(raw, dict) else ""
    if declared_season and declared_season != expected_season:
        raise GenerationError(f"mapping season mismatch in {path}: {declared_season} != {expected_season}")
    entries = _entries(raw, path)
    mappings = [parse_team_mapping(entry, expected_season, path) for entry in entries]
    teams = [mapping.team_key for mapping in mappings]
    if len(teams) != len(set(teams)):
        raise GenerationError(f"duplicate team mapping in {path}")
    if set(teams) != EXPECTED_TEAMS:
        missing = sorted(EXPECTED_TEAMS - set(teams))
        extra = sorted(set(teams) - EXPECTED_TEAMS)
        detail = f"missing={missing}" if missing else ""
        if extra:
            detail = f"{detail} extra={extra}".strip()
        raise GenerationError(f"mapping must cover exactly 16 expected teams in {path}: {detail}")
    available = {mapping.team_key for mapping in mappings if mapping.source_available}
    if len(available) != EXPECTED_AVAILABILITY[expected_season]:
        raise GenerationError(
            f"{expected_season} source availability must be {EXPECTED_AVAILABILITY[expected_season]}/16, got {len(available)}/16"
        )
    gaps = set(EXPECTED_TEAMS) - available
    if gaps != set(EXPECTED_GAPS[expected_season]):
        raise GenerationError(f"{expected_season} source gaps must be {sorted(EXPECTED_GAPS[expected_season])}, got {sorted(gaps)}")
    return sorted(mappings, key=lambda item: item.team_key), sha256_file(path)


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        content = None
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                content = path.read_text(encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        if content is None:
            raise GenerationError(f"unsupported CSV encoding: {path}")
        lines = content.splitlines()
        if not lines:
            raise GenerationError(f"CSV has no usable header: {path}")
        # One Bulls export quotes the whole header and doubles its inner
        # quotes. Its data rows are ordinary comma-separated values.
        if lines[0].startswith('"') and lines[0].endswith('"') and ',""' in lines[0]:
            lines[0] = lines[0][1:-1].replace('""', '"')
        choices: list[tuple[int, list[list[str]]]] = []
        for delimiter in (",", ";", "\t", "|"):
            parsed = list(csv.reader(lines, delimiter=delimiter))
            if parsed:
                choices.append((sum(len(row) for row in parsed), parsed))
        parsed_rows = max(choices, key=lambda item: item[0])[1]
        fields = [_text(field).strip() for field in parsed_rows[0]]
        while fields and not fields[-1]:
            fields.pop()
        if not fields or any(not field for field in fields):
            raise GenerationError(f"CSV has no usable header: {path}")
        if len(set(fields)) != len(fields):
            raise GenerationError(f"CSV has duplicate header fields: {path}")
        rows = []
        for row_number, values in enumerate(parsed_rows[1:], start=2):
            while len(values) > len(fields) and not _nonblank(values[-1]):
                values.pop()
            if len(values) == len(fields) + 1 and not _nonblank(values[0]):
                values = values[1:]
            if not any(_nonblank(value) for value in values):
                continue
            row = {field: _text(values[index] if index < len(values) else "") for index, field in enumerate(fields)}
            if len(values) > len(fields):
                row["__extra_values"] = json.dumps(values[len(fields):], ensure_ascii=False)
            row["__physical_row_number"] = str(row_number)
            rows.append(row)
    except FileNotFoundError as exc:
        raise GenerationError(f"input file is missing: {path}") from exc
    return fields, rows


def _read_xlsx(path: Path, sheet: str) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise GenerationError("openpyxl is required for XLSX HSR sources") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError as exc:
        raise GenerationError(f"input file is missing: {path}") from exc
    if sheet.isdigit() and 0 <= int(sheet) < len(workbook.worksheets):
        worksheet = workbook.worksheets[int(sheet)]
    elif sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
    else:
        workbook.close()
        raise GenerationError(f"sheet not found in {path}: {sheet}")
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration as exc:
        workbook.close()
        raise GenerationError(f"empty worksheet: {path} [{sheet}]") from exc
    fields = [_text(value).strip() for value in raw_headers]
    while fields and not fields[-1]:
        fields.pop()
    if not fields or len(set(fields)) != len(fields) or any(not field for field in fields):
        workbook.close()
        raise GenerationError(f"worksheet has invalid or duplicate headers: {path} [{sheet}]")
    rows = []
    for row_number, values in enumerate(rows_iter, start=2):
        values = list(values)
        if not any(_nonblank(value) for value in values):
            continue
        row = {field: _text(values[index] if index < len(values) else "") for index, field in enumerate(fields)}
        row["__physical_row_number"] = str(row_number)
        rows.append(row)
    workbook.close()
    return fields, rows


def read_table(path: Path, sheet: str = "") -> tuple[list[str], list[dict[str, str]]]:
    if path.suffix.casefold() in {".xlsx", ".xlsm", ".xls"}:
        return _read_xlsx(path, sheet)
    return _read_csv(path)


def _with_source_metadata(row: Mapping[str, str], physical_row: int, source: TeamMapping) -> dict[str, str]:
    copy = dict(row)
    copy["__physical_row_number"] = str(physical_row)
    copy["__source_sheet"] = source.source_sheet
    copy["__source_file_sha256"] = source.source_sha256 or ""
    if "source_row_number" not in copy or not _nonblank(copy.get("source_row_number")):
        copy["source_row_number"] = str(physical_row)
    if "source_sheet" not in copy or not _nonblank(copy.get("source_sheet")):
        copy["source_sheet"] = source.source_sheet
    if "source_file_sha256" not in copy or not _nonblank(copy.get("source_file_sha256")):
        copy["source_file_sha256"] = source.source_sha256 or ""
    if not _nonblank(copy.get("source_row_sha256")) and source.source_hash_fields:
        payload = {field: copy.get(field, "") for field in source.source_hash_fields}
        if source.source_hash_includes_physical_row:
            payload["__physical_row_number"] = str(physical_row)
        encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        copy["source_row_sha256"] = hashlib.sha256(encoded).hexdigest()
    return copy


def _locator_value(row: Mapping[str, str], field: str, physical_row: int | None = None) -> str:
    if field == "__physical_row_number":
        return _text(row.get(field, "")) or str(physical_row if physical_row is not None else "")
    return _text(row.get(field, ""))


def _validate_locators(rows: Sequence[Mapping[str, str]], fields: Sequence[str], label: str) -> None:
    seen: set[tuple[str, ...]] = set()
    for index, row in enumerate(rows, start=2):
        values = tuple(_locator_value(row, field, index) for field in fields)
        if any(not value.strip() for value in values):
            raise GenerationError(f"{label} has missing stable locator field(s) at row {index}: {fields}")
        if values in seen:
            raise GenerationError(f"{label} has duplicated stable locator at row {index}")
        seen.add(values)


def _locator_set_digest(rows: Sequence[Mapping[str, str]], fields: Sequence[str]) -> str:
    values = sorted(tuple(_locator_value(row, field) for field in fields) for row in rows)
    payload = json.dumps(values, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _source_lookup(source: TeamMapping, source_rows: Sequence[Mapping[str, str]]) -> dict[tuple[str, ...], Mapping[str, str]]:
    if source.mode == "observation_group":
        locator_fields = source.source_locator_fields
    else:
        locator_fields = tuple(right for _, right in source.join_fields)
    _validate_locators(source_rows, locator_fields, f"{source.team_key} HSR source")
    filtered = [row for row in source_rows if all(row.get(field, "") == expected for field, expected in source.source_filter.items())]
    # A source locator remains unique after filtering.  A group key may occur
    # more than once in the raw file, but a distance row must be unambiguous.
    lookup: dict[tuple[str, ...], Mapping[str, str]] = {}
    for row in filtered:
        key = tuple(_locator_value(row, field) for _, field in source.join_fields)
        if any(not value.strip() for value in key):
            continue
        if key in lookup:
            raise GenerationError(f"{source.team_key} HSR source has duplicated join key")
        lookup[key] = row
    return lookup


def _parse_date(value: str) -> date | None:
    text = _text(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _date_range(rows: Sequence[Mapping[str, str]], preferred: str | None) -> dict[str, str | None]:
    field = preferred
    if field is None:
        for candidate in ("session date", "session_date_clean", "cleaned_date", "date"):
            if rows and candidate in rows[0]:
                field = candidate
                break
    if field is None:
        return {"start": None, "end": None}
    values = [parsed for row in rows if (parsed := _parse_date(row.get(field, ""))) is not None]
    return {"start": min(values).isoformat() if values else None, "end": max(values).isoformat() if values else None}


def _provenance_values(source: TeamMapping, row: Mapping[str, str]) -> dict[str, str]:
    values = {
        "hsr_source_file_sha256": source.source_sha256 or "",
        "hsr_source_sheet": source.source_sheet,
        "hsr_source_row_number": _text(row.get("source_row_number", row.get("__physical_row_number", ""))),
        "hsr_source_row_sha256": _text(row.get("source_row_sha256", "")),
    }
    return {field: values.get(field, "") for field in source.output_provenance_fields}


def _row_payload_hash(fields: Sequence[str], row: Mapping[str, str]) -> str:
    payload = {field: _text(row.get(field, "")).strip() for field in fields}
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _resolve_source_files(mapping: TeamMapping, exposure_rows: Sequence[Mapping[str, str]]) -> dict[str, Path]:
    expected = {_text(row.get("source_file_sha256", "")).casefold() for row in exposure_rows}
    if "" in expected:
        raise GenerationError(f"{mapping.team_key} accepted exposure has a blank source file checksum")
    resolved: dict[str, Path] = {}
    for accepted_sha, (path, file_sha) in mapping.source_file_aliases.items():
        if accepted_sha not in expected:
            continue
        if sha256_file(path) != file_sha:
            raise GenerationError(f"{mapping.team_key} local source checksum differs from mapping: {path}")
        resolved[accepted_sha] = path
    candidates: list[Path] = []
    for root in mapping.source_roots:
        if root.is_file():
            candidates.append(root)
        elif root.is_dir():
            candidates.extend(
                path
                for path in root.rglob("*")
                if path.is_file() and path.suffix.casefold() in {".csv", ".xlsx", ".xlsm", ".xls"}
            )
        else:
            raise GenerationError(f"{mapping.team_key} source root is missing: {root}")
    for path in sorted(set(candidates)):
        digest = sha256_file(path)
        if digest in expected and digest not in resolved:
            resolved[digest] = path
    missing = sorted(expected - set(resolved))
    if missing:
        raise GenerationError(f"{mapping.team_key} local source file(s) missing for checksum(s): {missing}")
    return resolved


def _long_variable_lookup(
    mapping: TeamMapping,
    fields: Sequence[str],
    rows: Sequence[Mapping[str, str]],
) -> dict[str, Mapping[str, str]]:
    required = {"week_start", "player_display_name_harmonised", "variable", "match_day", "value"}
    if not required.issubset(fields):
        raise GenerationError(f"{mapping.team_key} long-format HSR source is missing required fields")
    hsr_variable = mapping.source_filter.get("variable", "")
    if not hsr_variable:
        raise GenerationError(f"{mapping.team_key} long-format HSR source requires a variable filter")
    groups: dict[tuple[str, str, str], dict[str, Mapping[str, str]]] = {}
    for row in rows:
        key = tuple(_text(row.get(field, "")) for field in ("week_start", "player_display_name_harmonised", "match_day"))
        variable = _text(row.get("variable", ""))
        values = groups.setdefault(key, {})
        if variable in values:
            raise GenerationError(f"{mapping.team_key} long-format HSR source has a duplicated variable group")
        values[variable] = row
    lookup: dict[str, Mapping[str, str]] = {}
    for values in groups.values():
        duration = values.get("total_time_minutes")
        distance = values.get("distance_total")
        hsr = values.get(hsr_variable)
        if duration is None or distance is None or hsr is None:
            continue
        duration_row = _text(duration.get("__physical_row_number", ""))
        if duration_row in lookup:
            raise GenerationError(f"{mapping.team_key} long-format HSR source has a duplicated duration row")
        enriched = dict(hsr)
        enriched["source_row_number"] = _text(hsr.get("__physical_row_number", ""))
        enriched["source_row_sha256"] = _row_payload_hash(fields, hsr)
        enriched["__duration_source_row_sha256"] = _row_payload_hash(fields, duration)
        enriched["__duration_value"] = _text(duration.get("value", ""))
        enriched["__distance_source_row_number"] = _text(distance.get("__physical_row_number", ""))
        enriched["__distance_source_row_sha256"] = _row_payload_hash(fields, distance)
        enriched["__distance_value"] = _text(distance.get("value", ""))
        enriched["__week_start"] = _text(duration.get("week_start", ""))
        enriched["__player_label"] = _text(duration.get("player_display_name_harmonised", ""))
        enriched["__match_day"] = _text(duration.get("match_day", ""))
        lookup[duration_row] = enriched
    return lookup


def _stable_player_uid(team_key: str, source_label: str) -> str:
    payload = f"ply\x1f{team_key}\x1f{source_label.strip()}"
    return "ply_" + hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]


def _parse_source_date(value: str, order: str) -> date | None:
    raw = _text(value).strip()
    if not raw:
        return None
    formats = ["%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"]
    formats.extend(
        ["%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%d/%m/%y"]
        if order == "month-first"
        else ["%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%m/%d/%y"]
    )
    for pattern in formats:
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            continue
    return None


def _same_number(left: str, right: str) -> bool:
    try:
        return abs(float(left) - float(right)) <= 0.000001
    except (TypeError, ValueError):
        return False


_MISSING_NUMERIC_TOKENS = frozenset({"na", "n/a", "nan", "null", "none", "not available", "missing"})
_NUMERIC_TEXT = re.compile(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?")


def _decimal_value(value: str) -> tuple[str, Decimal] | None:
    token = _text(value).strip()
    if not token or token.casefold() in _MISSING_NUMERIC_TOKENS:
        return None
    if "," in token:
        if "." in token or token.count(",") != 1:
            return None
        token = token.replace(",", ".")
    if not _NUMERIC_TEXT.fullmatch(token):
        return None
    try:
        number = Decimal(token)
    except InvalidOperation:
        return None
    if not number.is_finite():
        return None
    return token, number


def _canonical_hsr_value(value: str, team_key: str, accepted_row: Mapping[str, str]) -> tuple[str, str]:
    token = _text(value).strip()
    if not token:
        return "", MISSING_REASON_BLANK
    if token.casefold() in _MISSING_NUMERIC_TOKENS:
        return "", MISSING_REASON_TOKEN
    parsed = _decimal_value(token)
    if parsed is None:
        raise GenerationError(f"{team_key} HSR source value is not numeric: {token!r}")
    canonical, number = parsed
    if number < 0:
        raise GenerationError(f"{team_key} HSR source value is negative: {token!r}")
    total = None
    for field in ("distance_total_m_clean", "distance total", "distance_total", "total distance"):
        if _nonblank(accepted_row.get(field, "")):
            total = _decimal_value(_text(accepted_row.get(field, "")))
            break
    if total is not None and number - total[1] > Decimal("0.000001"):
        return "", MISSING_REASON_EXCEEDS_TOTAL
    return canonical, ""


def generate_team_artifact(mapping: TeamMapping, output_root: Path) -> dict[str, Any]:
    actual_exposure_sha = sha256_file(mapping.exposure_path)
    if actual_exposure_sha != mapping.exposure_sha256:
        raise GenerationError(f"{mapping.team_key} accepted exposure checksum differs from mapping")
    exposure_fields, exposure_rows = read_table(mapping.exposure_path)
    _validate_locators(exposure_rows, mapping.accepted_locator_fields, f"{mapping.team_key} accepted exposure")
    if not exposure_rows:
        raise GenerationError(f"{mapping.team_key} accepted exposure has no rows")
    if mapping.source_available and mapping.mode == "accepted_row" and mapping.source_field not in exposure_fields:
        raise GenerationError(f"{mapping.team_key} accepted exposure HSR source field is missing: {mapping.source_field}")

    source_rows: list[dict[str, str]] = []
    source_fields: list[str] = []
    lookup: dict[tuple[str, ...], Mapping[str, str]] = {}
    resolved_sources: dict[str, Path] = {}
    locator_tables: dict[tuple[str, str], tuple[list[str], dict[str, Mapping[str, str]]]] = {}
    long_tables: dict[tuple[str, str], dict[str, Mapping[str, str]]] = {}
    if mapping.source_available and mapping.mode in {"external_row", "observation_group"}:
        assert mapping.source_path is not None and mapping.source_sha256 is not None
        actual_source_sha = sha256_file(mapping.source_path)
        if actual_source_sha != mapping.source_sha256:
            raise GenerationError(f"{mapping.team_key} HSR source checksum differs from mapping")
        source_fields, raw_source_rows = read_table(mapping.source_path, mapping.source_sheet)
        source_rows = [
            _with_source_metadata(row, int(row.get("__physical_row_number", index)), mapping)
            for index, row in enumerate(raw_source_rows, start=2)
        ]
        if mapping.source_field not in source_fields:
            raise GenerationError(f"{mapping.team_key} HSR source field is missing: {mapping.source_field}")
        lookup = _source_lookup(mapping, source_rows)
    elif mapping.mode in {"locator_files", "long_variable_files"}:
        resolved_sources = _resolve_source_files(mapping, exposure_rows)

    output_fields = list(exposure_fields)
    if CANONICAL_HSR_FIELD not in output_fields:
        output_fields.append(CANONICAL_HSR_FIELD)
    # Add a reason column only when at least one row needs a controlled blank
    # reason.  External-row provenance is handled separately below.
    needs_blank_reason = False
    output_rows: list[dict[str, str]] = []
    output_blank_reasons: list[str] = []
    matched_rows = 0
    populated_rows = 0
    blank_rows = 0
    blank_reason_counts: dict[str, int] = {}
    source_trace_rows = 0
    month_day_swap_repair_count = 0

    for row in exposure_rows:
        out = dict(row)
        hsr_value = ""
        reason = ""
        source_row: Mapping[str, str] | None = None
        if not mapping.source_available or mapping.mode == "declared_source_only":
            reason = mapping.gap_reason or MISSING_REASON_GAP
        elif mapping.mode == "accepted_row":
            hsr_value = _text(row.get(mapping.source_field, ""))
            if not hsr_value.strip():
                hsr_value = ""
                reason = MISSING_REASON_BLANK
        elif mapping.mode in {"external_row", "observation_group"}:
            accepted_key = tuple(_locator_value(row, left) for left, _ in mapping.join_fields)
            if any(not value.strip() for value in accepted_key):
                reason = MISSING_REASON_NO_ROW
            else:
                source_row = lookup.get(accepted_key)
                if source_row is None:
                    reason = MISSING_REASON_NO_ROW
                else:
                    matched_rows += 1
                    out.update(_provenance_values(mapping, source_row))
                    hsr_value = _text(source_row.get(mapping.source_field, ""))
                    if hsr_value.strip() and "hsr_source_row_sha256" in mapping.output_provenance_fields and not _nonblank(source_row.get("source_row_sha256")):
                        raise GenerationError(f"{mapping.team_key} populated HSR source row has no source_row_sha256")
                    if not hsr_value.strip():
                        hsr_value = ""
                        reason = MISSING_REASON_BLANK
        elif mapping.mode == "locator_files":
            file_sha = _text(row.get("source_file_sha256", "")).casefold()
            sheet = _text(row.get("source_sheet", ""))
            row_number = _text(row.get("source_row_number", ""))
            table_key = (file_sha, sheet)
            if table_key not in locator_tables:
                source_sheet = mapping.source_sheet_aliases.get(sheet, sheet)
                fields, raw_rows = read_table(resolved_sources[file_sha], source_sheet)
                if mapping.source_field not in fields:
                    raise GenerationError(
                        f"{mapping.team_key} HSR source field is missing in {resolved_sources[file_sha]}: {mapping.source_field}"
                    )
                locator_tables[table_key] = (
                    fields,
                    {_text(item.get("__physical_row_number", "")): item for item in raw_rows},
                )
            _, rows_by_number = locator_tables[table_key]
            raw_source_row = rows_by_number.get(row_number)
            if raw_source_row is None:
                raise GenerationError(
                    f"{mapping.team_key} HSR source row is missing for {file_sha}:{sheet}:{row_number}"
                )
            if _nonblank(raw_source_row.get("__extra_values", "")):
                raise GenerationError(
                    f"{mapping.team_key} accepted source row has more values than headers for {file_sha}:{sheet}:{row_number}"
                )
            source_row = dict(raw_source_row)
            source_row.update(
                {
                    "source_file_sha256": file_sha,
                    "source_sheet": sheet,
                    "source_row_number": row_number,
                    "source_row_sha256": _text(row.get("source_row_sha256", "")),
                }
            )
            matched_rows += 1
            out.update(_provenance_values(mapping, source_row))
            hsr_value = _text(source_row.get(mapping.source_field, ""))
            if not hsr_value.strip():
                hsr_value = ""
                reason = MISSING_REASON_BLANK
        elif mapping.mode == "long_variable_files":
            file_sha = _text(row.get("source_file_sha256", "")).casefold()
            sheet = _text(row.get("source_sheet", ""))
            duration_row_number = _text(row.get("duration_source_row_number", ""))
            table_key = (file_sha, sheet)
            if table_key not in long_tables:
                source_sheet = mapping.source_sheet_aliases.get(sheet, sheet)
                fields, raw_rows = read_table(resolved_sources[file_sha], source_sheet)
                long_tables[table_key] = _long_variable_lookup(mapping, fields, raw_rows)
            source_row = long_tables[table_key].get(duration_row_number)
            if source_row is None:
                raise GenerationError(
                    f"{mapping.team_key} HSR observation group is missing for duration row {duration_row_number}"
                )
            if mapping.require_component_hash_match and _text(row.get("duration_source_row_sha256", "")) != _text(
                source_row.get("__duration_source_row_sha256", "")
            ):
                raise GenerationError(
                    f"{mapping.team_key} duration source row hash differs at accepted row {row.get('standardised_row_number', '')}"
                )
            if _text(row.get("distance_source_row_number", "")) != _text(
                source_row.get("__distance_source_row_number", "")
            ):
                raise GenerationError(
                    f"{mapping.team_key} distance source row differs at accepted row {row.get('standardised_row_number', '')}"
                )
            if mapping.require_component_hash_match and _text(row.get("distance_source_row_sha256", "")) != _text(
                source_row.get("__distance_source_row_sha256", "")
            ):
                raise GenerationError(
                    f"{mapping.team_key} distance source row hash differs at accepted row {row.get('standardised_row_number', '')}"
                )
            source_date = _parse_source_date(_text(source_row.get("__week_start", "")), mapping.source_date_order)
            accepted_date = _parse_date(_text(row.get("cleaned_date", row.get("session date", ""))))
            date_matches = source_date is not None and source_date == accepted_date
            used_month_day_swap_repair = False
            if (
                not date_matches
                and mapping.allow_month_day_swap_repair
                and source_date is not None
                and accepted_date is not None
            ):
                used_month_day_swap_repair = (
                    source_date.year == accepted_date.year
                    and source_date.month == accepted_date.day
                    and source_date.day == accepted_date.month
                )
                date_matches = used_month_day_swap_repair
            expected_setting = "match" if _text(source_row.get("__match_day", "")).casefold() in {"1", "1.0", "true", "yes"} else "training"
            semantic_match = (
                date_matches
                and _stable_player_uid(mapping.team_key, _text(source_row.get("__player_label", "")))
                == _text(row.get("player_uid", ""))
                and expected_setting == _text(row.get("setting", "")).casefold()
                and _same_number(_text(source_row.get("__duration_value", "")), _text(row.get("minutes_total_clean", "")))
                and _same_number(_text(source_row.get("__distance_value", "")), _text(row.get("distance_total_m_clean", "")))
            )
            if not semantic_match:
                raise GenerationError(
                    f"{mapping.team_key} long-format observation differs at accepted row {row.get('standardised_row_number', '')}"
                )
            if used_month_day_swap_repair:
                month_day_swap_repair_count += 1
            enriched_source_row = dict(source_row)
            enriched_source_row["source_file_sha256"] = file_sha
            enriched_source_row["source_sheet"] = sheet
            source_row = enriched_source_row
            matched_rows += 1
            out.update(_provenance_values(mapping, source_row))
            hsr_value = _text(source_row.get(mapping.source_field, ""))
            if not hsr_value.strip():
                hsr_value = ""
                reason = MISSING_REASON_BLANK
        else:
            raise GenerationError(f"{mapping.team_key} unsupported generation mode: {mapping.mode}")
        if hsr_value:
            hsr_value, canonical_reason = _canonical_hsr_value(hsr_value, mapping.team_key, row)
            if canonical_reason:
                reason = canonical_reason
            else:
                source_trace_rows += 1
        out[CANONICAL_HSR_FIELD] = hsr_value
        if hsr_value:
            populated_rows += 1
        else:
            blank_rows += 1
            blank_reason_counts[reason] = blank_reason_counts.get(reason, 0) + 1
        output_rows.append(out)
        output_blank_reasons.append(reason)

    before_locator_digest = _locator_set_digest(exposure_rows, mapping.accepted_locator_fields)
    after_locator_digest = _locator_set_digest(output_rows, mapping.accepted_locator_fields)
    if before_locator_digest != after_locator_digest:
        raise GenerationError(f"{mapping.team_key} accepted exposure locator set changed")
    preserved_fields = [field for field in exposure_fields if field != CANONICAL_HSR_FIELD]
    for row_number, (before, after) in enumerate(zip(exposure_rows, output_rows), start=2):
        if any(_text(before.get(field, "")) != _text(after.get(field, "")) for field in preserved_fields):
            raise GenerationError(f"{mapping.team_key} accepted field changed at row {row_number}")

    needs_blank_reason = blank_rows > 0 or mapping.blank_reason_field in output_fields
    if needs_blank_reason:
        for out, reason in zip(output_rows, output_blank_reasons):
            hsr_value = out.get(CANONICAL_HSR_FIELD, "")
            if hsr_value:
                out[mapping.blank_reason_field] = ""
                continue
            out[mapping.blank_reason_field] = reason

    if mapping.mode in {"external_row", "observation_group", "locator_files", "long_variable_files"}:
        for field in mapping.output_provenance_fields:
            if field not in output_fields:
                output_fields.append(field)
    if needs_blank_reason and mapping.blank_reason_field not in output_fields:
        output_fields.append(mapping.blank_reason_field)
    output_dir = output_root / mapping.season / mapping.team_key
    output_path = output_dir / "exposure_with_hsr.csv"
    output_dir.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=output_fields, extrasaction="raise", lineterminator="\n")
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in output_fields} for row in output_rows)
    if mapping.mode == "locator_files":
        source_row_count = sum(len(rows_by_number) for _, rows_by_number in locator_tables.values())
    elif mapping.mode == "long_variable_files":
        source_row_count = sum(len(rows_by_duration) for rows_by_duration in long_tables.values())
    elif mapping.source_available and mapping.mode != "accepted_row":
        source_row_count = len(source_rows)
    elif mapping.source_available:
        source_row_count = len(exposure_rows)
    else:
        source_row_count = 0
    return {
        "season": mapping.season,
        "team_key": mapping.team_key,
        "output_path": _display_path(output_path),
        "output_sha256": sha256_file(output_path),
        "before_row_count": len(exposure_rows),
        "after_row_count": len(output_rows),
        "matched_hsr_count": populated_rows,
        "blank_hsr_count": blank_rows,
        "source_rows": source_row_count,
        "source_rows_matched": matched_rows if mapping.mode != "accepted_row" and mapping.source_available else len(exposure_rows) if mapping.source_available else 0,
        "source_trace_rows": source_trace_rows,
        "locator_set_identity": {
            "identical": before_locator_digest == after_locator_digest,
            "before_sha256": before_locator_digest,
            "after_sha256": after_locator_digest,
        },
        "blank_reason_counts": dict(sorted(blank_reason_counts.items())),
        "date_range": _date_range(exposure_rows, mapping.date_field),
        "source_available": mapping.source_available,
        "local_source_available": mapping.local_source_available,
        "resolved_source_files": [
            {
                "path": _display_path(path),
                "accepted_source_sha256": digest,
                "local_file_sha256": sha256_file(path),
            }
            for digest, path in sorted(resolved_sources.items())
        ],
        "gap_reason": mapping.gap_reason,
        "month_day_swap_repair_count": month_day_swap_repair_count,
    }


def _manifest_entry(mapping: TeamMapping, result: Mapping[str, Any]) -> dict[str, Any]:
    source_location = None
    if mapping.mode in {"locator_files", "long_variable_files"}:
        source_location = {"files": result["resolved_source_files"]}
    elif mapping.mode == "declared_source_only":
        source_location = dict(mapping.declared_source_location)
    elif mapping.source_available and mapping.mode != "accepted_row":
        source_location = {
            "file": _display_path(mapping.source_path or Path("")),
            "sheet": mapping.source_sheet or None,
            "sha256": mapping.source_sha256,
        }
    elif mapping.source_available:
        source_location = {
            "file": _display_path(mapping.exposure_path),
            "sheet": None,
            "sha256": mapping.exposure_sha256,
        }
    if mapping.mode == "locator_files":
        stable_row_join = {
            "accepted_fields": list(DEFAULT_LOCATOR_FIELDS),
            "source_fields": [
                "checksum-bound local file",
                "resolved table or worksheet",
                "__physical_row_number",
                "accepted source-row lineage hash",
            ],
            "rule": "Exact source file, sheet and physical row; accepted source-row hash retained as provenance.",
        }
    elif mapping.mode == "long_variable_files":
        stable_row_join = {
            "accepted_fields": [
                "source_file_sha256",
                "source_sheet",
                "duration_source_row_number",
                "distance_source_row_number",
                "cleaned_date",
                "player_uid",
                "setting",
                "minutes_total_clean",
                "distance_total_m_clean",
            ],
            "source_fields": [
                "checksum-bound local file",
                "resolved table or worksheet",
                "total_time_minutes physical row",
                "distance_total physical row",
                "week_start",
                "pseudonym derived from player_display_name_harmonised",
                "match_day",
                "total_time_minutes value",
                "distance_total value",
            ],
            "rule": "Complete long-format observation group with an exact HSR distance variable row.",
            "date_repair": {
                "method": "swap month and day only when the repaired source date exactly matches the accepted date",
                "rows_repaired": result["month_day_swap_repair_count"],
                "rationale": mapping.month_day_swap_repair_rationale,
            }
            if mapping.allow_month_day_swap_repair
            else None,
        }
    else:
        stable_row_join = {
            "accepted_fields": list(mapping.accepted_locator_fields),
            "source_fields": [right for _, right in mapping.join_fields],
            "pairs": [{"accepted": left, "source": right} for left, right in mapping.join_fields],
        }
    return {
        "season": mapping.season,
        "team_key": mapping.team_key,
        "accepted_exposure": {
            "file": _display_path(mapping.exposure_path),
            "sha256": mapping.exposure_sha256,
        },
        "source_available": mapping.source_available,
        "local_source_available": mapping.local_source_available,
        "source_mode": mapping.mode,
        "source_field": mapping.source_metric_name or None,
        "source_value_field": mapping.source_field or None,
        "units": mapping.units,
        "threshold_or_zone": mapping.threshold_or_zone,
        "source_location": source_location,
        "stable_row_join": stable_row_join,
        "accepted_row_coverage": {
            "accepted_rows": result["before_row_count"],
            "matched_hsr_rows": result["matched_hsr_count"],
            "blank_hsr_rows": result["blank_hsr_count"],
        },
        "comparability_status": mapping.comparability_status,
        "gap_reason": mapping.gap_reason,
    }


def write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")


def _payload_number(value: Any, label: str) -> str | None:
    """Return a canonical numeric token without losing source precision."""
    text = _text(value).strip()
    if not text:
        return None
    parsed = _decimal_value(text)
    if parsed is None:
        raise GenerationError(f"{label} is not numeric: {text!r}")
    return parsed[0]


def _accepted_total_value(row: Mapping[str, str], team_key: str) -> str | None:
    for field in ("distance_total_m_clean", "distance total", "distance_total", "total distance"):
        if _nonblank(row.get(field, "")):
            text = _text(row[field]).strip()
            if text.casefold() in _MISSING_NUMERIC_TOKENS:
                return None
            parsed = _decimal_value(text)
            return parsed[0] if parsed is not None else None
    return None


def _database_parameter_payload(
    mappings: Sequence[TeamMapping],
    results: Mapping[tuple[str, str], Mapping[str, Any]],
    mapping_hashes: Mapping[str, str],
    output_root: Path,
) -> list[dict[str, Any]]:
    """Build the compact, private SQL parameter array from canonical outputs."""
    metadata: list[dict[str, Any]] = []
    observations: list[dict[str, Any]] = []
    for mapping in sorted(mappings, key=lambda item: (item.season, item.team_key)):
        result = results[(mapping.season, mapping.team_key)]
        metadata.append(
            {
                "kind": "team_season",
                "season": mapping.season,
                "team_key": mapping.team_key,
                "accepted_exposure_sha256": mapping.exposure_sha256,
                "canonical_output_sha256": result["output_sha256"],
                "mapping_sha256": mapping_hashes[mapping.season],
                "source_available": mapping.source_available,
                "local_source_available": mapping.local_source_available,
                "source_mode": mapping.mode,
                "units": mapping.units,
                "threshold_or_zone": mapping.threshold_or_zone,
                "comparability_status": mapping.comparability_status,
                "gap_reason": mapping.gap_reason,
                "accepted_locator_fields": list(mapping.accepted_locator_fields),
                "accepted_row_count": result["before_row_count"],
                "canonical_populated_row_count": result["matched_hsr_count"],
                "canonical_blank_row_count": result["blank_hsr_count"],
            }
        )
        output_path = output_root / mapping.season / mapping.team_key / "exposure_with_hsr.csv"
        output_fields, output_rows = read_table(output_path)
        if CANONICAL_HSR_FIELD not in output_fields:
            raise GenerationError(f"{mapping.team_key} canonical output is missing HSR")
        for row in output_rows:
            accepted_locator = {
                field: _text(row.get(field, "")) for field in mapping.accepted_locator_fields
            }
            if any(not value.strip() for value in accepted_locator.values()):
                raise GenerationError(f"{mapping.team_key} payload row has a missing accepted locator")
            hsr_distance = _payload_number(
                row.get(CANONICAL_HSR_FIELD, ""),
                f"{mapping.team_key} canonical HSR distance",
            )
            hsr_source_row_sha = _text(row.get("hsr_source_row_sha256", "")).strip() or None
            blank_reason = _text(row.get(mapping.blank_reason_field, "")).strip() or None
            observations.append(
                {
                    "kind": "observation",
                    "season": mapping.season,
                    "team_key": mapping.team_key,
                    "accepted_locator": accepted_locator,
                    "hsr_source_row_sha256": hsr_source_row_sha,
                    "hsr_distance_m": hsr_distance,
                    "blank_reason": blank_reason,
                    "accepted_total_distance_m": _accepted_total_value(row, mapping.team_key),
                }
            )
    observations.sort(
        key=lambda item: (
            item["season"],
            item["team_key"],
            tuple(item["accepted_locator"][field] for field in next(
                mapping.accepted_locator_fields
                for mapping in mappings
                if mapping.season == item["season"] and mapping.team_key == item["team_key"]
            )),
        )
    )
    return metadata + observations


def _write_database_parameter_payload(
    mappings: Sequence[TeamMapping],
    results: Mapping[tuple[str, str], Mapping[str, Any]],
    mapping_hashes: Mapping[str, str],
    output_root: Path,
    parameter_path: Path,
) -> None:
    payload = _database_parameter_payload(mappings, results, mapping_hashes, output_root)
    parameter_path.parent.mkdir(parents=True, exist_ok=True)
    parameter_path.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )


def generate_all(
    mapping_paths: Mapping[str, Path] | None = None,
    *,
    output_root: Path | None = None,
    manifest_path: Path | None = None,
    validation_path: Path | None = None,
    parameter_path: Path | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    mapping_paths = mapping_paths or {season: ROOT / "pipeline" / "hsr_mappings" / f"{season}.json" for season in SEASONS}
    output_root = output_root or ROOT / "data" / "hsr_exposure"
    manifest_path = manifest_path or ROOT / "docs" / "evidence" / "hsr_exposure_mapping_manifest.json"
    validation_path = validation_path or ROOT / "docs" / "evidence" / "hsr_exposure_validation_report.json"
    parameter_path = parameter_path or output_root / DATABASE_PARAMETER_FILENAME
    all_mappings: list[TeamMapping] = []
    mapping_hashes: dict[str, str] = {}
    for season in SEASONS:
        path = Path(mapping_paths[season])
        mappings, digest = load_mapping(path, season)
        all_mappings.extend(mappings)
        mapping_hashes[season] = digest
    results = [generate_team_artifact(mapping, output_root) for mapping in all_mappings]
    by_key = {(result["season"], result["team_key"]): result for result in results}
    manifest_entries = [_manifest_entry(mapping, by_key[(mapping.season, mapping.team_key)]) for mapping in all_mappings]
    manifest = {
        "schema_version": "urc_hsr_exposure_mapping_manifest_v1",
        "mapping_files": {season: {"path": _display_path(Path(mapping_paths[season])), "sha256": mapping_hashes[season]} for season in SEASONS},
        "team_seasons": manifest_entries,
        "source_availability_summary": {season: f"{EXPECTED_AVAILABILITY[season]}/16" for season in SEASONS},
        "source_availability": {season: {"teams_total": 16, "teams_with_source": sum(1 for mapping in all_mappings if mapping.season == season and mapping.source_available), "gaps": sorted(EXPECTED_GAPS[season])} for season in SEASONS},
        "local_materialisation": {
            season: {
                "teams_with_exact_local_source": sum(
                    1 for mapping in all_mappings if mapping.season == season and mapping.local_source_available
                ),
                "declared_source_not_retained_locally": sorted(
                    mapping.team_key
                    for mapping in all_mappings
                    if mapping.season == season and mapping.source_available and not mapping.local_source_available
                ),
            }
            for season in SEASONS
        },
        "privacy": {"row_identifiers_included": False, "player_values_included": False},
    }
    validation_rows = []
    for result in results:
        validation_rows.append(
            {
                "season": result["season"],
                "team_key": result["team_key"],
                "before_row_count": result["before_row_count"],
                "after_row_count": result["after_row_count"],
                "matched_hsr_count": result["matched_hsr_count"],
                "blank_hsr_count": result["blank_hsr_count"],
                "date_range": result["date_range"],
                "source_coverage": {
                    "source_available": result["source_available"],
                    "local_source_available": result["local_source_available"],
                    "source_rows": result["source_rows"],
                    "source_rows_matched": result["source_rows_matched"],
                    "source_trace_rows": result["source_trace_rows"],
                },
                "locator_set_identity": result["locator_set_identity"],
                "blank_reason_counts": result["blank_reason_counts"],
                "month_day_swap_repair_count": result["month_day_swap_repair_count"],
                "gap_reason": result["gap_reason"],
                "output_sha256": result["output_sha256"],
            }
        )
    validation = {
        "schema_version": "urc_hsr_exposure_validation_report_v1",
        "mapping_files": {season: {"path": _display_path(Path(mapping_paths[season])), "sha256": mapping_hashes[season]} for season in SEASONS},
        "source_availability_summary": {season: f"{EXPECTED_AVAILABILITY[season]}/16" for season in SEASONS},
        "source_availability": {season: {"teams_total": 16, "teams_with_source": sum(1 for mapping in all_mappings if mapping.season == season and mapping.source_available), "gaps": sorted(EXPECTED_GAPS[season])} for season in SEASONS},
        "local_materialisation": {
            season: {
                "teams_with_exact_local_source": sum(
                    1 for mapping in all_mappings if mapping.season == season and mapping.local_source_available
                ),
                "declared_source_not_retained_locally": sorted(
                    mapping.team_key
                    for mapping in all_mappings
                    if mapping.season == season and mapping.source_available and not mapping.local_source_available
                ),
            }
            for season in SEASONS
        },
        "team_seasons": validation_rows,
        "privacy": {"row_identifiers_included": False, "player_values_included": False},
    }
    write_json(manifest_path, manifest)
    write_json(validation_path, validation)
    _write_database_parameter_payload(all_mappings, by_key, mapping_hashes, output_root, parameter_path)
    return manifest, validation


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--mapping-2024-25", type=Path, default=ROOT / "pipeline/hsr_mappings/2024-25.json")
    parser.add_argument("--mapping-2025-26", type=Path, default=ROOT / "pipeline/hsr_mappings/2025-26.json")
    parser.add_argument("--output-root", type=Path, default=ROOT / "data/hsr_exposure")
    parser.add_argument("--manifest", type=Path, default=ROOT / "docs/evidence/hsr_exposure_mapping_manifest.json")
    parser.add_argument("--validation-report", type=Path, default=ROOT / "docs/evidence/hsr_exposure_validation_report.json")
    parser.add_argument("--database-parameters", type=Path, default=None)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        manifest, validation = generate_all(
            {"2024-25": args.mapping_2024_25, "2025-26": args.mapping_2025_26},
            output_root=args.output_root,
            manifest_path=args.manifest,
            validation_path=args.validation_report,
            parameter_path=args.database_parameters,
        )
    except GenerationError as exc:
        print(f"HSR generation failed closed: {exc}", file=sys.stderr)
        return 1
    parameter_path = args.database_parameters or args.output_root / DATABASE_PARAMETER_FILENAME
    print(json.dumps({
        "team_seasons": len(manifest["team_seasons"]),
        "validation_rows": len(validation["team_seasons"]),
        "parameter_path": _display_path(parameter_path),
        "parameter_sha256": sha256_file(parameter_path),
        "parameter_bytes": parameter_path.stat().st_size,
    }, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
