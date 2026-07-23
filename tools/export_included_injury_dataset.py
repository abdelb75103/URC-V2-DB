#!/usr/bin/env python3
"""Export included injury rows from a frozen human-review workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import tempfile
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


SOURCE_SHEET = "Injury Master"
EXCLUSION_HEADER = "Exclusion Reason"
DATE_FORMAT = "%d/%m/%Y"
EXCEL_ERROR_LIKE_TOKENS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def display_path(path: Path, repo_root: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(repo_root.resolve()))
    except ValueError:
        return str(resolved)


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def serialize_value(cell: Cell) -> str | int | float:
    value = cell.value
    if value is None:
        return ""
    if cell.data_type == "f":
        raise ValueError(f"Formula found in {SOURCE_SHEET}!{cell.coordinate}")
    if cell.data_type == "e":
        raise ValueError(
            f"Excel error value {value!r} found in {SOURCE_SHEET}!{cell.coordinate}"
        )
    if isinstance(value, (datetime, date)):
        return value.strftime(DATE_FORMAT)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (str, int, float)):
        return value
    return str(value)


def write_csv_atomic(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        newline="",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temporary_path = Path(handle.name)
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(headers)
        writer.writerows(rows)
    os.replace(temporary_path, path)


def write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temporary_path = Path(handle.name)
        json.dump(payload, handle, indent=2, ensure_ascii=False)
        handle.write("\n")
    os.replace(temporary_path, path)


def export_included_dataset(
    *,
    source: Path,
    output_csv: Path,
    manifest: Path,
    script_path: Path,
    repo_root: Path,
    season: str,
    generated_at: str,
    overwrite: bool = False,
) -> dict[str, Any]:
    named_paths = {
        "source workbook": source.resolve(),
        "CSV output": output_csv.resolve(),
        "manifest output": manifest.resolve(),
    }
    if len(set(named_paths.values())) != len(named_paths):
        collisions = ", ".join(
            f"{label}={path}" for label, path in named_paths.items()
        )
        raise ValueError(
            "Source workbook, CSV output, and manifest output must resolve to "
            f"three distinct paths: {collisions}"
        )

    for output_path in (output_csv, manifest):
        if output_path.exists() and not overwrite:
            raise FileExistsError(
                f"Refusing to overwrite existing output without --overwrite: {output_path}"
            )

    workbook = load_workbook(source, read_only=True, data_only=False)
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError(f"Required sheet is missing: {SOURCE_SHEET}")
    sheet = workbook[SOURCE_SHEET]

    header_cells = list(next(sheet.iter_rows(min_row=1, max_row=1)))
    if not header_cells:
        raise ValueError(f"{SOURCE_SHEET} has no header row")
    nonblank_header_indexes = [
        index for index, cell in enumerate(header_cells) if not is_blank(cell.value)
    ]
    if not nonblank_header_indexes:
        raise ValueError(f"{SOURCE_SHEET} has no populated headers")
    last_header_index = max(nonblank_header_indexes)
    trailing_cells = header_cells[last_header_index + 1 :]
    if any(not is_blank(cell.value) for cell in trailing_cells):
        raise ValueError("Unexpected populated header after the last retained column")

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in header_cells[: last_header_index + 1]
    ]
    if any(not header for header in headers):
        raise ValueError("Blank header found inside the retained column range")
    if len(headers) != len(set(headers)):
        raise ValueError("Duplicate headers found in the retained column range")
    if headers.count(EXCLUSION_HEADER) != 1:
        raise ValueError(f"Expected exactly one {EXCLUSION_HEADER!r} column")

    exclusion_index = headers.index(EXCLUSION_HEADER)
    source_rows = 0
    included_source_rows: list[int] = []
    included_rows: list[list[Any]] = []
    source_by_team: Counter[str] = Counter()
    included_by_team: Counter[str] = Counter()
    excluded_by_team: Counter[str] = Counter()
    output_warning_counts: Counter[tuple[str, str]] = Counter()

    for source_row, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        retained_cells = list(row_cells[: len(headers)])
        if len(retained_cells) != len(headers):
            raise ValueError(f"Row {source_row} is shorter than the header row")
        source_rows += 1
        team_value = retained_cells[0].value
        team = "" if team_value is None else str(team_value).strip()
        source_by_team[team] += 1

        for cell in retained_cells:
            if cell.data_type == "f":
                raise ValueError(
                    f"Formula found in {SOURCE_SHEET}!{cell.coordinate}; "
                    "the export requires fixed reviewed values"
                )

        exclusion_value = retained_cells[exclusion_index].value
        if is_blank(exclusion_value):
            included_source_rows.append(source_row)
            included_by_team[team] += 1
            serialized_row = [serialize_value(cell) for cell in retained_cells]
            included_rows.append(serialized_row)
            for header, value in zip(headers, serialized_row, strict=True):
                if isinstance(value, str) and value in EXCEL_ERROR_LIKE_TOKENS:
                    output_warning_counts[(header, value)] += 1
        else:
            excluded_by_team[team] += 1

    if not included_rows:
        raise ValueError("The inclusion rule selected zero rows")

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    manifest.parent.mkdir(parents=True, exist_ok=True)
    write_csv_atomic(output_csv, headers, included_rows)

    row_mapping_bytes = (
        "\n".join(str(row_number) for row_number in included_source_rows) + "\n"
    ).encode("utf-8")
    all_teams = sorted(source_by_team)
    counts_by_team = {
        team: {
            "source_rows": source_by_team[team],
            "included_rows": included_by_team[team],
            "excluded_rows": excluded_by_team[team],
        }
        for team in all_teams
    }
    payload: dict[str, Any] = {
        "artifact_type": "included_injury_dataset",
        "status": "working_dataset_for_further_cleanup",
        "season": season,
        "generated_at": generated_at,
        "source": {
            "workbook": display_path(source, repo_root),
            "workbook_sha256": sha256_file(source),
            "sheet": SOURCE_SHEET,
            "source_rows": source_rows,
            "columns": len(headers),
            "headers": headers,
        },
        "selection": {
            "rule": f"{EXCLUSION_HEADER} is blank after trimming whitespace",
            "included_rows": len(included_rows),
            "excluded_rows": source_rows - len(included_rows),
            "included_source_rows": included_source_rows,
            "included_source_rows_sha256": hashlib.sha256(
                row_mapping_bytes
            ).hexdigest(),
            "csv_row_mapping": (
                "included_source_rows is ordered; its first item maps to CSV row 2"
            ),
        },
        "output": {
            "csv": display_path(output_csv, repo_root),
            "csv_sha256": sha256_file(output_csv),
            "data_rows": len(included_rows),
            "columns": len(headers),
            "encoding": "UTF-8",
            "line_ending": "LF",
            "date_serialization": (
                "Native Excel date and datetime cells are written as DD/MM/YYYY; "
                "existing text values are preserved"
            ),
            "preserved_source_value_warnings": [
                {
                    "column": column,
                    "value": value,
                    "count": count,
                    "note": (
                        "Literal source value preserved; this is not a CSV formula "
                        "or export error"
                    ),
                }
                for (column, value), count in sorted(output_warning_counts.items())
            ],
        },
        "counts_by_team": counts_by_team,
        "generator": {
            "script": display_path(script_path, repo_root),
            "script_sha256": sha256_file(script_path),
        },
    }
    write_json_atomic(manifest, payload)
    return payload


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export rows with a blank Exclusion Reason from a frozen URC "
            "human-review workbook."
        )
    )
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--output-csv", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--season", default="2024-25")
    parser.add_argument(
        "--generated-at",
        default=datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
    )
    parser.add_argument("--repo-root", type=Path, default=Path.cwd())
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    script_path = Path(__file__).resolve()
    payload = export_included_dataset(
        source=args.source,
        output_csv=args.output_csv,
        manifest=args.manifest,
        script_path=script_path,
        repo_root=args.repo_root,
        season=args.season,
        generated_at=args.generated_at,
        overwrite=args.overwrite,
    )
    print(
        json.dumps(
            {
                "source_rows": payload["source"]["source_rows"],
                "included_rows": payload["selection"]["included_rows"],
                "excluded_rows": payload["selection"]["excluded_rows"],
                "columns": payload["output"]["columns"],
                "csv_sha256": payload["output"]["csv_sha256"],
            },
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
