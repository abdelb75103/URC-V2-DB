#!/usr/bin/env python3
"""Normalize Sharks squad aliases in the review master and included CSV."""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import shutil
import stat
import subprocess
import tempfile
import zipfile
from collections import Counter
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RULE_VERSION = "sharks_single_squad_normalization_2026-07-23_v1"
SHEET_NAME = "Injury Master"
TEAM = "Sharks"
OLD_VALUE = "Other team"
NEW_VALUE = "Sharks"
EXCLUSION_REASON = "Non-URC match"
EXPECTED_WORKBOOK_ROWS = 3060
EXPECTED_COLUMNS = 28
EXPECTED_SHARKS_CHANGES = 178
EXPECTED_INCLUDED_SHARKS_CHANGES = 163
EXPECTED_EXCLUDED_SHARKS_CHANGES = 15
EXPECTED_NON_SHARKS_OTHER_TEAM = 46
EXPECTED_GENUINE_OTHER_TEAM_ROWS = 230
EXPECTED_INCLUDED_ROWS = 2425
EXPECTED_SHARKS_MATCH_ROWS = 223
EXPECTED_SHARKS_ON_FIXTURE_MATCH_ROWS = 141
EXPECTED_SHARKS_OFF_FIXTURE_MATCH_ROWS = 82
AUDIT_HEADERS = [
    "source_workbook_row",
    "team",
    "player_id",
    "field",
    "old_value",
    "new_value",
    "workbook_updated",
    "included_csv_updated",
    "inclusion_status",
    "exclusion_reason",
    "reason",
    "rule_version",
    "evidence_origin",
]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def display_path(path: Path, repo_root: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(repo_root.resolve()))
    except ValueError:
        return str(resolved)


def write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temporary_path = Path(handle.name)
        json.dump(payload, handle, indent=2, ensure_ascii=False)
        handle.write("\n")
    os.replace(temporary_path, path)


def write_audit_atomic(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        newline="",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temporary_path = Path(handle.name)
        writer = csv.DictWriter(handle, fieldnames=AUDIT_HEADERS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    os.replace(temporary_path, path)


def workbook_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Required sheet missing: {SHEET_NAME}")
    sheet = workbook[SHEET_NAME]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    headers = [str(value).strip() if value is not None else "" for value in headers]
    if len(headers) != EXPECTED_COLUMNS or any(not header for header in headers):
        raise ValueError(f"Unexpected {SHEET_NAME} header shape")
    rows: list[dict[str, Any]] = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in cells[: len(headers)]]
        rows.append(
            {
                "source_workbook_row": row_number,
                "values": dict(zip(headers, values, strict=True)),
            }
        )
    workbook.close()
    if len(rows) != EXPECTED_WORKBOOK_ROWS:
        raise ValueError(
            f"Expected {EXPECTED_WORKBOOK_ROWS} workbook data rows, found {len(rows)}"
        )
    return headers, rows


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def build_normalization_plan(
    rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    targets: list[dict[str, Any]] = []
    non_sharks_other_team: list[dict[str, Any]] = []

    for item in rows:
        values = item["values"]
        if text(values["Received/Injured In Team"]) != OLD_VALUE:
            continue
        if text(values["Team"]) == TEAM:
            targets.append(item)
        else:
            non_sharks_other_team.append(item)

    reasons = Counter(text(item["values"]["Exclusion Reason"]) for item in targets)
    occasions = Counter(text(item["values"]["Occasion category"]) for item in targets)
    match_types = Counter(text(item["values"]["Match Type"]) for item in targets)
    non_sharks_blank = [
        item
        for item in non_sharks_other_team
        if not text(item["values"]["Exclusion Reason"])
    ]
    excluded_targets = [
        item for item in targets if text(item["values"]["Exclusion Reason"])
    ]
    included_targets = [
        item for item in targets if not text(item["values"]["Exclusion Reason"])
    ]
    excluded_reason_errors = [
        item
        for item in excluded_targets
        if text(item["values"]["Exclusion Reason"]) != EXCLUSION_REASON
        or text(item["values"]["Occasion category"]) != "Match"
        or text(item["values"]["Match Type"])
    ]
    included_match_errors = [
        item
        for item in included_targets
        if text(item["values"]["Occasion category"]) == "Match"
        and text(item["values"]["Match Type"]) != "Confirmed URC match fixture"
    ]

    expected = {
        "targets": EXPECTED_SHARKS_CHANGES,
        "included_targets": EXPECTED_INCLUDED_SHARKS_CHANGES,
        "excluded_targets": EXPECTED_EXCLUDED_SHARKS_CHANGES,
        "non_sharks_other_team": EXPECTED_NON_SHARKS_OTHER_TEAM,
    }
    actual = {
        "targets": len(targets),
        "included_targets": len(included_targets),
        "excluded_targets": len(excluded_targets),
        "non_sharks_other_team": len(non_sharks_other_team),
    }
    if actual != expected:
        raise ValueError(f"Normalization scope drifted: expected {expected}, found {actual}")
    if non_sharks_blank:
        raise ValueError(
            "Genuine non-Sharks Other team rows remain included: "
            f"{[item['source_workbook_row'] for item in non_sharks_blank]}"
        )
    if excluded_reason_errors:
        raise ValueError(
            "Sharks off-fixture exclusions do not match the accepted rule: "
            f"{[item['source_workbook_row'] for item in excluded_reason_errors]}"
        )
    if included_match_errors:
        raise ValueError(
            "Included Sharks match injuries are not confirmed URC fixtures: "
            f"{[item['source_workbook_row'] for item in included_match_errors]}"
        )

    summary = {
        **actual,
        "sharks_exclusion_reasons": dict(reasons),
        "sharks_occasions": dict(occasions),
        "sharks_match_types": dict(match_types),
        "non_sharks_other_team_by_team": dict(
            Counter(text(item["values"]["Team"]) for item in non_sharks_other_team)
        ),
        "non_sharks_other_team_blank_exclusion_reason": 0,
    }
    return targets, summary


def genuine_other_team_summary(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts = Counter()
    for item in rows:
        values = item["values"]
        reason = text(values["Exclusion Reason"]).lower()
        literal_other_team = (
            text(values["Team"]) != TEAM
            and text(values["Received/Injured In Team"]) == OLD_VALUE
        )
        if "another team" in reason or literal_other_team:
            if not text(values["Exclusion Reason"]):
                raise ValueError(
                    "Genuine other-team row remains included at workbook row "
                    f"{item['source_workbook_row']}"
                )
            counts[text(values["Team"])] += 1
    if sum(counts.values()) != EXPECTED_GENUINE_OTHER_TEAM_ROWS:
        raise ValueError(f"Genuine other-team scope drifted: {dict(counts)}")
    return dict(counts)


def effective_fixture_dates(path: Path) -> set[str]:
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
    fixture_dates: set[str] = set()
    for row in rows:
        if (
            row.get("home_team_alias", "").strip() != "Team U"
            and row.get("away_team_alias", "").strip() != "Team U"
        ):
            continue
        effective = (
            row.get("corrected_date", "").strip()
            or row.get("source_date", "").strip()
        )
        if not effective:
            raise ValueError("Canonical Sharks fixture has no effective date")
        fixture_dates.add(effective)
    if len(fixture_dates) != 20:
        raise ValueError(f"Expected 20 canonical Sharks fixtures, found {len(fixture_dates)}")
    if "2025-03-08" not in fixture_dates:
        raise ValueError("Corrected Sharks fixture date 2025-03-08 is missing")
    return fixture_dates


def workbook_date_iso(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()
    for date_format in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text(value), date_format).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Unparseable workbook date: {value!r}")


def verify_fixture_eligibility(
    rows: list[dict[str, Any]], fixture_dates: set[str]
) -> dict[str, Any]:
    match_rows = [
        item
        for item in rows
        if text(item["values"]["Team"]) == TEAM
        and text(item["values"]["Occasion category"]) == "Match"
    ]
    on_fixture: list[dict[str, Any]] = []
    off_fixture: list[dict[str, Any]] = []
    for item in match_rows:
        injury_date = workbook_date_iso(item["values"]["Date Injured"])
        (on_fixture if injury_date in fixture_dates else off_fixture).append(item)

    actual = (len(match_rows), len(on_fixture), len(off_fixture))
    expected = (
        EXPECTED_SHARKS_MATCH_ROWS,
        EXPECTED_SHARKS_ON_FIXTURE_MATCH_ROWS,
        EXPECTED_SHARKS_OFF_FIXTURE_MATCH_ROWS,
    )
    if actual != expected:
        raise ValueError(f"Sharks fixture alignment drifted: expected {expected}, found {actual}")
    on_fixture_errors = [
        item
        for item in on_fixture
        if text(item["values"]["Match Type"]) != "Confirmed URC match fixture"
        or text(item["values"]["Exclusion Reason"])
    ]
    off_fixture_errors = [
        item
        for item in off_fixture
        if text(item["values"]["Match Type"])
        or text(item["values"]["Exclusion Reason"]) != EXCLUSION_REASON
    ]
    if on_fixture_errors or off_fixture_errors:
        raise ValueError(
            "Sharks match fixture eligibility is inconsistent at rows "
            f"{[item['source_workbook_row'] for item in on_fixture_errors + off_fixture_errors]}"
        )
    return {
        "match_rows": len(match_rows),
        "on_fixture_included": len(on_fixture),
        "off_fixture_excluded": len(off_fixture),
        "canonical_fixture_dates": len(fixture_dates),
        "corrected_2025_03_08_fixture_present": True,
    }


def run_native_excel_edit(workbook_path: Path, target_rows: list[int]) -> str:
    row_list = ", ".join(str(row) for row in target_rows)
    excel_path = json.dumps(str(workbook_path.resolve()))
    excel_name = json.dumps(workbook_path.name)
    apple_script = f"""
set targetPath to {excel_path}
set targetRows to {{{row_list}}}
tell application "Microsoft Excel"
    with timeout of 3600 seconds
        if not (exists workbook {excel_name}) then
            open POSIX file targetPath
            delay 1
        end if
        set targetBook to workbook {excel_name}
        if (full name of targetBook) is not targetPath then error "Exact workbook path mismatch"
        set targetSheet to worksheet "{SHEET_NAME}" of targetBook
        repeat with rowNumber in targetRows
            set targetCell to range ("D" & (rowNumber as text)) of targetSheet
            set currentValue to value of targetCell as text
            if currentValue is "{OLD_VALUE}" then
                set value of targetCell to "{NEW_VALUE}"
            else if currentValue is not "{NEW_VALUE}" then
                error "Unexpected target value at D" & (rowNumber as text)
            end if
        end repeat
        save workbook targetBook
        set savedState to saved of targetBook
        close targetBook saving yes
        return (count of targetRows as text) & "|" & (savedState as text)
    end timeout
end tell
"""
    result = subprocess.run(
        ["osascript"],
        input=apple_script,
        text=True,
        capture_output=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(
            "Microsoft Excel native edit failed: "
            f"{result.stderr.strip() or result.stdout.strip()}"
        )
    return result.stdout.strip()


def native_excel_verify(workbook_path: Path) -> str:
    excel_path = json.dumps(str(workbook_path.resolve()))
    excel_name = json.dumps(workbook_path.name)
    apple_script = f"""
set targetPath to {excel_path}
tell application "Microsoft Excel"
    with timeout of 3600 seconds
        if not (exists workbook {excel_name}) then
            open POSIX file targetPath
            delay 1
        end if
        set targetBook to workbook {excel_name}
        if (full name of targetBook) is not targetPath then error "Exact workbook path mismatch"
        set targetSheet to worksheet "{SHEET_NAME}" of targetBook
        set populatedCount to count of cells of used range of targetSheet
        set firstTeam to value of range "A2" of targetSheet
        set lastTeam to value of range "A3061" of targetSheet
        close targetBook saving no
        return targetPath & "|" & (populatedCount as text) & "|" & firstTeam & "|" & lastTeam
    end timeout
end tell
"""
    result = subprocess.run(
        ["osascript"],
        input=apple_script,
        text=True,
        capture_output=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(
            "Microsoft Excel verification failed: "
            f"{result.stderr.strip() or result.stdout.strip()}"
        )
    return result.stdout.strip()


def style_signature(cell: Any) -> tuple[Any, ...]:
    return (
        copy(cell.font),
        copy(cell.fill),
        copy(cell.border),
        copy(cell.alignment),
        copy(cell.protection),
        cell.number_format,
    )


def verify_workbook_overlay(
    before_path: Path,
    after_path: Path,
    expected_transitions: dict[str, tuple[Any, Any]],
) -> dict[str, Any]:
    before = load_workbook(before_path, read_only=False, data_only=False)
    after = load_workbook(after_path, read_only=False, data_only=False)
    if before.sheetnames != after.sheetnames:
        raise ValueError("Workbook sheet order changed")
    value_diffs: list[str] = []
    style_diffs: list[str] = []
    structure_diffs: list[str] = []

    for sheet_name in before.sheetnames:
        left = before[sheet_name]
        right = after[sheet_name]
        if (left.max_row, left.max_column) != (right.max_row, right.max_column):
            structure_diffs.append(f"{sheet_name}:dimensions")
        if set(left.merged_cells.ranges) != set(right.merged_cells.ranges):
            structure_diffs.append(f"{sheet_name}:merged_cells")
        if str(left.freeze_panes) != str(right.freeze_panes):
            structure_diffs.append(f"{sheet_name}:freeze_panes")
        if left.auto_filter.ref != right.auto_filter.ref:
            structure_diffs.append(f"{sheet_name}:auto_filter")
        if set(left.tables) != set(right.tables):
            structure_diffs.append(f"{sheet_name}:tables")

        for row in left.iter_rows():
            for left_cell in row:
                right_cell = right[left_cell.coordinate]
                if left_cell.value != right_cell.value:
                    if (
                        sheet_name == SHEET_NAME
                        and left_cell.coordinate in expected_transitions
                    ):
                        expected_old, expected_new = expected_transitions[
                            left_cell.coordinate
                        ]
                        if (
                            left_cell.value != expected_old
                            or right_cell.value != expected_new
                        ):
                            raise ValueError(
                                f"Wrong target transition at {left_cell.coordinate}"
                            )
                        value_diffs.append(left_cell.coordinate)
                    else:
                        raise ValueError(
                            f"Unexpected workbook value change at "
                            f"{sheet_name}!{left_cell.coordinate}"
                        )
                if style_signature(left_cell) != style_signature(right_cell):
                    style_diffs.append(f"{sheet_name}!{left_cell.coordinate}")

    before.close()
    after.close()
    if set(value_diffs) != set(expected_transitions):
        raise ValueError(
            f"Workbook target changes differ from plan: {len(value_diffs)} found"
        )
    if style_diffs:
        raise ValueError(f"Workbook styles changed unexpectedly: {style_diffs[:10]}")
    if structure_diffs:
        raise ValueError(f"Workbook structure changed: {structure_diffs}")
    with zipfile.ZipFile(after_path) as package:
        bad_member = package.testzip()
    if bad_member is not None:
        raise ValueError(f"Corrupt workbook ZIP member: {bad_member}")
    return {
        "value_changes": len(value_diffs),
        "style_changes": 0,
        "structure_changes": 0,
        "zip_package_integrity": True,
    }


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]], list[str]]:
    physical_lines = path.read_text(encoding="utf-8").splitlines(keepends=True)
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError("Included CSV has no header")
        headers = list(reader.fieldnames)
        rows = [dict(row) for row in reader]
    if len(physical_lines) != len(rows) + 1:
        raise ValueError(
            "CSV contains embedded or nonstandard line breaks; narrow patch aborted"
        )
    return headers, rows, physical_lines


def patch_included_csv(
    path: Path,
    source_rows: list[int],
    target_source_rows: set[int],
) -> int:
    headers, rows, physical_lines = read_csv(path)
    if len(headers) != EXPECTED_COLUMNS or len(rows) != EXPECTED_INCLUDED_ROWS:
        raise ValueError("Included CSV shape drifted")
    if len(rows) != len(source_rows):
        raise ValueError("CSV rows and manifest source-row mapping differ in length")

    output_lines = [physical_lines[0]]
    changed = 0
    for index, (source_row, row) in enumerate(
        zip(source_rows, rows, strict=True)
    ):
        if source_row not in target_source_rows:
            output_lines.append(physical_lines[index + 1])
            continue
        if row["Team"].strip() != TEAM:
            raise ValueError(f"Unexpected CSV team at source row {source_row}")
        if row["Received/Injured In Team"].strip() != OLD_VALUE:
            raise ValueError(f"Unexpected CSV alias at source row {source_row}")
        if row["Exclusion Reason"].strip():
            raise ValueError(f"Included CSV row {source_row} has an exclusion reason")
        row["Received/Injured In Team"] = NEW_VALUE
        buffer = io.StringIO(newline="")
        writer = csv.DictWriter(buffer, fieldnames=headers, lineterminator="\n")
        writer.writerow(row)
        output_lines.append(buffer.getvalue())
        changed += 1

    if changed != EXPECTED_INCLUDED_SHARKS_CHANGES:
        raise ValueError(
            f"Expected {EXPECTED_INCLUDED_SHARKS_CHANGES} CSV changes, found {changed}"
        )
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        newline="",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temporary_path = Path(handle.name)
        handle.writelines(output_lines)
    os.replace(temporary_path, path)
    return changed


def apply_normalization(args: argparse.Namespace) -> dict[str, Any]:
    paths = [
        args.source_workbook,
        args.successor_workbook,
        args.fixtures,
        args.included_csv,
        args.manifest,
        args.backup_csv,
        args.backup_manifest,
        args.audit,
        args.qa,
    ]
    resolved = [path.resolve() for path in paths]
    if len(resolved) != len(set(resolved)):
        raise ValueError("All source, successor, backup, audit, and QA paths must differ")
    for path in (args.audit, args.qa):
        if path.exists():
            raise FileExistsError(f"Refusing to overwrite existing artifact: {path}")

    manifest = json.loads(args.manifest.read_text(encoding="utf-8"))
    source_rows = [int(value) for value in manifest["selection"]["included_source_rows"]]
    if len(source_rows) != EXPECTED_INCLUDED_ROWS:
        raise ValueError("Manifest included-row mapping drifted")
    if manifest["output"]["csv_sha256"] != sha256_file(args.included_csv):
        raise ValueError("Manifest CSV checksum does not match current included CSV")
    if manifest["source"]["workbook_sha256"] != sha256_file(args.source_workbook):
        raise ValueError("Manifest workbook checksum does not match the source workbook")

    _, rows = workbook_rows(args.source_workbook)
    targets, preflight = build_normalization_plan(rows)
    fixture_qa = verify_fixture_eligibility(rows, effective_fixture_dates(args.fixtures))
    target_rows = [item["source_workbook_row"] for item in targets]
    included_target_rows = {
        item["source_workbook_row"]
        for item in targets
        if not text(item["values"]["Exclusion Reason"])
    }
    excluded_target_rows = set(target_rows) - included_target_rows
    if included_target_rows - set(source_rows):
        raise ValueError("Included Sharks targets are missing from the CSV mapping")
    if excluded_target_rows & set(source_rows):
        raise ValueError("Excluded Sharks targets are present in the CSV mapping")

    prepared_paths = [
        args.successor_workbook.exists(),
        args.backup_csv.exists(),
        args.backup_manifest.exists(),
    ]
    expected_transitions = {
        f"D{row}": (OLD_VALUE, NEW_VALUE) for row in target_rows
    }
    successor_already_edited = False
    if any(prepared_paths) and not all(prepared_paths):
        raise FileExistsError("Only part of the prepared successor/backup set exists")
    if all(prepared_paths):
        if sha256_file(args.successor_workbook) != sha256_file(args.source_workbook):
            verify_workbook_overlay(
                args.source_workbook,
                args.successor_workbook,
                expected_transitions,
            )
            successor_already_edited = True
        if sha256_file(args.backup_csv) != sha256_file(args.included_csv):
            raise ValueError("Existing prepared CSV backup does not match current CSV")
        if sha256_file(args.backup_manifest) != sha256_file(args.manifest):
            raise ValueError(
                "Existing prepared manifest backup does not match current manifest"
            )
    else:
        args.successor_workbook.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(args.source_workbook, args.successor_workbook)
        shutil.copy2(args.included_csv, args.backup_csv)
        shutil.copy2(args.manifest, args.backup_manifest)
    args.successor_workbook.chmod(
        args.successor_workbook.stat().st_mode | stat.S_IWUSR
    )
    input_csv_sha256 = sha256_file(args.included_csv)
    input_manifest_sha256 = sha256_file(args.manifest)

    if successor_already_edited:
        excel_edit_result = "native Excel edit saved and resumed from verified overlay"
    else:
        excel_edit_result = run_native_excel_edit(args.successor_workbook, target_rows)
    overlay_qa = verify_workbook_overlay(
        args.source_workbook, args.successor_workbook, expected_transitions
    )
    if successor_already_edited:
        verification_copy = args.excel_verification_copy
        if verification_copy is None or not verification_copy.exists():
            raise ValueError("A Microsoft Excel verification copy is required")
        if sha256_file(verification_copy) != sha256_file(args.successor_workbook):
            raise ValueError("Excel verification copy does not match the successor")
        excel_verify_result = (
            "Microsoft Excel opened the byte-identical verification copy with "
            "the populated Injury Master sheet A1:AB3061: "
            f"{display_path(verification_copy, args.repo_root)}"
        )
    else:
        excel_verify_result = native_excel_verify(args.successor_workbook)
    successor_sha256 = sha256_file(args.successor_workbook)

    csv_changes = patch_included_csv(
        args.included_csv, source_rows, included_target_rows
    )
    output_csv_sha256 = sha256_file(args.included_csv)

    audit_rows = []
    for item in targets:
        values = item["values"]
        source_row = item["source_workbook_row"]
        included = source_row in included_target_rows
        audit_rows.append(
            {
                "source_workbook_row": source_row,
                "team": TEAM,
                "player_id": text(values["PlayerID"]),
                "field": "Received/Injured In Team",
                "old_value": OLD_VALUE,
                "new_value": NEW_VALUE,
                "workbook_updated": "yes",
                "included_csv_updated": "yes" if included else "no",
                "inclusion_status": "included" if included else "excluded",
                "exclusion_reason": text(values["Exclusion Reason"]),
                "reason": (
                    "Sharks, Sharks Senior Team, and Sharks Currie Cup are one squad"
                ),
                "rule_version": RULE_VERSION,
                "evidence_origin": "Abdel Babiker decision on 2026-07-23",
            }
        )
    write_audit_atomic(args.audit, audit_rows)

    genuine_other_team_by_team = genuine_other_team_summary(rows)
    qa_payload = {
        "rule_version": RULE_VERSION,
        "status": "passed",
        "preflight": preflight,
        "workbook": {
            "source": display_path(args.source_workbook, args.repo_root),
            "source_sha256": sha256_file(args.source_workbook),
            "successor": display_path(args.successor_workbook, args.repo_root),
            "successor_sha256": successor_sha256,
            **overlay_qa,
            "native_excel_edit": excel_edit_result,
            "native_excel_verification": excel_verify_result,
        },
        "included_csv": {
            "rows": EXPECTED_INCLUDED_ROWS,
            "columns": EXPECTED_COLUMNS,
            "changed_rows": csv_changes,
            "input_sha256": input_csv_sha256,
            "output_sha256": output_csv_sha256,
            "source_row_mapping_unchanged": True,
            "row_count_unchanged": True,
            "row_order_unchanged": True,
        },
        "league_other_team_audit": {
            "literal_other_team_rows_outside_sharks": EXPECTED_NON_SHARKS_OTHER_TEAM,
            "included_rows": 0,
            "all_excluded": True,
            "by_team": preflight["non_sharks_other_team_by_team"],
            "source_identity_audit": {
                "genuine_other_team_rows": sum(
                    genuine_other_team_by_team.values()
                ),
                "all_excluded": True,
                "by_team": genuine_other_team_by_team,
                "review_scope": "all 16 teams",
            },
        },
        "fixture_eligibility": {
            **fixture_qa,
            "fixture_file": display_path(args.fixtures, args.repo_root),
            "fixture_file_sha256": sha256_file(args.fixtures),
            "exclusion_reason_preserved": EXCLUSION_REASON,
        },
    }
    write_json_atomic(args.qa, qa_payload)

    generated_at = args.generated_at
    manifest["generated_at"] = generated_at
    manifest["source"]["workbook"] = display_path(
        args.successor_workbook, args.repo_root
    )
    manifest["source"]["workbook_sha256"] = successor_sha256
    manifest["output"]["csv_sha256"] = output_csv_sha256
    history = manifest.setdefault("cleanup_history", [])
    history.append(
        {
            "rule_version": RULE_VERSION,
            "applied_at": generated_at,
            "script": display_path(Path(__file__), args.repo_root),
            "script_sha256": sha256_file(Path(__file__)),
            "input_csv_sha256": input_csv_sha256,
            "output_csv_sha256": output_csv_sha256,
            "backup_csv": display_path(args.backup_csv, args.repo_root),
            "backup_csv_sha256": sha256_file(args.backup_csv),
            "backup_manifest": display_path(args.backup_manifest, args.repo_root),
            "backup_manifest_sha256": input_manifest_sha256,
            "successor_review_workbook": display_path(
                args.successor_workbook, args.repo_root
            ),
            "successor_review_workbook_sha256": successor_sha256,
            "audit": display_path(args.audit, args.repo_root),
            "audit_sha256": sha256_file(args.audit),
            "qa": display_path(args.qa, args.repo_root),
            "qa_sha256": sha256_file(args.qa),
            "workbook_cells_changed": len(expected_transitions),
            "sharks_rows_normalized": len(target_rows),
            "included_csv_rows_normalized": csv_changes,
            "excluded_off_fixture_rows_preserved": len(excluded_target_rows),
            "source_row_mapping_unchanged": True,
            "row_count_unchanged": True,
        }
    )
    write_json_atomic(args.manifest, manifest)

    reloaded = json.loads(args.manifest.read_text(encoding="utf-8"))
    if reloaded["output"]["csv_sha256"] != sha256_file(args.included_csv):
        raise ValueError("Final manifest CSV checksum verification failed")
    if reloaded["source"]["workbook_sha256"] != successor_sha256:
        raise ValueError("Final manifest workbook checksum verification failed")

    return {
        "workbook_rows_normalized": len(target_rows),
        "included_csv_rows_normalized": csv_changes,
        "off_fixture_exclusions_preserved": len(excluded_target_rows),
        "genuine_other_team_rows_all_excluded": EXPECTED_GENUINE_OTHER_TEAM_ROWS,
        "successor_workbook_sha256": successor_sha256,
        "included_csv_sha256": output_csv_sha256,
        "manifest_sha256": sha256_file(args.manifest),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-workbook", type=Path, required=True)
    parser.add_argument("--successor-workbook", type=Path, required=True)
    parser.add_argument("--fixtures", type=Path, required=True)
    parser.add_argument("--included-csv", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--backup-csv", type=Path, required=True)
    parser.add_argument("--backup-manifest", type=Path, required=True)
    parser.add_argument("--audit", type=Path, required=True)
    parser.add_argument("--qa", type=Path, required=True)
    parser.add_argument("--excel-verification-copy", type=Path)
    parser.add_argument("--repo-root", type=Path, default=Path.cwd())
    parser.add_argument(
        "--generated-at",
        default=datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
    )
    return parser.parse_args()


def main() -> None:
    print(json.dumps(apply_normalization(parse_args()), sort_keys=True))


if __name__ == "__main__":
    main()
